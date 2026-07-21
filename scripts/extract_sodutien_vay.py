# -*- coding: utf-8 -*-
"""Dẫn xuất SỐ DƯ TIỀN (03B_SODU_TIEN -> SDT) + DƯ NỢ VAY (04_VAY -> VAY) từ sheet
'TC01_SD TIỀN' của Báo cáo tiền tập đoàn (cùng file với extract_thuchi.py).

Cấu trúc nguồn (bảng dọc phân cấp):
  - dòng SECTION: cột 0 = 'I'/'II'/'III', cột 1 = LOẠI TIỀN ('TIỀN MẶT'/'TIỀN GỬI'/'TIỀN VAY')
  - dòng CÔNG TY: cột 2 = tên pháp nhân (subtotal của công ty trong section)
  - dòng NGÂN HÀNG: cột 3 = tên NH ('BIDV'), dòng con kỳ hạn ('BIDV vay ngắn hạn') bỏ qua.

SDT: mỗi dòng CÔNG TY -> 1 record 03B (cột tiền theo section). VAY: mỗi dòng NGÂN HÀNG cha
trong section TIỀN VAY -> 1 record 04_VAY (Đơn vị = công ty gần nhất phía trên).
Import qua template_filler.import_filled — idempotent theo (source_file, report_type),
KHÔNG đụng THUCHI của cùng file (delete scope theo report_type của file điền).

Chạy: .venv/bin/python scripts/extract_sodutien_vay.py <file.xlsx> --period 2026-01
"""
import argparse
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from servers import template_filler as tf  # noqa: E402
from servers.common import source_catalog as SC  # noqa: E402
from servers.common import be_bridge as bb  # noqa: E402

SHEET = "TC01_SD TIỀN"
# Tên trong nguồn không khớp fuzzy master (backend resolve được 4/8) -> alias bổ sung.
_NAME_ALIAS = {"an taxi": "AAG", "an ks": "AAG", "an khach san": "AAG", "global ai": "GA",
               "htx xanh tuyen quang": "HTX_XTQ", "htx xanh vinh phuc": "HTX_XVP",
               "xanh tuyen quang": "HTX_XTQ"}
_SEC_COL = {"tien mat": "Tiền mặt (tỷ)", "tien gui": "Tiền gửi NH (tỷ)",
            "tien vay": "Số dư tiền vay (tỷ) — đối chiếu 04_VAY"}
_CHILD_RE = re.compile(r"(ngan han|trung han|dai han)")


def _norm(s):
    return bb.remove_diacritics("" if s is None else str(s)).strip().lower()


def _resolve_co(name):
    code = bb.master.resolve_company_code(name)
    if code:
        return code
    return _NAME_ALIAS.get(_norm(name))


def _num(v):
    return float(v) if isinstance(v, (int, float)) else None


def extract(path: str, period: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET not in wb.sheetnames:
        return {"ok": False, "error": f"Không thấy sheet {SHEET}"}
    rows = [list(r) for r in wb[SHEET].iter_rows(values_only=True)]
    hdr_i = next((i for i, r in enumerate(rows[:15])
                  if any("dau ky" in _norm(c) for c in r) and any("den ngay" in _norm(c) for c in r)), None)
    if hdr_i is None:
        return {"ok": False, "error": "Không thấy dòng header (ĐẦU KỲ / ĐẾN NGÀY HIỆN TẠI)"}
    hdr = rows[hdr_i]
    c_dau = next(j for j, c in enumerate(hdr) if "dau ky" in _norm(c))
    c_cuoi = next(j for j, c in enumerate(hdr) if "den ngay" in _norm(c))

    sdt_recs, vay_recs, unresolved = [], [], set()
    dropped_orphans = []
    section = None   # normalized LOẠI TIỀN hiện tại
    company = None   # (tên, mã) công ty gần nhất trong section
    orphans = []     # dòng ngân hàng đứng TRƯỚC dòng công ty của nó (block TC trong file thật:
                     # bank 8-12 rồi mới 'Thịnh Cường' 13) — gán khi subtotal công ty khớp tổng.

    def _flush_orphans(code, subtotal_cuoi):
        """Gán orphan banks cho công ty vừa gặp NẾU tổng cuối kỳ khớp subtotal (±0.5%)."""
        nonlocal orphans
        if not orphans:
            return
        tong = sum(b["cuoi"] or 0.0 for b in orphans)
        if subtotal_cuoi is not None and abs(tong - subtotal_cuoi) <= max(abs(subtotal_cuoi) * 0.005, 1e4):
            for b in orphans:
                rec = {"Kỳ": period, "Đơn vị": code, "Ngân hàng": b["bank"]}
                if b["dau"] is not None:
                    rec["Dư nợ đầu kỳ (tỷ)"] = round(b["dau"] / 1e9, 9)
                if b["cuoi"] is not None:
                    rec["Dư nợ cuối kỳ (tỷ)"] = round(b["cuoi"] / 1e9, 9)
                vay_recs.append(rec)
        else:
            dropped_orphans.extend(b["bank"] for b in orphans)
        orphans = []

    for r in rows[hdr_i + 1:]:
        c0 = str(r[0]).strip() if r[0] not in (None, "") else ""
        c1 = str(r[1]).strip() if len(r) > 1 and r[1] not in (None, "") else ""
        c2 = str(r[2]).strip() if len(r) > 2 and r[2] not in (None, "") else ""
        c3 = str(r[3]).strip() if len(r) > 3 and r[3] not in (None, "") else ""
        if c0 and c1:                          # dòng section (I/II/III + tên loại tiền)
            section = next((k for k in _SEC_COL if k in _norm(c1)), None)
            company, orphans = None, []
            continue
        if not section:
            continue
        if c2:                                  # dòng CÔNG TY (subtotal)
            code = _resolve_co(c2)
            company = (c2, code)
            cuoi = _num(r[c_cuoi]) if c_cuoi < len(r) else None
            if code is None:
                unresolved.add(c2)
                orphans = []
                continue
            if section == "tien vay":
                _flush_orphans(code, cuoi)
            if cuoi is not None:
                sdt_recs.append({"Kỳ": period, "Đơn vị": code,
                                 _SEC_COL[section]: round(cuoi / 1e9, 9)})
            continue
        if section == "tien vay" and c3 and not _CHILD_RE.search(_norm(c3)):
            dau = _num(r[c_dau]) if c_dau < len(r) else None
            cuoi = _num(r[c_cuoi]) if c_cuoi < len(r) else None
            if dau is None and cuoi is None:
                continue
            if company and company[1]:          # bank ĐỨNG SAU công ty của nó (layout thường)
                rec = {"Kỳ": period, "Đơn vị": company[1], "Ngân hàng": c3}
                if dau is not None:
                    rec["Dư nợ đầu kỳ (tỷ)"] = round(dau / 1e9, 9)
                if cuoi is not None:
                    rec["Dư nợ cuối kỳ (tỷ)"] = round(cuoi / 1e9, 9)
                vay_recs.append(rec)
            else:                               # bank ĐỨNG TRƯỚC công ty (block TC) -> chờ subtotal
                orphans.append({"bank": c3, "dau": dau, "cuoi": cuoi})

    out = {"ok": True, "unresolved_companies": sorted(unresolved)}
    if dropped_orphans:
        out["dropped_orphan_banks"] = sorted(set(dropped_orphans))
    src_name = SC.source_id_from_path(path)
    if sdt_recs:
        p = os.path.join(tf.FILLED_DIR, f"SDT_{period}_03B_SODU_TIEN.xlsx")
        tf.fill("03B_SODU_TIEN", sdt_recs, p)
        imp = tf.import_filled(p, cong_ty=None, source_file=src_name)
        out["sdt"] = {"rows": imp.get("rows_imported"), "by_type": imp.get("by_type")}
    if vay_recs:
        p = os.path.join(tf.FILLED_DIR, f"VAY_{period}_04_VAY.xlsx")
        tf.fill("04_VAY", vay_recs, p)
        imp = tf.import_filled(p, cong_ty=None, source_file=src_name)
        out["vay"] = {"rows": imp.get("rows_imported"), "by_type": imp.get("by_type")}
    if not sdt_recs and not vay_recs:
        return {"ok": False, "error": "Không bóc được dòng SDT/VAY nào"}
    return out


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--period", required=True)
    a = ap.parse_args()
    import json
    print(json.dumps(extract(a.file, a.period), ensure_ascii=False))
