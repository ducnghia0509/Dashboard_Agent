# -*- coding: utf-8 -*-
"""ENGINE TRÍCH XUẤT THEO SPEC — thay cho việc viết 1 file Python cho mỗi nguồn báo cáo.

VÌ SAO CÓ FILE NÀY (2026-08-08): mọi deriver trước đây (`derive_hqkd_ngay.py`,
`derive_congno_tuoino.py`, `derive_vhkd_kqkd.py`…) đọc cột theo CHỈ SỐ CỨNG. Kế toán chèn/xoá 1 cột
là sai số im lặng, và mỗi lần mapping đổi là một vòng sửa Python → test → deploy. Đợt VHKD/XDV có
~10 nguồn nên chi phí đó nhân lên. Engine này đọc một **spec JSON** mô tả nguồn và ánh xạ cột, nên
đổi mapping = sửa JSON, không đụng code.

NGUYÊN TẮC SỐ 1 — DÒ CỘT THEO **TÊN HEADER**, KHÔNG THEO CHỮ CỘT.
Riêng điều này diệt cả một lớp lỗi đã gặp thật: file mapping ghi kho xe B2B "Trạng thái >30/45
ngày" ở cột AK, file thật để ở AM (AK là "Số ngày tồn kho thực tế"); kho xe B2C ghi "Tên KH" cột O,
thật ra ở P. Có thể khai `"cot_du_phong": "AM"` làm phao khi header rỗng, nhưng tên luôn thắng.

SPEC ĐẶT Ở: `Dashboard_Agent/extract_specs/<id>.json` (KHÔNG để trong `memory/` — thư mục đó
bị .gitignore vì là state runtime của agent; spec là CẤU HÌNH NGUỒN, phải được version). Đây là file CẤU HÌNH TIN CẬY của nội
bộ (không nhận từ ngoài) — `dan_xuat` dùng eval trong sandbox hẹp; đừng mở cho input người dùng.

CẤU TRÚC SPEC (khoá tiếng Việt cho kế toán/BA đọc được):
{
  "id": "vhkd_kqkd",                     // trùng tên file
  "report_type": "KDVH",                 // ghi vào raw_rows.report_type
  "row_index_base": 6400000,             // dải row_index riêng, tránh đụng deriver khác
  "khoi": "Khối KD Vinfast - Showroom",  // hằng số (hoặc bỏ, nếu suy từ cost_center)
  "nguon": {
    "folder": "SRVF/baocaokqkd",         // dưới received_reports/
    "file_glob": "*.xlsx",
    "sheet": {"batdau": "CHI TIẾT XHĐ"}  // | {"ten": "..."} | {"chua": "..."} | {"so": 0}
                                        // | {"theo_thang": "T{mm}"} — file 1 sheet/tháng (BCTC SRVF)
  },
  "header": {"dong": 1},                 // | {"dong": [1,2]} gộp 2 dòng header (lấy ô đầu khác rỗng)
                                         // | {"tim_o": "Mã số"} tự dò dòng header theo nhãn mốc
                                         //   (dùng khi vị trí header khác nhau giữa các sheet)
  "dong_bat_dau": 3,                     // tuỳ chọn — mặc định = dòng header cuối + 1
  "dong_ket_thuc": 19,                   // tuỳ chọn — chặn trên của dải dòng (bắt buộc khi dùng `vung`)
  "nam_tu_ten_file": {"regex": "\\.M\\.(\\d{4})\\."},   // cho `kieu: "thang_cuoi"`
  "vung": [                              // NHIỀU KHỐI trong CÙNG 1 sheet (xem `extract_file`).
    {"ten": "Sơn Tây", "header": {"dong": [6, 7]}, "dong_bat_dau": 8, "dong_ket_thuc": 19,
     "chieu_co_dinh": {"cost_center": "ST_AT"}, "cot": {"amount": {"header": "TXTX"}}}
  ],                                     // khoá trong vùng ghi đè spec; `cot`/`chieu_co_dinh` trộn
  "chieu_tu_ten_file": {"dim2": {"regex": "Xuathoadon_(B2B|B2C|GF)", "hoa": true},
                        // `chuan_hoa` chạy trên chính tên bắt được; hook trả dict thì trộn cả dict
                        "cost_center": {"regex": "BCDau(\\w+)", "chuan_hoa": "cc_qlts"}},
  "ngay_tu_ten_file": {"regex": "M\\.(\\d{4})\\.(\\d{1,2})\\.(\\d{1,2})", "thu_tu": "ymd"},
  "cot": {                               // đích -> cách lấy. Đích: ngay/cost_center/cong_ty/
    "cost_center": {"header": "Tên DVCS", "chuan_hoa": "sr_showroom"},   // amount/amount2/dim1..3/
    "ngay":   {"header": "Ngày hóa đơn", "kieu": "date"},                // payload.<khoá bất kỳ>
    "amount": {"header": "Giá bán", "kieu": "so", "he_so": 1e-9}
  },                                     // `kieu` khác: "thang_cuoi" (ô là SỐ THÁNG 1..12, năm
                                         // lấy từ tên file -> ngày cuối tháng) · "ngay_trong_thang"
                                         // · "ngay_dd_mm" (ô ghi "01/09 (Thứ 3)" — năm từ tên
                                         // file, tháng trong ô phải khớp tháng của file)
                                         // (ô là SỐ NGÀY, năm+tháng từ tên file). Hai kiểu này cho
                                         // báo cáo xếp mỗi kỳ MỘT DÒNG thay vì một cột.
  "cot_thang": {"tu": "E", "den": "J", "he_so": 1.0,
                "kieu": "nguong"},       // tuỳ chọn — ô là ngưỡng có toán tử ("≥95%", "<4%"):
                                         // amount = số, payload.toan_tu = '>='/'<='/'<'/'='
  "cot_ngay_dau_thang": "Kỳ này",       // bản ĐẦU THÁNG không có gì để trừ -> đọc cột này thay
                                        // (chỉ có nghĩa khi đi kèm `tru_ngay_truoc`)
  "doi_gia_tri": {"dim1": {"A100": "1000", "A300": "1047"}},   // đổi TÊN giá trị khi đổi nguồn
                                        // mà phải giữ hình dạng bản ghi cũ; mã ngoài bản đồ GIỮ
                                        // NGUYÊN, muốn lọc thì dùng `loc` điều kiện `thuoc`.
  "tinh_han_no": {"ngay_hoa_don": "payload.ngay_hoa_don", "cong_ngay": 15,   // suy đến hạn +
                  "den_han": "payload.den_han",            // số ngày quá hạn cho nguồn công nợ
                  "so_ngay_qua_han": "payload.so_ngay_qua_han", "phan_loai": "dim1"},
  "chi_lay_ngay_cua_file": true,         // bỏ dòng có `ngay` khác ngày suy từ TÊN FILE — cho nguồn
                                         // ngày cho lẫn sang hôm sau (xem chỗ dùng bên dưới)
  "tru_ngay_truoc": true,                // ô nguồn là LUỸ KẾ từ đầu tháng -> lấy hiệu với file
                                         // ngày trước để ra số CỦA RIÊNG NGÀY (xem `_tru_ngay_truoc`)
  "ban_ghi": "moi_dong",                 // | "moi_cot_gia_tri" | "moi_cot_ngay"
                                         // | "moi_cot_thang" (xem dưới)
  "cot_gia_tri": [                       // chỉ dùng khi ban_ghi = "moi_cot_gia_tri":
    {"header": "Công nợ trong hạn", "dim1": "Trong hạn", "he_so": 1e-9},
    {"header": "XDV Ocean Park", "cost_center": "OCP_XDV", "he_so": 1e-9},
    {"cot": "R", "dim1": "App An", "he_so": 1e-9,
     "amount2": {"cot": "M"}}            // tuỳ chọn — ĐO THỨ HAI của cùng cột giá trị, ghi vào
  ],                                     // amount2 (nhận "cot"/"header"/"he_so" như khai báo cột
                                         // thường). Cần khi một chiều mang HAI số đi liền nhau
                                         // (doanh thu + số cuốc của mỗi kênh An Taxi) và bản
                                         // NGÀY của cùng chiều đó đã dùng amount/amount2.
                                         // -> mỗi dòng nguồn đẻ N bản ghi, amount lấy từng cột;
                                         // ngoài dim1..3 còn gán được cost_center/cong_ty/khoi
                                         // khi CHIỀU ĐƠN VỊ nằm ở cột (mỗi xưởng một cột)
  "loc": [{"cot": "ngay", "dieu_kien": "khac_rong"}],
  "dan_xuat": {"payload.lng": "amount - payload.gia_von"},
  "payload_them": {"unit": "ty"}
}

`ban_ghi`:
  - "moi_dong"        : 1 dòng nguồn -> 1 bản ghi (bảng chi tiết: xuất hoá đơn, claim, kho xe).
  - "moi_cot_gia_tri" : 1 dòng nguồn -> N bản ghi, mỗi cột giá trị thành 1 dòng có dim1 riêng
                        (bảng ma trận: "Tuổi nợ phải thu" có sẵn cột trong hạn/1-30/>30-90/…).
  - "moi_cot_ngay"    : 1 dòng nguồn -> N bản ghi theo dải cột NGÀY trong tháng (`cot_ngay`).
  - "moi_cot_thang"   : 1 dòng nguồn -> N bản ghi theo dải cột THÁNG (`cot_thang`), mỗi bản ghi
                        neo vào ngày cuối tháng đó. Dùng cho bản KẾ HOẠCH năm: một file duy nhất
                        cấp số cho cả 12 kỳ, nên đừng đặt kỳ theo tên file.

Ghi DB idempotent theo `source_file` = "<FOLDER>::<tên file>" — cùng quy ước mọi deriver khác.
`--write` mới ghi; mặc định dry-run in ra tổng hợp để đối chiếu với file gốc.

HOOK (khi spec không tả nổi): thêm hàm vào `_CHUAN_HOA` bên dưới rồi gọi bằng tên trong
`"chuan_hoa"`. Hook trả str, hoặc dict để set nhiều trường cùng lúc (vd cost_center + cong_ty).
"""
import argparse
import calendar
import datetime as dt
import fnmatch
import glob
import json
import os
import re
import sys
import unicodedata

import openpyxl
import psycopg
from openpyxl.utils import column_index_from_string, get_column_letter

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.normpath(os.path.join(_HERE, ".."))
sys.path.insert(0, _ROOT)

from dotenv import load_dotenv  # noqa: E402

from servers.common import dataset_ky as _DSK  # noqa: E402

load_dotenv(os.path.join(_ROOT, ".env"))

DB_URL = os.environ.get("DATABASE_URL")
SPEC_DIR = os.path.join(_ROOT, "extract_specs")
REPORTS_DIR = os.path.normpath(os.path.join(_ROOT, "..", "Connect_VPS", "received_reports"))

_CHUAN_TRUONG = {"ngay", "cong_ty", "khoi", "cost_center", "amount", "amount2",
                 "dim1", "dim2", "dim3"}


# ─────────────────────────── tiện ích chuẩn hoá ───────────────────────────
def _nd(s):
    """Bỏ dấu + bỏ ký tự không phải chữ/số + thường hoá — dùng để so tên header/tên đơn vị."""
    s = str(s if s is not None else "").strip().lower().replace("đ", "d")
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]", "", s)


# Giá trị LỖI của Excel: VLOOKUP chưa tra được, ô tham chiếu bị xoá, chia 0… File nguồn vẫn gửi
# về nguyên trạng chuỗi này. Coi như Ô TRỐNG ở MỌI spec — nếu không, chuỗi "#N/A" đi thẳng vào
# dim/cost_center và trở thành một "giá trị nghiệp vụ" giả: đã dính 14/08/2026 ở claim B2C T8
# (cột Trạng thái 178 ô #N/A, cột Số tiền hỏng CẢ 605 ô) -> dashboard đếm 175 hồ sơ trạng thái
# "#N/A" giá trị 0đ và đòi bổ sung "#N/A" vào bảng quy ước trạng thái claim.
# Số ô lỗi được ĐẾM và báo ra `canh_bao` (xem cuối `extract_file`): ô lỗi là dấu hiệu file nguồn
# chưa cập nhật xong, phải nhìn thấy chứ không nuốt im lặng.
_LOI_EXCEL = {"#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!", "#GETTING_DATA"}


def _o_loi(v):
    return isinstance(v, str) and v.strip().upper() in _LOI_EXCEL


# Câu cảnh báo "đã đọc được file nhưng bộ lọc loại hết dòng" — `run()` dựa vào nó để phân biệt
# "file hết số" với "file không đọc được" (xem chú thích ở `run`). Một chỗ khai, hai chỗ dùng.
_W_BO_LOC = "dòng không qua bộ lọc"


def _so(v, he_so=1.0):
    if isinstance(v, bool):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v) * he_so
    s = str(v or "").strip().replace(" ", "")
    if not s:
        return 0.0
    # '1.234.567,89' (VN) và '1,234,567.89' (EN) đều gặp trong file kế toán
    s = s.replace(".", "").replace(",", ".") if re.search(r",\d{1,2}$", s) else s.replace(",", "")
    try:
        return float(s) * he_so
    except ValueError:
        return 0.0


def _so_co_rong(v, he_so=1.0):
    """Như `_so` nhưng ô TRỐNG/không đọc được -> None, KHÔNG phải 0.0.

    Dùng cho payload mà "0" là một giá trị có nghĩa thật, khác hẳn "chưa biết". Ca cụ thể (14/08/
    2026): 'Số ngày tồn thực tế' của file tồn kho GF — 12/62 xe ghi 'Ngày xe về SR' = 'Chưa' nên ô
    tuổi tồn là #VALUE!. Qua `_so` thì thành 0.0, tức 12 xe đó đọc thành "mới về hôm nay, tồn 0
    ngày" và lọt vào nhóm dưới 30 ngày -> tỷ lệ quá hạn TỰ HẠ, không cảnh báo gì. Còn xe về đúng
    ngày chốt thì 0 là số THẬT, nên không thể lấy `if not v` mà suy ra "chưa biết".
    ĐỪNG dùng cho `amount`: tầng đọc số cộng tiền bằng `r["amount"] or 0`, None ở đó vô hại nhưng
    cũng không thêm thông tin gì; giữ `so` cho amount để không đổi hành vi các spec đang chạy.
    """
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v) * he_so
    s = str(v).strip().replace(" ", "")
    s = s.replace(".", "").replace(",", ".") if re.search(r",\d{1,2}$", s) else s.replace(",", "")
    try:
        return float(s) * he_so
    except ValueError:
        return None


# Ô NGƯỠNG: nghiệp vụ đăng ký mục tiêu vận hành bằng chuỗi có TOÁN TỬ — "≥95%", "≤10%", "<4%",
# "<5". `_so` trả 0.0 cho những ô này nên `moi_cot_thang` bỏ qua chúng IM LẶNG: cả cụm chỉ tiêu
# chất lượng của bản kế hoạch An Taxi/XVP/Trạm sạc (vận doanh, hủy chuyến, hoàn thành trên app,
# xe nằm, bàn giao đúng hạn) rơi hết vào đây. Trước 18/08/2026 backend phải gõ cứng mấy ngưỡng
# này từ file mapping (`antaxi.py::KPI_*`) — và mapping ghi KHÁC bản đăng ký kế hoạch (xe nằm 5%
# vs 4%, hủy cuốc 15% vs 10%), tức số trên màn không phải chỉ tiêu nghiệp vụ đã ký.
# KHÁC "ngưỡng ghi bằng LỜI" ("Nhỏ hơn 30% tồn kho", "Bằng 0 - cứ có 1 khách hàng quá hạn trừ
# 5%…"): những ô đó KHÔNG parse được và vẫn phải chốt số với nghiệp vụ, `_nguong` trả None để
# chúng tiếp tục bị bỏ chứ không đoán bừa.
_RE_NGUONG = re.compile(r"^(>=|<=|=>|=<|>|<|≥|≤|=)?\s*([\d.,]+)\s*(%)?$")
_TOAN_TU = {"≥": ">=", "=>": ">=", "≤": "<=", "=<": "<=", "": "=", None: "="}


def _nguong(v, he_so=1.0):
    """Ô ngưỡng -> (giá trị, toán tử) hoặc None nếu không phải ngưỡng số.

    Toán tử được GIỮ LẠI (payload.toan_tu) chứ không quy hết về một chiều: "≥95%" và "<4%" chấm
    đèn ngược nhau, mất toán tử là giao diện phải đoán chiều theo tên chỉ tiêu — đúng kiểu suy
    diễn sinh lỗi im lặng khi nghiệp vụ đổi mục tiêu.
    """
    if isinstance(v, bool) or v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v) * he_so, "="
    s = str(v).strip().replace(" ", "").replace("\u00a0", "")
    if not s:
        return None
    m = _RE_NGUONG.match(s)
    if not m:
        return None
    # "%" trong ô chỉ là ĐƠN VỊ, không quy đổi: file ghi "≥95%" thì ngưỡng là 95 (đơn vị %),
    # đúng thang mà backend/FE đang dùng cho mọi tỷ lệ. Nhân/chia 100 ở đây là chỗ dễ lệch nhất.
    return _so(m.group(2)) * he_so, _TOAN_TU.get(m.group(1), m.group(1))


def _date(v):
    """-> 'YYYY-MM-DD' hoặc None. File có cả datetime, 'dd/mm/yyyy' lẫn 'yyyy-mm-dd'."""
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    s = str(v or "").strip()
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        d, mo, y = (int(x) for x in m.groups())
    else:
        m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)
        if not m:
            return None
        y, mo, d = (int(x) for x in m.groups())
    try:
        return dt.date(y, mo, d).isoformat()
    except ValueError:
        return None


# ─────────────────────────── hook chuẩn hoá ───────────────────────────
_CC_CACHE = {}


def _cc_showroom(ten):
    """Tên đơn vị trong file -> {cost_center, cong_ty} của khối Vinfast - Showroom.

    File ghi "Showroom OceanPark"/"Showroom Uông Bí", master ghi "Vinfast Ocean Park" -> bỏ tiền
    tố showroom/vinfast ở CẢ hai phía rồi so. Bản "(61)" (pháp nhân Xanh Vĩnh Phúc) chuẩn hoá
    thành "...61" nên không đụng bản gốc. "Vinfast B2B" là đội bán B2B tập trung -> B2B_SR.
    """
    return _cc_theo_khoi(ten, "Khối KD Vinfast - Showroom", {"b2b": ("B2B_SR", "TC")})


def _cc_hcns_xdv(ten):
    """Phòng/Ban của báo cáo NHÂN SỰ -> cost center XDV, CHỈ khi tên đó THỰC SỰ là một XƯỞNG.

    VÌ SAO KHÔNG DÙNG THẲNG `_cc_xdv` (bắt được 14/08/2026): `_bo_tien_to` bóc cả tiền tố
    "showroom", nên "Showroom Vinfast Smart City" quy về cùng khoá "smartcity" với xưởng
    "Vinfast Smart City" -> CẢ 9 SHOWROOM bị gán cost center XDV, cộng thêm 470 nhân sự BÁN HÀNG
    vào mẫu số "nhân sự xưởng" (994 thay vì 524, phồng 90%). Các spec XDV khác không dính vì file
    của chúng chỉ có xưởng; báo cáo nhân sự thì chứa MỌI khối trong một sheet.

    Báo cáo nhân sự gọi xưởng bằng đúng hai dạng: "Xưởng dịch vụ ..." và "Trung tâm Sửa chữa Pin,
    Động cơ ..." (xưởng HCM). Chốt bằng tiền tố đó là tách được Showroom khỏi xưởng.

    CÒN MỘT CA KHÔNG CHỐT ĐƯỢC Ở ĐÂY: "Xưởng dịch vụ Hà Khánh" xuất hiện ở CẢ khối "Khối Dịch vụ
    Hậu mãi" (33 người) và "Khối Kinh doanh Showroom Vinfast" (3 người) — hook chỉ thấy MỘT ô nên
    không biết dòng đang thuộc khối nào. Bên đọc PHẢI lọc thêm `dim1 = "Khối Dịch vụ Hậu mãi"`
    (xem `app/metrics/xdv.py::_nhan_su_xuong`), nếu không cộng đôi 3 người.
    """
    n = _nd(ten)
    if not (n.startswith("xuongdichvu") or n.startswith("trungtamsuachua")):
        return {"_khong_map": str(ten or "").strip()}
    return _cc_xdv(ten)


def _cc_xdv(ten):
    """Tên xưởng trong file -> cost center khối XDV. 13/14 xưởng khớp thẳng sau khi bỏ tiền tố
    ("Ocean Park" ↔ master "Vinfast Ocean Park"); riêng xưởng HCM có tới BA tên gọi nên phải
    khai alias: master "Vinfast Hồ Chí Minh", bản THỰC HIỆN ghi "HCM", bản KẾ HOẠCH
    (Baocaodoanhthukehoachngay, 10/08/2026) ghi theo quận — "Quận 12"."""
    # Khoá alias phải viết theo dạng ĐÃ CHUẨN HOÁ của `_nd` — bỏ dấu, bỏ cả khoảng trắng:
    # "Quận 12" -> "quan12". Viết "quan 12" thì không bao giờ khớp và xưởng đó lặng lẽ mất.
    # Alias thứ BA cho xưởng HCM (14/08/2026): báo cáo NHÂN SỰ (HCNS/baocaotongsonhansu, sheet
    # "Chi tiết") gọi nó là "Trung tâm Sửa chữa Pin, Động cơ Thành phố Hồ Chí Minh". Không khai
    # thì đúng 1 trong 14 xưởng rơi và mẫu số "nhân sự xưởng" hụt 35 người (~6,7%) — kiểu thiếu
    # vừa đủ nhỏ để không ai thấy.
    # Alias thứ TƯ (04/09/2026): bộ nguồn TỰ ĐỘNG của Cyber (TEST_XDV/DV01, DV02, DV41, DV42,
    # DV44) ghi đủ pháp nhân — "Chi nhánh Vinfast Hồ Chí Minh- Công ty CP Thịnh Cường" (chú ý
    # thiếu dấu cách trước gạch nối). Thiếu alias thì HCM là xưởng DUY NHẤT rơi khỏi mọi nguồn
    # ngày mới, mà tổng khối vẫn trông hợp lý nên rất khó thấy.
    return _cc_theo_khoi(ten, "Khối KD Vinfast - XDV",
                         {"hcm": ("HCM_XDV", "TC"), "quan12": ("HCM_XDV", "TC"),
                          "trungtamsuachuapindongcothanhphohochiminh": ("HCM_XDV", "TC"),
                          "chinhanhvinfasthochiminhcongtycpthinhcuong": ("HCM_XDV", "TC")})


def _master_loader():
    """Nạp `app/master_data/loader.py` của backend THEO ĐƯỜNG DẪN FILE, không theo tên package.

    Vì sao không `sys.path.insert(...)` rồi `from app.master_data import loader` như trước: khi
    hàm này chạy TRONG tiến trình đã import `servers/template_filler.py` (đường "Đồng bộ & nạp" và
    bước nạp lại của theo-dõi-thay-đổi), `app` ĐÃ nằm trong sys.modules và trỏ vào backend CŨ
    (`DashBoard_AI/backend`, theo BACKEND_PATH) — bản đó không có `master_data`, nên import ném
    ModuleNotFoundError và toàn bộ spec có chuẩn hoá cost center trả 0 dòng. Chạy CLI thì lại
    không lỗi (chưa ai import `app`), nên bug chỉ hiện khi bấm nút trên web — đúng kiểu khó tìm.
    Nạp theo đường dẫn file thì tên `app` của ai cũng không ảnh hưởng.

    loader.py chỉ dùng stdlib và đọc JSON trong `app/data/` nên nạp rời hoàn toàn an toàn.
    """
    import importlib.util
    ung_vien = [
        os.environ.get("MASTER_DATA_BACKEND"),
        os.path.join(_ROOT, "..", "AI_coding", "tc-admin-api"),      # bản test
        os.path.join(os.path.expanduser("~"), "apps/tc-console/tc-admin-api"),   # bản prod
    ]
    for base in ung_vien:
        if not base:
            continue
        p = os.path.normpath(os.path.join(base, "app", "master_data", "loader.py"))
        if not os.path.isfile(p):
            continue
        spec = importlib.util.spec_from_file_location("tc_master_data_loader", p)
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        return mod
    # Đường cũ làm lưới cuối: có thể chạy ở nơi bố cục thư mục khác hẳn dự đoán ở trên.
    sys.path.insert(0, os.path.join(_ROOT, "..", "AI_coding", "tc-admin-api"))
    from app.master_data import loader as master   # noqa: PLC0415
    return master


def _cc_theo_khoi(ten, khoi, alias):
    if khoi not in _CC_CACHE:
        master = _master_loader()
        m = {}
        for cc in master.master_data().get("costCenters", []):
            if (cc.get("khoi") or "") != khoi:
                continue
            m[_bo_tien_to(_nd(cc.get("ten")))] = (str(cc.get("ma") or "").strip(),
                                                 master.resolve_company_code(cc.get("congTy") or ""))
        _CC_CACHE[khoi] = m
    key = _bo_tien_to(_nd(ten))
    got = _CC_CACHE[khoi].get(key) or alias.get(key)
    if not got and "_" in str(ten or ""):
        # BIẾN THỂ "<đơn vị>_ HL" / "_ CP" (24/08/2026) — nhóm bán hàng nội bộ của CÙNG một
        # showroom, không phải đơn vị khác. Bắt được ở báo cáo công nợ SRVF T5/T6: 38 dòng dư nợ
        # hợp đồng thật (14,10 tỷ) mang tên "Showroom Uông Bí_ HL"/"_ CP" -> không khớp danh mục
        # -> `giu_khi_khong_map` giữ dòng nhưng cost_center/cong_ty rỗng. Hậu quả không phải mất
        # tổng (tổng khối vẫn đủ) mà là chúng thành MỘT NHÓM SNAPSHOT RIÊNG: `_rows` chọn ngày
        # chốt theo (cong_ty, khoi), nhóm rỗng đứng mãi ở 30/06 nên view nhiều kỳ cộng số dư T08
        # với phần sót T06 (38,94 thay vì 36,95 tỷ quá hạn) — sai mà không có dấu hiệu nào.
        #
        # Chúng thuộc Uông Bí, kiểm bằng chính file đó: mọi sheet TỔNG HỢP ("CN theo đơn vị",
        # "Tuổi nợ phải thu", "số dư nợ chuẩn") chỉ ghi "Showroom Uông Bí", và tra mã hợp đồng
        # của các dòng biến thể trong sheet "Tuổi nợ phải thu" đều ra đơn vị "Showroom Uông Bí".
        #
        # CẮT TRÊN TÊN THÔ, KHÔNG PHẢI SAU `_nd`: `_nd` xoá mọi ký tự không phải chữ/số nên dấu
        # `_` biến mất, và khớp kiểu tiền tố sau đó là gán nhầm pháp nhân — "Vinfast Cẩm Phả (61)"
        # thành `campha61`, khớp tiền tố `campha` của Thịnh Cường trong khi (61) là Xanh Vĩnh Phúc.
        # Tên thô của bản (61) không có `_` nên nhánh này không đụng tới nó.
        #
        # An toàn vì: (1) chỉ chạy SAU khi đã trượt -> không đổi được map nào đang đúng;
        # (2) không cost center nào của master có `_` trong TÊN (mã thì có: UB_SR) -> cắt tại `_`
        # không thể chặt mất tên thật; (3) tên thô vẫn nằm nguyên ở payload để đối chiếu.
        goc = _bo_tien_to(_nd(str(ten).split("_")[0]))
        if goc:
            got = _CC_CACHE[khoi].get(goc) or alias.get(goc)
    ma, cty = got or (None, None)
    # TRẢ LUÔN `khoi` (14/08/2026). Trước đây chỉ trả cost_center + cong_ty, và mọi spec dùng hook
    # này đều tự khai `"khoi"` ở cấp spec nên không ai thấy thiếu. Nhưng báo cáo NHÂN SỰ chứa NHIỀU
    # khối trong một sheet -> không khai được `khoi` cố định, mà thiếu `khoi` thì `filter_scope`
    # (màn xdv0 tự lọc khối) LỌC SẠCH các dòng đó và mẫu số "nhân sự xưởng" về None mà không có lỗi
    # nào. Giá trị trả về TRÙNG KHỚP với `khoi` mà các spec cũ vẫn khai, nên chúng không đổi hành vi.
    return ({"cost_center": ma, "cong_ty": cty, "khoi": khoi} if ma
            else {"_khong_map": str(ten or "").strip()})


def _bo_tien_to(n):
    # "thinhcuong": bản công nợ tuần 10/08/2026 đổi cách ghi tên xưởng, từ "Ocean Park" sang
    # "Vinfast Thịnh Cường Ocean Park" -> bỏ "vinfast" xong vẫn còn "thinhcuong" nên KHÔNG khớp
    # danh mục, cả 14 xưởng rơi hết và file nạp ra 0 dòng. Không cost center nào của master bắt
    # đầu bằng "Thịnh Cường" (mấy dòng Xe tải để ở ĐUÔI) nên bỏ tiền tố này là an toàn.
    # BÓC LẶP, không phải một lượt (sửa 14/08/2026). Vòng `for` một lượt xét tiền tố theo THỨ TỰ
    # khai, nên "Xưởng dịch vụ Vinfast Cẩm Phả" bóc được "xuongdichvu" rồi DỪNG, còn lại
    # "vinfastcampha" trong khi khoá master là "campha" -> 13/14 xưởng của báo cáo nhân sự HCNS
    # rơi hết. Bóc lặp thì cả hai vế (tên trong file và tên master) cùng quy về một dạng, và vì
    # danh mục cost center được chuẩn hoá bằng CHÍNH hàm này nên không sinh khoá lệch.
    doi = True
    while doi:
        doi = False
        for p in ("showroom", "vinfast", "thinhcuong", "xuongdichvu", "xdv"):
            if n.startswith(p) and len(n) > len(p):    # giữ nguyên nếu bóc xong thành rỗng
                n, doi = n[len(p):], True
    return n


# Nhãn dòng của sheet KHDT (bản kế hoạch) -> (mã nhóm, kênh). Khớp CHÍNH XÁC sau chuẩn hoá, KHÔNG
# khớp kiểu "chứa": ngay dưới các dòng chi tiết còn có dòng tổng "Tổng doanh thu kênh B2C" —
# khớp kiểu chứa là cộng đôi toàn bộ kế hoạch mà không có dấu hiệu gì.
# Mã nhóm theo ĐÚNG file kế hoạch: A230 = doanh thu khác, A250 = claim (thiết kế
# Dashboard.dc.html ghi ngược hai mã này — đã chốt với nghiệp vụ 2026-08-10, lấy theo file).
_KH_DONG = {}
for _ma, _nhan in (("A200", "Bán xe kênh {}"),
                   ("A230", "Doanh thu khác Kênh {}"),
                   ("A250", "Claim kênh {}")):
    for _k in ("B2C", "B2B", "GF"):
        _KH_DONG[_nd(_nhan.format(_k))] = (_ma, _k)


def _kh_dong(nhan):
    """Nhãn dòng kế hoạch -> {dim1: mã nhóm, dim2: kênh}. Dòng tổng/tiêu đề -> không đặt gì
    (bản ghi thiếu dim2 sẽ bị `loc` loại), nên chỉ các dòng CHI TIẾT theo kênh được giữ."""
    got = _KH_DONG.get(_nd(nhan))
    return {"dim1": got[0], "dim2": got[1]} if got else {}


# Dòng CHỈ TIÊU TỔNG của bản kế hoạch XDV (sheet "KHDT XHD" mục IV, V): nhãn nằm NGAY ở cột B của
# chính dòng có số, không có dòng tiêu đề riêng như mục I.1/I.2/II/III — nên không dùng
# `ngu_canh_dong` được (dòng tiêu đề không tự sinh bản ghi).
# THỨ TỰ QUAN TRỌNG: "kehoachtongchiphi" là TIỀN TỐ của "kehoachtongchiphivanhanh…", so kiểu
# "chứa" mà xét CP_TONG trước là mục V bị gán nhầm thành CP_TONG rồi hai dòng đè nhau.
_XDV_KH_CHI_TIEU = (
    ("kehoachtongchiphivanhanh", "CP_VH"),
    ("kehoachtongchiphi", "CP_TONG"),
)


def _xdv_kh_dong(nhan):
    """Cột B bản kế hoạch XDV -> chỉ tiêu tổng (mục IV/V) HOẶC xưởng.

    `dim2` là cờ phân tầng, BẮT BUỘC có để `loc` giữ được đúng dòng thật:
      · "Tổng"  — dòng cấp khối (mục IV/V, và dòng "Cộng" bắt ở cột A bởi `_xdv_kh_cong`)
      · "Xưởng" — dòng chi tiết theo cost center
    Không có cờ này thì mấy ô rác lạc giữa sheet (vd ô "1" ở H45 sheet "KHDT RO") vẫn mang dim1
    thừa hưởng từ ngữ cảnh mục đang mở và lọt vào DB thành một dòng kế hoạch vô nghĩa.
    """
    n = _nd(nhan)
    for key, ma in _XDV_KH_CHI_TIEU:
        if key in n:
            return {"dim1": ma, "dim2": "Tổng"}
    res = _cc_xdv(nhan)
    if res.get("cost_center"):
        res["dim2"] = "Xưởng"
    return res


# Sheet "CT Vận hành" của bản kế hoạch XDV — NGƯỠNG MỤC TIÊU vận hành T7..T12, nhãn ở cột B.
# CHỈ khai 5 chỉ tiêu KHÔNG trùng nguồn khác. 6 dòng đầu sheet (doanh thu, sản lượng, lợi nhuận
# gộp, LNST, doanh số & sản lượng lệnh nghiệm thu) lặp lại y hệt số của "KHDT XHD"/"KHDT RO" đã
# nạp thành XDV_KH*/XDV_KH_RO* — nạp thêm là tạo hai nguồn cho cùng một con số, sớm muộn lệch nhau.
# Khớp kiểu "chứa" sau `_nd`, nên chịu được cả lỗi gõ "Gía trị" lẫn đuôi giải thích dài trong ngoặc.
_XDV_CT_DONG = (
    ("giatrilenhnghiemthuchuaxuathoadon", "RO_CHUA_XHD"),
    ("lenhnghiemthuchuaxuathoadon30ngay", "RO_CHUA_XHD_30"),
    ("lenhsuachua10ngaychuanghiemthu", "LSC_10N"),
    ("giatribinhquandonhang", "BQ_DON"),
    ("nangsuatld", "NSLD"),
)


def _xdv_ct_dong(nhan):
    """Cột B sheet "CT Vận hành" -> mã ngưỡng. Dòng không khai -> {} (giữ nguyên, `loc` sẽ loại).

    Trả `{}` chứ KHÔNG trả None: hook trả dict thì engine gộp khoá rồi bỏ qua, nên dòng lạ không
    bị ghi đè dim1 thành None và cũng không sinh cảnh báo "không map được" vô nghĩa.
    """
    n = _nd(nhan)
    for key, ma in _XDV_CT_DONG:
        if key in n:
            return {"dim1": ma, "dim2": "Tổng"}
    return {}


# Sheet "CT Vận hành" của bản kế hoạch SHOWROOM (`1.SR.*.Kehoachthang.xlsx`) — KHÁC HẲN sheet
# cùng tên của bản XDV, đừng dùng lại `_xdv_ct_dong`: bản SR không có cột "Công thức tính" nên dải
# tháng lùi một cột (E..J thay vì F..K), và 10 nhãn dòng hoàn toàn khác.
# CHỈ khai 2 dòng NĂNG SUẤT. Tám dòng còn lại cố ý bỏ, mỗi dòng một lý do:
#   - "Doanh thu" (CT33) và "Sản lượng" (CT34) là công thức `=KHDT!F16` / `=KHDT!F21` — trỏ thẳng
#     vào sheet KHDT đã nạp thành VHKD_KH/VHKD_KH_SL. Đã đối chiếu T7: CT33 = 823.271.212.048 =
#     đúng tổng A200 ba kênh, CT34 = 1.414 = đúng tổng sản lượng ba kênh. Nạp thêm là hai nguồn
#     cho một con số.
#   - "Gía trị bình quân đơn hàng" (CT36) là `=D5/D7`, `app/metrics/vhkd.py` đã tự suy ra.
#   - "Lợi nhuận" bỏ trống toàn bộ trong file.
#   - 4 dòng cuối (xe tồn > 30 ngày, claim chưa xác nhận, claim chưa làm hồ sơ, công nợ quá hạn
#     đã có COC) ghi ngưỡng BẰNG LỜI ("Nhỏ hơn 30% tồn kho", "Bằng 0 - cứ có 1 khách hàng quá
#     hạn, trừ 5% mức độ hoàn thành") -> `_so` ra 0 nên `moi_cot_thang` tự bỏ. CHỦ Ý: muốn chấm
#     %HT cho 4 nhóm đó thì phải chốt ngưỡng thành SỐ với nghiệp vụ trước, không phải đoán ở đây.
# MÃ Ở CỘT A CỦA FILE ĐỀ LỆCH MAPPING — bám nhãn dòng, không bám mã: mapping ghi năng suất sản
# lượng = CT37 và giá trị BQ đơn hàng = CT38, còn file ghi CT38 và CT36. `app/metrics/vhkd.py`
# đang theo mã của FILE, nên hook cũng gắn mã theo file.
_SR_CT_DONG = (
    ("nangsuatdoanhthu", "NS_DT"),
    ("nangsuatsanluong", "NS_SL"),
)


def _sr_ct_dong(nhan):
    """Cột B sheet "CT Vận hành" bản Showroom -> mã chỉ tiêu năng suất. Dòng khác -> {}.

    Khớp kiểu "chứa" sau `_nd` để chịu được dấu cách thừa của file ("Năng suất Sản lượng /1NV").
    Thứ tự trong `_SR_CT_DONG` không quan trọng ở đây vì hai khoá không lồng nhau, nhưng ĐỪNG
    thêm khoá ngắn kiểu "sanluong": nó nằm trong "nangsuatsanluong1nvbanhang" và sẽ bắt sai dòng.
    """
    n = _nd(nhan)
    for key, ma in _SR_CT_DONG:
        if key in n:
            return {"dim1": ma, "dim2": "Tổng"}
    return {}


# ── Bản kế hoạch của 4 khối về VPS 17/08/2026: 3.TS (Trạm sạc) · 5.HT (Xe tải) · 6.XVP (Vận tải
# Taxi XVP) · 7.AnTX (An Taxi). Bốn file DÙNG CHUNG một biểu mẫu sheet "KHDT": cột C là nhãn dòng,
# cột G..R là 12 tháng, khối tách làm hai phần A GIÁ TRỊ / B SẢN LƯỢNG.
# BẢNG PHÂN CẤP 3 TẦNG, và tầng nào cũng có số ở CÙNG một cột — cộng nhầm tầng là nhân đôi:
#     Tổng   : dòng "Khối KD …" (số chính thức của khối)
#     Nhóm   : các nguồn thu / nhóm sản lượng, Σ nhóm = Tổng (đã verify từng file, xem `_kiem_chung`)
#     Chi tiết: dòng con của một nhóm (An Taxi: 3 dòng "DT cung cấp dịch vụ bằng xe …" nằm TRONG
#               "Doanh thu từ hoạt động kinh doanh taxi"; Xe tải: 6 dòng đời xe nằm trong A200)
# `dim2` mang đúng ba nhãn đó để backend chọn tầng, KHÔNG phải tự đoán theo tên dòng.
# Khớp CHÍNH XÁC sau `_nd` (không "chứa"): "Doanh thu khác" là nhãn của một NHÓM ở An Taxi/Xe tải,
# mà "chứa" thì nó cũng khớp "Doanh thu khác Kênh B2C" của bản Showroom — hai bảng mã khác nhau.
_KH_KHOI = {
    # ── 7.AnTX — Khối KD Dịch vụ An Taxi ──
    "doanhthubanxethuongquyen": ("DT_BANXE_TQ", "Nhóm"),
    "doanhthutuhoatdongkinhdoanhtaxi": ("DT_TAXI", "Nhóm"),
    "dtcungcapdichvubangxecongty": ("DT_TAXI_XECT", "Chi tiết"),
    "dtcungcapdichvubangxethuongquyen": ("DT_TAXI_XETQ", "Chi tiết"),
    "dtcungcapdichvubangxethueapp": ("DT_TAXI_XEAPP", "Chi tiết"),
    "doanhthukhac": ("DT_KHAC", "Nhóm"),
    "soluongxechay": ("SL_XE", "Nhóm"),
    # ── 3.TS — Khối KD Trạm sạc Vgreen ──
    "doanhthubanhang": ("DT_BANHANG", "Nhóm"),
    "tramsac": ("SL_TRAMSAC", "Chi tiết"),
    "dichvukhac": ("SL_DVKHAC", "Chi tiết"),
    # ── 6.XVP — Khối KD Vận tải Taxi XVP ──
    "doanhthudichvutaxi": ("DT_TAXI", "Nhóm"),
    "doanhthuhtkd": ("DT_HTKD", "Nhóm"),
    "doanhthuxethuekhoan": ("DT_THUEKHOAN", "Nhóm"),
    "doanhthuphiquanlyxehtkd": ("DT_PQL_HTKD", "Nhóm"),
    "doanhthuphiquanlyhtx": ("DT_PQL_HTX", "Nhóm"),
    "dichvutaxi": ("SL_TAXI", "Chi tiết"),
    "htkd": ("SL_HTKD", "Chi tiết"),
    "xethuekhoan": ("SL_THUEKHOAN", "Chi tiết"),
    "phiquanlyxehtkd": ("SL_PQL_HTKD", "Chi tiết"),
    "phiquanlyhtx": ("SL_PQL_HTX", "Chi tiết"),
    # ── 5.HT — Khối KD Xe tải (Hưng Thịnh) ──
    "doanhthubanxe": ("A200", "Nhóm"),
    "sanluongbanxe": ("A200", "Nhóm"),
    "xedaukeo440": ("T101_1", "Chi tiết"),
    "xeben84": ("T101_2", "Chi tiết"),
    "xediensanny": ("T101_3_SANNY", "Chi tiết"),
    "xedienshacman": ("T101_3_SHACMAN", "Chi tiết"),
    "xedienfaw": ("T101_3_FAW", "Chi tiết"),
    "xediendongfeng": ("T101_3_DONGFENG", "Chi tiết"),
    "doanhthutrusac": ("T101_4", "Nhóm"),
    "sanluongtrusac": ("T101_4", "Nhóm"),
    "doanhthuhoatdongtaichinh": ("T102", "Nhóm"),
}
# Nhãn quá dài để gõ nguyên văn (kèm cả công thức trong ngoặc) -> khớp theo TIỀN TỐ.
_KH_KHOI_TIEN_TO = (
    ("doanhthuhoahong", ("DT_HOAHONG", "Nhóm")),   # "…được hưởng( chia sẻ 750đ/kw, chi hộ, DV…)"
)


def _kh_khoi_dong(nhan):
    """Nhãn cột C sheet KHDT (4 bản kế hoạch khối) -> {dim1: mã, dim2: cấp}.

    Dòng không khai -> {} (bị `loc` loại). CỐ Ý không cảnh báo "không map được": sheet còn dòng
    trống và dòng tiêu đề phần B, kêu lên là nhiễu. Đổi lại, mọi spec đều có `_kiem_chung` chốt
    Σ Nhóm = dòng Tổng — nhãn mới xuất hiện mà không được khai sẽ lộ ra ở đó chứ không im.
    """
    n = _nd(nhan)
    if n.startswith("khoikd"):        # "Khối KD Dịch vụ An Taxi", "Khối KD Xe tải", …
        return {"dim1": "TONG", "dim2": "Tổng"}
    got = _KH_KHOI.get(n)
    if not got:
        got = next((v for k, v in _KH_KHOI_TIEN_TO if n.startswith(k)), None)
    return {"dim1": got[0], "dim2": got[1]} if got else {}


# ── Sheet "CT Vận hành" của 4 bản kế hoạch trên. Nhãn ở cột B, dải tháng E..J (T7..T12) —
# giống bản Showroom, KHÁC bản XDV (có thêm cột "Công thức tính" nên dải là F..K).
# Mã ở cột A của FILE KHÔNG tin được: bản Trạm sạc gắn CT37 cho dòng "Năng suất Doanh thu/1NV"
# nhưng giá trị lại là 42.726.780.000/1.054 bộ = giá trị bình quân MỘT BỘ, còn dòng CT35 "Năng
# suất Sản lượng/1NV" (đvt ghi "bộ") lại mang 42.726.780.000/12 NV = tiền. Hai dòng bị hoán vị so
# với nhãn. Vì vậy mã ở đây gắn theo Ý NGHĨA SỐ đã đối chiếu, và `payload.ma_file` giữ nguyên mã
# của file để đối chất khi nghiệp vụ sửa biểu mẫu.
_KH_CT_DONG = (
    ("soluongcuocchay", "SL_CUOC"),
    ("soluongxechay", "SL_XE"),
    ("songayphatsinhdoanhthu", "SO_NGAY"),
    ("doanhthubinhquanxengay", "DT_BQ_XE_NGAY"),
    ("doanhthubinhquanngay", "DT_BQ_NGAY"),
    ("doanhthubinhquancuoc", "DT_BQ_CUOC"),
    ("giatribinhquanxeban", "GT_BQ_XE"),
    ("tylehoanthanhtrenapp", "TL_APP"),
    ("tylehoanthanhchuyen", "TL_HT_CHUYEN"),
    ("tylevandoanh", "TL_VAN_DOANH"),
    ("tylexenam", "TL_XE_NAM"),
    ("tylehuychuyen", "TL_HUY_CHUYEN"),
    ("tylebangiaothicongdunghan", "TL_DUNG_HAN"),
    ("nangsuatdoanhthu", "NS_DT"),
    ("nangsuatsanluong", "NS_SL"),
    ("sanluongnvbanhang", "SL_NV"),
    ("sanluongbanxe", "SL_BAN_XE"),
    ("sanluongtramsacbanduoc", "SL_TRAMSAC"),
    ("dontonta", "DON_TON"),
    ("donton", "DON_TON"),
    ("tonkho30ngay", "TON_30N"),
    ("xetonkho30ngay", "TON_30N"),
    ("hoahongduochuongdaxacnhan", "HH_DA_XN"),
    ("hoahongduochuongchuaduocxacnhan", "HH_CHUA_XN"),
    ("congnoquahandacococ", "CN_QH_COC"),
    ("loinhuan", "LOI_NHUAN"),
    ("doanhthu", "DOANH_THU"),
)


def _kh_ct_dong(nhan):
    """Cột B sheet "CT Vận hành" -> mã chỉ tiêu. Dòng không khai -> {}.

    Khớp kiểu "chứa" (nhãn có đuôi giải thích dài: "Số lượng xe chạy/xe phát sinh doanh thu",
    "Số lượng cuốc chạy (cuốc hoàn thành)"), nên THỨ TỰ trong `_KH_CT_DONG` là thứ tự ưu tiên:
    "doanhthubinhquanngay" phải đứng TRƯỚC "doanhthu", "soluongcuocchay" trước "soluongxechay"…
    Bỏ quy tắc đó là mọi dòng doanh thu bình quân đều bị gắn DOANH_THU và đè lên nhau.
    """
    n = _nd(nhan)
    for key, ma in _KH_CT_DONG:
        if key in n:
            return {"dim1": ma}
    return {}


def _xdv_kh_cong(v):
    """Cột A -> đánh dấu dòng "Cộng" (tổng khối) của từng mục.

    Dòng tổng của bản kế hoạch XDV ghi chữ "Cộng" ở cột A và BỎ TRỐNG cột B (khác hẳn các dòng
    xưởng). Mục III (lợi nhuận sau thuế) CHỈ có đúng dòng này — bỏ qua là mất trắng cả mục.
    """
    return {"dim2": "Tổng"} if _nd(v) in ("cong", "tong", "tongcong") else {}


_KHO_KENH = {
    "2001": "B2C",          # Kho ô tô showroom
    "1010": "B2C",          # Xe ô tô Showroom Quảng Ninh
    "2006": "GF",           # Kho ô tô cũ
    "2010": "B2B",          # Kho ô tô (Xe thương quyền)_Vinfast TC
    "2011": "B2B",          # Kho ô tô (Xe thương quyền)_Xanh Vĩnh Phúc
    "2013": "B2B",          # Kho ô tô B2B_KD Mạnh
    "2014": "B2B",          # Kho ô tô B2B_KD Quỳnh
    "2015": "B2B",          # Kho ô tô B2B_KD Hiếu
    "2004": "Xe demo",      # Kho ô tô demo      — KHÔNG thuộc 3 kênh bán
    "2008": "Xe ký gửi",    # Kho ô tô ký gửi    — KHÔNG thuộc 3 kênh bán
}


def _kho_kenh(v):
    """Mã kho -> kênh bán, cho báo cáo tồn kho xe TỰ ĐỘNG (KD36).

    VÌ SAO CẦN: bản tay `Tonkhoxevatly_T{n}_{kênh}.xlsx` mang kênh trong TÊN FILE (3 file/kỳ);
    bản tự động gộp một file và KHÔNG có cột kênh nào. Mã kho là thứ duy nhất suy ra được.

    ĐỘ TIN CẬY (đo 03/09/2026, đối chiếu bản tay chốt 28/08 với bản tự động cùng ngày): 1.217 xe
    khớp VIN -> 1.217 ĐÚNG / 0 SAI. Không mã kho nào chứa hai kênh khác nhau. Toàn bộ 1.214 VIN
    của bản tay đều có mặt trong bản tự động.

    "Xe demo" (2004) VÀ "Xe ký gửi" (2008) LÀ NHÓM RIÊNG, KHÔNG PHẢI KÊNH BÁN — cùng lý lẽ với
    `_coc_trang_thai` (ba nhóm chứ không phải hai). Chính chúng là 146 VIN mà bản tay không có:
    3 file kênh cũ đã lọc chúng ra. Dồn vào B2B/B2C là thổi phồng tồn kênh; bỏ hẳn là hụt tồn
    tổng. Để riêng thì màn lọc theo 3 kênh vẫn đúng mà tổng tồn vẫn đủ.

    Mã lạ -> "Khác" (nhìn thấy được trên màn), KHÔNG im lặng gán về một kênh: kho mới mở là
    chuyện bình thường và phải lộ ra để bổ sung vào bảng trên.
    """
    s = re.sub(r"\D", "", str(v or ""))
    if not s:
        return None
    return _KHO_KENH.get(s, "Khác")


_KENH_TK = [("13111", "B2C"), ("13116", "GF"), ("1316", "B2B")]


def _kenh_tk(v):
    """Số hiệu tài khoản -> kênh bán. Khớp theo TIỀN TỐ, không khớp tuyệt đối.

    Mapping ghi kênh B2B = "1316" nhưng tài khoản THẬT trong file là 13161 và 13163 (1.053 dòng
    trên tổng 1.911). Khớp tuyệt đối "1316" thì toàn bộ B2B rơi vào nhóm "Khác" mà vẫn chạy trơn.
    Thứ tự trong `_KENH_TK` quan trọng: "13116" (GF) phải đứng TRƯỚC "1316" (B2B), nếu không GF
    bị tiền tố B2B nuốt.
    """
    s = re.sub(r"\D", "", str(v or ""))
    for tien_to, kenh in _KENH_TK:
        if s.startswith(tien_to):
            return kenh
    return "Khác" if s else None


def _qua_han(v):
    """Cột "số ngày quá hạn" -> "Quá hạn" / "Trong hạn".

    Mapping (DASHBOARD 2 dòng 3-4) định nghĩa quá hạn = cột AE > 0, trong hạn = còn lại. Ô rỗng
    hoặc không phải số -> "Trong hạn": đó là hợp đồng chưa tới hạn nên file bỏ trắng, không phải
    dữ liệu thiếu.
    """
    n = _so(v)
    return "Quá hạn" if n is not None and n > 0 else "Trong hạn"


def _tk_cap(v):
    """Số hiệu tài khoản -> "Cấp 1".."Cấp 4" theo SỐ CHỮ SỐ (111 -> Cấp 1, 1111 -> Cấp 2…).

    Bảng CĐPS liệt kê CẢ CÂY tài khoản: dòng 111 rồi 1111 rồi 11114, mỗi cấp đã bao trọn cấp
    dưới. Cộng hết là gấp 3-4 lần. File phân cấp bằng ĐỘ THỤT ĐẦU DÒNG, mà engine `strip()` ô
    text trước khi tới hook nên không đọc được thụt lề — dùng số chữ số thay, quy tắc hệ thống
    tài khoản Việt Nam là cố định nên tương đương. Màn hình phải lọc theo cấp, đừng SUM cả cột.
    """
    s = re.sub(r"\D", "", str(v or ""))
    if not s:
        return None
    return f"Cấp {max(1, len(s) - 2)}"


def _coc_ngay(v):
    """Cột "Ngày nhập COC" của nguồn TỰ ĐỘNG -> "Đã có COC" / "Chưa có COC".

    Khác `_coc_trang_thai`: bên bản TAY cột COC là ô TỰ DO (kế toán gõ đủ kiểu chữ) nên phải đoán
    qua 15 biến thể; bản tự động của Cyber là một cột NGÀY sạch, có ngày = đã nhập COC, bỏ trắng =
    chưa. Đưa ô trắng vào `_coc_trang_thai` sẽ ra "Không xác định" (nhánh cuối) — đúng cho bản tay
    (ô trắng ở đó thật sự là không biết) nhưng SAI cho bản tự động, và sẽ dồn gần hết dòng vào một
    nhóm vô nghĩa.
    """
    if isinstance(v, (dt.datetime, dt.date)):
        return "Đã có COC"
    s = str(v if v is not None else "").strip()
    if not s:
        return "Chưa có COC"
    return "Đã có COC" if _date(s) else _coc_trang_thai(v)


_COC_CHUA = ("chuacococ", "chuave", "cocchuave", "chuaco", "chua", "xechuacococ",
             "dangst", "chohsgf", "chuacohosogf", "chuacohoso")
_COC_DA = ("dacococ", "dave", "dacohoso")
_COC_NGOAI = ("tq", "xanhtq", "thuongquyen", "xanhthuongquyen", "xegf", "gf", "xedemo", "demo")


def _coc_trang_thai(v):
    """Cột "COC Ngày về/ chưa về" -> "Đã có COC" / "Chưa có COC" / "Không xác định".

    Cột này là Ô TỰ DO, kế toán mỗi kỳ ghi một kiểu. Trên 8 kỳ đã gặp: datetime thật (669 dòng),
    chuỗi "dd/mm/yyyy" (~120), SỐ SERIAL Excel chưa format (46137, 46120…), và ~15 biến thể chữ
    ("Chưa có COC", "CHƯA CÓ COC", "COC chưa về", "Đã về", "chưa có"…). Đếm bằng
    `isinstance(datetime)` như bản dò tay đầu tiên là bỏ sót gần một nửa số dòng ĐÃ CÓ COC.

    Ba nhóm chứ không phải hai: "TQ"/"Xanh TQ"/"Thương quyền"/"Xe demo"/"xe GF" là xe KHÔNG thuộc
    diện COC — chính file cũng tách riêng ("Xanh Thương quyền 28,03 tỷ" đứng ngoài bảng COC ở T5).
    Dồn chúng vào "Chưa có COC" là thổi phồng phần chưa có. Chuỗi lạ -> "Không xác định" và được
    đếm vào cảnh báo, KHÔNG im lặng gán về một phía.
    """
    if isinstance(v, (dt.datetime, dt.date)):
        return "Đã có COC"
    s = str(v if v is not None else "").strip()
    if not s:
        return None
    if _date(s):                       # "22/01/2026", "21/1/2025", "2026-07-14 00:00:00"
        return "Đã có COC"
    # Serial Excel chưa format (`_lay_o` đã đổi ô thành chuỗi trước khi gọi hook, nên 46137 tới
    # đây là "46137"). 40000 ≈ 2009-07, 60000 ≈ 2064; ngoài dải là số rác, phần lớn là 0.
    if re.fullmatch(r"\d+(\.0+)?", s):
        return "Đã có COC" if 40000 <= float(s) <= 60000 else "Không xác định"
    n = _nd(s)
    if any(k in n for k in _COC_NGOAI):
        return "Không xác định"
    if any(k in n for k in _COC_CHUA):
        return "Chưa có COC"
    if any(k in n for k in _COC_DA):
        return "Đã có COC"
    # "ST 16/04/2026", "DBP 28/5/26" — ngày kèm chú thích. Bắt buộc có DẤU GẠCH CHÉO: "Đã giải
    # ngân 30.01" cũng có dạng số-chấm-số nhưng là ngày GIẢI NGÂN của ngân hàng, không phải COC.
    if re.search(r"\d{1,2}/\d{1,2}", s):
        return "Đã có COC"
    return "Không xác định"


# QLTS gọi cùng một địa điểm ở nhiều nghiệp vụ khác nhau, mà master_data lại có tới 3 cost center
# cùng địa danh (vd Vĩnh Phúc: `VP_SR` showroom · `VP_XDV` xưởng · `VP_DP` depot taxi). Vì vậy KHÔNG
# bóc hết tiền tố rồi khớp theo địa danh trần — sẽ va nhau và gán bừa. Thay vào đó dùng CHÍNH TỪ
# LOẠI trong tên đơn vị để khoanh hậu tố mã cần tìm.
_QLTS_LOAI = (
    (r"^xưởng\s*dịch\s*vụ\s*", ("_XDV",)),
    (r"^showroom\s*",          ("_SR",)),
    # Taxi: thử depot (Xanh) trước, rồi An Taxi, rồi HTX — nhờ vậy "Taxi Vĩnh Phúc" ra `VP_DP`
    # (Khối KD Vận tải Taxi Xanh) còn "Taxi Sơn Tây" ra `ST_AT` (Khối KD Dịch vụ An Taxi). Đây
    # chính là chỗ tách "Khối taxi" của QLTS thành 2 khối chuẩn mà không cần bảng tay.
    (r"^taxi\s*",              ("_DP", "_AT", "_HTX")),
    (r"^trạm\s*sạc\s*",        ("_TS",)),
    # VIẾT TẮT — các nguồn QLTS khác (bảo hiểm cột "Vị trí tài sản", bảo dưỡng xe demo cột "Địa
    # điểm BD") ghi tắt "XDV Sơn Tây" / "SR Ocean Park" thay vì viết đủ như sheet tài sản. Đặt SAU
    # các mẫu viết đủ để không chen ngang, và neo bằng `\b` để không cắt nhầm tên bắt đầu bằng
    # "sr"/"dp" (vd không có ca nào hiện tại, nhưng luật phải chặt).
    (r"^xdv\b\s*",             ("_XDV",)),
    (r"^sr\b\s*",              ("_SR",)),
    (r"^(?:depot|dp)\b\s*",    ("_DP", "_AT", "_HTX")),
)
# File QLTS hay ghi "<địa điểm> - <tỉnh>" trong khi master chỉ có địa điểm ("Xưởng dịch vụ Việt Trì
# - Phú Thọ" vs "Vinfast Việt Trì"). Cắt phần sau dấu gạch bằng regex TỔNG QUÁT thay vì liệt kê
# từng ca — bản đầu tôi khai alias tay và trỏ SAI đích (`viettriphutho`->`vietri` trong khi khoá
# master là `viettri`), nên 974 tài sản xưởng Việt Trì rơi mất mà cảnh báo chỉ ghi "ngoài danh mục".
_QLTS_BO_TINH = re.compile(r"\s*-\s*[^-]+$")
# Số thứ tự khu/lô dính cuối tên công trường ("Yên Bình 3" -> "yenbinh"). Chạy SAU khi khoá đủ đã
# trượt, xem `_cc_qlts`.
_QLTS_KHU_SO = re.compile(r"\d+$")
# Còn lại chỉ là khác CÁCH GỌI, không suy được bằng luật.
_QLTS_ALIAS = {"saigon": "hochiminh", "ocenpark": "oceanpark"}
# Gán THẲNG mã cost center cho các tên mà luật chung không với tới. Bảng này khai TÊN -> MÃ nên
# không thể lan sang tên khác, khác hẳn cách nới `hau_to`: có lúc tôi thêm "_GD" vào danh sách hậu
# tố của tên trơn để bắt "An An Garden", và nó gán luôn "Tài sản Sơn Tây (gồm văn phòng, xưởng,
# Nhà Xanh...)" — 2.930 tài sản khối văn phòng — vào `ST_GD` khối Dịch vụ An KS.
# `xelanhdao`: sheet "th XE DEMO" ghi cột "Phân loại" là ĐƠN VỊ GIỮ XE ("SR Hạ Long", "SR Sơn
# Tây"…), riêng 10 xe của ban lãnh đạo ghi "xe lãnh đạo". Bản đầu cố ý KHÔNG map vì file không nói
# xe đứng tên đơn vị nào — nhưng cột đó đọc theo đơn vị GIỮ xe chứ không phải bên đứng tên, và
# danh mục có sẵn `BLĐ` "Ban lãnh đạo". Để trống thì 10 xe rơi khỏi mọi bộ lọc Đơn vị mà mapping
# QLTS đòi cho nhóm chỉ tiêu bảo trì (#14-#17) — kế toán báo thiếu ngày 26/08/2026.
_QLTS_CC_TRUC_TIEP = {"anangarden": "ST_GD", "xelanhdao": "BLĐ"}
_QLTS_CACHE = {}

# 8 khối của file QLTS -> 10 khối chuẩn của master_data. QLTS phân loại theo BẢN CHẤT TÀI SẢN nên
# tên không trùng khối kinh doanh; phải map tay. Dùng làm NỀN cho cột `khoi`: `_cc_qlts` sẽ ghi đè
# bằng khối của cost center khi map được (nhờ vậy "Khối taxi" tách đúng thành Vận tải Taxi Xanh /
# Dịch vụ An Taxi theo từng đơn vị), còn đơn vị chưa có mã chuẩn thì vẫn giữ được khối từ đây ->
# bộ lọc Khối phủ 100% dòng thay vì rơi mất phần dư.
_QLTS_KHOI = {
    "khoiduan": "Khối KD Dự án",
    "khoihaumai": "Khối KD Vinfast - XDV",          # 'hậu mãi' CHÍNH LÀ khối XDV (xem hcns_nhansu_phongban)
    "khoikinhdoanh": "Khối KD Vinfast - Showroom",
    "khoitramsac": "Khối KD Trạm sạc Vgreen",
    "khoikinhdoanhdichvu": "Khối KD Dịch vụ An KS",
    "khoivanphongkhoxuong": "Khối hỗ trợ tập đoàn",
    # Mặc định cho "Khối taxi": phần lớn là Taxi Xanh (2.102/2.500 tài sản). Đơn vị nào map được
    # cost center thì `_cc_qlts` ghi đè đúng khối của nó.
    "khoitaxi": "Khối KD Vận tải Taxi Xanh",
    # "Bất động sản" (5 tài sản) KHÔNG có khối chuẩn tương ứng -> để trống, không nhét bừa vào
    # "Khối hỗ trợ tập đoàn" cho đủ.
    #
    # BA FILE BẢO DƯỠNG (31/08/2026) ghi khối theo CÁCH KHÁC sheet tài sản — phải khai riêng,
    # KHÔNG lấy nguyên văn: "Khối KD Dự án" tình cờ đã đúng tên chuẩn, nhưng "Khối KD xe điện
    # Vinfast - SR" thì KHÔNG (chuẩn là "Khối KD Vinfast - Showroom"). Lấy nguyên văn là đẻ thêm
    # một tên khối thứ 11 trong DB, bộ lọc Khối tách làm hai dòng cho cùng một khối.
    "khoikdduan": "Khối KD Dự án",
    "khoikdxedienvinfastsr": "Khối KD Vinfast - Showroom",
}


def _cty_qlts(ten):
    """Cột "TÊN PHÁP NHÂN" của file QLTS -> mã công ty chuẩn, hoặc None.

    Đây là NƯỚC LÙI cho `cong_ty`, KHÔNG phải nguồn chính: spec đặt cột này TRƯỚC `cost_center`
    nên khi `_cc_qlts` khớp được đơn vị thì công ty của cost center GHI ĐÈ lên đây. Lý do phải
    tách hai nguồn: đó là HAI KHÁI NIỆM khác nhau và lệch nhau 2.171 dòng ở kỳ T8/2026 — cột
    Pháp nhân là bên ĐỨNG TÊN tài sản (Café VinFast Sơn Tây đứng tên Thịnh Cường), còn công ty
    của cost center là pháp nhân VẬN HÀNH đơn vị (An An's Garden). Bảng chỉ tiêu đọc theo cơ cấu
    vận hành nên cost center thắng; cột Pháp nhân chỉ để vá 3.687 dòng mà đơn vị chưa có mã chuẩn
    (chủ yếu 2.738 tài sản "Sơn Tây (gồm văn phòng, xưởng, Nhà Xanh…)") — nhờ nó bộ lọc Công ty
    của màn tài sản phủ 99% thay vì 81%.

    TÊN NGOÀI DANH MỤC TẬP ĐOÀN THÌ TRẢ None, KHÔNG đoán: cột này lẫn cả cá nhân đứng tên
    (BÙI HÙNG THỊNH 143 dòng), bên cho thuê tài chính (Chailease, BIDV SumiTrust) và rác `0x2a`
    (90 dòng). Tổng 264 dòng — vẫn nằm trong tổng tài sản, chỉ không thuộc pháp nhân nào.
    """
    t = str(ten or "").strip()
    if not t:
        return None
    return _master_loader().resolve_company_code(t) or None


def _khoi_qlts(ten):
    """Tên khối trong file QLTS -> tên khối chuẩn. Không có trong bảng -> None (giữ trống)."""
    return _QLTS_KHOI.get(_nd(ten))


def _qlts_khoa(ten):
    """Khoá khớp cho PHÍA MASTER: bóc luôn từ loại của master ("Depot Vĩnh Phúc" -> "vinhphuc").

    Bắt buộc phải có, KHÔNG dùng trực tiếp `_bo_tien_to`: hàm đó bóc `vinfast`/`xuongdichvu`/`xdv`
    nhưng KHÔNG bóc `depot`/`duan`/`tramsac`. Hệ quả đã gặp: "Vinfast Tuyên Quang" (TQ_XDV) ra khoá
    `tuyenquang` còn "Depot Tuyên Quang" (TQ_DP) ra `depottuyenquang` -> tra hậu tố `_DP` cho
    "Tài sản Taxi Tuyên Quang" TRƯỢT và rơi vào ứng viên duy nhất còn lại là TQ_XDV, gán 328 tài
    sản taxi vào khối Xưởng dịch vụ mà không có lỗi nào. Không sửa `_bo_tien_to` vì nó dùng chung
    với spec HCNS/XDV/công nợ.
    """
    n = _nd(ten)
    doi = True
    while doi:
        doi = False
        for p in ("depot", "duan", "tramsac", "garden"):
            if n.startswith(p) and len(n) > len(p):
                n, doi = n[len(p):], True
    return _bo_tien_to(n)


def _cc_qlts(ten):
    """Đơn vị sử dụng của QLTS ("Tài sản Xưởng dịch vụ Ocean Park") -> cost center chuẩn.

    Trả `{cost_center, cong_ty, khoi}` lấy TỪ master_data, nên một phép map này giải luôn cả ba
    chiều mà mapping QLTS đòi (Công ty · Đơn vị · và khối để lọc). CỐ Ý không dùng cột "TÊN PHÁP
    NHÂN" của file làm `cong_ty`: cột đó là PHÁP NHÂN ĐỨNG TÊN tài sản nên lẫn cả `0x2a` (198
    dòng), tên cá nhân (BÙI HÙNG THỊNH 145 dòng), và bên cho thuê tài chính (Chailease, BIDV
    SumiTrust) — đưa vào bộ lọc Công ty là sinh ra "công ty" không tồn tại. Bản nguyên văn vẫn giữ
    ở payload để truy vết.

    Không khớp -> `{"_khong_map": <tên gốc>}` giống `_cc_theo_khoi`, để `giu_khi_khong_map` giữ
    nguyên văn. 15 đơn vị hiện chưa có mã chuẩn (An An Garden, VP G2, Café VinFast, Sân
    Pickleball, Núi Pháo, Bình Phước, Quảng Ngãi, Cù Vân, Bất động sản, trạm sạc theo địa
    điểm...) — user chốt 17/08/2026: giữ nguyên tên, KHÔNG gộp vào "Khác", để tổng không hụt và
    biết chính xác cần bổ sung mã nào vào master_data.
    """
    goc = str(ten or "").strip()
    if not goc:
        return None
    if not _QLTS_CACHE:
        master = _master_loader()
        for cc in master.master_data().get("costCenters", []):
            ma = str(cc.get("ma") or "").strip()
            _QLTS_CACHE.setdefault(_qlts_khoa(cc.get("ten")), []).append(
                (ma, master.resolve_company_code(cc.get("congTy") or ""), cc.get("khoi") or ""))
    # Bỏ phần trong ngoặc: "Tài sản Sơn Tây (gồm văn phòng, xưởng, Nhà Xanh...)" và
    # "Taxi Vĩnh Phúc (thương quyền)" — phần chú thích không tham gia khớp tên.
    t = re.sub(r"\s*\(.*$", "", goc).strip()
    t = re.sub(r"^\s*tài\s*sản\s*", "", t, flags=re.I).strip()
    # Tên TRƠN (không có từ loại) chỉ thử DỰ ÁN. CỐ Ý không thử `_SR`: showroom trong file QLTS
    # luôn ghi rõ "Showroom …", nên tên trơn mà khớp `_SR` gần như chắc chắn là gán bừa. Bản đầu
    # có `_SR` ở đây và nó gán "Tài sản Sơn Tây (gồm văn phòng, xưởng, Nhà Xanh, Bảo Long, Đồng
    # Vừng)" — 2.936 tài sản thuộc *Khối văn phòng, kho, xưởng* — vào `ST_SR` khối Showroom.
    hau_to = ("_DA",)
    for pat, suf in _QLTS_LOAI:
        if re.match(pat, t, re.I):
            t = re.sub(pat, "", t, flags=re.I).strip()
            hau_to = suf
            break
    key = _qlts_khoa(t)
    ma_thang = _QLTS_CC_TRUC_TIEP.get(key)
    if ma_thang:
        for ds in _QLTS_CACHE.values():
            for ma, cty, khoi in ds:
                if ma == ma_thang:
                    return {"cost_center": ma, "cong_ty": cty, "khoi": khoi}
    if key not in _QLTS_CACHE:
        key = _QLTS_ALIAS.get(key) or _qlts_khoa(_QLTS_BO_TINH.sub("", t)) or key
        key = _QLTS_ALIAS.get(key, key)
    if key not in _QLTS_CACHE:
        # SỐ THỨ TỰ KHU/LÔ ở cuối tên công trường: báo cáo bảo dưỡng ghi "Yên Bình 3" trong khi
        # danh mục chỉ có "Dự án Yên Bình" (13 thiết bị rơi mất vì đúng một chữ số). Chỉ thử ở
        # nước cuối, sau khi khoá đầy đủ đã không khớp — cắt sớm là gộp bừa các khu khác nhau.
        key = _QLTS_KHU_SO.sub("", key) or key
    for suf in hau_to:
        for ma, cty, khoi in _QLTS_CACHE.get(key) or []:
            if ma.endswith(suf):
                return {"cost_center": ma, "cong_ty": cty, "khoi": khoi}
    # CỐ Ý KHÔNG có nhánh dự phòng "địa danh khớp nhưng khác loại thì lấy ứng viên duy nhất".
    # Bản đầu có nhánh đó và nó gán "Tài sản Taxi Tuyên Quang" vào `TQ_XDV` (xưởng dịch vụ) —
    # sai khối, sai cả cách đọc số. Không map được thì để nguyên tên còn hơn gán bừa.
    return {"_khong_map": goc}


# ── BÁO CÁO QTVH XANH TAXI (`B.6.XVP.PKDVH.M.*.Baocaotonghop`) ─────────────────────────────
# Bảng NGANG: cột A = khối/depot (ô GỘP, chỉ ghi ở dòng đầu -> cột khai `lap_lai`), cột B = tên
# chỉ tiêu, C..AG = mỗi ngày một cột. Mỗi sheet là MỘT THÁNG và tháng mới lại thêm sheet vào chính
# file cũ (`moi_sheet_chua`), nên chỉ tiêu phải dò theo TÊN chứ không theo số dòng: sheet "BÁO CÁO
# THÁNG 7" thiếu hẳn 3 dòng nhân sự và 3 dòng cảnh báo của sheet tháng 8, mọi số dòng đều lệch.
#
# BẪY: `_nd()` bỏ hết ký tự không phải chữ/số nên "GMV ngày" (dòng 19, tiền của NGÀY ĐÓ) và
# "GMV/ Ngày" (dòng 22, tiền chia SỐ XE HOẠT ĐỘNG) chuẩn hoá về CÙNG một chuỗi "gmvngay". Hai
# dòng khác nhau cả bậc độ lớn (899.802.460 vs 978.046). Vì thế bảng này so bằng `_nd_gach` —
# GIỮ LẠI dấu "/" — và khớp TUYỆT ĐỐI, không khớp kiểu "chứa".
def _nd_gach(s):
    """Như `_nd` nhưng giữ dấu '/' — xem bẫy GMV ngày vs GMV/ Ngày ở trên."""
    s = str(s if s is not None else "").strip().lower().replace("đ", "d")
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9/]", "", s)


# nhãn cột B -> (mã chỉ tiêu, nhóm đơn vị). Nhóm đi vào `dim3` để backend biết phải quy đổi thế
# nào: "tien" = ĐỒNG (chia 1e9 khi hiển thị tỷ) · "dem" = số đếm (xe/cuốc/người) · "tyle" = phân
# số 0..1 (nhân 100). Một report_type ôm cả ba vì đây là MỘT bảng của nghiệp vụ; tách ba spec chỉ
# để "đơn vị sạch" thì mỗi lần biểu mẫu thêm dòng lại phải sửa ba chỗ.
_XVP_CT = {
    # nhân sự
    "tongnhansu": ("NS_TONG", "dem"), "nhansu": ("NS_TONG", "dem"),
    "ho": ("NS_HO", "dem"), "depot": ("NS_DEPOT", "dem"),
    # đội xe
    "tongsoxe": ("XE_TONG", "dem"), "tongsoxetaidepot": ("XE_TONG", "dem"),
    "soluongxedudieukienvandoanh": ("XE_DU_DK", "dem"),
    "soluongxethieutai": ("XE_THIEU_TAI", "dem"),
    "xethuongquyen": ("XE_TQ", "dem"), "xethuekhoan": ("XE_KHOAN", "dem"),
    "soxehoatdongtb/ngaytaxi": ("XE_HOAT_DONG", "dem"),
    "soxenamxuongvctnloikythuat": ("XE_NAM_XUONG", "dem"),
    "soxetaidoncongan": ("XE_CONG_AN", "dem"),
    "soxedanglamlaigiaytodobimatbihong": ("XE_GIAY_TO", "dem"),
    "soxevctntrongngay": ("XE_VCTN", "dem"),
    "tilexenamdepotxdvcqca": ("TL_XE_NAM_FILE", "tyle"),
    "xesaphetdangkiem": ("XE_DANGKIEM", "dem"),
    "xesaphethanbaohiem": ("XE_BAOHIEM", "dem"),
    # doanh thu (GMV)
    "gmvngay": ("DT_NGAY", "tien"), "targetngay": ("TARGET_NGAY", "tien"),
    "gmvluyke": ("DT_LUYKE", "tien"),
    "gmv/ngay": ("DT_TREN_XE_NGAY", "tien"), "gmv/tongxe": ("DT_TREN_XE", "tien"),
    "aov": ("DT_TREN_CUOC", "tien"),
    "doanhsoxethuongquyen/xekhoan": ("DS_TQ_KHOAN", "tien"),
    "doanhsoappan": ("DS_APP_AN", "tien"),
    "tylehoanthanhkpi": ("TL_HT_KPI", "tyle"),
    # sản lượng · chất lượng
    "tongcuoctaxi": ("CUOC_TONG", "dem"), "tongcuoc": ("CUOC_TONG", "dem"),
    "cuochoanthanh": ("CUOC_HT", "dem"), "cuochoanthanhtaxi": ("CUOC_HT", "dem"),
    "cuochuy": ("CUOC_HUY", "dem"), "cuochuytaxi": ("CUOC_HUY", "dem"),
    "tylehoanthanhcuoc": ("TL_HT_CUOC", "tyle"),
    "tylehoanthanhcuoctaxi": ("TL_HT_CUOC", "tyle"),
    "cuocbidanhgiaduoi5": ("KHIEU_NAI", "dem"),
    # khối KINH DOANH (bán xe) và HỢP TÁC XÃ — cùng nhãn nhưng khác khối, phân biệt bằng `dim2`
    "soluongxeban": ("KD_SL_XE", "dem"), "doanhthu": ("KD_DOANHTHU", "tien"),
    "taixethamgiahtx": ("HTX_TAIXE", "dem"),
    "doanhthutuphixavien": ("HTX_DT_PHI", "tien"),
    "sotaixethanhlyhopdong": ("HTX_THANHLY", "dem"),
}
# Dòng CỐ Ý BỎ (không phải "chưa khai"): "Lũy kế" xuất hiện 3 lần trong file (KINH DOANH dòng
# 121, HỢP TÁC XÃ dòng 124 và 126) với CÙNG nhãn, nên nạp vào là hai dòng HTX chồng khoá nhau và
# cộng đôi. Số lũy kế vốn là tổng dồn của chính dòng ngay trên nó -> backend tự cộng lấy.
_XVP_BO_QUA = {"luyke"}


def _xvp_chi_tieu(nhan):
    """Cột B -> {dim1: mã, dim3: nhóm đơn vị}. Nhãn lạ -> `_khong_map` để engine kêu lên."""
    n = _nd_gach(nhan)
    if not n or n in _XVP_BO_QUA:
        return {}
    got = _XVP_CT.get(n)
    return {"dim1": got[0], "dim3": got[1]} if got else {"_khong_map": str(nhan).strip()}


_XVP_DON_VI = {
    "vanhanhxanh": (None, "Tổng"),        # dòng TỔNG khối — quy tắc 3, không gắn cost_center
    "vinhphuc": ("VP_DP", "Depot"),
    "phutho": ("PT_DP", "Depot"),
    "tuyenquang": ("TQ_DP", "Depot"),
    "kinhdoanh": (None, "Kinh doanh"),    # bán xe — khối riêng trong cùng sheet
    "hoptacxa": (None, "HTX"),
}


def _xvp_don_vi(nhan):
    """Cột A (ô GỘP, đọc kèm `lap_lai`) -> {cost_center, dim2}.

    `dim2` là thứ phân biệt DÒNG TỔNG với dòng depot: cả hai đều mang cùng `dim1`, cộng lẫn là
    nhân đôi đội xe (Σ 3 depot = dòng tổng ở mọi chỉ tiêu đếm — đã verify T8).
    """
    got = _XVP_DON_VI.get(_nd(nhan))
    if not got:
        return {"_khong_map": str(nhan).strip()} if str(nhan or "").strip() else {}
    return {"cost_center": got[0], "dim2": got[1]}


# Cột "MÃ SỐ" của báo cáo HQKD NGÀY Xanh Taxi (`B.6.XVP.D.*.Baocaohqkdngay`) — 5 dòng doanh thu
# THUẦN theo nguồn thu, đúng 5 mục "Cơ cấu doanh thu" mà mapping dòng 31-35 liệt kê. Neo theo MÃ
# chứ không theo nhãn/dòng: nhãn ghi tay ("3.2 Doanh thu thuần HTKD" = doanh thu BÁN XE hợp tác
# kinh doanh) còn số dòng thì đổi theo từng kỳ. Mã ngoài bảng -> {} (dòng bị `loc` loại): sheet
# còn cả phần giá vốn 4.x và chi phí, cố ý không nạp ở đây (đã có `derive_hqkd_ngay.py`).
_XVP_MA_DT = {
    "3.1": "DT_TAXI", "3.2": "DT_BANXE", "3.3": "DT_THUEXE",
    "3.4": "DT_PLF", "3.5": "DT_HTX",
}


def _xvp_ma_doanh_thu(v):
    return {"dim1": _XVP_MA_DT[k]} if (k := str(v or "").strip()) in _XVP_MA_DT else {}


# "Mã kiểu xe" -> "Loại xe" (dòng xe) cho bán xe Showroom. CẦN vì nguồn tự động KD73
# (`Bangkehoadonbanxe`, ổ TESTBAOCAOTUDONG) CHỈ có "Mã kiểu xe"; bản kế toán tự dựng
# (`Xuathoadon_*`) có sẵn cột "Loại xe" và `vhkd_kqkd` đọc thẳng cột đó vào dim1. Không có hook
# này thì dim1 của nguồn mới là "VF305"/"VF304" đứng lẫn với "VF3" của dữ liệu cũ -> biểu đồ
# theo dòng xe tách một dòng xe thành nhiều cột, im lặng.
#
# BẢNG DỰNG TỪ CHÍNH DỮ LIỆU CŨ, không phải suy luận: quét toàn bộ file `Xuathoadon_*` lấy cặp
# (Mã kiểu xe, Loại xe) -> 33 mã, 0 mã cho ra 2 loại xe khác nhau (kiểm 27/08/2026). Vì vậy đây
# là ánh xạ 1-1 tái hiện ĐÚNG cách kế toán đang phân loại, không phải quy ước mới của Dashboard.
#
# KHÔNG rút gọn thành một regex: quy tắc của nguồn KHÔNG nhất quán — MNO1 -> MNO và HRO1 -> HRO
# (bỏ số) nhưng LM1 -> LM1 và LH1 -> LH1 (giữ số). Regex "cắt chữ số cuối" sẽ biến LM1 thành LM,
# tức đẻ ra một dòng xe không tồn tại.
_SR_LOAI_XE = {
    "ECVAN": "ECV", "ECVANNC": "ECV", "ECVANNCCT": "ECV",
    "HRO": "HRO", "HRO1": "HRO", "LH1": "LH1", "LM1": "LM1",
    "MNO1": "MNO", "MPV701": "MPV",
    "VF301": "VF3", "VF302": "VF3", "VF304": "VF3", "VF305": "VF3",
    "VF501": "VF5", "VF502": "VF5",
    "VF602": "VF6", "VF604": "VF6",
    "VF702": "VF7", "VF703": "VF7", "VF708": "VF7", "VF709": "VF7",
    "VF801": "VF8", "VF802": "VF8", "VF803": "VF8", "VF804": "VF8", "VF805": "VF8",
    "VF903": "VF9", "VF904": "VF9", "VF905": "VF9", "VF908": "VF9", "VF909": "VF9",
    "VF910": "VF9", "VF911": "VF9",
}
_RE_VF_DONG = re.compile(r"^VF(\d)\d*$")


def _sr_loai_xe(v):
    """'VF305' -> 'VF3'. Mã lạ -> GIỮ nguyên mã + kêu lên (spec khai `giu_khi_khong_map`).

    Đường lùi `VF<số>` cho model VinFast MỚI ra sau bảng này (VF3xx/VF9xx đã có 12 mã, hãng còn
    thêm): bắt được thì im lặng cho qua, đúng dòng xe. Mã ngoài cả hai -> giữ nguyên để số không
    hụt, kèm cảnh báo để người bổ sung bảng — thà hiện "MPV702" lạ mắt còn hơn mất doanh thu.
    """
    ma = str(v or "").strip().upper()
    if not ma:
        return {}
    if ma in _SR_LOAI_XE:
        return {"dim1": _SR_LOAI_XE[ma]}
    if m := _RE_VF_DONG.match(ma):
        return {"dim1": f"VF{m.group(1)}"}
    return {"dim1": ma, "_khong_map": ma}


_CHUAN_HOA = {
    "cc_qlts": _cc_qlts,
    "khoi_qlts": _khoi_qlts,
    "cty_qlts": _cty_qlts,
    "sr_showroom": _cc_showroom,
    "kenh_tk": _kenh_tk,
    "kho_kenh": _kho_kenh,
    "qua_han": _qua_han,
    "coc_trang_thai": _coc_trang_thai,
    "coc_ngay": _coc_ngay,
    "tk_cap": _tk_cap,
    "xdv": _cc_xdv,
    "hcns_xdv": _cc_hcns_xdv,
    "kh_dong": _kh_dong,
    "xdv_kh_dong": _xdv_kh_dong,
    "xdv_kh_cong": _xdv_kh_cong,
    "xdv_ct_dong": _xdv_ct_dong,
    "sr_ct_dong": _sr_ct_dong,
    "kh_khoi_dong": _kh_khoi_dong,
    "kh_ct_dong": _kh_ct_dong,
    "xvp_chi_tieu": _xvp_chi_tieu,
    "xvp_ma_doanh_thu": _xvp_ma_doanh_thu,
    "xvp_don_vi": _xvp_don_vi,
    "sr_loai_xe": _sr_loai_xe,
    "hoa": lambda v: str(v or "").strip().upper() or None,
    "cat": lambda v: str(v or "").strip() or None,
}


# ─────────────────────────── đọc spec & sheet ───────────────────────────
def load_spec(ref):
    path = ref if os.path.isfile(ref) else os.path.join(SPEC_DIR, f"{ref}.json")
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def specs_for_path(path):
    """Các spec phụ trách FILE này, suy từ `nguon.folder` — [] nếu file không thuộc nguồn nào.

    Vì sao cần: `agent_cli.cmd_autofill` (nút "Phân tích AI") và `template_filler.autofill_file`
    ("Đồng bộ & nạp" + bước nạp lại của theo-dõi-thay-đổi) là hai cửa DUY NHẤT mọi nguồn đi qua,
    và cả hai điều phối bằng một bảng viết tay: mỗi loại báo cáo một deriver. 18 nguồn khai bằng
    JSON spec ra sau nên chưa có tên trong bảng đó — bấm ingest thì báo thành công mà 0 dòng, im
    lặng, không ai hiểu vì sao (đúng lỗi user gặp 10/08/2026). Hàm này là mảnh nối còn thiếu.

    MỘT THƯ MỤC CÓ THỂ CÓ NHIỀU SPEC: `XDV/baocaodoanhthungay` cấp cho 3 spec (doanh thu vụ
    việc / doanh thu RO / số lượng RO) và `KEHOACH/baocaokehoachthang` cấp cho 2. Phải trả về
    TẤT CẢ, chạy thiếu một cái là màn hình thiếu đúng một khối số.
    """
    p = os.path.normpath(os.path.abspath(path)).replace("\\", "/")
    out = []
    for f in sorted(glob.glob(os.path.join(SPEC_DIR, "*.json"))):
        try:
            with open(f, encoding="utf-8") as fh:
                sp = json.load(fh)
        except Exception:
            continue                      # spec hỏng cú pháp KHÔNG được làm chết luồng nạp
        folder = ((sp.get("nguon") or {}).get("folder") or "").strip("/")
        if not folder:
            continue
        if f"/{folder}/" in p:
            out.append(sp)
    return out


def spec_doc_tron_file(path) -> bool:
    """Các spec của file này có ĐỌC TRỌN file không, hay chỉ bóc một lát của nó?

    True  -> caller được phép ra sớm: spec là nguồn DUY NHẤT của file (mọi nguồn VHKD/XDV, file
             sinh ra chỉ để cấp đúng bộ chỉ tiêu mà spec tả).
    False -> file còn cấp report_type khác qua đường tất định (deriver CĐKT/CĐPS/tồn kho/công nợ…),
             caller PHẢI chạy tiếp sau khi spec xong, không được return.

    Đánh dấu bằng `doc_tron_file: false` trong chính spec — mặc định True để 26 spec hiện có giữ
    NGUYÊN hành vi cũ, chỉ spec nào tự khai "tôi chỉ bóc một lát" mới đổi. Suy đoán tự động (vd
    "file có sheet CĐKT thì chắc là BCTC") nghe tiện nhưng sai ở cả hai chiều và không ai đọc code
    ra được vì sao một file lại đi hai đường; khai tường minh trong spec thì đọc spec là biết.

    Xem agent_cli._cmd_autofill_impl (ca SRVF T05/T06 mất 13 report_type, 14/08/2026).
    """
    specs = specs_for_path(path)
    return all(sp.get("doc_tron_file", True) for sp in specs) if specs else True


def run_for_path(path, write=False):
    """Chạy mọi spec phụ trách file này. Trả [] nếu KHÔNG spec nào phụ trách -> caller đi đường cũ.

    Từng spec bọc riêng try/except: một spec lỗi không được kéo theo các spec còn lại, và tuyệt
    đối không được ném ra ngoài — hàm này nằm trên đường nạp CHUNG của mọi báo cáo.
    """
    ket_qua = []
    for sp in specs_for_path(path):
        # BỎ QUA BẢN ĐÃ BỊ THAY THẾ. Hàm này nạp ĐÚNG file được đưa vào, không nhìn sang các file
        # khác cùng kỳ — nên "Nạp lại tất cả" (sync_orchestrator reprocess) sẽ ghi lại dòng của
        # bản CŨ mà lượt quét cả thư mục đã cố tình loại, và kỳ đó lập tức đếm đôi trở lại. Đúng
        # 3.029 dòng claim T3-T6 vừa phải xoá tay ngày 12/08/2026 sẽ sống dậy chỉ bằng một cú bấm.
        if sp.get("moi_ky_lay_file_moi_nhat"):
            try:
                giu, _w = quet_nguon(sp)
                giu, _bo = loc_file_moi_nhat(sp, giu)
                if os.path.abspath(path) not in {os.path.abspath(x) for x in giu}:
                    ket_qua.append({"file": os.path.basename(path), "dong": 0,
                                    "spec": sp.get("id"), "report_type": sp.get("report_type"),
                                    "canh_bao": ["BỎ QUA — đã có bản MỚI HƠN của cùng kỳ; nạp bản "
                                                 "này vào là đếm đôi. Xoá tài liệu của nó thay vì nạp."]})
                    continue
            except Exception:                                      # noqa: BLE001
                pass          # không kiểm được thì cứ chạy như cũ, đừng chặn đường nạp
        try:
            r = run(sp, path, write=write)
        except Exception as e:                                     # noqa: BLE001
            r = {"file": os.path.basename(path), "dong": 0,
                 "canh_bao": [f"LỖI spec {sp.get('id')}: {type(e).__name__}: {e}"]}
        r["spec"] = sp.get("id")
        r["report_type"] = sp.get("report_type")
        ket_qua.append(r)
    return ket_qua


def _chon_sheet(wb, cfg, thang=None):
    """Chọn sheet theo cfg. `theo_thang` (vd "T{mm}") dành cho file có MỘT SHEET MỖI THÁNG.

    BCTC của SRVF là ca đó: cùng một file tháng 7 chứa cả T01..T07, mỗi sheet là một tháng. Lấy
    sheet đầu hay khớp theo tiền tố "T" đều sai — "T01." và "T01" chuẩn hoá về cùng chuỗi, còn
    "T01BC"/"TSCD" cũng bắt đầu bằng T. Phải neo theo ĐÚNG tháng của kỳ đang nạp và khớp CHÍNH XÁC.
    """
    cfg = cfg or {}
    if "so" in cfg:
        return wb.sheetnames[int(cfg["so"])]
    if "theo_thang" in cfg:
        if not thang:
            return None
        muon = _nd(str(cfg["theo_thang"]).replace("{mm}", f"{int(thang):02d}"))
        # ƯU TIÊN khớp tuyệt đối tên gốc trước, vì "T01." và "T01" cùng _nd() -> "t01".
        for name in wb.sheetnames:
            if str(name).strip() == str(cfg["theo_thang"]).replace("{mm}", f"{int(thang):02d}"):
                return name
        for name in wb.sheetnames:
            if _nd(name) == muon:
                return name
        return None
    # `ten` nhận LIST tên ứng viên, thử theo ĐÚNG THỨ TỰ KHAI (không theo thứ tự sheet trong
    # file). Công nợ phải thu SRVF đổi tên sheet chi tiết giữa các kỳ ("Chi tiết theo hợp đồng"
    # T1-T4/T6/T7 -> "Chi tiết công nợ" T5/T8) và T7 còn có một BẢN SAO tên "Chi tiết theo hợp
    # đồng (2)" lệch 1 cột, cột COC rỗng sạch. Khớp kiểu `chua: "Chi tiết"` sẽ vớ đúng bản sao đó
    # vì nó là sheet đầu tiên trong file -> ra một tháng toàn "chưa có COC" mà không lỗi gì.
    ung_vien = cfg["ten"] if isinstance(cfg.get("ten"), list) else \
        ([cfg["ten"]] if "ten" in cfg else [])
    for t in ung_vien:
        for name in wb.sheetnames:
            if _nd(name) == _nd(t):
                return name
    for name in wb.sheetnames:
        n = _nd(name)
        if "batdau" in cfg and n.startswith(_nd(cfg["batdau"])):
            return name
        if "chua" in cfg and _nd(cfg["chua"]) in n:
            return name
    return None


def _map_header(rows, dong):
    """{tên header đã chuẩn hoá: chỉ số cột 0-based}. `dong` là số hoặc list số (1-based).

    Gộp nhiều dòng: lấy ô KHÁC RỖNG ĐẦU TIÊN theo thứ tự khai — file Claim để nhóm cột ở dòng 1
    ("XE"/"PIN") còn tên cột thật ở dòng 2 (Z "SR", AA "Trạng thái", AB "Số tiền").
    """
    dongs = dong if isinstance(dong, list) else [dong]
    ncol = max((len(rows[d - 1]) for d in dongs if d - 1 < len(rows)), default=0)
    out, theo_dong = {}, {d: {} for d in dongs}
    for j in range(ncol):
        xong = False
        for d in dongs:
            r = rows[d - 1] if d - 1 < len(rows) else ()
            v = r[j] if j < len(r) else None
            n = _nd(v)
            if n:
                theo_dong[d].setdefault(n, j)
                if not xong:
                    out.setdefault(n, j)
                    xong = True
    return out, theo_dong


def _resolve_cot(spec, hmap, warn):
    """{đích: (chỉ số cột, cfg)} — dò theo TÊN, phao là `cot_du_phong` (chữ cột)."""
    out = {}
    for dich, cfg in (spec.get("cot") or {}).items():
        j = _tim_cot(hmap, cfg, dich, warn)
        if j is None:
            continue
        out[dich] = (j, cfg)
    return out


def _tim_cot(hmaps, cfg, nhan, warn):
    """Chỉ số cột cho 1 khai báo. `header` nhận CHUỖI hoặc LIST tên ứng viên (thử theo thứ tự).

    Vì sao cần list: cùng một cột nhưng file đổi tên giữa các kỳ — báo cáo Claim ghi "Check" ở
    T1/T2 rồi đổi thành "Trạng thái" từ T3. Khai `["Check", "Trạng thái"]` chạy được cả 6 file.

    Vì sao cần `dong_header`: T1/T2 còn có một cột MỒI ở dòng 1 cũng tên "Trạng thái" nhưng chứa
    ĐỊA CHỈ khách. Không giới hạn dòng header thì trạng thái claim biến thành "Thôn Xám, Xã Vĩnh
    Phú…" — sai mà vẫn chạy trơn. Khai `"dong_header": 2` để chỉ dò trên đúng dòng tiêu đề thật.
    """
    hmap, theo_dong = hmaps
    if cfg.get("dong_header"):
        hmap = theo_dong.get(cfg["dong_header"], {})
    ten = cfg.get("header")
    ung_vien = ten if isinstance(ten, list) else ([ten] if ten else [])
    for t in ung_vien:
        j = hmap.get(_nd(t))
        if j is not None:
            return j
    if cfg.get("cot") or cfg.get("cot_du_phong"):
        chu = cfg.get("cot") or cfg["cot_du_phong"]
        if cfg.get("cot_du_phong") and ung_vien:
            warn.append(f"cột '{ung_vien[0]}' không thấy theo tên -> dùng phao {chu}")
        return column_index_from_string(chu) - 1
    if cfg.get("bat_buoc", True):
        warn.append(f"THIẾU cột bắt buộc '{ten}' ({nhan})")
    return None


def _thang_cuoi(v, nam):
    """Ô chứa SỐ THÁNG (1..12) -> 'YYYY-MM-DD' ngày cuối tháng đó, năm lấy từ tên file.

    Dùng cho báo cáo xếp 12 DÒNG THÁNG (QTVH An Taxi: một file cấp số cho cả năm nên kỳ KHÔNG
    suy được từ tên file). Neo CUỐI THÁNG để `period_month` rơi đúng tháng — cùng quy ước với
    `moi_cot_thang` của bản kế hoạch. Ô là chữ ("Quý", "06 Tháng đầu năm") -> None, nhờ vậy các
    dòng tổng quý/nửa năm ngay dưới dải 12 tháng tự bị bộ lọc `ngay khác_rong` loại đi thay vì
    cộng đôi vào một kỳ nào đó.
    """
    if nam is None:
        return None
    m = re.match(r"^\s*(\d{1,2})(?:\.0)?\s*$", str(v if v is not None else "").strip())
    if not m or not 1 <= int(m.group(1)) <= 12:
        return None
    mo = int(m.group(1))
    return dt.date(nam, mo, calendar.monthrange(nam, mo)[1]).isoformat()


def _ngay_trong_thang(v, ky):
    """Ô chứa SỐ NGÀY (1..31) -> 'YYYY-MM-DD', (năm, tháng) lấy từ TÊN FILE (`_ky_thang`).

    Bản THEO DÒNG của `moi_cot_ngay` (báo cáo ngày An Taxi xếp mỗi ngày MỘT DÒNG, cột A là số
    ngày). Ngày 30/31 ở tháng ngắn -> None (dòng thừa của mẫu in sẵn), không phải lỗi.
    """
    if not ky:
        return None
    m = re.match(r"^\s*(\d{1,2})(?:\.0)?\s*$", str(v if v is not None else "").strip())
    if not m:
        return None
    try:
        return dt.date(ky[0], ky[1], int(m.group(1))).isoformat()
    except ValueError:
        return None


def _ngay_dd_mm(v, ky):
    """Ô ghi 'dd/mm' CÓ ĐUÔI -> 'YYYY-MM-DD'. Năm lấy từ tên file (`_ky_thang`); THÁNG trong ô
    phải khớp tháng của file, lệch thì trả None.

    Anh em theo-DÒNG của phần `moi_cot_ngay` đọc tiêu đề 'dd/mm (thứ)': kế hoạch xuất hoá đơn 15
    ngày xếp mỗi ngày một DÒNG ở sheet tổng hợp ('01/09 (Thứ 3)', '05/09 (Thứ 7 - Cao điểm)') và
    mỗi dòng xe một CỘT, nên chiều ngày nằm ở cột A chứ không ở tiêu đề.

    Khác `ngay_trong_thang` ở chỗ ô có SẴN tháng: phải kiểm nó, vì một file kế hoạch tràn sang
    tháng sau mà vẫn nạp theo tháng của tên file là gán số vào sai kỳ, im lặng.
    """
    if not ky:
        return None
    m = re.match(r"^\s*(\d{1,2})\s*/\s*(\d{1,2})\b", str(v if v is not None else "").strip())
    if not m or int(m.group(2)) != ky[1]:
        return None
    try:
        return dt.date(ky[0], ky[1], int(m.group(1))).isoformat()
    except ValueError:
        return None


def _lay_o(row, j, cfg, dem_loi=None):
    v = row[j] if j < len(row) else None
    if _o_loi(v):
        if dem_loi is not None:
            k = str(v).strip().upper()
            dem_loi[k] = dem_loi.get(k, 0) + 1
        v = None
    kieu = cfg.get("kieu", "text")
    if kieu == "so":
        return _so(v, float(cfg.get("he_so", 1.0)))
    if kieu == "so_co_rong":
        return _so_co_rong(v, float(cfg.get("he_so", 1.0)))
    if kieu == "date":
        return _date(v)
    if kieu == "thang_cuoi":
        return _thang_cuoi(v, cfg.get("_nam"))
    if kieu == "ngay_trong_thang":
        return _ngay_trong_thang(v, cfg.get("_ky"))
    if kieu == "ngay_dd_mm":
        return _ngay_dd_mm(v, cfg.get("_ky"))
    s = str(v).strip() if v is not None else None
    return s or None


def _dat(rec, dich, val):
    if dich.startswith("payload."):
        rec.setdefault("payload", {})[dich[len("payload."):]] = val
    else:
        rec[dich] = val


def _lay(rec, dich):
    if dich.startswith("payload."):
        return (rec.get("payload") or {}).get(dich[len("payload."):])
    return rec.get(dich)


def _qua_loc(rec, loc):
    for f in loc or []:
        v = _lay(rec, f["cot"])
        dk, gt = f.get("dieu_kien", "khac_rong"), f.get("gia_tri")
        if dk == "khac_rong" and (v is None or v == "" or v == 0):
            return False
        if dk == "bang" and str(v) != str(gt):
            return False
        if dk == "khac" and str(v) == str(gt):
            return False
        if dk == "thuoc" and str(v) not in [str(x) for x in gt]:
            return False
        if dk == "khong_thuoc" and str(v) in [str(x) for x in gt]:
            return False
        if dk == "lon_hon" and not (isinstance(v, (int, float)) and v > gt):
            return False
        if dk == "nho_hon" and not (isinstance(v, (int, float)) and v < gt):
            return False
        if dk == "chua" and _nd(gt) not in _nd(v):
            return False
    return True


def _qua_loc_o(row, dieu_kien):
    """Bản `_qua_loc` chạy THẲNG trên ô của dòng nguồn (địa chỉ theo CHỮ CỘT), dùng cho
    `ngu_canh_dong` — lúc đó bản ghi chưa được dựng nên chưa có trường nào để so.

    `khong_regex` là điều kiện chủ lực: dòng tiêu đề đơn vị nhận biết bằng "cột A có chữ nhưng
    KHÔNG phải mã chỉ tiêu" (mã dạng B110/B120…). Không có nó thì các dòng chi tiết không mã
    (Doanh thu công gò/công sơn) cũng bị nhận nhầm là tiêu đề đơn vị.
    """
    if not dieu_kien:
        return False
    for f in dieu_kien:
        j = column_index_from_string(f["cot"]) - 1
        v = row[j] if j < len(row) else None
        s = str(v).strip() if v is not None else ""
        dk, gt = f.get("dieu_kien", "khac_rong"), f.get("gia_tri")
        if dk == "khac_rong" and not s:
            return False
        if dk == "rong" and s:
            return False
        if dk == "bang" and s != str(gt):
            return False
        if dk == "khac" and s == str(gt):
            return False
        if dk == "regex" and not re.search(gt, s):
            return False
        if dk == "khong_regex" and re.search(gt, s):
            return False
        if dk == "chua" and _nd(gt) not in _nd(s):
            return False
    return True


def _dan_xuat(rec, cong_thuc):
    """eval trong sandbox hẹp — spec là cấu hình nội bộ tin cậy (xem docstring đầu file)."""
    if not cong_thuc:
        return
    env = {k: v for k, v in rec.items() if isinstance(v, (int, float))}
    env.update({f"payload.{k}": v for k, v in (rec.get("payload") or {}).items()})
    safe = {re.sub(r"\W", "_", k): (v or 0) for k, v in env.items()}
    for dich, bt in cong_thuc.items():
        try:
            # Đổi `payload.x` -> `payload_x` NHƯNG CHỈ TRONG TÊN BIẾN. Bản cũ chạy
            # `re.sub(r"\W", "_", bt)` trên CẢ biểu thức nên thay luôn dấu cách và toán tử:
            # "payload.a + payload.b" biến thành MỘT tên biến "payload_a___payload_b" -> NameError
            # -> `dan_xuat` trả None mà không ai thấy (chỉ ghi vào `rec["_loi"]`). Công thức một
            # biến hoặc không có dấu chấm thì cả hai cách cho kết quả như nhau, nên đổi ở đây
            # không ảnh hưởng spec cũ (đã rà: chỉ `amount + amount2`, `1`, `amount - payload_gia_von`).
            _dat(rec, dich, eval(  # noqa: S307
                re.sub(r"\b[A-Za-z_][\w.]*\b", lambda m: m.group(0).replace(".", "_"), bt),
                {"__builtins__": {}}, safe))
        except Exception as ex:                                    # noqa: BLE001
            _dat(rec, dich, None)
            rec.setdefault("_loi", []).append(f"{dich}: {ex}")


def _doi_gia_tri(rec, cfg):
    """Đổi giá trị một trường theo bản đồ khai trong spec — dùng khi ĐỔI NGUỒN nhưng phải GIỮ
    NGUYÊN hình dạng bản ghi của nguồn cũ.

    Bối cảnh: mapping VHKD chốt "file A bị thay bởi file B". File B mang mã riêng của nó (A100,
    A300, A600…) trong khi mọi màn hình đang đọc mã của file A ('1000', '1047', '1112'). Không có
    khoá này thì phải viết một hook Python cho MỖI report_type — bốn hook gần giống hệt nhau chỉ
    khác vài dòng bản đồ.

    Giá trị KHÔNG có trong bản đồ được GIỮ NGUYÊN (không bỏ trắng): bản đồ ở đây là phép ĐỔI TÊN,
    không phải bộ lọc. Muốn giữ đúng vài mã thì lọc bằng `loc` với điều kiện `thuoc` — tách hai
    việc ra để đọc spec là thấy ngay mã nào lên, mã nào không.
    """
    for truong, bando in (cfg or {}).items():
        v = _lay(rec, truong)
        if v is None:
            continue
        k = str(v).strip()
        if k in bando:
            _dat(rec, truong, bando[k])


def _tinh_han_no(rec, cfg, ngay_file):
    """Suy `đến hạn` + `số ngày quá hạn` cho nguồn công nợ KHÔNG có sẵn hai cột đó.

    Bản công nợ TAY (`SRVF/baocaocongnophaithu`) do kế toán tính sẵn hai cột này trong sheet chi
    tiết. Bản TỰ ĐỘNG của Cyber (`TEST_SR/baocaocongnophaithungay`) chỉ có `Ngày hóa đơn` — đúng
    quy ước nghiệp vụ đã đối chiếu 03/09/2026: `đến hạn` = Ngày hoá đơn + 15 ngày (khớp 472/472
    dòng của bản tay), `số ngày quá hạn` = ngày chốt của FILE − đến hạn.

    Vì sao là một khoá riêng chứ không phải `dan_xuat`: `dan_xuat` chạy `eval` trên môi trường
    CHỈ CÓ SỐ (`isinstance(v, (int, float))`) nên không đụng được vào ngày, và nó cũng không biết
    ngày chốt của file. Hai thứ đó đều bắt buộc ở đây.

    Chưa tới hạn -> `so_ngay_qua_han` ÂM (giữ nguyên dấu, không kẹp về 0): hook `qua_han` phân loại
    theo `> 0` nên số âm ra "Trong hạn", đúng như ô rỗng của bản tay; giữ dấu để sau này dựng được
    "còn mấy ngày tới hạn" mà không phải đọc lại nguồn.
    """
    if not cfg:
        return
    hd = _lay(rec, cfg.get("ngay_hoa_don", "payload.ngay_hoa_don"))
    d_hd = _date(hd)
    dich_dh = cfg.get("den_han", "payload.den_han")
    dich_qh = cfg.get("so_ngay_qua_han", "payload.so_ngay_qua_han")
    if not d_hd:
        # Không có ngày hoá đơn -> để trống chứ KHÔNG đoán. Dòng vẫn vào DB (số dư vẫn đúng),
        # chỉ mất chiều tuổi nợ — giống hệt cách bản tay xử ô "đến hạn" bỏ trắng.
        _dat(rec, dich_dh, None)
        _dat(rec, dich_qh, None)
        if cfg.get("phan_loai"):
            _dat(rec, cfg["phan_loai"], _qua_han(None))
        return
    y, mo, d = (int(x) for x in d_hd.split("-"))
    den_han = dt.date(y, mo, d) + dt.timedelta(days=int(cfg.get("cong_ngay", 15)))
    _dat(rec, dich_dh, den_han.isoformat())
    moc = _date(cfg.get("_moc") or ngay_file)
    so_ngay = None
    if moc:
        ym, mm, dd = (int(x) for x in moc.split("-"))
        so_ngay = (dt.date(ym, mm, dd) - den_han).days
    _dat(rec, dich_qh, so_ngay)
    if cfg.get("phan_loai"):
        _dat(rec, cfg["phan_loai"], _qua_han(so_ngay))


def _tuan_truoc(bc):
    """Ngày phát hành báo cáo -> (thứ Hai, Chủ nhật) của TUẦN LIỀN TRƯỚC.

    Không suy từ số tuần trong tên file (8.1 / 8.2 — "tuần thứ mấy của tháng" mỗi nơi đếm một
    kiểu) mà suy từ chính ngày phát hành: lùi về thứ Hai của tuần chứa nó rồi lùi tiếp 7 ngày.
    Báo cáo ra đúng thứ Hai 03/08 -> tuần 27/07..02/08; ra muộn tới thứ Tư 05/08 vẫn ra đúng
    tuần đó, nên file về trễ không nhảy kỳ.
    """
    thu_hai = bc - dt.timedelta(days=bc.weekday())
    return thu_hai - dt.timedelta(days=7), thu_hai - dt.timedelta(days=1)


def chu_ky_tuan(spec, path):
    """-> {'ngay_bao_cao','tuan_tu','tuan_den'} để nhét vào payload, hoặc {} nếu spec không khai.

    Số trong file là DƯ NỢ TẠI THỜI ĐIỂM chốt tuần, không phải phát sinh trong tuần. Giao diện
    phải nói được "ảnh chụp tuần nào, phát hành ngày nào" — thiếu ba mốc này thì người xem không
    phân biệt được số của tuần trước với số mới nhất.
    """
    c = spec.get("ky_tu_ten_file") or {}
    if not (c.get("regex_ngay") and c.get("tuan_truoc_ngay_bao_cao")):
        return {}
    ten = os.path.basename(path)
    mn, mt = re.search(c["regex_nam"], ten), re.search(c["regex_thang"], ten)
    md = re.search(c["regex_ngay"], ten)
    if not (mn and mt and md):
        return {}
    try:
        bc = dt.date(int(mn.group(1)), int(mt.group(1)), int(md.group(1)))
    except ValueError:
        return {}
    tu, den = _tuan_truoc(bc)
    return {"ngay_bao_cao": bc.isoformat(), "tuan_tu": tu.isoformat(), "tuan_den": den.isoformat()}


def ngay_tu_ten_file(spec, path):
    """-> ('YYYY-MM-DD' | None, [cảnh báo]). Dùng cho bảng ẢNH CHỤP (không có cột ngày từng dòng).

    `ky_tu_ten_file` tách RIÊNG regex năm và tháng — file công nợ đặt tên
    "…M.2026.07.22_Baocaocongnophaithu_T1.xlsx": cụm 2026.07.22 là NGÀY LẬP báo cáo (giống hệt
    nhau ở mọi kỳ), kỳ thật nằm ở hậu tố _T1/_T7. Lấy nhầm cụm đầu là dồn cả 12 tháng vào tháng 7.
    `ngay_tu_ten_file` dùng khi tên file có ngày chốt thật (kho xe B2B: "…2026.7.31.KHO XE.xlsx").

    `regex_ngay` (tuỳ chọn, thêm 12/08/2026): tên file báo cáo TUẦN của KSCL ghi đủ
    "W.<năm>.<tháng>.<tuần>.<ngày báo cáo>" — vd "W.2026.8.2.10" = tuần 2 tháng 8, phát hành
    ngày 10/08. Trước đây chỉ đọc năm+tháng rồi gán CUỐI THÁNG, nên mọi bản tuần của một tháng
    đổ chung vào 31/08: lọc theo tuần không ra gì, và nhiều tuần thì phải vứt bớt file. Có
    `regex_ngay` thì kỳ lấy đúng theo ngày (xem thêm `tuan_truoc_ngay_bao_cao`).
    """
    ten = os.path.basename(path)
    warn = []
    if spec.get("ky_tu_ten_file"):
        c = spec["ky_tu_ten_file"]
        mn, mt = re.search(c["regex_nam"], ten), re.search(c["regex_thang"], ten)
        md = re.search(c["regex_ngay"], ten) if c.get("regex_ngay") else None
        if mn and mt and c.get("regex_ngay") and not md:
            warn.append(f"không dò được NGÀY báo cáo từ tên file: {ten} — tạm lấy cuối tháng")
        if mn and mt and md:
            try:
                bc = dt.date(int(mn.group(1)), int(mt.group(1)), int(md.group(1)))
            except ValueError:
                warn.append(f"ngày báo cáo trong tên file không hợp lệ: {md.group(0)}")
            else:
                if not c.get("tuan_truoc_ngay_bao_cao"):
                    return bc.isoformat(), warn
                tu, den = _tuan_truoc(bc)
                if bc.weekday() != 0:
                    # Quy ước: báo cáo phát hành THỨ HAI, chốt số của tuần liền trước. File ra
                    # muộn 1-2 hôm vẫn quy về đúng tuần đó (lùi về thứ Hai của tuần chứa nó), còn
                    # ra vào cuối tuần thì rất có thể là chốt CHÍNH tuần đang chạy -> phải kêu,
                    # đừng lặng lẽ gán nhầm một tuần.
                    warn.append(f"ngày báo cáo {bc.isoformat()} KHÔNG phải thứ Hai "
                                f"({bc.strftime('%A')}) — vẫn quy về tuần {tu}..{den}, kiểm tra lại")
                return den.isoformat(), warn
        if mn and mt:
            y, mo = int(mn.group(1)), int(mt.group(1))
            return dt.date(y, mo, calendar.monthrange(y, mo)[1]).isoformat(), warn
        warn.append(f"không dò được kỳ (năm/tháng) từ tên file: {ten}")
    if spec.get("ngay_tu_ten_file"):
        m = re.search(spec["ngay_tu_ten_file"]["regex"], ten)
        if m:
            try:
                return dt.date(*(int(m.group(i + 1)) for i in range(3))).isoformat(), warn
            except ValueError:
                warn.append(f"ngày trong tên file không hợp lệ: {m.group(0)}")
        else:
            warn.append(f"không dò được ngày từ tên file: {ten}")
    return None, warn


def _ky_thang(spec, path):
    """-> ((năm, tháng), [cảnh báo]) cho `ban_ghi = moi_cot_ngay`. Kỳ lấy từ TÊN FILE
    (vd '...202608.Baocaodoanhthungay.xlsx'), vì các cột chỉ ghi số ngày, không ghi tháng."""
    c = spec.get("ky_thang_tu_ten_file") or {"regex": r"\.(\d{4})(\d{2})\."}
    m = re.search(c["regex"], os.path.basename(path))
    if not m:
        return None, [f"không dò được kỳ (năm/tháng) từ tên file: {os.path.basename(path)}"]
    # NHÓM CÓ TÊN `nam`/`thang` (04/09/2026) — cho nguồn ghi THÁNG TRƯỚC NĂM trong tên file
    # ('Kế hoạch xhđ 15 ngày T9.2026.xlsx'). Mặc định vẫn là nhóm 1 = năm, nhóm 2 = tháng, nên
    # mọi spec cũ không đổi một ly. Không có đường nào khác: thứ tự nhóm của regex là thứ tự
    # xuất hiện trong chuỗi, không đảo được.
    g = m.groupdict()
    if g.get("nam") and g.get("thang"):
        return (int(g["nam"]), int(g["thang"])), []
    return (int(m.group(1)), int(m.group(2))), []


def _chuyen_xls_cu(duong_dan):
    """`.xls` đời cũ (BIFF) -> sinh bản `.xlsx` cạnh nó, trả đường dẫn mới (None nếu hỏng).

    openpyxl KHÔNG đọc được định dạng Excel 97-2003; nghiệp vụ thỉnh thoảng lưu nhầm (công nợ
    phải thu T5 gửi 12/08/2026). Không có bước này thì file nằm im, không dòng nào vào DB và
    KHÔNG có cảnh báo nào — đúng kiểu thiếu số mà không ai biết.

    Bản chuyển đổi lưu CẠNH file gốc (cùng thư mục -> `source_file` trong DB vẫn là
    "<FOLDER>::<tên>.xlsx" đúng quy ước) và dùng lại ở lần quét sau, chỉ chuyển khi thiếu hoặc
    cũ hơn bản .xls.

    Dùng xlrd + openpyxl chứ KHÔNG gọi `soffice`: máy này chỉ cài libreoffice-core/common, thiếu
    hẳn gói `libreoffice-calc` nên soffice trả "source file could not be loaded" với MỌI bảng
    tính, kể cả .xlsx hợp lệ (đã thử 12/08/2026). Đường thuần Python còn khỏi cần sudo để cài gói
    và khỏi đẻ tiến trình con giữa lượt nạp.

    Chỉ bê GIÁ TRỊ ô sang, không giữ định dạng — engine chỉ đọc giá trị (`data_only=True`).
    """
    dich = os.path.splitext(duong_dan)[0] + ".xlsx"
    if os.path.exists(dich) and os.path.getmtime(dich) >= os.path.getmtime(duong_dan):
        return dich
    try:
        import xlrd
    except ImportError:
        return None
    try:
        nguon = xlrd.open_workbook(duong_dan)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for sh in nguon.sheets():
            # Tên sheet: Excel giới hạn 31 ký tự và cấm : \ / ? * [ ]
            ws = wb.create_sheet(title=re.sub(r"[:\\/?*\[\]]", "-", sh.name)[:31] or "Sheet")
            for r in range(sh.nrows):
                for c in range(sh.ncols):
                    o = sh.cell(r, c)
                    if o.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK, xlrd.XL_CELL_ERROR):
                        continue
                    v = o.value
                    if o.ctype == xlrd.XL_CELL_DATE:
                        try:
                            v = dt.datetime(*xlrd.xldate_as_tuple(v, nguon.datemode))
                        except Exception:                              # noqa: BLE001
                            pass                                       # số ngày hỏng -> để nguyên
                    elif o.ctype == xlrd.XL_CELL_BOOLEAN:
                        v = bool(v)
                    ws.cell(row=r + 1, column=c + 1, value=v)
        wb.save(dich)
    except Exception:                                                  # noqa: BLE001
        return None
    return dich


def _chuyen_xlsb(duong_dan):
    """`.xlsb` (Excel nhị phân) -> sinh bản `.xlsx` cạnh nó, trả đường dẫn mới (None nếu hỏng).

    Cùng lý do và cùng cách làm với `_chuyen_xls_cu`: openpyxl KHÔNG đọc được `.xlsb`, mà nguồn
    QLTS (`B.9.TC.H5.M.<yyyy>.<m>.Taisan.xlsb`, 17/08/2026) lại lưu ở đúng định dạng đó — 21k dòng
    sheet "CHI TIẾT". Không có bước này thì file nằm im, không dòng nào vào DB và KHÔNG có cảnh báo.

    Dùng `pyxlsb` thuần Python, KHÔNG gọi `soffice` (máy này thiếu `libreoffice-calc`, xem ghi chú
    dài ở `_chuyen_xls_cu`).

    NGÀY THÁNG: `.xlsb` trả ngày dưới dạng SỐ SERIAL của Excel (vd 46226 = 2026-08-01) và pyxlsb
    không cho biết ô nào đang định dạng ngày. Bản chuyển vì vậy giữ nguyên SỐ — spec đọc nguồn này
    phải lấy kỳ từ TÊN FILE (`ky_tu_ten_file`), đừng khai `"kieu": "date"` cho cột lấy từ `.xlsb`.
    """
    dich = os.path.splitext(duong_dan)[0] + ".xlsx"
    if os.path.exists(dich) and os.path.getmtime(dich) >= os.path.getmtime(duong_dan):
        return dich
    try:
        from pyxlsb import open_workbook as _mo_xlsb
    except ImportError:
        return None
    try:
        from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        with _mo_xlsb(duong_dan) as nguon:
            for ten_sh in nguon.sheets:
                ws = wb.create_sheet(title=re.sub(r"[:\\/?*\[\]]", "-", ten_sh)[:31] or "Sheet")
                with nguon.get_sheet(ten_sh) as sh:
                    for row in sh.rows():
                        for o in row:
                            v = o.v
                            if v is None or v == "":
                                continue
                            if isinstance(v, str):
                                # Ô kế toán gõ tay hay dính ký tự điều khiển (vd tên tài sản
                                # "Camera quan sát …\x07gồm thẻ nhớ"). XML của .xlsx cấm hẳn nhóm
                                # này -> openpyxl ném IllegalCharacterError và HỎNG CẢ FILE, chỉ
                                # vì một ô. Bỏ ký tự đó đi, phần chữ giữ nguyên.
                                v = ILLEGAL_CHARACTERS_RE.sub("", v)
                                if not v:
                                    continue
                            ws.cell(row=o.r + 1, column=o.c + 1, value=v)
        wb.save(dich)
    except Exception:                                                  # noqa: BLE001
        return None
    return dich


def quet_nguon(spec):
    """Danh sách file của nguồn -> [đường dẫn], [cảnh báo].

    KHỚP ĐUÔI KHÔNG PHÂN BIỆT HOA/THƯỜNG. `glob` phân biệt hoa thường trên Linux, nên
    `"*.xlsx"` bỏ qua sạch `.Xlsx` — nghiệp vụ đặt tên lẫn lộn cả hai kiểu. Đã mất 4/7 kỳ công
    nợ phải thu SRVF (T2/T3/T4/T6 đều `.Xlsx`) suốt từ lúc dựng nguồn tới 12/08/2026, không một
    dòng cảnh báo: file có trên đĩa, spec chạy "thành công", màn hình chỉ trống mấy tháng giữa.
    """
    warn = []
    thu_muc = os.path.join(REPORTS_DIR, (spec.get("nguon") or {}).get("folder", ""))
    if not os.path.isdir(thu_muc):
        return [], [f"không có thư mục nguồn: {thu_muc}"]
    mau = ((spec.get("nguon") or {}).get("file_glob") or "*.xlsx").lower()
    ten = sorted(os.listdir(thu_muc))
    # File tạm Excel sinh ra khi ai đó đang MỞ file trên máy chia sẻ. Khớp "*.xlsx" nhưng không
    # phải workbook thật -> đọc vào là ném lỗi khó hiểu giữa lượt nạp.
    ten = [n for n in ten if not n.startswith("~$")]
    out = [os.path.join(thu_muc, n) for n in ten if fnmatch.fnmatch(n.lower(), mau)]
    if mau.endswith(".xlsx"):
        da_co = {os.path.splitext(p)[0].lower() for p in out}
        for n in ten:
            g = os.path.join(thu_muc, n)
            if n.lower().endswith(".xls") and os.path.splitext(g)[0].lower() not in da_co:
                moi = _chuyen_xls_cu(g)
                if moi:
                    warn.append(f"{n}: Excel 97-2003, đã chuyển sang {os.path.basename(moi)}")
                    out.append(moi)
                else:
                    warn.append(f"{n}: Excel 97-2003 và CHUYỂN ĐỔI HỎNG — file này chưa vào DB")
    # `.xlsb` (Excel nhị phân) -> chuyển sang `.xlsx` rồi đọc bản đã chuyển. Đặt SAU nhánh `.xls`
    # để cả hai định dạng lạ đều quy về một đường: engine phía dưới chỉ biết openpyxl.
    doi = []
    for p in out:
        if p.lower().endswith(".xlsb"):
            moi = _chuyen_xlsb(p)
            if moi:
                doi.append(moi)
            else:
                warn.append(f"{os.path.basename(p)}: .xlsb và CHUYỂN ĐỔI HỎNG — file này chưa vào DB")
        else:
            doi.append(p)
    return sorted(doi), warn


def loc_file_moi_nhat(spec, files):
    """Mỗi KỲ chỉ giữ file có ngày chốt MỚI NHẤT.

    Kho xe B2B có 2 ảnh chụp cùng tháng 7 ("…7.24.KHO XE" và "…7.31.KHO XE"). Nạp cả hai là
    ĐẾM ĐÔI tồn kho tháng 7 (mỗi VIN xuất hiện 2 lần) — số tồn phình gần gấp đôi mà không có
    cảnh báo nào. Bật `"moi_ky_lay_file_moi_nhat": true` trong spec cho mọi nguồn ảnh chụp.

    `"moi_ky_lay_file_moi_nhat": "ngay"` — gộp theo NGÀY CHỐT thay vì theo tháng, cho nguồn báo
    cáo TUẦN (KSCL): các bản tuần trong cùng một tháng là những ảnh chụp KHÁC NHAU, phải giữ đủ
    để lọc theo tuần ra được số của đúng tuần đó. Gộp theo tháng ở đây là vứt mất tuần cũ. Việc
    không cộng đôi khi xem cả tháng do tầng đọc lo (`_SNAP_RT` -> chỉ lấy ngày chốt mới nhất).

    `"moi_ky_lay_file_moi_nhat": "mot_file"` (14/08/2026) — giữ ĐÚNG MỘT file mới nhất trong CẢ
    thư mục, bất kể kỳ. Cho nguồn mà MỖI FILE THÁNG ĐÃ CHỨA TOÀN BỘ CÁC KỲ TỪ ĐẦU NĂM: sheet
    "Chi tiết" của báo cáo nhân sự HCNS có đủ 6 tháng trong file tháng 1 lẫn file tháng 6. Hai chế
    độ trên không cứu được vì mỗi file suy ra một kỳ KHÁC NHAU nên đều được giữ -> nạp 6 file là
    ghi cùng dữ liệu 6 lần dưới 6 `source_file`, và `_ghi` (xoá theo source_file) không chặn nổi:
    tổng nhân sự phình 6 lần.

    `"moi_ky_slot_regex"` (29/08/2026) — MỘT KỲ CÓ NHIỀU LÁT SONG SONG, mỗi lát vẫn phải giữ bản
    mới nhất của RIÊNG nó. Không có nó thì cờ trên là con dao hai lưỡi: `vhkd_tonkho_vatly` có
    3 kênh (B2B/B2C/GF) × nhiều ảnh chụp trong tháng 8, `qlts_dau_bqnl` có 4 trạm cùng kỳ tháng 7
    — bật cờ trần là giữ ĐÚNG MỘT file rồi vứt 2 kênh / 3 trạm dữ liệu thật, im lặng. Khai regex
    bắt phần định danh lát (kênh, trạm…); các nhóm bắt được ghép vào khoá gộp. KHÔNG khớp -> lát
    đứng RIÊNG một mình (theo tên file), không bị gộp vào lát khác: tên lạ mà gộp nhầm là mất
    nguyên một lát. Cùng quy ước với `slot` của `cron_qtvh_core._slot`.
    """
    che_do = spec.get("moi_ky_lay_file_moi_nhat")
    if not che_do:
        return files, []
    pat = spec.get("moi_ky_slot_regex")
    ky_pat = spec.get("moi_ky_ky_regex")
    giu, bo = {}, []
    for f in files:
        ngay, _ = ngay_tu_ten_file(spec, f)
        tu_ky_pat = False
        if not ngay and ky_pat:
            # `moi_ky_ky_regex` — KỲ LẤY TỪ TÊN FILE CHỈ ĐỂ GỘP BẢN CHỐT, không đụng gì tới `ngay`
            # của từng dòng. Cần cho nguồn suy kỳ từ CỘT NGÀY (`vhkd_kqkd`): `ngay_tu_ten_file`
            # trả None nên trước đây hàm này giữ nguyên MỌI file, và chạy tay
            # `spec_extract.py vhkd_kqkd --write` nạp cả 7 bản luỹ kế T8 chồng lên nhau — đúng 7
            # lần số thật (dính 29/08/2026 trên DB test, phải xoá 3.047 dòng). Cron `LUY_KE` vốn
            # chặn được, nhưng luật chống trùng nằm ở MỘT đường vào thì đường kia là cửa mở.
            # Khai `ky_tu_ten_file` để lấp chỗ này thì KHÔNG được: khoá đó đổi luôn cách gán ngày
            # cho từng dòng, tức đổi số của nguồn đang chạy đúng.
            m = re.search(ky_pat, os.path.basename(f), re.IGNORECASE)
            if m:
                ngay = "|".join(x or "" for x in (m.groups() or (m.group(0),)))
                tu_ky_pat = True
        if not ngay:
            giu[f] = (f, "")      # không suy được kỳ -> giữ nguyên, đừng im lặng loại
            continue
        # `tu_ky_pat`: khoá đã LÀ định danh kỳ rồi (vd "2026|8"), cắt [:7] là băm nát nó.
        ky = ("" if che_do == "mot_file"
              else ngay if (che_do == "ngay" or tu_ky_pat) else ngay[:7])
        if pat:
            m = re.search(pat, os.path.basename(f), re.IGNORECASE)
            ky = (ky, tuple(x or "" for x in m.groups()) if m else os.path.basename(f))
        cu = giu.get(ky)
        # HOÀ ngày chốt -> lấy file VỀ SAU (mtime). Nghiệp vụ gửi lại bản SỬA của cùng một kỳ với
        # tên khác (claim T3-T6: "…7.24. BaocaoClaim_B2C_T3" rồi "…8.11. BaocaoClaim_B2C_T3"),
        # hai tên cùng suy ra một kỳ nên so ngày là hoà. Không phá hoà thì thứ tự sorted() quyết
        # định, mà "7.24" đứng trước "8.11" -> giữ đúng bản CŨ và vứt bản đã sửa.
        # Giữ KÈM khoá đã suy được (`cu[1]`) chứ không suy lại từ tên file: với `moi_ky_ky_regex`
        # thì `ngay_tu_ten_file` trả None, suy lại là so `str > None` -> TypeError giữa chừng mẻ.
        if cu is None:
            moi_hon = True
        else:
            moi_hon = (ngay, os.path.getmtime(f)) > (cu[1], os.path.getmtime(cu[0]))
        if moi_hon:
            if cu is not None:
                bo.append(os.path.basename(cu[0]))
            giu[ky] = (f, ngay)
        else:
            bo.append(os.path.basename(f))
    return sorted(v[0] for v in giu.values()), bo


# ─────────────────────────── trích 1 file ───────────────────────────
def _tron_vung(spec, v):
    """Trộn 1 khai báo `vung` lên spec gốc -> spec con chỉ đọc ĐÚNG khối đó.

    Khoá trong `vung` GHI ĐÈ khoá cùng tên ở spec, riêng `cot` / `chieu_co_dinh` thì trộn theo
    từng khoá con (vùng chỉ cần khai lại cột nào LỆCH, không phải chép cả bảng cột).
    """
    sp = {k: x for k, x in spec.items() if k != "vung"}
    for k, x in v.items():
        if k in ("cot", "chieu_co_dinh"):
            sp[k] = {**(spec.get(k) or {}), **(x or {})}
        elif k != "ten":
            sp[k] = x
    return sp


def extract_file(spec, path):
    """Trích 1 file. `vung` (nếu có) = NHIỀU KHỐI trong CÙNG một sheet, đọc lần lượt.

    VÌ SAO CÓ `vung` (14/08/2026, báo cáo QTVH An Taxi): file `B.7.AAG.PKDVH.M.2026...` xếp 6 khối
    nghiệp vụ (A nhân sự · B đội xe · C chuyến · D1 doanh thu · D2 thuê xe) NỐI TIẾP nhau trong
    một sheet, mỗi khối có dòng tiêu đề riêng và 12 dòng tháng, lại chia tiếp theo depot
    (I Sơn Tây · II Thái Nguyên · III Tổng 2 depot). Một spec = một dải dòng nên nếu không có
    `vung` thì phải đẻ ~10 file spec gần trùng nhau cho CÙNG một file nguồn — mỗi lần mapping đổi
    là sửa 10 chỗ. Nay 1 spec khai `cot` chung + liệt kê các vùng lệch.

    Mỗi vùng vẫn tự dò cột THEO TÊN HEADER trong đúng dòng tiêu đề của nó (nguyên tắc số 1) và
    vẫn kiểm được `kiem_tra_o` riêng, nên chèn/xoá cột trong một khối không lây sang khối khác.
    """
    # `moi_sheet_chua` (19/08/2026, báo cáo QTVH Xanh Taxi): MỘT file chứa NHIỀU SHEET THÁNG
    # ("BÁO CÁO THÁNG 8", "BÁO CÁO THÁNG 7"…) và tháng sau lại thêm một sheet nữa vào chính file
    # đó. Khai `sheet.ten` cứng là mỗi kỳ phải sửa spec, còn `theo_thang` chỉ lấy đúng tháng của
    # TÊN FILE nên các tháng cũ nằm im trong file mà không ai biết. Ở đây quét MỌI sheet có tên
    # chứa mốc, mỗi sheet đọc như một file con; kỳ của từng sheet lấy từ chính ô ngày ở dòng tiêu
    # đề (`cot_ngay.ky_tu_o`), không suy từ tên file.
    _sh = (spec.get("nguon") or {}).get("sheet") or {}
    if "moi_sheet_chua" in _sh:
        ten_sheet = [n for n in _mo_wb(path).sheetnames
                     if _nd(_sh["moi_sheet_chua"]) in _nd(n)]
        if not ten_sheet:
            return [], [f"không sheet nào chứa '{_sh['moi_sheet_chua']}'"]
        recs, warn = [], []
        for ten in ten_sheet:
            con = {**spec, "nguon": {**(spec.get("nguon") or {}), "sheet": {"ten": ten}}}
            r, w = extract_file(con, path)
            recs += r
            warn += [f"[sheet {ten}] {x}" for x in w]
        return recs, _gop_canh_bao(warn)

    # `moi_sheet_ngay` (19/08/2026): MỖI SHEET LÀ MỘT NGÀY, tên sheet chính là số ngày ("01".."17"
    # ở báo cáo HQKD ngày của Xanh Taxi). Khác `moi_cot_ngay` (ngày nằm ở CỘT) và khác
    # `theo_thang` (một sheet cho cả tháng). Năm+tháng lấy từ tên file, ngày từ tên sheet, rồi
    # ghim vào `chieu_co_dinh.ngay` để mọi dòng của sheet mang đúng ngày đó.
    if _sh.get("moi_sheet_ngay"):
        ky, w = _ky_thang(spec, path)
        if not ky:
            return [], w
        y, mo = ky
        ten_sheet = [n for n in _mo_wb(path).sheetnames if str(n).strip().isdigit()
                     and 1 <= int(str(n).strip()) <= 31]
        if not ten_sheet:
            return [], [*w, "không sheet nào mang tên là SỐ NGÀY (01..31)"]
        recs, warn = [], list(w)
        for ten in sorted(ten_sheet, key=lambda x: int(str(x).strip())):
            try:
                ngay = dt.date(y, mo, int(str(ten).strip())).isoformat()
            except ValueError:                       # ngày 30/31 ở tháng ngắn -> sheet thừa
                warn.append(f"sheet '{ten}' không phải ngày hợp lệ của {y}-{mo:02d} -> bỏ")
                continue
            con = {**spec,
                   "nguon": {**(spec.get("nguon") or {}), "sheet": {"ten": ten}},
                   "chieu_co_dinh": {**(spec.get("chieu_co_dinh") or {}), "ngay": ngay}}
            r, w2 = extract_file(con, path)
            recs += r
            warn += [f"[sheet {ten}] {x}" for x in w2]
        return recs, _gop_canh_bao(warn)

    # `chi_nap_tu_ngay`: chặn NẠP các bản cũ hơn mốc, nhưng KHÔNG chặn khi đang được
    # `_tru_ngay_truoc` đọc làm mốc trừ — bản ngay trước mốc vẫn phải đọc được, nếu không thì
    # ngày đầu tiên sau mốc mất số trừ và ôm trọn phần luỹ kế từ đầu tháng.
    if spec.get("chi_nap_tu_ngay") and not spec.get("_dang_doc_ngay_truoc"):
        _nf, _ = ngay_tu_ten_file(spec, path)
        if _nf and _nf < spec["chi_nap_tu_ngay"]:
            return [], [f"BỎ QUA — bản {_nf} cũ hơn mốc `chi_nap_tu_ngay` "
                        f"{spec['chi_nap_tu_ngay']}"]

    vung = spec.get("vung")
    if not vung:
        recs, warn = _extract_vung(spec, path)
    else:
        recs, warn = [], []
        for i, v in enumerate(vung, 1):
            r, w = _extract_vung(_tron_vung(spec, v), path)
            recs += r
            warn += [f"[vùng {v.get('ten') or i}] {x}" for x in w]
    if spec.get("tru_ngay_truoc") and not spec.get("_dang_doc_ngay_truoc"):
        recs, w3 = _tru_ngay_truoc(spec, path, recs)
        warn += w3
    return recs, warn


def _tru_ngay_truoc(spec, path, recs):
    """Ô nguồn là LUỸ KẾ TỪ ĐẦU THÁNG -> đổi thành số CỦA RIÊNG NGÀY bằng hiệu hai file liên tiếp.

    Vì sao phải làm ở tầng nạp chứ không để tầng đọc trừ: `raw_rows` là sổ CỘNG ĐƯỢC — mọi màn đều
    SUM theo kỳ/đơn vị. Nạp thẳng luỹ kế vào là chọn 3 ngày bất kỳ rồi cộng lại ra gấp mấy lần số
    thật, và không có chỗ nào chặn được. Trừ ngay lúc nạp thì DB chứa đúng "số phát sinh trong
    ngày", cộng bao nhiêu ngày cũng đúng.

    Đây chính là cách kế toán chốt trong mapping XDV/VHKD (cột "Mô tả cách lấy API từ Cyber"):
    "Lấy mã số B100: Báo cáo ngày sau - báo cáo ngày hôm trước". Đã đối chứng 30->31/08/2026 trên
    báo cáo lợi nhuận khối XDV: hiệu luỹ kế = hiệu cột 'Kỳ này', LỆCH 0 ở cả B100/B110/B120/B130/
    B410, và tách được tới từng xưởng.

    BA QUY TẮC, đừng bỏ cái nào:
      1. CHỈ TRỪ TRONG CÙNG MỘT THÁNG khi `tru_ngay_truoc: true`. Cột luỹ kế reset về 0 đầu mỗi
         tháng, nên file ngày 01 phải giữ nguyên giá trị (nó ĐÃ là số của ngày), trừ với ngày 31
         tháng trước là ra số âm khổng lồ.
         NGOẠI LỆ `tru_ngay_truoc: "lien_thang"` — cho cột luỹ kế KHÔNG reset theo tháng (cột
         "Lũy kế" của báo cáo LN toàn khối, luỹ kế từ khi thành lập). Ở đó CHÍNH ngày 01 mới cần
         trừ với bản cuối tháng trước, còn giữ nguyên là nạp cả nghìn tỷ vào một ngày. Khai nhầm
         khoá này cho cột reset theo tháng thì ngày 01 ra số ÂM bằng cả tháng trước — kiểm bằng
         cách xem giá trị bản ngày 01 có nhỏ hơn bản ngày 31 tháng trước hay không.
      2. FILE TRƯỚC = file có ngày LỚN NHẤT còn nhỏ hơn ngày đang nạp, KHÔNG phải "hôm qua".
         Nguồn nghỉ cuối tuần/lễ nên chuỗi ngày đứt quãng; lấy cứng d-1 là mất trắng phần phát
         sinh giữa hai lần nộp. Khi khoảng cách > 1 ngày thì hiệu là số GỘP của cả quãng, gán vào
         ngày cuối quãng — có cảnh báo để người soi biết, vì phân bố theo ngày lúc đó không thật.
      3. DÒNG BIẾN MẤT so với file trước vẫn phải đẻ bản ghi ÂM: nếu kỳ trước một xưởng có số mà
         kỳ này hết (bị điều chỉnh giảm về 0), bỏ qua là để lại phần dôi vĩnh viễn trong DB.
    """
    warn = []
    ngay_nay, _ = ngay_tu_ten_file(spec, path)
    if not ngay_nay:
        return recs, ["tru_ngay_truoc: không đọc được ngày từ tên file -> giữ nguyên luỹ kế"]
    nguon = spec.get("nguon") or {}
    thu_muc = os.path.dirname(path)
    ung_vien = []
    for f in glob.glob(os.path.join(thu_muc, nguon.get("file_glob") or "*.xlsx")):
        d, _ = ngay_tu_ten_file(spec, f)
        cung_thang = spec.get("tru_ngay_truoc") != "lien_thang"
        if d and d < ngay_nay and (d[:7] == ngay_nay[:7] or not cung_thang):
            ung_vien.append((d, f))
    if not ung_vien:
        # NGÀY ĐẦU THÁNG KHÔNG CÓ GÌ ĐỂ TRỪ (user chốt 05/09/2026: "cái của ngày đầu tháng thì
        # không cần trừ đi"). Nhưng nếu cột đang đọc là luỹ kế LIÊN THÁNG thì ô của chính nó là
        # luỹ kế từ khi thành lập — hàng nghìn tỷ, không phải số của ngày. `cot_ngay_dau_thang`
        # khai cột thay thế CHỈ dùng cho bản đầu tháng: trên bản ngày 01, cột "Kỳ này" CHÍNH LÀ
        # luỹ kế trong tháng tính tới ngày 01, tức đúng con số mà phép trừ lẽ ra phải cho ra.
        # Không phải trộn hai thước đo — vẫn là "luỹ kế trong tháng", chỉ đọc ở cột mang nó.
        cot_dt = spec.get("cot_ngay_dau_thang")
        if cot_dt:
            spec2 = json.loads(json.dumps({k: v for k, v in spec.items()
                                           if k not in ("tru_ngay_truoc",)}))
            for c in spec2.get("cot_gia_tri") or []:
                c["header"] = cot_dt
            recs, _w2 = extract_file({**spec2, "_dang_doc_ngay_truoc": True}, path)
            return ([r for r in recs if abs(r.get("amount") or 0) > 1e-9],
                    [f"tru_ngay_truoc: {ngay_nay} là bản đầu tiên có trong tháng -> KHÔNG trừ, "
                     f"đọc cột '{cot_dt}' của chính bản này (xem `cot_ngay_dau_thang`)"])
        return ([r for r in recs if abs(r.get("amount") or 0) > 1e-9],
                [f"tru_ngay_truoc: {ngay_nay} là bản đầu tiên "
                 + ("có trong thư mục" if spec.get("tru_ngay_truoc") == "lien_thang"
                    else "có trong tháng")
                 + " -> giữ nguyên luỹ kế"
                 + ("" if spec.get("tru_ngay_truoc") != "lien_thang" else
                    " — VỚI CỘT LUỸ KẾ LIÊN THÁNG ĐÂY LÀ SỐ RÁC (luỹ kế từ khi thành lập), "
                    "đặt `chi_nap_tu_ngay` sau bản này để bỏ nó")])
    ngay_truoc, file_truoc = max(ung_vien)
    truoc, _w = extract_file({**spec, "_dang_doc_ngay_truoc": True}, file_truoc)
    if (dt.date.fromisoformat(ngay_nay) - dt.date.fromisoformat(ngay_truoc)).days > 1:
        warn.append(f"tru_ngay_truoc: bản liền trước là {ngay_truoc}, cách {ngay_nay} hơn 1 ngày "
                    f"-> số ghi vào {ngay_nay} là GỘP của cả quãng")

    def khoa(r):
        return (r.get("cost_center"), r.get("cong_ty"), r.get("khoi"),
                r.get("dim1"), r.get("dim2"), r.get("dim3"))
    cu = {}
    for r in truoc:
        cu[khoa(r)] = cu.get(khoa(r), 0.0) + (r.get("amount") or 0.0)
    ra, EPS = [], 1e-9
    for r in recs:
        k = khoa(r)
        r["amount"] = (r.get("amount") or 0.0) - cu.pop(k, 0.0)
        if abs(r["amount"]) > EPS:
            ra.append(r)
    # Khoá còn sót trong `cu` = có ở file trước, mất ở file này -> phần giảm, phải ghi âm (quy tắc 3)
    mat = 0
    for k, v in cu.items():
        if abs(v) <= EPS:
            continue
        mau = dict(recs[0]) if recs else {}
        mau.pop("payload", None)
        r = {**mau, "cost_center": k[0], "cong_ty": k[1], "khoi": k[2],
             "dim1": k[3], "dim2": k[4], "dim3": k[5], "amount": -v,
             "payload": {**(spec.get("payload_them") or {}), "chi_con_o_ban_truoc": True}}
        r["ngay"] = ngay_nay
        ra.append(r)
        mat += 1
    if mat:
        warn.append(f"tru_ngay_truoc: {mat} chỉ tiêu có ở bản {ngay_truoc} nhưng mất ở bản "
                    f"{ngay_nay} -> ghi bản ghi ÂM để tổng luỹ kế vẫn khớp")
    return ra, warn


# Bộ nhớ đệm MỘT workbook (19/08/2026). `vung` và hai chế độ `moi_sheet_*` đọc CÙNG một file
# nhiều lượt — báo cáo HQKD ngày Xanh Taxi có 17 sheet × 4 vùng = 68 lượt mở, mà riêng
# `load_workbook` của file đó mất 7,5 giây: 8 phút cho một file 790 KB. Giữ đúng 1 workbook (theo
# path + mtime) là đủ, vì cả `extract_file` lẫn `run` đều xử lý xong một file rồi mới sang file
# sau. Không đóng workbook đang nằm trong đệm — `_dong_wb` gọi ở cuối mỗi file trong `run`.
_WB_CACHE = {}


def _gop_canh_bao(warn):
    """Gộp cảnh báo GIỐNG NHAU lặp qua nhiều sheet thành một dòng.

    File HQKD ngày có 17 sheet × 4 vùng, mỗi lượt đẻ đúng một câu "bỏ 101 dòng không qua bộ lọc"
    -> 68 dòng cảnh báo y hệt nhau che mất cảnh báo THẬT (sheet lệch layout, mã lạ). Gộp lại giữ
    nguyên thông tin mà đọc được.
    """
    thu_tu, nhom = [], {}
    for w in warn:
        m = re.match(r"^\[sheet (.+?)\] (.*)$", w, re.S)
        khoa, sheet = (m.group(2), m.group(1)) if m else (w, None)
        if khoa not in nhom:
            nhom[khoa] = []
            thu_tu.append(khoa)
        if sheet:
            nhom[khoa].append(sheet)
    ra = []
    for khoa in thu_tu:
        sh = nhom[khoa]
        if not sh:
            ra.append(khoa)
        elif len(sh) == 1:
            ra.append(f"[sheet {sh[0]}] {khoa}")
        else:
            ra.append(f"[{len(sh)} sheet: {sh[0]}…{sh[-1]}] {khoa}")
    return ra


def _mo_wb(path):
    khoa = (os.path.abspath(path), os.path.getmtime(path))
    if _WB_CACHE.get("khoa") != khoa:
        _dong_wb()
        _WB_CACHE["khoa"] = khoa
        _WB_CACHE["wb"] = openpyxl.load_workbook(path, data_only=True, read_only=True)
    return _WB_CACHE["wb"]


def _dong_wb():
    wb = _WB_CACHE.pop("wb", None)
    _WB_CACHE.pop("khoa", None)
    if wb is not None:
        try:
            wb.close()
        except Exception:                                          # noqa: BLE001
            pass


def _extract_vung(spec, path):
    warn, recs = [], []
    wb = _mo_wb(path)
    try:
        _sh_cfg = (spec.get("nguon") or {}).get("sheet") or {}
        _thang = None
        if "theo_thang" in _sh_cfg:
            _ky, _w = _ky_thang(spec, path)
            warn += _w
            _thang = _ky[1] if _ky else None
        sheet = _chon_sheet(wb, _sh_cfg, thang=_thang)
        if not sheet:
            return [], [f"không chọn được sheet (có: {wb.sheetnames})"]
        ws = wb[sheet]
        hdr_cfg = spec.get("header") or {"dong": 1}
        if hdr_cfg.get("khong_co"):
            # Sheet KHÔNG có dòng tiêu đề (vd "Tồn kho xe vật lý": dữ liệu chạy thẳng từ dòng 2,
            # 3 cột không tên). Khi đó mọi khai báo cột PHẢI dùng `"cot": "<chữ cột>"`.
            max_hdr, hmap = 0, ({}, {})
        elif hdr_cfg.get("tim_o"):
            # TỰ DÒ dòng header theo một nhãn mốc. Cần khi cùng một spec phải đọc nhiều sheet mà
            # dòng header nằm ở vị trí KHÁC NHAU: BCTC SRVF để header ở dòng 1 tại sheet T01 nhưng
            # dòng 8 tại T07 (T07 có thêm 7 dòng tiêu đề đơn vị/địa chỉ phía trên). Khai `dong`
            # cứng thì một trong hai sheet chắc chắn trượt; gộp `dong: [1,8]` cũng sai vì ô đầu
            # dòng 1 của T07 là tên chi nhánh, không phải "Mã số".
            moc = _nd(hdr_cfg["tim_o"])
            toi_da = int(hdr_cfg.get("toi_da", 30))
            quet = [list(r) for r in ws.iter_rows(min_row=1, max_row=toi_da, values_only=True)]
            max_hdr = 0
            for i, r in enumerate(quet, start=1):
                if any(_nd(c) == moc for c in r if c not in (None, "")):
                    max_hdr = i
                    break
            if not max_hdr:
                return [], [f"không tìm được dòng header chứa '{hdr_cfg['tim_o']}' "
                            f"trong {toi_da} dòng đầu sheet '{sheet}'"]
            hmap = _map_header(quet[:max_hdr], max_hdr)
        else:
            hdr_dong = hdr_cfg.get("dong", 1)
            max_hdr = max(hdr_dong) if isinstance(hdr_dong, list) else hdr_dong
            head_rows = [list(r) for r in ws.iter_rows(min_row=1, max_row=max_hdr,
                                                       values_only=True)]
            hmap = _map_header(head_rows, hdr_dong)
        cot = _resolve_cot(spec, hmap, warn)

        # Kỳ nằm ở TÊN FILE nhưng số tháng/ngày nằm TRONG Ô -> ghép hai nửa lại ngay tại đây rồi
        # nhét vào cfg của cột, để `_lay_o` (chỉ thấy 1 ô) dựng được ngày đầy đủ. Bản kế hoạch báo
        # cáo QTVH An Taxi là MỘT file cho cả năm ("…M.2026.Baocaotonghop"), nên kỳ tuyệt đối
        # không được suy từ tên file như các nguồn ảnh-chụp khác.
        _nam, _ky_ngay = None, None
        if any(c.get("kieu") == "thang_cuoi" for _, c in cot.values()):
            cn = spec.get("nam_tu_ten_file") or {"regex": r"\.M\.(\d{4})\."}
            mn = re.search(cn["regex"], os.path.basename(path))
            if mn:
                _nam = int(mn.group(1))
            else:
                warn.append(f"không dò được NĂM từ tên file: {os.path.basename(path)}")
        if any(c.get("kieu") in ("ngay_trong_thang", "ngay_dd_mm") for _, c in cot.values()):
            _ky_ngay, w5 = _ky_thang(spec, path)
            warn.extend(w5)
        for dich, (j, c) in list(cot.items()):
            if c.get("kieu") == "thang_cuoi":
                cot[dich] = (j, {**c, "_nam": _nam})
            elif c.get("kieu") in ("ngay_trong_thang", "ngay_dd_mm"):
                cot[dich] = (j, {**c, "_ky": _ky_ngay})

        chieu = {}
        for dich, cfg in (spec.get("chieu_tu_ten_file") or {}).items():
            m = re.search(cfg["regex"], os.path.basename(path), re.I)
            v = (m.group(int(cfg.get("nhom", 1))) if m else cfg.get("mac_dinh"))
            v = (str(v).upper() if v and cfg.get("hoa") else v)
            # `chuan_hoa` (17/08/2026): CHÍNH cái tên bắt được từ file cũng cần chuẩn hoá. Báo cáo
            # dầu không có cột đơn vị nào — dự án chỉ nằm trong TÊN FILE (`BCDauCaoBang`) — nên
            # không hook được ở `cot`, và thiếu nó thì màn nhiên liệu rỗng hoàn toàn khi bật lọc.
            # Hook trả dict (vd `cc_qlts` -> cost_center + cong_ty + khoi) thì trộn cả 3 vào, khi
            # đó `dich` do chính dict quyết định.
            hook = cfg.get("chuan_hoa")
            if hook and v is not None:
                res = _CHUAN_HOA[hook](v)
                if isinstance(res, dict):
                    if res.get("_khong_map"):
                        warn.append(f"{dich}: không map được {res['_khong_map']!r} (từ tên file)")
                    chieu.update({k: x for k, x in res.items() if k != "_khong_map"})
                    continue
                v = res
            chieu[dich] = v

        # CHỐT CHẶN LAYOUT: nguồn chỉ đúng khi các ô mốc khớp. Bắt buộc với spec dùng CHỮ CỘT —
        # file công nợ T1 và T7 tuy cùng tên sheet nhưng bố cục KHÁC HẲN (T1 không có phân tách
        # kênh, header ở dòng 5). Không có chốt này thì T1 vẫn "chạy được" và đẻ ra 9,86 tỷ vô
        # nghĩa, im lặng — kiểu sai nguy hiểm nhất.
        for ktr in spec.get("kiem_tra_o") or []:
            thuc = ws[ktr["o"]].value
            if _nd(ktr.get("bang")) not in _nd(thuc):
                return [], [f"BỎ QUA — layout khác spec: ô {ktr['o']} = "
                            f"{str(thuc)[:40]!r}, cần chứa {ktr.get('bang')!r}"]

        # `chi_doc_khi_o`: giống `kiem_tra_o` nhưng KHÔNG khớp thì bỏ qua IM LẶNG, và `bang`
        # nhận `{thang}`/`{mm}` thay bằng tháng của kỳ. Dùng cho sheet xếp NHIỀU KHỐI THÁNG nối
        # tiếp nhau (kế hoạch ngày của Showroom: THÁNG 8 ở dòng 2, THÁNG 9 dòng 13, … tới THÁNG
        # 12) — mỗi khối khai một `vung`, và chỉ khối trùng tháng của file mới được đọc. Không có
        # chốt này thì cột "Ngày 01..31" của khối THÁNG 9 sẽ bị gán vào tháng 8 (kỳ suy từ TÊN
        # FILE), tức kế hoạch tháng sau đè lên tháng này. Im lặng vì 4/5 khối không khớp là
        # chuyện BÌNH THƯỜNG mỗi lần chạy, báo ra chỉ làm loãng cảnh báo thật.
        if spec.get("chi_doc_khi_o"):
            _kyt, w6 = _ky_thang(spec, path)
            warn.extend(w6)
            for ktr in spec["chi_doc_khi_o"]:
                can = str(ktr.get("bang", ""))
                if _kyt:
                    can = can.replace("{thang}", str(_kyt[1])).replace("{mm}", f"{_kyt[1]:02d}")
                if _nd(can) not in _nd(ws[ktr["o"]].value):
                    return [], []

        ngay_file, w2 = ngay_tu_ten_file(spec, path)
        warn.extend(w2)
        # Nguồn báo cáo TUẦN: kèm mốc tuần vào payload từng dòng để giao diện nói rõ đang xem
        # ảnh chụp tuần nào (`ngay` chỉ mang ngày CHỐT — Chủ nhật cuối tuần được phủ).
        payload_tuan = chu_ky_tuan(spec, path)

        bat_dau = spec.get("dong_bat_dau") or (max_hdr + 1)
        # `dong_ket_thuc` (14/08/2026): CHẶN TRÊN của dải dòng — bắt buộc khi nhiều khối nằm nối
        # tiếp trong một sheet (xem `vung`). Không có nó thì khối A quét luôn xuống khối B: cột
        # "Tổng" của bảng nhân sự và cột "Tổng số xe" của bảng đội xe cùng nằm ở E/C nên số của
        # khối sau vẫn "đọc được" và cộng vào khối trước — sai mà không một dấu hiệu nào.
        ket_thuc = spec.get("dong_ket_thuc")
        gia_tri_cols = spec.get("cot_gia_tri") or []
        gt_idx = [(_tim_cot(hmap, c, f"cột giá trị {c.get('dim1') or c.get('header')}", warn), c)
                  for c in gia_tri_cols]
        gt_idx = [(j, c) for j, c in gt_idx if j is not None]

        # Cột-theo-ngày: đọc SỐ NGÀY từ chính dòng tiêu đề của từng cột rồi ghép với kỳ của file.
        # KHÔNG đánh số ngày theo thứ tự cột: tháng 2 chỉ có 28-29 cột có nghĩa, và vài file chèn
        # thêm cột phụ giữa dải ngày — suy theo vị trí sẽ lệch ngày mà không có dấu hiệu gì.
        # 3 khoá TUỲ CHỌN thêm 10/08/2026 cho bản KẾ HOẠCH doanh thu theo ngày của XDV
        # (Baocaodoanhthukehoachngay, sheet "KHDT theo ngày") — mặc định giữ nguyên hành vi cũ
        # nên mọi spec đang chạy không đổi một ly:
        #   "dong": 9       -> đọc mốc ngày từ DÒNG KHÁC dòng tiêu đề. File đó ghi ngày ở dòng 9
        #                      (ô gộp 2 cột), còn dòng 10 là tiêu đề con "Số lượng"/"Doanh thu".
        #   ô là NGÀY THẬT  -> chấp nhận cell datetime/date, không chỉ số "1".."31". Kèm kiểm
        #                      tháng/năm khớp kỳ của file: lệch thì BỎ cột đó và cảnh báo, vì
        #                      nhận sai là gán số của tháng khác vào kỳ này mà không ai thấy.
        #   "lech_amount"   -> giá trị KHÔNG nằm ở cột mang mốc ngày mà lệch sang phải n cột.
        #                      Ở file này mỗi ngày chiếm 2 cột: (Số lượng, Doanh thu).
        #   "lech_amount2"  -> cột thứ hai của cặp, ghi vào amount2 (giữ luôn KH số lượng RO để
        #                      tính "doanh thu bình quân/RO mục tiêu" mà không cần nạp 2 lần).
        nc_ngay = spec.get("cot_ngay") or {}
        ngay_theo_cot = []
        if spec.get("ban_ghi") == "moi_cot_ngay":
            # `ky_tu_o` (19/08/2026): kỳ nằm ở CHÍNH Ô NGÀY của cột, không ở tên file. Cần cho
            # file một-sheet-một-tháng (`moi_sheet_chua`): tên file chỉ ghi tháng mới nhất nên
            # sheet tháng cũ sẽ bị loại sạch bởi phép kiểm "ngày không thuộc kỳ của file".
            # Chỉ chấp ô NGÀY THẬT — ô ghi số trần "1".."31" không nói được nó thuộc tháng nào.
            ky_tu_o = bool(nc_ngay.get("ky_tu_o"))
            nam_thang, w3 = (None, []) if ky_tu_o else _ky_thang(spec, path)
            warn.extend(w3)
            if ky_tu_o or nam_thang:
                y, mo = nam_thang if nam_thang else (None, None)
                j1 = column_index_from_string(nc_ngay.get("tu", "G")) - 1
                j2 = column_index_from_string(nc_ngay.get("den", "AK")) - 1
                dong_ngay = nc_ngay.get("dong")
                if dong_ngay:
                    moc_row = [c.value for c in ws[int(dong_ngay)]]
                else:
                    moc_row = head_rows[(hdr_dong[-1] if isinstance(hdr_dong, list) else hdr_dong) - 1]
                for j in range(j1, j2 + 1):
                    raw = moc_row[j] if j < len(moc_row) else None
                    ngay_cot = None
                    if isinstance(raw, (dt.datetime, dt.date)):
                        if ky_tu_o or (raw.year, raw.month) == (y, mo):
                            ngay_cot = dt.date(raw.year, raw.month, raw.day).isoformat()
                        else:
                            warn.append(f"cột {get_column_letter(j + 1)} ghi ngày {raw} không thuộc "
                                        f"kỳ {y}-{mo:02d} của file -> bỏ cột")
                    else:
                        # Chấp cả "01" lẫn "Ngày 01": bản kế hoạch ngày của Showroom ghi tiêu
                        # đề cột là "Ngày 01".."Ngày 31" chứ không phải số trần.
                        txt = str(raw or "")
                        m = None if ky_tu_o else \
                            re.match(r"^\s*(?:ng[àa]y\s*)?(\d{1,2})\s*$", txt, re.I)
                        thang_cot = None
                        if m is None and not ky_tu_o:
                            # DẠNG "dd/mm" CÓ ĐUÔI (04/09/2026): kế hoạch xuất hoá đơn 15 ngày ghi
                            # tiêu đề cột là "01/09 (T3)", "05/09 (T7)*" — có cả thứ trong tuần và
                            # dấu * đánh dấu ngày cao điểm. Mẫu cũ neo `$` nên không khớp, và cột
                            # ngày không dựng được thì spec BỎ QUA CẢ FILE.
                            # Bắt buộc có dấu `/`: KHÔNG nới mẫu cũ thành "số + rác phía sau", vì
                            # thế là mọi tiêu đề bắt đầu bằng chữ số ("15 Ngày", "9 Showroom")
                            # đều thành một cột ngày và bảng đẻ ra ngày không có thật.
                            m = re.match(r"^\s*(\d{1,2})\s*/\s*(\d{1,2})\b", txt)
                            if m:
                                thang_cot = int(m.group(2))
                        if m:
                            # Có tháng trong chính tiêu đề thì PHẢI khớp tháng của file — cùng
                            # phép kiểm mà nhánh ô-ngày-thật ở trên đang làm. Bỏ qua bước này là
                            # cột của tháng khác lặng lẽ rơi vào kỳ này.
                            if thang_cot is not None and thang_cot != mo:
                                warn.append(f"cột {get_column_letter(j + 1)} ghi {txt.strip()!r} "
                                            f"không thuộc tháng {mo:02d} của file -> bỏ cột")
                            else:
                                try:
                                    ngay_cot = dt.date(y, mo, int(m.group(1))).isoformat()
                                except ValueError:
                                    ngay_cot = None   # ngày 30/31 ở tháng ngắn -> cột thừa, bỏ
                    if ngay_cot:
                        ngay_theo_cot.append((j, ngay_cot))
            if not ngay_theo_cot:
                return [], [*warn, "BỎ QUA — không dựng được dải cột theo ngày"]

        # Cột-theo-THÁNG: bảng KẾ HOẠCH xếp 12 cột "Tháng 1".."Tháng 12" — MỘT file cấp kế hoạch
        # cho cả năm, nên một lần nạp phải đẻ ra 12 kỳ. Tháng đọc từ CHÍNH tiêu đề cột (không suy
        # theo vị trí): trước dải tháng còn có "Tổng năm" / "6 tháng đầu năm" / "6 tháng cuối năm"
        # và mấy cột tổng này rất hay bị chèn thêm giữa các bản kế hoạch. Năm lấy từ tên file.
        nc_thang = spec.get("cot_thang") or {}
        thang_theo_cot = []
        if spec.get("ban_ghi") == "moi_cot_thang":
            nam_thang, w4 = _ky_thang(spec, path)
            warn.extend(w4)
            if nam_thang:
                y = nam_thang[0]
                j1 = column_index_from_string(nc_thang.get("tu", "G")) - 1
                j2 = column_index_from_string(nc_thang.get("den", "R")) - 1
                hdr_row = head_rows[(hdr_dong[-1] if isinstance(hdr_dong, list) else hdr_dong) - 1]
                # `regex` (tuỳ chọn, 13/08/2026): bản kế hoạch XDV không ghi tiêu đề cột trần là
                # "Tháng 1" mà kèm tên chỉ tiêu — "Sản lượng Tháng 1", "Doanh thu Tháng 1", "Lợi
                # nhuận gộp\nTháng 1" (mỗi mục trong sheet một kiểu). Mặc định giữ nguyên khớp
                # CHẶT `^thang\d+$` để các spec đang chạy không đổi hành vi; spec nào cần thì nới
                # thành `thang(\d{1,2})$`. Vẫn chỉ dò trong dải cột `tu`..`den`, nên "Tổng 6 tháng
                # đầu năm" (không kết thúc bằng số) và các cột tổng khác không lọt vào.
                re_thang = re.compile(nc_thang.get("regex") or r"^thang(\d{1,2})$")
                for j in range(j1, j2 + 1):
                    m = re_thang.search(_nd(hdr_row[j] if j < len(hdr_row) else None))
                    if not m or not 1 <= int(m.group(1)) <= 12:
                        continue
                    mo = int(m.group(1))
                    # Kỳ của kế hoạch là CẢ THÁNG -> neo vào ngày cuối tháng, giống mọi nguồn
                    # ảnh chụp khác, để `period_month` rơi đúng tháng đó.
                    thang_theo_cot.append(
                        (j, dt.date(y, mo, calendar.monthrange(y, mo)[1]).isoformat()))
            if not thang_theo_cot:
                return [], [*warn, "BỎ QUA — không dựng được dải cột theo tháng"]

        khong_map, khong_map_giu, bo_loc, o_loi = {}, {}, 0, {}
        bo_khac_ngay = 0       # xem `chi_lay_ngay_cua_file`
        ngu_canh = {}          # ngữ cảnh mang từ dòng tiêu đề xuống, xem `ngu_canh_dong`
        lap_lai_cuoi = {}      # giá trị gần nhất của cột khai `lap_lai`, xem ngay dưới
        nc_cfg = spec.get("ngu_canh_dong")
        for row in ws.iter_rows(min_row=bat_dau, max_row=ket_thuc, values_only=True):
            # BẢNG PHÂN CẤP: một số báo cáo không lặp lại tên đơn vị trên từng dòng mà đặt nó ở
            # DÒNG TIÊU ĐỀ riêng, các dòng bên dưới ngầm hiểu là của đơn vị đó (báo cáo doanh thu
            # XDV: dòng "3S có đồng sơn | Ocean Park" rồi 8 dòng mã B110..B150 bên dưới).
            # Dòng tiêu đề KHÔNG tự sinh bản ghi — nó chỉ đặt ngữ cảnh.
            # `ngu_canh_dong` nhận 1 quy tắc (dict) hoặc NHIỀU quy tắc (list) — file doanh thu XDV
            # có 2 TẦNG ngữ cảnh lồng nhau: dòng "II/ DOANH THU LỆNH W (BẢO HÀNH)" mở một khối
            # loại lệnh, bên trong lại có các dòng tên xưởng. Quy tắc xét theo thứ tự, khớp cái
            # đầu tiên; `xoa` liệt kê các trường phải quên khi sang khối mới (đổi khối loại lệnh
            # thì cost_center của xưởng cuối khối trước KHÔNG được mang sang).
            hit = False
            for rule in ([nc_cfg] if isinstance(nc_cfg, dict) else (nc_cfg or [])):
                if not _qua_loc_o(row, rule.get("khi") or []):
                    continue
                for k in rule.get("xoa") or []:
                    ngu_canh.pop(k, None)
                for dich, c in (rule.get("gan") or {}).items():
                    v = _lay_o(row, column_index_from_string(c["cot"]) - 1, c, o_loi)
                    if c.get("anh_xa"):
                        # Nhãn trong file dài dòng ("II/ DOANH THU LỆNH W (BẢO HÀNH)") -> quy về
                        # giá trị ngắn dùng cho dim. So theo kiểu "chứa", bỏ dấu.
                        v = next((out for key, out in c["anh_xa"].items() if _nd(key) in _nd(v)),
                                 c.get("mac_dinh"))
                    hook = c.get("chuan_hoa")
                    if hook and v is not None:
                        res = _CHUAN_HOA[hook](v)
                        if isinstance(res, dict):
                            if res.get("_khong_map"):
                                khong_map[res["_khong_map"]] = khong_map.get(res["_khong_map"], 0) + 1
                            ngu_canh.update({k: x for k, x in res.items() if k != "_khong_map"})
                            continue
                        v = res
                    ngu_canh[dich] = v
                hit = True
                break
            if hit:
                continue          # dòng tiêu đề KHÔNG tự sinh bản ghi
            base = {"payload": {**(spec.get("payload_them") or {}), **payload_tuan}}
            if spec.get("khoi"):
                base["khoi"] = spec["khoi"]
            base.update(spec.get("chieu_co_dinh") or {})   # vd {"dim2": "B2B"} cho file 1 kênh
            for dich, v in ngu_canh.items():
                _dat(base, dich, v)
            base.update({k: v for k, v in chieu.items()})
            if ngay_file:
                base["ngay"] = ngay_file
            for dich, (j, cfg) in cot.items():
                val = _lay_o(row, j, cfg, o_loi)
                # `lap_lai` (19/08/2026): Ô GỘP theo chiều dọc — tên khối chỉ ghi ở dòng ĐẦU của
                # khối (cột A báo cáo QTVH Xanh Taxi: "VẬN HÀNH XANH" ở dòng 5 rồi bỏ trắng 29
                # dòng dưới). openpyxl trả None cho phần thân ô gộp, nên không có nó thì 29/30
                # dòng mất đơn vị. Khác `ngu_canh_dong` ở chỗ dòng mang nhãn VẪN sinh bản ghi:
                # ở đây dòng đầu khối cũng là một dòng số liệu thật, không phải dòng tiêu đề.
                if cfg.get("lap_lai"):
                    if val in (None, ""):
                        val = lap_lai_cuoi.get(dich)
                    else:
                        lap_lai_cuoi[dich] = val
                hook = cfg.get("chuan_hoa")
                # `chuan_hoa_khi_rong`: chạy hook CẢ KHI ô trống, vì với cột này "để trống" là
                # một TRẠNG THÁI chứ không phải thiếu dữ liệu. Cột "số ngày quá hạn" của công nợ
                # phải thu T8 bỏ trắng 46/273 hợp đồng chưa tới hạn (16,99 tỷ); bỏ qua hook thì
                # dim1 rỗng và 17 tỷ đó biến mất khỏi cả "Quá hạn" lẫn "Trong hạn" trên màn hình.
                if hook and (val is not None or cfg.get("chuan_hoa_khi_rong")):
                    res = _CHUAN_HOA[hook](val)
                    if isinstance(res, dict):
                        # Hook hỏng KHÔNG loại dòng ngay: phải để `loc` chạy trước, nếu không mọi
                        # dòng RÁC (vùng dán thừa, không có ngày) sẽ bị gộp vào cảnh báo "không map
                        # được đơn vị" và che mất đơn vị hỏng THẬT. Đánh dấu rồi xử lý sau bộ lọc.
                        if res.get("_khong_map"):
                            # `giu_khi_khong_map`: dòng KHÔNG khớp danh mục vẫn là dòng THẬT và
                            # phải giữ để tổng không hụt — vd 16 xe tồn đứng tên "CHI NHÁNH
                            # VINFAST HÀ NỘI" trong báo cáo tồn vật lý: không thuộc SR nào nhưng
                            # là xe có thật. Giữ lại, cost_center để trống, tên thô nằm ở payload.
                            # Giữ dòng thì VẪN PHẢI BÁO. Trước đây `giu_khi_khong_map` nuốt luôn
                            # cảnh báo: 38 dòng công nợ mang tên "Showroom Uông Bí_ HL"/"_ CP"
                            # (8,0 tỷ ở T5, 6,1 tỷ ở T6) vào DB với cost_center rỗng — tổng khối
                            # đúng nhưng biến mất khỏi mọi biểu đồ theo SR, không dấu hiệu gì.
                            base["_khong_map_giu" if cfg.get("giu_khi_khong_map")
                                 else "_khong_map"] = res["_khong_map"]
                        for k2, v2 in res.items():
                            if k2 != "_khong_map":
                                base[k2] = v2
                        continue
                    val = res
                _dat(base, dich, val)
            if all(base.get(k) in (None, "") for k in _CHUAN_TRUONG if k != "khoi") \
                    and not (base.get("payload") or {}):
                continue
            # DỪNG QUÉT khi gặp dòng thoả `dung_khi` — sheet "CN theo đơn vị" có dòng "Total" rồi
            # BÊN DƯỚI còn một bảng phụ (công nợ quá hạn đã có COC) cũng liệt kê tên Showroom ở
            # cột A. Không dừng là gộp luôn bảng phụ vào -> cộng đôi mà không có dấu hiệu gì.
            if spec.get("dung_khi") and _qua_loc(base, spec["dung_khi"]):
                break
            outs = []
            if spec.get("ban_ghi") == "moi_cot_ngay":
                # Bảng có MỘT CỘT MỖI NGÀY (báo cáo doanh thu XDV: G..AK = ngày 01..31, tiêu đề
                # cột chính là số ngày). Mỗi ô có số -> 1 bản ghi mang đúng `ngay` của cột đó,
                # nhờ vậy vừa vẽ được biểu đồ theo ngày vừa cộng lên tháng mà không đếm đôi.
                # Ô rỗng/0 bị bỏ: tháng đang dở thì các ngày chưa tới đều là 0, giữ lại chỉ làm
                # phình bảng và đẻ ra "ngày có số 0" giả.
                lech = int(nc_ngay.get("lech_amount", 0))
                lech2 = nc_ngay.get("lech_amount2")
                for j, ngay_cot in ngay_theo_cot:
                    jv = j + lech
                    v = _so(row[jv] if jv < len(row) else None, float(nc_ngay.get("he_so", 1.0)))
                    if not v:
                        continue
                    r2 = json.loads(json.dumps(base))
                    r2["ngay"] = ngay_cot
                    r2["amount"] = v
                    if lech2 is not None:
                        j2v = j + int(lech2)
                        # he_so RIÊNG: cột thứ hai của cặp là SỐ LƯỢNG RO (đơn vị lệnh), không
                        # phải tiền — dùng chung he_so 1e-9 của tiền là ra 0.0000005 lệnh.
                        r2["amount2"] = _so(row[j2v] if j2v < len(row) else None,
                                            float(nc_ngay.get("he_so_amount2", 1.0)))
                    outs.append(r2)
            elif spec.get("ban_ghi") == "moi_cot_thang":
                # Ô rỗng/0 bị bỏ: bản kế hoạch để trống các tháng chưa đăng ký (A230/A250 trống
                # hết T1-T6). Giữ lại là đẻ ra "kế hoạch = 0" giả, và %HT sẽ chia cho 0.
                # `kieu: "nguong"` — ô là chuỗi có toán tử ("≥95%", "<4%"), xem `_nguong`.
                nguong_kieu = nc_thang.get("kieu") == "nguong"
                for j, ngay_cot in thang_theo_cot:
                    o = row[j] if j < len(row) else None
                    if nguong_kieu:
                        ng = _nguong(o, float(nc_thang.get("he_so", 1.0)))
                        if ng is None:
                            continue
                        v, toan_tu = ng
                    else:
                        v, toan_tu = _so(o, float(nc_thang.get("he_so", 1.0))), None
                        if not v:
                            continue
                    r2 = json.loads(json.dumps(base))
                    r2["ngay"] = ngay_cot
                    r2["amount"] = v
                    if toan_tu:
                        _dat(r2, "payload.toan_tu", toan_tu)
                    outs.append(r2)
            elif spec.get("ban_ghi") == "moi_cot_gia_tri":
                for j, c in gt_idx:
                    r2 = json.loads(json.dumps(base))
                    r2["amount"] = _so(row[j] if j < len(row) else None,
                                       float(c.get("he_so", 1.0)))
                    # `amount2` (26/08/2026): ĐO THỨ HAI đi kèm cùng một chiều. Báo cáo tháng An
                    # Taxi xếp mỗi kênh thu một CỤM cột (App An: số cuốc M · hoàn thành N · hủy O
                    # · doanh thu R) và bản NGÀY của đúng chiều đó đã quy ước amount = doanh thu,
                    # amount2 = số cuốc. Không có khoá này thì phải tách kênh thành 2 dim2 riêng,
                    # tức hai nguồn cùng một chiều lại có hình dạng khác nhau — builder phải viết
                    # hai nhánh và rất dễ cộng nhầm doanh thu với số cuốc.
                    if c.get("amount2"):
                        j2 = _tim_cot(hmap, c["amount2"],
                                      f"amount2 của {c.get('dim1') or c.get('cot')}", warn)
                        if j2 is not None:
                            r2["amount2"] = _so(row[j2] if j2 < len(row) else None,
                                                float(c["amount2"].get("he_so", 1.0)))
                    # `cost_center`/`cong_ty`/`khoi` trong khai báo cột (20/08/2026): báo cáo lợi
                    # nhuận khối XDV xếp MỖI XƯỞNG MỘT CỘT (I->V) bên phải cột "Kỳ này" của cả
                    # khối, còn chiều "chỉ tiêu" thì nằm ở CỘT MÃ SỐ của từng dòng. Không có khoá
                    # này thì phải khai 14 spec gần giống hệt nhau, mỗi spec một xưởng.
                    for k2 in ("dim1", "dim2", "dim3", "cost_center", "cong_ty", "khoi"):
                        if c.get(k2):
                            r2[k2] = c[k2]
                    outs.append(r2)
            else:
                outs.append(base)
            for r2 in outs:
                _doi_gia_tri(r2, spec.get("doi_gia_tri"))
                _tinh_han_no(r2, spec.get("tinh_han_no"), ngay_file)
                _dan_xuat(r2, spec.get("dan_xuat"))
                hong = r2.pop("_khong_map", None)
                giu = r2.pop("_khong_map_giu", None)
                if giu and (r2.get("amount") or 0):
                    khong_map_giu[giu] = khong_map_giu.get(giu, 0) + 1
                # Đơn vị KHÔNG map được mà dòng CÓ SỐ -> luôn báo ra, KHÔNG để bộ lọc nuốt trước.
                # Bắt được 2026-08-09: spec lọc `cost_center khác rỗng`, nên xưởng "Quận 12" (có
                # thật trong file DMS, chưa có trong master_data) rơi vào nhánh "bỏ N dòng không
                # qua bộ lọc" — mất 19 lệnh mà không một cảnh báo nào. Dòng không map mà TOÀN 0
                # thì vẫn coi là rác (dòng tổng/дòng trống), đếm vào bo_loc cho đỡ nhiễu.
                if hong and (r2.get("amount") or 0):
                    khong_map[hong] = khong_map.get(hong, 0) + 1
                elif hong or not _qua_loc(r2, spec.get("loc")):
                    bo_loc += 1
                elif (spec.get("chi_lay_ngay_cua_file") and ngay_file
                        and r2.get("ngay") and r2["ngay"] != ngay_file):
                    # CHỈ GIỮ DÒNG THUỘC ĐÚNG NGÀY CỦA FILE (04/09/2026). Vài nguồn ngày của Cyber
                    # cho lẫn sang ngày HÔM SAU: `TEST_XDV/bangkehoadondv` bản 26/08 chứa 61 hoá
                    # đơn ghi ngày 27/08, mà bản 27/08 cũng có đủ 61 cái đó -> ngày 27/08 bị cộng
                    # đôi 178.253.949 đ. Không khử được bằng `loc` vì `loc` không biết ngày file.
                    # Dòng bị bỏ ở đây LUÔN xuất hiện lại trong file của đúng ngày nó, nên không
                    # mất số; đã kiểm 61/61 số hoá đơn đều có trong bản 27/08.
                    bo_khac_ngay += 1
                else:
                    recs.append(r2)
        if bo_loc:
            warn.append(f"bỏ {bo_loc} {_W_BO_LOC}")
        if bo_khac_ngay:
            # Đếm RIÊNG chứ không gộp vào `bo_loc`: đây là dòng ĐÚNG dữ liệu nhưng thuộc file
            # khác, số phải nhìn thấy được để biết nguồn đang lẫn ngày tới mức nào.
            warn.append(f"bỏ {bo_khac_ngay} dòng thuộc ngày khác ngày của file "
                        f"(chi_lay_ngay_cua_file, ngày file = {ngay_file})")
        if o_loi:
            # Ô lỗi Excel đã bị coi là trống ở trên — báo ra để biết FILE NGUỒN chưa cập nhật
            # xong, đừng đi tìm lỗi ở spec/deriver.
            warn.append("Ô LỖI EXCEL trong file nguồn (đã coi như ô trống): "
                        + ", ".join(f"{k} ({v} ô)"
                                    for k, v in sorted(o_loi.items(), key=lambda x: -x[1])))
        if khong_map:
            warn.append("KHÔNG MAP ĐƯỢC đơn vị (đã qua bộ lọc, tức là dòng THẬT): "
                        + ", ".join(f"{k} ({v} dòng)"
                                    for k, v in sorted(khong_map.items(), key=lambda x: -x[1])[:10]))
        if khong_map_giu:
            warn.append("ĐƠN VỊ NGOÀI DANH MỤC — đã GIỮ dòng (tổng khối đúng) nhưng cost_center "
                        "rỗng nên không lên được biểu đồ theo đơn vị: "
                        + ", ".join(f"{k} ({v} dòng)" for k, v in
                                    sorted(khong_map_giu.items(), key=lambda x: -x[1])[:10]))
        return recs, warn
    finally:
        pass          # workbook nằm trong `_WB_CACHE`, đóng ở cuối mỗi file (xem `_mo_wb`)


def _source_id(path):
    parts = os.path.normpath(path).split(os.sep)
    return f"{parts[-3]}::{os.path.basename(path)}" if len(parts) >= 3 else os.path.basename(path)


def _ghi(spec, path, recs):
    source_file = _source_id(path)
    conn = psycopg.connect(DB_URL)
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM raw_rows WHERE report_type=%s AND source_file=%s",
                    (spec["report_type"], source_file))
        by_ky, thieu_ky = {}, 0
        for r in recs:
            if not r.get("ngay"):
                thieu_ky += 1
                continue
            by_ky.setdefault(r["ngay"][:7], []).append(r)
        rows, skipped, i = [], [], 0
        base_i = int(spec.get("row_index_base", 6500000))
        ky_moi = []
        for period, items in sorted(by_ky.items()):
            # KỲ: nguồn spec là SỐ THỰC TẾ -> được khai sinh kỳ đã tới. Kỳ TƯƠNG LAI vẫn bị chặn
            # (nguồn KẾ HOẠCH NĂM ghi trọn 12 kỳ trong một lần nạp — cho tạo là hôm nay mọc ngay
            # 2026-10/11/12 rỗng). Chi tiết 4 chốt: servers/common/dataset_ky.py.
            ds_ky, _tt = _DSK.lay_hoac_tao_ky(cur, period, nguon=source_file)
            if not ds_ky:
                skipped.append(period)
                continue
            if _tt == _DSK.MOI_TAO:
                ky_moi.append(period)
            got = (ds_ky,)
            for r in items:
                i += 1
                rows.append((got[0], spec["report_type"], base_i + i, r.get("ngay"),
                             r.get("cong_ty"), r.get("khoi"), r.get("cost_center"), period,
                             r.get("amount"), r.get("amount2"), r.get("dim1"), r.get("dim2"),
                             r.get("dim3"),
                             json.dumps(r.get("payload") or {}, ensure_ascii=False), source_file))
        if rows:
            cur.executemany(
                "INSERT INTO raw_rows (dataset_id, report_type, row_index, ngay, cong_ty, khoi, "
                "cost_center, period_month, amount, amount2, dim1, dim2, dim3, payload, "
                "source_file) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", rows)
        conn.commit()
        return {"written": len(rows),
                **({"ky_moi_tao": ky_moi} if ky_moi else {}),
                # Sau 03/09/2026 khoá này CHỈ còn chứa kỳ KHÔNG ĐƯỢC PHÉP tự tạo (tương lai / cũ
                # hơn 24 tháng / sai dạng) — không còn nghĩa "kỳ chưa ai nạp báo cáo tháng" như
                # tên cũ `bo_qua_chua_co_dataset`. Đổi tên để đọc log không kết luận nhầm.
                **({"bo_qua_ky_khong_tao_duoc": skipped} if skipped else {}),
                **({"bo_qua_khong_co_ngay": thieu_ky} if thieu_ky else {})}
    finally:
        conn.close()


_W_LECH_THANG = "SỐ TRONG FILE KHÔNG THUỘC THÁNG MÀ TÊN FILE KHAI"
_W_CONG_DOI = "CỘNG ĐÔI TRONG MỘT FILE"
_TEN_DON_VI_CACHE = {}


def _moi_ten_don_vi():
    """Khoá tên đã chuẩn hoá của MỌI cost center trong master_data (mọi khối) -> mã.

    Dùng `_bo_tien_to(_nd(...))` đúng như `_cc_theo_khoi` để hai vế cùng một dạng khoá.
    Hỏng master_data thì trả rỗng -> chốt ở `_kiem_cong_doi` tự tắt, KHÔNG được chặn đường nạp.
    """
    if not _TEN_DON_VI_CACHE:
        _TEN_DON_VI_CACHE["_da_nap"] = ""
        try:
            for cc in _master_loader().master_data().get("costCenters", []):
                k = _bo_tien_to(_nd(cc.get("ten")))
                if len(k) >= 4:
                    _TEN_DON_VI_CACHE[k] = str(cc.get("ma") or "").strip()
        except Exception:                                          # noqa: BLE001
            pass
    return _TEN_DON_VI_CACHE


def _spec_phan_cap(spec):
    """Các quy tắc `ngu_canh_dong` GÁN cost_center — tức spec tự khai "file này phân cấp:
    tên đơn vị nằm ở DÒNG TIÊU ĐỀ riêng, các dòng số bên dưới thuộc về nó"."""
    nc = spec.get("ngu_canh_dong")
    return [r for r in ([nc] if isinstance(nc, dict) else (nc or []))
            if "cost_center" in (r.get("gan") or {})]


def _kiem_cong_doi(spec, recs):
    """Chốt CỘNG ĐÔI TRONG MỘT FILE cho spec phân cấp. Trả cảnh báo, hoặc None.

    VÌ SAO (bắt 31/08/2026, `Baocaodoanhthungay` khối I): kế toán đổi bố cục — dòng tên xưởng
    trước ở `cột A = "3S có đồng sơn" · cột B = "Ocean Park"` (KHÔNG mã), nay mang luôn mã
    `B100` ở cột A. Quy tắc "dòng tên xưởng" đòi cột A KHÔNG khớp mẫu mã `^B[0-9]+$` nên nay
    không nổ nữa:
    (1) cả khối I mất cost_center, (2) chính dòng đó thành một dòng SỐ, nằm chung xô cấp khối
    với dòng tổng của file -> doanh thu XDV T8 lên 85,098 tỷ trong khi file ghi 42,549 tỷ,
    ĐÚNG GẤP ĐÔI vì hai vế bằng nhau tuyệt đối.

    KHÔNG một lớp nào cũ bắt được: `_ghi` xoá theo `source_file` rồi nạp lại nên không phải nạp
    chồng; `moi_ky_lay_file_moi_nhat` chỉ soi nhiều file cùng kỳ; bước "RÀ SOÁT CỘNG ĐÔI" của cron
    chỉ hỏi "có 2 file cùng đóng góp một lát không"; `kiem_tra_o` thì ô mốc A9/C9 vẫn nguyên. Sai
    số 2× sống nửa ngày trên dashboard chủ tịch mà không có một dòng cảnh báo nào.

    HAI BẤT BIẾN, chỉ áp cho spec CÓ khai `ngu_canh_dong` gán cost_center (hôm nay là 3 spec của
    file doanh thu XDV) — spec phẳng không có khái niệm "dòng tiêu đề đơn vị" nên không đụng tới:
      1. Đã khai dòng tiêu đề đơn vị thì phải có ÍT NHẤT một dòng nhận được cost_center.
      2. Dòng ở CẤP KHỐI (cost_center rỗng) KHÔNG được mang nhãn của một đơn vị trong master_data.
         Ở trạng thái lành, nhãn cấp khối là "TỔNG DOANH THU XDV" / "Doanh thu công việc (XHĐ)" /
         "Lệnh bảo hành (W)" — không cái nào là tên xưởng.
    Vi phạm -> `run()` trả rỗng kèm cảnh báo và KHÔNG ghi, tức GIỮ NGUYÊN dữ liệu cũ đang đúng.
    Cùng triết lý `_kiem_thang_ten_file`: thà đứng lại ở số của hôm qua còn hơn ghi đè bằng số
    gấp đôi mà không ai biết.
    """
    if not recs or not _spec_phan_cap(spec):
        return None
    if not any(r.get("cost_center") for r in recs):
        return (f"{_W_CONG_DOI}: spec khai dòng tiêu đề đơn vị nhưng KHÔNG MỘT dòng nào trong "
                f"{len(recs)} dòng nhận được cost_center -> file đã đổi bố cục, mọi dòng đang "
                f"đứng ở cấp khối và sẽ cộng chung với dòng tổng. BỎ QUA cả file: sửa "
                f"`ngu_canh_dong` của spec cho khớp bố cục mới rồi nạp lại.")
    ten_dv = _moi_ten_don_vi()
    if len(ten_dv) <= 1:
        return None           # không đọc được master_data -> không chấm, đừng chặn đường nạp
    lac = {}
    for r in recs:
        if r.get("cost_center"):
            continue
        for k in ("dim1", "dim2", "dim3"):
            v = r.get(k)
            if v and _bo_tien_to(_nd(v)) in ten_dv:
                lac[f"{k}={v}"] = lac.get(f"{k}={v}", 0) + 1
                break
    if not lac:
        return None
    top = ", ".join(f"{k!r} ({v} dòng)" for k, v in sorted(lac.items(), key=lambda x: -x[1])[:6])
    return (f"{_W_CONG_DOI}: {sum(lac.values())} dòng ở CẤP KHỐI (cost_center rỗng) lại mang nhãn "
            f"của một ĐƠN VỊ — {top}. Dòng tiêu đề đơn vị đã lọt vào dữ liệu, cộng nó cùng dòng "
            f"tổng của file là gấp đôi. BỎ QUA cả file: sửa `ngu_canh_dong` của spec cho khớp bố "
            f"cục mới rồi nạp lại.")


def _kiem_thang_ten_file(spec, path, recs):
    """Tên file khai tháng nào thì file phải CÓ số của tháng đó. Trả cảnh báo, hoặc None.

    VÌ SAO (bắt 29/08/2026, `Xuathoadon_GF_T7.xlsx`): file mang tên T7 nhưng ruột là BẢN SAO của
    T6 — 11 hoá đơn, 5,5331 tỷ, ngày hoá đơn toàn tháng 6. Kỳ của nguồn này suy từ CỘT NGÀY chứ
    không từ tên file, nên nó nạp "thành công" và cộng thêm một lần nữa vào tháng 6 (GF tháng 6
    hiện 11,0662 = đúng gấp đôi), còn tháng 7 thì trống trơn. Không một lớp nào bắt được: tên file
    khác nhau nên `_ghi` không đè, kỳ suy ra khác nhau nên `moi_ky_lay_file_moi_nhat` coi là hai
    kỳ độc lập, và `_SNAP_RT` thì KDVH vốn là báo cáo dòng nên không đụng tới.

    Chốt LỎNG có chủ ý — chỉ từ chối khi KHÔNG MỘT DÒNG NÀO rơi vào tháng được khai. Hoá đơn xuất
    tràn sang đầu tháng sau là chuyện bình thường, siết chặt hơn là chặn nhầm dữ liệu thật; còn
    "không có lấy một dòng của chính tháng mình" thì chắc chắn là file gửi nhầm.
    """
    pat = spec.get("thang_tu_ten_file_regex")
    if not pat or not recs:
        return None
    m = re.search(pat, os.path.basename(path), re.IGNORECASE)
    if not m:
        return None
    try:
        thang = int(m.group(1))
    except (ValueError, IndexError):
        return None
    co = sorted({(r.get("ngay") or "")[:7] for r in recs if r.get("ngay")})
    if not co or any(k[5:7] == f"{thang:02d}" for k in co):
        return None
    return (f"{_W_LECH_THANG}: tên khai T{thang} nhưng {len(recs)} dòng đọc được đều thuộc "
            f"{', '.join(co)} -> BỎ QUA cả file. Đây là lỗi FILE NGUỒN (gửi nhầm bản của tháng "
            f"khác), nạp vào là cộng đôi tháng kia và để trống tháng này. Báo kế toán gửi lại.")


def run(spec, path, write=False):
    try:
        recs, warn = extract_file(spec, path)
    finally:
        _dong_wb()          # xong một file thì nhả workbook trong đệm (xem `_mo_wb`)
    lech = _kiem_thang_ten_file(spec, path, recs) or _kiem_cong_doi(spec, recs)
    if lech:
        # Trả rỗng KÈM cảnh báo và KHÔNG có `_W_BO_LOC` -> nhánh ghi ở dưới bỏ qua, dữ liệu cũ của
        # chính file này (nếu có) giữ nguyên. Xoá là việc của người đọc cảnh báo, không phải của
        # một chốt tự động: file có thể từng nạp đúng rồi mới bị gửi đè bằng bản nhầm.
        return {"file": os.path.basename(path), "dong": 0, "ky": {},
                "canh_bao": [lech, *warn]}
    by_ky = {}
    for r in recs:
        by_ky.setdefault((r.get("ngay") or "?")[:7], []).append(r)
    out = {"file": os.path.basename(path), "dong": len(recs),
           "ky": {k: {"dong": len(v),
                      "amount": round(sum(x.get("amount") or 0 for x in v), 6)}
                  for k, v in sorted(by_ky.items())}}
    if warn:
        out["canh_bao"] = warn
    # KHÔNG có bản ghi vẫn phải ghi (tức là XOÁ phần cũ của chính file này) khi file ĐÃ ĐỌC ĐƯỢC
    # mà bộ lọc loại hết dòng — đó là "file nguồn hết số", vd claim B2C T8 nhận 13/08/2026 có cột
    # Trạng thái + Số tiền toàn #N/A. Không xoá thì DB giữ nguyên bản nạp trước đó và dashboard
    # hiện số cũ như thể vẫn đúng.
    # Ngược lại, file KHÔNG đọc được (sai sheet, thiếu cột bắt buộc) trả rỗng mà KHÔNG kèm dòng bị
    # lọc -> giữ nguyên dữ liệu cũ: spec hỏng thì im lặng còn hơn xoá trắng dữ liệu đang đúng.
    if write and (recs or any(_W_BO_LOC in w for w in warn)):
        out.update(_ghi(spec, path, recs))
    return out


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Trích xuất báo cáo theo spec JSON.")
    ap.add_argument("spec", help="id spec (memory/extract_specs/<id>.json) hoặc đường dẫn file")
    ap.add_argument("--file", help="chỉ chạy 1 file (mặc định: quét cả folder trong spec)")
    ap.add_argument("--write", action="store_true", help="ghi DB (mặc định dry-run)")
    a = ap.parse_args()
    sp = load_spec(a.spec)
    if a.file:
        files, bo = [a.file], []
    else:
        files, w = quet_nguon(sp)
        for x in w:
            print(f"[!] {x}", file=sys.stderr)
        files, bo = loc_file_moi_nhat(sp, files)
    ket_qua = [run(sp, f, write=a.write) for f in files]
    if bo:
        ket_qua.append({"_bo_qua_anh_chup_cu": bo})
    print(json.dumps(ket_qua, ensure_ascii=False, indent=2))
