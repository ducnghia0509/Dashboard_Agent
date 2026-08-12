# -*- coding: utf-8 -*-
"""Dẫn xuất CƠ CẤU CHI PHÍ (02_CHIPHI) từ Báo cáo KQKD (cùng nguồn 01_HQKD).

02_CHIPHI KHÔNG có sheet nguồn riêng — chi phí nằm trong KQKD (phần TỔNG CHI PHÍ, mã A3xx).
Bóc các dòng NHÓM chi phí cấp cao (mã ^A3[1-9]0$, BỎ A300 tổng để khỏi cộng trùng), phân loại
'Nhóm CP (chuẩn mực KT)' theo từ khoá, rồi điền 02_CHIPHI -> report_type CHIPHI (trang Chi phí).

Chạy: .venv/bin/python scripts/extract_chiphi.py <file_kqkd.xlsx> [--sheet LNQ1] --period 2026-01 --cong-ty TC
"""
import argparse
import os
import re
import sys
import unicodedata

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from servers import template_filler as tf          # noqa: E402
from servers.common import source_catalog as SC  # noqa: E402
from servers.common import contract as C           # noqa: E402


def _nm(s):
    s = str(s or "").lower().replace("đ", "d")
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if not unicodedata.combining(c))


# Phân loại Nhóm CP theo chuẩn mực kế toán (từ khoá trên tên khoản mục). Thứ tự = ưu tiên.
def _nhom_cp(name: str, ma: str = "") -> str:
    # Ưu tiên MÃ TT200 (chắc chắn) trước, rồi mới tới từ khoá tên.
    m = str(ma).strip()
    if m == "11":
        return "Giá vốn hàng bán"
    if m in ("12", "26"):          # 12 = CP quản lý kinh doanh (mẫu B02-HTX/TT133); 26 = CP QLDN (TT200)
        return "Chi phí QLDN"
    if m == "22":
        return "Chi phí tài chính"
    if m == "25":
        return "Chi phí bán hàng"
    n = _nm(name)
    if "gia von" in n:
        return "Giá vốn hàng bán"
    if "lai vay" in n or "tai chinh" in n:
        return "Chi phí tài chính"
    if any(k in n for k in ("luong", "bhxh", "quan ly", "phan bo ho", " ho", "quan tri")):
        return "Chi phí QLDN"
    if any(k in n for k in ("ban hang", "thuong", "mat bang", "marketing", "van hanh", "showroom", "show room", "tvbh", "hoa hong")):
        return "Chi phí bán hàng"
    return "Chi phí khác"


# Phân loại theo YẾU TỐ (bản chất chi phí: nhân sự/khấu hao/DV mua ngoài/khác) — dùng cho dim2
# "Yếu tố chi phí" (chart "Chi phí theo yếu tố", KHÁC _nhom_cp ở trên: đó là phân loại theo CHUẨN
# KẾ TOÁN TT200 bằng MÃ, đây theo TỪ KHOÁ trên TÊN dòng — vì mã khác nhau hoàn toàn giữa 12 đơn vị
# còn tên tiếng Việt tái diễn từ khoá giống nhau dù đổi qua tháng/đơn vị, vd Dự án đổi "Khấu hao
# phân bổ" (T01-05) -> "Chi phí khấu hao HO" (T06), cả 2 đều chứa "khấu hao"). "Giá vốn" trả thẳng
# "CP Giá vốn" (không bóc con — nguyên tắc #1 kế hoạch "Chi phí theo yếu tố", spec user chỉ nói lấy
# nguyên Mục Giá vốn); các khoản OPEX còn lại phân theo từ khoá, phần KHÔNG khớp rơi vào "Hoạt động
# khác" (đúng tinh thần spec: 5 nhóm phủ hết OPEX — "khác" bắt phần dư: lãi vay/phí NH/phân bổ
# HO/khoản lạ).
def _yeuto_cp(name: str) -> str:
    n = _nm(name)
    if "gia von" in n:
        return "CP Giá vốn"
    if any(k in n for k in ("nhan su", "nhan vien", "luong", "bhxh", "bhyt", "phuc loi", "tuyen dung",
                             "dao tao", "phu cap", "cong doan", "thuong")):
        return "CP nhân sự"
    if any(k in n for k in ("khau hao", "hao mon", "phan bo ccdc")):
        return "CP khấu hao TSCĐ"
    if ("dien" in n and "nuoc" in n) or any(  # "điện - nước"/"điện, nước"/"điện nước" đều khớp
            k in n for k in ("marketing", "quang cao", "truyen thong", "khuyen mai",
                              "mat bang", "thue vp", "thue bai", "thue mat bang",
                              "dien thoai", "internet", "itn", "cpn",
                              "sua chua", "bao duong", "bao hanh")):
        return "CP DV mua ngoài"
    return "Hoạt động khác"


def _find_kqkd_sheet(wb):
    """Sheet KQKD = sheet có mã TỔNG CHI PHÍ 'A300' (SRVF) HOẶC mã TT200 '10'/'50' (chuẩn/HT)."""
    for ws in wb.worksheets:
        for r in ws.iter_rows(min_row=1, max_row=200, values_only=True):
            vals = {str(c).strip() for c in r if c is not None}
            if "A300" in vals or ("10" in vals and any(v in vals for v in ("50", "60"))):
                return ws.title
    return None


def _header_row(ws):
    """Dòng header = dòng chứa cả 'chỉ tiêu' và ('kỳ này' hoặc 'mã'). Trả (idx, {col: role})."""
    rows = list(ws.iter_rows(min_row=1, max_row=30, values_only=True))
    for i, r in enumerate(rows):
        cells = {_nm(c): j for j, c in enumerate(r) if c not in (None, "")}
        has_ct = any(k.startswith("chi tieu") for k in cells)
        if has_ct and (any("ky nay" in k for k in cells) or any(k == "ma so" or k.startswith("ma") for k in cells)):
            ma = next((j for k, j in cells.items() if k in ("ma so", "ma")), 0)
            ct = next((j for k, j in cells.items() if k.startswith("chi tieu")), 1)
            val = next((j for k, j in cells.items() if "ky nay" in k), None)
            if val is None:  # fallback: cột giá trị kỳ này thường ngay sau 'lũy kế'
                val = next((j for k, j in cells.items() if "ky nay" in k or "thuc hien" in k), 3)
            return i, ma, ct, val
    return None


def extract(path, sheet, period, cong_ty=None):
    # read_only=True: CHỈ đọc (iter_rows). File BCTC có thể chứa sheet phantom hàng trăm nghìn dòng
    # (vd HT 'Sổ chi tiết vật tư theo lô' 200k dòng) -> mở full-parse mất ~23s/lượt; read_only ~0.1s.
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = sheet or _find_kqkd_sheet(wb)
    if not sheet or sheet not in wb.sheetnames:
        return {"ok": False, "error": "Không thấy sheet KQKD (mã A300)"}
    ws = wb[sheet]
    hr = _header_row(ws)
    if not hr:
        return {"ok": False, "error": "Không thấy header (Chỉ tiêu/Kỳ này)"}
    hi, cma, cct, cval = hr
    rows = list(ws.iter_rows(values_only=True))
    recs = []
    a_pat = re.compile(r"^A3[1-9]0$")   # SRVF: nhóm chi phí cấp cao A310..A390 (bỏ A300 tổng/A3x1 con)
    tt200 = {"11", "12", "22", "25", "26", "32"}   # TT200: giá vốn/CP tài chính/bán hàng/QLDN/khác; 12=CP QLKD (B02-HTX/TT133)

    def _row_name_val(r):
        nm_ = str(r[cct]).strip() if cct < len(r) and r[cct] not in (None, "") else ""
        vl_ = r[cval] if cval < len(r) else None
        return nm_, vl_

    n_rows = len(rows)
    i = hi + 1
    while i < n_rows:
        r = rows[i]
        ma = str(r[cma]).strip() if cma < len(r) and r[cma] not in (None, "") else ""
        if not (a_pat.match(ma) or ma in tt200):   # nhận CẢ A-series LẪN TT200
            i += 1
            continue
        name, val = _row_name_val(r)
        nhom = _nhom_cp(name, ma)
        # Gom DÒNG CON ngay sau (mã rỗng) tới khi gặp mã tiếp -> tách yếu tố chi tiết hơn (vd
        # GlobalAI 'Chi phí lương bộ phận bán hàng' nằm dưới mã '25' nhưng cột mã của DÒNG CON để
        # trống). Bỏ qua bóc con cho Giá vốn (dim2 luôn 'CP Giá vốn' nguyên khối — xem _yeuto_cp).
        j = i + 1
        children = []
        while j < n_rows:
            rj = rows[j]
            maj = str(rj[cma]).strip() if cma < len(rj) and rj[cma] not in (None, "") else ""
            if maj:
                break
            nj, vj = _row_name_val(rj)
            if nj and isinstance(vj, (int, float)) and vj:
                children.append((nj, vj))
            j += 1
        if nhom == "Giá vốn hàng bán":
            if name and isinstance(val, (int, float)) and val:
                recs.append({"Kỳ (yyyy-mm)": period, "Nhóm CP (chuẩn mực KT)": nhom,
                             "Khoản mục chi tiết": name, "Yếu tố chi phí": "CP Giá vốn",
                             "Thực hiện (tỷ)": round(val / 1e9, 9)})
        elif children:
            for nj, vj in children:
                recs.append({"Kỳ (yyyy-mm)": period, "Nhóm CP (chuẩn mực KT)": nhom,
                             "Khoản mục chi tiết": nj, "Yếu tố chi phí": _yeuto_cp(nj),
                             "Thực hiện (tỷ)": round(vj / 1e9, 9)})
        elif name and isinstance(val, (int, float)) and val:
            recs.append({"Kỳ (yyyy-mm)": period, "Nhóm CP (chuẩn mực KT)": nhom,
                         "Khoản mục chi tiết": name, "Yếu tố chi phí": _yeuto_cp(name),
                         "Thực hiện (tỷ)": round(val / 1e9, 9)})
        i = j
    if not recs:
        return {"ok": False, "error": "Không bóc được dòng chi phí (A3x0 / TT200 11/22/25/26)"}
    out = os.path.join(tf.FILLED_DIR, f"{cong_ty or 'X'}_{period}_02_CHIPHI.xlsx")
    tf.fill("02_CHIPHI", recs, out)
    # cong_ty: ưu tiên alias THƯ MỤC NGUỒN (received_reports/<folder>) như source_file/khoi — 3 pháp
    # nhân Taxi Xanh cùng tên file 'B.6.XVP' chỉ phân biệt bằng thư mục (HTXXANHVINHPHUC->HTX_XVP,
    # HTXXANHTUYENQUANG->HTX_XTQ). Trước lấy từ TÊN FILE -> ra 'XVP' sai cho 2 HTX. Không có thư mục
    # received_reports -> giữ hành vi cũ (resolve theo tên file). Chỉ đổi 2 HTX; đơn vị khác giữ nguyên.
    _parts = os.path.normpath(path).split(os.sep)
    _folder = _parts[_parts.index("received_reports") + 1] if "received_reports" in _parts \
        and _parts.index("received_reports") + 1 < len(_parts) else None
    imp = tf.import_filled(out, cong_ty=C.resolve_company(cong_ty or _folder, os.path.basename(path)),
                           khoi=C.khoi_from_path(path), source_file=SC.source_id_from_path(path))
    return {"ok": True, "sheet": sheet, "rows": imp.get("rows_imported"), "by_type": imp.get("by_type"),
            "nhom": sorted({r["Nhóm CP (chuẩn mực KT)"] for r in recs}), "out": out}


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--period", required=True)
    ap.add_argument("--cong-ty", dest="cong_ty", default=None)
    a = ap.parse_args()
    import json
    print(json.dumps(extract(a.file, a.sheet, a.period, a.cong_ty), ensure_ascii=False))
