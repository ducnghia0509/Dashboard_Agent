# -*- coding: utf-8 -*-
"""Deriver CHUYÊN BIỆT cho SRVF (CHI NHÁNH VINFAST) — CĐPS 1 TẦNG (format riêng, deriver chung
không đọc được). Sheet 'CĐPS': A=Tài khoản · B=Tên · C=Nợ đầu · D=Có đầu · E=PS Nợ · F=PS Có ·
G=Dư nợ cuối · H=Dư có cuối. Bóc theo TK (chị Điệp: công nợ CĐPS TK131/331):
  · TK 131 -> PTHU (05): dư nợ (đầu C / cuối G, tăng=PS Nợ E, giảm=PS Có F).
  · TK 331 -> PTRA (06): dư có (đầu D / cuối H, tăng=PS Có F, giảm=PS Nợ E).
  · TK 152/153/154/156 -> TỒN KHO (09): đầu=Nợ đầu C, Nhập=PS Nợ E, Xuất=PS Có F, cuối=Dư nợ G
    (spec #43 chốt 2026-07-22, thay cách cũ đọc sheet 156_xe per-model).

cong_ty=TC, khối Vinfast Showroom (SRVF thuộc Cổ phần Thịnh Cường). Idempotent theo source_file.
Chạy: .venv/bin/python scripts/derive_srvf_cdps.py <file.xlsx> --period 2026-06
"""
import argparse
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from servers import template_filler as tf  # noqa: E402
from servers.common import be_bridge as bb  # noqa: E402
import agent_cli as A  # noqa: E402

norm = lambda v: bb.normalize_header(v, True)  # noqa: E731
TY = lambda v: round(v * 1e-9, 9) if isinstance(v, (int, float)) else None  # noqa: E731


def _cols(rows):
    """Dò dòng HEADER CỘT (không phải dòng tiêu đề) + index cột theo VAI. Header cột = dòng mà cột
    'phát sinh nợ' + 'dư nợ cuối' dò được thành ô RIÊNG (dòng tiêu đề gộp 1 ô -> loại)."""
    for i, r in enumerate(rows[:12]):
        low = [norm(c) for c in r]
        def col(*kw):
            return next((j for j, c in enumerate(low) if all(k in c for k in kw)), None)
        m = {"tk": col("tai khoan"), "no_dau": col("no dau"), "co_dau": col("co dau"),
             "ps_no": col("phat sinh no"), "ps_co": col("phat sinh co"),
             "no_cuoi": col("du no cuoi"), "co_cuoi": col("du co cuoi")}
        if m["ps_no"] is not None and m["no_cuoi"] is not None:   # đúng dòng header cột
            return i, m
    return None, None


def _cdkt_ma(wbk, norm, ma):
    """Đọc 1 mã số CĐKT (Số đầu kỳ, Số cuối kỳ) từ sheet CĐKT chuẩn TT200 (cột 'Số cuối kỳ'/'Số đầu
    kỳ'). Trả (dau, cuoi) VND hoặc (None, None) nếu không thấy sheet/mã."""
    sn = next((s for s in wbk.sheetnames if "cdkt" in norm(s) or "can doi ke toan" in norm(s)
               or "đkt" in s.lower()), None)
    if sn is None:
        return (None, None)
    rows = [list(r) for r in wbk[sn].iter_rows(values_only=True)]
    hi = next((i for i, r in enumerate(rows[:15])
               if any(norm(c) == "ma so" for c in r) and any("cuoi" in norm(c) for c in r)), None)
    if hi is None:
        return (None, None)
    hdr = rows[hi]
    ma_j = next((j for j, c in enumerate(hdr) if norm(c) == "ma so"), None)
    cuoi_j = next((j for j, c in enumerate(hdr) if "cuoi" in norm(c)), None)
    dau_j = next((j for j, c in enumerate(hdr) if "dau" in norm(c) and "nam" not in norm(c)), None)
    if ma_j is None or cuoi_j is None:
        return (None, None)
    for r in rows[hi + 1:]:
        if ma_j < len(r) and str(r[ma_j]).strip() == ma:
            _num = lambda j: r[j] if (j is not None and j < len(r) and isinstance(r[j], (int, float))) else None
            return (_num(dau_j), _num(cuoi_j))
    return (None, None)


def extract(path, period, cong_ty="TC"):
    wb = bb.fast_load_workbook(path, data_only=True, read_only=True)
    if "CĐPS" not in wb.sheetnames:
        return {"ok": False, "error": "không thấy sheet CĐPS"}
    rows = [list(r) for r in wb["CĐPS"].iter_rows(values_only=True)]
    # SỐ DƯ công nợ lấy từ CĐKT (hướng dẫn C Điệp cập nhật: #30 phải thu = CĐKT Mã 131; #36 phải trả
    # = CĐKT Mã 311). CĐPS chỉ dùng cho PS tăng/giảm (#31/#37). CĐPS≠CĐKT ở tháng có phân loại lệch
    # (vd T05: CĐPS 131=198,464 vs CĐKT=198,460). None -> fallback CĐPS (giữ tương thích cũ).
    _ck131_dau, _ck131_cuoi = _cdkt_ma(wb, norm, "131")   # phải thu KH ngắn hạn
    _ck311_dau, _ck311_cuoi = _cdkt_ma(wb, norm, "311")   # phải trả người bán ngắn hạn
    wb.close()
    hi, c = _cols(rows)
    if hi is None:
        return {"ok": False, "error": "không dò được header CĐPS"}
    src, khoi = A._source_id(path), A._khoi_of(path)

    def find_tk(tk):
        for r in rows[hi + 1:]:
            if r and str(bb.parse_text(r[c["tk"]])).strip() == tk:
                return r
        return None

    def val(r, key):
        j = c[key]
        return TY(r[j]) if (j is not None and j < len(r)) else None

    out = {"ok": True}
    # PTHU (TK 131, dư NỢ)
    r131 = find_tk("131")
    if r131:
        rec = [{"Kỳ": period, "Đơn vị": cong_ty, "Khách hàng": "Phải thu khách hàng (tổng)",
                "Dư cuối kỳ (tỷ)": TY(_ck131_cuoi) if _ck131_cuoi is not None else val(r131, "no_cuoi"),
                "Dư đầu kỳ (tỷ)": TY(_ck131_dau) if _ck131_dau is not None else val(r131, "no_dau"),
                "PS tăng - Nợ (tỷ)": val(r131, "ps_no"), "PS giảm - Có (tỷ)": val(r131, "ps_co")}]
        p = os.path.join(tf.FILLED_DIR, f"SRVF_{period}_05_PHAITHU.xlsx")
        tf.fill("05_PHAITHU", rec, p)
        out["pthu"] = tf.import_filled(p, cong_ty=cong_ty, khoi=khoi, source_file=src).get("rows_imported")
    # PTRA (TK 331, dư CÓ)
    r331 = find_tk("331")
    if r331:
        rec = [{"Kỳ": period, "Đơn vị": cong_ty, "Nhà cung cấp": "Phải trả nhà cung cấp (tổng)",
                "Dư cuối kỳ (tỷ)": TY(_ck311_cuoi) if _ck311_cuoi is not None else val(r331, "co_cuoi"),
                "Dư đầu kỳ (tỷ)": TY(_ck311_dau) if _ck311_dau is not None else val(r331, "co_dau"),
                "PS tăng - Có (tỷ)": val(r331, "ps_co"), "PS giảm - Nợ (tỷ)": val(r331, "ps_no")}]
        p = os.path.join(tf.FILLED_DIR, f"SRVF_{period}_06_PHAITRA.xlsx")
        tf.fill("06_PHAITRA", rec, p)
        out["ptra"] = tf.import_filled(p, cong_ty=cong_ty, khoi=khoi, source_file=src).get("rows_imported")
    # CHIỀU ĐẢO công nợ (ADDITIVE, cho thẻ "Trả trước NCC"/"Người mua trả tiền trước" màn Công nợ —
    # FE đọc report_type PTRA_ADV/PTHU_ADV): TK331 dư NỢ cuối = mình ứng trước cho NCC (mã 132);
    # TK131 dư CÓ cuối = khách ứng trước cho mình (mã 312/321). Nguồn CĐPS (khớp CĐKT). Insert trực
    # tiếp (report_type ADDITIVE, không có template); idempotent theo source_file+period+report_type.
    _adv = [("PTRA_ADV", r331, "no_cuoi", "Trả trước NCC (tổng)"),         # TK331 dư Nợ = ứng trước NCC
            ("PTHU_ADV", r131, "co_cuoi", "Người mua trả tiền trước (tổng)")]  # TK131 dư Có = khách ứng trước
    _dbh = bb.db.get_db()
    for _rt, _r, _key, _lbl in _adv:
        if not _r:
            continue
        _v = val(_r, _key)
        _dbh.execute("DELETE FROM raw_rows WHERE source_file=? AND period_month=? AND report_type=?",
                     (src, period, _rt))
        if _v is None or abs(_v) < 1e-9:
            continue
        _twin = _dbh.execute(
            "SELECT dataset_id, ngay FROM raw_rows WHERE source_file=? AND period_month=? "
            "AND report_type IN ('PTRA','PTHU') LIMIT 1", (src, period)).fetchone()
        if not _twin:
            continue
        import json as _json
        _dbh.execute(
            "INSERT INTO raw_rows(dataset_id,report_type,row_index,ngay,cong_ty,khoi,cost_center,"
            "period_month,amount,amount2,dim1,dim2,dim3,payload,source_file) "
            "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (_twin["dataset_id"], _rt, 6000000, _twin["ngay"], cong_ty, khoi, None, period,
             _v, None, _lbl, None, None,
             _json.dumps({"unit": "ty", "nguon": f"CĐPS {_key}"}, ensure_ascii=False), src))
        out[_rt.lower()] = _v
    _dbh.commit()
    # THUẾ: TK 133 (GTGT được khấu trừ = phải thu, dư NỢ) + TK 333 (thuế phải nộp, dư CÓ). Điền đủ
    # đầu/PS tăng/PS giảm để bảng cân (cuối = đầu + tăng − giảm). Số dư RÒNG (chiều chính − ngược).
    #   133: đầu=Nợ đầu−Có đầu, tăng=PS Nợ, giảm=PS Có, cuối=Nợ cuối−Có cuối.
    #   333: đầu=Có đầu−Nợ đầu, tăng=PS Có, giảm=PS Nợ, cuối=Có cuối−Nợ cuối.
    def _net(r, pos, neg):
        p, n = val(r, pos), val(r, neg)
        return None if (p is None and n is None) else round((p or 0) - (n or 0), 9)
    thue = []
    for tk, pt, dpos, dneg, cpos, cneg, inc, dec in (
            ("133", "Phải thu", "no_dau", "co_dau", "no_cuoi", "co_cuoi", "ps_no", "ps_co"),
            ("333", "Phải nộp", "co_dau", "no_dau", "co_cuoi", "no_cuoi", "ps_co", "ps_no")):
        r = find_tk(tk)
        if not r:
            continue
        cuoi, dau, tang, giam = _net(r, cpos, cneg), _net(r, dpos, dneg), val(r, inc), val(r, dec)
        if any(abs(v or 0) > 1e-9 for v in (cuoi, dau, tang, giam)):   # bỏ dòng toàn 0
            ten = bb.parse_text(r[c["tk"] + 1]) if c["tk"] + 1 < len(r) else None
            thue.append({"Kỳ": period, "Đơn vị": cong_ty,
                         "Loại thuế (GTGT ra/vào, TNCN, TNDN, NK, khác)": ten or f"TK {tk}",
                         "Phải thu/Phải nộp": pt, "Dư đầu kỳ (tỷ)": dau,
                         "PS tăng (tỷ)": tang, "PS giảm (tỷ)": giam, "Dư cuối kỳ (tỷ)": cuoi})
    if thue:
        p = os.path.join(tf.FILLED_DIR, f"SRVF_{period}_10_THUE.xlsx")
        tf.fill("10_THUE", thue, p)
        out["thue"] = tf.import_filled(p, cong_ty=cong_ty, khoi=khoi, source_file=src).get("rows_imported")

    # TỒN KHO (spec #43 chốt 2026-07-22): các TK KHO 152 (NVL) / 153 (CCDC) / 154 (SPDD) / 156 (hàng
    # hóa) TRÊN CĐPS. Mỗi TK 1 dòng: đầu = Nợ đầu kỳ (C), Nhập = Phát sinh Nợ (E), Xuất = Phát sinh
    # Có (F), cuối = Dư nợ cuối kỳ (G). Tổng nhập/xuất màn Tồn kho = Σ 4 TK -> khớp #43. Nhất quán
    # _derive_tonkho_cdps (đơn vị khác). TRƯỚC lấy 156_xe per-model + dòng "khác" cân theo CĐKT mã140
    # -> nhập/xuất CHỈ phản ánh xe (thiếu movement 152/153/154 & 156 ngoài xe) -> sai #43.
    _TK_KHO = {"152": "Nguyên vật liệu", "153": "Công cụ, dụng cụ",
               "154": "Chi phí SXKD dở dang", "156": "Hàng hóa"}
    recs = []
    for _tk, _default_ten in _TK_KHO.items():
        r = find_tk(_tk)
        if not r:
            continue
        _dau, _cuoi = val(r, "no_dau"), val(r, "no_cuoi")
        # Giữ TK có tồn ĐẦU hoặc CUỐI > 0 (bỏ TK chạy-qua đầu=cuối=0 dù có PS) — như _derive_tonkho_cdps.
        if not ((_dau and abs(_dau) > 1e-9) or (_cuoi and abs(_cuoi) > 1e-9)):
            continue
        _ten = bb.parse_text(r[c["tk"] + 1]) if c["tk"] + 1 < len(r) else None
        recs.append({"Kỳ": period, "Đơn vị": cong_ty,
                     "Loại HTK (NVL/Vật tư/Hàng hóa…)": _ten or _default_ten,
                     "TK (151-156)": _tk, "Dư đầu kỳ (tỷ)": _dau,
                     "Nhập trong kỳ (tỷ)": val(r, "ps_no"), "Xuất trong kỳ (tỷ)": val(r, "ps_co"),
                     "Dư cuối kỳ (tỷ)": _cuoi})
    if recs:
        p = os.path.join(tf.FILLED_DIR, f"SRVF_{period}_09_TONKHO.xlsx")
        tf.fill("09_TONKHO", recs, p)
        out["tonkho"] = tf.import_filled(p, cong_ty=cong_ty, khoi=khoi, source_file=src).get("rows_imported")

    if not any(k in out for k in ("pthu", "ptra", "thue", "tonkho")):
        return {"ok": False, "error": "không bóc được gì"}
    return out


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--period", required=True)
    a = ap.parse_args()
    import json
    print(json.dumps(extract(a.file, a.period), ensure_ascii=False))
