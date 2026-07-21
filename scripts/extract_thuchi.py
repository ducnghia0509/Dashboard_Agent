# -*- coding: utf-8 -*-
"""Dẫn xuất DÒNG TIỀN THU-CHI THEO TỪNG PHÁP NHÂN (03_DONGTIEN -> THUCHI) từ Báo cáo tiền tập đoàn.

File thu chi có sheet RIÊNG mỗi pháp nhân ('BC THU CHI_T<m>_<CTY>'): mục I = Tổng thu, II = Tổng chi,
dưới đó là các KHOẢN MỤC cấp 1 (1,2,3...). Bóc từng khoản mục -> dim1 'A. <km>' (thu)/'B. <km>' (chi),
gắn Mã Công ty theo sheet -> dashboard tách dòng tiền theo công ty (lọc công ty hoạt động).

Chạy: .venv/bin/python scripts/extract_thuchi.py <file_thuchi.xlsx> --period 2026-01
"""
import argparse
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from servers import template_filler as tf  # noqa: E402
from servers.common import source_catalog as SC  # noqa: E402
from servers.common import org  # noqa: E402

# Hậu tố sheet 'BC THU CHI_T1_<SUFFIX>' -> mã pháp nhân. T6+ tách 'AN' thành
# 'ANTAXI'/'ANKS' — cùng pháp nhân AAG (2 khối 7/10; THUCHI cố ý không gắn khối).
_SUFFIX_CO = {"TC": "TC", "VFQN": "VFQN", "XANH": "XVP", "HUNGTHINH": "HT", "AN": "AAG",
              "ANTAXI": "AAG", "ANKS": "AAG"}
_COL_KY = "Kỳ / Ngày"
_COL_CTY = "Mã Công ty (auto từ CC)"
_COL_LOAI = "Loại (Thu/Chi)"
_COL_KM = "Khoản mục (Thu bán hàng, Thu đầu tư, Chi NCC, Chi tài chính, Chi đầu tư TS…)"
_COL_TH = "Thực hiện (tỷ)"


def _company_of(sheet: str):
    """Suy mã pháp nhân từ tên sheet (hậu tố sau '_'). None nếu là sheet TỔNG/không khớp."""
    up = sheet.upper()
    if "TỔNG" in up or "TONG" in up:
        return None
    suf = up.rsplit("_", 1)[-1].strip()
    return _SUFFIX_CO.get(suf)


def _val_col(rows):
    """Cột giá trị = 'TM + TG + T.VAY' (tổng kỳ). Trả index; mặc định 3 (đã quan sát)."""
    for r in rows[:12]:
        for j, c in enumerate(r):
            s = str(c).strip().upper() if c is not None else ""
            if "TM" in s and "VAY" in s:
                return j
    return 3


def extract(path: str, period: str, cong_ty: str = None) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = [s for s in wb.sheetnames if re.search(r"THU CHI_T\d", s.upper()) and _company_of(s)]
    if not sheets:
        return {"ok": False, "error": "Không thấy sheet thu chi theo pháp nhân (BC THU CHI_T*_<CTY>)"}
    recs = []
    for sh in sheets:
        co = _company_of(sh)
        ws = wb[sh]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        vc = _val_col(rows)
        sec = None   # 'A' (thu) khi gặp mục I; 'B' (chi) khi gặp mục II
        for r in rows:
            c0 = str(r[0]).strip() if r[0] not in (None, "") else ""
            c1 = str(r[1]).strip() if len(r) > 1 and r[1] not in (None, "") else ""
            if c0 == "I":
                sec = "A"; continue
            if c0 == "II":
                sec = "B"; continue
            if sec and re.fullmatch(r"\d+", c0) and c1:   # khoản mục CẤP 1 (1,2,3...) — bỏ 1.1/'+...'
                v = r[vc] if vc < len(r) and isinstance(r[vc], (int, float)) else None
                if v is None:
                    continue
                recs.append({_COL_KY: period, _COL_CTY: co,
                             _COL_LOAI: "Thu" if sec == "A" else "Chi",
                             _COL_KM: c1, _COL_TH: round(v / 1e9, 9)})
    if not recs:
        return {"ok": False, "error": "Không bóc được khoản mục thu/chi nào"}
    out = os.path.join(tf.FILLED_DIR, f"THUCHI_{period}_03_DONGTIEN.xlsx")
    tf.fill("03_DONGTIEN", recs, out)
    # cong_ty=None: giữ Mã Công ty THEO DÒNG (mỗi khoản mục đã gắn pháp nhân), không stamp đè.
    imp = tf.import_filled(out, cong_ty=None, source_file=SC.source_id_from_path(path))
    return {"ok": True, "sheets": sheets, "companies": sorted({r[_COL_CTY] for r in recs}),
            "rows": imp.get("rows_imported"), "by_type": imp.get("by_type"), "out": out}


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--period", required=True)
    ap.add_argument("--cong-ty", dest="cong_ty", default=None)
    a = ap.parse_args()
    import json
    print(json.dumps(extract(a.file, a.period, a.cong_ty), ensure_ascii=False))
