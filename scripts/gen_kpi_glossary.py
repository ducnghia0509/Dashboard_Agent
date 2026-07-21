# -*- coding: utf-8 -*-
"""Sinh kpi_glossary.json TỪ knowledge/50_chi_tieu.yaml (business-authored, 2026-07-09).

NGUỒN SỰ THẬT MỚI cho nội dung 50 chỉ tiêu (tên/công thức/sheet đích/màn FE/trạng thái dữ
liệu): knowledge/50_chi_tieu.yaml — thay cho sheet Template_chuan!00_50CHITIEU trước đây
(sheet đó KHÔNG bị xoá, các pipeline khác — contract.chi_tieu_50(), DashBoard_AI
metrics_extra.py — vẫn đọc trực tiếp Excel, KHÔNG đổi theo file này).

Ngưỡng cảnh báo (nguong_vang/nguong_do/trang_thai_den/gia_tri_mau/chieu_phan_tich/
chieu_danh_gia) CHƯA có trong 50_chi_tieu.yaml (tính năng chỉnh ngưỡng phát triển sau) —
JOIN bổ sung từ sheet Excel cũ theo tên chỉ tiêu đã chuẩn hoá (_norm), best-effort. Chỉ
tiêu nào không khớp tên ở Excel thì các field ngưỡng để rỗng — QA (glossary_lookup) khi đó
nói "chưa có ngưỡng cấu hình" thay vì bịa số.

Chạy: python scripts/gen_kpi_glossary.py   -> ghi đè kpi_glossary.json
"""
import json
import os
import re
import sys
import unicodedata

import yaml
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
# Ưu tiên env GOLDEN_TEMPLATE (khớp contract.py), mặc định ~/template_trust/Template_chuan.xlsx
TEMPLATE_XLSX = os.environ.get("GOLDEN_TEMPLATE") or os.path.normpath(
    os.path.join(ROOT, "..", "template_trust", "Template_chuan.xlsx"))
KNOWLEDGE_DIR = os.environ.get("KNOWLEDGE_DIR", "/home/sysadmin/knowledge")
CHI_TIEU_YAML = os.path.join(KNOWLEDGE_DIR, "50_chi_tieu.yaml")
OUT_PATH = os.path.join(ROOT, "kpi_glossary.json")
SHEET_NAME = "00_50CHITIEU"


def _norm(s) -> str:
    s = str(s or "").strip().lower().replace("đ", "d")  # NFKD KHÔNG tách đ(U+0111)
    s = unicodedata.normalize("NFKD", s)
    return " ".join("".join(c for c in s if not unicodedata.combining(c)).split())


# Nhãn header -> khoá chuẩn (khớp theo _norm, bền với đổi thứ tự/typo cột).
_COL_ALIASES = {
    "stt": "stt",
    "nhom chi tieu": "nhom_con", "nhom chi tieu (theo html)": "nhom_con",
    "chi tieu": "chi_tieu", "chi tieu (ten dung tren html)": "chi_tieu",
    "gia tri mau": "gia_tri_mau", "gia tri mau tren html": "gia_tri_mau",
    "trang thai den": "trang_thai_den", "trang thai den mau": "trang_thai_den",
    "cong thuc": "cong_thuc", "cong thuc / logic": "cong_thuc",
    "dvt": "don_vi",
    "nguon du lieu": "nguon_du_lieu",
    "sheet nhap lieu": "sheet_nhap",
    "cot nguon": "cot_nguon", "cot nguon trong sheet": "cot_nguon",
    "chieu phan tich": "chieu_phan_tich",
    "nguong vang": "nguong_vang",
    "nguong do": "nguong_do",
    "chieu danh gia": "chieu_danh_gia",
}


def _find_header(ws):
    """Tìm dòng header (dòng có ô 'Chỉ tiêu' + 'Công thức'); trả (row_idx, {col_idx: key})."""
    for r in range(1, min(ws.max_row, 6) + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        keys = {}
        for ci, v in enumerate(vals, start=1):
            k = _COL_ALIASES.get(_norm(v))
            if k:
                keys[ci] = k
        if "chi_tieu" in keys.values() and "cong_thuc" in keys.values():
            return r, keys
    raise RuntimeError("Không tìm thấy dòng header trong 00_50CHITIEU")


_THRESHOLD_FIELDS = ("nguong_vang", "nguong_do", "trang_thai_den", "gia_tri_mau",
                      "chieu_phan_tich", "chieu_danh_gia", "don_vi", "nhom_bao_cao")


def _load_threshold_lookup(path=None, sheet=SHEET_NAME):
    """Đọc sheet Excel cũ CHỈ để lấy field ngưỡng cảnh báo (chưa có trong yaml mới), trả
    {ten_chuan_hoa: {field ngưỡng}}. Best-effort — thiếu file/sheet thì trả {} (không chặn build)."""
    path = path or TEMPLATE_XLSX
    if not os.path.exists(path):
        return {}
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb[sheet]
        hdr_row, colmap = _find_header(ws)
    except Exception as e:
        print(f"CẢNH BÁO: không đọc được ngưỡng từ {sheet} ({e}) — bỏ qua, để rỗng.", file=sys.stderr)
        return {}

    lookup = {}
    for r in range(hdr_row + 1, ws.max_row + 1):
        rec = {}
        for ci, key in colmap.items():
            v = ws.cell(r, ci).value
            rec[key] = "" if v is None else str(v).strip()
        chi_tieu = rec.get("chi_tieu", "")
        if not chi_tieu or chi_tieu.startswith("▶"):
            continue
        lookup[_norm(chi_tieu)] = {f: rec.get(f, "") for f in _THRESHOLD_FIELDS}
    return lookup


def build_glossary(yaml_path=None, xlsx_path=None):
    yaml_path = yaml_path or CHI_TIEU_YAML
    with open(yaml_path, encoding="utf-8") as fh:
        doc = yaml.safe_load(fh)

    threshold_lookup = _load_threshold_lookup(xlsx_path)
    unmatched = 0

    items = []
    for nhom_key, nhom_label in (
        ("chi_tieu_tai_chinh", None), ("chi_tieu_nhan_su", "Nhân sự"),
        ("chi_tieu_van_hanh", "Vận hành"),
    ):
        for item in doc.get(nhom_key) or []:
            items.append((item, nhom_label))
    items.sort(key=lambda pair: pair[0].get("stt", 0))

    records = []
    for item, group_fallback in items:
        chi_tieu = str(item.get("ten") or "").strip()
        if not chi_tieu:
            continue
        co_tren = str(item.get("co_tren_dashboard") or "").strip()
        ghi_chu = " ".join(x for x in [item.get("ghi_chu"), item.get("ghi_chu_srvf")] if x)
        rec = {
            "id": item.get("stt"),
            "stt": str(item.get("stt") or ""),
            "nhom_con": item.get("nhom") or group_fallback or "",
            "chi_tieu": chi_tieu,
            "cong_thuc": str(item.get("cong_thuc") or ""),
            "nguon_du_lieu": str(item.get("cot_lay") or ""),
            "sheet_nhap": str(item.get("template_sheet") or ""),
            "cot_nguon": str(item.get("cot_lay") or ""),
            "y_nghia": str(item.get("y_nghia") or ""),
            "man_hien_thi": item.get("man_hien_thi") or [],
            "khoi_ap_dung": item.get("khoi_ap_dung"),
            "co_tren_dashboard": co_tren,
            "ghi_chu": ghi_chu,
            # Ngưỡng cảnh báo: JOIN best-effort từ Excel cũ (xem module docstring).
            "nguong_vang": "", "nguong_do": "", "trang_thai_den": "", "gia_tri_mau": "",
            "chieu_phan_tich": "", "chieu_danh_gia": "", "don_vi": "", "nhom_bao_cao": "",
        }
        th = threshold_lookup.get(_norm(chi_tieu))
        if th:
            rec.update(th)
        else:
            unmatched += 1
        # 'canh_bao_do' cũ ~ ngưỡng đỏ (để glossary_lookup vẫn khớp từ khoá cảnh báo).
        rec["canh_bao_do"] = rec.get("nguong_do", "")
        # needs_followup: trạng thái business tự khai (co_tren_dashboard == "chua") HOẶC
        # ghi_chu cảnh báo "CHƯA CÓ NGUỒN" toàn phần/đa số đơn vị. Trước đây chỉ nhìn
        # co_tren_dashboard nên 8 chỉ tiêu ghi rõ "CHƯA CÓ NGUỒN" vẫn wired=true -> QA
        # khẳng định có số thật (bịa số). Thiếu nguồn MỘT PHẦN (vd "SRVF: CHƯA CÓ NGUỒN,
        # đơn vị khác: có") vẫn wired nhưng mang canh_bao_nguon để QA nói rõ phần thiếu.
        gnorm = _norm(ghi_chu)
        thieu_nguon = "chua co nguon" in gnorm
        thieu_toan_bo = thieu_nguon and (
            "tat ca" in gnorm or "da so" in gnorm
            or not re.search(r"(co tu|con lai|khac\s*:|van co)", gnorm))
        rec["canh_bao_nguon"] = ghi_chu if thieu_nguon else ""
        rec["needs_followup"] = (co_tren == "chua") or thieu_toan_bo
        rec["wired"] = bool(rec["sheet_nhap"]) and not rec["needs_followup"]
        records.append(rec)

    return records, unmatched


def main():
    if not os.path.exists(CHI_TIEU_YAML):
        print(f"Không tìm thấy {CHI_TIEU_YAML}", file=sys.stderr)
        sys.exit(1)
    records, unmatched = build_glossary()
    with open(OUT_PATH, "w", encoding="utf-8") as fh:
        json.dump(records, fh, ensure_ascii=False, indent=2)
    n_follow = sum(1 for r in records if r["needs_followup"])
    n_wired = sum(1 for r in records if r["wired"])
    print(f"Đã ghi {OUT_PATH}: {len(records)} chỉ tiêu "
          f"({n_wired} đã nối sheet, {n_follow} cần chốt nguồn, {unmatched} thiếu ngưỡng cảnh báo "
          f"do không khớp tên ở {SHEET_NAME}). Nguồn nội dung: {os.path.basename(CHI_TIEU_YAML)}.")


if __name__ == "__main__":
    main()
