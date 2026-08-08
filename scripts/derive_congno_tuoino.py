# -*- coding: utf-8 -*-
"""Deriver: cụm "Phân loại phải thu theo kỳ hạn thu nợ (aging)" — Khối 9, MÀN HÌNH III.1 "Công nợ"
(spec CHÍNH THỨC: /home/itadmin/MoTaChiTiet_Dashboard_TaiChinh.docx).

CHỐT LẠI 2026-08-03 (sau khi đối chiếu docx — bản đầu 2026-08-03 tự dựng bảng "Tuổi nợ phải thu
theo đơn vị" KHÔNG có trong spec, đã bỏ): docx định nghĩa field "aging" như sau —
  Thành phần: < 1 tháng / 1–3 / 3–6 / > 6 tháng
  Lấy trường: Dư PT cuối kỳ chia theo tuổi nợ
  Nguồn: Sổ chi tiết công nợ 131 của đơn vị (ngày phát sinh)
  Công thức: = Ngày báo cáo − Ngày phát sinh; nhóm theo dải
Đây CHÍNH LÀ field `aging` (payload keys `tuoi_no_1t`/`aging_13`/`aging_36`/`aging_6p`) đã có sẵn
trong `debt.py::debt_extras()` — trước giờ luôn 0 vì KHÔNG deriver nào ghi các key này (verify: field
tồn tại ở MỌI dòng PTHU nhưng luôn =0, là placeholder từ template, không phải nguồn riêng công ty
nào). KHÔNG phải "tuổi nợ theo hạn nợ" (trong hạn/đến hạn/quá hạn) — đó là khái niệm KHÁC (dựa Ngày
đến hạn) mà 3 đơn vị này không có (Thời hạn nợ ghi 1825 ngày = 5 năm -> luôn "trong hạn").

Nguồn dữ liệu (per-khách-hàng): `received_reports/<FOLDER>/baocaotuoinophaithu/*.xlsx`, sheet "Báo
cáo tuổi nợ_Mẫu(Tuổi nợ phả...)". Cột cần: "Ngày hóa đơn" (= Ngày phát sinh, cột G) + "Tổng nợ phải
thu" (= dư PT cuối kỳ của dòng đó, cột J) — dò theo TÊN header, không cứng chữ cột.

Tính PER DÒNG (không dùng cột quá-hạn-theo-hạn-nợ N-Q có sẵn trong file — khác trục với "aging"):
  age_days = Ngày báo cáo (CUỐI kỳ, vd 2026-06 -> 2026-06-30) − Ngày hóa đơn
  b1  (< 1 tháng):  age_days < 30
  b13 (1–3 tháng):  30 <= age_days < 90
  b36 (3–6 tháng):  90 <= age_days < 180
  b6p (> 6 tháng):  age_days >= 180
rồi CỘNG "Tổng nợ phải thu" của dòng đó vào đúng dải.

report_type MỚI 'PTHU_TUOINO' (additive, không đụng PTHU/PTRA vàng — cùng pattern
derive_congno_advance.py PTHU_ADV/PTRA_ADV). 1 dòng TỔNG/công ty/kỳ, payload dùng ĐÚNG tên field
`debt.py` đã đọc sẵn (`tuoi_no_1t`/`aging_13`/`aging_36`/`aging_6p`) — debt.py cần sửa để GỘP
report_type này vào cùng `pt` khi tính `aging` (xem debt.py, dòng "aging = ...").

_UNITS: dict folder-nguồn -> (cong_ty, khoi, mode). Chốt 2026-08-03 (bản 2, mở rộng thêm HO/TRẠM SẠC/
GLOBAL AI theo spec mới — CHƯA có file thật, chỉ thêm cấu hình trước) — CHÚ Ý: HO và TRẠM SẠC
CÙNG `cong_ty='TC'` (khác `khoi`) -> PHẢI lưu khoi tường minh ở đây (KHÔNG suy qua "twin" theo
cong_ty như bản đầu, vì twin sẽ vớ nhầm khoi của 1 trong 2 khi cong_ty trùng nhau). DELETE khi ghi
lại cũng scope theo `source_file` (KHÔNG theo cong_ty+period) — cùng lý do, tránh HO đè/xoá nhầm
dòng của TRẠM SẠC. Thêm đơn vị mới: 1 dòng vào dict này (verify cong_ty/khoi qua
`SELECT DISTINCT cong_ty,khoi FROM raw_rows WHERE source_file ILIKE '<FOLDER>::%' LIMIT 5` trên DB
coding trước khi thêm, KHÔNG đoán).

`mode` chọn cách tính, vì 2 nhóm đơn vị có file nguồn KHÁC TRỤC nhau (2026-08-03, thêm SRVF):
  - "age": XVP/HTX_XTQ/HTX_XVP/HO/TRẠM SẠC/GLOBAL AI — "Thời hạn nợ" ghi 1825 ngày (5 năm) nên
    cột quá-hạn-theo-hạn-nợ (K..Q, xem dưới) LUÔN 0/rỗng, không dùng được -> tính age_days từ
    "Ngày hóa đơn" như spec docx (per dòng, xem _find_cols/vòng lặp bucket b1/b13/b36/b6p).
  - "hanno": SRVF — file có "Thời hạn nợ"/"Ngày đến hạn" THẬT (không phải 5 năm), sheet đã tự
    tính sẵn theo khách hàng các cột (dò theo TÊN, xem _find_cols_hanno):
      J "Tổng nợ phải thu" (= K+L+M, verify được ở dòng 0 = dòng tổng của sheet)
      K "Công nợ trong hạn"      L "Công nợ đến hạn"       M "Công nợ quá hạn" (= N+O+P+Q)
      N quá hạn 1-30 ngày · O >30-90 · P >90-180 · Q >180 ngày (sub-header dòng dưới header chính)
  - "hanno_tong" / "hanno_tong_xdv": ĐỌC THẲNG DÒNG TỔNG nằm NGAY TRÊN header (HT/Dự án/XDV) —
    spec user chỉ đích danh ô ở dòng đó, xem `_agg_hanno_tong`. Riêng XDV bố cục cột khác template
    (không có "đến hạn", 5 dải quá hạn thay vì 4) -> resolver riêng `_find_cols_xdv`.
    -> KHÔNG tính age_days per dòng (khác hẳn "age") — CỘNG THẲNG giá trị các cột trên qua mọi
    dòng khách hàng. Payload dùng field riêng (tong_no/trong_han/den_han/qua_han/qh_1_30/qh_30_90/
    qh_90_180/qh_180p) — KHÔNG tái dùng key `tuoi_no_1t`/`aging_13`/`aging_36`/`aging_6p` của mode
    "age" vì ý nghĩa khác nhau (age = tuổi TOÀN BỘ dư nợ tính từ hoá đơn; hanno.N..Q chỉ là phần
    QUÁ HẠN, thiếu phần K+L "chưa đến hạn" — gộp lẫn vào cùng field sẽ làm sai lệch donut "Σ dải =
    Tổng dư PT" của debt.py cho các đơn vị mode "age"). debt.py cần đọc thêm field mới này riêng
    khi muốn hiển thị cho SRVF — CHƯA sửa debt.py ở đây (out of scope, agent khác lo phần đọc/UI).

Chạy (dry-run in tổng, KHÔNG ghi):
  .venv/bin/python scripts/derive_congno_tuoino.py <file.xlsx> --period 2026-06
Ghi thật:
  .venv/bin/python scripts/derive_congno_tuoino.py <file.xlsx> --period 2026-06 --write
"""
import argparse
import calendar
import datetime as dt
import json
import os
import re
import sys
import unicodedata

import openpyxl
import psycopg

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from servers.common import source_catalog as _SC  # noqa: E402

DB_URL = (os.environ.get("DATABASE_URL") or os.environ.get("TC_DATABASE_URL")
          or "postgresql://tc:tc@localhost:5433/tc_dashboard")

# folder nguồn (thư mục received_reports/<FOLDER>) -> (cong_ty, khoi). CHỈ đơn vị đã có nguồn trong
# spec — đơn vị khác không nằm trong dict này -> derive() no-op (skip=True), không ảnh hưởng gì.
# (cong_ty, khoi) verify qua raw_rows thật kỳ 2026-06 (DB coding 5435) trước khi thêm.
_UNITS = {
    # XVP/HTX_XTQ/HTX_XVP/TRẠM SẠC/GLOBAL AI đổi 'age'->'hanno' (2026-08-08, QA đối chiếu chart
    # "Phân loại phải thu theo kỳ hạn thu nợ" ở màn Công nợ với chỉ tiêu #14 Tổng quan): giả định cũ
    # "Thời hạn nợ ghi 1825 ngày = 5 năm nên cột quá hạn luôn 0" đã lỗi thời — file thật hiện có ĐỦ
    # cột trong hạn/đến hạn/quá hạn với số liệu KHÁC 0 (vd Trạm sạc T06/2026 quá hạn 476,35 triệu).
    # `_agg_hanno` giờ tính LUÔN age-bucket (payload tuoi_no_1t/aging_13/aging_36/aging_6p) song song
    # field hạn-nợ nên đổi mode KHÔNG làm mất số ở chart "aging" như trước khi thêm tính năng đó.
    "XANHVINHPHUC": ("XVP", "Khối KD Vận tải Taxi Xanh", "hanno"),
    "HTXXANHTUYENQUANG": ("HTX_XTQ", "Khối KD Vận tải Taxi Xanh", "hanno"),
    "HTXXANHVINHPHUC": ("HTX_XVP", "Khối KD Vận tải Taxi Xanh", "hanno"),
    # HO: file "Báo cáo tuổi nợ" thật đã có (2026-08-08) nhưng KHÔNG có dòng khách hàng nào (toàn bộ
    # cột Mã khách/Tổng nợ phải thu rỗng ở mọi kỳ 202601-202606, verify trực tiếp trên file) -> để
    # nguyên 'age' (ra 0 y hệt 'hanno' vì không có dữ liệu, không đổi gì để giảm rủi ro không cần
    # thiết); đổi khi có dữ liệu chi tiết khách hàng thật.
    "HO": ("TC", "Khối hỗ trợ tập đoàn", "age"),
    "TRAMSAC": ("TC", "Khối KD Trạm sạc Vgreen", "hanno"),
    "GLOBALAI": ("GA", "Khối KD Công nghệ", "hanno"),
    # SRVF: verify 2026-08-03 qua raw_rows thật (PTHU/CDPS) trên DB coding — cong_ty='TC',
    # khoi='Khối KD Vinfast - Showroom' (folder SRVF cũng có dữ liệu cong_ty='VFQN' ở nguồn khác,
    # nhưng file baocaotuoino/*.xlsx hiện có TẤT CẢ đều tên "B.1.TC...." -> chỉ company TC).
    "SRVF": ("TC", "Khối KD Vinfast - Showroom", "hanno"),
    # An Taxi/An Khách sạn: verify 2026-08-04 qua raw_rows thật trên DB coding — CÙNG cong_ty='AAG'
    # nhưng khác khoi (giống HO/TRẠM SẠC cùng 'TC') -> khoi lấy tường minh ở đây. File "Báo cáo tuổi
    # nợ" của 2 đơn vị này có "Ngày đến hạn" thật + cột trong-hạn/đến-hạn/quá-hạn (giống SRVF) ->
    # mode 'hanno'. Cột lệch chữ so với SRVF (thêm cột NỢ/CÓ chen giữa J và L) nhưng _find_cols_hanno
    # dò theo TÊN header nên không cần sửa gì thêm, chỉ _find_sheet cần khớp thêm dạng tên sheet
    # "AN T{mm}.26" (xem _find_sheet). ANKHACHSAN CHƯA có file "Báo cáo tuổi nợ" nào (2026-08-04,
    # chỉ có báo cáo tài chính + tài sản cố định) — cấu hình sẵn, chưa verify được cong_ty/khoi qua
    # report_type PTHU_TUOINO (verify qua PTHU/CDPS khác thay, khớp raw_rows thật).
    "ANTAXI": ("AAG", "Khối KD Dịch vụ An Taxi", "hanno"),
    "ANKHACHSAN": ("AAG", "Khối KD Dịch vụ An KS", "hanno"),
    # ---- mode 'hanno_tong': ĐỌC THẲNG DÒNG TỔNG (spec user 2026-08-06, bản 2 — có file thật) ----
    # Phần tử thứ 4 = GỢI Ý TÊN SHEET (đã chuẩn hoá `_nd`): 2 file này KHÔNG có sheet tên "tuổi nợ"
    # nên `_find_sheet` không tự tìm ra, phải chỉ tên.
    #
    # XE TẢI HƯNG THỊNH — `HUNGTHINH/baocaotuoino/B5.HT.TCKT.M.<YYYYMM>.Baocaocongnoxetai.xlsx`,
    # sheet "Phải thu" (KHÔNG phải "Báo cáo tuổi nợ_Mẫu" như spec bản 1 đoán). Spec chỉ ĐÍCH DANH
    # ô ở DÒNG TỔNG (dòng 4, nằm NGAY TRÊN header dòng 5): H4 tổng · J4 trong hạn · K4 đến hạn ·
    # L4 quá hạn · M4/N4/O4/P4 = 4 dải. Cột I = 'CÓ' (khách ứng trước) -> KHÔNG thuộc aging.
    #
    # DỰ ÁN — `DUAN/baocaotuoino/B.4.TC.TCKT.M.<YYYYM>.Baocaotuoinophaithu.xlsx`, sheet "Tháng {m}".
    # Cùng kiểu nhưng cột lệch: D3 tổng · F3 trong hạn · G3 đến hạn · H3 quá hạn · I3/J3/K3/L3 dải
    # (E = 'Có'). Header dòng 4, 4 dải nằm NGAY TRÊN header chính chứ không ở sub-header như HT ->
    # `_find_cols_hanno` phải dò dải ở CẢ 2 dòng (xem hàm đó).
    # (cong_ty, khoi) verify 2026-08-06 trên DB coding 5435: HT|'Khối KD Xe tải', TC|'Khối KD Dự án'.
    "HUNGTHINH": ("HT", "Khối KD Xe tải", "hanno_tong", ("phai thu",)),
    "DUAN": ("TC", "Khối KD Dự án", "hanno_tong", ("thang",)),
    # XƯỞNG DỊCH VỤ VINFAST (spec user 2026-08-06) — `XDV/baocaotuoino/B9.TC.TCKT.M.<YYYYMM>.
    # TUOINOPHAITHU.xlsx`, sheet "Tuổi nợ ". Cũng đọc DÒNG TỔNG nằm TRÊN header như HT/Dự án, nhưng
    # bố cục cột KHÁC HẲN template -> mode riêng 'hanno_tong_xdv' (xem `_find_cols_xdv`).
    # (cong_ty, khoi) verify 2026-08-06 trên DB coding 5435: TC | 'Khối KD Vinfast - XDV'.
    "XDV": ("TC", "Khối KD Vinfast - XDV", "hanno_tong_xdv", ("tuoi no",)),
}


def _nd(s):
    s = str(s or "").strip().lower().replace("đ", "d")
    return "".join(ch for ch in unicodedata.normalize("NFD", s)
                   if unicodedata.category(ch) != "Mn")


def _num(v):
    return float(v) if isinstance(v, (int, float)) else None


def _source_id(path):
    """folder::basename — DÙNG THẲNG source_catalog.source_id_from_path() (agent_cli._source_id
    cũng gọi hàm này) thay vì tự đoán tên thư mục con. BUG đã gặp (2026-08-03): bản tự viết trước
    loại trừ cứng tên thư mục con 'BAOCAOTUOINOPHAITHU' — HTX_XTQ/HTX_XVP đổi thư mục con thành
    'baocaotuoino' (rút gọn) khi nạp thêm 6 tháng, làm _source_id cũ lấy NHẦM 'baocaotuoino' làm
    folder-đơn-vị thay vì 'HTXXANHTUYENQUANG'/'HTXXANHVINHPHUC'. Hàm thật lấy đúng parts[0] sau
    RECEIVED_DIR (không quan tâm tên/số cấp thư mục con), không lặp lại lỗi này."""
    folder = (_SC.raw_company_from_path(path) or "").upper()
    return f"{folder}::{os.path.basename(path)}"


def _find_sheet(wb, period=None, hints=None):
    """`hints` (2026-08-06): gợi ý tên sheet khai ở `_UNITS` cho đơn vị mà tên sheet KHÔNG chứa
    'tuổi nợ' và cũng không theo dạng 'T{mm}' — HT dùng sheet "Phải thu", Dự án dùng "Tháng {m}".
    Ưu tiên khớp '<hint> {mm}' (đúng THÁNG của kỳ, tránh vớ nhầm 'Tháng 6' khi đang xử lý kỳ 07 nếu
    file có nhiều sheet tháng — cùng bẫy `_sheet_thang` ở derive_tscd_hetkhauhao), sau đó mới khớp
    mở đầu tên. Hết hints thì rơi về logic cũ y nguyên (không đổi hành vi đơn vị đang chạy).

    Tên sheet KHÔNG cố định — file kỳ 06/2026 đặt tên đầy đủ 'Báo cáo tuổi nợ_Mẫu(...)', các
    kỳ trước (T01-T05, phát hiện 2026-08-03 khi HTX_XTQ/HTX_XVP có thêm 6 tháng) chỉ đặt 'T1'..'T6'
    (số trùng THÁNG); An Taxi (phát hiện 2026-08-04) đặt kiểu 'AN T6.26' (có tiền tố + hậu tố năm).
    Ưu tiên tên mô tả; fallback khớp token 't{mm}' ở BẤT KỲ ĐÂU trong tên (biên không phải chữ số,
    tránh 'T1' khớp nhầm 'T10'/'T11') thay vì đòi tên sheet == 'T{mm}' tuyệt đối — không lấy 'T1'
    đầu tiên bừa (cùng bẫy đã gặp ở derive_tscd_hetkhauhao._sheet_thang với sheet 'Tháng N')."""
    if hints:
        mm = int(period[5:7]) if period else None
        for h in hints:
            sn = next((s for s in wb.sheetnames if _nd(s) == f"{h} {mm}"), None)
            if sn:
                return sn
        for h in hints:
            # khớp BẰNG tên hint: XDV kỳ 02 có CẢ 'Tuổi nợ ' và 'Tuổi nợ, ', kỳ 04 có 'Tuổi nợ..'
            # (bản nháp, cùng số liệu) -> startswith sẽ ra 2 ứng viên rồi trả None ở dưới; sheet
            # ĐÚNG luôn là cái tên sạch = hint.
            sn = next((s for s in wb.sheetnames if _nd(s) == h), None)
            if sn:
                return sn
        for h in hints:
            cand = [s for s in wb.sheetnames if _nd(s).startswith(h)]
            if len(cand) == 1:
                return cand[0]
            if cand:
                return None      # NHIỀU sheet khớp mà không cái nào đúng tháng -> KHÔNG đoán bừa
        return None
    sn = next((s for s in wb.sheetnames if "tuoi no" in _nd(s) or "tuoi_no" in _nd(s)), None)
    if sn or not period:
        return sn
    mm = int(period[5:7])
    pat = re.compile(rf"(?<![0-9])t0?{mm}(?![0-9])")
    return next((s for s in wb.sheetnames if pat.search(_nd(s))), None)


def _period_end(period):
    """'YYYY-MM' -> ngày cuối tháng (Ngày báo cáo)."""
    y, m = int(period[:4]), int(period[5:7])
    return dt.date(y, m, calendar.monthrange(y, m)[1])


def _is_tong_hdr(s):
    """Header cột TỔNG dư nợ. Template chuẩn (SRVF/XVP/HTX/HO/An Taxi) + Dự án ghi 'Tổng nợ phải
    thu'; spec 2026-08-06 gọi chỉ tiêu này là 'Tổng số dư nợ'; file thật của HT (sheet "Phải thu")
    ghi gọn 'Tổng nợ' -> nhận cả 3 cách gọi (dò theo TÊN, không cứng chữ cột: mỗi đơn vị một vị trí).
    'tong no' để KHỚP BẰNG chứ không startswith — nếu không sẽ vớ nhầm 'Tổng nợ quá hạn'."""
    return s == "tong no" or s.startswith("tong no phai thu") or s.startswith("tong so du no")


def _find_cols(rows):
    """Dò header theo TÊN — chỉ cần 'Ngày hóa đơn' (ngày phát sinh) + 'Tổng nợ phải thu' (dư PT).
    Trả (data_start, ngay_i, tong_i) hoặc (None, None, None)."""
    hdr_i = next((i for i, r in enumerate(rows[:6])
                  if any(_is_tong_hdr(_nd(c)) for c in r if c)), None)
    if hdr_i is None:
        return None, None, None
    h = rows[hdr_i]
    ngay_i = next((j for j, c in enumerate(h) if c and _nd(c).startswith("ngay hoa don")), None)
    tong_i = next((j for j, c in enumerate(h) if c and _is_tong_hdr(_nd(c))), None)
    ma_i = next((j for j, c in enumerate(h) if c and _nd(c) == "ma khach"), None)
    if ngay_i is None or tong_i is None:
        return None, None, None
    return hdr_i + 2, ngay_i, tong_i, (ma_i if ma_i is not None else 4)


def _find_cols_hanno(rows):
    """Dò cột cho mode 'hanno' (SRVF) — header CHÍNH ở hàng chứa 'Tổng nợ phải thu' (giống
    _find_cols), nhưng 4 dải quá hạn theo ngày nằm ở SUB-HEADER hàng NGAY DƯỚI (cell 'Công nợ quá
    hạn' bị merge ngang N:Q ở hàng chính, tên dải thật '1-30 ngày'/'>30-90'/'>90-180'/'>180 ngày'
    chỉ xuất hiện ở hàng phụ) — verify bằng file thật SRVF/baocaotuoino T06/2026.
    Trả dict cột hoặc None nếu không đủ cột (fallback sang mode 'age' ở derive())."""
    hdr_i = next((i for i, r in enumerate(rows[:6])
                  if any(_is_tong_hdr(_nd(c)) for c in r if c)), None)
    if hdr_i is None or hdr_i + 1 >= len(rows):
        return None
    h, sub = rows[hdr_i], rows[hdr_i + 1]

    def _idx(row, pred):
        return next((j for j, c in enumerate(row) if c and pred(_nd(c))), None)

    def _idx2(pred):
        """4 dải quá hạn: SRVF/An Taxi/HT để ở SUB-HEADER, Dự án để THẲNG trên header chính
        (2026-08-06) -> tìm sub trước rồi mới tới hàng chính."""
        return _idx(sub, pred) if _idx(sub, pred) is not None else _idx(h, pred)

    cols = {
        "tong": _idx(h, _is_tong_hdr),
        "trong_han": _idx(h, lambda s: s.startswith("cong no trong han")),
        "den_han": _idx(h, lambda s: s.startswith("cong no den han")),
        "qua_han": _idx(h, lambda s: s.startswith("cong no qua han")),
        "qh_1_30": _idx2(lambda s: s.startswith("1-30")),
        "qh_30_90": _idx2(lambda s: "30-90" in s),
        "qh_90_180": _idx2(lambda s: "90-180" in s),
        "qh_180p": _idx2(lambda s: s.startswith("180") or s.startswith(">180")),
    }
    if any(v is None for v in cols.values()):
        return None
    ma_i = _idx(h, lambda s: s == "ma khach")
    cols["ma"] = ma_i if ma_i is not None else 4
    cols["hdr_i"] = hdr_i
    cols["data_start"] = hdr_i + 2
    # 'Ngày hóa đơn' (2026-08-08): CÓ trên MỌI file mode 'hanno' đã khảo sát (SRVF/An Taxi/An KS/
    # Trạm sạc/XVP/HTX_XTQ/HTX_XVP) — cho phép _agg_hanno TÍNH THÊM age-bucket (tuoi_no_1t/aging_13/
    # aging_36/aging_6p, field "aging" đọc ở debt.py) CÙNG LÚC với field hạn-nợ, để đổi mode 1 đơn vị
    # từ 'age' sang 'hanno' KHÔNG làm mất số liệu ở màn "Phân loại phải thu theo kỳ hạn thu nợ" đang
    # có sẵn. Optional (không có -> None, _agg_hanno chỉ ghi field hạn-nợ như cũ, không lỗi).
    cols["ngay"] = _idx(h, lambda s: s.startswith("ngay hoa don"))
    return cols


# XDV: file chia 5 DẢI quá hạn (1-30 · 30-60 · 60-90 · 90-180 · >180) còn dashboard chỉ có 4 dải
# -> dải ">30-90" = 30-60 + 60-90 (CỘNG 2 cột, đúng ý spec user "Công nợ quá hạn >30-90 ngày").
# Mỗi dải nhận NHIỀU cách bố trí: mỗi phần tử của tuple ngoài là MỘT phương án, phương án là
# tuple các tên cột phải CỘNG lại. Dò lần lượt, lấy phương án khớp đầu tiên.
# Vì sao cần nhiều phương án: file XDV đổi bố cục theo tháng — T01/02/03/05/06/07 có MỘT cột
# "Quá hạn 30-90 ngày" (đúng mapping), riêng T04 tách thành "30-60" + "60-90". Bản cũ chỉ khai
# phương án của T04 nên 6 tháng còn lại rớt hết ở bước dò cột (đó chính là lỗi "không dò được cột").
_XDV_DAI = {
    "qh_1_30": (("qua han 1-30",),),
    "qh_30_90": (("qua han 30-90",), ("qua han 30-60", "qua han 60-90")),
    "qh_90_180": (("qua han 90-180",),),
    "qh_180p": (("qua han >180",),),
}


def _find_cols_xdv(rows):
    """Dò cột cho mode 'hanno_tong_xdv' (XƯỞNG DỊCH VỤ VINFAST), sheet "Tuổi nợ ".

    MAPPING (bản kế toán gửi 2026-08-08) — đây là CHUẨN cần bám:
      D Tổng số dư nợ · E Công nợ trong hạn · F Công nợ đến hạn ·
      G quá hạn 1-30 · H quá hạn 30-90 · I quá hạn 90-180 · J quá hạn >180
      Công nợ quá hạn = G+H+I+J   ·   Tỷ lệ nợ quá hạn = quá hạn / tổng dư nợ

    THỰC TẾ FILE (khảo sát lại cả 7 kỳ 202601..202607, 2026-08-08) — 4/7 kỳ đúng y mapping,
    3 kỳ lệch, nên KHÔNG đọc theo chữ cột mà dò theo TÊN header rồi quy về schema của mapping:
      T01,02,03,05 : D..J đúng mapping.
      T06          : đúng VỊ TRÍ mapping nhưng cột F ghi nhãn "Công nợ trong hạn" (gõ nhầm,
                     trùng tên cột E) -> nhận theo VỊ TRÍ, nhãn thật lưu ở `den_han_theo_vi_tri`.
      T04          : KHÔNG có cột "đến hạn", và tách "30-90" thành "30-60" + "60-90" (5 dải)
                     -> cộng 2 cột đó lại thành qh_30_90 để cùng schema với 6 kỳ kia.
      T07          : thêm cột "Tên đối tượng" (mọi cột dịch phải 1) và header xuống DÒNG 3.

    BẢN TRƯỚC (2026-08-06) SAI Ở ĐÂU: nó được viết bám theo T04 — mà T04 lại là kỳ lệch chuẩn
    nhất — nên khai cứng "phải có cả 30-60 lẫn 60-90" và gán `den_han = None`. Hệ quả: chạy được
    đúng kỳ sai chuẩn, còn 4 kỳ ĐANG ĐÚNG mapping thì rớt ở bước dò cột ("không dò được cột").

    Hai chỗ mapping ghi không khớp file, đã xử lý mà không sửa mapping:
      - Mapping ghi ô ở DÒNG 2; thực tế dòng tổng ở dòng 1 (T01-06) hoặc dòng 2 (T07)
        -> `_agg_hanno_tong` quét ngược lên trên header, không hardcode số dòng.
      - Mapping không có cột tổng "Công nợ quá hạn" trong file -> `qua_han=None`, cộng 4 dải
        đúng như công thức mapping.
    Tên cột trong file có typo "Trog hạn" (thiếu 'n') -> nhận cả 2 cách viết.
    Trả dict cột (giá trị có thể là tuple nhiều cột phải CỘNG, hoặc None nếu file không có cột đó)."""
    hdr_i = next((i for i, r in enumerate(rows[:8])
                  if any(_nd(c).startswith("tong cn") for c in r if c)), None)
    if hdr_i is None:
        return None
    h = rows[hdr_i]

    def _idx(pre):
        return next((j for j, c in enumerate(h) if c and _nd(c).startswith(pre)), None)

    cols = {"tong": _idx("tong cn"),
            # "Trog hạn" là typo có thật trong file nguồn (thiếu 'n') -> nhận cả 2 cách viết.
            "trong_han": next((j for j in (_idx("trog han"), _idx("trong han")) if j is not None), None),
            "qua_han": None}     # file không có cột tổng quá hạn -> cộng 4 dải (đúng mapping)
    if cols["tong"] is None or cols["trong_han"] is None:
        return None
    for k, phuong_an in _XDV_DAI.items():
        chon = None
        for names in phuong_an:
            idxs = tuple(j for j in (_idx(n) for n in names) if j is not None)
            if len(idxs) == len(names):
                chon = idxs if len(idxs) > 1 else idxs[0]
                break
        if chon is None:
            return None          # thiếu dải -> coi như đổi layout, KHÔNG ghi số lệch
        cols[k] = chon

    # "Công nợ đến hạn" — mapping đặt nó ở cột NGAY GIỮA "Trong hạn" và "Quá hạn 1-30 ngày".
    # Dò theo TÊN trước; không thấy thì dùng đúng quy tắc VỊ TRÍ của mapping. Cần bước vị trí vì
    # T06 ghi nhãn cột này là "Công nợ trong hạn" (gõ nhầm — trùng tên cột bên cạnh), còn T04/T07
    # thì KHÔNG có cột này thật (giữa hai cột kia không còn khe nào) -> để None, không bịa số 0.
    j1, j2 = cols["trong_han"], (cols["qh_1_30"][0] if isinstance(cols["qh_1_30"], tuple) else cols["qh_1_30"])
    dh = _idx("cong no den han")
    if dh is None and j2 - j1 == 2:
        dh = j1 + 1
        cols["den_han_theo_vi_tri"] = str(h[dh] or "").strip()   # ghi lại nhãn thật để đối chiếu
    cols["den_han"] = dh
    cols["hdr_i"] = hdr_i
    cols["data_start"] = hdr_i + 1   # KHÔNG có sub-header như template chuẩn
    return cols


def _twin_attrs(cur, period):
    """dataset_id/ngay từ 1 dòng BẤT KỲ đã nạp trong CÙNG period_month (dataset/ngay là cấp ĐỘ KỲ,
    dùng chung mọi công ty trong tháng đó — KHÔNG cần lọc theo cong_ty/khoi). `khoi` LẤY THẲNG từ
    `_UNITS` (không suy qua twin): nhiều đơn vị CÙNG cong_ty khác khoi (vd HO/TRẠM SẠC đều 'TC') ->
    twin theo cong_ty sẽ vớ nhầm khoi của đơn vị khác cùng mã công ty.
    `ngay` BẮT BUỘC không None: report_type nằm trong `_SNAP_RT` (metrics/_shared.py) nên `_rows()`
    chọn snapshot theo MAX(ngay); ngay=None -> _rows() trả [] câm (không lỗi)."""
    cur.execute(
        "SELECT dataset_id, ngay FROM raw_rows WHERE period_month=%s "
        "AND ngay IS NOT NULL ORDER BY (report_type='PTHU') DESC LIMIT 1",
        (period,))
    return cur.fetchone()


def _agg_age(rows, report_date):
    """mode 'age' (XVP/HTX/HO/TRẠM SẠC/GLOBAL AI) — bucket theo tuổi hoá đơn, xem docstring đầu file."""
    data_start, ngay_i, tong_i, ma_i = _find_cols(rows)
    if data_start is None:
        return None
    agg = {"b1": 0.0, "b13": 0.0, "b36": 0.0, "b6p": 0.0}
    n_rows = 0
    for r in rows[data_start:]:
        if not r or ma_i >= len(r) or not r[ma_i]:
            continue
        ngay = r[ngay_i] if ngay_i < len(r) else None
        tong = _num(r[tong_i]) if tong_i < len(r) else None
        if not isinstance(ngay, (dt.date, dt.datetime)) or tong is None:
            continue
        n_rows += 1
        age_days = (report_date - (ngay.date() if isinstance(ngay, dt.datetime) else ngay)).days
        bucket = "b1" if age_days < 30 else "b13" if age_days < 90 else "b36" if age_days < 180 else "b6p"
        agg[bucket] += tong
    tong_all = round(sum(agg.values()) * 1e-9, 9)
    payload = {"tuoi_no_1t": round(agg["b1"] * 1e-9, 9), "aging_13": round(agg["b13"] * 1e-9, 9),
               "aging_36": round(agg["b36"] * 1e-9, 9), "aging_6p": round(agg["b6p"] * 1e-9, 9),
               "unit": "ty"}
    return {"n_rows": n_rows, "tong_all": tong_all,
            "agg_ty": {k: round(v * 1e-9, 9) for k, v in agg.items()}, "payload": payload}


def _agg_hanno(rows, report_date=None):
    """mode 'hanno' (SRVF/An Taxi/An KS + Trạm sạc/XVP/HTX_XTQ/HTX_XVP/Global AI từ 2026-08-08) —
    CỘNG THẲNG các cột trong-hạn/đến-hạn/quá-hạn (J..Q) đã có sẵn theo khách hàng trong sheet.

    Kèm age-bucket (2026-08-08): file mode 'hanno' NÀO CŨNG có cột 'Ngày hóa đơn' (verify SRVF/An
    Taxi/Trạm sạc/XVP/HTX_XTQ/HTX_XVP) nên tính LUÔN tuoi_no_1t/aging_13/aging_36/aging_6p per dòng
    (giống hệt `_agg_age`, dùng CHUNG report_date) và gộp vào CÙNG payload — để field "aging" ở
    debt.py (màn "Phân loại phải thu theo kỳ hạn thu nợ") không bị mất số khi 1 đơn vị đổi mode từ
    'age' sang 'hanno'. `report_date=None` (gọi cũ, nếu có) -> bỏ qua age-bucket, giữ hành vi cũ."""
    cols = _find_cols_hanno(rows)
    if cols is None:
        return None
    keys = ("tong", "trong_han", "den_han", "qua_han", "qh_1_30", "qh_30_90", "qh_90_180", "qh_180p")
    agg = {k: 0.0 for k in keys}
    age_agg = {"b1": 0.0, "b13": 0.0, "b36": 0.0, "b6p": 0.0}
    ngay_i = cols.get("ngay")
    n_rows = 0
    for r in rows[cols["data_start"]:]:
        if not r or cols["ma"] >= len(r) or not r[cols["ma"]]:
            continue
        vals = {k: _num(r[cols[k]]) if cols[k] < len(r) else None for k in keys}
        if vals["tong"] is None:
            continue
        n_rows += 1
        for k in keys:
            agg[k] += vals[k] or 0.0
        if report_date is not None and ngay_i is not None:
            ngay = r[ngay_i] if ngay_i < len(r) else None
            if isinstance(ngay, (dt.date, dt.datetime)):
                age_days = (report_date - (ngay.date() if isinstance(ngay, dt.datetime) else ngay)).days
                bucket = "b1" if age_days < 30 else "b13" if age_days < 90 else "b36" if age_days < 180 else "b6p"
                age_agg[bucket] += vals["tong"]
    tong_all = round(agg["tong"] * 1e-9, 9)
    payload = {"tong_no": round(agg["tong"] * 1e-9, 9), "trong_han": round(agg["trong_han"] * 1e-9, 9),
               "den_han": round(agg["den_han"] * 1e-9, 9), "qua_han": round(agg["qua_han"] * 1e-9, 9),
               "qh_1_30": round(agg["qh_1_30"] * 1e-9, 9), "qh_30_90": round(agg["qh_30_90"] * 1e-9, 9),
               "qh_90_180": round(agg["qh_90_180"] * 1e-9, 9), "qh_180p": round(agg["qh_180p"] * 1e-9, 9),
               "unit": "ty"}
    if report_date is not None and ngay_i is not None:
        payload.update({"tuoi_no_1t": round(age_agg["b1"] * 1e-9, 9), "aging_13": round(age_agg["b13"] * 1e-9, 9),
                         "aging_36": round(age_agg["b36"] * 1e-9, 9), "aging_6p": round(age_agg["b6p"] * 1e-9, 9)})
    return {"n_rows": n_rows, "tong_all": tong_all, "agg_ty": {k: round(v * 1e-9, 9) for k, v in agg.items()},
            "payload": payload}


_HANNO_KEYS = ("tong", "trong_han", "den_han", "qua_han",
               "qh_1_30", "qh_30_90", "qh_90_180", "qh_180p")


def _hanno_payload(agg):
    p = {("tong_no" if k == "tong" else k): round(agg[k] * 1e-9, 9) for k in _HANNO_KEYS}
    p["unit"] = "ty"
    return p


def _cell(row, spec):
    """Giá trị 1 chỉ tiêu ở `row` theo `spec` của dict cột:
      int   -> 1 cột (mọi đơn vị mode 'hanno_tong')
      tuple -> CỘNG nhiều cột (XDV: dải '>30-90' = '30-60' + '60-90')
      None  -> cột KHÔNG tồn tại trong file -> 0 (XDV: 'Công nợ đến hạn')."""
    if spec is None:
        return 0.0
    idxs = spec if isinstance(spec, tuple) else (spec,)
    return sum((_num(row[j]) or 0.0) if j < len(row) else 0.0 for j in idxs)


def _has_num(row, spec):
    """Dòng có SỐ ở cột `spec` (dùng để tìm dòng tổng / lọc dòng chi tiết)."""
    idxs = (spec,) if isinstance(spec, int) else (spec or ())
    return any(j < len(row) and _num(row[j]) is not None for j in idxs)


def _agg_hanno_tong(rows, cols=None):
    """mode 'hanno_tong' (HƯNG THỊNH / DỰ ÁN, spec user 2026-08-06) và 'hanno_tong_xdv' (XƯỞNG DỊCH
    VỤ VINFAST — truyền `cols` từ `_find_cols_xdv`) — ĐỌC THẲNG DÒNG TỔNG thay vì cộng từng khách
    hàng như `_agg_hanno`.

    Spec chỉ đích danh ô ở dòng tổng (HT: H4/J4/K4/L4/M4-P4 · Dự án: D3/F3/G3/H3/I3-L3) — dòng đó
    nằm NGAY TRÊN header, nên KHÔNG dò được bằng `data_start` (vốn tính xuôi xuống dưới header).
    Cách dò: sau khi có cột từ `_find_cols_hanno`, quét NGƯỢC các dòng TRƯỚC header, lấy dòng đầu
    tiên có số ở cột `tong`. Không hardcode số dòng/chữ cột — file đổi bố cục nhẹ vẫn chạy.

    Vẫn cộng luôn phần chi tiết để ĐỐI CHIẾU (`sum_chi_tiet_ty` trong dry-run): 2026-08-06 khớp
    tuyệt đối 8/8 cột ở cả 2 file. Nếu sau này LỆCH -> vẫn LẤY DÒNG TỔNG (đúng spec) nhưng cờ
    `lech_tong_vs_chi_tiet` bật lên để biết mà đi hỏi kế toán, không âm thầm ra số khác."""
    cols = cols if cols is not None else _find_cols_hanno(rows)
    if cols is None:
        return None
    hdr_i = cols["hdr_i"]
    tot_i = next((i for i in range(hdr_i - 1, -1, -1) if _has_num(rows[i], cols["tong"])), None)
    if tot_i is None:
        return None
    tr = rows[tot_i]
    agg = {k: _cell(tr, cols[k]) for k in _HANNO_KEYS}

    chi_tiet = {k: 0.0 for k in _HANNO_KEYS}
    n_rows = 0
    for r in rows[cols["data_start"]:]:
        if not r or not _has_num(r, cols["tong"]):
            continue
        n_rows += 1
        for k in _HANNO_KEYS:
            chi_tiet[k] += _cell(r, cols[k])
    # File không có cột tổng "Công nợ quá hạn" (XDV) -> = Σ 4 dải. Suy CẢ ở dòng tổng và ở phần chi
    # tiết để cờ `lech_tong_vs_chi_tiet` vẫn so đúng cùng cách tính.
    if cols["qua_han"] is None:
        for d in (agg, chi_tiet):
            d["qua_han"] = d["qh_1_30"] + d["qh_30_90"] + d["qh_90_180"] + d["qh_180p"]
    lech = any(abs(agg[k] - chi_tiet[k]) > 1 for k in _HANNO_KEYS)

    return {"n_rows": n_rows, "tong_all": round(agg["tong"] * 1e-9, 9),
            "dong_tong": tot_i + 1,          # số dòng Excel (1-based) đã lấy — đối chiếu với spec
            "agg_ty": {k: round(v * 1e-9, 9) for k, v in agg.items()},
            "sum_chi_tiet_ty": {k: round(v * 1e-9, 9) for k, v in chi_tiet.items()},
            "lech_tong_vs_chi_tiet": lech, "payload": _hanno_payload(agg)}


def derive(path, period, write=False):
    folder = _source_id(path).split("::", 1)[0]
    unit = _UNITS.get(folder)
    if not unit:
        return {"ok": False, "skip": True}
    cong_ty, khoi, mode = unit[:3]
    sheet_hints = unit[3] if len(unit) > 3 else None
    # agent_cli gọi deriver này cho MỌI file của folder (không chỉ file tuổi nợ) -> phải phân biệt
    # "file tuổi nợ mà hỏng" (báo error) với "file loại khác" (skip câm). Mọi file tuổi nợ hiện có
    # đều nằm ở thư mục con/tên file chứa 'tuoino' (baocaotuoino/baocaotuoinophaithu/TUOINOPHAITHU).
    # Với file KHÁC loại, CHỈ nhận sheet có tên mô tả 'tuổi nợ' — KHÔNG cho fallback 'T{mm}' của
    # _find_sheet chạy: nếu bật, file B5.HT.TCTC.M.202605 (Hưng Thịnh) có sheet "chi tiết xe tc xuất
    # ht5 tháng" sẽ khớp NHẦM 't5' rồi báo lỗi cột rác mỗi lượt nạp báo cáo tài chính.
    is_tuoino_file = "tuoino" in _nd(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        sn = _find_sheet(wb, period if is_tuoino_file else None,
                         sheet_hints if is_tuoino_file else None)
        if not sn:
            # Đơn vị có `sheet_hints` = tên sheet XÁC ĐỊNH -> file trong thư mục tuoino mà không có
            # sheet đó là BÁO CÁO KHÁC, skip CÂM (không báo lỗi mỗi lượt nạp). Cụ thể
            # XDV/baocaotuoino còn chứa B.2.TC.TCKT.[DM].20260500.Baocaocongnophaithu*.xlsx (đối soát
            # công nợ VF, 10-50 sheet, không có sheet 'Tuổi nợ') — cùng bệnh đã fix cho HT ở bản 4.
            # Layout đổi THẬT (sheet còn đó, cột lệch) vẫn báo error ở bước _agg_* bên dưới.
            if sheet_hints:
                return {"ok": False, "skip": True}
            return ({"ok": False, "error": "không thấy sheet 'Báo cáo tuổi nợ'"}
                    if is_tuoino_file else {"ok": False, "skip": True})
        rows = [list(r) for r in wb[sn].iter_rows(values_only=True)]
    finally:
        wb.close()

    result = (_agg_hanno_tong(rows, _find_cols_xdv(rows)) if mode == "hanno_tong_xdv"
              else _agg_hanno_tong(rows) if mode == "hanno_tong"
              else _agg_hanno(rows, _period_end(period)) if mode == "hanno"
              else _agg_age(rows, _period_end(period)))
    if result is None:
        err = ("không dò được cột 'Tổng cn'/'Trog hạn'/5 dải quá hạn, hoặc không thấy DÒNG TỔNG phía "
               "trên header" if mode == "hanno_tong_xdv"
               else "không dò được cột trong-hạn/đến-hạn/quá-hạn, hoặc không thấy DÒNG TỔNG phía trên "
               "header" if mode == "hanno_tong"
               else "không dò được cột trong-hạn/đến-hạn/quá-hạn (J..Q)" if mode == "hanno"
               else "không dò được cột 'Ngày hóa đơn'/'Tổng nợ phải thu'")
        return {"ok": False, "error": err}
    n_rows, tong_all, payload = result["n_rows"], result["tong_all"], result["payload"]

    out = {"file": os.path.basename(path), "period": period, "cong_ty": cong_ty, "mode": mode,
           "sheet": sn, "n_rows": n_rows, "tong_ty": tong_all, "agg_ty": result["agg_ty"]}
    for k in ("dong_tong", "sum_chi_tiet_ty", "lech_tong_vs_chi_tiet"):
        if k in result:
            out[k] = result[k]

    if write:
        source_file = _source_id(path)
        conn = psycopg.connect(DB_URL)
        try:
            cur = conn.cursor()
            attrs = _twin_attrs(cur, period)
            if not attrs:
                out["error"] = f"không thấy dòng nào có ngay (period={period}) để suy dataset_id/ngay -> BỎ ghi"
                return out
            dataset_id, ngay = attrs
            # idempotent: xoá bản cũ CÙNG source_file (KHÔNG cong_ty+period — vài đơn vị dùng CHUNG
            # cong_ty như HO/TRẠM SẠC đều 'TC', xoá theo cong_ty sẽ xoá NHẦM sang đơn vị khác cùng mã)
            cur.execute("DELETE FROM raw_rows WHERE report_type='PTHU_TUOINO' AND source_file=%s",
                        (source_file,))
            cur.execute(
                "INSERT INTO raw_rows (dataset_id, report_type, row_index, ngay, cong_ty, khoi, "
                "cost_center, period_month, amount, amount2, dim1, dim2, dim3, payload, source_file) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (dataset_id, "PTHU_TUOINO", 6100000, ngay, cong_ty, khoi, None, period,
                 tong_all, None, cong_ty, None, None,
                 json.dumps(payload, ensure_ascii=False), source_file))
            conn.commit()
            out["written"] = 1
        finally:
            conn.close()
    return out


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--period", required=True)
    ap.add_argument("--write", action="store_true", help="ghi DB (mặc định dry-run)")
    a = ap.parse_args()
    print(json.dumps(derive(a.file, a.period, a.write), ensure_ascii=False, indent=2))
