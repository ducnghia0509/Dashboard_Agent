# -*- coding: utf-8 -*-
"""Kế hoạch doanh thu NĂM theo KHỐI -> cột `amount2` của report_type DTHU (spec bảng 50 chỉ tiêu,
dòng 1 "Kế hoạch doanh thu", màn "Doanh thu – Giá vốn").

NGUỒN: A:\\0.BANLANHDAO\\Kehoachnam2026\\0.KH.GR.Y.2026.Kehoachdoanhthu.xlsx, sheet "KHDT".
Bố cục: cột B "Mã Khối" (1..10), cột C "Khối", cột D "Tổng năm", E/F 6 tháng đầu/cuối, G..R
"Tháng 1".."Tháng 12". Dòng cuối "Tổng" (Mã Khối rỗng) -> BỎ (nếu cộng vào sẽ đếm đôi).

VÌ SAO map theo MÃ KHỐI chứ không theo TÊN: mã 6 trong file ghi "Khối KD Vận tải Taxi XVP" còn
master (`master_data.khoi`) ghi "Khối KD Vận tải Taxi Xanh" — khớp tên sẽ rớt đúng khối này. Cả
10 mã 1..10 của file trùng khít mã master nên map theo mã là tất định.

VÌ SAO ghi vào `amount2` của DTHU (không tạo report_type riêng): toàn bộ đường ống kế hoạch ĐÃ có
sẵn — `revenue.build_revenue` trả `plan = flow_sum(ds,"DTHU",...,col="amount2")`, FE Revenue.tsx
đã có ô "TỶ LỆ HOÀN THÀNH KH" đọc `d.plan`; `expense.build_expense` cũng lấy `rev_plan` từ đây để
suy ngân sách chi phí. Chỉ thiếu dữ liệu.

CÁC DÒNG NÀY KHÔNG LÀM SAI SỐ THỰC HIỆN:
  - `amount` = NULL -> mọi tổng thực hiện (flow_sum/flow_by_khoi/top_by_costcenter dùng col
    "amount") cộng vào 0; `_per_file_resolved` gom theo source_file nên nhóm kế hoạch đứng riêng,
    không chen vào logic "dòng trực tiếp vs breakdown cost_center" của file thật.
  - `dim1` = NULL -> `sum_by_dim1` (nuôi biểu đồ "byType" của màn Doanh thu) có điều kiện
    `AND dim1 IS NOT NULL` nên loại hẳn dòng kế hoạch, không đẻ ra hạng mục lạ.

Mỗi tháng ghi vào ĐÚNG dataset kind='month' của kỳ đó. Tháng chưa có dataset -> bỏ qua và báo
trong `skipped` (KHÔNG tự tạo dataset rỗng). Từ 03/09/2026 quy tắc đó nằm ở
`servers/common/dataset_ky.py` dưới dạng cờ `tao=False`, dùng chung với hai đường ghi kia
(báo cáo ngày và spec_extract) — hai đường ấy ghi SỐ THỰC TẾ nên ĐƯỢC khai sinh kỳ, còn kế
hoạch thì không: file này phủ trọn 12 kỳ, cho tạo là mọc ra các kỳ chưa tới.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.normpath(os.path.join(_HERE, ".."))
sys.path.insert(0, _ROOT)

from dotenv import load_dotenv  # noqa: E402

from servers.common import dataset_ky as _DSK  # noqa: E402

load_dotenv(os.path.join(_ROOT, ".env"))

DB_URL = os.environ.get("DATABASE_URL")
SHEET = "KHDT"
REPORT_TYPE = "DTHU"
ROW_INDEX_BASE = 6300000          # dải riêng cho dòng kế hoạch (ngày dùng 6200000, xem derive_hqkd_ngay)


def _nd(s):
    import unicodedata
    s = str(s if s is not None else "").strip().lower().replace("đ", "d")
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def _khoi_by_ma():
    """{mã khối master: tên khối master} — nguồn duy nhất cho tên khối ghi vào raw_rows."""
    sys.path.insert(0, os.path.join(_ROOT, "..", "AI_coding", "tc-admin-api"))
    from app.master_data import loader as master
    return {str(k.get("ma") or "").strip(): (k.get("ten") or "").strip()
            for k in master.master_data().get("khoi", [])}


def _source_id(path):
    """'<THƯ MỤC NGUỒN>::<tên file>' — cùng quy ước với các deriver khác (dùng để idempotent)."""
    parts = os.path.normpath(path).split(os.sep)
    folder = parts[-3] if len(parts) >= 3 else "KEHOACH"
    return f"{folder}::{os.path.basename(path)}"


def read_plan(path):
    """-> {'2026-01': {tên khối: số tiền VND}, ...} + danh sách mã không map được."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        if SHEET not in wb.sheetnames:
            return None, [f"không thấy sheet '{SHEET}'"]
        rows = [list(r) for r in wb[SHEET].iter_rows(min_row=1, max_row=60, max_col=30,
                                                     values_only=True)]
    finally:
        wb.close()

    hdr_i = next((i for i, r in enumerate(rows) if any(_nd(c) == "ma khoi" for c in r)), None)
    if hdr_i is None:
        return None, ["không thấy dòng tiêu đề có cột 'Mã Khối'"]
    hdr = rows[hdr_i]
    ma_j = next(j for j, c in enumerate(hdr) if _nd(c) == "ma khoi")
    # Cột tháng dò theo TÊN ('Tháng 1'..'Tháng 12'), KHÔNG theo vị trí cố định — file có thêm
    # 'Tổng năm'/'6 tháng đầu năm'/'6 tháng cuối năm' xen giữa, vị trí có thể đổi giữa các bản.
    month_j = {}
    for j, c in enumerate(hdr):
        n = _nd(c)
        for mm in range(1, 13):
            if n == f"thang {mm}":
                month_j[mm] = j

    khoi_map = _khoi_by_ma()
    year = None
    for r in rows[:hdr_i + 1]:
        for c in r:
            import re
            m = re.search(r"20\d{2}", str(c or ""))
            if m:
                year = m.group(0)
                break
        if year:
            break
    year = year or os.environ.get("DASHBOARD_YEAR", "2026")

    plan, unknown = {}, []
    for r in rows[hdr_i + 1:]:
        ma = str(r[ma_j]).strip() if ma_j < len(r) and r[ma_j] is not None else ""
        if not ma or ma.lower() in ("none", "tong"):   # dòng 'Tổng' không có mã -> bỏ, tránh đếm đôi
            continue
        ma = ma.split(".")[0]                          # '1.0' (ô số) -> '1'
        ten = khoi_map.get(ma)
        if not ten:
            unknown.append(ma)
            continue
        for mm, j in month_j.items():
            v = r[j] if j < len(r) else None
            if isinstance(v, (int, float)) and v:
                plan.setdefault(f"{year}-{mm:02d}", {})[ten] = float(v)
    return plan, unknown


def derive(path, write=False):
    plan, problems = read_plan(path)
    if plan is None:
        return {"ok": False, "error": "; ".join(problems)}
    out = {"ok": True, "file": os.path.basename(path),
           "ky": {k: {"so_khoi": len(v), "tong_ty": round(sum(v.values()) * 1e-9, 9)}
                  for k, v in sorted(plan.items())}}
    if problems:
        out["ma_khoi_khong_map_duoc"] = sorted(set(problems))
    if not write:
        return out

    source_file = _source_id(path)
    conn = psycopg.connect(DB_URL)
    try:
        cur = conn.cursor()
        # idempotent: xoá bản kế hoạch cũ CÙNG source_file (1 file phủ trọn 12 tháng).
        cur.execute("DELETE FROM raw_rows WHERE report_type=%s AND source_file=%s",
                    (REPORT_TYPE, source_file))
        payload = json.dumps({"unit": "ty", "loai": "ke_hoach"}, ensure_ascii=False)
        recs, skipped, i = [], [], 0
        for period, by_khoi in sorted(plan.items()):
            # `tao=False` CỐ Ý: kế hoạch KHÔNG được khai sinh kỳ. File này ghi trọn 12 kỳ
            # trong một lần nạp, cho tạo là đẻ ngay các kỳ chưa tới, rỗng số thực tế, mà vẫn
            # nằm trong ô chọn kỳ. Xem servers/common/dataset_ky.py (nguyên tắc: chỉ SỐ THỰC
            # TẾ của kỳ ĐÃ TỚI mới khai sinh được kỳ).
            ds_ky, _tt = _DSK.lay_hoac_tao_ky(cur, period, nguon=source_file, tao=False)
            if not ds_ky:
                skipped.append(period)
                continue
            row = (ds_ky,)
            for ten, v in by_khoi.items():
                i += 1
                recs.append((row[0], REPORT_TYPE, ROW_INDEX_BASE + i, None, None, ten, None,
                             period, None, round(v * 1e-9, 9), None, None, None,
                             payload, source_file))
        if recs:
            cur.executemany(
                "INSERT INTO raw_rows (dataset_id, report_type, row_index, ngay, cong_ty, khoi, "
                "cost_center, period_month, amount, amount2, dim1, dim2, dim3, payload, source_file) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", recs)
        conn.commit()
        out["written"] = len(recs)
        if skipped:
            out["bo_qua_chua_co_dataset"] = skipped
    finally:
        conn.close()
    return out


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--write", action="store_true", help="ghi DB (mặc định dry-run)")
    a = ap.parse_args()
    print(json.dumps(derive(a.file, write=a.write), ensure_ascii=False, indent=2))
