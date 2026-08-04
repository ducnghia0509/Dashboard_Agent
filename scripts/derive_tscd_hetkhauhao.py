# -*- coding: utf-8 -*-
"""Deriver CHUYÊN BIỆT: "TS hết khấu hao còn dùng / chờ thanh lý" (Khối 12, TSCĐ Chart 3/3).

Trước đây KHÔNG có nguồn tự động (xem memory tscd-chart3-hetkhauhao-blocked) -> payload
08_TSCD.het_kh_con_sd/het_kh_thanh_ly luôn 0. Nay có spec cột/sheet chính xác theo đơn vị
(user cung cấp 2026-07-29, đã đối chiếu khớp file thật kỳ T06/2026):

  · Layout PHẲNG (header 1 tầng): An Khách sạn, GA, Xanh VP, Trạm sạc, Dự án, SRVF (.xlsb) —
    lọc DÒNG TÀI SẢN có "Giá trị còn lại" (cuối kỳ) = 0, cộng "Nguyên giá" (cuối kỳ).
  · Layout NHÓM (header 2 tầng 'Tài sản'/'Khấu hao'/'Giá trị sổ sách'): An Taxi, Hưng Thịnh,
    Trạm sạc (mapping v2 2026-07-31) — lọc GTSS = 0, cộng cột 'Tài sản' CUỐI KỲ (An Taxi/HT
    cột I, Trạm sạc cột L — đều dò động theo sub-header ngày cuối, không cứng chữ cột); dòng
    tài sản = có cột đặc tính (ngày mua lại/phương thức), dòng loại/tổng (vd "24213 Chi phí
    phân bổ CCDC") chỉ có cột A + số gộp -> bỏ. SRVF (mapping v2): sheet 'Bảng tính KH chi
    tiết', lọc 'Giá trị còn lại (Cyber Tx.26)' = 0, cộng 'Nguyên giá cuối năm'.
  · HO (spec cập nhật 2026-07-30): sheet 'Theo dõi KH tài sản HO' trong CHÍNH file B.9 —
    lọc GTCL=0, cộng "NG cuối kỳ", dò cột theo TÊN header vì VỊ TRÍ TRÔI THEO THÁNG (mỗi kỳ
    chèn thêm 1 cột KH tháng: GTCL ở AN kỳ 12/25 -> AX kỳ 6/26). Lấy CẢ TSCĐ + CCDC (spec không
    loại trừ "Phân Loại"). THAY proxy mã 223 BCĐKT cũ (~27.8 tỷ, sai bản chất — số thật ~100 tỷ);
    proxy đã GỠ khỏi _derive_tscd_cdkt (agent_cli.py 2026-07-30) để không ghi đôi.
  · Chỉ An Khách sạn có cột "Tình trạng sử dụng" tách được còn dùng/chờ thanh lý (đã kiểm tra
    các đơn vị khác không có cột trạng thái tương đương) -> đơn vị khác đổ hết vào "còn sử dụng".
  · Depot PT/VP/TQ, HTX Xanh VP, HTX Xanh Tuyên Quang, XDV: CHƯA có nguồn -> bỏ qua (no-op).

API 2 lớp (tránh 2 lần import_filled CÙNG source_file đè nhau — xem agent_cli.py dispatcher):
  · compute(path) -> (con_sd, thanh_ly) RAW VND hoặc None — KHÔNG ghi DB.
  · extract(path, period, cong_ty) -> compute() rồi TỰ GHI (source_file riêng, 08_TSCD CHỈ 2
    cột hết-KH). Chỉ an toàn cho đơn vị mà _derive_tscd/_derive_tscd_duan KHÔNG ghi cho CHÍNH
    source_file này (An Khách sạn/GA/Xanh VP/Trạm sạc/Dự án/SRVF). An Taxi/Hưng Thịnh (layout
    NHÓM, _derive_tscd ghi thành công) -> dispatcher tự gắn kết quả compute() vào record của
    _derive_tscd, KHÔNG gọi extract()."""
import os
import re
import sys

from openpyxl.utils import column_index_from_string as _ci

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from servers import template_filler as tf  # noqa: E402
from servers.common import be_bridge as bb  # noqa: E402
import agent_cli as A  # noqa: E402

norm = lambda v: bb.normalize_header(v, True) if v is not None else ""  # noqa: E731
_EPS = 1000.0   # 1.000đ — coi GTCL trong ngưỡng này là hết khấu hao (chống lệch làm tròn)


def _num(v):
    return v if isinstance(v, (int, float)) else None


def _numz(v):
    """Cột LỌC (GTCL/GTSS) — QUY ƯỚC 2026-07-31: rỗng/None/"-"/bất kỳ chuỗi không phải số nào
    đều = 0.0 (kế toán hay để trống hoặc ghi "-" khi tài sản đã hết khấu hao hoàn toàn, thay vì
    ghi số 0 tường minh). Có số thật (kể cả số 0) -> giữ nguyên. CHỈ áp dụng cho cột lọc dùng để
    xét hết-khấu-hao; cột Nguyên giá vẫn dùng _num() (rỗng = thiếu dữ liệu dòng đó -> bỏ dòng),
    không áp dụng quy ước này."""
    return float(v) if isinstance(v, (int, float)) else 0.0


def _is_zero(v):
    return v is not None and abs(v) < _EPS


# Layout PHẲNG — đối chiếu khớp file thật T06/2026 (xem docstring trên).
_FLAT = {
    "ANKHACHSAN": dict(sheet="danh sách sổ tài sản cố định", header_row=3, ng="J", gtcl="M", status="Q"),
    # GA: spec "lấy mã TSxx" -> chỉ dòng Mã tài sản (cột B) bắt đầu 'TS', bỏ CCDC. Sheet T1-T5
    # xuất tên 'Tháng 1'/'Tháng 5' (T2-T4 vẫn đề 'Tháng 1'!) thay vì 'Tài sản, CCDC' — cùng
    # layout cột -> fallback _sheet_thang (chỉ bật cho GA, xem compute()).
    "GLOBALAI": dict(sheet="tài sản, ccdc", header_row=2, ng="D", gtcl="J", ma="B"),
    "XANHVINHPHUC": dict(sheet="tài sản", header_row=6, ng="I", gtcl="AK"),
}
_DUAN = dict(header_row=6, ng="N", gtcl="P", loai="H", hm="O")
# TRAMSAC chuyển sang GROUPED (mapping v2 2026-07-31): file đổi sang layout nhóm 2 tầng như An
# Taxi/HT — spec lọc "Giá trị sổ sách"=0 (cột Q), cộng "Tài sản" CUỐI KỲ (cột L). Đã kiểm cả 6
# tháng T1-T6/2026 đều là layout mới (file flat cũ ng=I/gtcl=N không còn).
_GROUPED = {"ANTAXI", "HUNGTHINH", "TRAMSAC"}
_SKIP_NAMES = ("tai san", "tong", "cong")


def _flat_rows(ws, cfg):
    hdr_r = cfg["header_row"]
    ng_i, gtcl_i = _ci(cfg["ng"]) - 1, _ci(cfg["gtcl"]) - 1
    status_i = _ci(cfg["status"]) - 1 if cfg.get("status") else None
    ma_i = _ci(cfg["ma"]) - 1 if cfg.get("ma") else None
    out = []
    for row in ws.iter_rows(min_row=hdr_r + 1, values_only=True):
        if not row or ng_i >= len(row) or gtcl_i >= len(row):
            continue
        ng, gtcl = _num(row[ng_i]), _numz(row[gtcl_i])
        if ng is None:
            continue
        if norm(row[0]) in _SKIP_NAMES:      # dòng tổng/tiêu đề (vd Trạm sạc dòng "Tài sản")
            continue
        if ma_i is not None and not str(row[ma_i] or "").strip().upper().startswith("TS"):
            continue                          # spec GA: chỉ mã TSxx (bỏ CCDC/dòng tổng)
        status = row[status_i] if (status_i is not None and status_i < len(row)) else None
        out.append((ng, gtcl, status))
    return out


def _grouped_rows(rows, header_scan=8, ng="end"):
    """ng='end' -> cột 'Tài sản' CUỐI KỲ (Hưng Thịnh, spec 'tổng cột I'); ng='start' -> ĐẦU KỲ
    (An Taxi, spec 'tổng cột Tài sản (cột F)' — sub-header '01/xx'). Vị trí F/I/N đã kiểm ổn
    định cả 6 tháng T1-T6/2026 nhưng vẫn dò động theo header cho chắc."""
    grp_i = next((i for i, r in enumerate(rows[:header_scan])
                  if any("tai san" in norm(c) for c in r) and any("khau hao" in norm(c) for c in r)), None)
    if grp_i is None:
        return None
    group = A._forward_fill(rows[grp_i])
    sub = rows[grp_i + 1] if grp_i + 1 < len(rows) else []
    data = rows[grp_i + 2:]

    def _pick(kw, mode="end"):
        cands = [j for j in range(len(sub)) if j < len(group) and kw in norm(group[j])]
        # forward_fill lan nhãn nhóm sang cột TRỐNG bên phải (An Taxi/Hưng Thịnh: 'Giá trị sổ
        # sách' số ở 1 cột, cột kế rỗng) -> nếu không loại, `[-1]` chọn đúng cột rỗng và mọi dòng
        # bị bỏ (chỉ tiêu luôn 0 bất kể dữ liệu). Chỉ giữ cột thật sự CÓ SỐ ở vùng dữ liệu.
        cands = [j for j in cands if any(j < len(r) and isinstance(r[j], (int, float)) for r in data)]
        if not cands:
            return None
        if mode == "start":
            first = [j for j in cands if str(sub[j] or "").strip().startswith("01")]
            return (first or cands)[0]
        end = [j for j in cands if "31" in str(sub[j] or "") or "30" in str(sub[j] or "")]
        return (end or cands)[-1]
    ng_i = _pick("tai san", ng)
    gtsl_i = _pick("gia tri so sach") or _pick("gia tri")
    if ng_i is None or gtsl_i is None:
        return None
    out = []
    for r in data:
        if not r or all(c is None for c in r):
            continue
        # dòng CHI TIẾT tài sản có cột đặc tính (ngày mua lại/khấu hao lần đầu/phương thức, cột
        # B/C/D) — dòng loại/tổng (vd "24213 Chi phí phân bổ CCDC") chỉ có cột A + số gộp.
        has_detail = any(r[j] not in (None, "") for j in (1, 2, 3) if j < len(r))
        if not has_detail:
            continue
        ng = r[ng_i] if ng_i < len(r) else None
        gtsl = r[gtsl_i] if gtsl_i < len(r) else None
        ng, gtsl = _num(ng), _numz(gtsl)
        if ng is None:
            continue
        out.append((ng, gtsl, None))
    return out


def _split(rows_ngc):
    con_sd = thanh_ly = 0.0
    for ng, gtcl, status in rows_ngc:
        if not _is_zero(gtcl):
            continue
        st = norm(status)
        if "thanh ly" in st:
            thanh_ly += ng
        else:
            con_sd += ng
    return con_sd, thanh_ly


def _write(period, cong_ty, khoi, src, con_sd, thanh_ly):
    rec = [{"Kỳ": period, "Đơn vị": cong_ty,
            "TS hết KH còn sử dụng (NG, tỷ)": round(con_sd * 1e-9, 9),
            "TS hết KH chờ thanh lý (NG, tỷ)": round(thanh_ly * 1e-9, 9)}]
    out = os.path.join(tf.FILLED_DIR, f"TSHETKH_{period}_{cong_ty or 'NA'}_08_TSCD.xlsx")
    tf.fill("08_TSCD", rec, out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=khoi, source_file=src)
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"),
            "con_sd": round(con_sd * 1e-9, 9), "thanh_ly": round(thanh_ly * 1e-9, 9)}


def _sheet_thang(names, path):
    """Sheet 'Tháng N' ĐÚNG KỲ theo tên file (…M2026mm…); không khớp -> sheet 'Tháng*' đầu tiên.
    BẮT BUỘC ưu tiên đúng kỳ: file DUAN T4/2026 chứa đủ 6 sheet 'Tháng 1..6' — lấy sheet đầu
    tiên là dính kỳ khác (đã gây 643.7 tỷ thay vì ~238 tỷ trong DB 2026-07-30)."""
    m = re.search(r"\.M\d{4}(\d{2})\.", os.path.basename(path or ""))
    if m:
        sn = names.get(f"tháng {int(m.group(1))}") or names.get(f"thang {int(m.group(1))}")
        if sn:
            return sn
    return next((orig for low, orig in names.items()
                 if low.startswith("tháng") or low.startswith("thang")), None)


def compute(path):
    """Trả (con_sd, thanh_ly) RAW VND (chưa chia tỷ) từ báo cáo 'baocaotaisancodinhcongcudungcu',
    hoặc None nếu file/đơn vị này không khớp layout đã biết. Không ghi DB — dùng để GẮN vào record
    của _derive_tscd/_derive_tscd_duan (đơn vị mà 2 hàm đó đã ghi thành công cho CHÍNH source_file
    này, tránh 2 lần import_filled cùng source_file đè nhau) hoặc để extract() tự ghi standalone
    (đơn vị KHÔNG có dòng nào khác ghi cho source_file này — gồm cả HO từ 2026-07-30)."""
    folder = A._source_id(path).split("::", 1)[0].upper()
    if folder == "HO":
        try:
            return _compute_ho(path)
        except Exception:  # noqa: BLE001
            return None
    if folder == "SRVF" and path.lower().endswith(".xlsb"):
        try:
            return _compute_srvf_xlsb(path)
        except Exception:  # noqa: BLE001
            return None
    if folder not in _FLAT and folder != "DUAN" and folder not in _GROUPED and folder != "SRVF":
        return None
    try:
        wb = bb.fast_load_workbook(path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return None
    try:
        names = {s.strip().lower(): s for s in wb.sheetnames}
        rows_out = None
        if folder in _FLAT:
            cfg = _FLAT[folder]
            # GA: T1-T5 xuất tên sheet 'Tháng N' (cùng layout) -> fallback; đơn vị khác giữ STRICT
            # theo tên sheet spec (fallback mù layout lạ dễ đọc nhầm cột).
            sn = names.get(cfg["sheet"]) or (_sheet_thang(names, path) if folder == "GLOBALAI" else None)
            if sn:
                rows_out = _flat_rows(wb[sn], cfg)
        elif folder == "DUAN":
            sn = _sheet_thang(names, path)
            if sn:
                rows_out = _flat_rows(wb[sn], _DUAN)
        elif folder in _GROUPED:
            # Hưng Thịnh đặt tên sheet 'Tháng N' (không phải 'biểu khấu hao' như An Taxi/Trạm sạc).
            sn = names.get("biểu khấu hao") or _sheet_thang(names, path)
            if sn:
                rows = [list(r) for r in wb[sn].iter_rows(values_only=True)]
                # mapping v2 2026-07-31: NG = cột 'Tài sản' CUỐI KỲ cho CẢ 3 (An Taxi đổi từ cột F
                # đầu kỳ sang cột I cuối kỳ; HT cột I; Trạm sạc cột L — đều là sub-header ngày cuối).
                rows_out = _grouped_rows(rows)
        elif folder == "SRVF":
            # SRVF/XDV gửi .Xlsx (đuôi viết HOA — đường .xlsb cũ đã xử lý ở trên). Mapping v2
            # 2026-07-31: sheet 'Bảng tính KH chi tiết', lọc 'Giá trị còn lại (Cyber Tx.26)' = 0
            # (spec ghi nhãn 'Giá trị KH lũy kế (Cyber)' nhưng chữ cột BK + công thức GTCL=0 trỏ
            # vào cột GTCL Cyber ngay cạnh), cộng 'Nguyên giá cuối năm' (cột L). Fallback lần lượt:
            # sheet kỳ 'T<m>.<yy>' (mapping cũ) -> sheet 'Bao cao TS…' (layout .xlsb cũ).
            sn = next((s for s in wb.sheetnames if norm(s).startswith("bang tinh kh chi tiet")), None)
            if sn:
                rows = [list(r) for r in wb[sn].iter_rows(values_only=True)]
                rows_out = _srvf_khct_rows(rows)
            else:
                m = re.search(r"\.M(\d{4})(\d{2})\.", os.path.basename(path))
                sn = None
                if m:
                    key = f"t{int(m.group(2))}.{m.group(1)[2:]}"
                    sn = next((s for s in wb.sheetnames if s.strip().lower().rstrip(".") == key), None)
                if sn is None:
                    sn = next((s for s in wb.sheetnames if norm(s).startswith("bao cao ts")), None)
                if sn:
                    rows = [list(r) for r in wb[sn].iter_rows(values_only=True)]
                    rows_out = _srvf_rows(rows)
    finally:
        wb.close()
    if rows_out is None:
        return None
    return _split(rows_out)


def _compute_ho(path):
    """HO — sheet 'Theo dõi KH tài sản HO' (spec 2026-07-30): lọc GTCL=0, cộng 'NG cuối kỳ'.
    Dò cột theo TÊN header ('NG cuối kỳ'/'GTCL') vì vị trí TRÔI theo tháng (mỗi kỳ chèn 1 cột KH
    tháng: GTCL ở AN kỳ 12/25 -> AX kỳ 6/26 — spec ghi 'cột AX' chỉ đúng file T6). Dòng tài sản
    = Stt (cột A) là SỐ; lấy cả TSCĐ + CCDC (spec không loại trừ 'Phân Loại'). Không có cột
    trạng thái -> thanh_ly = 0."""
    wb = bb.fast_load_workbook(path, read_only=True, data_only=True)
    try:
        sn = next((s for s in wb.sheetnames if norm(s).startswith("theo doi kh tai san")), None)
        if not sn:
            return None
        rows = [list(r) for r in wb[sn].iter_rows(values_only=True)]
    finally:
        wb.close()
    hdr_i = next((i for i, r in enumerate(rows[:6]) if any(norm(c) == "stt" for c in r if c)), None)
    if hdr_i is None:
        return None
    h = rows[hdr_i]
    ng_j = next((j for j, c in enumerate(h) if c and norm(c).startswith("ng cuoi ky")), None)
    g_j = next((j for j, c in enumerate(h) if c and norm(c) == "gtcl"), None)
    if ng_j is None or g_j is None:
        return None
    con_sd = 0.0
    for r in rows[hdr_i + 1:]:
        if len(r) <= max(ng_j, g_j) or not isinstance(r[0], (int, float)):
            continue
        ng, g = _num(r[ng_j]), _numz(r[g_j])
        if ng and _is_zero(g):
            con_sd += ng
    return (con_sd, 0.0)


def _compute_srvf_xlsb(path):
    from pyxlsb import open_workbook
    wb = open_workbook(path)
    try:
        sn = next((s for s in wb.sheets if norm(s).startswith("bao cao ts")), None)
        if not sn:
            return None
        with wb.get_sheet(sn) as sh:
            rows = []
            for row in sh.rows():
                cells = {c.c: c.v for c in row}
                maxc = max(cells) if cells else -1
                rows.append([cells.get(i) for i in range(maxc + 1)])
    finally:
        wb.close()
    rows_out = _srvf_rows(rows)
    if rows_out is None:
        return None
    return _split(rows_out)


def _srvf_khct_rows(rows):
    """SRVF/XDV — sheet 'Bảng tính KH chi tiết' (mapping v2 2026-07-31): header dòng có 'Mã tài
    sản'; lọc cột 'Giá trị còn lại (Cyber Tx.26)' = 0 (dò theo TÊN — vị trí đổi theo tháng),
    cộng 'Nguyên giá cuối năm'. Dòng tài sản thật = có Mã tài sản (bỏ dòng tổng đầu sheet)."""
    hdr = next((i for i, r in enumerate(rows[:12])
                if any(norm(c) == "ma tai san" for c in r if c)), None)
    if hdr is None:
        return None
    h = rows[hdr]
    ma_i = next((j for j, c in enumerate(h) if c and norm(c) == "ma tai san"), None)
    ng_i = next((j for j, c in enumerate(h) if c and norm(c).startswith("nguyen gia cuoi nam")), None)
    g_i = next((j for j, c in enumerate(h) if c and "con lai" in norm(c) and "cyber" in norm(c)), None)
    if None in (ma_i, ng_i, g_i):
        return None
    rows_out = []
    for r in rows[hdr + 1:]:
        if ma_i >= len(r) or r[ma_i] in (None, "") or not str(r[ma_i]).strip():
            continue
        ng = _num(r[ng_i]) if ng_i < len(r) else None
        gtcl = _numz(r[g_i]) if g_i < len(r) else 0.0
        if ng is None:
            continue
        rows_out.append((ng, gtcl, None))
    return rows_out or None


def _srvf_rows(rows):
    """SRVF/XDV (chung file): dòng TSCĐ (LOẠI TRỪ 'Công cụ dụng cụ' — CCDC), GTCL cuối kỳ = 0,
    cộng 'Nguyên giá cuối kỳ'. Dò cột theo TÊN header — chữ cột I/Z/T trong spec là của layout
    cũ; file .Xlsx thật (2026, sheet T<m>.<yy>, header dòng 8): D='Loại tài sản',
    L='Nguyên giá cuối kỳ', R='Giá trị còn lại'. 'Loại tài sản' chứa TÊN NHÓM ('Máy móc, thiết
    bị', 'TSCĐ vô hình', 'Nhà cửa, vật kiến trúc'…, vài dòng bỏ trống) -> filter FLEX theo yêu
    cầu user 2026-07-30: loại CCDC thay vì đòi đúng chữ 'TSCĐ'; dòng tài sản thật phải có
    'Mã tài sản' (bỏ dòng tổng/rác cuối sheet)."""
    hdr = next((i for i, r in enumerate(rows[:12])
                if any(norm(c) == "loai tai san" for c in r if c)), None)
    if hdr is None:
        return None
    h = rows[hdr]

    def _col(pred):
        return next((j for j, c in enumerate(h) if c and pred(norm(c))), None)
    d_i = _col(lambda n: n == "loai tai san")
    t_i = _col(lambda n: n.startswith("nguyen gia cuoi"))
    z_i = _col(lambda n: n.startswith("gia tri con lai"))
    ma_i = _col(lambda n: n == "ma tai san")
    if None in (d_i, t_i, z_i):
        return None
    rows_out = []
    for r in rows[hdr + 1:]:
        if max(d_i, t_i, z_i) >= len(r):
            continue
        if ma_i is not None:
            ma = r[ma_i] if ma_i < len(r) else None
            if ma is None or not str(ma).strip():
                continue
        t = norm(r[d_i])
        if "cong cu" in t or t == "ccdc":     # CCDC — không phải TSCĐ
            continue
        ng, gtcl = _num(r[t_i]), _numz(r[z_i])
        if ng is None:
            continue
        rows_out.append((ng, gtcl, None))
    return rows_out or None


def _khoi_of(path):
    """KHÔNG dùng A._khoi_of() thẳng: mọi file 'baocaotaisancodinhcongcudungcu' đặt tên
    'B.9.<mã>...' — số '9' trùng MÃ KHỐI 9 = "Khối hỗ trợ tập đoàn" trong quy ước
    khoi_from_filename (số 9 ở đây là SỐ THỨ TỰ BÁO CÁO trong bộ hồ sơ kế toán, không phải khối!)
    -> _khoi_of() ưu tiên filename sẽ gán NHẦM "Khối hỗ trợ tập đoàn" cho MỌI đơn vị (đã phát
    hiện 2026-07-29: An Taxi/An KS/GA/HT/XVP/Dự án/Trạm sạc/SRVF đều bị gộp nhầm vào khối HO khi
    lọc theo Khối trên FE). Dò theo ĐƯỜNG DẪN THƯ MỤC trước (đáng tin hơn cho báo cáo này); GA
    không map được qua path -> fallback bảng cứng theo khối THẬT (khớp dòng "TSCĐ theo CĐKT" của
    chính đơn vị đó, đã verify trong DB 2026-07-29)."""
    from servers.common import contract as C
    byp = C.khoi_from_path(path)
    if byp:
        return byp
    folder = A._source_id(path).split("::", 1)[0].upper()
    return {"GLOBALAI": "Khối KD Công nghệ"}.get(folder)


def extract(path, period, cong_ty=None):
    """Standalone: TỰ TÍNH + GHI (source_file riêng). Chỉ dùng khi CHẮC CHẮN không có deriver nào
    khác ghi report_type TS cho CHÍNH source_file này trong CÙNG lần chạy. MỌI đơn vị (kể cả HO
    từ 2026-07-30 — proxy 223 trong _derive_tscd_cdkt đã gỡ; và An Taxi/Hưng Thịnh — dispatcher
    CHỦ ĐỘNG bỏ qua _derive_tscd cho các khối dùng CĐKT làm nguồn chính, tránh đếm đôi #17/#18
    Overview) đều ghi qua extract() này, source_file riêng, an toàn."""
    r = compute(path)
    if r is None:
        return {"ok": False, "skip": True}
    con_sd, thanh_ly = r
    khoi = _khoi_of(path)
    return _write(period, cong_ty, khoi, A._source_id(path), con_sd, thanh_ly)


# ── "Cơ cấu TSCĐ theo loại" (NG) + "Biến động tài sản theo loại" (Hao mòn LK), Khối Dự án ──────
# Chốt 2026-08-01. Spec user: filter cột "Loại tài sản" (H) = 1 trong 4 nhãn, rồi tổng
#   · cột "Nguyên giá" (N)                -> Chart 1 "Cơ cấu TSCĐ theo loại (NG)"
#   · cột "Giá trị khấu hao lũy kế" (O)   -> Chart 2 "Biến động tài sản theo loại", cột Hao mòn LK
# của CHÍNH sheet "Tháng N" (cùng file/sheet compute() đã đọc ở trên). THAY nguồn CĐPS TK 2112/
# 2113/2114 cũ trong agent_cli._derive_cdkt (nhánh DUAN "5 nhóm TSCĐ") — số CĐPS bị đứng yên/lệch
# nhiều tháng (vd 'Máy móc, thiết bị' ra Y HỆT 388,328701461 tỷ ở cả 6 kỳ — nghi đọc nhầm cột/kỳ),
# trong khi sổ tài sản chi tiết dao động đúng theo tháng (388,7-392,4 tỷ). Tag T5_NG_2/3/4 +
# T5_HM_2/3/4 khớp thứ tự _T5_CATS đã dùng ở agent_cli.py; T5_NG_6/T5_HM_6 "Tài sản khác" là tag
# MỚI (asset.py._T5_ORDER đã thêm) vì sổ TS có nhóm này nhưng không map vào TK 211x/213x nào trong
# 5 nhóm chuẩn. Nhánh DUAN trong agent_cli KHÔNG ghi T5_HM_* nào -> không đè/đếm đôi.
# "Tăng NG" (Chart 2) và nhóm "Nhà cửa, vật kiến trúc": spec user = "ko có" -> KHÔNG ghi (sổ TS
# không có cột PS tăng; nhãn "Nhà cửa, kiến trúc" chỉ xuất hiện ở file T04 và đang nghi lỗi nhập,
# xem memory duan-tscd-cocau-theo-loai) -> asset_extras trả tang=0, không có nhóm Nhà cửa.
_COCAU_T5_TAG = {"Máy móc, thiết bị": "2", "Phương tiện vận tải, truyền dẫn": "3",
                 "Thiết bị, dụng cụ quản lý": "4", "Tài sản khác": "6"}


def _cocau_by_loai_duan(ws):
    """Tổng Nguyên giá + Giá trị khấu hao lũy kế (RAW VND) theo "Loại tài sản" (cột H,
    _DUAN['loai']) của sheet "Tháng N" -> {loại: {"ng": …, "hm": …}}.

    "#REF!" (công thức cột Loại tài sản bị vỡ — gặp ở file T05/2026: 268/1066 dòng, 35,66 tỷ,
    KHÔNG gặp ở 5 tháng còn lại) GOM vào "Phương tiện vận tải, truyền dẫn": đối chiếu T01-T04+T06,
    nhóm PTVT dao động 148-172 tỷ; T05 (không tính #REF!) chỉ còn 135,7 tỷ — cộng lại 35,66 tỷ ra
    ~171,4 tỷ, khớp đúng dải bình thường, trong khi 3 nhóm còn lại ở T05 đều nằm trong dải bình
    thường của chúng (không thiếu hụt) -> #REF! gần như chắc chắn là PTVT bị lỗi đọc, không phải
    nhóm khác. Nhãn trống/lạ KHÁC "#REF!" -> "(chưa phân loại)" (không đoán, tránh gán sai nhóm
    khi chưa có bằng chứng như trên).

    NG và Hao mòn LK cộng ĐỘC LẬP nhau (dòng chỉ có 1 trong 2 ô vẫn tính cho ô có số; bỏ dòng khi
    CẢ HAI rỗng). Thực tế cả 6 file T01-T06/2026 mọi dòng đều có đủ 2 ô -> 2 cách cộng ra cùng số,
    quy ước này chỉ để không âm thầm mất số nếu kế toán bỏ trống 1 cột về sau."""
    ng_i, hm_i = _ci(_DUAN["ng"]) - 1, _ci(_DUAN["hm"]) - 1
    loai_i = _ci(_DUAN["loai"]) - 1
    agg = {}
    for row in ws.iter_rows(min_row=_DUAN["header_row"] + 1, values_only=True):
        if not row or max(ng_i, hm_i, loai_i) >= len(row):
            continue
        ng, hm = _num(row[ng_i]), _num(row[hm_i])
        if ng is None and hm is None:
            continue
        loai = str(row[loai_i]).strip() if row[loai_i] not in (None, "") else ""
        if loai == "#REF!":
            loai = "Phương tiện vận tải, truyền dẫn"
        elif not loai:
            loai = "(chưa phân loại)"
        a = agg.setdefault(loai, {"ng": 0.0, "hm": 0.0})
        a["ng"] += ng or 0.0
        a["hm"] += hm or 0.0
    return agg


def extract_cocau_duan(path, period, cong_ty=None):
    """Ghi "Cơ cấu TSCĐ theo loại" (NG, tag T5_NG_2/3/4/6) + "Hao mòn LK" (tag T5_HM_2/3/4/6) —
    Khối Dự án — vào 07_TAISAN_NV (đọc lại qua asset.py::asset_extras). source_file = CHÍNH file
    B.9 này (khác B.4 CĐKT/CĐPS mà agent_cli._derive_cdkt ghi) -> import_filled xoá-theo-
    source_file không đụng dòng CĐKT khác, không đếm đôi. Hao mòn ghi ÂM vào "Cuối kỳ (tỷ)" theo
    đúng quy ước agent_cli._t5_hm (asset_extras lấy abs()). Chỉ chạy cho folder DUAN (no-op —
    {"skip": True} — với đơn vị khác)."""
    folder = A._source_id(path).split("::", 1)[0].upper()
    if folder != "DUAN":
        return {"ok": False, "skip": True}
    try:
        wb = bb.fast_load_workbook(path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return {"ok": False, "error": "TSCĐ cơ cấu DUAN: không mở được file"}
    try:
        names = {s.strip().lower(): s for s in wb.sheetnames}
        sn = _sheet_thang(names, path)
        if not sn:
            return {"ok": False, "error": "TSCĐ cơ cấu DUAN: không thấy sheet 'Tháng N'"}
        agg = _cocau_by_loai_duan(wb[sn])
    finally:
        wb.close()
    records = []
    for ten, idx in _COCAU_T5_TAG.items():
        a = agg.get(ten) or {}
        if a.get("ng"):
            records.append({"Kỳ": period, "Đơn vị": cong_ty,
                            "Khoản mục (theo CĐKT)": f"[T5] {ten} - Nguyên giá (theo sổ TS)",
                            "Ghi chú": f"T5_NG_{idx}", "Cuối kỳ (tỷ)": round(a["ng"] * 1e-9, 9)})
        if a.get("hm"):
            records.append({"Kỳ": period, "Đơn vị": cong_ty,
                            "Khoản mục (theo CĐKT)": f"[T5] {ten} - Hao mòn LK (theo sổ TS)",
                            "Ghi chú": f"T5_HM_{idx}",
                            "Cuối kỳ (tỷ)": round(-abs(a["hm"]) * 1e-9, 9)})
    if not records:
        return {"ok": False, "error": "TSCĐ cơ cấu DUAN: không khớp nhóm nào trong sổ TS"}
    out = os.path.join(tf.FILLED_DIR, f"TSCOCAU_{period}_{cong_ty or 'NA'}_07_TAISAN_NV.xlsx")
    tf.fill("07_TAISAN_NV", records, out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(path), source_file=A._source_id(path))
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"),
            "by_loai_vnd": agg}


# ── "Cơ cấu TSCĐ theo loại" (NG) + "Biến động tài sản theo loại" (Hao mòn LK), HO ──────────────
# Chốt 2026-08-03, spec user. Sheet "Theo dõi KH tài sản HO" (CHÍNH sheet _compute_ho() dùng cho
# hết-khấu-hao, trong CHÍNH file B.9 này): filter cột "Loại tài sản" = 1 trong 5 nhãn, rồi tổng
#   · cột "NG cuối kỳ"   -> Chart 1 "Cơ cấu TSCĐ theo loại (NG)"
#   · cột "Tổng KHLK"    -> Chart 2 "Biến động tài sản theo loại", cột Hao mòn LK
# THAY nguồn CĐPS cũ trong agent_cli._derive_cdkt nhánh `_src == "HO"` — nhánh đó CHỈ có Chart 1
# (NG, tag T5_NG_0..3 theo _T5_CATS, KHÔNG có "Tài sản khác") và KHÔNG CÓ Chart 2 (HM) — spec cũ
# ghi "Chưa có (yêu cầu cung cấp)". Đã đổi nhánh đó thành `pass` (agent_cli.py) để tránh đếm đôi
# NG (2 nguồn cùng ghi T5_NG_1..4 cho HO). "Tăng NG"/"KH kỳ" (Chart 2) tính snapshot-diff ở metrics
# layer (asset.py::asset_extras, chốt 2026-08-03) — không cần cột PS tăng ở sheet này.
# "Chương trình phần mềm": spec user = "ko có" -> KHÔNG ghi (giống Dự án — không map được nhóm
# này trong sổ TS HO, đã verify quét toàn bộ cột "Loại tài sản" kỳ 06/2026 chỉ ra đúng 5 nhãn).
_COCAU_HO_TAG = {"Nhà cửa, vật kiến trúc": "1", "Máy móc, thiết bị": "2",
                 "Phương tiện vận tải, truyền dẫn": "3", "Thiết bị, dụng cụ quản lý": "4",
                 "Tài sản khác": "6"}


def _cocau_by_loai_ho(ws):
    """Tổng "NG cuối kỳ" + "Tổng KHLK" theo "Loại tài sản" của sheet 'Theo dõi KH tài sản HO' ->
    {loại: {"ng":…, "hm":…}}. Dò cột theo TÊN (KHÔNG cứng cột H/Z/BC): "Loại tài sản"/"NG cuối kỳ"
    đứng TRƯỚC vùng cột KH-theo-tháng nên vị trí ổn định, nhưng "Tổng KHLK" đứng SAU vùng đó nên
    TRÔI dần mỗi kỳ (mỗi tháng chèn thêm 1 cột KH, xem _compute_ho — cùng file, cùng hiện tượng).
    Dòng tài sản thật = Stt (cột A) là SỐ (bỏ dòng tổng/trống đầu sheet)."""
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    hdr_i = next((i for i, r in enumerate(rows[:6]) if any(norm(c) == "stt" for c in r if c)), None)
    if hdr_i is None:
        return {}
    h = rows[hdr_i]

    def _col(pred):
        return next((j for j, c in enumerate(h) if c and pred(norm(c))), None)
    loai_i = _col(lambda n: n.startswith("loai tai san"))
    ng_i = _col(lambda n: n.startswith("ng cuoi ky"))
    khlk_i = _col(lambda n: n.startswith("tong khlk"))
    if None in (loai_i, ng_i, khlk_i):
        return {}
    agg = {}
    for r in rows[hdr_i + 1:]:
        if not r or not isinstance(r[0], (int, float)) or max(loai_i, ng_i, khlk_i) >= len(r):
            continue
        loai = r[loai_i].strip() if isinstance(r[loai_i], str) else r[loai_i]
        if not loai:
            continue
        ng = r[ng_i] if isinstance(r[ng_i], (int, float)) else 0.0
        hm = r[khlk_i] if isinstance(r[khlk_i], (int, float)) else 0.0
        a = agg.setdefault(loai, {"ng": 0.0, "hm": 0.0})
        a["ng"] += ng
        a["hm"] += hm
    return agg


def extract_cocau_ho(path, period, cong_ty=None):
    """Ghi "Cơ cấu TSCĐ theo loại" (NG, tag T5_NG_1/2/3/4/6) + "Hao mòn LK" (tag T5_HM_1/2/3/4/6)
    — HO — vào 07_TAISAN_NV (đọc lại qua asset.py::asset_extras). source_file = CHÍNH file B.9
    này (khác B.4 CĐKT/CĐPS mà agent_cli._derive_cdkt ghi) -> import_filled xoá-theo-source_file
    không đụng dòng CĐKT khác, không đếm đôi (xem chú thích khối trên). Hao mòn ghi ÂM vào "Cuối
    kỳ (tỷ)" theo đúng quy ước agent_cli._t5_hm (asset_extras lấy abs()). Chỉ chạy cho folder HO
    (no-op — {"skip": True} — với đơn vị khác)."""
    folder = A._source_id(path).split("::", 1)[0].upper()
    if folder != "HO":
        return {"ok": False, "skip": True}
    try:
        wb = bb.fast_load_workbook(path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return {"ok": False, "error": "TSCĐ cơ cấu HO: không mở được file"}
    try:
        sn = next((s for s in wb.sheetnames if norm(s).startswith("theo doi kh tai san")), None)
        if not sn:
            return {"ok": False, "error": "TSCĐ cơ cấu HO: không thấy sheet 'Theo dõi KH tài sản HO'"}
        agg = _cocau_by_loai_ho(wb[sn])
    finally:
        wb.close()
    if not agg:
        return {"ok": False, "error": "TSCĐ cơ cấu HO: không đọc được cột Loại tài sản/NG cuối kỳ/Tổng KHLK"}
    records = []
    for ten, idx in _COCAU_HO_TAG.items():
        a = agg.get(ten) or {}
        if a.get("ng"):
            records.append({"Kỳ": period, "Đơn vị": cong_ty,
                            "Khoản mục (theo CĐKT)": f"[T5] {ten} - Nguyên giá (theo sổ TS HO)",
                            "Ghi chú": f"T5_NG_{idx}", "Cuối kỳ (tỷ)": round(a["ng"] * 1e-9, 9)})
        if a.get("hm"):
            records.append({"Kỳ": period, "Đơn vị": cong_ty,
                            "Khoản mục (theo CĐKT)": f"[T5] {ten} - Hao mòn LK (theo sổ TS HO)",
                            "Ghi chú": f"T5_HM_{idx}",
                            "Cuối kỳ (tỷ)": round(-abs(a["hm"]) * 1e-9, 9)})
    if not records:
        return {"ok": False, "error": "TSCĐ cơ cấu HO: không khớp nhóm nào trong sổ TS"}
    out = os.path.join(tf.FILLED_DIR, f"TSCOCAU_{period}_{cong_ty or 'NA'}_07_TAISAN_NV.xlsx")
    tf.fill("07_TAISAN_NV", records, out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(path), source_file=A._source_id(path))
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"),
            "by_loai_vnd": agg}


if __name__ == "__main__":
    import argparse
    import json
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--period", required=True)
    ap.add_argument("--cong-ty", dest="cong_ty", default=None)
    a = ap.parse_args()
    print(json.dumps(extract(a.file, a.period, a.cong_ty), ensure_ascii=False, default=str))
