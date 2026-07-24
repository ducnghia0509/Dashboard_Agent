# -*- coding: utf-8 -*-
"""Deriver CHUYÊN BIỆT cho GA (Global AI, Khối Công nghệ) — công nợ layout "BÁO CÁO CÔNG NỢ"
2 TẦNG header (khác sheet 131/331 nên deriver chung không nhận; trước để LLM -> chỉ nạp T01/T02).
Sheet 'PThu'/'PTra': [0]Tài khoản [1]Mã KH [2]Tên KH; nhóm cột Nợ/Có: Dư đầu · DƯ ĐẦU KỲ ·
PHÁT SINH TRONG KỲ · DƯ CUỐI KỲ. Map theo spec 50 chỉ tiêu (#36/#37):
  · CHỈ nhận TK 131 (PThu) / 331 (PTra) — LOẠI 3361 (phải trả nội bộ), 244 (ký quỹ),
    136/138/336/338… (sheet GA trộn nhiều TK trong cùng bảng).
  · Số dư = GROSS MỘT CHIỀU (khớp đối soát kế toán + CĐKT Mã 311/131):
      PTHU(05): dư cuối = dư Nợ 131  (dư Có = KH trả trước -> Mã 312 người mua trả trước, NGOÀI phạm vi).
      PTRA(06): dư cuối = dư Có 331  (dư Nợ = trả trước NCC -> Mã 132 tài sản, NGOÀI phạm vi).
    (KHÔNG net K−J: trả trước là dòng CĐKT riêng, không khấu trừ vào phải thu/phải trả.)
  · GIỮ dòng có phát sinh dù dư cuối = 0 (để #37 Tăng/Giảm không mất PS).
Per-customer (Tên KH). cong_ty=GA. Idempotent theo source_file.
Chạy: .venv/bin/python scripts/derive_ga_congno.py <file.xlsx> --period 2026-06
"""
import argparse
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from servers import template_filler as tf  # noqa: E402
from servers.common import be_bridge as bb  # noqa: E402
import agent_cli as A  # noqa: E402

norm = lambda v: bb.normalize_header(v, True)  # noqa: E731
TY = lambda v: round(v * 1e-9, 9) if isinstance(v, (int, float)) else None  # noqa: E731


def _find_cols(rows):
    """Dò header 2 tầng: r_lbl (nhóm 'DƯ CUỐI KỲ'...) + r_side (Nợ/Có). Trả (hdr_side_idx, cols).
    Nhóm là cặp Nợ/Có liên tiếp; thứ tự chuẩn: [đầu] · DƯ ĐẦU · PHÁT SINH · DƯ CUỐI (cuối cùng).
    Lấy DƯ CUỐI theo label (duy nhất), suy PS/ĐẦU theo offset -2/-4 (bền, tránh nhầm 'dư đầu' lặp)."""
    for i, r in enumerate(rows[:12]):
        low = [norm(c) for c in r]
        if not any("ten khach hang" in c for c in low):
            continue
        ten_kh = next((j for j, c in enumerate(low) if "ten khach hang" in c), 2)
        j_cuoi = next((j for j, c in enumerate(low) if "du cuoi ky" in c), None)
        if j_cuoi is None:
            continue
        side = [norm(c) for c in rows[i + 1]] if i + 1 < len(rows) else []
        # xác định Nợ/Có tại cụm DƯ CUỐI (mặc định cuoi_no=j_cuoi, cuoi_co=j_cuoi+1; hoán nếu ngược)
        no, co = j_cuoi, j_cuoi + 1
        if co < len(side) and side[no].startswith("co") and side[co].startswith("no"):
            no, co = co, no
        cols = {"ten": ten_kh, "tk": 0,
                "cuoi_no": no, "cuoi_co": co,
                "ps_no": no - 2, "ps_co": co - 2,
                "dau_no": no - 4, "dau_co": co - 4}
        return i + 1, cols
    return None, None


def _rows_of(rows, hdr_side, cols):
    """Sinh (tên, tk, dau_no, dau_co, ps_no, ps_co, cuoi_no, cuoi_co) cho mỗi dòng KH có tên."""
    out = []
    for r in rows[hdr_side + 1:]:
        if cols["ten"] >= len(r):
            continue
        ten = bb.parse_text(r[cols["ten"]]) if r[cols["ten"]] is not None else None
        ten = (str(ten).strip() if ten is not None else "")
        if not ten:
            continue

        def g(k):
            j = cols[k]
            return r[j] if (0 <= j < len(r) and isinstance(r[j], (int, float))) else None
        tk = bb.parse_text(r[cols["tk"]]) if cols["tk"] < len(r) and r[cols["tk"]] is not None else None
        out.append((ten, str(tk or "").strip(), g("dau_no"), g("dau_co"),
                    g("ps_no"), g("ps_co"), g("cuoi_no"), g("cuoi_co")))
    return out


def extract(path, period, cong_ty="GA"):
    wb = bb.fast_load_workbook(path, data_only=True, read_only=True)
    names = {s.strip().lower(): s for s in wb.sheetnames}
    src, khoi = A._source_id(path), A._khoi_of(path)
    out = {"ok": True}

    def sheet_rows(want):
        sn = names.get(want)
        if not sn:
            return None
        return [list(r) for r in wb[sn].iter_rows(values_only=True)]

    # ---- PThu -> 05_PHAITHU (dư cuối NỢ) ----
    pr = sheet_rows("pthu")
    if pr:
        hs, c = _find_cols(pr)
        if hs is not None:
            rec = []
            for ten, tk, dn, dc, pn, pc, cn, cc in _rows_of(pr, hs, c):
                if not tk.startswith("131"):   # chỉ TK 131 (KH); loại 244 ký quỹ, 136/138…
                    continue
                if not (cn or pn or pc):   # dư Nợ = phải thu; dư Có (KH trả trước) & dòng trống -> bỏ
                    continue
                rec.append({"Kỳ": period, "Đơn vị": cong_ty, "Khách hàng": ten,
                            "Dư cuối kỳ (tỷ)": TY(cn), "Dư đầu kỳ (tỷ)": TY(dn),
                            "PS tăng - Nợ (tỷ)": TY(pn), "PS giảm - Có (tỷ)": TY(pc)})
            if rec:
                p = os.path.join(tf.FILLED_DIR, f"GA_{period}_05_PHAITHU.xlsx")
                tf.fill("05_PHAITHU", rec, p)
                out["pthu"] = tf.import_filled(p, cong_ty=cong_ty, khoi=khoi, source_file=src).get("rows_imported")
                out["pthu_tong"] = round(sum(x["Dư cuối kỳ (tỷ)"] or 0 for x in rec), 6)

    # ---- PTra -> 06_PHAITRA (dư cuối CÓ) ----
    tr = sheet_rows("ptra")
    if tr:
        hs, c = _find_cols(tr)
        if hs is not None:
            rec = []
            for ten, tk, dn, dc, pn, pc, cn, cc in _rows_of(tr, hs, c):
                if not tk.startswith("331"):   # chỉ TK 331 (NCC); loại 3361 nội bộ, 336/338…
                    continue
                if not (cc or pc or pn):   # dư Có = phải trả; dư Nợ (trả trước NCC) & dòng trống -> bỏ
                    continue
                rec.append({"Kỳ": period, "Đơn vị": cong_ty, "Nhà cung cấp": ten,
                            "Dư cuối kỳ (tỷ)": TY(cc), "Dư đầu kỳ (tỷ)": TY(dc),
                            "PS tăng - Có (tỷ)": TY(pc), "PS giảm - Nợ (tỷ)": TY(pn)})
            if rec:
                p = os.path.join(tf.FILLED_DIR, f"GA_{period}_06_PHAITRA.xlsx")
                tf.fill("06_PHAITRA", rec, p)
                out["ptra"] = tf.import_filled(p, cong_ty=cong_ty, khoi=khoi, source_file=src).get("rows_imported")
                out["ptra_tong"] = round(sum(x["Dư cuối kỳ (tỷ)"] or 0 for x in rec), 6)
    wb.close()
    return out


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--period", required=True)
    ap.add_argument("--cong-ty", dest="cong_ty", default="GA")
    a = ap.parse_args()
    import json
    print(json.dumps(extract(a.file, a.period, a.cong_ty), ensure_ascii=False, default=str))
