# -*- coding: utf-8 -*-
"""Test hồi quy cho derive_congno_tuoino.py — verify cột đọc THẬT trên file nguồn cố định
(T06/2026), không mock. Chạy: python scripts/test_derive_congno_tuoino.py

Chốt 2026-08-08: mode 'hanno' được mở rộng để tính LUÔN age-bucket (tuoi_no_1t/aging_13/aging_36/
aging_6p) song song field hạn-nợ (tong_no/trong_han/den_han/qua_han/qh_*) — test này khoá cả 2 bộ
số bằng file thật, để lần sau ai sửa cột/regex không âm thầm làm lệch 1 trong 2.
"""
import glob
import os
import sys

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.normpath(os.path.join(_HERE, ".."))
sys.path.insert(0, _ROOT)
sys.path.insert(0, _HERE)

import derive_congno_tuoino as d  # noqa: E402
import openpyxl  # noqa: E402

_RECEIVED = "/home/itadmin/AI_Dashboard_QT/Connect_VPS/received_reports"
_failed = False


def _ok(label):
    print(f"[OK] {label}")


def _fail(label, detail=""):
    global _failed
    _failed = True
    print(f"[FAIL] {label} {detail}")


def _close(a, b, tol=1e-6):
    return a is not None and b is not None and abs(a - b) <= tol


def _payload_for(folder, filename, period):
    path = os.path.join(_RECEIVED, folder, "baocaotuoino", filename)
    if not os.path.exists(path):
        return None, f"file không tồn tại: {path}"
    unit = d._UNITS.get(folder)
    hints = unit[3] if len(unit) > 3 else None
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        sn = d._find_sheet(wb, period, hints)
        if not sn:
            return None, "không tìm được sheet 'tuổi nợ'"
        rows = [list(r) for r in wb[sn].iter_rows(values_only=True)]
    finally:
        wb.close()
    mode = unit[2]
    if mode == "hanno":
        res = d._agg_hanno(rows, d._period_end(period))
    elif mode == "hanno_tong_xdv":
        res = d._agg_hanno_tong(rows, d._find_cols_xdv(rows))
    elif mode == "hanno_tong":
        res = d._agg_hanno_tong(rows)
    else:
        res = d._agg_age(rows, d._period_end(period))
    if res is None:
        return None, "aggregator trả None (không dò được cột)"
    return res["payload"], None


# (folder, file T06/2026, period, mode kỳ vọng, field cần khoá số -> giá trị verify 2026-08-08)
_CASES = [
    ("TRAMSAC", "B.3.TC.TCKT.M.202606.Baocaotuoino.xlsx", "2026-06", "hanno", {
        "tong_no": 5.447868233, "qua_han": 0.476352001, "trong_han": 5.148316232,
        "tuoi_no_1t": 5.148316232, "aging_36": 0.249512001, "aging_6p": 0.05004,
    }),
    ("XANHVINHPHUC", "B.6.XVP.M.202606.Baocaotuoinophaithu.xlsx", "2026-06", "hanno", {
        "tong_no": 82.569209353, "qua_han": 0.0,
        "tuoi_no_1t": 50.041490386, "aging_13": 5.504841629,
        "aging_36": 13.468811434, "aging_6p": 13.554065904,
    }),
    ("HTXXANHTUYENQUANG", "B.6.HTX_XTQ.M.202606.Baocaotuoinophaithu.xlsx", "2026-06", "hanno", {
        "tong_no": 0.030356457, "qua_han": 0.0, "tuoi_no_1t": 0.030356457,
    }),
    ("HTXXANHVINHPHUC", "B.6.HTX_XVP.M.202606.Baocaotuoinophaithu.xlsx", "2026-06", "hanno", {
        "tong_no": 1.443199828, "qua_han": 0.0, "tuoi_no_1t": 1.443199828,
    }),
    ("GLOBALAI", "B.8.GA.TCKT.M.202606.Baocaotuoino.xlsx", "2026-06", "hanno", {
        "tong_no": 0.0, "qua_han": 0.0, "tuoi_no_1t": 0.0,
    }),
    ("SRVF", "B.1.TC.TCKT.M.202606.Baocaotuoino.xlsx", "2026-06", "hanno", {
        "tong_no": 194.273764018, "qua_han": 28.528111818, "qh_1_30": 9.27765,
        # SRVF không có 'Ngày hóa đơn' điền thật -> age-bucket vẫn 0 (không phải regression, xem
        # docstring _agg_hanno) — khoá số 0 để phát hiện nếu sau này có ai vô tình bơm giá trị giả.
        "tuoi_no_1t": 0.0,
    }),
]


def test_units_mode():
    print("\n== mode + số liệu theo đơn vị (file T06/2026 thật) ==")
    for folder, filename, period, exp_mode, expect in _CASES:
        unit = d._UNITS.get(folder)
        if unit is None:
            _fail(folder, "không có trong _UNITS")
            continue
        if unit[2] != exp_mode:
            _fail(folder, f"mode={unit[2]!r} (kỳ vọng {exp_mode!r})")
            continue
        payload, err = _payload_for(folder, filename, period)
        if err:
            _fail(folder, err)
            continue
        bad = [(k, payload.get(k), v) for k, v in expect.items() if not _close(payload.get(k), v)]
        if bad:
            _fail(folder, f"lệch: {bad}")
        else:
            _ok(f"{folder}: {len(expect)} field khớp (mode={unit[2]})")


def test_ho_khong_co_du_lieu():
    """HO để nguyên mode 'age' vì file thật (2026-08-08) không có dòng khách hàng nào — nếu sau
    này ai bơm dữ liệu chi tiết vào file mà quên đổi mode, test KHÔNG bắt được (chỉ cảnh báo n_rows
    vẫn 0 để nhắc soát lại), không fail cứng (tránh vỡ pipeline khi kế toán chưa gửi file mới)."""
    print("\n== HO (kỳ vọng vẫn rỗng, mode 'age') ==")
    unit = d._UNITS.get("HO")
    if unit[2] != "age":
        _fail("HO", f"mode đã đổi thành {unit[2]!r} — nhớ kiểm tra lại file trước khi đổi mode")
        return
    path = os.path.join(_RECEIVED, "HO", "baocaotuoino", "B.9.TC.TCKT.M.202606.Baocaotuoino.xlsx")
    if not os.path.exists(path):
        _ok("HO: không có file để kiểm (bỏ qua)")
        return
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    has_data = any(ws.cell(row=r, column=5).value for r in range(4, ws.max_row + 1))
    wb.close()
    if has_data:
        print("[WARN] HO giờ ĐÃ có dòng khách hàng — cân nhắc đổi mode 'age' -> 'hanno' như 5 khối kia")
    else:
        _ok("HO: file vẫn rỗng (đúng lý do giữ mode 'age')")


if __name__ == "__main__":
    test_units_mode()
    test_ho_khong_co_du_lieu()
    if _failed:
        print("\n== CÓ TEST FAIL ==")
        sys.exit(1)
    print("\n== TẤT CẢ TEST PASS ==")
