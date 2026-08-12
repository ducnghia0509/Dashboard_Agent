#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Deterministic extractor cho "phần Vay trở xuống" của màn Dòng tiền.

Đọc báo cáo ngân hàng trong THUCHI/baocaonganhang, xuất JSON theo schema mục 8 của
spec: nguyen_tac_lay_no_den_han_dashboard.md (thư mục anh em của repo)

Thuần rule-based (openpyxl), KHÔNG LLM. Mỗi bản ghi = (đơn vị, tháng).

LỚP 1 (số tổng)  : sheet 'BCTH 2'  -> credit[] / lc[] / baolanh[] / tien[] / intercompany
LỚP 2 (bổ sung)  : nợ gốc đến hạn theo lịch:
                     - NH/TH: ma trận cột-tháng 1..12 (dòng tổng ngay dưới header 1..12)
                     - Thịnh Cường trung hạn: cột rộng 'Trả gốc T<MM>/YYYY'
                   LC đến hạn (sheet LC 'Hạn TT LC'); bảo lãnh (sheet 'Dư BL' 'Ngày hết hạn')
"""
import openpyxl, glob, json, os, re, datetime, sys

_WS = os.path.normpath(os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".."))  # cha của Dashboard_Agent/
DATA_DIR = os.path.join(_WS, "Connect_VPS", "received_reports", "THUCHI", "baocaonganhang")
# Gốc quét THẬT: cả THUCHI/, không riêng baocaonganhang. Từ 11/08/2026 agent .253 đổi cách phân
# loại, đẩy 10 file ngân hàng T07/T08 sang `baocaothuchi/<Báo cáo X>/` (qua sub_folder) — glob một
# thư mục không đệ quy nên bỏ qua sạch, chain VAY vẫn đọc bản cũ nằm ở chỗ cũ và tháng 8 thiếu
# 86,9 tỷ dư nợ (Thịnh Cường lệch 85,09 tỷ). Không ai thấy vì cả hai bản đều tồn tại.
SCAN_ROOT = os.path.join(_WS, "Connect_VPS", "received_reports", "THUCHI")
# Ghi cạnh DATA_DIR (thư mục anh em của repo) — trước hard-code /home/sysadmin nên user khác
# chạy là PermissionError. Env OUT_JSON để ghi chỗ khác.
OUT_JSON = os.environ.get("OUT_JSON") or os.path.join(_WS, "cashflow_vay_extract.json")

# ---- nhận diện đơn vị từ TÊN FILE (thứ tự: chuỗi đặc trưng trước) ----
def unit_of(fn):
    U = fn.upper()
    if "HTX XANH" in U:        return "HTX_Xanh"
    if "XANH VĨNH PHÚC" in U:  return "XanhVP"
    if "HƯNG THỊNH" in U:      return "HungThinh"
    if "THỊNH CƯỜNG" in U:     return "ThinhCuong"
    if "QUẢNG NINH" in U:      return "QuangNinh"
    if "AAG" in U:             return "AN"
    return None

def month_of(fn):
    m = re.search(r'2026(0[1-9]|1[0-2])', fn)
    if m: return int(m.group(1))
    m = re.search(r'T\s*(\d{1,2})\.2026', fn) or re.search(r'THÁNG\s*0?(\d{1,2})\.2026', fn)
    return int(m.group(1)) if m else None

# ---- helpers ----
def load(f): return openpyxl.load_workbook(f, read_only=True, data_only=True)
def s(v):    return str(v).strip().lower() if v is not None else ""
def num(v):  return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None

def rows_of(ws, maxr=None):
    return list(ws.iter_rows(min_row=1, max_row=maxr, values_only=True))

def find_col(cells, *kws):
    for j, c in enumerate(cells):
        cs = s(c)
        if cs and all(k in cs for k in kws): return j
    return None

def get(row, j):
    return row[j] if (j is not None and j < len(row)) else None

# ---- LỚP 2a: ma trận cột-tháng (NH/TH) ----
def matrix_month(ws, month):
    rs = rows_of(ws, 8)
    for i, r in enumerate(rs):
        pos = {c: j for j, c in enumerate(r) if isinstance(c, int) and 1 <= c <= 12 and j >= 10}
        if 1 in pos and 12 in pos and len(pos) >= 12:
            tr = rs[i + 1] if i + 1 < len(rs) else None
            if tr is None: return None
            return num(get(tr, pos[month])) or 0.0
    return None  # không có ma trận

def sheet_kind(ws):
    for r in rows_of(ws, 6):
        for c in r:
            u = str(c).upper() if isinstance(c, str) else ""
            if "DƯ NỢ NGẮN HẠN" in u: return "NH"
            # trung/dài hạn: đủ chữ "TRUNG HẠN" HOẶC viết tắt "DƯ NỢ TH " / "DƯ NỢ TH-" (vd 'TH BAB')
            if "DƯ NỢ TRUNG HẠN" in u or "DƯ NỢ TH " in u or "DƯ NỢ TH-" in u: return "TH"
    return None

# ---- LỚP 2b: Thịnh Cường cột rộng 'Trả gốc T<MM>/YYYY' ----
def wide_tra_goc(ws, month, year=2026):
    rs = rows_of(ws, 5)
    hi = next((i for i, r in enumerate(rs)
               if sum(1 for c in r if isinstance(c, str) and "trả gốc" in c.lower()) >= 3), None)
    if hi is None: return None, "no_wide_header"
    hdr = rs[hi]
    tok = (f"t{month:02d}/{year}", f"t{month}/{year}")
    cols = [j for j, c in enumerate(hdr)
            if isinstance(c, str) and "trả gốc" in c.lower()
            and any(t in c.lower().replace(" ", "").replace("\n", "") for t in tok)
            and "tk" not in c.lower() and "539908" not in c.lower()]
    if not cols: return None, "missing_month_column"
    total = 0.0
    for r in ws.iter_rows(min_row=hi + 3, values_only=True):   # bỏ hàng tổng 'Dư nợ hiện tại'
        for j in cols:
            v = num(get(r, j))
            if v: total += v
    return total, "ok"

# ---- LỚP 1: đọc BCTH 2 ----
def is_total_row(row):
    """Dòng 'TỔNG CỘNG' — chữ có thể nằm ở bất kỳ cột nào (col2/4/5)."""
    for c in row:
        cs = s(c)
        if "tổng cộng" in cs or "tổng cong" in cs: return True
    return False

def is_hung_thinh(code, tt):
    """Dòng thuộc Hưng Thịnh (để loại khỏi file Thịnh Cường, tránh đếm đôi)."""
    if s(tt) == "hưng thịnh": return True
    cu = str(code).upper() if code else ""
    return "_HT_" in cu or re.search(r'\bHT\b', cu) is not None

def classify(code, bank, name):
    t = " ".join(s(x) for x in (code, bank, name))
    if "dư bl" in t or "bảo lãnh" in t: return "baolanh"
    if re.search(r'\blc\b', t): return "lc"
    return "credit"

def read_bcth2(wb, unit):
    if "BCTH 2" not in wb.sheetnames:
        return None, ["missing_bcth2"]
    ws = wb["BCTH 2"]
    rs = rows_of(ws)
    warn = []

    # ranh giới khối tiền
    tien_start = next((i for i, r in enumerate(rs)
                       if any(isinstance(c, str) and "TÀI KHOẢN TIỀN" in c.upper() for c in r)), len(rs))

    # header khối tín dụng (chứa 'nợ đầu kỳ' & 'nợ cuối kỳ')
    ch = next((i for i, r in enumerate(rs[:tien_start])
               if find_col(r, "nợ đầu") is not None and find_col(r, "nợ cuối") is not None), None)
    credit, lc, baolanh, intercompany = [], [], [], None
    if ch is None:
        warn.append("no_credit_header")
    else:
        H = rs[ch]
        c_code = find_col(H, "stt");  c_tt = find_col(H, "tt") if find_col(H, "tt") != c_code else None
        c_bank = find_col(H, "dư nợ tại nh"); c_name = find_col(H, "thời hạn")
        c_dau = find_col(H, "nợ đầu"); c_vay = find_col(H, "tổng số vay")
        c_tt2 = find_col(H, "đã thanh toán"); c_cuoi = find_col(H, "nợ cuối")
        # fallback vị trí cố định (spec 3.2): code=2 tt=3 bank=4 name=5 dau=6 vay=7 tt=8 cuoi=9 (1-based)
        c_code = c_code if c_code is not None else 1
        c_tt   = c_tt   if c_tt   is not None else 2
        c_bank = c_bank if c_bank is not None else 3
        c_name = c_name if c_name is not None else 4
        c_dau  = c_dau  if c_dau  is not None else 5
        c_vay  = c_vay  if c_vay  is not None else 6
        c_tt2  = c_tt2  if c_tt2  is not None else 7
        c_cuoi = c_cuoi if c_cuoi is not None else 8
        for r in rs[ch + 1: tien_start]:
            code, tt = get(r, c_code), get(r, c_tt)
            bank, name = get(r, c_bank), get(r, c_name)
            dau = num(get(r, c_dau))
            if dau is None and num(get(r, c_cuoi)) is None:  # dòng trống/không số
                continue
            if is_total_row(r):                              # bỏ dòng TỔNG CỘNG (chữ ở cột bất kỳ)
                continue
            rec = {"code": (str(code).strip() if code else None),
                   "tt": (str(tt).strip() if tt else None),
                   "bank": (str(bank).strip() if bank else None),
                   "name": (str(name).strip() if name else None),
                   "no_dau_ky": dau or 0.0, "tong_vay": num(get(r, c_vay)) or 0.0,
                   "da_thanh_toan": num(get(r, c_tt2)) or 0.0, "no_cuoi_ky": num(get(r, c_cuoi)) or 0.0}
            kind = classify(code, bank, name)
            # --- chống đếm đôi / hợp nhất ---
            if unit == "ThinhCuong" and is_hung_thinh(code, tt):
                continue                                     # HT lấy từ file HungThinh
            if unit == "XanhVP" and str(code).strip().upper() == "TC":
                intercompany = rec; continue                 # vay nội bộ -> tách
            (credit if kind == "credit" else lc if kind == "lc" else baolanh).append(rec)

    # khối tiền
    tien = []
    if tien_start < len(rs):
        bh = next((i for i in range(tien_start, min(tien_start + 4, len(rs)))
                   if find_col(rs[i], "số dư đầu") is not None), None)
        if bh is not None:
            H = rs[bh]
            c_tk = find_col(H, "tài khoản") or 3
            c_num = find_col(H, "số tk"); c_dau = find_col(H, "số dư đầu")
            c_thu = find_col(H, "thu"); c_chi = find_col(H, "chi"); c_cuoi = find_col(H, "số dư cuối")
            c_dau = c_dau if c_dau is not None else 5
            c_thu = c_thu if c_thu is not None else 6
            c_chi = c_chi if c_chi is not None else 7
            c_cuoi = c_cuoi if c_cuoi is not None else 8
            for r in rs[bh + 1:]:
                dau = num(get(r, c_dau))
                if dau is None and num(get(r, c_thu)) is None and num(get(r, c_cuoi)) is None:
                    continue
                tk = get(r, c_tk); tkno = get(r, c_num)
                nm = " ".join(x for x in (str(tk) if tk else "", str(tkno) if tkno else "") if x).strip()
                tien.append({"tk": nm or None,
                             "so_du_dau": dau or 0.0, "thu": num(get(r, c_thu)) or 0.0,
                             "chi": num(get(r, c_chi)) or 0.0, "so_du_cuoi": num(get(r, c_cuoi)) or 0.0,
                             "is_total": ("tài khoản ngân hàng" in s(tk))})
    return {"credit": credit, "lc": lc, "baolanh": baolanh, "tien": tien,
            "intercompany": intercompany}, warn

# ---- LỚP 2c: LC đến hạn (sheet LC 'Hạn TT LC') ----
def lc_due(wb, unit, month, year=2026):
    # chọn sheet LC của chính đơn vị (bỏ 'LC Hưng Thịnh' khi không phải HungThinh)
    names = [n for n in wb.sheetnames if n.upper().startswith("LC")]
    if unit != "HungThinh":
        names = [n for n in names if "HƯNG THỊNH" not in n.upper()] or names
    out = []
    for nm in names:
        ws = wb[nm]; rs = rows_of(ws, 60)
        hi = next((i for i, r in enumerate(rs) if find_col(r, "hạn tt lc") is not None), None)
        if hi is None: continue
        H = rs[hi]
        c_han = find_col(H, "hạn tt lc"); c_open = find_col(H, "ngày mở")
        c_val = find_col(H, "trị giá lc (vnd)") or find_col(H, "vnd")
        c_paid = find_col(H, "đã thanh toán"); c_name = find_col(H, "tên lc")
        for r in ws.iter_rows(min_row=hi + 2, values_only=True):
            han = get(r, c_han)
            # 'Hạn TT LC' có thể là số ngày HOẶC ngày; chỉ nhận datetime để lọc theo tháng
            if isinstance(han, datetime.datetime) and han.year == year and han.month == month:
                out.append({"sheet": nm, "name": (str(get(r, c_name)).strip() if c_name is not None and get(r, c_name) else None),
                            "han_tt": han.strftime("%Y-%m-%d"),
                            "tri_gia_vnd": num(get(r, c_val)) or 0.0,
                            "da_thanh_toan": num(get(r, c_paid)) or 0.0})
    return out

# ---- LỚP 2d: bảo lãnh (Dư BL) ----
def baolanh_detail(wb):
    if "Dư BL" not in wb.sheetnames: return []
    ws = wb["Dư BL"]; rs = rows_of(ws, 300)
    hi = next((i for i, r in enumerate(rs) if find_col(r, "ngày hết hạn") is not None), None)
    if hi is None: return []
    H = rs[hi]
    c_val = find_col(H, "giá trị"); c_exp = find_col(H, "ngày hết hạn")
    c_no = find_col(H, "số thư"); c_left = find_col(H, "còn lại")
    out = []
    for r in ws.iter_rows(min_row=hi + 1, values_only=True):
        val = num(get(r, c_val)); exp = get(r, c_exp)
        if val is None and not isinstance(exp, datetime.datetime): continue
        out.append({"so_thu": (str(get(r, c_no)).strip() if c_no is not None and get(r, c_no) else None),
                    "gia_tri": val or 0.0,
                    "ngay_het_han": exp.strftime("%Y-%m-%d") if isinstance(exp, datetime.datetime) else None,
                    "con_lai": num(get(r, c_left)) or 0.0})
    return out

# ---- LỚP 2a+2b gắn đến hạn vào credit + tính tổng đến hạn ----
def den_han(wb, unit, month):
    short = medium = 0.0; had = False; notes = []
    for nm in wb.sheetnames:
        if nm.startswith("foxz"): continue
        k = sheet_kind(wb[nm])
        if k == "NH":
            v = matrix_month(wb[nm], month)
            if v is not None: short += v; had = True; notes.append(f"NH:{nm}")
        elif k == "TH":
            v = matrix_month(wb[nm], month)
            if v is not None: medium += v; had = True; notes.append(f"TH:{nm}")
    if unit == "ThinhCuong":
        for nm in ("Trung hạn 2022", "Trung hạn BIDV"):
            if nm in wb.sheetnames:
                v, st = wide_tra_goc(wb[nm], month)
                if v is not None: medium += v; had = True; notes.append(f"{nm}:{st}")
                else: notes.append(f"{nm}:{st}")
    return (short, medium, had, notes)

# ---- DATA totals (để verify Thu/Chi) ----
def data_totals(wb, month, year=2026):
    """Tổng THU/CHI của sheet DATA LỌC theo tháng báo cáo (DATA là lũy kế từ đầu năm)."""
    if "DATA" not in wb.sheetnames: return None
    ws = wb["DATA"]
    rs = rows_of(ws, 2)
    if len(rs) < 2: return None
    H = rs[1]  # hàng 2 = tiêu đề
    c_date = find_col(H, "ngày"); c_thu = find_col(H, "thu"); c_chi = find_col(H, "chi")
    if c_thu is None or c_chi is None or c_date is None: return None
    thu = chi = 0.0
    for r in ws.iter_rows(min_row=3, values_only=True):
        d = get(r, c_date)
        if not (isinstance(d, datetime.datetime) and d.year == year and d.month == month):
            continue
        thu += num(get(r, c_thu)) or 0.0
        chi += num(get(r, c_chi)) or 0.0
    return {"thu": thu, "chi": chi}

# ---- build 1 bản ghi ----
def extract_file(f):
    unit, month = unit_of(os.path.basename(f)), month_of(os.path.basename(f))
    rec = {"period": f"2026-{month:02d}" if month else None, "unit": unit,
           "file_path": f, "credit": [], "lc": [], "baolanh": [], "tien": [],
           "intercompany": None, "totals": {}, "warnings": []}
    wb = load(f)
    if "BCTH" in wb.sheetnames and "BCTH 2" not in wb.sheetnames:
        rec["warnings"].append("stale_bcth_only")
    b2, w = read_bcth2(wb, unit)
    rec["warnings"] += w
    if b2:
        rec.update({k: b2[k] for k in ("credit", "lc", "baolanh", "tien", "intercompany")})
    # LỚP 2: đến hạn
    short, medium, had, notes = den_han(wb, unit, month)
    rec["den_han_sources"] = notes
    if not had:
        if unit in ("HTX_Xanh", "HungThinh"): rec["warnings"].append("missing_detail_sheets")
    # LC + BL
    rec["lc_den_han"] = lc_due(wb, unit, month)
    if not rec["baolanh"]:
        rec["baolanh"] = baolanh_detail(wb)
    # cờ hợp nhất
    if unit == "HTX_Xanh": rec["warnings"].append("htx_no_own_source")
    # totals
    no_cuoi = sum(c["no_cuoi_ky"] for c in rec["credit"])
    thu = sum(t["thu"] for t in rec["tien"] if not t.get("is_total"))
    chi = sum(t["chi"] for t in rec["tien"] if not t.get("is_total"))
    so_du_cuoi = sum(t["so_du_cuoi"] for t in rec["tien"] if not t.get("is_total"))
    rec["totals"] = {"no_cuoi_ky": no_cuoi, "den_han_ngan": short, "den_han_trung": medium,
                     "den_han_thang": short + medium, "thu": thu, "chi": chi, "so_du_cuoi": so_du_cuoi}
    dt = data_totals(wb, month)              # đối soát DATA (lọc tháng) vs BCTH2 (integrity check)
    rec["_data_totals"] = dt                  # để verify, không thuộc schema
    if dt and (thu or chi):
        # đối soát tư vấn: chênh > 3% VÀ > 100 triệu => DATA không khớp BCTH2 (vd DATA bị copy)
        def big(a, b): return abs(a - b) > 0.03 * max(abs(a), abs(b), 1) and abs(a - b) > 100e6
        if big(dt["thu"], thu) or big(dt["chi"], chi):
            rec["warnings"].append("data_reconcile_diff")
    wb.close()
    return rec

def files_can_doc(root=None):
    """Danh sách file ngân hàng CẦN ĐỌC — quét đệ quy, mỗi (đơn vị, tháng) giữ bản MỚI NHẤT.

    Ba lý do phải làm chỗ này thay vì glob thẳng:
      1. File có thể nằm ở `baocaonganhang/` HOẶC `baocaothuchi/<Báo cáo X>/` — agent đổi cách
         phân loại là chỗ cũ đứng yên, không báo lỗi.
      2. Cùng một (đơn vị, tháng) có mặt ở CẢ HAI nơi với nội dung KHÁC nhau; đọc nhầm bản cũ là
         sai số thật, không phải sai nhãn.
      3. Đuôi `.Xlsx` viết hoa cũng phải bắt — cùng bẫy đã diệt ở spec_extract.

    Bản mới nhất chọn theo mtime, không theo thư mục: không giả định nơi nào "đúng" hơn.
    """
    root = root or SCAN_ROOT
    tot = {}
    for f in glob.glob(os.path.join(root, "**", "*"), recursive=True):
        ten = os.path.basename(f)
        if not ten.lower().endswith((".xlsx", ".xlsm")) or ten.startswith("~$"):
            continue
        u = unit_of(ten)
        if u is None:
            continue                        # không nhận ra đơn vị -> không phải báo cáo ngân hàng
        khoa = (u, month_of(ten))
        cu_hon = tot.get(khoa)
        if cu_hon is None or os.path.getmtime(f) > os.path.getmtime(cu_hon):
            tot[khoa] = f
    return sorted(tot.values())


def main():
    files = files_can_doc()
    out = []
    for f in files:
        if unit_of(os.path.basename(f)) is None: continue
        out.append(extract_file(f))
    dst = OUT_JSON
    with open(dst, "w", encoding="utf-8") as fh:
        json.dump(out, fh, ensure_ascii=False, indent=2, default=str)
    print(f"OK: {len(out)} bản ghi -> {dst}")
    return out

if __name__ == "__main__":
    main()
