# -*- coding: utf-8 -*-
"""Deriver BỔ SUNG chiều NGƯỢC của công nợ (số dư đảo bên) mà pipeline vàng CỐ Ý bỏ:

  · TK 331 "công nợ phải trả": số dư CUỐI KỲ **Nợ** = TRẢ TRƯỚC NCC (mình ứng tiền cho NCC)
        -> report_type 'PTRA_ADV'.  (chiều Có = phải trả đã có ở PTRA)
  · TK 131 "công nợ phải thu": số dư CUỐI KỲ **Có** = NGƯỜI MUA TRẢ TIỀN TRƯỚC
        -> report_type 'PTHU_ADV'.  (chiều Nợ = phải thu đã có ở PTHU)

Bối cảnh: importer/template vàng gộp mỗi đối tượng thành 1 số dư ròng và chỉ giữ chiều
"thuận" (PTHU=dư Nợ 131, PTRA=dư Có 331). Chiều đảo bị "KHÔNG tính" (xem derive_ga_congno.py
dòng 6-7) nên chỉ tiêu "Trả trước NCC" luôn = 0 dù báo cáo có (HT T06 = 28,2 tỷ). Deriver này
đọc THẲNG sổ tổng hợp công nợ (2 tầng header Đầu kỳ/Phát sinh/Cuối kỳ × Nợ/Có), bóc chiều đảo
per-đối-tượng và ghi thành report_type MỚI (ADDITIVE — không đụng PTHU/PTRA cũ).

Thuộc tính dataset_id/ngay/khoi/cong_ty được LẤY THEO twin đã nạp (PTRA/PTHU cùng source_file+
period) -> đảm bảo scoping/khớp filter y hệt bản gốc. Idempotent: xoá ADV cũ cùng source_file
trước khi ghi.

Chạy (dry-run in tổng, KHÔNG ghi):
  .venv/bin/python scripts/derive_congno_advance.py <file.xlsx> --period 2026-06
Ghi thật:
  .venv/bin/python scripts/derive_congno_advance.py <file.xlsx> --period 2026-06 --write
"""
import argparse
import json
import os
import sys
import unicodedata

import openpyxl
import psycopg

DB_URL = (os.environ.get("DATABASE_URL") or os.environ.get("TC_DATABASE_URL")
          or "postgresql://tc:tc@localhost:5433/tc_dashboard")


def _nd(s):
    s = str(s or "").strip().lower()
    return "".join(ch for ch in unicodedata.normalize("NFD", s)
                   if unicodedata.category(ch) != "Mn")


def _num(v):
    return float(v) if isinstance(v, (int, float)) else None


def _find_header(rows):
    """Dò 2 tầng header: dòng nhãn có nhóm 'Cuối kỳ' + dòng dưới có 'Nợ'/'Có'.
    Trả (data_start, {ten, ma, cuoi_no, cuoi_co}). None nếu không nhận."""
    for i, r in enumerate(rows[:15]):
        low = [_nd(c) for c in r]
        j_cuoi = next((j for j, c in enumerate(low) if c.startswith("cuoi ky")), None)
        if j_cuoi is None:
            continue
        side = [_nd(c) for c in rows[i + 1]] if i + 1 < len(rows) else []
        no, co = j_cuoi, j_cuoi + 1
        # xác định đúng cột Nợ/Có trong cụm 'Cuối kỳ' (mặc định Nợ trước Có; hoán nếu ngược)
        if co < len(side) and side[no].startswith("co") and side[co].startswith("no"):
            no, co = co, no
        elif not (no < len(side) and side[no].startswith("no")):
            # header phụ không rõ Nợ/Có -> vẫn giả định [Nợ, Có] theo chuẩn sổ tổng hợp
            pass
        ma = next((j for j, c in enumerate(low) if c.startswith("ma")), 0)
        ten = next((j for j, c in enumerate(low) if c.startswith("ten")), 1)
        return i + 2, {"ma": ma, "ten": ten, "cuoi_no": no, "cuoi_co": co}
    return None, None


def _detect_tk(rows):
    """Nhận diện tài khoản 131/331 từ dòng 'Tài khoản: 3xx' đầu sheet."""
    for r in rows[:8]:
        for c in r:
            t = _nd(c)
            if t.startswith("tai khoan"):
                if "331" in t:
                    return "331"
                if "131" in t:
                    return "131"
    return None


def _entities(rows, data_start, cols):
    """Sinh (ma, ten, cuoi_no, cuoi_co) cho mỗi dòng ĐỐI TƯỢNG (bỏ dòng 'Tổng'/không mã)."""
    out = []
    for r in rows[data_start:]:
        ma = r[cols["ma"]] if cols["ma"] < len(r) else None
        ten = r[cols["ten"]] if cols["ten"] < len(r) else None
        ma_s = str(ma).strip() if ma is not None else ""
        ten_s = str(ten).strip() if ten is not None else ""
        if not ma_s and not ten_s:
            continue
        if _nd(ma_s).startswith("tong") or _nd(ten_s).startswith("tong"):
            continue          # dòng tổng cộng
        if not ma_s:
            continue          # phải có mã đối tượng mới là 1 NCC/KH thực
        cn = _num(r[cols["cuoi_no"]]) if cols["cuoi_no"] < len(r) else None
        cc = _num(r[cols["cuoi_co"]]) if cols["cuoi_co"] < len(r) else None
        out.append((ma_s, ten_s, cn, cc))
    return out


def _twin_attrs(cur, source_file, period, twin_rt):
    """Lấy dataset_id/ngay/khoi/cong_ty từ dòng đã nạp cùng source_file+period (report_type thuận)."""
    cur.execute(
        "SELECT dataset_id, ngay, khoi, cong_ty FROM raw_rows "
        "WHERE source_file=%s AND period_month=%s AND report_type=%s LIMIT 1",
        (source_file, period, twin_rt))
    return cur.fetchone()


def derive(path, period, write=False, source_file=None):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    names = {_nd(s): s for s in wb.sheetnames}
    conn = psycopg.connect(DB_URL)
    cur = conn.cursor()
    out = {"file": os.path.basename(path), "period": period, "write": write, "blocks": []}

    # source_file: dùng đúng nhãn đã nạp (thư mục::tên) nếu không truyền vào -> suy từ DB theo basename
    base = os.path.basename(path)
    if not source_file:
        cur.execute("SELECT DISTINCT source_file FROM raw_rows WHERE source_file LIKE %s AND period_month=%s",
                    ("%" + base, period))
        hits = [r[0] for r in cur.fetchall()]
        source_file = hits[0] if len(hits) == 1 else None
    out["source_file"] = source_file

    # (báo cáo, sheet-hint, twin report_type thuận, report_type ADV, chiều lấy, tên chỉ tiêu)
    JOBS = [
        ("331", "phai tra", "PTRA", "PTRA_ADV", "cuoi_no", "Trả trước NCC"),
        ("131", "phai thu", "PTHU", "PTHU_ADV", "cuoi_co", "Người mua trả tiền trước"),
    ]

    for tk, hint, twin_rt, adv_rt, side, label in JOBS:
        sn = next((names[k] for k in names if hint in k), None)
        if not sn:
            out["blocks"].append({"tk": tk, "skip": f"không thấy sheet '{hint}'"})
            continue
        rows = [list(r) for r in wb[sn].iter_rows(values_only=True)]
        if _detect_tk(rows) not in (None, tk):
            out["blocks"].append({"tk": tk, "skip": f"sheet '{sn}' không phải TK {tk}"})
            continue
        data_start, cols = _find_header(rows)
        if data_start is None:
            out["blocks"].append({"tk": tk, "skip": "không dò được header Cuối kỳ Nợ/Có"})
            continue

        recs = []
        for ma, ten, cn, cc in _entities(rows, data_start, cols):
            val = cn if side == "cuoi_no" else cc     # chiều ĐẢO cần bóc
            if not val or abs(val) < 1:               # bỏ 0 / nhiễu < 1đ
                continue
            recs.append({"ma": ma, "ten": ten or ma, "ty": round(val * 1e-9, 9)})
        tong = round(sum(r["ty"] for r in recs), 6)

        blk = {"tk": tk, "adv_rt": adv_rt, "label": label, "sheet": sn,
               "so_doi_tuong": len(recs), "tong_ty": tong,
               "top": sorted(recs, key=lambda x: -x["ty"])[:5]}

        if write and recs:
            if not source_file:
                blk["error"] = "thiếu source_file (không map được twin) -> BỎ ghi"
                out["blocks"].append(blk)
                continue
            attrs = _twin_attrs(cur, source_file, period, twin_rt)
            if not attrs:
                blk["error"] = f"không thấy twin {twin_rt} cùng source_file+period -> BỎ ghi"
                out["blocks"].append(blk)
                continue
            dataset_id, ngay, khoi, cong_ty = attrs
            # idempotent: xoá ADV cũ cùng source_file+period
            cur.execute("DELETE FROM raw_rows WHERE source_file=%s AND period_month=%s AND report_type=%s",
                        (source_file, period, adv_rt))
            for k, r in enumerate(recs):
                cur.execute(
                    "INSERT INTO raw_rows (dataset_id, report_type, row_index, ngay, cong_ty, khoi, "
                    "cost_center, period_month, amount, amount2, dim1, dim2, dim3, payload, source_file) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    (dataset_id, adv_rt, 6000000 + k, ngay, cong_ty, khoi, None, period,
                     r["ty"], None, r["ten"], None, None,
                     json.dumps({"ma_dt": r["ma"], "unit": "ty", "nguon": f"TK{tk} dư {side}"},
                                ensure_ascii=False), source_file))
            blk["written"] = len(recs)
        out["blocks"].append(blk)

    if write:
        conn.commit()
    conn.close()
    wb.close()
    return out


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--period", required=True)
    ap.add_argument("--source-file", default=None, help="nhãn source_file đã nạp (mặc định suy từ DB)")
    ap.add_argument("--write", action="store_true", help="ghi DB (mặc định dry-run)")
    a = ap.parse_args()
    print(json.dumps(derive(a.file, a.period, a.write, a.source_file), ensure_ascii=False, indent=2))
