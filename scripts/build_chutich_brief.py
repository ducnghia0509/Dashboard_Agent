# -*- coding: utf-8 -*-
"""Dựng BRIEF_CHUTICH.xlsx — số liệu tính SẴN cho agent qa_chutich trả 71 câu của Chủ tịch.

Vì sao có file này: agent chỉ đọc Excel (quyết định 07/08 chưa mở sql_query). Nếu để agent tự
mở hàng chục file nguồn mỗi lần được hỏi thì vừa chậm vừa dễ lệch số. Tầng này tính tất định
bằng Python — KHÔNG có LLM — nên số không bao giờ trôi; agent chỉ việc đọc và diễn giải.

Đặt ở `Tài liệu/` vì đó là 1 trong 3 thư mục `qa_server._allowed_read_dirs()` cho phép
`source_inspect` đọc, và là thư mục itadmin ghi được (khác `~/.openclaw` của sysadmin). Nhờ vậy
vòng làm mới không cần docker exec, không cần tool MCP mới, không cần DB.

Số liệu lấy bằng cách chạy LẠI chính các extractor production ở chế độ offline (xem
`chutich_sources.py`) — không viết lại logic đọc Excel, nên brief và dashboard cùng một nguồn sự thật.

Chạy:
    .venv/bin/python scripts/build_chutich_brief.py --full
    .venv/bin/python scripts/build_chutich_brief.py --only L5_CONGNO L10_NGHIVAN
"""
import argparse
import json
import os
import shutil
import statistics
import sys
import tempfile
import traceback
from collections import defaultdict
from datetime import datetime, timedelta, timezone

import openpyxl

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.normpath(os.path.join(_HERE, ".."))
for _p in (_ROOT, _HERE):
    if _p not in sys.path:
        sys.path.insert(0, _p)

from scripts import chutich_sources as CS  # noqa: E402
from servers.common import be_bridge as bb  # noqa: E402

OUT_PATH = os.path.normpath(os.path.join(_ROOT, "..", "Tài liệu", "BRIEF_CHUTICH.xlsx"))
INBOX_EVENTS = os.path.normpath(
    os.path.join(_ROOT, "..", "Connect_VPS", "inbox_events.jsonl"))

TZ = timezone(timedelta(hours=7))          # giờ VN — mốc as_of Chủ tịch đọc
STALE_AFTER_HOURS = 2

# Thư mục nguồn -> (tên đơn vị hiển thị, mã pháp nhân, khối). Nguồn: sơ đồ tổ chức trong
# agents/qa/SKILL.md + `Tài liệu/Các khối và công ty.xlsx`. Giữ TƯỜNG MINH ở đây thay vì suy
# từ tên file: nhiều nguồn khác nhau nộp cùng mã trong tên file (3 HTX đều là 'B.6.XVP...').
FOLDER_MAP = {
    "SRVF":              ("Showroom Vinfast", "TC", "Vinfast - Showroom"),
    "XDV":               ("Xưởng dịch vụ Vinfast", "TC", "Vinfast - XDV"),
    "DUAN":              ("Khối Dự án", "TC", "Dự án"),
    "TRAMSAC":           ("Trạm sạc Vgreen", "TC", "Trạm sạc Vgreen"),
    "HO":                ("Hỗ trợ tập đoàn", "TC", "hỗ trợ tập đoàn"),
    "HUNGTHINH":         ("Hưng Thịnh (xe tải)", "HT", "Xe tải"),
    "GLOBALAI":          ("Global AI", "GA", "Công nghệ"),
    "ANTAXI":            ("Dịch vụ An Taxi", "AAG", "Dịch vụ An Taxi"),
    "ANKHACHSAN":        ("Dịch vụ An KS", "AAG", "Dịch vụ An KS"),
    "XANHVINHPHUC":      ("Xanh Vĩnh Phúc", "XVP", "Vận tải Taxi Xanh"),
    "HTXXANHTUYENQUANG": ("HTX Xanh Tuyên Quang", "HTX_XTQ", "Vận tải Taxi Xanh"),
    "HTXXANHVINHPHUC":   ("HTX Xanh Vĩnh Phúc", "HTX_XVP", "Vận tải Taxi Xanh"),
}

# report_type nào về thì phải dựng lại sheet nào (watcher dùng bảng này).
REPORT_TYPE_SHEETS = {
    "baocaohqkdngay":       ["L1_CANHBAO", "L2_DOANHTHU", "L3_LOINHUAN", "L7_CHIPHI",
                             "L8_XEPHANG", "L9_DUAN"],
    "baocaotaichinhrieng":  ["L8_XEPHANG", "L9_DUAN", "L3_LOINHUAN"],
    "baocaotuoino":         ["L5_CONGNO", "L10_NGHIVAN"],
    "baocaotuoinophaithu":  ["L5_CONGNO", "L10_NGHIVAN"],
    "baocaothuchi":         ["L4_DONGTIEN", "L1_CANHBAO"],
    "baocaonganhang":       ["L4_DONGTIEN"],
    "baocaokehoachdoanhthu": ["L2_DOANHTHU", "L3_LOINHUAN"],
}

# 10 câu KHÔNG có nguồn dữ liệu -> ủy quyền. Đồng bộ với eval/chutich/questions.json.
ROUTING_ROWS = [
    ("L6-02", "Bao nhiêu tồn kho chậm luân chuyển?", "CUVT (Cung ứng - Vật tư) + TCKT",
     "Báo cáo tồn kho chi tiết theo mặt hàng, kèm ngày nhập kho và số lượng xuất trong kỳ"),
    ("L6-03", "Hàng nào tồn trên 90 ngày?", "CUVT (Cung ứng - Vật tư) + TCKT",
     "Báo cáo tồn kho chi tiết theo mặt hàng, kèm ngày nhập kho"),
    ("L6-04", "Bao nhiêu tiền đang nằm chết trong kho?", "CUVT (Cung ứng - Vật tư) + TCKT",
     "Tồn kho chi tiết theo mặt hàng + tuổi tồn + giá trị ghi sổ"),
    ("L6-05", "Nếu bán thanh lý tồn kho sẽ thu về bao nhiêu?", "CUVT (Cung ứng - Vật tư) + TCKT",
     "Tồn kho chi tiết + giá trị thuần có thể thực hiện được (NRV) hoặc giá thanh lý tham chiếu"),
    ("L7-03", "Đơn vị nào vượt ngân sách?", "TCKT",
     "Kế hoạch CHI PHÍ theo đơn vị/khối (hiện chỉ có kế hoạch doanh thu 0.KH.GR.Y.2026)"),
    ("L7-04", "Chi phí nào chưa có chứng từ?", "TCKT",
     "Trạng thái chứng từ theo bút toán (đã/chưa có hóa đơn, chứng từ gốc)"),
    ("L9-05", "Máy nào gây chi phí lớn nhất?", "KDVH (khối Dự án / Xe tải)",
     "Chi phí theo đầu máy/thiết bị (mã máy, giờ vận hành, chi phí sửa chữa - nhiên liệu)"),
    ("L9-06", "Nhiên liệu vượt định mức ở đâu?", "KDVH (khối Dự án / Xe tải)",
     "Bảng định mức nhiên liệu theo đầu xe/máy + sản lượng nhiên liệu thực dùng"),
    ("L10-03", "Có thanh toán nào vượt quyền phê duyệt không?", "KSNB",
     "Ma trận hạn mức phê duyệt theo chức danh + log duyệt chi (ai duyệt, số tiền, thời điểm)"),
    ("L10-08", "Có giao dịch ngoài giờ hành chính?", "KSNB + CNTT",
     "Sổ giao dịch có TIMESTAMP (giờ:phút). Dữ liệu hiện tại chỉ có ngày, 'Ngày hóa đơn' luôn 00:00"),
]


def _now():
    return datetime.now(TZ)


def _ts(dt=None):
    return (dt or _now()).strftime("%Y-%m-%d %H:%M:%S")


def _norm(s):
    return bb.remove_diacritics(str(s or "")).strip().lower()


def _num(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def _ty(v, nd=3):
    """Làm tròn CHỈ Ở ĐÂY (lúc ghi ra), không round trong lúc tính — quy ước full precision."""
    return round(v, nd) if isinstance(v, (int, float)) else v


# ---- Thu thập dữ liệu thô ---------------------------------------------------------

def _collect_hqkd_ngay(errors, skipped):
    """{folder: {'period','days':{'YYYY-MM-DD': {doanh_thu_ty, chi_phi_ty, lntt_ty}}}}

    Đơn vị nào deriver trả ok=False thì GHI VÀO `skipped` kèm lý do, không bỏ im lặng — brief
    thiếu 1 đơn vị mà không nói ra thì Chủ tịch sẽ đọc bảng xếp hạng như thể đã đủ mặt.
    """
    from scripts import derive_hqkd_ngay as m
    out = {}
    for folder, e in CS.latest_per_folder("baocaohqkdngay").items():
        try:
            res, _ = CS.capture(m.derive, e["path"], write=False)
            if res.get("ok") and res.get("tong_theo_ngay"):
                out[folder] = {"period": res.get("period"), "file": e["file"],
                               "days": res["tong_theo_ngay"]}
            else:
                skipped.append((folder, e["file"],
                                str(res.get("error") or res.get("skip") or "không rõ")[:200]))
        except Exception as ex:
            errors.append(f"hqkd_ngay/{folder}: {type(ex).__name__}: {ex}")
    return out


def _norm_aging(res):
    """Chuẩn hoá 2 schema tuổi nợ khác nhau của deriver về cùng một hình.

    mode 'hanno*' -> có ngày đến hạn: tong/trong_han/qua_han/qh_1_30/qh_30_90/qh_90_180/qh_180p
    mode 'age'    -> chỉ có TUỔI nợ:  b1 (<1 tháng) / b13 (1-3) / b36 (3-6) / b6p (>6 tháng)

    Nguồn 'age' KHÔNG có ngày đến hạn nên KHÔNG suy ra được 'quá hạn' — trả None chứ không
    đoán bằng 0 (đọc nhầm 0 thành 'không có nợ quá hạn' là sai nguy hiểm). Riêng '>180 ngày'
    thì hai schema tương đương: qh_180p ↔ b6p (đều là trên 6 tháng).
    """
    agg = res.get("agg_ty") or {}
    tong = res.get("tong_ty")
    if "qua_han" in agg:
        return {"tong": tong if tong is not None else agg.get("tong"),
                "trong_han": agg.get("trong_han"), "den_han": agg.get("den_han"),
                "qua_han": agg.get("qua_han"), "qh_180p": agg.get("qh_180p"),
                "chi_tiet": (f"trong hạn {_ty(agg.get('trong_han'))} · quá hạn "
                             f"{_ty(agg.get('qua_han'))} · 1-30 {_ty(agg.get('qh_1_30'))} · "
                             f"30-90 {_ty(agg.get('qh_30_90'))} · 90-180 "
                             f"{_ty(agg.get('qh_90_180'))} · >180 {_ty(agg.get('qh_180p'))}"),
                "schema": "han_no"}
    return {"tong": tong, "trong_han": None, "den_han": None,
            "qua_han": None, "qh_180p": agg.get("b6p"),
            "chi_tiet": (f"<1 tháng {_ty(agg.get('b1'))} · 1-3 tháng {_ty(agg.get('b13'))} · "
                         f"3-6 tháng {_ty(agg.get('b36'))} · >6 tháng {_ty(agg.get('b6p'))} "
                         f"— nguồn chỉ có TUỔI nợ, không có ngày đến hạn nên không tính được "
                         f"'quá hạn'"),
            "schema": "tuoi_no"}


def _collect_aging(errors):
    """{folder: {'period','agg' (đã chuẩn hoá), 'mode', 'file', 'path'}} — đơn vị tỷ."""
    from scripts import derive_congno_tuoino as m
    out = {}
    for folder, e in CS.latest_per_folder("baocaotuoino").items():
        try:
            res, _ = CS.capture(m.derive, e["path"], e["period"], write=False)
            if res.get("agg_ty") or res.get("tong_ty") is not None:
                out[folder] = {"period": res.get("period") or e["period"], "file": e["file"],
                               "agg": _norm_aging(res), "mode": res.get("mode"),
                               "n_rows": res.get("n_rows"), "path": e["path"]}
        except Exception as ex:
            errors.append(f"aging/{folder}: {type(ex).__name__}: {ex}")
    return out


def _collect_sodu_tien(errors, notes):
    """Số dư tiền/vay theo pháp nhân từ Báo cáo Tiền tập đoàn (1 file, nhiều pháp nhân).

    KHÔNG lấy thẳng file kỳ mới nhất: đầu tháng, file kỳ mới đã có mặt nhưng cột 'ĐẾN NGÀY
    HIỆN TẠI' còn để trống/ghi 0, extract ra số dư = 0 cho TOÀN BỘ pháp nhân. Trả 0 cho câu
    'tiền đang nằm ở đâu' là sai nguy hiểm hơn nhiều so với trả số của kỳ liền trước. Vì vậy
    duyệt từ kỳ mới nhất lùi dần, lấy kỳ ĐẦU TIÊN có số khác 0, và ghi rõ đã dùng kỳ nào.
    """
    from scripts import extract_sodu_tien as m
    cands = sorted((e for e in CS.discover(report_type="baocaothuchi") if e["period"]),
                   key=lambda e: e["period"], reverse=True)
    if not cands:
        errors.append("sodu_tien: không thấy file baocaothuchi")
        return [], {}
    skipped_periods = []
    for e in cands[:4]:
        try:
            _res, tpl = CS.capture(m.extract, e["path"], e["period"])
            recs = tpl.get("03B_SODU_TIEN", [])
        except Exception as ex:
            errors.append(f"sodu_tien/{e['period']}: {type(ex).__name__}: {ex}")
            continue
        tong = sum(abs(v) for r in recs for k, v in r.items()
                   if isinstance(v, (int, float)) and k != "Kỳ")
        if recs and tong > 0:
            if skipped_periods:
                notes.append(
                    f"Số dư tiền: kỳ {', '.join(skipped_periods)} có file nhưng toàn số 0 "
                    f"(chưa điền cột 'đến ngày hiện tại') — brief dùng kỳ {e['period']}")
            return recs, {"file": e["file"], "period": e["period"]}
        skipped_periods.append(e["period"])
    notes.append(f"Số dư tiền: cả {len(skipped_periods)} kỳ gần nhất "
                 f"({', '.join(skipped_periods)}) đều ra 0 — KHÔNG có số dư tiền để báo cáo")
    return [], {}


_DETAIL_SHEET_HINTS = ("chi tiet theo hop dong", "chi tiet doi soat", "chi tiet cong no",
                       "chi tiet")

# Hai nguồn lớn nhất dùng HAI schema cột khác nhau cho cùng một loại dữ liệu:
#   SRVF (showroom): 'Mã khách' / 'Tên kh'  / 'Số hóa đơn' / 'Số khung'    / 'số ngày quá hạn'
#   XDV  (xưởng DV): 'Mã đối tượng'/'Tên VAT'/'Số HĐ'     / 'Biển số xe'  / 'Tuổi nợ'
# Khoá theo alias thay vì tên cứng để không mất nguồn XDV (17k dòng, đơn vị lớn thứ 2).
_DETAIL_ALIASES = {
    "ma_kh":        ("ma khach", "ma doi tuong", "ma kh"),
    "ten_kh":       ("ten kh", "ten khach hang", "ten vat", "ten doi tuong"),
    "don_vi":       ("ten don vi", "don vi"),
    "no_cuoi":      ("no cuoi ky",),
    "co_cuoi":      ("co cuoi ky",),
    "ps_no":        ("phat sinh no",),
    "so_hd":        ("so hoa don", "so hd", "so hddt"),
    "ngay_hd":      ("ngay hoa don", "ngay hd"),
    "so_khung":     ("so khung", "bien so xe"),
    "ngay_qua_han": ("so ngay qua han", "tuoi no"),
    "nhan_vien":    ("ten nhan vien", "co van dich vu"),
}
# Header hợp lệ phải có định danh khách + ít nhất 1 cột số dư — tránh nhận nhầm dòng tiêu đề phụ.
_DETAIL_REQUIRED = (("ma khach", "ma doi tuong", "ma kh"), ("no cuoi ky", "co cuoi ky"))


def _read_congno_detail(path, limit_rows=40000):
    """Đọc sheet chi tiết công nợ -> list dict theo khách hàng/hóa đơn.

    Đây là chỗ DUY NHẤT trong toàn bộ nguồn có dữ liệu tới mức KHÁCH HÀNG + SỐ HÓA ĐƠN +
    SỐ KHUNG/BIỂN SỐ, nên nó gánh cả L5 (top khách nợ) lẫn L10 (trùng hóa đơn/trùng xe).
    Deriver aging không dùng được cho việc này vì nó đã gộp sẵn theo dải tuổi nợ.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = None
        for hint in _DETAIL_SHEET_HINTS:            # ưu tiên hint cụ thể trước hint chung
            for cand in wb.worksheets:
                if hint in _norm(cand.title):
                    ws = cand
                    break
            if ws is not None:
                break
        if ws is None:
            return []
        hdr_idx, out = None, []
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i > limit_rows:
                break
            if hdr_idx is None:
                cells = [_norm(c) for c in r]
                if all(any(a in cells for a in grp) for grp in _DETAIL_REQUIRED):
                    pos = {name: j for j, name in enumerate(cells) if name}
                    hdr_idx = {}
                    for key, aliases in _DETAIL_ALIASES.items():
                        for a in aliases:
                            if a in pos:
                                hdr_idx[key] = pos[a]
                                break
                continue
            g = lambda k: (r[hdr_idx[k]] if k in hdr_idx and hdr_idx[k] < len(r) else None)  # noqa: E731
            ma_kh = g("ma_kh")
            if not ma_kh:
                continue
            out.append({
                "ma_kh": str(ma_kh).strip(),
                "ten_kh": str(g("ten_kh") or "").strip(),
                "don_vi": str(g("don_vi") or "").strip(),
                "no_cuoi": _num(g("no_cuoi")) or 0.0,
                "co_cuoi": _num(g("co_cuoi")) or 0.0,
                "ps_no": _num(g("ps_no")) or 0.0,
                "so_hd": str(g("so_hd") or "").strip(),
                "ngay_hd": g("ngay_hd"),
                "so_khung": str(g("so_khung") or "").strip(),
                "ngay_qua_han": _num(g("ngay_qua_han")),
                "nhan_vien": str(g("nhan_vien") or "").strip(),
            })
        return out
    finally:
        wb.close()


def _collect_congno_detail(aging, errors):
    """{folder: [rec,...]} — chỉ những nguồn có sheet chi tiết (SRVF/XDV... có, DUAN không)."""
    out = {}
    for folder, info in aging.items():
        try:
            det = _read_congno_detail(info["path"])
            if det:
                out[folder] = det
        except Exception as ex:
            errors.append(f"congno_detail/{folder}: {type(ex).__name__}: {ex}")
    return out


# ---- Bộ dựng từng sheet -----------------------------------------------------------

def _mtd(days):
    """Cộng dồn doanh thu/chi phí/LNTT của cả kỳ từ dict theo ngày."""
    dt = sum((d.get("doanh_thu_ty") or 0) for d in days.values())
    cp = sum((d.get("chi_phi_ty") or 0) for d in days.values())
    ln = sum((d.get("lntt_ty") or 0) for d in days.values())
    return dt, cp, ln


def _latest_day(days):
    ks = sorted(k for k in days if days[k])
    return ks[-1] if ks else None


def build_L1_CANHBAO(ctx):
    """Bất thường hôm nay = ngày mới nhất lệch quá 2 lần độ lệch chuẩn so các ngày còn lại."""
    rows = [("Mức", "Đơn vị", "Chỉ tiêu", "Ngày", "Giá trị (tỷ)",
             "Trung bình các ngày trước (tỷ)", "Lệch (tỷ)", "Lệch (%)", "Nguồn")]
    items = []
    for folder, info in ctx["hqkd"].items():
        days = info["days"]
        last = _latest_day(days)
        if not last:
            continue
        ten = FOLDER_MAP.get(folder, (folder, "", ""))[0]
        for field, label in (("doanh_thu_ty", "Doanh thu"), ("chi_phi_ty", "Chi phí"),
                             ("lntt_ty", "LNTT")):
            hist = [(days[d].get(field) or 0) for d in sorted(days) if d != last]
            cur = days[last].get(field) or 0
            if len(hist) < 2:
                continue
            avg = statistics.fmean(hist)
            sd = statistics.pstdev(hist)
            lech = cur - avg
            if sd > 0 and abs(lech) >= 2 * sd and abs(lech) > 0.05:
                muc = "ĐỎ" if abs(lech) >= 3 * sd else "VÀNG"
                pct = (lech / avg * 100) if avg else None
                items.append((abs(lech), (muc, ten, label, last, _ty(cur), _ty(avg), _ty(lech),
                                          _ty(pct, 1) if pct is not None else "",
                                          info["file"])))
    items.sort(key=lambda x: -x[0])
    rows += [it[1] for it in items[:40]]
    if len(rows) == 1:
        rows.append(("—", "—", "Không có chỉ tiêu nào lệch quá 2 độ lệch chuẩn", "", "", "", "", "", ""))
    return rows


def build_L2_DOANHTHU(ctx):
    rows = [("Đơn vị", "Pháp nhân", "Khối", "Kỳ", "Doanh thu kỳ (tỷ)",
             "Doanh thu ngày mới nhất (tỷ)", "Ngày mới nhất", "% đóng góp toàn nhóm", "Nguồn")]
    tot = 0.0
    tmp = []
    for folder, info in ctx["hqkd"].items():
        dt, _cp, _ln = _mtd(info["days"])
        last = _latest_day(info["days"])
        ten, cty, khoi = FOLDER_MAP.get(folder, (folder, "", ""))
        tot += dt
        tmp.append((dt, ten, cty, khoi, info["period"],
                    (info["days"].get(last, {}).get("doanh_thu_ty") if last else None),
                    last, info["file"]))
    tmp.sort(key=lambda x: -x[0])
    for dt, ten, cty, khoi, per, dlast, last, f in tmp:
        rows.append((ten, cty, khoi, per, _ty(dt), _ty(dlast), last,
                     _ty(dt / tot * 100, 1) if tot else "", f))
    rows.append(("TỔNG", "", "", "", _ty(tot), "", "", "100.0",
                 f"Σ {len(tmp)} đơn vị CÓ số báo cáo ngày"))
    for folder, f, ly_do in ctx["skipped"]:
        ten = FOLDER_MAP.get(folder, (folder, "", ""))[0]
        rows.append((ten, FOLDER_MAP.get(folder, ("", "", ""))[1],
                     FOLDER_MAP.get(folder, ("", "", ""))[2], "", "KHÔNG CÓ SỐ", "", "",
                     "", f"{f} — {ly_do}"))
    return rows


def build_L3_LOINHUAN(ctx):
    rows = [("Đơn vị", "Pháp nhân", "Kỳ", "Doanh thu (tỷ)", "Chi phí (tỷ)", "LNTT (tỷ)",
             "Biên LNTT (%)", "Nếu cắt 10% chi phí → LNTT (tỷ)", "Nguồn")]
    tmp = []
    for folder, info in ctx["hqkd"].items():
        dt, cp, ln = _mtd(info["days"])
        ten, cty, _k = FOLDER_MAP.get(folder, (folder, "", ""))
        tmp.append((ln, ten, cty, info["period"], dt, cp, ln, info["file"]))
    tmp.sort(key=lambda x: -x[0])
    T = [0.0, 0.0, 0.0]
    for _s, ten, cty, per, dt, cp, ln, f in tmp:
        T[0] += dt
        T[1] += cp
        T[2] += ln
        rows.append((ten, cty, per, _ty(dt), _ty(cp), _ty(ln),
                     _ty(ln / dt * 100, 1) if dt else "", _ty(ln + cp * 0.1), f))
    rows.append(("TỔNG", "", "", _ty(T[0]), _ty(T[1]), _ty(T[2]),
                 _ty(T[2] / T[0] * 100, 1) if T[0] else "", _ty(T[2] + T[1] * 0.1),
                 "Σ các đơn vị có báo cáo ngày"))
    return rows


def build_L7_CHIPHI(ctx):
    """Chi phí theo đơn vị + đánh dấu 'chi phí tăng nhưng doanh thu không tăng'."""
    rows = [("Đơn vị", "Kỳ", "Chi phí kỳ (tỷ)", "Doanh thu kỳ (tỷ)",
             "Chi phí ngày mới nhất (tỷ)", "TB chi phí/ngày trước đó (tỷ)",
             "Cờ: CP tăng - DT không tăng", "Cờ: CP > 0 nhưng DT = 0", "Nguồn")]
    tmp = []
    for folder, info in ctx["hqkd"].items():
        days = info["days"]
        last = _latest_day(days)
        if not last:
            continue
        dt, cp, _ln = _mtd(days)
        cp_last = days[last].get("chi_phi_ty") or 0
        dt_last = days[last].get("doanh_thu_ty") or 0
        hist_cp = [(days[d].get("chi_phi_ty") or 0) for d in sorted(days) if d != last]
        hist_dt = [(days[d].get("doanh_thu_ty") or 0) for d in sorted(days) if d != last]
        avg_cp = statistics.fmean(hist_cp) if hist_cp else 0
        avg_dt = statistics.fmean(hist_dt) if hist_dt else 0
        co1 = "CÓ" if (cp_last > avg_cp and dt_last <= avg_dt) else ""
        co2 = "CÓ" if (cp > 0 and dt <= 0) else ""
        ten = FOLDER_MAP.get(folder, (folder, "", ""))[0]
        tmp.append((cp, ten, info["period"], cp, dt, cp_last, avg_cp, co1, co2, info["file"]))
    tmp.sort(key=lambda x: -x[0])
    for _s, ten, per, cp, dt, cl, ac, co1, co2, f in tmp:
        rows.append((ten, per, _ty(cp), _ty(dt), _ty(cl), _ty(ac), co1, co2, f))
    return rows


def build_L8_XEPHANG(ctx):
    """Xếp hạng theo PHÁP NHÂN (gộp các đơn vị cùng mã công ty)."""
    rows = [("Hạng", "Pháp nhân", "Đơn vị gộp", "Doanh thu (tỷ)", "Chi phí (tỷ)", "LNTT (tỷ)",
             "Biên LNTT (%)", "Số dư tiền (tỷ)", "Dư vay (tỷ)", "Công nợ phải thu (tỷ)")]
    agg = defaultdict(lambda: {"dt": 0.0, "cp": 0.0, "ln": 0.0, "units": []})
    for folder, info in ctx["hqkd"].items():
        dt, cp, ln = _mtd(info["days"])
        ten, cty, _k = FOLDER_MAP.get(folder, (folder, folder, ""))
        a = agg[cty]
        a["dt"] += dt
        a["cp"] += cp
        a["ln"] += ln
        a["units"].append(ten)
    tien = {r.get("Đơn vị"): r for r in ctx["sodu"]}
    congno = defaultdict(float)
    for folder, info in ctx["aging"].items():
        cty = FOLDER_MAP.get(folder, (folder, folder, ""))[1]
        congno[cty] += (info["agg"].get("tong") or 0)
    ordered = sorted(agg.items(), key=lambda kv: -kv[1]["ln"])
    for i, (cty, a) in enumerate(ordered, 1):
        t = tien.get(cty, {})
        tien_ty = (t.get("Tiền mặt (tỷ)") or 0) + (t.get("Tiền gửi NH (tỷ)") or 0)
        rows.append((i, cty, " + ".join(sorted(a["units"])), _ty(a["dt"]), _ty(a["cp"]),
                     _ty(a["ln"]), _ty(a["ln"] / a["dt"] * 100, 1) if a["dt"] else "",
                     _ty(tien_ty), _ty(t.get("Số dư tiền vay (tỷ) — đối chiếu 04_VAY") or 0),
                     _ty(congno.get(cty, 0))))
    return rows


def build_L9_DUAN(ctx):
    rows = [("Nguồn/Khối", "Kỳ", "Doanh thu (tỷ)", "Chi phí (tỷ)", "LNTT (tỷ)",
             "Tỷ lệ chi phí/doanh thu (%)", "Trạng thái", "Nguồn")]
    for folder in ("DUAN", "TRAMSAC", "XDV", "SRVF"):
        info = ctx["hqkd"].get(folder)
        if not info:
            continue
        dt, cp, ln = _mtd(info["days"])
        ten = FOLDER_MAP.get(folder, (folder, "", ""))[0]
        rows.append((ten, info["period"], _ty(dt), _ty(cp), _ty(ln),
                     _ty(cp / dt * 100, 1) if dt else "",
                     "LỖ" if ln < 0 else "Lãi", info["file"]))
    rows.append(("", "", "", "", "", "", "", ""))
    rows.append(("GHI CHÚ", "Báo cáo ngày gộp theo NGUỒN NỘP, chưa tách được 7 dự án "
                 "(CB/LS/QS/YB/TT/PQ/TC_DA). Muốn tách từng dự án phải đọc BCTC tháng "
                 "khối Dự án theo cost center.", "", "", "", "", "", ""))
    return rows


def build_L4_DONGTIEN(ctx):
    rows = [("Mục", "Pháp nhân/Chi tiết", "Giá trị (tỷ)", "Ghi chú", "Nguồn")]
    src = ctx["sodu_meta"].get("file", "")
    ky_tien = ctx["sodu_meta"].get("period")
    rows.append(("KỲ SỐ LIỆU TIỀN", ky_tien or "KHÔNG CÓ", "",
                 "Số dư tiền/vay dưới đây là của kỳ này — phải nói rõ khi trả lời, "
                 "đừng để Chủ tịch hiểu là số hôm nay", src))
    tm = tg = vay = bl = 0.0
    for r in ctx["sodu"]:
        cty = r.get("Đơn vị")
        a = r.get("Tiền mặt (tỷ)") or 0
        b = r.get("Tiền gửi NH (tỷ)") or 0
        v = r.get("Số dư tiền vay (tỷ) — đối chiếu 04_VAY") or 0
        g = r.get("Bảo lãnh thanh toán (tỷ)") or 0
        tm += a
        tg += b
        vay += v
        bl += g
        rows.append(("Tiền theo pháp nhân", cty, _ty(a + b),
                     f"tiền mặt {_ty(a)} + gửi NH {_ty(b)}; dư vay {_ty(v)}", src))
    rows.append(("TIỀN — TỔNG", "Tiền mặt + tiền gửi", _ty(tm + tg),
                 f"tiền mặt {_ty(tm)} · gửi NH {_ty(tg)}", src))
    rows.append(("VAY — TỔNG", "Dư nợ vay", _ty(vay), f"bảo lãnh thanh toán {_ty(bl)}", src))

    tong_cn = sum((i["agg"].get("tong") or 0) for i in ctx["aging"].values())
    han = [i for i in ctx["aging"].values() if i["agg"]["schema"] == "han_no"]
    qua_han = sum((i["agg"].get("qua_han") or 0) for i in han)
    den_han = sum((i["agg"].get("den_han") or 0) for i in han)
    rows.append(("TIỀN KẸT — công nợ phải thu", "Tổng phải thu", _ty(tong_cn),
                 f"quá hạn {_ty(qua_han)} (chỉ đo được trên {len(han)}/{len(ctx['aging'])} "
                 f"đơn vị có ngày đến hạn)", "Σ baocaotuoino các đơn vị"))
    rows.append(("SẼ VỀ 30 NGÀY", "Khoản đến hạn", _ty(den_han),
                 f"Chỉ gồm khoản có ngày đến hạn trong nguồn ({len(han)} đơn vị); "
                 f"KHÔNG phải dự báo dòng tiền", "Σ baocaotuoino các đơn vị"))
    rows.append(("NẾU THU HẾT CÔNG NỢ", "Tiền sau khi thu", _ty(tm + tg + tong_cn),
                 f"= tiền hiện có {_ty(tm + tg)} + phải thu {_ty(tong_cn)}", "tính từ 2 dòng trên"))
    rows.append(("NẾU THU 30% CÔNG NỢ", "Tiền sau khi thu", _ty(tm + tg + tong_cn * 0.3),
                 f"+{_ty(tong_cn * 0.3)}", "tính từ 2 dòng trên"))
    rows.append(("", "", "", "", ""))
    rows.append(("GIỚI HẠN", "Tồn kho / đầu tư / tài sản", "",
                 "Không có báo cáo tồn kho chi tiết; tồn kho chỉ có 1 dòng tổng trên CĐKT. "
                 "Câu hỏi tồn kho chi tiết -> xem sheet ROUTING.", ""))
    return rows


def build_L5_CONGNO(ctx):
    rows = [("Phần", "Khoá", "Tên", "Đơn vị", "Giá trị (tỷ)", "Ghi chú", "Nguồn")]
    for folder, info in sorted(ctx["aging"].items()):
        a = info["agg"]
        ten = FOLDER_MAP.get(folder, (folder, "", ""))[0]
        rows.append(("Aging theo đơn vị", folder, ten, info["period"], _ty(a.get("tong")),
                     a.get("chi_tiet"), info["file"]))
    tot = sum((i["agg"].get("tong") or 0) for i in ctx["aging"].values())
    q180 = sum((i["agg"].get("qh_180p") or 0) for i in ctx["aging"].values())
    han = [i for i in ctx["aging"].values() if i["agg"]["schema"] == "han_no"]
    qh = sum((i["agg"].get("qua_han") or 0) for i in han)
    tot_han = sum((i["agg"].get("tong") or 0) for i in han)
    n_tuoi = len(ctx["aging"]) - len(han)
    rows.append(("TỔNG", "", "Toàn bộ đơn vị có báo cáo tuổi nợ", "", _ty(tot),
                 f">180 ngày (trên 6 tháng) {_ty(q180)} — tính được cho cả {len(ctx['aging'])} "
                 f"đơn vị", "Σ baocaotuoino"))
    rows.append(("QUÁ HẠN", "", f"Chỉ {len(han)}/{len(ctx['aging'])} đơn vị có ngày đến hạn", "",
                 _ty(qh),
                 f"trên nền {_ty(tot_han)} tỷ của {len(han)} đơn vị đó "
                 f"({_ty(qh / tot_han * 100, 1) if tot_han else 0}%). "
                 f"{n_tuoi} đơn vị còn lại nguồn chỉ có TUỔI nợ nên KHÔNG tính được quá hạn — "
                 f"đừng cộng con số này với tổng toàn tập đoàn", "Σ baocaotuoino"))
    rows.append(("NGUY CƠ MẤT VỐN", "", "Nợ trên 180 ngày / trên 6 tháng", "", _ty(q180),
                 "Hai schema nguồn tương đương nhau ở dải này", "Σ baocaotuoino"))
    rows.append(("", "", "", "", "", "", ""))

    # Top 20 khách nợ — chỉ nguồn nào có sheet chi tiết theo hợp đồng
    agg_kh = defaultdict(lambda: {"no": 0.0, "ten": "", "dv": set(), "qh": 0})
    for folder, det in ctx["detail"].items():
        for r in det:
            k = (folder, r["ma_kh"])
            a = agg_kh[k]
            a["no"] += (r["no_cuoi"] or 0) - (r["co_cuoi"] or 0)
            a["ten"] = a["ten"] or r["ten_kh"]
            if r["don_vi"]:
                a["dv"].add(r["don_vi"])
            if r["ngay_qua_han"]:
                a["qh"] = max(a["qh"], int(r["ngay_qua_han"]))
    top = sorted(agg_kh.items(), key=lambda kv: -kv[1]["no"])[:20]
    if top:
        for i, ((folder, ma), a) in enumerate(top, 1):
            rows.append((f"Top khách nợ #{i}", ma, a["ten"], ", ".join(sorted(a["dv"]))[:60],
                         _ty(a["no"] / 1e9), f"quá hạn tối đa {a['qh']} ngày",
                         ctx["aging"].get(folder, {}).get("file", folder)))
    else:
        rows.append(("Top khách nợ", "—", "Không nguồn nào có sheet chi tiết theo hợp đồng "
                     "trong kỳ này", "", "", "", ""))
    return rows


def build_L10_NGHIVAN(ctx):
    """Nghi vấn kiểm toán — CHỈ báo cái đã loại xong nhiễu cấu trúc.

    Đã đối chiếu dữ liệu thật (SRVF T06, 4.497 dòng / 2.542 VIN) trước khi chốt luật, vì hai
    kiểu "trùng" hiển nhiên nhất lại hầu hết là hợp lệ:

    - Trùng VIN mà CÙNG mã khách: là cặp bút toán đối ứng (Nợ 495.040.000 / Có 495.040.000
      của cùng số hóa đơn) — ghi nhận rồi tất toán, KHÔNG phải ghi hai lần. 20/2.542 VIN trùng
      thì gần như toàn bộ thuộc dạng này.
    - Trùng số hóa đơn mà KHÁC đơn vị: 119/154 trường hợp — mỗi showroom đánh số hóa đơn theo
      dãy RIÊNG nên số trùng nhau giữa các chi nhánh là bình thường.

    Báo cả hai kiểu trên thì Chủ tịch nhận ~134 báo động giả. Chỉ giữ lại: trùng VIN KHÁC
    khách, và trùng số hóa đơn KHÁC khách nhưng CÙNG một đơn vị (19 trường hợp).
    """
    rows = [("Loại nghi vấn", "Mức tin cậy", "Khoá", "Chi tiết", "Số bản ghi",
             "Giá trị ròng (tỷ)", "Nguồn")]
    for folder, det in sorted(ctx["detail"].items()):
        src = ctx["aging"].get(folder, {}).get("file", folder)
        by_hd, by_vin = defaultdict(list), defaultdict(list)
        for r in det:
            if r["so_hd"] and r["so_hd"] not in ("0", "-"):
                by_hd[r["so_hd"]].append(r)
            if r["so_khung"] and len(r["so_khung"]) >= 10:
                by_vin[r["so_khung"]].append(r)

        def _net(recs):
            return _ty(sum((r["no_cuoi"] or 0) - (r["co_cuoi"] or 0) for r in recs) / 1e9)

        vin_khac_kh = [(k, v) for k, v in by_vin.items()
                       if len(v) > 1 and len({r["ma_kh"] for r in v}) > 1]
        for vin, recs in sorted(vin_khac_kh, key=lambda kv: -len(kv[1]))[:15]:
            rows.append(("Trùng SỐ KHUNG, KHÁC khách hàng", "Cần xác minh", vin,
                         "; ".join(sorted({r["ten_kh"] for r in recs}))[:80], len(recs),
                         _net(recs), src))

        hd_nghi = [(k, v) for k, v in by_hd.items()
                   if len(v) > 1 and len({r["ma_kh"] for r in v}) > 1
                   and len({r["don_vi"] for r in v}) == 1]
        for hd, recs in sorted(hd_nghi, key=lambda kv: -len(kv[1]))[:20]:
            rows.append(("Trùng SỐ HÓA ĐƠN trong cùng đơn vị", "Cần xác minh", hd,
                         f"{list({r['don_vi'] for r in recs})[0][:28]} — "
                         + "; ".join(sorted({r["ten_kh"] for r in recs}))[:60],
                         len(recs), _net(recs), src))

        n_vin_cung_kh = sum(1 for v in by_vin.values()
                            if len(v) > 1 and len({r["ma_kh"] for r in v}) == 1)
        n_hd_khac_dv = sum(1 for v in by_hd.values()
                           if len(v) > 1 and len({r["don_vi"] for r in v}) > 1)
        rows.append((f"[{folder}] Đã loại nhiễu", "—", "",
                     f"{n_vin_cung_kh} VIN trùng nhưng cùng khách (cặp bút toán đối ứng) và "
                     f"{n_hd_khac_dv} số hóa đơn trùng giữa các đơn vị khác nhau (mỗi đơn vị "
                     f"đánh số riêng) — đã loại, KHÔNG phải bỏ sót", "", "", src))
        # Dồn ngày hóa đơn vào 3 ngày cuối tháng -> dấu hiệu ghi nhận sớm
        cnt_by_day = defaultdict(int)
        for r in det:
            d = r["ngay_hd"]
            if isinstance(d, datetime):
                cnt_by_day[d.day] += 1
        if cnt_by_day:
            tail = sum(v for k, v in cnt_by_day.items() if k >= 28)
            allc = sum(cnt_by_day.values())
            if allc and tail / allc > 0.35:
                rows.append(("Dồn hóa đơn cuối tháng", "Cần xác minh — có thể do chu kỳ bán hàng",
                             "ngày >= 28", f"{tail}/{allc} hóa đơn rơi vào 3 ngày cuối tháng",
                             tail, "", src))
    if len(rows) == 1:
        rows.append(("—", "", "", "Không phát hiện trùng số khung/số hóa đơn trong kỳ", "", "", ""))
    rows.append(("", "", "", "", "", "", ""))
    rows.append(("PHẠM VI", "", "", "Chỉ quét được phía PHẢI THU (baocaotuoino). Thanh toán cho "
                 "nhà cung cấp, hạn mức phê duyệt và giờ giao dịch KHÔNG có dữ liệu — "
                 "xem sheet ROUTING.", "", "", ""))
    return rows


def build_ROUTING(_ctx):
    rows = [("Mã câu", "Câu hỏi", "Phòng phụ trách", "Dữ liệu cần bổ sung",
             "Mẫu nội dung yêu cầu để Chủ tịch chuyển tiếp")]
    for ma, q, phong, du_lieu in ROUTING_ROWS:
        mau = (f"Kính gửi {phong}, để Dashboard trả lời được câu \"{q}\", đề nghị bộ phận "
               f"cung cấp định kỳ: {du_lieu}. Định dạng file Excel, gửi theo luồng báo cáo "
               f"hiện tại về Connect_VPS.")
        rows.append((ma, q, phong, du_lieu, mau))
    return rows


BUILDERS = {
    "L1_CANHBAO": build_L1_CANHBAO,
    "L2_DOANHTHU": build_L2_DOANHTHU,
    "L3_LOINHUAN": build_L3_LOINHUAN,
    "L4_DONGTIEN": build_L4_DONGTIEN,
    "L5_CONGNO": build_L5_CONGNO,
    "L7_CHIPHI": build_L7_CHIPHI,
    "L8_XEPHANG": build_L8_XEPHANG,
    "L9_DUAN": build_L9_DUAN,
    "L10_NGHIVAN": build_L10_NGHIVAN,
    "ROUTING": build_ROUTING,
}
SHEET_ORDER = ["_META"] + list(BUILDERS.keys())

# Sheet nào cần nguồn nào — chỉ nạp đúng nguồn cần khi chạy --only (tiết kiệm thời gian mở file)
SHEET_NEEDS = {
    "L1_CANHBAO": {"hqkd"},
    "L2_DOANHTHU": {"hqkd"},
    "L3_LOINHUAN": {"hqkd"},
    "L7_CHIPHI": {"hqkd"},
    "L9_DUAN": {"hqkd"},
    "L8_XEPHANG": {"hqkd", "sodu", "aging"},
    "L4_DONGTIEN": {"sodu", "aging"},
    "L5_CONGNO": {"aging", "detail"},
    "L10_NGHIVAN": {"aging", "detail"},
    "ROUTING": set(),
}


# ---- Ghi workbook -----------------------------------------------------------------

def _last_inbox_event():
    """Dòng cuối inbox_events.jsonl -> (at, file_name). Đọc đuôi file, không nạp cả 838+ dòng."""
    try:
        with open(INBOX_EVENTS, "rb") as fh:
            fh.seek(0, os.SEEK_END)
            size = fh.tell()
            fh.seek(max(0, size - 8192))
            tail = fh.read().decode("utf-8", "ignore").strip().splitlines()
        for line in reversed(tail):
            try:
                ev = json.loads(line)
                return ev.get("at"), ev.get("file_name")
            except Exception:
                continue
    except FileNotFoundError:
        pass
    return None, None


def _write(sheets_built, prev_meta, errors, notes, out_path):
    """Ghi NGUYÊN TỬ: dựng bản tạm rồi os.replace — agent đang đọc không bao giờ vớ file dở."""
    wb = openpyxl.load_workbook(out_path) if os.path.exists(out_path) else openpyxl.Workbook()
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        del wb["Sheet"]

    built_at = {}
    built_at.update(prev_meta.get("built_at", {}))
    for name, rows in sheets_built.items():
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)
        for r in rows:
            ws.append(list(r))
        built_at[name] = _ts()

    # _META dựng lại mỗi lần
    if "_META" in wb.sheetnames:
        del wb["_META"]
    meta = wb.create_sheet("_META", 0)
    ev_at, ev_file = _last_inbox_event()
    now = _now()
    stale = ""
    oldest = min(built_at.values()) if built_at else None
    if oldest:
        try:
            age_h = (now - datetime.strptime(oldest, "%Y-%m-%d %H:%M:%S").replace(tzinfo=TZ)
                     ).total_seconds() / 3600
            if age_h > STALE_AFTER_HOURS:
                stale = f"STALE — sheet cũ nhất đã {age_h:.1f} giờ chưa dựng lại"
        except Exception:
            pass
    meta.append(["Trường", "Giá trị"])
    meta.append(["as_of", _ts(now)])
    meta.append(["trang_thai", stale or "OK"])
    meta.append(["last_event_at", ev_at or "—"])
    meta.append(["last_file_received", ev_file or "—"])
    meta.append(["so_sheet", len(built_at)])
    meta.append([])
    meta.append(["Sheet", "built_at"])
    for k in SHEET_ORDER[1:]:
        if k in built_at:
            meta.append([k, built_at[k]])
    if notes:
        meta.append([])
        meta.append(["Nguồn THIẾU trong kỳ này (phải nói ra khi trả lời)", ""])
        for n in notes[:40]:
            meta.append([n, ""])
    if errors:
        meta.append([])
        meta.append(["Lỗi khi dựng (sheet liên quan có thể thiếu dòng)", ""])
        for e in errors[:40]:
            meta.append([e, ""])

    wb._sheets.sort(key=lambda s: SHEET_ORDER.index(s.title)
                    if s.title in SHEET_ORDER else 99)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=os.path.dirname(out_path))
    os.close(fd)
    try:
        wb.save(tmp)
        # mkstemp tạo file 0600 và shutil.move GIỮ NGUYÊN quyền đó -> agent OpenClaw (chạy user
        # `node` trong container) không đọc được brief, dù đường dẫn hoàn toàn hợp lệ. Đã dính
        # thật 07/08: agent hỏi ngược người dùng xin chmod thay vì trả lời. Đặt 0644 cho khớp
        # các file khác trong `Tài liệu/`.
        os.chmod(tmp, 0o644)
        shutil.move(tmp, out_path)
    finally:
        if os.path.exists(tmp):
            os.unlink(tmp)
    return built_at, stale


def _read_prev_meta(out_path):
    if not os.path.exists(out_path):
        return {}
    try:
        wb = openpyxl.load_workbook(out_path, read_only=True, data_only=True)
        if "_META" not in wb.sheetnames:
            return {}
        rows = list(wb["_META"].iter_rows(values_only=True))
        wb.close()
        built, seen_hdr = {}, False
        for r in rows:
            if r and r[0] == "Sheet" and len(r) > 1 and r[1] == "built_at":
                seen_hdr = True
                continue
            if seen_hdr and r and r[0] and r[0] in BUILDERS:
                built[r[0]] = r[1]
        return {"built_at": built}
    except Exception:
        return {}


def build(only=None, out_path=OUT_PATH, verbose=True):
    targets = [s for s in BUILDERS if not only or s in only]
    if only:
        unknown = [s for s in only if s not in BUILDERS]
        if unknown:
            raise SystemExit(f"Sheet không tồn tại: {unknown}. Hợp lệ: {list(BUILDERS)}")
    needs = set()
    for t in targets:
        needs |= SHEET_NEEDS.get(t, set())

    errors, extra_notes = [], []
    ctx = {"hqkd": {}, "aging": {}, "detail": {}, "sodu": [], "sodu_meta": {}, "skipped": []}
    if "hqkd" in needs:
        ctx["hqkd"] = _collect_hqkd_ngay(errors, ctx["skipped"])
    if "aging" in needs or "detail" in needs:
        ctx["aging"] = _collect_aging(errors)
    if "detail" in needs:
        ctx["detail"] = _collect_congno_detail(ctx["aging"], errors)
    if "sodu" in needs:
        ctx["sodu"], ctx["sodu_meta"] = _collect_sodu_tien(errors, extra_notes)

    sheets = {}
    for name in targets:
        try:
            sheets[name] = BUILDERS[name](ctx)
        except Exception as ex:
            errors.append(f"sheet {name}: {type(ex).__name__}: {ex}")
            if verbose:
                traceback.print_exc()

    notes = [f"{FOLDER_MAP.get(f, (f,))[0]} ({f}): KHÔNG có số báo cáo ngày — {ly_do}"
             for f, _file, ly_do in ctx["skipped"]] + extra_notes
    built_at, stale = _write(sheets, _read_prev_meta(out_path), errors, notes, out_path)
    if verbose:
        print(f"[brief] ghi {out_path}")
        print(f"[brief] sheet dựng lần này: {', '.join(sorted(sheets)) or '(không)'}")
        print(f"[brief] nguồn: hqkd={len(ctx['hqkd'])} đơn vị · aging={len(ctx['aging'])} · "
              f"detail={len(ctx['detail'])} · sodu={len(ctx['sodu'])} pháp nhân")
        for n in notes:
            print(f"[brief][THIẾU] {n}")
        if stale:
            print(f"[brief] {stale}")
        for e in errors:
            print(f"[brief][LỖI] {e}")
    return {"sheets": sorted(sheets), "errors": errors, "notes": notes,
            "built_at": built_at, "stale": stale}


def main():
    ap = argparse.ArgumentParser(description="Dựng BRIEF_CHUTICH.xlsx")
    ap.add_argument("--full", action="store_true", help="dựng lại toàn bộ sheet")
    ap.add_argument("--only", nargs="+", metavar="SHEET",
                    help=f"chỉ dựng sheet chỉ định: {', '.join(BUILDERS)}")
    ap.add_argument("--out", default=OUT_PATH)
    a = ap.parse_args()
    if not a.full and not a.only:
        ap.error("phải chọn --full hoặc --only SHEET [SHEET...]")
    r = build(only=a.only, out_path=a.out)
    return 1 if r["errors"] and not r["sheets"] else 0


if __name__ == "__main__":
    sys.exit(main())
