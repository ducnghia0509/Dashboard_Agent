# -*- coding: utf-8 -*-
"""Dẫn xuất SỐ DƯ VAY (04_VAY) từ Báo cáo Tiền tập đoàn (sheet '..._SD TIỀN', mục I = TIỀN VAY).

Mapping nghiệp vụ: VAY / SỐ DƯ VAY lấy từ Baocaothuchi (sheet ngân hàng). Trong SD TIỀN, mục I có
theo CÔNG TY × NGÂN HÀNG: cột 'ĐẦU KỲ' = dư đầu kỳ, 'ĐẾN NGÀY HIỆN TẠI' = dư cuối kỳ, các cột
'+/- NGÀY dd' = biến động ngày (Σ dương = VAY THÊM, Σ|âm| = TRẢ NỢ). Đủ để lên #24 Nợ vay + trang Vay.
KHÔNG có: kỳ hạn (NH/TH/DH), lãi vay, nợ đến hạn (mapping ghi "hỏi phòng Thanh toán") -> để trống.

Chạy: .venv/bin/python scripts/extract_vay.py <file_thuchi.xlsx> --period 2026-01
"""
import argparse
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from servers import template_filler as tf  # noqa: E402
from servers.common import source_catalog as SC  # noqa: E402
from servers.common import org  # noqa: E402  — danh mục tổ chức (nguồn define duy nhất)


def _sd_sheet(wb):
    for ws in wb.worksheets:
        if "SD TI" in ws.title.upper() or "SỐ DƯ TI" in ws.title.upper():
            return ws
    return None


def _cols(ws):
    """Tìm dòng header + chỉ số cột 'đầu kỳ'/'đến ngày' + danh sách cột '+/-'. None nếu không thấy."""
    for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
        low = [str(c).strip().lower() if c is not None else "" for c in row]
        i_dau = next((i for i, c in enumerate(low) if c.startswith("đầu kỳ")), None)
        i_cuoi = next((i for i, c in enumerate(low) if "đến ngày" in c), None)
        if i_dau is not None and i_cuoi is not None:
            deltas = [i for i, c in enumerate(low) if c.startswith("+/-")]
            return i_dau, i_cuoi, deltas
    return None


def extract(path: str, period: str, cong_ty: str = None) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = _sd_sheet(wb)
    if ws is None:
        return {"ok": False, "error": "Không thấy sheet SD TIỀN"}
    cc = _cols(ws)
    if not cc:
        return {"ok": False, "error": "Không thấy cột ĐẦU KỲ / ĐẾN NGÀY HIỆN TẠI"}
    i_dau, i_cuoi, deltas = cc

    import re
    _KYHAN = re.compile(r"ngắn hạn|trung hạn|dài hạn")   # dòng con kỳ hạn (cha đã có tổng) -> bỏ
    num = lambda v: v if isinstance(v, (int, float)) else 0.0  # noqa: E731
    recs, cur_co, in_vay, vay_seen = [], None, False, False
    for r in ws.iter_rows(values_only=True):
        c1 = str(r[1]).strip() if len(r) > 1 and r[1] not in (None, "") else ""
        c2 = str(r[2]).strip() if len(r) > 2 and r[2] not in (None, "") else ""
        c3 = str(r[3]).strip() if len(r) > 3 and r[3] not in (None, "") else ""
        # Đổi mục khi cột LOẠI TIỀN (index1) có nhãn mục. Chỉ xử lý mục I = TIỀN VAY, và CHỈ khối
        # ĐẦU TIÊN: nguồn liệt kê 'TIỀN VAY' 2 lần (cùng 'đến ngày hiện tại') -> đọc cả 2 sẽ NHÂN ĐÔI.
        if c1:
            if "VAY" in c1.upper():
                in_vay = not vay_seen
                vay_seen = True
            else:
                in_vay = False
            continue
        if not in_vay:
            continue
        if c2 and not c3:                    # dòng CÔNG TY (tổng) -> đặt pháp nhân hiện tại
            cur_co = org.company_code(c2)
            continue
        if c3 and cur_co:                    # dòng NGÂN HÀNG (chi tiết) dưới công ty
            if _KYHAN.search(c3.lower()):    # 'BIDV vay ngắn/trung hạn' -> con của 'BIDV', BỎ tránh cộng trùng
                continue
            dau = num(r[i_dau] if i_dau < len(r) else None)
            cuoi = num(r[i_cuoi] if i_cuoi < len(r) else None)
            vay_them = sum(num(r[i]) for i in deltas if i < len(r) and num(r[i]) > 0)
            tra_no = -sum(num(r[i]) for i in deltas if i < len(r) and num(r[i]) < 0)
            if not any((dau, cuoi, vay_them, tra_no)):
                continue
            recs.append({"Kỳ": period, "Đơn vị": cur_co, "Ngân hàng": c3,
                         "Dư nợ đầu kỳ (tỷ)": round(dau / 1e9, 9),
                         "Vay thêm trong kỳ (tỷ)": round(vay_them / 1e9, 9),
                         "Trả nợ trong kỳ (tỷ)": round(tra_no / 1e9, 9),
                         "Dư nợ cuối kỳ (tỷ)": round(cuoi / 1e9, 9)})
    if not recs:
        return {"ok": False, "error": "Không bóc được dòng vay nào (mục I)"}
    out = os.path.join(tf.FILLED_DIR, f"THUCHI_{period}_04_VAY.xlsx")
    tf.fill("04_VAY", recs, out)
    imp = tf.import_filled(out, cong_ty=cong_ty, source_file=SC.source_id_from_path(path))
    return {"ok": True, "rows": imp.get("rows_imported"), "by_type": imp.get("by_type"),
            "banks": len(recs), "companies": sorted({x["Đơn vị"] for x in recs}), "out": out}


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--period", required=True)
    ap.add_argument("--cong-ty", dest="cong_ty", default=None)
    a = ap.parse_args()
    import json
    print(json.dumps(extract(a.file, a.period, a.cong_ty), ensure_ascii=False))
