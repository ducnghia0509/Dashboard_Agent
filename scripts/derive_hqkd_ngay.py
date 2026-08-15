# -*- coding: utf-8 -*-
"""Deriver: BÁO CÁO NGÀY (HQKD theo từng ngày) — nguồn `received_reports/<FOLDER>/baocaohqkdngay/
B.<n>.<MÃ>.D.<YYYYMM>.<Tên>.xlsx` (ký tự **D** trong tên = Day; bản **M** là báo cáo tháng đã có
pipeline riêng, KHÔNG đụng tới).

PHẠM VI (chốt theo Mapping_Dashboard_QTTC.xlsx — cột "Đường link lấy dữ liệu ngày tạm thời trên
EXCEL", ngay bên phải cột "Map màn hình"): CHỈ đơn vị có ô này khác "ko có" mới lên được báo cáo
ngày, và CHỈ cụm chỉ tiêu P&L (doanh thu / giá vốn / chi phí / lợi nhuận). Mọi chỉ tiêu khác của
màn Công nợ · Tồn kho · Tài sản · Thuế · Dòng tiền ghi "ko có" -> KHÔNG dựng số ngày cho chúng.
Hiện có 10 đơn vị: SRVF, XANHVINHPHUC, HTXXANHTUYENQUANG, HTXXANHVINHPHUC, ANTAXI, ANKHACHSAN,
GLOBALAI, TRAMSAC, DUAN, HO.

report_type RIÊNG có hậu tố `_D` (HQKD_D / PNLT_D / CHIPHI_D) — KHÔNG ghi đè HQKD/PNLT/CHIPHI của
báo cáo tháng. LÝ DO BẮT BUỘC: dòng ngày nằm CÙNG dataset tháng (để bộ lọc "Khoảng ngày" của FE
vẫn resolve ra đúng dataset kỳ). Nếu dùng chung report_type thì chế độ Tháng — vốn KHÔNG gửi
from/to nên cộng hết mọi dòng trong dataset — sẽ cộng cả 31 dòng ngày lẫn dòng tháng = số gấp đôi.
Backend chọn `_D` hay bản tháng theo `grain` của request (xem app/metrics/repository.py::_rt).

Đơn vị & layout (verify trên file thật kỳ 2026-08):
  · SRVF  (`B.1.TC.TCKT.D.202608.BaocaoHQKD.xlsx`) — layout "srvf": mỗi ngày 1 sheet tên "{d}.{m}"
    (vd "1.8" = 01/08). Header dòng 1: Mục | Khối | Mã số | Chỉ tiêu | <Showroom> | % DT | …
    Mã A-series y hệt sheet tháng T{mm}BC (A100 doanh thu, A300 tổng CP, A310 giá vốn, A600 lợi
    nhuận, U302 LNST). KHÔNG có cột tổng (cột "CHI NHÁNH VINFAST HÀ NỘI" luôn = 0) -> chỉ ghi dòng
    theo cost center, tổng khối = Σ showroom (đúng nhánh cc_v của repository._per_file_resolved).
    ⚠ VỊ TRÍ CỘT LỆCH GIỮA CÁC SHEET (sheet "1.8" showroom đầu ở cột 5, "2.8"/"3.8" ở cột 6 vì
    thêm 1 cột "% DT") -> BẮT BUỘC dò cột theo TÊN header từng sheet, không hardcode index.
  · XANHVINHPHUC (`B.6.XVP.D.<YYYYMM>.Baocaohqkdngay.xlsx`) — layout "kqkd": sheet "01".."31" theo
    ngày. Header dòng 4: CHỈ TIÊU | MÃ SỐ | Tổng cộng | HO | Depot Phú Thọ | Depot Vĩnh Phúc |
    Depot Tuyên Quang. Có cột tổng -> ghi thêm dòng "trực tiếp" (không cost center) như bản tháng.
  · HTXXANHTUYENQUANG / HTXXANHVINHPHUC (`B.6.HTX_*.D.<YYYYMM>.Baocaotaichinhrieng.xlsx`) — cũng
    layout "kqkd" nhưng header ở dòng 6 và CHỈ 1 cột giá trị (không có cost center).
  · ANTAXI (`B.7.AAG.TCKT.**M**.<YYYYMM>.Baocaotaichinhrieng.xlsx`) — layout "antaxi": sheet
    "1".."31" theo ngày (+ sheet "TH" tổng hợp, bỏ qua). Header dòng 6: TT | CHỈ TIÊU | Tỉ lệ |
    Sơn Tây | Tỉ lệ | Thái Nguyên | Tỉ lệ | Tổng cộng | Xe thương quyền | … | Lũy kế tháng.
    ⚠ TÊN FILE ghi `.M.` (trùng hệt báo cáo THÁNG ở `baocaotaichinhrieng/`) — chỉ THƯ MỤC
    `baocaohqkdngay/` phân biệt được, xem `_period_of`/`_in_day_dir`.
    ⚠ Số mục La Mã ở cột TT riêng + lệch so với XVP -> bảng anchor riêng, xem `_ANTAXI_ANCHOR`.
  · ANKHACHSAN (`B.10.AAG.TCKT.D.<YYYYMM>.Baocaotaichinhrieng.xlsx`) — layout "anks", KHÁC HẲN
    4 layout trên: CHỈ 1 SHEET, mỗi NGÀY là 1 CỘT (không phải mỗi ngày 1 sheet). Header dòng 3:
    TT | Nội dung chi phí | ĐVT | Tổng cộng | 1 | 2 | … | 31 — số ngày nằm sẵn ở header, không
    suy từ tên sheet. Chỉ 1 cơ sở (Garden Sơn Tây), không cost center. Xem `_anks_all_days`.
  · GLOBALAI (`B.8.GA.TCKT.D.<YYYYMM><DD>.Baocaotaichinhrieng.xlsx`) — layout "tcode": mỗi ngày 1
    sheet tên "D1".."D31" (file LUỸ KẾ trong tháng — số sheet tăng dần theo ngày đã qua, KHÔNG
    zero-pad "D1" chứ không phải "D01"). Mã T100/T200/T300 ở cột A y hệt sheet THÁNG "HQKD HỢP
    NHẤT GA" (cột B = Chỉ tiêu, cột giá trị header ghi "D01".."D31" — CÓ zero-pad, khác tên sheet).
    KHÔNG cost center (GA 1 pháp nhân, không showroom/depot).
    ⚠ Tên file ghi thêm 2 số NGÀY snapshot ở cuối (".D.20260601." = kỳ 2026-06, KHÔNG PHẢI báo cáo
    của riêng ngày 01 — file là snapshot LUỸ KẾ cả tháng tính đến ngày lưu) — `_period_of` chỉ bắt
    6 số đầu sau ".D." nên vẫn ra đúng "2026-06" dù tên có thêm 2 số dư.
    Chỉ nạp 4 chỉ tiêu CÓ NGUỒN theo Mapping (DT thuần T101, Giá vốn T201, Tổng chi phí T200, LNTT
    T300) + breakdown chi phí T201-T204 (giá vốn/tài chính/bán hàng/QLDN, khớp nhãn TT200 mà bản
    THÁNG của GA đang dùng — xem `_TCODE_CANON`). KHÔNG suy diễn "Lợi nhuận sau thuế"
    từ T300: file NGÀY không tách dòng thuế TNDN, khác bản THÁNG (đang đọc sheet 'KQKD' TT200 hợp
    lệ, có dòng thuế riêng mã 51/52) — gán LNST=T300 như HT (T-series) sẽ SAI khi GA phát sinh
    thuế thật; để trống, tương tự các chỉ tiêu khác ghi "ko có" trong Mapping cho GA.
  · TRAMSAC (`B.3.TC.TCKT.D.<YYYYMMDD>.Baocaotaichinhrieng.xlsx`) — CÙNG layout "tcode" với GA
    (mã T100/T200/T300, sheet "D1".."D31" luỹ kế) nhưng KHÁC 1 ĐIỂM: quy ước PNLT LNTT/LNST.
    Verify DB 2026-06: PNLT của Trạm sạc CHỈ có đúng 1 dim1 'Lợi nhuận sau thuế' (không có 'Lợi
    nhuận trước thuế' riêng), giá trị BẰNG TUYỆT ĐỐI HQKD mã 1112 (-0,3844 = -0,3844) -> T300 ở
    Trạm sạc ĐÃ LÀ LNST (không như GA còn lăn tăn thuế TNDN chưa tách). Gán nhãn 'Lợi nhuận trước
    thuế' cho Trạm sạc như GA sẽ SAI TÊN cột trên bảng Cấu trúc Doanh thu/Chi phí (dim1 không khớp
    quy ước THÁNG). Xem tham số `profit_pnlt` của `_tcode_facts` — set qua `_UNITS["TRAMSAC"]`.
    ⚠ Tên file có NGÀY ĐẦY ĐỦ ".D.20260801." (8 số, khác GA chỉ dư 2 số) nhưng `_period_of` chỉ
    bắt 6 số đầu sau ".D." nên vẫn ra đúng "2026-08"; sheet "D1" KHÔNG zero-pad (khác header
    'D1' cũng không zero-pad, khác GA có header zero-pad 'D01' — không ảnh hưởng vì `_tcode_facts`
    dò val_j có fallback `ten_j + 1` khi không tìm thấy header dạng d\\d{2}).
  · DUAN (Khối Dự án, `B.4.TC.TCKT.D.<YYYYM>.BaocaoHQKD.xlsx`) — layout "duan": mỗi ngày 1 sheet
    tên "1".."31" (không zero-pad, giống "kqkd"). ⚠ THÁNG TRONG TÊN FILE KHÔNG ZERO-PAD ("
    .D.20268." = kỳ 2026-08, CHỈ 5 CHỮ SỐ — khác 6 chữ số của mọi đơn vị khác) → `_period_of`
    cần regex fallback riêng, xem bên dưới. Header CHIA 2 DÒNG khác nhau (dòng "STT|CHỈ TIÊU|
    DỰ ÁN" rồi dòng SAU MỚI liệt kê cột: "Tổng Dự án|HO Dự án|Cao Bằng|Tân Thịnh|Lạng Sơn|Yên
    Bình 3|Phú Quốc|Bình phước|Quang Sơn") — khác mọi layout khác (đều gộp nhãn cột + tên cost
    center CÙNG 1 dòng) nên KHÔNG dùng chung `_kqkd_scan`, có `_duan_facts` riêng. Cột "Tổng Dự
    án" = Σ 7 dự án (verify ngày 01/08: 392.421.525 + 239.351.852 = 631.773.377 = đúng cột Tổng)
    và "HO Dự án" luôn 0 ở dữ liệu đã verify → BỎ CẢ HAI, chỉ lấy 7 cột dự án (E:K) làm cost
    center, đúng theo Mapping ("E-K tương ứng cho các Costcenter"). Số La Mã (III/IV/V/VI/VIII/
    IX/X/XI/XII) nằm ở cột STT RIÊNG (giống "antaxi"), nhãn cột "CHỈ TIÊU" là text trần không số
    — neo theo nhãn CHUẨN HOÁ, đa số EXACT MATCH (không startswith) vì "Chi phí khác" (X.2, mã
    neo `cp_khac`) startswith sẽ trúng NHẦM dòng con "Chi phí khác tại dự án" (mục 1.7 của Giá
    vốn, đứng TRƯỚC trong sheet) nếu dùng prefix "chi phi khac" lỏng lẻo — chỉ 2 mã LNTT/LNST
    dùng startswith (nhãn có hậu tố "(EBT)"/"(EAT)" đổi được). Tổng chi phí = Giá vốn(IV) + Chi
    phí biến đổi(VI) + Chi phí cố định(VIII) + Chi phí tài chính(IX.2) + Chi phí khác(X.2) — ĐÚNG
    công thức "E25+E46+E74+E80+E83" của Mapping, cùng 5 nhóm CP y hệt ANTAXI (không có mục
    "phân bổ chung" cấp I riêng — đã nằm lồng trong "Chi phí khác"). Mã cost center Cao Bằng/
    Lạng Sơn/Phú Quốc/Quang Sơn lấy y hệt bản THÁNG (agent_cli._DA_PROJECT_CC: CB_DA/LS_DA/
    PQ_DA/QS_DA); ⚠ "Tân Thịnh"→YB_DA và "Yên Bình"→TT_DA là NGƯỢC viết tắt đã xác nhận nguồn
    (xem agent_cli.py dòng ~728, ĐỪNG "sửa cho xuôi"). "Bình phước" CHƯA có trong master_data
    (giống Núi Pháo/Quảng Ngãi bản tháng) → mã tự đặt BINHPHUOC_DA, backfill cong_ty qua
    import_filled.

Neo dòng chỉ tiêu của layout "kqkd" theo TIỀN TỐ SỐ LA MÃ / số mục đã chuẩn hoá bỏ dấu, KHÔNG theo
"Mã số" và KHÔNG theo địa chỉ ô cứng (C17/C28/C100… như ghi chú trong file mapping): mã số bị TRÙNG
(HTX có 2 dòng mã "10": "1. Doanh thu bán hàng" và "3. Doanh thu thuần") và số thứ tự dòng LỆCH
giữa XVP (dòng 20) với HTX (dòng 22).

Tổng chi phí = Σ 6 cấu phần (IV giá vốn + VI biến đổi + VIII cố định + IX.2 tài chính + X.2 khác +
XII phân bổ chung) — verify khớp ĐÚNG dòng "18. Tổng chi phí" có sẵn trong file XVP ngày 01/08
(1.026.059.132 = 850.225.145 + 40.124.348 + 0 + 122.161.253 + 0 + 13.548.387). Lấy Σ cấu phần thay
vì dòng 18.1 để Σ lát CHIPHI_D luôn == tổng chi phí HQKD_D (file HTX không có dòng 18.1).

Chạy (dry-run, in tổng theo ngày, KHÔNG ghi):
  .venv/bin/python scripts/derive_hqkd_ngay.py <file.xlsx>
Ghi thật:
  .venv/bin/python scripts/derive_hqkd_ngay.py <file.xlsx> --write
"""
import argparse
import json
import os
import re
import sys
import unicodedata

import openpyxl
import psycopg

_HERE = os.path.dirname(os.path.abspath(__file__))
sys.path[:0] = [p for p in (os.path.dirname(_HERE), _HERE) if p not in sys.path]
from servers.common import source_catalog as _SC  # noqa: E402

DB_URL = (os.environ.get("DATABASE_URL") or os.environ.get("TC_DATABASE_URL")
          or "postgresql://tc:tc@localhost:5433/tc_dashboard")

RT_HQKD, RT_PNLT, RT_CHIPHI, RT_DTHU = "HQKD_D", "PNLT_D", "CHIPHI_D", "DTHU_D"
# Doanh thu BÁN XE theo ngày × kênh (chỉ layout "srvf"). Tách khỏi RT_DTHU vì đó là doanh
# thu thuần toàn khối, còn cái này là cụm A200 chia B2C/B2B/GF — vế thực hiện của bảng
# điểm vhkd0. Tên khớp `KDVH` (bản THÁNG, nguồn BaocaoKQKD) + hậu tố _D theo quy ước ngày.
RT_KDVH = "KDVH_D"
# Lợi nhuận theo NGÀY x KENH (15/08/2026). Tach khoi PNLT_D: PNLT_D giu dong muc SHOWROOM
# (dim2 rong), con day la breakdown theo kenh — de chung mot report_type thi ai cong ca hai se
# dem doi. Verify sheet 13.8: A401 (B2C) -368.300.254 + A402 (B2B) -39.218.994 = -407.519.248
# = dung o A600 "LOI NHUAN SHOW ROOM".
RT_LN_KENH = "VHKD_LN_D"
REPORT_TYPES = (RT_HQKD, RT_PNLT, RT_CHIPHI, RT_DTHU, RT_KDVH, RT_LN_KENH)

# Mã chỉ tiêu 01_HQKD (khớp app/metrics/repository.py: HQKD_REVENUE/COST/PROFIT_AT).
MA_DT, MA_CP, MA_LNTT = "1000", "1047", "1112"

_UNITS = {
    "SRVF": {"layout": "srvf", "cong_ty": "TC", "khoi": "Khối KD Vinfast - Showroom"},
    "XANHVINHPHUC": {"layout": "kqkd", "cong_ty": "XVP", "khoi": "Khối KD Vận tải Taxi Xanh"},
    "HTXXANHTUYENQUANG": {"layout": "kqkd", "cong_ty": "HTX_XTQ", "khoi": "Khối KD Vận tải Taxi Xanh"},
    "HTXXANHVINHPHUC": {"layout": "kqkd", "cong_ty": "HTX_XVP", "khoi": "Khối KD Vận tải Taxi Xanh"},
    "ANTAXI": {"layout": "antaxi", "cong_ty": "AAG", "khoi": "Khối KD Dịch vụ An Taxi"},
    "ANKHACHSAN": {"layout": "anks", "cong_ty": "AAG", "khoi": "Khối KD Dịch vụ An KS"},
    # pnlt_skip T101: GA bản THÁNG chạy extractor TT200 (không phải T-series) -> không có dim1
    # 'Doanh thu bán hàng'; xem `_tcode_facts`.
    "GLOBALAI": {"layout": "tcode", "cong_ty": "GA", "khoi": "Khối KD Công nghệ",
                 "pnlt_skip": ("T101",)},
    "TRAMSAC": {"layout": "tcode", "cong_ty": "TC", "khoi": "Khối KD Trạm sạc Vgreen",
               "profit_pnlt": ("Lợi nhuận sau thuế",)},
    "DUAN": {"layout": "duan", "cong_ty": "TC", "khoi": "Khối KD Dự án"},
    "HO": {"layout": "ho_kqkd", "cong_ty": "TC", "khoi": "Khối hỗ trợ tập đoàn"},
    # XE TẢI HƯNG THỊNH (spec user 2026-08-06) — layout "ht", xem `_ht_facts`.
    "HUNGTHINH": {"layout": "ht", "cong_ty": "HT", "khoi": "Khối KD Xe tải"},
    # XƯỞNG DỊCH VỤ VINFAST (spec user 2026-08-06) — layout "xdv", xem `_xdv_facts`.
    "XDV": {"layout": "xdv", "cong_ty": "TC", "khoi": "Khối KD Vinfast - XDV"},
}

# Cost center theo TỪ KHOÁ trong tên cột (dò theo tên, không theo vị trí — xem docstring).
# Mã CC lấy y hệt bản tháng (agent_cli._SR_SHOWROOM_CC / raw_rows thật của XVP).
_CC_SRVF = [("uong bi", "UB_SR"), ("b2b", "B2B_SR"), ("oceanpark", "OCP_SR"), ("long bien", "LB_SR"),
            ("smart city", "SMC_SR"), ("ha long", "HL_SR"), ("cam pha", "CP_SR"),
            ("vinh phuc", "VP_SR"), ("son tay", "ST_SR"), ("xuan mai", "XM_SR")]
_CC_XVP = [("depot phu tho", "PT_DP"), ("depot vinh phuc", "VP_DP"), ("depot tuyen quang", "TQ_DP"),
           ("ho", "HO_XVP")]
# An Taxi: header ghi tên tỉnh trần ("Sơn Tây" / "Thái Nguyên"), mã CC lấy y hệt master_data
# (app/data/master_data.json: ST_AT 'Depot Sơn Tây', TN_AT 'Depot Thái Nguyên').
_CC_ANTAXI = [("son tay", "ST_AT"), ("thai nguyen", "TN_AT")]

# Showroom Uông Bí thuộc pháp nhân VFQN (SRVF là thư mục ĐA-pháp-nhân) — giống bản tháng.
_CC_CONGTY = {"UB_SR": "VFQN"}


def _nd(s):
    """Chuẩn hoá: bỏ dấu, thường hoá, gộp khoảng trắng (nguồn hay gõ sai dấu: 'Gíá vốn', 'LỢI NHUÂN')."""
    s = str(s or "").strip().lower().replace("đ", "d")
    s = "".join(ch for ch in unicodedata.normalize("NFD", s) if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", s)


def _num(v):
    return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def _source_id(path):
    folder = (_SC.raw_company_from_path(path) or "").upper()
    return f"{folder}::{os.path.basename(path)}"


_DAY_DIR = "baocaohqkdngay"


def _in_day_dir(path):
    """File có nằm trong thư mục `baocaohqkdngay/` không."""
    return _nd(os.path.basename(os.path.dirname(os.path.abspath(path)))).replace(" ", "") == _DAY_DIR


def _period_of(file_name, in_day_dir=False):
    """'B.1.TC.TCKT.D.202608.BaocaoHQKD.xlsx' -> '2026-08'. None nếu không phải file kỳ .D.<YYYYMM>.

    An Taxi gửi báo cáo NGÀY nhưng đặt tên '.M.<YYYYMM>' Y HỆT báo cáo THÁNG
    ('B.7.AAG.TCKT.M.202608.Baocaotaichinhrieng.xlsx' — cùng tên với file trong
    `baocaotaichinhrieng/`) -> chỉ THƯ MỤC phân biệt được. Vì vậy '.M.' CHỈ được coi là kỳ báo
    cáo ngày khi `in_day_dir`; nếu nới lỏng theo tên file thì báo cáo THÁNG của An Taxi sẽ bị
    gate ngày bắt và pipeline tháng không bao giờ chạy."""
    m = re.search(r"\.D\.(\d{4})(\d{2})", file_name)
    if m is None:
        # DUAN: tháng KHÔNG zero-pad ('.D.20268.' = 2026-08, chỉ 5 chữ số) — bắt buộc neo dấu
        # '.' liền sau để không lẫn với 6-số của các đơn vị khác (regex trên đã tự fail cho
        # trường hợp đó nên không tranh chấp thứ tự thử).
        m = re.search(r"\.D\.(\d{4})(\d{1,2})\.", file_name)
    if m is None and in_day_dir:
        m = re.search(r"\.M\.(\d{4})(\d{2})", file_name)
    return f"{m.group(1)}-{int(m.group(2)):02d}" if m else None


def is_daily_report(path):
    """File này có phải BÁO CÁO NGÀY của đơn vị đã cấu hình không (dùng cho gate ở agent_cli)."""
    folder = _source_id(path).split("::", 1)[0]
    return bool(_UNITS.get(folder)) and bool(_period_of(os.path.basename(path), _in_day_dir(path)))


# ---------------------------------------------------------------------------------------------
# Layout "srvf" — mã A-series, mỗi ngày 1 sheet "{d}.{m}"
# ---------------------------------------------------------------------------------------------
# Cấu phần TRỰC TIẾP của A300 (bản tháng parse từ công thức A300; workbook data_only KHÔNG còn
# công thức nên dùng thẳng danh sách fallback y hệt agent_cli._chiphi_recs_srvf).
# Mã CON của A200 -> kênh bán. A211/A211A/A213 đều là B2B (xe khối B2B, xe khối B2B bản
# phụ, và B2B đã xuất hoá đơn) — gộp về một kênh cho khớp chiều phân tích của mapping
# ("Kênh B2B, B2C, GF"), mã gốc vẫn giữ ở dim3 để soát ngược từng dòng với file.
_SRVF_BANXE = [("A210", "B2C"), ("A211", "B2B"), ("A211A", "B2B"),
               ("A212", "GF"), ("A213", "B2B")]

# Ma CON cua A600 (LNTT) va U302 (LNST) -> kenh. Cung nguyen tac: chi doc dong CON, dong cha
# da nam o PNLT_D roi.
_SRVF_LN_KENH = [("A401", "B2C", "A600"), ("A402", "B2B", "A600"), ("A403", "GF", "A600"),
                 ("A407", "B2C", "U302"), ("A408", "B2B", "U302"), ("A409", "GF", "U302")]

_SRVF_CP_CODES = ["A310", "A320", "A325", "A330", "A340", "A350", "A360", "A500"]


def _srvf_day_sheets(wb, period):
    """[(sheet_name, 'YYYY-MM-DD')] — chỉ sheet tên '{ngày}.{tháng}' khớp THÁNG của kỳ."""
    y, mm = int(period[:4]), int(period[5:7])
    out = []
    for s in wb.sheetnames:
        m = re.fullmatch(r"(\d{1,2})\.(\d{1,2})", s.strip())
        if m and int(m.group(2)) == mm and 1 <= int(m.group(1)) <= 31:
            out.append((s, f"{y:04d}-{mm:02d}-{int(m.group(1)):02d}"))
    return sorted(out, key=lambda x: x[1])


def _srvf_facts(rows):
    """rows -> [(cost_center, report_type, dim1, dim3, value_VND)] cho 1 ngày. [] nếu sai layout."""
    hdr = next((r for r in rows[:8] if any(_nd(c) == "ma so" for c in r if c is not None)), None)
    if hdr is None:
        return []
    ma_j = next(j for j, c in enumerate(hdr) if _nd(c) == "ma so")
    ten_j = next((j for j, c in enumerate(hdr) if _nd(c).startswith("chi tieu")), ma_j + 1)
    cols = [(cc, j) for kw, cc in _CC_SRVF
            for j in [next((j for j, c in enumerate(hdr) if isinstance(c, str) and kw in _nd(c)), None)]
            if j is not None]
    if len(cols) < 8:      # thiếu showroom = layout lạ -> không đoán
        return []
    by_code = {}
    for r in rows:
        c = str(r[ma_j]).strip().upper() if ma_j < len(r) and r[ma_j] not in (None, "") else ""
        if c and c not in by_code:
            by_code[c] = r

    def val(code, j):
        r = by_code.get(code)
        return _num(r[j]) if r is not None and j < len(r) else None

    from extract_chiphi import _nhom_cp
    facts = []
    for cc, j in cols:
        for code, rt, dim1 in (("A100", RT_HQKD, MA_DT), ("A300", RT_HQKD, MA_CP),
                               ("A600", RT_HQKD, MA_LNTT), ("A310", RT_PNLT, "Giá vốn hàng bán"),
                               ("A100", RT_PNLT, "Doanh thu HH, DV"),
                               ("A600", RT_PNLT, "Lợi nhuận trước thuế"),
                               ("A100", RT_DTHU, "Doanh thu thuần")):
            v = val(code, j)
            if v:
                facts.append((cc, rt, dim1, dim1, v))
        lnst = val("U302", j)
        lnst = lnst if lnst else val("A600", j)
        if lnst:
            facts.append((cc, RT_PNLT, "Lợi nhuận sau thuế", "Lợi nhuận sau thuế", lnst))
        # ── Doanh thu BÁN XE theo KÊNH (mapping VHKD dòng 37: "Số liệu BC ngày: Folder
        # BAOCAOHQKDNGAY - Sheet 01,02.. tương ứng ngày, tương ứng các cột costcenter").
        # CHỈ lấy các mã CON, KHÔNG lấy A200: A200 = A210+A211+A211A+A212+A213 (verify sheet 13.8:
        # 7.352.955.365 + 387.727.273 + 1.697.181.819 = 9.437.864.457 = đúng ô A200). Lấy cả hai
        # là gấp đôi doanh thu bán xe của mọi ngày.
        for code, kenh in _SRVF_BANXE:
            v = val(code, j)
            if v:
                facts.append((cc, RT_KDVH, "A200", code, v, kenh))
        # Loi nhuan truoc/sau thue theo KENH. dim1 = ma CHA (A600 / U302) de tang chung mot
        # truc voi PNLT_D, dim3 = ma con de soat nguoc tung dong voi file.
        for code, kenh, cha in _SRVF_LN_KENH:
            v = val(code, j)
            if v:
                facts.append((cc, RT_LN_KENH, cha, code, v, kenh))

        for code in _SRVF_CP_CODES:
            v = val(code, j)
            if not v:
                continue
            r = by_code[code]
            ten = str(r[ten_j]).strip() if ten_j < len(r) and r[ten_j] not in (None, "") else code
            facts.append((cc, RT_CHIPHI, _nhom_cp(ten, code), ten, v))
    return facts


# ---------------------------------------------------------------------------------------------
# Layout "kqkd" — XVP / HTX, mỗi ngày 1 sheet "01".."31"
# ---------------------------------------------------------------------------------------------
# (khoá, tiền tố nhãn đã chuẩn hoá) — neo theo NHÃN, xem docstring đầu file.
# ⚠ "lntt"/"lnst" dùng CONTAINS (tuple 1 phần tử), KHÔNG theo tiền tố La Mã như các dòng khác:
# verify file thật 2026-08 (XVP/HTX_XTQ/HTX_XVP) — cột "Mã số" đúng thứ tự (…11,12,13,14,15…) nhưng
# CỘT NHÃN kế toán ghi SAI/TRÙNG số La Mã từ dòng 13 trở đi (vd XVP: mã 13 "Lợi nhuận trước thuế
# TNDN (EBT)" lại ghi "XI." — trùng roman của chính mã 11 "LN TRỰC TIẾP TRƯỚC THUẾ TNDN" đứng trước;
# mã 15 "Lợi nhuận sau thuế TNDN (EAT)" ghi "XIII." thay vì "XV"; HTX ghi "XII."/"XIII." tuỳ đơn vị)
# -> neo prefix "xiii."/"xv." cũ KHÔNG BAO GIỜ khớp, HQKD/PNLT mất trắng 1112 + "Lợi nhuận trước/sau
# thuế" ở cả 3 đơn vị (phát hiện 2026-08-07, dữ liệu TEST xác nhận thiếu qua raw_rows). "tndn" bắt
# buộc trong "lntt" để KHÔNG trúng nhầm dòng "...TRỰC TIẾP TRƯỚC THUẾ TNDN" (mã 11, đứng TRƯỚC dòng
# đúng trong sheet) — 2 dòng đều chứa "loi nhuan" + "truoc thue" nhưng chỉ dòng EBT có liền cụm
# "loi nhuan truoc thue tndn" (chữ "truc tiep" chen giữa ở dòng kia phá vỡ tính liền mạch).
_KQKD_ANCHOR = {
    "dt_thuan": "3. doanh thu thuan",
    "gia_von": "iv. gia von hang ban",
    "ln_gop": "v. loi nhuan gop",
    "cp_bien_doi": "vi. chi phi bien doi",
    "cp_co_dinh": "viii. chi phi co dinh",
    "cp_tai_chinh": "2. chi phi tai chinh",
    "cp_khac": "2. chi phi khac",
    "pb_chung": "xii. phan bo chi phi chung",
    "lntt": ("loi nhuan truoc thue tndn",),
    "lnst": ("loi nhuan sau thue",),
}
# Cấu phần tổng chi phí -> (nhóm CP chuẩn mực, nhãn hiển thị). Nhóm khớp bản THÁNG của cùng đơn vị
# (raw_rows CHIPHI của XVP chỉ có 4 nhóm: Giá vốn / Chi phí tài chính / Chi phí bán hàng / Chi phí
# QLDN) để biểu đồ cơ cấu chi phí đọc được như nhau ở cả 2 chế độ; nhãn gốc giữ ở dim3.
_KQKD_CP = [
    ("gia_von", "Giá vốn hàng bán", "Giá vốn hàng bán"),
    ("cp_bien_doi", "Chi phí bán hàng", "Chi phí biến đổi"),
    ("cp_co_dinh", "Chi phí QLDN", "Chi phí cố định"),
    ("cp_tai_chinh", "Chi phí tài chính", "Chi phí tài chính"),
    ("cp_khac", "Chi phí khác", "Chi phí khác"),
    ("pb_chung", "Chi phí QLDN", "Phân bổ chi phí chung"),
]


def _kqkd_day_sheets(wb, period):
    y, mm = int(period[:4]), int(period[5:7])
    out = []
    for s in wb.sheetnames:
        t = s.strip()
        if t.isdigit() and 1 <= int(t) <= 31:
            out.append((s, f"{y:04d}-{mm:02d}-{int(t):02d}"))
    return sorted(out, key=lambda x: x[1])


def _kqkd_scan(rows, ccs, anchors):
    """Dò header + cột giá trị + neo dòng chỉ tiêu cho layout dạng KQKD (dùng cho cả 'kqkd' và
    'antaxi'). -> (cols, anchored, val) ; cols=[] nếu sai layout."""
    hdr_i = next((i for i, r in enumerate(rows[:10])
                  if any(_nd(c) == "chi tieu" for c in r if c is not None)), None)
    if hdr_i is None:
        return [], {}, None
    hdr = rows[hdr_i]
    ten_j = next(j for j, c in enumerate(hdr) if _nd(c) == "chi tieu")
    # Cột giá trị: ƯU TIÊN các cột cost center (Depot/HO/tỉnh). CỐ Ý BỎ cột "Tổng cộng" khi đã có
    # cột cost center — Σ cost center khớp ĐÚNG cột tổng (verify XVP 01/08: 283.756.022 +
    # 480.309.773 + 140.030.764 = 904.096.560; An Taxi 01/08 DT thuần: 76.225.625 + 34.226.640 =
    # 110.452.265), nên ghi thêm dòng tổng chỉ tạo nguy cơ đếm đôi ở các truy vấn KHÔNG đi qua
    # repository._per_file_resolved (daily_series/sum_by_dim1 — biểu đồ xu hướng ngày).
    # Đơn vị 1 cột (HTX, không có cost center) -> dùng cột tổng / cột ngay sau "MÃ SỐ".
    cols = [(cc, j) for j, c in enumerate(hdr) if j != ten_j
            for cc in [next((cc for kw, cc in ccs if _nd(c) == kw), None)] if cc]
    if not cols:
        tong_j = next((j for j, c in enumerate(hdr) if _nd(c).startswith("tong cong")), None)
        if tong_j is None:
            ma_j = next((j for j, c in enumerate(hdr) if _nd(c) == "ma so"), ten_j)
            tong_j = ma_j + 1
        cols = [(None, tong_j)]

    anchored = {}
    for r in rows[hdr_i + 1:]:
        if not r or ten_j >= len(r):
            continue
        n = _nd(r[ten_j])
        for key, pref in anchors.items():
            if key in anchored:
                continue
            # pref là tuple 1 phần tử -> khớp CONTAINS (bất kể tiền tố La Mã đứng trước, xem
            # comment `_KQKD_ANCHOR` lntt/lnst); chuỗi thường -> khớp STARTSWITH như cũ.
            hit = (pref[0] in n) if isinstance(pref, tuple) else n.startswith(pref)
            if hit:
                anchored[key] = r

    def val(key, j):
        r = anchored.get(key)
        return _num(r[j]) if r is not None and j < len(r) else None

    return cols, anchored, val


def _kqkd_facts(rows):
    """rows -> [(cost_center|None, report_type, dim1, dim3, value_VND)]. [] nếu sai layout."""
    cols, anchored, val = _kqkd_scan(rows, _CC_XVP, _KQKD_ANCHOR)
    if not cols or "dt_thuan" not in anchored:
        return []

    facts = []
    for cc, j in cols:
        dt = val("dt_thuan", j)
        if dt:
            facts.append((cc, RT_HQKD, MA_DT, MA_DT, dt))
            facts.append((cc, RT_PNLT, "Doanh thu HH, DV", "Doanh thu HH, DV", dt))
            facts.append((cc, RT_DTHU, "Doanh thu thuần", "Doanh thu thuần", dt))
        tong_cp = 0.0
        for key, nhom, ten in _KQKD_CP:
            v = val(key, j)
            if not v:
                continue
            tong_cp += v
            facts.append((cc, RT_CHIPHI, nhom, ten, v))
        if tong_cp:
            facts.append((cc, RT_HQKD, MA_CP, MA_CP, tong_cp))
        for key, rt, dim1 in (("lntt", RT_HQKD, MA_LNTT), ("gia_von", RT_PNLT, "Giá vốn hàng bán"),
                              ("ln_gop", RT_PNLT, "Lợi nhuận gộp"),
                              ("lntt", RT_PNLT, "Lợi nhuận trước thuế"),
                              ("lnst", RT_PNLT, "Lợi nhuận sau thuế")):
            v = val(key, j)
            if v:
                facts.append((cc, rt, dim1, dim1, v))
    return facts


# ---------------------------------------------------------------------------------------------
# Layout "antaxi" — An Taxi (ANTAXI/baocaohqkdngay), mỗi ngày 1 sheet "1".."31" + sheet "TH"
# ---------------------------------------------------------------------------------------------
# KHÁC layout "kqkd" ở 2 điểm (verify file thật B.7.AAG.TCKT.M.202608, sheet 1-31, header dòng 6):
#  1. Số mục La Mã nằm ở cột "TT" RIÊNG, KHÔNG dính vào nhãn: dòng 33 = TT "III" + CHỈ TIÊU
#     "DOANH THU THUẦN" (XVP ghi "3. Doanh thu thuần" trong CÙNG 1 ô) -> neo theo NHÃN TRẦN.
#  2. Số La Mã của An Taxi lệch hẳn XVP (LNTT = XI ở An Taxi vs XIII ở XVP; LNST = XIII vs XV;
#     An Taxi KHÔNG có mục "phân bổ chi phí chung") -> phải có bảng anchor riêng, không dùng chung.
# Đã verify nhãn nào cũng khớp DUY NHẤT 1 dòng trên cả 31 sheet (các dòng con "Giá vốn bán xe…",
# "Chi phí hoạt động khác", "LỢI NHUẬN TRƯỚC KHẤU HAO…(EBITDA)" đều KHÔNG startswith nhãn neo).
_ANTAXI_ANCHOR = {
    "dt_gross": "doanh thu ban hang",              # I  (mã 100) — DT GỘP
    "giam_tru": "cac khoan giam tru doanh thu",    # II (mã 110)
    "dt_thuan": "doanh thu thuan",                 # III (mã 120)
    "gia_von": "gia von hang ban",                 # IV (mã 130)
    "ln_gop": "lai gop",                           # V  (mã 140) — An Taxi ghi "LÃI GỘP"
    "cp_bien_doi": "chi phi bien doi",             # VI (mã 150)
    "cp_co_dinh": "chi phi co dinh",               # VIII (mã 170)
    "dt_tai_chinh": "doanh thu tai chinh",         # IX.1
    "cp_tai_chinh": "chi phi tai chinh",           # IX.2 (mã 182)
    "tn_khac": "thu nhap khac",                    # X.1
    "cp_khac": "chi phi khac",                     # X.2 (mã 192)
    "lntt": "loi nhuan truoc thue",                # XI (mã 200)
    "lnst": "loi nhuan sau thue",                  # XIII (mã 220)
}
# Tổng chi phí = IV + VI + VIII + IX.2 + X.2 (đúng công thức cột "Công thức" của mapping An Taxi;
# file KHÔNG in dòng tổng chi phí nào để đối chiếu). Verify vòng kín theo LNTT có sẵn trong file —
# 01/08 Sơn Tây: 76.225.625 (DTT) − 93.342.155 (Σ 5 cấu phần) + 2.965.749 (thu nhập khác) + 0
# (DT tài chính) = −14.150.781 = ĐÚNG dòng "XI. LỢI NHUÂN TRƯỚC THUẾ TNDN (EBT)".
# dim1 (nhóm CP) lấy Y HỆT bản THÁNG của CHÍNH An Taxi — raw_rows CHIPHI T6 có đúng 5 nhóm:
# Giá vốn hàng bán / Chi phí biến đổi / Chi phí cố định / Chi phí tài chính / Chi phí khác
# (KHÁC XVP dùng 'Chi phí bán hàng'/'Chi phí QLDN') để biểu đồ cơ cấu chi phí đọc được như nhau
# ở cả 2 chế độ Ngày/Tháng.
_ANTAXI_CP = [
    ("gia_von", "Giá vốn hàng bán", "Giá vốn hàng bán"),
    ("cp_bien_doi", "Chi phí biến đổi", "Chi phí biến đổi"),
    ("cp_co_dinh", "Chi phí cố định", "Chi phí cố định"),
    ("cp_tai_chinh", "Chi phí tài chính", "Chi phí tài chính"),
    ("cp_khac", "Chi phí khác", "Chi phí khác"),
]


def _antaxi_facts(rows):
    """rows -> [(cost_center, report_type, dim1, dim3, value_VND)]. [] nếu sai layout."""
    cols, anchored, val = _kqkd_scan(rows, _CC_ANTAXI, _ANTAXI_ANCHOR)
    if not cols or "dt_thuan" not in anchored:
        return []

    facts = []
    for cc, j in cols:
        dt = val("dt_thuan", j)
        if dt:
            facts.append((cc, RT_HQKD, MA_DT, MA_DT, dt))
            facts.append((cc, RT_DTHU, "Doanh thu thuần", "Doanh thu thuần", dt))
        tong_cp = 0.0
        for key, nhom, ten in _ANTAXI_CP:
            v = val(key, j)
            if not v:
                continue
            tong_cp += v
            facts.append((cc, RT_CHIPHI, nhom, ten, v))
        if tong_cp:
            facts.append((cc, RT_HQKD, MA_CP, MA_CP, tong_cp))
        # PNLT: giữ ĐÚNG bộ dim1 bản THÁNG của An Taxi. Lưu ý An Taxi là đơn vị DUY NHẤT có dòng
        # giảm trừ tách riêng, và bản tháng ghi CẢ HAI tên cho DT GỘP ("Doanh thu bán hàng và cung
        # cấp dịch vụ" = tên đích danh cho bảng "Cấu trúc Doanh thu" ở revenue.py::_gross_of, và
        # "Doanh thu HH, DV" = tên chung cho KPI #1 ở overview.py) — cùng 1 giá trị, KHÔNG cộng đôi
        # vì 2 truy vấn lọc dim1_ilike khác nhau. Bỏ 1 trong 2 là mất số ở đúng 1 trong 2 chỗ đó.
        # ⚠ "Doanh thu HH, DV" của An Taxi là DT GỘP (mã 100), KHÁC XVP/HTX (= DT thuần).
        for key, rt, dim1 in (("lntt", RT_HQKD, MA_LNTT),
                              ("dt_gross", RT_PNLT, "Doanh thu bán hàng và cung cấp dịch vụ"),
                              ("dt_gross", RT_PNLT, "Doanh thu HH, DV"),
                              ("giam_tru", RT_PNLT, "Các khoản giảm trừ doanh thu"),
                              ("gia_von", RT_PNLT, "Giá vốn hàng bán"),
                              ("ln_gop", RT_PNLT, "Lợi nhuận gộp"),
                              ("dt_tai_chinh", RT_PNLT, "Doanh thu tài chính"),
                              ("tn_khac", RT_PNLT, "Thu nhập khác"),
                              ("lntt", RT_PNLT, "Lợi nhuận trước thuế"),
                              ("lnst", RT_PNLT, "Lợi nhuận sau thuế")):
            v = val(key, j)
            if v:
                facts.append((cc, rt, dim1, dim1, v))
    return facts


# ---------------------------------------------------------------------------------------------
# Layout "tcode" — Global AI (GLOBALAI/baocaohqkdngay), mỗi ngày 1 sheet "D1".."D31"
# ---------------------------------------------------------------------------------------------
# Neo theo MÃ SỐ T-series (cột A, y hệt sheet THÁNG "HQKD HỢP NHẤT GA") — KHÁC layout "kqkd"/
# "antaxi" (neo theo NHÃN đã chuẩn hoá) vì GA không bị trùng mã như HTX/XVP, xem docstring đầu file.
# T201/T202/T203/T204 = CON TRỰC TIẾP của T200 (Σ 4 dòng = T200, verify sheet THÁNG khi có số thật)
# — cùng 4 nhóm CP mà bản THÁNG của GA hiện dùng qua sheet TT200 'KQKD' (mã 11/22/25/26 -> nhãn
# 'Giá vốn hàng bán'/'Chi phí tài chính'/'Chi phí bán hàng'/'Chi phí quản lý doanh nghiệp') để biểu
# đồ cơ cấu chi phí đọc được như nhau ở cả 2 chế độ Ngày/Tháng — xem agent_cli._TT200_CHITIEU.
# Chuẩn hoá nhãn Y HỆT `agent_cli._derive_kqkd_tseries._canon` (lý do xem trong `_tcode_facts`).
_TCODE_CANON = {"T201": "Giá vốn hàng bán", "T103": "Thu nhập khác"}


def _tcode_day_sheets(wb, period):
    """[(sheet_name, 'YYYY-MM-DD')] — sheet 'D1'..'D31' (file LUỸ KẾ: số sheet tăng dần theo ngày
    đã qua trong tháng, tên sheet KHÔNG zero-pad)."""
    y, mm = int(period[:4]), int(period[5:7])
    out = []
    for s in wb.sheetnames:
        m = re.fullmatch(r"[Dd](\d{1,2})", s.strip())
        if m and 1 <= int(m.group(1)) <= 31:
            out.append((s, f"{y:04d}-{mm:02d}-{int(m.group(1)):02d}"))
    return sorted(out, key=lambda x: x[1])


def _tcode_facts(rows, profit_pnlt=("Lợi nhuận trước thuế",), pnlt_skip=()):
    """rows -> [(None, report_type, dim1, dim3, value_VND)] cho 1 ngày. [] nếu sai layout.
    Cột giá trị dò theo header khớp 'D\\d{2}' (zero-pad) — TỰ suy ra cột đúng của sheet này, không
    cần biết trước số ngày, vì mỗi sheet chỉ có DUY NHẤT 1 cột như vậy.

    `profit_pnlt`: (các) tên dim1 PNLT ghi từ T300 — THAM SỐ HOÁ vì 2 đơn vị dùng chung mã T-series
    nhưng quy ước LNTT/LNST bản THÁNG KHÁC NHAU (xem docstring khối 'tcode' + đơn vị TRAMSAC):
    GA ghi 'Lợi nhuận trước thuế' (chưa chắc = LNST thật khi phát sinh thuế); Trạm sạc bản THÁNG
    CHỈ có đúng 1 dim1 PNLT 'Lợi nhuận sau thuế' (không có 'Lợi nhuận trước thuế' riêng, verify DB
    2026-06: PNLT 'Lợi nhuận sau thuế' = HQKD 1112 = -0,3844, bằng nhau tuyệt đối -> T300 ở Trạm
    sạc ĐÃ LÀ LNST, gán nhãn 'Lợi nhuận trước thuế' cho nó sẽ SAI tên cột trên bảng Cấu trúc CP."""
    hdr_i = next((i for i, r in enumerate(rows[:10]) if any(_nd(c) == "ma so" for c in r if c is not None)), None)
    if hdr_i is None:
        return []
    hdr = rows[hdr_i]
    ma_j = next(j for j, c in enumerate(hdr) if _nd(c) == "ma so")
    ten_j = next((j for j, c in enumerate(hdr) if _nd(c).startswith("chi tieu")), ma_j + 1)
    val_j = next((j for j, c in enumerate(hdr) if j not in (ma_j, ten_j)
                  and re.fullmatch(r"d\d{2}", _nd(c))), ten_j + 1)

    byco = {}
    for r in rows[hdr_i + 1:]:
        c = str(r[ma_j]).strip() if ma_j < len(r) and r[ma_j] not in (None, "") else ""
        if re.fullmatch(r"T\d{3}", c) and c not in byco:
            lab = str(r[ten_j]).strip() if ten_j < len(r) and r[ten_j] not in (None, "") else c
            byco[c] = (lab, _num(r[val_j]) if val_j < len(r) else None)
    if "T101" not in byco or "T200" not in byco or "T300" not in byco:
        return []

    def val(code):
        return byco.get(code, (None, None))[1]

    facts = []
    dt = val("T101")
    if dt:
        facts.append((None, RT_HQKD, MA_DT, MA_DT, dt))
        facts.append((None, RT_PNLT, "Doanh thu HH, DV", "Doanh thu HH, DV", dt))
        facts.append((None, RT_DTHU, "Doanh thu thuần", "Doanh thu thuần", dt))
    if val("T200"):
        facts.append((None, RT_HQKD, MA_CP, MA_CP, val("T200")))
    lntt = val("T300")
    if lntt:
        facts.append((None, RT_HQKD, MA_LNTT, MA_LNTT, lntt))
        for nm in profit_pnlt:
            facts.append((None, RT_PNLT, nm, nm, lntt))
    # Lợi nhuận gộp = T100 − T201 — BỔ SUNG 2026-08-06. Bản THÁNG của CẢ HAI đơn vị đều có dim1 này
    # (verify DB: Trạm sạc T07 'Lợi nhuận gộp' = 1,297842 · GA T06 có 'Lợi nhuận gộp'), bản ngày
    # trước đây thiếu -> thẻ LN gộp trống ở chế độ Ngày. LƯU Ý: quyết định cũ "GA không suy diễn
    # LNST/LN gộp" chỉ đúng cho LNST (T300 chưa chắc đã trừ thuế TNDN — GA bản tháng có dòng 'Chi
    # phí thuế TNDN' riêng); LN gộp = T100 − T201 KHÔNG liên quan thuế nên áp cho cả hai.
    #
    # SỬA 11/08/2026 — lấy T101 (doanh thu thuần) chứ KHÔNG phải T100. `T100 = T101 + T102 (DT tài
    # chính) + T103 (thu nhập khác)`; chính hàm này đã dùng `dt = val("T101")` cho doanh thu ở trên,
    # riêng dòng lãi gộp lấy T100 -> trộn doanh thu tài chính vào LÃI GỘP. Bản THÁNG
    # (`agent_cli._derive_kqkd_tseries`) sửa cùng lượt để hai bản không lệch định nghĩa.
    if val("T201") is not None and dt is not None and (dt - val("T201")):
        facts.append((None, RT_PNLT, "Lợi nhuận gộp", "Lợi nhuận gộp", dt - val("T201")))
    # MỌI mã còn lại -> PNLT giữ NHÃN GỐC, chuẩn hoá 2 nhãn y hệt bản THÁNG
    # (`agent_cli._derive_kqkd_tseries._canon`): T201 -> 'Giá vốn hàng bán' (nguồn gõ 'Gía vốn' làm
    # metrics lọc ILIKE '%giá vốn%' ACCENT-SENSITIVE trượt) · T103 -> 'Thu nhập khác'.
    # `pnlt_skip`: GA bản tháng chạy extractor TT200 (KHÔNG phải T-series) nên KHÔNG có dim1
    # 'Doanh thu bán hàng' (T101 nhãn gốc) — thêm vào sẽ đẻ chỉ tiêu bản tháng không có, trùng giá
    # trị với 'Doanh thu HH, DV'. Trạm sạc bản tháng CÓ dòng đó -> không skip.
    for code, (lab, v) in byco.items():
        if code in ("T100", "T200", "T300") or code in pnlt_skip or not v:
            continue
        ten = _TCODE_CANON.get(code, lab)
        facts.append((None, RT_PNLT, ten, ten, v))
    # CHIPHI: mã con TRỰC TIẾP của T200 = '^T2\d{2}$' trừ T200, NHÃN GỐC (bản tháng dùng nhãn gốc —
    # verify DB Trạm sạc T07: CHIPHI 'Gía vốn hàng bán' GIỮ typo, trong khi PNLT là bản chuẩn hoá;
    # bản ngày trước đây hardcode nhãn chuẩn -> cùng 1 khoản bị tách 2 nhóm khác tên giữa Tháng/Ngày).
    # Chỉ dựng khi Σ con PHỦ HẾT T200 (1%) — y hệt điều kiện `_covers` bản tháng, tránh breakdown thiếu.
    cp = [(byco[c][0], byco[c][1]) for c in byco
          if re.fullmatch(r"T2\d{2}", c) and c != "T200" and byco[c][1]]
    if cp and val("T200") and abs(sum(x[1] for x in cp) - val("T200")) <= abs(val("T200")) * 0.01:
        for nhom, v in cp:
            facts.append((None, RT_CHIPHI, nhom, nhom, v))
    return facts


# ---------------------------------------------------------------------------------------------
# Layout "ht" — XE TẢI HƯNG THỊNH (HUNGTHINH/baocaohqkdngay/B5.HT.TCTC.M.<YYYYMM>.baocaongay.xlsx)
# ---------------------------------------------------------------------------------------------
# Mỗi ngày 1 sheet "{dd}.{mm}" ("01.08".."04.08" — spec user viết "sheet 01, 02, 03..." nên
# `_ht_day_sheets` nhận CẢ dạng số trần "01".."31"). Ô lấy số theo spec user 2026-08-06 (đã đối
# chiếu file thật T08, cột E = "Ngày {dd}" = tổng của F "Hưng Thịnh" + G "Thịnh Cường"):
#   Doanh thu HH, DV = E8+E9+E10+E11   ·   Tổng chi phí = E14   ·   Giá vốn = E15
#   CP tài chính = E21   ·   CP vận hành = E22   ·   LNTT = E41   ·   LNST = E43
# KHÔNG hardcode số dòng: dò theo MÃ ở cột mã số (T101.x/T200/T201/T202/T203/T300), vì mã ổn định
# còn thứ tự dòng đổi theo tháng (T203.1..T203.18 co giãn theo khoản mục kế toán phát sinh).
# ⚠ 3 điểm layout HT KHÁC hẳn "tcode" (GA/Trạm sạc) nên KHÔNG dùng lại `_tcode_facts` được:
#   1. Header cột mã số ghi "v" (không phải "Mã số") -> `_tcode_facts` không tìm ra header, chết
#      ngay bước đầu. Ở đây dò cột mã theo NỘI DUNG (cột nào chứa T100/T200/T300).
#   2. KHÔNG có mã "T101" tổng — doanh thu chỉ nằm ở 4 mã con T101.1..T101.4 (đúng spec E8..E11);
#      `_tcode_facts` đòi có "T101" nên sẽ trả [] cho mọi sheet.
#   3. LNST (E43) là dòng KHÔNG CÓ MÃ -> neo theo nhãn "Lợi nhuận sau thuế" ở cột Chỉ tiêu.
# dim1 giữ ĐÚNG tên bản THÁNG của HT đang dùng (verify raw_rows: 'Doanh thu HH, DV' / 'Giá vốn
# hàng bán' / 'Chi phí tài chính' / 'Chi phí vận hành' / 'Lợi nhuận sau thuế'; HQKD dùng mã
# 1000/1047/1112) để dòng NGÀY và dòng THÁNG khớp nhãn trên cùng một bảng.
# KHÔNG dựng "Lợi nhuận gộp"/các tỷ lệ: spec ghi rõ chúng "Tính toán trên Dashboard".
# Chuẩn hoá nhãn Y HỆT `agent_cli._derive_kqkd_tseries._canon` (2 mã, lý do xem trong `_ht_facts`).
_HT_CANON = {"T201": "Giá vốn hàng bán", "T103": "Thu nhập khác"}


def _ht_day_sheets(wb, period):
    """[(sheet_name, 'YYYY-MM-DD')] — sheet "{dd}.{mm}" (khớp ĐÚNG tháng của kỳ) hoặc "{dd}" trần.
    DÙNG CHUNG cho layout "ht" và "xdv" (cả 2 spec đều ghi "sheet 01, 02, 03... tương đương số ngày
    trong tháng", file thật đặt "01.08".."04.08")."""
    y, mm = int(period[:4]), int(period[5:7])
    out = []
    for s in wb.sheetnames:
        t = s.strip()
        m = re.fullmatch(r"(\d{1,2})(?:\.(\d{1,2}))?", t)
        if not m or not 1 <= int(m.group(1)) <= 31:
            continue
        if m.group(2) and int(m.group(2)) != mm:      # "05.07" = ngày 5 THÁNG 7 -> khác kỳ, bỏ
            continue
        out.append((s, f"{y:04d}-{mm:02d}-{int(m.group(1)):02d}"))
    return sorted(out, key=lambda x: x[1])


def _ht_facts(rows):
    """rows -> [(None, report_type, dim1, dim3, value_VND)] cho 1 ngày. [] nếu sai layout."""
    hdr_i = next((i for i, r in enumerate(rows[:10])
                  if any(_nd(c) == "chi tieu" for c in r if c is not None)), None)
    if hdr_i is None:
        return []
    hdr = rows[hdr_i]
    ten_j = next(j for j, c in enumerate(hdr) if _nd(c) == "chi tieu")
    # Cột giá trị: header "Ngày {dd}"; fallback = cột ngay sau "Chỉ tiêu" (cột E ở file mẫu).
    val_j = next((j for j, c in enumerate(hdr) if j > ten_j and re.fullmatch(r"ngay ?\d{0,2}", _nd(c))),
                 ten_j + 1)
    # Cột mã số: header ghi "v" (vô nghĩa) -> tìm cột NÀO chứa các mã T-series trong phần dữ liệu.
    body = rows[hdr_i + 1:]
    ma_j = next((j for j in range(0, ten_j)
                 if sum(1 for r in body if j < len(r) and re.fullmatch(r"T\d{3}(\.\d+)?",
                                                                      str(r[j] or "").strip())) >= 3), None)
    if ma_j is None:
        return []

    # byco: mã -> (nhãn gốc, giá trị). Nhận CẢ mã con "T203.7" (bản tháng cũng gom vậy).
    byco, lnst = {}, None
    for r in body:
        v = _num(r[val_j]) if val_j < len(r) else None
        code = str(r[ma_j] or "").strip() if ma_j < len(r) else ""
        if re.fullmatch(r"T\d{3}(\.\d+)?", code) and code not in byco:
            lab = str(r[ten_j]).strip() if ten_j < len(r) and r[ten_j] not in (None, "") else code
            byco[code] = (lab, v)
        if lnst is None and ten_j < len(r) and _nd(r[ten_j]).startswith("loi nhuan sau thue"):
            lnst = v
    if "T200" not in byco or "T300" not in byco:
        return []

    def val(code):
        return byco.get(code, (None, None))[1]

    facts = []
    # "Doanh thu thuần" = Σ T101.x (hoặc T101 gộp nếu có), KHÔNG phải T100 — Y HỆT bản THÁNG
    # (agent_cli._derive_kqkd_tseries, chốt user 2026-07-19 đối chiếu file gốc T05: T100 = ΣT101.x +
    # T102 DT tài chính + T103 thu nhập khác, nên T100 THỪA cho chỉ tiêu "doanh thu thuần"). Sheet
    # NGÀY của HT không có dòng T101 gộp, chỉ có T101.1..T101.4 = đúng "E8+E9+E10+E11" của spec.
    t101 = [v for c, (_, v) in byco.items() if re.fullmatch(r"T101\.\d+", c) and v is not None]
    dt = val("T101") if val("T101") is not None else (sum(t101) if t101 else val("T100"))
    if dt:
        facts.append((None, RT_HQKD, MA_DT, MA_DT, dt))
        facts.append((None, RT_DTHU, "Doanh thu thuần", "Doanh thu thuần", dt))
        facts.append((None, RT_PNLT, "Doanh thu HH, DV", "Doanh thu HH, DV", dt))
    if val("T200"):
        facts.append((None, RT_HQKD, MA_CP, MA_CP, val("T200")))
    if val("T300"):
        facts.append((None, RT_HQKD, MA_LNTT, MA_LNTT, val("T300")))
    # LNST: bản THÁNG gán = T300 vì sheet tháng KHÔNG có dòng LNST/thuế TNDN riêng. Sheet NGÀY thì CÓ
    # (dòng "LỢI NHUẬN SAU THUẾ TNDN" = E43, ngay dưới dòng "Thuế TNDN") -> ưu tiên dòng thật, chỉ
    # rơi về T300 khi thiếu. Hôm nay thuế = 0 nên 2 cách ra cùng số; khi HT phát sinh thuế thật thì
    # dòng thật MỚI đúng, còn T300 sẽ là trước thuế.
    lnst = lnst if lnst is not None else val("T300")
    if lnst:
        facts.append((None, RT_PNLT, "Lợi nhuận sau thuế", "Lợi nhuận sau thuế", lnst))
    # Lợi nhuận gộp = DOANH THU THUẦN (`dt` = Σ T101.x, tính ở trên) − giá vốn T201.
    # SỬA 11/08/2026 — trước lấy T100, tức cộng luôn T102 DT tài chính + T103 thu nhập khác vào lãi
    # gộp. Chú thích cũ ghi "bản tháng dùng T100 ở ĐÚNG chỗ này" là MÔ TẢ SAI: bản tháng cũng sai y
    # hệt và đã sửa cùng lượt (agent_cli._derive_kqkd_tseries). Đo kỳ 06/2026: HT 4,535576 -> đúng
    # 4,189939 tỷ, thổi lên 345,6 triệu.
    if val("T201") is not None and dt is not None and (dt - val("T201")):
        facts.append((None, RT_PNLT, "Lợi nhuận gộp", "Lợi nhuận gộp", dt - val("T201")))
    # MỌI mã còn lại -> PNLT giữ NHÃN GỐC (T101.x dòng xe · T102 · T201.x giá vốn chi tiết · T202 ·
    # T203 + T203.x 18 khoản chi phí vận hành). Chuẩn hoá 2 nhãn y hệt bản tháng: T201 -> 'Giá vốn
    # hàng bán' (nguồn HT gõ 'Gía vốn' — dấu sắc trên i — làm metrics.build_revenue lọc
    # ILIKE '%giá vốn%' ACCENT-SENSITIVE trượt) và T103 -> 'Thu nhập khác' (HT gõ 'Doanh thu khác').
    # Mã CON giữ nguyên nhãn typo -> KHÔNG khớp filter -> KHÔNG đếm đôi với dòng tổng đã chuẩn hoá.
    for code, (lab, v) in byco.items():
        if code in ("T100", "T200", "T300") or not v:
            continue
        ten = _HT_CANON.get(code, lab)
        facts.append((None, RT_PNLT, ten, ten, v))
    # 02_CHIPHI: các mã CON TRỰC TIẾP của T200 = '^T2\d{2}$' trừ T200 (T201 giá vốn · T202 CP tài
    # chính · T203 CP vận hành; T2xx.y là con của chúng -> loại, tránh cộng trùng). Chỉ dựng khi Σ
    # con PHỦ HẾT T200 (sai số 1%) — y hệt điều kiện `_covers` của bản tháng, tránh breakdown thiếu.
    cp = [(c, byco[c][0], byco[c][1]) for c in byco
          if re.fullmatch(r"T2\d{2}", c) and c != "T200" and byco[c][1]]
    if cp and val("T200") and abs(sum(x[2] for x in cp) - val("T200")) <= abs(val("T200")) * 0.01:
        for _c, nhom, v in cp:
            facts.append((None, RT_CHIPHI, nhom, nhom, v))
    return facts


# ---------------------------------------------------------------------------------------------
# Layout "xdv" — XƯỞNG DỊCH VỤ VINFAST (XDV/baocaohqkdngay/B.2.TC.TCKT.M.<YYYYMM>.BaocaoHQKD.xlsx)
# ---------------------------------------------------------------------------------------------
# Mỗi ngày 1 sheet "{dd}.{mm}". Header dòng 8: A "Mã số" · B "Chỉ tiêu" · C "Kỳ này" (TỔNG) ·
# D..H (TK nợ/TK có/Mã phí/Mã NS/Công thức, BỎ QUA) · I→V = 14 cost center.
# Mã B-series (spec user 2026-08-06): B100 DT · B210 DT thuần · B300 giá vốn · B410 lãi gộp ·
# B500 tổng CP xưởng (= B600 nhân sự + B700 hoạt động) · B810 CP cố định · B822 lãi vay ·
# B833 CP khác · B840 LNTT · B900 LNST.
#
# ⚠ LỆCH GIỮA 2 CỘT CỦA CHÍNH SPEC — "Tổng chi phí": cột "Cách lấy (nguyên văn)" ghi
# `B500+B810+B822+B833` (KHÔNG có giá vốn), nhưng cột "Công thức" ngay bên cạnh lại ghi
# "Tổng CP = **Giá vốn** + CP QLDN + CP bán hàng + CP khác + CP tài chính" (CÓ giá vốn). Bản THÁNG
# của XDV (agent_cli `_kqkd_recs_xdv`/`_chiphi_recs_xdv`) đã chốt 2026-07-17 là **GỒM B300**, có
# audit đối soát BCHN Tập đoàn (37,71 khớp hợp nhất) và để `byNhom == byKhoi`. Deriver NGÀY theo
# ĐÚNG bản tháng (1047 = B300+B500+B810+B822+B833): nếu bỏ B300 thì cùng thẻ "Chi phí" sẽ mang 2
# nghĩa khác nhau khi bấm qua lại Tháng/Ngày (ngày ~0,2 tỷ vs tháng ~13,3 tỷ) — sai nghiêm trọng
# hơn nhiều so với việc lệch chữ với cột "nguyên văn".
#
# Mã cost center lấy Y HỆT bản tháng (`agent_cli._XDV_BRANCH_CC`, khớp trust_me_bro.xlsx) — dò theo
# TỪ KHOÁ trong tên cột, không theo vị trí. Cả 14 đều pháp nhân TC (không có case chéo pháp nhân
# như SRVF UB_SR) nên không cần `_CC_CONGTY`. Verify file T08: 14/14 cột nhận diện được và
# **Σ 14 CC = ĐÚNG cột C** ở cả 10 mã × 3 ngày -> ghi THEO CC (không ghi thêm dòng tổng, tránh
# đếm đôi; tổng khối = Σ cost center, đúng nhánh cc_v của repository._per_file_resolved).
_XDV_CC = [
    ("ocean park", "OCP_XDV"), ("long bien", "LB_XDV"), ("smart city", "SMC_XDV"),
    ("ha long", "HL_XDV"), ("cam pha", "CP_XDV"), ("xuan mai", "XM_XDV"),
    ("uong bi", "UB_XDV"), ("tuyen quang", "TQ_XDV"), ("vinh phuc", "VP_XDV"),
    ("son tay", "ST_XDV"), ("dai tu", "ĐT_XDV"), ("viet tri", "VT_XDV"),
    ("ha khanh", "HK_XDV"), ("ho chi minh", "HCM_XDV"),
]
_XDV_CP_CODES = ("B300", "B500", "B810", "B822", "B833")     # cấu thành 1047 (GỒM giá vốn B300)


def _xdv_facts(rows):
    """rows -> [(cost_center, report_type, dim1, dim3, value_VND)] cho 1 ngày. [] nếu sai layout."""
    hdr_i = next((i for i, r in enumerate(rows[:14])
                  if any(_nd(c) == "ma so" for c in r if c is not None)), None)
    if hdr_i is None:
        return []
    hdr = rows[hdr_i]
    ma_j = next(j for j, c in enumerate(hdr) if _nd(c) == "ma so")
    ten_j = next((j for j, c in enumerate(hdr) if _nd(c).startswith("chi tieu")), ma_j + 1)
    cc_cols = [(cc, j) for kw, cc in _XDV_CC
               for j in [next((k for k, c in enumerate(hdr) if isinstance(c, str) and kw in _nd(c)),
                              None)] if j is not None]
    if len(cc_cols) < 10:        # thiếu chi nhánh = đổi layout -> báo sai layout, KHÔNG ghi số lệch
        return []

    byco = {}
    for r in rows[hdr_i + 1:]:
        c = str(r[ma_j] or "").strip() if ma_j < len(r) else ""
        if re.fullmatch(r"B\d{3}", c) and c not in byco:
            byco[c] = r
    if "B840" not in byco or "B300" not in byco:
        return []

    def v(code, j):
        r = byco.get(code)
        return _num(r[j]) if r is not None and j < len(r) else None

    def vsum(codes, j):
        xs = [v(c, j) for c in codes]
        return sum(x for x in xs if x) if any(x is not None for x in xs) else None

    facts = []
    for cc, j in cc_cols:
        dt = v("B210", j) if v("B210", j) is not None else v("B100", j)
        for rt, dim1, val in (
                (RT_HQKD, MA_DT, dt),
                (RT_DTHU, "Doanh thu thuần", dt),
                (RT_HQKD, MA_CP, vsum(_XDV_CP_CODES, j)),
                (RT_HQKD, MA_LNTT, v("B840", j)),
                (RT_PNLT, "Doanh thu thuần", dt),
                (RT_PNLT, "Giá vốn hàng bán", v("B300", j)),
                (RT_PNLT, "Lợi nhuận gộp", v("B410", j)),
                (RT_PNLT, "Lợi nhuận sau thuế", v("B900", j))):
            if val:
                facts.append((cc, rt, dim1, dim1, val))
        # Cơ cấu chi phí (spec #10) — tách B500 -> B600 nhân sự + B700 hoạt động NẾU cộng khớp,
        # y hệt `agent_cli._chiphi_recs_xdv`; không khớp thì để nguyên B500.
        b500, b600, b700 = v("B500", j), v("B600", j), v("B700", j)
        cp_groups = [("B300", "Giá vốn hàng bán")]
        if None not in (b500, b600, b700) and abs((b600 + b700) - b500) < 1000:
            cp_groups += [("B600", "Chi phí nhân sự"), ("B700", "Chi phí hoạt động xưởng")]
        else:
            cp_groups += [("B500", "Chi phí xưởng dịch vụ")]
        cp_groups += [("B810", "Chi phí cố định"), ("B822", "Chi phí tài chính"),
                      ("B833", "Chi phí khác")]
        for code, nhom in cp_groups:
            x = v(code, j)
            if x:
                ten = str(byco[code][ten_j]).strip() if ten_j < len(byco[code]) and byco[code][ten_j] else nhom
                facts.append((cc, RT_CHIPHI, nhom, ten, x))
    return facts


# ---------------------------------------------------------------------------------------------
# Layout "duan" — Khối Dự án (DUAN/baocaohqkdngay), mỗi ngày 1 sheet "1".."31"
# ---------------------------------------------------------------------------------------------
# Cột cost center dò theo TỪ KHOÁ (chứa, không cần khớp hệt) vì nhãn cột đổi nhẹ theo tháng (vd
# "Yên Bình 3"). Mã CC lấy Y HỆT bản THÁNG (agent_cli._DA_PROJECT_CC) — kể cả quy ước NGƯỢC viết
# tắt Tân Thịnh<->Yên Bình đã xác nhận nguồn, xem docstring đầu file. "Bình phước" mới, chưa có
# trong master_data -> mã tự đặt (giống Núi Pháo/Quảng Ngãi bản tháng).
_CC_DUAN = [("cao bang", "CB_DA"), ("tan thinh", "YB_DA"), ("lang son", "LS_DA"),
            ("yen binh", "TT_DA"), ("phu quoc", "PQ_DA"), ("quang son", "QS_DA"),
            ("nui phao", "NUIPHAO_DA"), ("quang ngai", "QUANGNGAI_DA"), ("tho chu", "TC_DA"),
            ("binh phuoc", "BINHPHUOC_DA")]

# (khoá -> (nhãn chuẩn hoá, exact?)). Đa số EXACT (không startswith) vì nhãn ngắn dễ bị dòng con
# "nuốt" nhầm — vd "chi phi khac" (mục X.2, mã neo cp_khac) là PREFIX của "Chi phí khác tại dự
# án" (mục con 1.7 của Giá vốn IV, đứng TRƯỚC trong sheet) -> startswith sẽ khoá nhầm dòng đó.
# Chỉ lntt/lnst dùng startswith vì nhãn gốc có hậu tố đổi được "(EBT)"/"(EAT)".
_DUAN_ANCHOR = {
    "dt_thuan": ("doanh thu thuan", True),        # III
    "gia_von": ("gia von hang ban", True),        # IV
    "ln_gop": ("lai gop", True),                  # V (nhãn gốc "Lãi gộp")
    "cp_bien_doi": ("chi phi bien doi", True),    # VI
    "cp_co_dinh": ("chi phi co dinh", True),      # VIII
    "cp_tai_chinh": ("chi phi tai chinh", True),  # IX.2
    "cp_khac": ("chi phi khac", True),            # X.2
    "lntt": ("loi nhuan truoc thue", False),      # XI "...(EBT)"
    "lnst": ("loi nhuan sau thue", False),        # XII "...(EAT)"
}
# Tổng chi phí = IV + VI + VIII + IX.2 + X.2 (đúng công thức "E25+E46+E74+E80+E83" của Mapping) —
# 5 nhóm Y HỆT ANTAXI (không có mục "phân bổ chung" cấp I riêng, đã lồng trong "Chi phí khác").
_DUAN_CP = [
    ("gia_von", "Giá vốn hàng bán", "Giá vốn hàng bán"),
    ("cp_bien_doi", "Chi phí biến đổi", "Chi phí biến đổi"),
    ("cp_co_dinh", "Chi phí cố định", "Chi phí cố định"),
    ("cp_tai_chinh", "Chi phí tài chính", "Chi phí tài chính"),
    ("cp_khac", "Chi phí khác", "Chi phí khác"),
]


def _duan_facts(rows):
    """rows -> [(cost_center, report_type, dim1, dim3, value_VND)]. [] nếu sai layout.

    KHÁC mọi layout khác: nhãn cột "CHỈ TIÊU" (dòng "STT|CHỈ TIÊU|DỰ ÁN") và tên cost center
    ("Tổng Dự án|HO Dự án|<từng dự án>") nằm ở 2 DÒNG KHÁC NHAU (dòng sau) — `_kqkd_scan` giả
    định cùng 1 dòng nên không dùng chung được, phải dò riêng."""
    hdr_i = next((i for i, r in enumerate(rows[:10])
                  if any(_nd(c) == "tong du an" for c in r if c is not None)), None)
    ten_i = next((i for i, r in enumerate(rows[:10])
                  if any(_nd(c) == "chi tieu" for c in r if c is not None)), None)
    if hdr_i is None or ten_i is None:
        return []
    hdr = rows[hdr_i]
    ten_j = next(j for j, c in enumerate(rows[ten_i]) if _nd(c) == "chi tieu")
    # BỎ cột "Tổng Dự án"/"HO Dự án" — chỉ lấy 7 cột dự án (E:K, đúng theo Mapping) để Σ cost
    # center không đếm đôi với cột tổng (verify 01/08: Cao Bằng 392.421.525 + Phú Quốc
    # 239.351.852 = 631.773.377 = đúng "Tổng Dự án"; "HO Dự án" luôn 0 các ngày đã verify).
    cols = [(cc, j) for j, c in enumerate(hdr) if isinstance(c, str)
            for cc in [next((cc for kw, cc in _CC_DUAN if kw in _nd(c)), None)] if cc]
    if len(cols) < 3:
        return []

    anchored = {}
    for r in rows[ten_i + 1:]:
        if not r or ten_j >= len(r):
            continue
        n = _nd(r[ten_j])
        for key, (pref, exact) in _DUAN_ANCHOR.items():
            if key not in anchored and (n == pref if exact else n.startswith(pref)):
                anchored[key] = r
    if "dt_thuan" not in anchored:
        return []

    def val(key, j):
        r = anchored.get(key)
        return _num(r[j]) if r is not None and j < len(r) else None

    facts = []
    for cc, j in cols:
        dt = val("dt_thuan", j)
        if dt:
            facts.append((cc, RT_HQKD, MA_DT, MA_DT, dt))
            facts.append((cc, RT_PNLT, "Doanh thu HH, DV", "Doanh thu HH, DV", dt))
            facts.append((cc, RT_DTHU, "Doanh thu thuần", "Doanh thu thuần", dt))
        gv = val("gia_von", j)
        if gv:
            facts.append((cc, RT_PNLT, "Giá vốn hàng bán", "Giá vốn hàng bán", gv))
        ln_gop = val("ln_gop", j)
        if ln_gop:
            facts.append((cc, RT_PNLT, "Lợi nhuận gộp", "Lợi nhuận gộp", ln_gop))
        tong_cp = 0.0
        for key, nhom, ten in _DUAN_CP:
            v = val(key, j)
            if not v:
                continue
            tong_cp += v
            facts.append((cc, RT_CHIPHI, nhom, ten, v))
        if tong_cp:
            facts.append((cc, RT_HQKD, MA_CP, MA_CP, tong_cp))
        lntt = val("lntt", j)
        if lntt:
            facts.append((cc, RT_HQKD, MA_LNTT, MA_LNTT, lntt))
            facts.append((cc, RT_PNLT, "Lợi nhuận trước thuế", "Lợi nhuận trước thuế", lntt))
        lnst = val("lnst", j)
        if lnst:
            facts.append((cc, RT_PNLT, "Lợi nhuận sau thuế", "Lợi nhuận sau thuế", lnst))
    return facts


_FACTS_FN = {"srvf": _srvf_facts, "kqkd": _kqkd_facts, "antaxi": _antaxi_facts, "tcode": _tcode_facts,
             "duan": _duan_facts, "ht": _ht_facts, "xdv": _xdv_facts}


# ---------------------------------------------------------------------------------------------
# Layout "ho_kqkd" — HO (HO/baocaohqkdngay), sheet "D1".."D31" luỹ kế GIỐNG HỆT "tcode" (GA/
# TRAMSAC) nhưng KHÔNG dùng mã T-series: cột A ("Mã số") của 3 DÒNG TỔNG (Doanh thu/Chi phí/Lợi
# nhuận) TRỐNG, chỉ có mã ở các dòng CHI TIẾT (511_TS/5118/515.01... — mã tài khoản kế toán, không
# phải mã báo cáo) -> PHẢI neo theo NHÃN cột B ("Chỉ tiêu"), không theo mã như "tcode".
# Verify file thật (kiểm 2026-08-05, kỳ 2026-08 D1/D2): header dòng 5: Mã số | Chỉ tiêu | D01 |
# %DT | T07 | %DT | T08 | … | T12 | %DT — cột GIÁ TRỊ luôn ngay SAU "Chỉ tiêu" (khớp header dạng
# d\d{2} zero-pad, dò bằng CHÍNH cơ chế val_j của "tcode"). Cột T07-T12 là TEMPLATE RÁC (mọi kỳ,
# mọi sheet đều =0/#DIV0!/#REF!, không đổi theo ngày thật) -> CHỈ lấy cột D01 (ngay sau Chỉ tiêu).
# 3 dòng tổng (không mã): "Tổng Doanh thu" (dòng 6) / "Tổng chi phí" (dòng 14) / "Tổng lợi nhuận"
# (dòng 26) — user tự ghi rõ ô C6/C14/C26 trong Mapping, nhưng neo THEO NHÃN thay vì địa chỉ ô
# cứng (số dòng có thể lệch giữa các tháng khi kế toán chèn/xoá dòng chi tiết).
# LNTT/LNST: verify DB 2026-01..07 — PNLT của HO CHỈ có đúng 1 dim1 'Lợi nhuận sau thuế', giá trị
# BẰNG TUYỆT ĐỐI HQKD mã 1112 ở CẢ 7 kỳ (vd T07: -4,5950 = -4,5950) -> 'Tổng lợi nhuận' ĐÃ LÀ LNST
# (giống Trạm sạc, KHÁC GA) -> dùng profit_pnlt=('Lợi nhuận sau thuế',) như Trạm sạc.
# KHÔNG 'Giá vốn hàng bán'/'Lợi nhuận gộp': HO là khối hỗ trợ tập đoàn (chi phí quản lý/lãi vay),
# monthly KHÔNG có dim1 'Giá vốn hàng bán' cho HO (đơn vị không có COGS) — daily cũng không suy.
# Cơ cấu chi phí (CHIPHI_D): monthly HIỆN CHƯA CÓ breakdown nào cho HO (0 dòng CHIPHI) nhưng
# Mapping yêu cầu rõ (#10: "lấy từng loại chi phí trong file BC (C15-C25)") và file NGÀY CÓ sẵn 11
# dòng chi tiết (511_TS/5118/… ở phần DT không tính, CHỈ 11 dòng CHI PHÍ nằm giữa 'Tổng chi phí' và
# 'Tổng lợi nhuận') -> lấy NHÃN GỐC (cột B) làm dim1 THẲNG (không có nhóm chuẩn nào để map vào vì
# monthly chưa từng phân loại) — daily "đi trước" monthly ở khoản này, chấp nhận, không tự đặt tên
# nhóm mới có thể sai. Quét theo VÙNG (dòng SAU anchor 'Tổng chi phí', TRƯỚC anchor 'Tổng lợi
# nhuận'), KHÔNG hardcode "dòng 15-25" vì số dòng chi tiết có thể đổi giữa các tháng.
# ⚠ SỬA 2026-08-06 — "Doanh thu" KHÔNG lấy nguyên văn dòng "Tổng Doanh thu": dòng đó trong FILE
# THẬT cộng cả 515.xx (DT tài chính) + 7111 (thu nhập khác). Quy ước ĐÃ CHỐT với user 2026-07-30 cho
# bản THÁNG (agent_cli._derive_kqkd_ho::code511_sum, QA #10 T2/2026: dòng 6 = 32.833.798 ≠ Σ511 =
# 30.832.363 là số kế toán mong đợi) là **Doanh thu = Σ MỌI dòng mã 511*** (511_TS · 5118 · 5119 ·
# 511_LN). Bản NGÀY trước đây neo nhầm vào dòng tổng -> sẽ thổi phồng doanh thu HO đúng bằng
# 515+7111 ngay khi file có số. Anchor 'dt' giữ lại CHỈ để fallback khi sheet không có cột mã.
_HO_ANCHOR = {"dt": "tong doanh thu", "tong_cp": "tong chi phi", "lntt": "tong loi nhuan"}


def _ho_facts(rows):
    """rows -> [(None, report_type, dim1, dim3, value_VND)] cho 1 ngày. [] nếu sai layout."""
    hdr_i = next((i for i, r in enumerate(rows[:10]) if any(_nd(c) == "chi tieu" for c in r if c is not None)), None)
    if hdr_i is None:
        return []
    hdr = rows[hdr_i]
    ten_j = next(j for j, c in enumerate(hdr) if _nd(c) == "chi tieu")
    val_j = next((j for j, c in enumerate(hdr) if j > ten_j and re.fullmatch(r"d\d{2}", _nd(c))),
                 ten_j + 1)

    anchor_i = {}
    for i, r in enumerate(rows[hdr_i + 1:], hdr_i + 1):
        if not r or ten_j >= len(r):
            continue
        n = _nd(r[ten_j])
        for key, pref in _HO_ANCHOR.items():
            if key not in anchor_i and n == pref:
                anchor_i[key] = i
    if "dt" not in anchor_i or "tong_cp" not in anchor_i or "lntt" not in anchor_i:
        return []

    def val(key):
        i = anchor_i.get(key)
        r = rows[i] if i is not None else None
        return _num(r[val_j]) if r is not None and val_j < len(r) else None

    # Doanh thu = Σ dòng mã 511* (xem chú thích ở _HO_ANCHOR). Cột mã = header 'Mã số', fallback cột
    # ngay TRƯỚC 'Chỉ tiêu'. Excel lưu mã dạng số ('5118' -> 5118.0) nên phải bỏ đuôi '.0'.
    ma_j = next((j for j, c in enumerate(hdr) if _nd(c) == "ma so"), ten_j - 1)
    dt_511, co_511 = 0.0, False
    if ma_j >= 0:
        for r in rows[hdr_i + 1:]:
            if not r or ma_j >= len(r) or r[ma_j] in (None, ""):
                continue
            code = str(r[ma_j]).strip()
            if code.endswith(".0"):
                code = code[:-2]
            v = _num(r[val_j]) if val_j < len(r) else None
            if code.startswith("511") and v is not None:
                dt_511 += v
                co_511 = True

    facts = []
    # Không dòng 511* nào có số -> rơi về dòng "Tổng Doanh thu" (sheet đổi layout, mất cột mã).
    dt = dt_511 if co_511 else val("dt")
    if dt:
        facts.append((None, RT_HQKD, MA_DT, MA_DT, dt))
        facts.append((None, RT_DTHU, "Doanh thu thuần", "Doanh thu thuần", dt))
        facts.append((None, RT_PNLT, "Doanh thu HH, DV", "Doanh thu HH, DV", dt))
        # LN gộp = DT thuần (HO không có giá vốn) — Y HỆT bản THÁNG, chốt Mapping 2026-07-18.
        # Bản ngày trước đây THIẾU dòng này -> thẻ "Lợi nhuận gộp" của HO luôn trống ở chế độ Ngày.
        facts.append((None, RT_PNLT, "Lợi nhuận gộp", "Lợi nhuận gộp", dt))
    tong_cp = val("tong_cp")
    if tong_cp:
        facts.append((None, RT_HQKD, MA_CP, MA_CP, tong_cp))
    lntt = val("lntt")
    if lntt:
        facts.append((None, RT_HQKD, MA_LNTT, MA_LNTT, lntt))
        facts.append((None, RT_PNLT, "Lợi nhuận sau thuế", "Lợi nhuận sau thuế", lntt))
    # Cơ cấu chi phí: mọi dòng có nhãn NẰM GIỮA anchor 'tong_cp' và 'lntt' -> 1 dim1 riêng theo
    # nhãn gốc (xem docstring khối trên).
    for i in range(anchor_i["tong_cp"] + 1, anchor_i["lntt"]):
        r = rows[i]
        if not r or ten_j >= len(r):
            continue
        ten = str(r[ten_j]).strip() if r[ten_j] not in (None, "") else None
        v = _num(r[val_j]) if val_j < len(r) else None
        if ten and v:
            facts.append((None, RT_CHIPHI, ten, ten, v))
    return facts


_FACTS_FN["ho_kqkd"] = _ho_facts


# ---------------------------------------------------------------------------------------------
# Layout "anks" — An KS (ANKHACHSAN), KHÁC HẲN 3 layout trên: chỉ 1 SHEET DUY NHẤT, mỗi NGÀY là
# 1 CỘT (không phải mỗi ngày 1 sheet). Header dòng 3: TT | Nội dung chi phí | ĐVT | Tổng cộng |
# 1 | 2 | … | 31 (số ngày nằm NGAY Ở HEADER, không cần suy). Nhãn chỉ tiêu ở cột "Nội dung chi
# phí", neo theo TIỀN TỐ đã chuẩn hoá (mã La Mã I/II/III ở cột riêng bên trái, không dính nhãn).
# CỘT "Tổng cộng" = Σ các cột ngày trong tháng (verify ngày 1-4/08: 2.975.000+1.600.000+1.615.000
# +1.400.000 = 7.590.000 = đúng ô Tổng cộng dòng I) -> BỎ cột này, chỉ lấy cột NGÀY để không đếm
# đôi khi gộp nhiều ngày (giống lý do bỏ cột "Tổng cộng" ở layout kqkd).
# KHÔNG cost center: An KS chỉ 1 cơ sở (Garden Sơn Tây) — verify raw_rows bản THÁNG: cột
# cost_center của khối An KS luôn NULL ở mọi report_type/mọi kỳ.
# dim1 PNLT/CHIPHI lấy Y HỆT bản THÁNG của An KS (rà DB 2026-06):
#   CHIPHI: 'Giá vốn hàng bán' / 'Chi phí chung' / 'Chi phí lương + CP khác cho CNV' (3 nhóm,
#     KHÁC An Taxi/XVP) — đúng 3 dòng con II.1/II.2/II.3 của file.
#   PNLT: ghi CẢ HAI tên gross ('Doanh thu HH, DV' VÀ 'Doanh thu bán hàng và cung cấp dịch vụ',
#     cùng giá trị — như An Taxi, xem doanhthu-cautructructure-gross-giamtru) + 'Giá vốn hàng
#     bán' + CHỈ 'Lợi nhuận sau thuế' (An KS KHÔNG có dòng thuế TNDN -> LNTT=LNST, bản THÁNG
#     không ghi dim1 'Lợi nhuận trước thuế' riêng, xem An.xlsx spec — daily giữ đúng convention,
#     KHÔNG tự thêm dòng monthly không có). 'Lợi nhuận gộp' KHÔNG ghi tường minh — bản THÁNG lưu
#     nó nhưng do BACKEND tự tính (revenue.py::pnl.py fallback DT-giá vốn theo khối khi thiếu
#     dòng riêng), daily để trống cho backend tự suy, tránh trùng 2 nguồn.
_ANKS_ANCHOR = {
    "dt": "tong doanh thu",              # I
    "tong_cp": "tong chi phi",           # II (nhãn gốc có dấu ':' cuối -> startswith vẫn khớp)
    "gia_von": "chi phi gia von",        # II.1
    "chi_chung": "chi phi chung",        # II.2
    "luong": "chi phi luong",            # II.3 "CHI PHÍ LƯƠNG + CP KHÁC CHO CNV"
    "lntt": "loi nhuan (i-ii)",          # III — An KS: LNTT = LNST (không có dòng thuế TNDN)
}


def _anks_all_days(rows, period):
    """1 sheet, mỗi cột 1 ngày -> [(ngay, facts), ...] cho TRỌN THÁNG trong 1 lần gọi (khác hẳn
    3 layout trên vốn mỗi ngày gọi facts_fn riêng — xem docstring khối trên)."""
    hdr_i = next((i for i, r in enumerate(rows[:10])
                  if any(_nd(c) == "tong cong" for c in r if c is not None)), None)
    if hdr_i is None:
        return []
    hdr = rows[hdr_i]
    tong_j = next(j for j, c in enumerate(hdr) if _nd(c) == "tong cong")
    y, mm = int(period[:4]), int(period[5:7])
    day_cols = [(j, int(c)) for j, c in enumerate(hdr) if j > tong_j and isinstance(c, int)
                and not isinstance(c, bool) and 1 <= c <= 31]

    label_j = next((j for j, c in enumerate(hdr) if _nd(c).startswith("noi dung")), 2)
    anchored = {}
    for r in rows[hdr_i + 1:]:
        if not r or label_j >= len(r):
            continue
        n = _nd(r[label_j])
        for key, pref in _ANKS_ANCHOR.items():
            if key not in anchored and n.startswith(pref):
                anchored[key] = r

    def val(key, j):
        r = anchored.get(key)
        return _num(r[j]) if r is not None and j < len(r) else None

    per_day = []
    for j, d in day_cols:
        facts = []
        dt = val("dt", j)
        if dt:
            facts.append((None, RT_HQKD, MA_DT, MA_DT, dt))
            facts.append((None, RT_DTHU, "Doanh thu thuần", "Doanh thu thuần", dt))
            facts.append((None, RT_PNLT, "Doanh thu HH, DV", "Doanh thu HH, DV", dt))
            facts.append((None, RT_PNLT, "Doanh thu bán hàng và cung cấp dịch vụ",
                          "Doanh thu bán hàng và cung cấp dịch vụ", dt))
        tong_cp = val("tong_cp", j)
        if tong_cp:
            facts.append((None, RT_HQKD, MA_CP, MA_CP, tong_cp))
        for key, dim1 in (("gia_von", "Giá vốn hàng bán"), ("chi_chung", "Chi phí chung"),
                          ("luong", "Chi phí lương + CP khác cho CNV")):
            v = val(key, j)
            if v:
                facts.append((None, RT_CHIPHI, dim1, dim1, v))
        gia_von = val("gia_von", j)
        if gia_von:
            facts.append((None, RT_PNLT, "Giá vốn hàng bán", "Giá vốn hàng bán", gia_von))
        lntt = val("lntt", j)
        if lntt:
            facts.append((None, RT_HQKD, MA_LNTT, MA_LNTT, lntt))
            facts.append((None, RT_PNLT, "Lợi nhuận sau thuế", "Lợi nhuận sau thuế", lntt))
        if facts:
            per_day.append((f"{y:04d}-{mm:02d}-{d:02d}", facts))
    return per_day


# ---------------------------------------------------------------------------------------------
def derive(path, write=False):
    folder = _source_id(path).split("::", 1)[0]
    unit = _UNITS.get(folder)
    period = _period_of(os.path.basename(path), _in_day_dir(path))
    if not unit or not period:
        return {"ok": False, "skip": True}

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheets = None            # None = layout "anks" (1 sheet, không dò theo ngày)
    try:
        if unit["layout"] == "anks":
            rows = [list(r) for r in wb[wb.sheetnames[0]].iter_rows(values_only=True)]
            per_day = _anks_all_days(rows, period)
        else:
            layout = unit["layout"]
            facts_fn = _FACTS_FN[layout]
            if layout == "tcode":
                # 2 đơn vị chung layout "tcode" (GA/TRAMSAC) nhưng quy ước PNLT LNTT/LNST khác
                # nhau — xem docstring `_tcode_facts` param `profit_pnlt`.
                pp = unit.get("profit_pnlt", ("Lợi nhuận trước thuế",))
                ps = unit.get("pnlt_skip", ())
                facts_fn = lambda rows, _pp=pp, _ps=ps: _tcode_facts(rows, _pp, _ps)  # noqa: E731
            if layout == "srvf":
                sheets = _srvf_day_sheets(wb, period)
            elif layout in ("tcode", "ho_kqkd"):   # cùng kiểu sheet "D1".."D31" luỹ kế
                sheets = _tcode_day_sheets(wb, period)
            elif layout in ("ht", "xdv"):          # "{dd}.{mm}" hoặc "{dd}" trần
                sheets = _ht_day_sheets(wb, period)
            else:
                sheets = _kqkd_day_sheets(wb, period)
            per_day = []
            for sheet, ngay in sheets:
                rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
                facts = facts_fn(rows)
                if facts:
                    per_day.append((ngay, facts))
    finally:
        wb.close()

    if not per_day:
        # PHÂN BIỆT 2 nguyên nhân — bản cũ gộp chung 1 câu "không đọc được sheet ngày nào" khiến
        # chẩn đoán đi nhầm hướng (2026-08-06, Trạm sạc/GA/HO kỳ 08: tưởng hỏng dò sheet, hoá ra
        # sheet đọc tốt nhưng file nguồn ghi 0 CỨNG ở mọi mã tổng — xem `sheets_ngay` trả kèm).
        if sheets is not None and sheets:
            names = ", ".join(s for s, _ in sheets)
            return {"ok": False, "period": period, "layout": unit["layout"],
                    "sheets_ngay": [s for s, _ in sheets],
                    "error": f"đọc được {len(sheets)} sheet ngày ({names}) nhưng không sheet nào ra "
                             f"chỉ tiêu có số — kiểm tra file nguồn còn để trống/ghi 0 ở các mã tổng "
                             f"(layout {unit['layout']}, kỳ {period})"}
        return {"ok": False, "error": f"không thấy sheet ngày nào khớp kỳ {period} "
                                      f"(layout {unit['layout']})"}

    out = {"ok": True, "file": os.path.basename(path), "period": period, "cong_ty": unit["cong_ty"],
           "layout": unit["layout"], "days": len(per_day),
           "tong_theo_ngay": {
               # `x[:5]` chứ không giải nén cứng 5 phần tử: fact của layout "srvf" có thêm
               # dim2 (kênh bán) ở vị trí thứ 6.
               ngay: {"doanh_thu_ty": round(sum(x[4] for x in f
                                                if x[1] == RT_HQKD and x[2] == MA_DT) * 1e-9, 9),
                      "chi_phi_ty": round(sum(x[4] for x in f
                                              if x[1] == RT_HQKD and x[2] == MA_CP) * 1e-9, 9),
                      "lntt_ty": round(sum(x[4] for x in f
                                           if x[1] == RT_HQKD and x[2] == MA_LNTT) * 1e-9, 9),
                      "ban_xe_ty": round(sum(x[4] for x in f if x[1] == RT_KDVH) * 1e-9, 9)}
               for ngay, f in per_day}}
    if not write:
        return out

    source_file = _source_id(path)
    conn = psycopg.connect(DB_URL)
    try:
        cur = conn.cursor()
        cur.execute("SELECT id FROM datasets WHERE kind='month' AND period=%s "
                    "ORDER BY created_at DESC LIMIT 1", (period,))
        row = cur.fetchone()
        if not row:
            out["ok"] = False
            out["error"] = f"chưa có dataset tháng {period} — nạp báo cáo THÁNG trước rồi chạy lại"
            return out
        dataset_id = row[0]
        # idempotent: xoá MỌI dòng cũ CÙNG source_file (mỗi file phủ trọn 1 tháng ngày).
        # KHÔNG giới hạn ở REPORT_TYPES nữa (sửa 2026-08-06): file báo cáo ngày của đơn vị CHƯA
        # khai trong `_UNITS` sẽ rơi xuống pipeline THÁNG (gate `is_daily_report` trả False) và để
        # lại dòng report_type THÁNG mang đúng source_file này — vd HT kỳ 2026-08 có 40 dòng
        # HQKD/PNLT/CHIPHI/DTHU sinh từ chính file ngày, hiện lên dashboard như số CẢ THÁNG trong
        # khi thực chất là vài ngày cộng lại. Khi đơn vị được thêm vào `_UNITS`, deriver này là chủ
        # DUY NHẤT của file (agent_cli return sớm ở gate) -> mọi dòng khác cùng source_file đều là
        # rác của lần phân loại nhầm trước đó, phải dọn cùng lượt nạp.
        cur.execute("DELETE FROM raw_rows WHERE source_file=%s", (source_file,))
        payload = json.dumps({"unit": "ty", "grain": "day"}, ensure_ascii=False)
        recs, i = [], 0
        for ngay, facts in per_day:
            for f in facts:
                # Phần tử thứ 6 (dim2) là TUỲ CHỌN — chỉ layout "srvf" dùng, để gắn kênh bán cho
                # cụm bán xe. Giải nén theo lát cắt thay vì đổi mọi layout sang tuple 6 phần tử.
                cc, rt, dim1, dim3, v = f[:5]
                dim2 = f[5] if len(f) > 5 else None
                i += 1
                recs.append((dataset_id, rt, 6200000 + i, ngay,
                             _CC_CONGTY.get(cc) or unit["cong_ty"], unit["khoi"], cc, period,
                             round(v * 1e-9, 9), None, dim1, dim2, dim3, payload, source_file))
        cur.executemany(
            "INSERT INTO raw_rows (dataset_id, report_type, row_index, ngay, cong_ty, khoi, "
            "cost_center, period_month, amount, amount2, dim1, dim2, dim3, payload, source_file) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", recs)
        conn.commit()
        out["written"] = len(recs)
    finally:
        conn.close()
    return out


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--write", action="store_true", help="ghi DB (mặc định dry-run)")
    a = ap.parse_args()
    print(json.dumps(derive(a.file, a.write), ensure_ascii=False, indent=2))
