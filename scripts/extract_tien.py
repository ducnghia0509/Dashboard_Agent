# -*- coding: utf-8 -*-
"""DẪN XUẤT KHỐI TIỀN (GỘP 1 MỐI) từ Báo cáo tiền tập đoàn — 1 lần đọc, neo theo MÃ:
  · SDT  (03B_SODU_TIEN) — số dư Tiền mặt/Tiền gửi/Tiền vay theo CÔNG TY (sheet 'TC01_SD TIỀN').
  · VAY  (04_VAY)        — dư nợ đầu/cuối + vay thêm/trả nợ theo CÔNG TY × NGÂN HÀNG (mục TIỀN VAY).
  · THUCHI (03_DONGTIEN) — thu/chi theo KHOẢN MỤC (mã 1-9) từng pháp nhân (sheet 'BC THU CHI_T*_<CTY>').

Thay 3 extractor rời (extract_sodu_tien / extract_vay / extract_thuchi) — cùng nguồn, gom để nhất
quán + hết nhân đôi. Neo theo MÃ/section (không cứng vị trí). Import idempotent theo (source_file,
report_type) qua template_filler.import_filled.

Chạy: .venv/bin/python scripts/extract_tien.py <file.xlsx> --period 2026-06
"""
import argparse
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from servers import template_filler as tf  # noqa: E402
from servers.common import source_catalog as SC  # noqa: E402
from servers.common import be_bridge as bb  # noqa: E402

SD_SHEET = "TC01_SD TIỀN"
# Hậu tố sheet 'BC THU CHI_T<m>_<SUFFIX>' -> mã pháp nhân (T6+ tách AN thành ANTAXI/ANKS, cùng AAG).
_SUFFIX_CO = {"TC": "TC", "VFQN": "VFQN", "XANH": "XVP", "HUNGTHINH": "HT", "AN": "AAG",
              "ANTAXI": "AAG", "ANKS": "AAG"}

# ─────────── CỜ NẠP THEO NGÀY (THUCHI_DAILY) — MẶC ĐỊNH TẮT ───────────
# Nguồn có sẵn dữ liệu TỪNG NGÀY: 'BC THU CHI' mỗi ngày 1 khối cột [dd, TM, NH, TVAY];
# 'TC01_SD TIỀN' mỗi ngày 1 cột số dư. Bật cờ -> bóc theo ngày (Kỳ = yyyy-mm-dd) và nạp vào
# ĐÚNG DATASET THÁNG (force_grain='month'), vì Σ các ngày == số cả kỳ (đã đối chiếu 493/493
# khoá (công ty × loại × khoản mục × hình thức) trên 6 kỳ, lệch 0) nên KPI tháng không đổi.
#
# Cờ đặt trong .env của API GỌI extractor (subprocess kế thừa env), nên TEST bật được mà PROD
# không đổi — TEST và PROD dùng CHUNG thư mục Dashboard_Agent này.
def _daily_on() -> bool:
    return os.environ.get("THUCHI_DAILY", "").strip().lower() in ("1", "true", "yes", "on")


def _days_in(period: str) -> int:
    """Số ngày thật của kỳ 'yyyy-mm'. Mọi sheet đều có ĐỦ 31 cột ngày kể cả tháng 28/30 ngày —
    cột dư là RỖNG; không chặn thì số dư ngày cuối đọc ra 0 và làm sai cả chuỗi."""
    import calendar
    y, m = int(period[:4]), int(period[5:7])
    return calendar.monthrange(y, m)[1]


def _day_cols(hdr_row, ndays: int, subs=()) -> dict:
    """{ngày: {sub_label: col}} dò theo TEXT header (KHÔNG dùng offset cố định).

    Bắt buộc dò theo text: sheet 'ANKS' (T06) dùng khối 6 cột/ngày thay vì 4, và 'TC01_SD TIỀN'
    T01 thiếu cột '+/- NGÀY 27'/'+/- NGÀY 30' -> đếm bước cố định sẽ lệch cột.
    subs=() -> mỗi ngày chỉ 1 cột (số dư): trả {ngày: {'': col}}.
    """
    hdr = {i: str(x).strip() for i, x in enumerate(hdr_row) if x is not None and str(x).strip() != ""}
    idx = sorted(hdr)
    out = {}
    for k, i in enumerate(idx):
        if not re.fullmatch(r"\d{1,2}", hdr[i]):
            continue
        d = int(hdr[i])
        if not 1 <= d <= ndays or d in out:
            continue
        if not subs:
            out[d] = {"": i}
            continue
        cols = {}
        for j in idx[k + 1:]:
            if re.fullmatch(r"\d{1,2}", hdr[j]):      # sang ngày kế -> dừng
                break
            for lbl, name in subs:
                if hdr[j] == lbl and name not in cols:
                    cols[name] = j
        if cols:
            out[d] = cols
    return out
# Cột số dư theo LOẠI TIỀN (SD TIỀN) -> cột template 03B.
_SEC_COL = {"tien mat": "Tiền mặt (tỷ)", "tien gui": "Tiền gửi NH (tỷ)",
            "tien vay": "Số dư tiền vay (tỷ) — đối chiếu 04_VAY"}
# NGOẠI BẢNG: mục IV BẢO LÃNH / V LC trong TC01_SD TIỀN -> cột tương ứng của template 03B.
# (Nguồn KHÔNG có mục TSĐB nào -> cột 'Tài sản đảm bảo (tỷ)' để trống, không bịa số.)
_OFF_SEC = {"bao lanh": "bao_lanh", "lc": "lc"}
_OFF_COL = {"lc": "Ngoại bảng: LC (tỷ)", "bao_lanh": "Bảo lãnh thanh toán (tỷ)"}
_NAME_ALIAS = {"an taxi": "AAG", "an ks": "AAG", "an khach san": "AAG", "global ai": "GA",
               "htx xanh tuyen quang": "HTX_XTQ", "htx xanh vinh phuc": "HTX_XVP",
               "xanh tuyen quang": "HTX_XTQ"}
_KYHAN = re.compile(r"(ngan han|trung han|dai han)")   # dòng con kỳ hạn dưới ngân hàng -> bỏ (cộng trùng)


def _norm(s):
    return bb.remove_diacritics("" if s is None else str(s)).strip().lower()


def _resolve_co(name):
    return bb.master.resolve_company_code(name) or _NAME_ALIAS.get(_norm(name))


def _num(v):
    return float(v) if isinstance(v, (int, float)) else None


# ───────────────────────── SD TIỀN -> SDT + VAY ─────────────────────────
def _sd_extract(rows, period):
    """Neo theo MÃ/section: dòng SECTION (cột0+cột1=LOẠI TIỀN), CÔNG TY (cột2), NGÂN HÀNG (cột3).
    SDT = số dư cuối theo công ty × loại tiền. VAY = đầu/cuối + vay thêm/trả theo công ty × bank."""
    hdr_i = next((i for i, r in enumerate(rows[:15])
                  if any("dau ky" in _norm(c) for c in r) and any("den ngay" in _norm(c) for c in r)), None)
    if hdr_i is None:
        return [], [], set(), None, "Không thấy header (ĐẦU KỲ / ĐẾN NGÀY HIỆN TẠI)"
    hdr = rows[hdr_i]
    c_dau = next(j for j, c in enumerate(hdr) if "dau ky" in _norm(c))
    c_cuoi = next(j for j, c in enumerate(hdr) if "den ngay" in _norm(c))
    deltas = [j for j, c in enumerate(hdr) if _norm(c).startswith("+/-")]
    # THEO NGÀY: mỗi ngày 1 cột SỐ DƯ (mục ở cột B đã là loại tiền nên không có cột con).
    # Chặn theo số ngày thật của kỳ; dò theo TEXT vì T01 thiếu vài cột '+/- NGÀY'.
    sd_days = _day_cols(hdr, _days_in(period)) if _daily_on() else {}

    def _last_cell(r):
        """Cột chứa SỐ DƯ CHỐT KỲ của dòng r: ƯU TIÊN cột NGÀY CUỐI KỲ, chỉ rơi về cột F
        ('ĐẾN NGÀY HIỆN TẠI') khi ô ngày cuối trống/0.

        Cột ngày neo vào MỘT ngày xác định; F chỉ là "đến lúc kế toán sửa lần cuối" nên có thể
        chưa chốt tới ngày cuối kỳ:
         · T01–T06 (đã chốt): hai nguồn bằng nhau -> không đổi số nào.
         · T07: cột ngày 31 cho dư nợ vay 1524.058 tỷ, khớp CẢ (a) ngày 30 + dòng vay 31/07 của
           từng công ty (5/5) và (b) số mở đầu 01/08 của file T08; còn F = 1499.5 (HT để trống)
           -> tin F thì 31/07 sai và số dư nhảy vô cớ 12.77 tỷ sang 01/08.
         · T08 (đang dở): F trống, chỉ cột ngày có số -> vẫn đọc được.
        Ô ngày cuối = 0 coi như CHƯA CHỐT (kỳ đang dở điền 0 sẵn) -> để F đảm nhiệm, đúng quy ước
        `if cuoi:` / `if _v:` dùng sẵn trong file này."""
        if sd_days:
            _j = (sd_days.get(_days_in(period)) or {}).get("")
            if _j is not None and _j < len(r) and _num(r[_j]):
                return _j
        return c_cuoi

    def _cells(r):
        """[(Kỳ, cột, ghi_0)] cần đọc cho dòng r. Dùng CHUNG cho số dư loại tiền, ngoại bảng (IV/V)
        và dư nợ cuối kỳ của report VAY -> ba chỗ không thể lệch quy tắc nhau.

        ghi_0=False cho Ô NGÀY CUỐI KỲ: 0 ở đó nghĩa CHƯA CHỐT (kỳ đang dở điền 0 sẵn) nên phải BỎ,
        không thì snapshot_sum lấy nó làm "số dư mới nhất" = 0 và KPI của kỳ đang dở về 0 (T08).
        Nhánh không-daily giữ đúng hành vi cũ (`cuoi is not None` -> 0 vẫn ghi) vì PROD chạy nhánh này."""
        if not sd_days:
            return [(period, _last_cell(r), True)]
        _last = _days_in(period)
        out = [(f"{period}-{_d:02d}", _c[""], True) for _d, _c in sorted(sd_days.items()) if _d != _last]
        out.append((f"{period}-{_last:02d}", _last_cell(r), False))
        return out

    def _bank_vals(r):
        dau = _num(r[c_dau]) if c_dau < len(r) else None
        _jc = _last_cell(r)
        cuoi = _num(r[_jc]) if _jc < len(r) else None
        vt = sum(v for j in deltas if j < len(r) and (v := _num(r[j])) and v > 0)
        tn = -sum(v for j in deltas if j < len(r) and (v := _num(r[j])) and v < 0)
        return dau, cuoi, vt, tn

    def _vay_rec(code, bank, dau, cuoi, vt, tn):
        # DƯ NỢ đầu/cuối theo BANK (giữ, nguồn SD TIỀN). Vay thêm/trả=0 ở đây — nay lấy PER-BANK từ
        # báo cáo ngân hàng BCTH 2 (cột Tổng vay/Đã thanh toán) trong extract_vay_kyhan (chạy sau).
        # KHÔNG dùng delta SD TIỀN (sai) cũng KHÔNG dùng BC THU CHI mức công ty (làm bẩn chart theo NH).
        return {"Kỳ": period, "Đơn vị": code, "Ngân hàng": bank,
                "Dư nợ đầu kỳ (tỷ)": round((dau or 0) / 1e9, 9),
                "Vay thêm trong kỳ (tỷ)": 0.0,
                "Trả nợ trong kỳ (tỷ)": 0.0,
                "Dư nợ cuối kỳ (tỷ)": round((cuoi or 0) / 1e9, 9)}

    sdt, vay, unresolved, dachi = [], [], set(), None
    off_by_ky = {}          # {Kỳ: {mã công ty: {"lc": VND, "bao_lanh": VND}}} — mục IV/V THEO NGÀY
    off_sec = None           # đang ở trong khối ngoại bảng nào (None = không)
    section = company = None
    vay_seen = False        # 'TIỀN VAY' liệt kê 2 lần cùng số -> chỉ đọc khối ĐẦU (tránh nhân đôi)
    in_vay = False
    orphans = []            # bank ĐỨNG TRƯỚC dòng công ty (layout TC) -> gán khi subtotal cuối khớp

    def _flush(code, subtotal_cuoi):
        nonlocal orphans
        if orphans and subtotal_cuoi is not None:
            tong = sum(b[2] or 0.0 for b in orphans)   # b=(bank,dau,cuoi,vt,tn)
            if abs(tong - subtotal_cuoi) <= max(abs(subtotal_cuoi) * 0.005, 1e4):
                for b in orphans:
                    vay.append(_vay_rec(code, *b))
        orphans = []

    for r in rows[hdr_i + 1:]:
        c0 = str(r[0]).strip() if r[0] not in (None, "") else ""
        c1 = str(r[1]).strip() if len(r) > 1 and r[1] not in (None, "") else ""
        c2 = str(r[2]).strip() if len(r) > 2 and r[2] not in (None, "") else ""
        c3 = str(r[3]).strip() if len(r) > 3 and r[3] not in (None, "") else ""
        if c0 and c1:                                   # dòng SECTION
            _n1 = _norm(c1)
            section = next((k for k in _SEC_COL if k in _n1), None)
            # Mục IV BẢO LÃNH / V LC: KHÔNG phải loại tiền -> section=None, đọc riêng vào off_by_co.
            off_sec = None if section else next((v for k, v in _OFF_SEC.items() if k in _n1), None)
            in_vay = ("tien vay" == section) and not vay_seen
            if section == "tien vay":
                vay_seen = True
            company, orphans = None, []
            continue
        # Dòng công ty trong khối NGOẠI BẢNG. Chỉ nhận dòng resolve được ra MÃ công ty — các dòng
        # con như 'Công trường' / 'Xe Vinfast' (chi tiết bên trong Thịnh Cường) không resolve được
        # nên tự bị bỏ, tránh cộng trùng với dòng tổng của công ty.
        if off_sec and c2:
            _code = _resolve_co(c2)
            if _code:
                for _ky, _j, _ in _cells(r):       # ngoại bảng cũng có cột theo NGÀY, không chỉ cột F
                    _v = _num(r[_j]) if _j < len(r) else None
                    if _v:
                        _d = off_by_ky.setdefault(_ky, {}).setdefault(_code, {})
                        _d[off_sec] = _d.get(off_sec, 0.0) + _v
            continue
        if not section:
            continue
        # 'Đã chi nhưng chưa có chứng từ' (memo dưới TIỀN MẶT) -> cảnh báo TC (guide 21/7). Bắt theo TÊN
        # dòng (c2), chặn TRƯỚC nhánh công ty (nếu không sẽ rơi vào unresolved). Lấy cả đầu & cuối kỳ.
        if section == "tien mat" and c2 and "da chi" in _norm(c2) and "chung tu" in _norm(c2):
            _jc = _last_cell(r)
            dachi = {"cuoi": _num(r[_jc]) if _jc < len(r) else None,
                     "dau": _num(r[c_dau]) if c_dau < len(r) else None}
            continue
        if c2:                                          # dòng CÔNG TY (subtotal)
            code = _resolve_co(c2)
            _jc = _last_cell(r)                          # cùng quy tắc với _bank_vals -> _flush khớp
            cuoi = _num(r[_jc]) if _jc < len(r) else None
            if code is None:
                unresolved.add(c2)
                company, orphans = None, []
                continue
            # KHỐI 'TIỀN VAY' LẶP LẠI: sheet có BẢNG THỨ HAI (dòng ~165) với 'STT/LOẠI TIỀN' +
            # 'I TIỀN VAY' nữa. Cờ vay_seen trước đây chỉ chặn nhân đôi cho bản ghi VAY, KHÔNG chặn
            # SDT -> mỗi công ty ra 2 dòng '6. TIỀN VAY' và snapshot_sum CỘNG cả hai, làm KPI
            # "Số dư tiền vay" phồng lên (T05: 3506.4 thay vì 1753.2 = gấp đôi; T04: +38.6).
            # Bug có TỪ TRƯỚC bản nạp-theo-ngày; nay chặn cả SDT theo đúng ý cờ vay_seen.
            if section == "tien vay" and not in_vay:
                company, orphans = None, []
                continue
            if in_vay:
                _flush(code, cuoi)                        # gán các bank orphan (đứng trước) cho cty này
            company = code
            # Số dư TỪNG NGÀY: 1 dòng/ngày (ngày cuối kỳ theo quy tắc _last_cell). Ô rỗng -> BỎ
            # (không ghi 0) vì SDT là báo cáo SỐ DƯ: ghi 0 cho ngày chưa chốt sẽ bị snapshot_sum
            # lấy làm "số dư mới nhất" = 0. Không bật daily -> _cells trả 1 ô chốt kỳ.
            # Chỉ 1 dòng/ngày (không 2) vì snapshot_sum CỘNG các dòng cùng ngày.
            for _ky, _j, _z0 in _cells(r):
                _v = _num(r[_j]) if _j < len(r) else None
                if _v is None or (not _z0 and not _v):
                    continue
                sdt.append({"Kỳ": _ky, "Đơn vị": code, _SEC_COL[section]: round(_v / 1e9, 9)})
            continue
        if in_vay and c3 and not _KYHAN.search(_norm(c3)):   # dòng NGÂN HÀNG
            dau, cuoi, vt, tn = _bank_vals(r)
            if not any((dau, cuoi, vt, tn)):
                continue
            if company:                                  # bank đứng SAU công ty của nó (layout thường)
                vay.append(_vay_rec(company, c3, dau, cuoi, vt, tn))
            else:                                        # bank đứng TRƯỚC công ty -> chờ subtotal
                orphans.append((c3, dau, cuoi, vt, tn))
    # Gắn ngoại bảng vào dòng SDT của ĐÚNG công ty đó. cashflow_extras dedup theo cong_ty rồi cộng
    # 1 dòng/công ty, nên mọi dòng của công ty mang cùng giá trị là đúng (không cộng trùng).
    for rec in sdt:
        _o = (off_by_ky.get(rec.get("Kỳ")) or {}).get(rec.get("Đơn vị")) or {}
        for _k, _col_name in _OFF_COL.items():
            if _o.get(_k):
                rec[_col_name] = round(_o[_k] / 1e9, 9)
    return sdt, vay, unresolved, dachi, None


# ───────────────────────── BC THU CHI -> THUCHI ─────────────────────────
_COL_KY = "Kỳ / Ngày"
_COL_CTY = "Mã Công ty (auto từ CC)"
_COL_LOAI = "Loại (Thu/Chi)"
_COL_KM = "Khoản mục (Thu bán hàng, Thu đầu tư, Chi NCC, Chi tài chính, Chi đầu tư TS…)"
_COL_HT = "Hình thức (TM/TG/Đối trừ CN/Vay)"   # -> dim2 (import map sẵn); tách TM/gửi/vay
_COL_TH = "Thực hiện (tỷ)"


def _co_of_sheet(sheet):
    up = sheet.upper()
    if "TỔNG" in up or "TONG" in up:
        return None
    return _SUFFIX_CO.get(up.rsplit("_", 1)[-1].strip())


def _val_col(rows):
    """Cột giá trị tổng = 'TM + TG + T.VAY' (gồm vay, theo yêu cầu). Mặc định 3 (cột D)."""
    for r in rows[:12]:
        for j, c in enumerate(r):
            s = str(c).strip().upper() if c is not None else ""
            if "TM" in s and "VAY" in s:
                return j
    return 3


_HT_SUBS = (("TM", "Tiền mặt"), ("NH", "Tiền gửi"), ("TVAY", "Tiền vay"))
# Offset cột CẢ KỲ so với cột tổng D (vc): E=TM, F=NH, G=TVAY — dùng để đối chiếu bản ngày.
_HT_PERIOD_OFF = ((1, "Tiền mặt"), (2, "Tiền gửi"), (3, "Tiền vay"))


def _thuchi_extract(wb, period):
    """-> (records, sheets, recon). recon = các khoản mục mà Σ CÁC NGÀY != số CẢ KỲ của nguồn.

    Chỉ có ý nghĩa khi bật cờ ngày. Nguồn KHÔNG bảo đảm 2 con số này luôn bằng nhau: cột cả kỳ
    là công thức cộng DỌC theo dòng con (vd ANKS T06: E9 = SUM(E10:E11)) còn các cột ngày nhập
    tay -> có sheet lệch thật (ANKS T06 lệch 284.000đ ở 'Thu từ hoạt động kinh doanh'). PHẢI báo
    ra chứ không nuốt, vì nạp theo ngày sẽ làm KPI tháng lệch đúng phần đó so với báo cáo giấy.
    """
    daily = _daily_on()
    nd = _days_in(period)
    recs = []
    recon = []
    sheets = [s for s in wb.sheetnames if re.search(r"THU CHI_T\d", s.upper()) and _co_of_sheet(s)]
    for sh in sheets:
        co = _co_of_sheet(sh)
        # HT (Xe tải Hưng Thịnh): DÒNG TIỀN nay lấy từ 'BC THU CHI_T*_HUNGTHINH' NHƯ MỌI PHÁP NHÂN
        # (ĐẢO 2026-07-21 — xem memory ht-dongtien-from-lctt: bỏ GỌI _derive_lctt_ht, lấy chung nguồn
        # để khớp báo cáo tập đoàn + hết nạp đôi). TRƯỚC ĐÂY skip HT ở đây (lấy từ LCTT); nay KHÔNG skip
        # nữa — _derive_lctt_ht đã không được gọi nên KHÔNG đếm đôi. (SDT/VAY HT vẫn ở _sd_extract.)
        rows = [list(r) for r in wb[sh].iter_rows(values_only=True)]
        vc = _val_col(rows)
        # Header thật = dòng có cột 'TM + TG + T.VAY' (thường dòng 7) — cũng là dòng mang mã ngày.
        hdr_row = next((r for r in rows[:12]
                        if any(isinstance(c, str) and "TM" in c.upper() and "VAY" in c.upper() for c in r)), None)
        dcols = _day_cols(hdr_row, nd, _HT_SUBS) if (daily and hdr_row is not None) else {}
        sec = None                                       # 'A' (thu, mục I) / 'B' (chi, mục II)
        for r in rows:
            c0 = str(r[0]).strip() if r[0] not in (None, "") else ""
            c1 = str(r[1]).strip() if len(r) > 1 and r[1] not in (None, "") else ""
            if c0 == "I":
                sec = "A"; continue
            if c0 == "II":
                sec = "B"; continue
            if sec and re.fullmatch(r"\d+", c0) and c1:  # khoản mục CẤP 1 (1,2,3…) — bỏ 1.1 / '+…'
                # TÁCH THEO HÌNH THỨC (spec 2026-07-23): cột D (vc) = E(TM) + F(NH/gửi) + G(TVAY) — verify
                # khớp per khoản mục. Emit 1 dòng/hình thức (Hình thức -> dim2) THAY dòng tổng D: Σ = D nên
                # inflow/outflow (Σ amount) KHÔNG đổi; có thêm 'tiền mặt/gửi thu-chi' lọc theo dim2. Bỏ dòng =0.
                loai = "Thu" if sec == "A" else "Chi"
                if dcols:
                    # THEO NGÀY: cùng khoản mục × hình thức, 1 dòng/ngày. KHÔNG dùng cột tổng ngày
                    # (nguồn tự sai: T01 sheet TC dòng TỔNG CHI ngày 13 lệch 27.5 tỷ so với TM+NH+TVAY)
                    # -> chỉ cộng các cột thành phần, vốn khớp đúng số cả kỳ.
                    _sum = {}
                    for _d, _cols in dcols.items():
                        for _ht, _j in _cols.items():
                            _v = r[_j] if _j < len(r) and isinstance(r[_j], (int, float)) else None
                            if not _v:
                                continue
                            _sum[_ht] = _sum.get(_ht, 0.0) + _v
                            recs.append({_COL_KY: f"{period}-{_d:02d}", _COL_CTY: co, _COL_LOAI: loai,
                                         _COL_KM: c1, _COL_HT: _ht, _COL_TH: round(_v / 1e9, 9)})
                    # đối chiếu với cột CẢ KỲ (vc+1/2/3) — ngưỡng 1.000đ để bỏ nhiễu làm tròn
                    for _off, _ht in _HT_PERIOD_OFF:
                        _j = vc + _off
                        _p = r[_j] if _j < len(r) and isinstance(r[_j], (int, float)) else 0.0
                        _delta = _sum.get(_ht, 0.0) - (_p or 0.0)
                        if abs(_delta) > 1e3:
                            recon.append({"sheet": sh, "cong_ty": co, "loai": loai, "khoan_muc": c1,
                                          "hinh_thuc": _ht, "sum_ngay": round(_sum.get(_ht, 0.0) / 1e9, 9),
                                          "ca_ky": round((_p or 0.0) / 1e9, 9),
                                          "lech": round(_delta / 1e9, 9)})
                    continue
                for _off, _ht in ((1, "Tiền mặt"), (2, "Tiền gửi"), (3, "Tiền vay")):
                    _j = vc + _off
                    _v = r[_j] if _j < len(r) and isinstance(r[_j], (int, float)) else None
                    if not _v:
                        continue
                    recs.append({_COL_KY: period, _COL_CTY: co, _COL_LOAI: loai,
                                 _COL_KM: c1, _COL_HT: _ht, _COL_TH: round(_v / 1e9, 9)})
    return recs, sheets, recon


def extract(path: str, period: str, cong_ty: str = None) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    src = SC.source_id_from_path(path)
    out = {"ok": True}
    # Mốc "dòng có TRƯỚC lượt nạp này" + tập report_type lượt này SINH RA. Cuối hàm dùng để dọn
    # loại CŨ không còn sinh ra nữa (import_filled chỉ xoá theo loại của TỪNG lần gọi — xem
    # template_filler.prune_stale_types).
    _before_id = tf.next_row_id()
    _made, _ds = set(), None

    # 1) SDT + VAY
    if SD_SHEET in wb.sheetnames:
        rows = [list(r) for r in wb[SD_SHEET].iter_rows(values_only=True)]
        sdt, vay, unresolved, dachi, err = _sd_extract(rows, period)
        # VAY: dòng per-BANK giữ dư nợ đầu/cuối (SD TIỀN); vay thêm/trả=0 ở đây — nay lấy PER-BANK từ
        # báo cáo ngân hàng (BCTH 2) trong extract_vay_kyhan (chạy sau, wired autofill). KHÔNG còn emit
        # dòng '(Cả <cty>)' mức công ty (làm bẩn chart theo ngân hàng + chart đi vay/trả trống).
        if err:
            out["sd_error"] = err
        if unresolved:
            out["unresolved_companies"] = sorted(unresolved)
        if sdt:
            p = os.path.join(tf.FILLED_DIR, f"TIEN_{period}_03B_SODU_TIEN.xlsx")
            tf.fill("03B_SODU_TIEN", sdt, p)
            # force_grain=month khi bóc theo ngày: giữ 1 DATASET/KỲ chứa dòng theo ngày
            # (hướng A) thay vì để importer suy grain=day rồi tạo 1 dataset cho MỖI NGÀY.
            _r = tf.import_filled(p, cong_ty=None, source_file=src,
                                  force_grain="month" if _daily_on() else None)
            out["sdt"] = _r.get("rows_imported")
            _ds = _r.get("dataset_id") or _ds
            _made |= set((_r.get("by_type") or {}))
        if vay:
            p = os.path.join(tf.FILLED_DIR, f"TIEN_{period}_04_VAY.xlsx")
            tf.fill("04_VAY", vay, p)
            _r = tf.import_filled(p, cong_ty=None, source_file=src)
            out["vay"] = _r.get("rows_imported")
            _ds = _r.get("dataset_id") or _ds
            _made |= set((_r.get("by_type") or {}))
        # 'Đã chi nhưng chưa có chứng từ' (cảnh báo TC) -> DACHI_CCT direct-insert (KHÔNG qua template).
        # amount=cuối, amount2=đầu. Idempotent (DELETE trước). Twin lấy dataset_id/ngay/khoi từ dòng SDT
        # TC vừa nạp ở trên. Port từ extract_sodu_tien (legacy) -> nay tất định trong extract_tien (wired).
        if dachi is not None:
            from servers.common import be_bridge as bb
            import json as _json
            _db = bb.db.get_db()
            _db.execute("DELETE FROM raw_rows WHERE source_file=? AND period_month=? AND report_type=?",
                        (src, period, "DACHI_CCT"))
            # ORDER BY ngay DESC: bản nạp THEO NGÀY có ~30 dòng SDT/công ty -> phải neo vào NGÀY
            # CUỐI KỲ (số dư chốt), nếu LIMIT 1 tuỳ ý thì cảnh báo rơi vào một ngày giữa tháng.
            _tw = _db.execute("SELECT dataset_id, ngay, khoi FROM raw_rows WHERE source_file=? AND "
                              "period_month=? AND report_type='SDT' AND cong_ty='TC' "
                              "ORDER BY ngay DESC LIMIT 1",
                              (src, period)).fetchone()
            if _tw:
                _db.execute(
                    "INSERT INTO raw_rows(dataset_id,report_type,row_index,ngay,cong_ty,khoi,cost_center,"
                    "period_month,amount,amount2,dim1,dim2,dim3,payload,source_file) "
                    "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (_tw["dataset_id"], "DACHI_CCT", 6000000, _tw["ngay"], "TC", _tw["khoi"], None, period,
                     round((dachi["cuoi"] or 0) * 1e-9, 9), round((dachi["dau"] or 0) * 1e-9, 9),
                     "Đã chi nhưng chưa có chứng từ", None, None,
                     _json.dumps({"unit": "ty", "nguon": "TC01_SD TIỀN - Đã chi chưa có chứng từ"},
                                 ensure_ascii=False), src))
                _db.commit()
                out["dachi_cct"] = round((dachi["cuoi"] or 0) * 1e-9, 9)
                _made.add("DACHI_CCT")
                _ds = _ds or _tw["dataset_id"]
    else:
        out["sd_error"] = f"Không thấy sheet {SD_SHEET}"

    # 2) THUCHI
    tc, sheets, recon = _thuchi_extract(wb, period)
    if tc:
        p = os.path.join(tf.FILLED_DIR, f"TIEN_{period}_03_DONGTIEN.xlsx")
        tf.fill("03_DONGTIEN", tc, p)
        _r = tf.import_filled(p, cong_ty=None, source_file=src,
                              force_grain="month" if _daily_on() else None)
        out["thuchi"] = _r.get("rows_imported")
        _ds = _r.get("dataset_id") or _ds
        _made |= set((_r.get("by_type") or {}))
        out["thuchi_sheets"] = sheets
        if recon:
            # Nguồn tự lệch giữa cột cả kỳ và các cột ngày -> báo ra để người dùng biết KPI
            # tháng sẽ lệch đúng phần này so với báo cáo giấy (KHÔNG tự ý bù/ép cho khớp).
            out["daily_recon_warnings"] = recon
            out["daily_recon_total_lech"] = round(sum(x["lech"] for x in recon), 9)
    if not any(k in out for k in ("sdt", "vay", "thuchi")):
        return {"ok": False, "error": "Không bóc được SDT/VAY/THUCHI nào", **out}
    # Dọn report_type CŨ của CHÍNH file này mà lượt nạp này không còn sinh ra (vd extractor trước
    # đây có emit DTU, nay bỏ -> dòng DTU cũ sẽ sống mãi nếu không dọn).
    pr = tf.prune_stale_types(_ds, src, _made, _before_id)
    if pr.get("pruned"):
        out["pruned_stale"] = pr
    return out


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--period", required=True)
    ap.add_argument("--cong-ty", dest="cong_ty", default=None)
    a = ap.parse_args()
    import json
    print(json.dumps(extract(a.file, a.period, a.cong_ty), ensure_ascii=False))
