# -*- coding: utf-8 -*-
"""CLI cau noi cho BE (DashBoard_AI) goi qua subprocess -> luon in ra 1 dong JSON tren stdout.
Tach biet tien trinh BE (khong import cheo), chay trong venv Dashboard_Agent (co du deps).

Subcommands:
  plan/profile <file>             -> liet ke sheet + canonical_kind_guess + mapping DA HOC +
                                     de xuat scope/basis/man hinh (chua LLM, deterministic)
  propose  <file> --sheet S       -> heuristic (TK131/TK331) hoac goi analyst agent (LLM)
  execute  <file> --sheet S --mapping-file M.json [--period P] [--dry-run]
                                  -> generic_import_execute (ghi GEN_*), tra rows + sample
  autobatch <file> [--propose] [--dry-run] -> nap tat ca sheet wired 1 lan
  confirm <canonical_kind> --scope --basis --target-screen --chi-tieu
                                  -> luu quyet dinh phan loai (1 lan, dung lai cho moi file)

Vi du:
  .venv/bin/python scripts/agent_cli.py profile /abs/f.xlsx
  .venv/bin/python scripts/agent_cli.py execute /abs/f.xlsx --sheet 131 --mapping-file /tmp/m.json --period 2026-01 --dry-run
"""
import argparse
import json
import os
import subprocess
import sys

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.normpath(os.path.join(_HERE, ".."))
sys.path.insert(0, _ROOT)

from dotenv import load_dotenv  # noqa: E402

load_dotenv(os.path.join(_ROOT, ".env"))


def _out(obj):
    """In dung 1 dong JSON (BE doc dong cuoi)."""
    print(json.dumps(obj, ensure_ascii=False, default=str))


def _source_id(path):
    """ĐỊNH DANH NGUỒN ghi vào raw_rows.source_file = '<công_ty_thư_mục>::<tên_file>' (khớp
    source_key UI) để 3 nguồn trùng basename (vd B.6.XVP nhận từ 3 thư mục) KHÔNG đè/che nhau.
    File ngoài received_reports -> basename như cũ."""
    from servers.common import source_catalog as SC
    return SC.source_id_from_path(path)


def cmd_profile(args):
    return cmd_plan(args)  # alias — plan đầy đủ hơn (kèm kpi_hints từ guideline)


def _learned_sheet_mappings(sheet: str, canonical_kind: str | None, all_specs: list) -> list:
    """Khớp SheetMapping đã học TỪ danh sách catalog đã nạp sẵn (B3: không đọc đĩa lại mỗi sheet).
    Khớp theo canonical_kind (tái dùng chéo công ty) hoặc đúng tên sheet. Ưu tiên bản đã
    verified=True (ghi thật thành công) lên đầu — bản chỉ qua dry-run (verified=False) vẫn
    dùng được (khép vòng học sớm) nhưng đáng tin cậy thấp hơn. Trả nguyên SheetMapping (không
    nhét thêm field verified vào — tránh lẫn với schema mapping thật khi dùng lại để execute)."""
    seen, candidates = set(), []
    for r in all_specs:
        sm = r.get("sheet_mapping") or {}
        match = (canonical_kind and sm.get("canonical_kind") == canonical_kind) or sm.get("sheet") == sheet
        fp = r.get("fingerprint")
        if not match or fp in seen:
            continue
        seen.add(fp)
        candidates.append((bool(r.get("verified", True)), sm))
    candidates.sort(key=lambda t: t[0], reverse=True)
    return [sm for _verified, sm in candidates]


def _kpi_hints_from_guideline(sheet: str, canonical_kind: str | None, gl_cache: dict) -> list:
    """Khớp sheet/canonical_kind với kpi_glossary.json (sinh từ guideline.xlsx).
    B3: memoize glossary_lookup theo term (nhiều sheet có thể trùng canonical_kind)."""
    from servers import qa_server as qa

    seen_ids, hints = set(), []
    for term in [canonical_kind, sheet]:
        if not term:
            continue
        if term not in gl_cache:
            gl_cache[term] = qa.glossary_lookup(term)
        gl = gl_cache[term]
        for rec in gl.get("kpi_glossary", []):
            rid = rec.get("id")
            if rid in seen_ids:
                continue
            seen_ids.add(rid)
            hints.append({
                "id": rid,
                "chi_tieu": rec.get("chi_tieu"),
                "nguon_du_lieu": rec.get("nguon_du_lieu"),
                "needs_followup": rec.get("needs_followup", False),
                "nhom_bao_cao": rec.get("nhom_bao_cao"),
                "nhom_con": rec.get("nhom_con"),
            })
    return hints[:8]


def _dashboard_relevance(learned: list, kpi_hints: list, canonical_kind: str | None) -> str:
    """wired = khớp guideline hoặc đã học; optional = có loại chuẩn nhưng chưa có KPI; unused = còn lại."""
    if learned or kpi_hints:
        return "wired"
    if canonical_kind:
        return "optional"
    return "unused"


def cmd_plan(args):
    """Profile + khớp guideline (kpi_glossary) + mapping đã học — không cần LLM.

    B1/B2: chỉ mở workbook 1 lần qua sheet_profile(sheet=None) — nhẹ (chỉ sheetnames +
    canonical_kind_guess). BỎ template_analyze (parse thứ 2): plan chỉ chạy cho template LẠ
    (handoff /import đã lọc file khớp 9 báo cáo / Tháng) nên fixed_report_type luôn null ở đây.
    B3: nạp catalog report_specs + cache glossary 1 lần, khớp trong bộ nhớ cho mọi sheet.
    """
    from servers import ingest_server as ing
    from servers.common import canonical, memory

    routed = ing.sheet_routes(args.file)         # 1 lần mở workbook — route theo TÊN + NỘI DUNG
    all_specs = memory.report_spec_search()      # 1 lần đọc catalog
    gl_cache: dict = {}                          # memoize glossary theo term
    file_name = routed.get("file_name") or os.path.basename(args.file)
    scope_from_filename = canonical.guess_scope(file_name)  # rẻ — chỉ so tên file, không mở sheet
    sheets = []
    counts = {"wired": 0, "optional": 0, "unused": 0}
    unrecognized = []   # KHÔNG drop im lặng: sheet có dữ liệu mà chưa route được -> cần analyst/người
    for r in routed.get("routes", []):
        name = r["sheet"]
        ck = r.get("canonical_kind")             # đã ưu tiên tên, fallback nội dung
        target_sheet = r.get("target_sheet")
        if r.get("status") == "unknown":
            unrecognized.append(name)
        learned = _learned_sheet_mappings(name, ck, all_specs)
        kpi_hints = _kpi_hints_from_guideline(name, ck, gl_cache)
        rel = _dashboard_relevance(learned, kpi_hints, ck)
        counts[rel] = counts.get(rel, 0) + 1
        # Đề xuất chỉ tiêu/màn hình phục vụ (đọc guideline, KHÔNG tự ghi số vào KPI nào —
        # chỉ để hiện gợi ý cho người dùng xác nhận, xem AgentImportPanel + /import/agent/confirm).
        proposed_kpi = kpi_hints[0] if kpi_hints else None
        proposed_screen = (canonical.guess_screen(proposed_kpi.get("nhom_bao_cao", ""),
                                                  proposed_kpi.get("nhom_con", ""))
                          if proposed_kpi else None)
        binding = memory.binding_get(ck) if ck else None  # đã xác nhận từ trước (công ty khác/lần trước)
        sheets.append({
            "sheet": name,
            "canonical_kind_guess": ck,
            "target_sheet": target_sheet,
            "route_via": r.get("route_via"),
            "route_status": r.get("status"),
            "learned_mappings": learned,
            "kpi_hints": kpi_hints,
            "guideline_match": bool(kpi_hints),
            "ready": bool(learned),
            "dashboard_relevance": rel,
            "scope_guess": scope_from_filename,
            "proposed_kpi": proposed_kpi,
            "proposed_screen": proposed_screen,
            "binding": binding,
        })
    _out({
        "ok": True,
        "file_name": file_name,
        "fixed_report_type": None,  # plan chỉ cho template lạ; handoff đã lọc template chuẩn
        "sheets": sheets,
        "sheet_summary": counts,
        "unrecognized_sheets": unrecognized,   # sheet lạ có dữ liệu — hiển thị để không bỏ sót
        "guideline_source": "DashBoard_AI/guideline.xlsx → kpi_glossary.json",
    })


# canonical_kind mà heuristic tất định (không LLM) áp được — sổ tổng hợp công nợ dạng
# header 2 tầng (Đầu kỳ/Phát sinh/Cuối kỳ) x (Nợ/Có) + cột Mã ĐT/Tên ĐT. Đã verify đúng số
# liệu (KH 02010001) trên file thật cho TK131; TK331 cùng mẫu S31-DN nên cấu trúc giống hệt.
_HEURISTIC_CANONICAL_KINDS = {"TK131", "TK331"}

# NGƯỠNG SỐ CỘT BỆNH LÝ: 1 số file (vd CĐPS của HO — 341×16350 cột) có HÀNG NGHÌN cột "ma"
# (phantom/định dạng thừa) -> deriver materialize hàng triệu ô, ngốn ~25s/lượt và bị gọi nhiều
# lượt (thue + tonkho_cdps) -> autofill nghẽn ~5 phút/file. KHÔNG báo cáo tài chính hợp lệ nào có
# >1000 cột DỮ LIỆU -> sheet rộng hơn ngưỡng này bị bỏ qua trong autofill (log rõ, không nuốt thầm).
_MAX_SHEET_COLS = 1000

# SHEET BỎ QUA THEO TÊN (chuẩn hoá: bỏ dấu + lower): 'Sổ nhật ký chung'/NKC = SỔ CÁI THÔ (liệt kê mọi
# bút toán, hàng nghìn dòng) — KHÔNG bao giờ là nguồn chỉ tiêu dashboard (metric lấy từ BCTC:
# KQKD/CĐKT/131/331…). Deriver/LLM đụng vào chỉ tốn thời gian (Trạm sạc/DUAN/HO đều có). Bỏ qua để
# tăng tốc autofill. (Nếu sau này làm 'tuổi nợ'/aging cần ngày phát sinh -> khai thác NKC riêng, gỡ khỏi đây.)
_SKIP_SHEET_NAMES = {"so nhat ky chung", "nkc", "so nhat ky"}

# SỔ CHI TIẾT GIAO DỊCH bỏ qua THEO NGUỒN (thư mục nhận, phần trước '::' của source_file). SRVF
# (VinFast Service) đính kèm sổ chi tiết khổng lồ (154 ~18k dòng, PTTC/PTB2B ~2,5k, 5111/13692/
# 33692 ~1k) KHÔNG phải nguồn chỉ tiêu dashboard — đối chiếu T01–T05: cả 13 report_type đều đến
# từ sheet TẤT ĐỊNH (CĐKT/LNQ1/131/331/152/156/TSCĐ/CĐPS), các sổ này đóng góp 0 dòng nhưng gửi
# LLM thì mỗi lượt cực chậm (18k dòng) + kết quả bị bỏ/trùng. Bỏ qua theo THƯ MỤC 'SRVF' (KHÔNG
# theo cong_ty vì SRVF resolve → 'TC' mà Thịnh Cường có file THẬT riêng). Tên chuẩn hoá tf._norm
# + bỏ khoảng trắng. Hoàn tác: xoá khỏi set.
_SKIP_DETAIL_SHEETS_BY_SOURCE = {
    "SRVF": {"154", "153", "5111", "13692", "33692", "1362", "b2c",
             "ptgf", "ptub", "pttc", "ptb2b", "211+242"},
}


def _khoi_of(file_path: str):
    """Khối của file bctc/kqkd suy từ TÊN FILE 'B.<khối>.<cty>.' (B.5.HT -> Khối Xe tải). None
    nếu tên file không có token khối (vd file thu chi 'BÁO CÁO TIỀN' -> None, dòng tiền theo pháp
    nhân không gán khối). Truyền vào import_filled để STAMP khoi -> file HT hiện ở filter Khối."""
    from servers.common import contract as C
    fn = os.path.basename(file_path)
    return C.khoi_from_filename(fn) or C.khoi_from_path(file_path)


def _forward_fill(cells: list) -> list:
    out, last = [], None
    for c in cells:
        if c not in (None, ""):
            last = c
        out.append(last)
    return out


def _guess_month_token(file_path: str):
    """Tháng (1-12, int) của kỳ suy từ TÊN FILE — dùng để nhận cột 'đầu kỳ này' theo nhãn tháng.
    None nếu không suy được."""
    from servers import template_filler as tf
    p = tf.guess_period(os.path.basename(file_path or ""))  # 'YYYY-MM'
    if p and "-" in p:
        try:
            return int(p.split("-")[1])
        except ValueError:
            return None
    return None


def _label_has_month(norm_label: str, month: int) -> bool:
    """Nhãn cột (đã chuẩn hoá) có ám chỉ ĐÚNG tháng `month` không: bắt 't1'/'thang 1'/'t01'/
    'thang 01' (word-ish). KHÔNG khớp 't10'..'t12' khi month=1 (biên từ)."""
    import re
    return bool(re.search(rf"\b(?:t|thang)\s*0?{month}\b", norm_label))


def _heuristic_tk_mapping(file_path: str, sheet: str, canonical_kind: str):
    """Heuristic TẤT ĐỊNH (không LLM) cho sheet dạng 'sổ tổng hợp công nợ' TK131/TK331:
    dò dòng phụ đề 'Nợ/Có', dòng nhóm kỳ ngay phía trên ('Đầu kỳ'/'Phát sinh'/'Cuối kỳ',
    forward-fill do cell merge), và 2 cột 'Mã ĐT'/'Tên ĐT'. Trả (None, {}) nếu không khớp
    cấu trúc mong đợi (để cmd_propose/cmd_autobatch tự fallback sang LLM) — KHÔNG đoán liều.

    Trả (mapping, meta): meta gồm basis_guess ('luyke'|'thang'|None) suy từ chính text tiêu
    đề đã đọc được trong row_sample (KHÔNG mở thêm file — tận dụng lại lượt đọc này)."""
    from servers import ingest_server as ing
    from servers.common import be_bridge as bb
    from servers.common import canonical

    prof = ing.sheet_profile(file_path, sheet=sheet, max_rows=20, max_cols=15)
    rows = prof.get("row_sample", [])
    norm = lambda v: bb.normalize_header(v, True)  # noqa: E731

    basis_guess = None
    for row in rows[:10]:
        for cell in row:
            if isinstance(cell, str):
                basis_guess = canonical.guess_basis(cell)
                if basis_guess:
                    break
        if basis_guess:
            break

    # Phụ đề Nợ/Có — có bản ghi 'PHẢI THU/ĐÃ THU' (131) hoặc 'PHẢI TRẢ/ĐÃ TRẢ' (331) thay Nợ/Có
    # (DUAN T05/T06 máy tạo). Map theo NGHĨA: phải thu=Nợ (số dư), đã thu=Có; phải trả=Có (số dư),
    # đã trả=Nợ. Giữ nhận 'no'/'co' cũ làm fallback.
    def _side(v):
        t = norm(v)
        # Cột LŨY KẾ (YTD) KHÔNG phải Nợ/Có TRONG KỲ -> loại khỏi nhận vai (nếu không: CĐPS An KS có
        # cả 'Phát sinh Nợ/Có' (trong kỳ, E/F) LẪN 'Nợ/Có lũy kế' (G/H) dưới cùng nhóm 'Phát sinh'
        # (forward-fill) -> cả 2 nhận role phat_sinh_{no,co}, cột lũy kế đứng SAU ghi đè -> tồn kho
        # đọc NHẦM nhập/xuất lũy kế thay vì trong kỳ). Chỉ tiêu #43 cần PHÁT SINH TRONG KỲ.
        if "luy ke" in t:
            return None
        if t.startswith("phai thu") or t.startswith("da tra"):
            return "no"
        if t.startswith("phai tra") or t.startswith("da thu"):
            return "co"
        if t.startswith("no"):
            return "no"
        if t.startswith("co"):
            return "co"
        return None
    sub_idx = None
    for i, row in enumerate(rows):
        hits = sum(1 for c in row if _side(c))
        if hits >= 2:
            sub_idx = i
            break
    if sub_idx is None or sub_idx == 0:
        return None, {"basis_guess": basis_guess}

    group_row = _forward_fill(rows[sub_idx - 1])
    sub_row = rows[sub_idx]
    code_idx = name_idx = None
    for idx, cell in enumerate(rows[sub_idx - 1]):
        t = norm(cell)
        if code_idx is None and (t == "ma" or t.startswith("ma ")
                                 or t.startswith("so tai khoan") or t.startswith("so tk")
                                 or t.startswith("so hieu") or t.startswith("shtk")):   # 'Số tài khoản'/'SHTK'
            code_idx = idx
        elif name_idx is None and (t == "ten" or t.startswith("ten ")):
            name_idx = idx
    if code_idx is None or name_idx is None:
        return None, {"basis_guess": basis_guess}

    # Tháng của kỳ (từ tên file, vd M.202601 -> '1') — dùng để nhận cột 'đầu kỳ này' khi nhãn
    # ghi theo tháng ('SỐ DƯ ĐẦU T1'), phân biệt với cột tháng khác (stale) / đầu năm.
    file_period = _guess_month_token(file_path)

    # TRIẾT LÝ (theo yêu cầu 2026-07-10): CHẮC thì lấy tất định; KHÔNG chắc thì để agent quyết
    # dựa ngữ cảnh tháng tên file. Neo là CUỐI KỲ (thẻ dashboard cần) — hầu như luôn rõ, lấy chắc.
    #   cuối kỳ  : 'cuoi'                         -> LUÔN chắc (anchor)
    #   phát sinh: 'phat sinh'/'ps'               -> chắc
    #   đầu kỳ   : chỉ CHẮC khi (a) đúng 1 cột 'đầu', hoặc (b) nhãn khớp THÁNG file, hoặc
    #              (c) nhãn nói 'năm' (đầu năm — loại khỏi đầu kỳ). Nhiều cột 'đầu' mập mờ
    #              (vd '331' ghi 'ĐẦU T3' trong file THÁNG 1) -> KHÔNG đoán, đánh dấu để AGENT xử.
    columns = [{"index": code_idx, "role": "entity_code"}, {"index": name_idx, "role": "entity_name"}]
    openings = []   # [(idx, side, gt)] các cột 'đầu ...' để phân giải sau
    for idx in range(len(sub_row)):
        if idx in (code_idx, name_idx):
            continue
        side = _side(sub_row[idx])
        if side is None:
            continue
        gt = norm(group_row[idx]) if idx < len(group_row) else ""
        if "cuoi" in gt:
            columns.append({"index": idx, "role": f"cuoi_ky_{side}"})
        elif "phat sinh" in gt or gt == "ps":
            columns.append({"index": idx, "role": f"phat_sinh_{side}"})
        elif "dau" in gt:
            openings.append((idx, side, gt))
        # nhãn khác -> bỏ (không đoán)

    # Phân giải các cột 'đầu': tách nhóm theo index (mỗi nhóm 2 ô Nợ/Có liền nhau).
    opening_idxs = sorted({i for i, _, _ in openings})
    ambiguous_openings = []
    if opening_idxs:
        def _emit(role, only_idxs):
            for i, side, _ in openings:
                if i in only_idxs:
                    columns.append({"index": i, "role": f"{role}_{side}"})
        if len(opening_idxs) <= 2:              # (a) 1 nhóm 'đầu' duy nhất -> chắc chắn đầu kỳ
            _emit("dau_ky", set(opening_idxs))
        else:
            # >1 nhóm: dùng NGHĨA để tách. Nhóm nhãn 'năm' -> đầu năm; nhãn khớp THÁNG file ->
            # đầu kỳ. Nếu vẫn không phân định được đầu-kỳ -> KHÔNG đoán, để agent quyết.
            by_idx = {i: gt for i, _, gt in openings}
            nam = {i for i in opening_idxs if "nam" in by_idx[i]}
            month_hit = {i for i in opening_idxs
                         if file_period and _label_has_month(by_idx[i], file_period)}
            _emit("dau_nam", nam)
            if month_hit:
                _emit("dau_ky", month_hit)       # (b) khớp tháng file -> chắc
                ambiguous_openings = [i for i in opening_idxs if i not in nam | month_hit]
            elif len(opening_idxs) - len(nam) <= 2 and (opening_idxs and set(opening_idxs) - nam):
                # còn đúng 1 nhóm 'đầu' sau khi loại 'năm' -> nhóm đó là đầu kỳ (chắc)
                _emit("dau_ky", set(opening_idxs) - nam)
            else:
                # (c) QUY ƯỚC nhãn: nhãn TRƠN nhất ('số dư đầu' KHÔNG kèm tháng/qualifier) là
                # đầu kỳ NÀY; nhãn kèm token thừa ('đầu T3') là kỳ khác/gốc. Chỉ nhận khi có
                # DUY NHẤT 1 nhãn ngắn nhất và mọi nhãn khác dài hơn hẳn -> vẫn "chắc theo nghĩa".
                non_nam = [i for i in opening_idxs if i not in nam]
                labs = {i: by_idx[i] for i in non_nam}
                if non_nam and len(set(labs.values())) > 1:
                    minlen = min(len(s.split()) for s in labs.values())
                    plain = {i for i in non_nam if len(labs[i].split()) == minlen}
                    plain_labels = {labs[i] for i in plain}
                    if len(plain_labels) == 1 and all(len(labs[i].split()) > minlen
                                                      for i in non_nam if i not in plain):
                        _emit("dau_ky", plain)
                        _emit("dau_nam", set(non_nam) - plain)
                    else:
                        ambiguous_openings = non_nam   # vẫn mập mờ -> để agent
                else:
                    # nhiều cột 'đầu' cùng nhãn/không tách được -> ĐỂ AGENT (không đoán)
                    ambiguous_openings = non_nam

    # BẮT BUỘC có cột CUỐI KỲ (số dư cuối = số thẻ dashboard đọc); thiếu -> không tin cấu trúc.
    if not any(c["role"].startswith("cuoi_ky") for c in columns):
        return None, {"basis_guess": basis_guess}

    # BỎ dòng subtotal ('Tổng cộng'/'Cộng') ở đầu bảng -> tránh đếm đôi (dòng 'Tổng cộng' của
    # sheet 131 HT nằm NGAY dòng dữ liệu đầu). Dời data_start_row qua các dòng subtotal liền kề.
    data_start = sub_idx + 1
    while data_start < len(rows):
        # Dò subtotal 'Cộng'/'Tổng cộng' PHẢI phân biệt DẤU: 'cộng'(tổng) ≠ 'công'(công ty). norm bỏ
        # dấu -> 'cong ty'(khách hàng) bị nhận nhầm 'cộng' -> skip sạch data (lỗi DUAN 131: mọi KH
        # tên 'Công ty …' bị bỏ, data_start nhảy tới cuối). Dùng chuỗi RAW có dấu.
        raw = str(rows[data_start][name_idx] or "").strip().lower() if name_idx < len(rows[data_start]) else ""
        code_raw = str(rows[data_start][code_idx] or "").strip() if code_idx < len(rows[data_start]) else ""
        # 'cộng <x>' chỉ là dòng tổng khi KHÔNG có mã đối tượng ('Cộng tác viên …' là NCC thật, có mã)
        if (raw == "cộng" or raw.startswith("tổng cộng")
                or (raw.startswith("cộng ") and not code_raw)):
            data_start += 1
        else:
            break

    mapping = {
        "orientation": "row_major", "data_start_row": data_start,
        "target_report_type": canonical.canonical_gen_code(canonical_kind), "canonical_kind": canonical_kind,
        "columns": columns, "header_rows": [sub_idx - 1, sub_idx],
        "notes": ["heuristic tất định: header 2 tầng (kỳ) x (Nợ/Có) — không qua LLM"],
    }
    # ambiguous_openings != [] -> có cột 'đầu' KHÔNG chắc vai (đầu kỳ/kỳ trước) -> đường trên
    # nên để AGENT phân giải các cột này dựa file_period (tháng của file). CUỐI KỲ vẫn tất định
    # (thẻ dashboard chính đã đủ). partial=True báo "lấy chắc phần cuối kỳ, đầu kỳ cần agent".
    meta = {"basis_guess": basis_guess, "file_period": file_period}
    if ambiguous_openings:
        meta["ambiguous_opening_cols"] = ambiguous_openings
        meta["partial"] = True
    return mapping, meta


# Identity KẾ TOÁN theo loại tài khoản (chuẩn, KHÔNG hardcode per-file): cột nào là số dư/PS
# được CHỌN THEO VAI ngữ nghĩa detector dò ra, còn công thức (nợ/có cho phải thu/phải trả) là
# đẳng thức kế toán cố định theo canonical_kind.
#   TK131 (phải thu, TK dư NỢ): số dư = cuối kỳ NỢ; tăng = PS NỢ; giảm = PS CÓ.
#   TK331 (phải trả, TK dư CÓ): số dư = cuối kỳ CÓ; tăng = PS CÓ; giảm = PS NỢ.
# Số dư đầu/cuối = NET hai chiều (guide #30/#36): TK131 phải thu = Nợ − Có; TK331 phải trả = Có − Nợ
# (khoản dư NGƯỢC chiều = trả trước NCC / người mua trả trước, tính ÂM). Bắt buộc để bảng "Biến động"
# cân: Cuối = Đầu + Tăng − Giảm (Tăng/Giảm là PS GỘP 1 chiều; nếu Đầu/Cuối chỉ lấy 1 cột thì PS chiều
# kia (vd trả trước NCC) làm lệch — HT T01 đầu 5,37 + tăng 2,25 − giảm 33,09 ≠ cuối 2,84). bal_opp/
# open_opp = cột NGƯỢC chiều để trừ ra số dư ròng. Trả trước vẫn hiện riêng ở thẻ PTRA_ADV/PTHU_ADV.
_CONGNO_IDENTITY = {
    "TK131": {"target": "05_PHAITHU", "ten_col": "Khách hàng",
              "bal": "cuoi_ky_no", "bal_opp": "cuoi_ky_co", "open": "dau_ky_no", "open_opp": "dau_ky_co",
              "inc": "phat_sinh_no", "dec": "phat_sinh_co",
              "inc_col": "PS tăng - Nợ (tỷ)", "dec_col": "PS giảm - Có (tỷ)"},
    "TK331": {"target": "06_PHAITRA", "ten_col": "Nhà cung cấp",
              "bal": "cuoi_ky_co", "bal_opp": "cuoi_ky_no", "open": "dau_ky_co", "open_opp": "dau_ky_no",
              "inc": "phat_sinh_co", "dec": "phat_sinh_no",
              "inc_col": "PS tăng - Có (tỷ)", "dec_col": "PS giảm - Nợ (tỷ)"},
}


# Mã số TT200 -> Chỉ tiêu KQKD chuẩn (tên importer khớp để sinh DTHU/HQKD/PNLT + thẻ LNST).
# 'Doanh thu thuần'->DTHU+HQKD1000; 'Tổng chi phí'->HQKD1047; 'Lợi nhuận trước thuế'->HQKD1112;
# 'Lợi nhuận sau thuế'->PNLT (thẻ LNST khớp ILIKE '%lợi nhuận%sau thu%'); còn lại -> PNLT cơ cấu.
_TT200_CHITIEU = {
    "10": "Doanh thu thuần", "11": "Giá vốn hàng bán", "20": "Lợi nhuận gộp",
    "21": "Doanh thu hoạt động tài chính", "22": "Chi phí tài chính",
    "25": "Chi phí bán hàng", "26": "Chi phí quản lý doanh nghiệp",
    "50": "Lợi nhuận trước thuế", "51": "Chi phí thuế TNDN", "60": "Lợi nhuận sau thuế",
}


def _derive_kqkd_tseries(rows, period, cong_ty, file_path):
    """TẤT ĐỊNH cho P&L layout T-SERIES (HT 'kqkdQT'/'kqkd hưng thịnh'): mã T100/T200/T201/T202/
    T203/T300 ở CỘT A; header có 'Năm 2026' (LŨY KẾ) + 'Tháng 1'..'Tháng N'. LẤY cột 'Tháng {mm}'
    = SỐ TỪNG THÁNG (mm = tháng của kỳ, đọc từ period) — KHÔNG lấy 'Năm 2026' (lũy kế). Map ->
    01_HQKD y như đường TT200: 'Doanh thu thuần'(T100)->1000+DTHU · 'Tổng chi phí'(T200)->1047 ·
    'Lợi nhuận trước thuế'(T300)->1112 · các mã còn lại giữ nhãn gốc -> PNLT.
    Trả None nếu KHÔNG phải T-series (để _derive_kqkd chạy tiếp nhánh TT200 — 0 ảnh hưởng cty khác)."""
    import re as _re2
    from servers import template_filler as tf
    from servers.common import be_bridge as bb
    _norm = lambda v: bb.normalize_header(v, True)  # noqa: E731
    # 1) Nhận diện T-series + CỘT MÃ. HT: mã T100/T200/T300 ở CỘT A. Trạm sạc ('BC HQKD'): mã ở CỘT C
    #    (cột A/B là số/'MHN'). Quét cột 0..4, lấy cột ĐẦU có mã T\d{3}. Không có -> None (cty khác).
    def _has_tcode_col(j):
        return any(_re2.fullmatch(r"T\d{3}", str(r[j]).strip()) for r in rows
                   if r and j < len(r) and r[j] not in (None, ""))
    code_j = next((j for j in range(5) if _has_tcode_col(j)), None)
    if code_j is None:
        return None
    lab_j = code_j + 1              # nhãn chỉ tiêu ngay sau cột mã (HT cột B / Trạm sạc cột D 'Chỉ tiêu')
    try:
        mm = int(str(period)[-2:])
    except Exception:
        return {"ok": False, "error": "T-series: period không hợp lệ"}
    # 2) Cột giá trị THÁNG. HT: 'Tháng {mm}'. Trạm sạc: 'T{mm}' (vd 'T01'). Cột 'Năm 2026' = LŨY KẾ -> loại.
    _mth = lambda s: _norm(s).startswith("thang ") or bool(_re2.fullmatch(r"t\d{1,2}", _norm(s)))  # noqa: E731
    hdr_i = next((i for i, r in enumerate(rows[:14])
                  if any(_norm(c) == "nam 2026" or _mth(c) for c in r)), None)
    if hdr_i is None:
        return {"ok": False, "error": "T-series: không thấy header Năm 2026/Tháng/T{mm}"}
    hdr = rows[hdr_i]
    _cur = (f"thang {mm}", f"t{mm:02d}", f"t{mm}")
    # val_j phải Ở SAU cột mã (j>code_j) -> loại cột A rác (Trạm sạc header col A cố định 'T04' bất kể
    # tháng, khớp nhầm 't04' ở file T04). Trạm sạc: cột tháng dồn tích (T01..T12) -> khớp ĐÚNG 'T{mm}'.
    val_j = next((j for j, c in enumerate(hdr) if j > code_j and _norm(c) in _cur), None)
    if val_j is None:               # fallback: cột tháng CUỐI sau cột mã (loại 'Năm 2026' lũy kế)
        mcols = [j for j, c in enumerate(hdr) if j > code_j and _mth(c)]
        val_j = mcols[-1] if mcols else None
    if val_j is None:
        return {"ok": False, "error": f"T-series: không thấy cột 'Tháng {mm}'/'T{mm}'"}

    def tnum(r):
        v = r[val_j] if val_j < len(r) else None
        return round(v * 1e-9, 9) if isinstance(v, (int, float)) else None
    # 3) Gom mã T -> (nhãn gốc, value tỷ). Lấy dòng ĐẦU cho mỗi mã.
    byco = {}
    for r in rows[hdr_i + 1:]:
        c = str(r[code_j]).strip() if r and code_j < len(r) and r[code_j] not in (None, "") else ""
        if _re2.fullmatch(r"T\d{3}(\.\d+)?", c) and c not in byco:
            lab = str(r[lab_j]).strip() if lab_j < len(r) and r[lab_j] not in (None, "") else c
            byco[c] = (lab, tnum(r))
    if "T100" not in byco or "T300" not in byco:
        return {"ok": False, "error": "T-series: thiếu T100/T300"}
    records = []

    def add(ten, val):
        if val is not None:
            records.append({"Kỳ (yyyy-mm)": period, "Chỉ tiêu KQKD": ten, "Thực hiện (tỷ)": val})
    # Doanh thu thuần = Σ T101.x (DT bán hàng theo dòng xe), KHÔNG dùng T100 (dòng TỔNG DOANH THU
    # = ΣT101.x + T102 DT tài chính + T103 thu nhập khác -> thừa; chốt user 2026-07-19 đối chiếu file
    # gốc T05: T100=239,3438 = ΣT101.x(238,6787)+T102(0,0469)+T103(0,6182)). CỘT giá trị vẫn val_j =
    # ĐÚNG THÁNG đang xử lý (dò theo cột 'Tháng {mm}'/'T{mm}' ở trên) — KHÔNG đổi cách chọn cột.
    _t101_subs = [val for c, (lab, val) in byco.items()
                  if _re2.fullmatch(r"T101\.\d+", c) and val is not None]
    _dt_ban_hang = byco["T101"][1] if "T101" in byco else (
        round(sum(_t101_subs), 9) if _t101_subs else None)
    # 3 dòng CHUẨN -> HQKD 1000/1047/1112 (tên khớp luật importer). T200 = T100 − T300 (đẳng thức).
    add("Doanh thu thuần", _dt_ban_hang if _dt_ban_hang is not None else byco["T100"][1])
    if "T200" in byco:
        add("Tổng chi phí", byco["T200"][1])
    add("Lợi nhuận trước thuế", byco["T300"][1])
    add("Lợi nhuận sau thuế", byco["T300"][1])   # -> PNLT (nuôi thẻ LNST). HT che_do_phan_tich: TỔNG
    # CHI PHÍ T200 ĐÃ GỒM thuế TNDN (kế toán HT xác nhận) -> T300 "trước thuế" thực chất ĐÃ SAU THUẾ
    # -> LNST = T300. Nhất quán HO/SRVF (P&L quản trị không tách dòng thuế riêng).
    # Chi tiết còn lại -> PNLT (giữ nhãn gốc); BỎ T100/T200/T300 đã emit (tránh trùng 1000/1047/1112).
    # T201 (TỔNG giá vốn) -> CHUẨN HOÁ chính tả nhãn thành 'Giá vốn hàng bán': nguồn HT gõ 'Gía' (dấu
    # sắc trên i, không phải a) nên metrics build_revenue (cogs = PNLT ILIKE '%giá vốn%', ACCENT-SENSITIVE)
    # KHÔNG khớp -> giá vốn + tỷ lệ GV trống MỌI kỳ. Breakdown T201.x GIỮ nhãn gốc (typo) -> KHÔNG khớp
    # filter -> KHÔNG đếm đôi (cogs chỉ cộng ĐÚNG dòng TỔNG chuẩn hoá này).
    # CHUẨN HOÁ nhãn để metrics gom NHẤT QUÁN theo chỉ tiêu (thẻ DT tài chính / Thu nhập khác 2 màn):
    # T201->'Giá vốn hàng bán'; T103 (thu nhập khác) nguồn gõ khác nhau ('Thu nhập khác' Trạm sạc vs
    # 'Doanh thu khác' HT) -> ép 'Thu nhập khác'. T102 nhãn gốc ('Doanh thu hoạt động tài chính') đã
    # khớp ILIKE '%doanh thu%tài chính%' nên GIỮ NGUYÊN (surgical). Mã khác giữ nhãn gốc.
    _canon = {"T201": "Giá vốn hàng bán", "T103": "Thu nhập khác"}
    for c, (lab, val) in byco.items():
        if c in ("T100", "T200", "T300"):
            continue
        add(_canon.get(c, lab), val)
    # Lợi nhuận gộp = T100 − T201 (sheet HT KHÔNG có dòng LN gộp riêng, spec #6 = DT thuần − giá vốn).
    # -> PNLT (metrics gross = ILIKE '%lợi nhuận gộp%'; biên LNG = LNG/DT thuần tính ở BE). Chỉ khi có
    # T201 -> đơn vị T-series không có giá vốn (Trạm sạc) KHÔNG bị thêm dòng (0 ảnh hưởng).
    _gv201, _dt100 = byco.get("T201", (None, None))[1], byco["T100"][1]
    if _gv201 is not None and _dt100 is not None:
        add("Lợi nhuận gộp", round(_dt100 - _gv201, 9))
    # Doanh thu HH, DV (chỉ tiêu #1 bảng 50 = doanh thu bán hàng) = CÙNG giá trị vừa dùng cho
    # "Doanh thu thuần" (_dt_ban_hang, Σ T101.x hoặc T101 gộp) — 2 khái niệm trùng nhau ở đơn vị này.
    if _dt_ban_hang is not None:
        add("Doanh thu HH, DV", _dt_ban_hang)
    out = os.path.join(tf.FILLED_DIR, f"KQKD_{period}_{cong_ty or 'NA'}_01_HQKD.xlsx")
    tf.fill("01_HQKD", records, out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    # ---- 02_CHIPHI: các mã CON TRỰC TIẾP của T200 (^T2\d{2}$ trừ T200: T201 giá vốn, T202 tài
    # chính, T203 vận hành) — tổng = T200 = 1047 (khớp byNhom==byKhoi). Thay CHIPHI LLM (dump cả P&L
    # gồm doanh thu -> đếm trội ~7×). Nhãn nhóm = tên gốc; cột giá trị = val_j 'Tháng {mm}' như P&L.
    # 02_CHIPHI byNhom: dựng khi các mã T2xx con TRỰC TIẾP PHỦ HẾT T200 (Σ con ≈ T200) — đúng cho HT
    # (T201 giá vốn + T202 tài chính + T203 vận hành = T200) BẤT KỂ mã ở cột nào (trước gate code_j==0
    # để phân biệt HT cột A / Trạm sạc cột C; nhưng HT dựng lại file -> mã sang CỘT C nên gate cũ làm
    # 02_CHIPHI HT RỖNG). Trạm sạc: giá vốn KHÔNG có T-code -> T2xx con không phủ hết T200 -> Σ lệch ->
    # BỎ byNhom (tránh breakdown thiếu; chi phí TỔNG 1047 vẫn ĐÚNG từ T200).
    _cp_recs = _chiphi_recs_tseries(rows, code_j, lab_j, val_j, period)
    _t200 = byco.get("T200", (None, None))[1]
    _cp_sum = round(sum(r["Thực hiện (tỷ)"] for r in _cp_recs), 6)
    _covers = _t200 not in (None, 0) and abs(_cp_sum - _t200) <= abs(_t200) * 0.01   # trong 1%
    _cpr = _fill_import_chiphi(_cp_recs, period, cong_ty, file_path) if (_cp_recs and _covers) else None
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"), "target": "01_HQKD",
            "value_col_header": f"T{mm:02d}", "chiphi": _cpr}


def _chiphi_recs_tseries(rows, code_j, lab_j, val_j, period):
    """Dòng 02_CHIPHI cho HT (P&L T-series): các mã CON TRỰC TIẾP của T200 = '^T2\\d{2}$' (trừ T200)
    -> T201 giá vốn / T202 CP tài chính / T203 CP vận hành; tổng = T200. Mã ở CỘT code_j, nhãn ở CỘT
    lab_j (HT dựng lại file: mã sang cột C nên KHÔNG hard-code cột 0/1). Giá trị tại val_j (cột tháng,
    đúng cột P&L). dim1=dim3=nhãn gốc. Bỏ dòng 0/None. (T2xx.y là con của T2xx -> loại để không cộng trùng.)"""
    import re as _re
    recs = []
    for r in rows:
        cd = str(r[code_j]).strip() if r and code_j < len(r) and r[code_j] not in (None, "") else ""
        if not _re.fullmatch(r"T2\d{2}", cd) or cd == "T200":
            continue
        val = r[val_j] if val_j < len(r) else None
        if not isinstance(val, (int, float)) or val == 0:
            continue
        nm = str(r[lab_j]).strip() if lab_j < len(r) and r[lab_j] not in (None, "") else cd
        recs.append({"Kỳ (yyyy-mm)": period, "Nhóm CP (chuẩn mực KT)": nm,
                     "Khoản mục chi tiết": nm, "Thực hiện (tỷ)": round(val * 1e-9, 9)})
    return recs


def _sheet_has_tcodes(file_path, sheet):
    """True nếu CỘT A của sheet có mã P&L T-series (T100/T200/T300...) — tức sheet P&L 'thật' của HT
    (dù ĐỔI TÊN theo tháng). Dùng để KHÔNG bỏ nhầm sheet P&L T-series khi tên không nằm trong whitelist
    p_and_l (whitelist chỉ nhằm chặn sheet 'kqkd' TT200 ĐÓNG BĂNG — sheet đó không có T-code). Cty TT200
    khác không có T-code -> trả False -> whitelist giữ nguyên hành vi cũ (0 ảnh hưởng)."""
    import re as _re2
    from servers.common import be_bridge as bb
    try:
        wb = bb.fast_load_workbook(file_path, read_only=True, data_only=True)
        try:
            if sheet not in wb.sheetnames:
                return False
            for i, row in enumerate(wb[sheet].iter_rows(values_only=True)):
                if i > 80:
                    break
                c0 = row[0] if row else None
                if c0 is not None and _re2.fullmatch(r"T\d{3}", str(c0).strip()):
                    return True
        finally:
            wb.close()
    except Exception:
        return False
    return False


def _derive_kqkd_duan(rows, period, cong_ty, file_path):
    """TẤT ĐỊNH cho P&L Khối DỰ ÁN (DUAN sheet 'HQKD'): dòng CÓ TÊN ở cột B (idx 1), cột giá trị =
    cột header "Tổng dự án" (tổng toàn Khối; thứ tự cột dự án đổi theo tháng → khớp theo TÊN, không
    theo vị trí). Dòng: 'Tổng doanh thu ròng'/'Giá vốn'/'Lợi nhuận gộp'/'Chi phí phân bổ HO'/'Thu nhập
    trước thuế'/'Thu nhập ròng'. Map -> 01_HQKD: DThu(1000)=Tổng DT ròng; Tổng chi phí(1047)=Tổng DT
    ròng − Thu nhập trước thuế (=giá vốn+CP phân bổ HO); LNTT(1112)=Thu nhập trước thuế; giá vốn/LN
    gộp/CP phân bổ HO/LNST -> PNLT. Trả None nếu KHÔNG phải layout DUAN (0 ảnh hưởng cty khác)."""
    from servers import template_filler as tf
    from servers.common import be_bridge as bb
    _norm = lambda v: bb.remove_diacritics("" if v is None else str(v)).strip().lower()  # noqa: E731

    def _lbl(r):
        return _norm(r[1]) if len(r) > 1 and r[1] not in (None, "") else ""
    # 1) Nhận diện DUAN HQKD: cột B có "tong doanh thu rong" VÀ "chi phi phan bo ho".
    labs = [_lbl(r) for r in rows]
    if "tong doanh thu rong" not in labs or not any("chi phi phan bo ho" in k for k in labs):
        return None
    # 2) CỘT giá trị = cột có header "Tổng dự án" (dò trong 12 dòng đầu).
    val_j = next((j for r in rows[:12] for j, c in enumerate(r) if _norm(c) == "tong du an"), None)
    if val_j is None:
        return {"ok": False, "error": "DUAN HQKD: không thấy cột 'Tổng dự án'"}

    def val_of(*subs):
        for r in rows:
            if any(s in _lbl(r) for s in subs):
                v = r[val_j] if val_j < len(r) else None
                return round(v * 1e-9, 9) if isinstance(v, (int, float)) else None
        return None
    dt, lntt = val_of("tong doanh thu rong"), val_of("thu nhap truoc thue")
    if dt is None or lntt is None:
        return {"ok": False, "error": "DUAN HQKD: thiếu Tổng DT ròng / Thu nhập trước thuế"}

    def _sum_between(start_kw, end_kw):
        """Tổng cột val_j các dòng NẰM GIỮA dòng start_kw và dòng end_kw (không gồm 2 mốc)."""
        si = next((i for i, k in enumerate(labs) if start_kw in k), None)
        if si is None:
            return None
        ei = next((i for i in range(si + 1, len(labs)) if end_kw in labs[i]), len(labs))
        s, got = 0.0, False
        for r in rows[si + 1:ei]:
            v = r[val_j] if val_j < len(r) else None
            if isinstance(v, (int, float)):
                s += v
                got = True
        return round(s * 1e-9, 9) if got else None
    records = []

    def add(ten, v):
        if v is not None:
            records.append({"Kỳ (yyyy-mm)": period, "Chỉ tiêu KQKD": ten, "Thực hiện (tỷ)": v})
    # BREAKDOWN THEO DỰ ÁN (chốt 2026-07-23, 5 chỉ tiêu): 8 cột dự án (Cao Bằng/Tân Thịnh/Lạng Sơn/
    # Núi Pháo/Quang Sơn/Quảng Ngãi/Yên Bình/Phú Quốc) đứng TRƯỚC cột "Tổng dự án" — verify Σ 8 dự
    # án = ĐÚNG cột tổng mọi chỉ tiêu gốc (DT ròng/Giá vốn/LN gộp/CP phân bổ HO, cả 6 tháng). VỊ TRÍ
    # CỘT ĐỔI THEO THÁNG (T06 hoán vị Tân Thịnh<->Quang Sơn) -> PHẢI dò theo TÊN, không hardcode index.
    # ⚠️ "Thu nhập trước thuế"/"Thu nhập ròng" theo TỪNG DỰ ÁN KHÔNG lấy trực tiếp ô gốc: T06 phát
    # hiện ô per-dự-án SAI lệch 1,5 tỷ so với cân đối ngược từ LN gộp/CP phân bổ HO/lãi vay (cột
    # TỔNG vẫn đúng — lỗi nằm ở NGUỒN, chỉ ô per-dự-án). Tính LẠI: LNTT_dự_án = LN gộp − CP phân bổ
    # HO − Chi phí lãi vay (dòng RIÊNG "Chi phí lãi vay", khác "...lãi vay HO" đã gộp trong CP phân
    # bổ HO; chỉ có ở 1 số tháng) -> verify khớp cột Tổng dự án CẢ 6 THÁNG (kể cả T06).
    # Mã dự án theo trust_me_bro.xlsx: CB_DA/LS_DA/QS_DA/PQ_DA; ⚠️TT_DA="Yên Bình", YB_DA="Tân Thịnh"
    # (NGƯỢC viết tắt, xác nhận nguồn — [[audit-fixes-2026-07-09]], ĐỪNG "sửa cho xuôi"). Núi Pháo/
    # Quảng Ngãi KHÔNG có trong MD_COSTCENTER -> mã tự đặt NUIPHAO_DA/QUANGNGAI_DA (backfill cong_ty
    # =TC qua import_filled, giống pattern HO_XVP/B2B_SR).
    _DA_PROJECT_CC = [
        ("cao bang", "CB_DA"), ("tan thinh", "YB_DA"), ("lang son", "LS_DA"),
        ("nui phao", "NUIPHAO_DA"), ("quang son", "QS_DA"), ("quang ngai", "QUANGNGAI_DA"),
        ("yen binh", "TT_DA"), ("phu quoc", "PQ_DA"),
    ]

    def _find_da_col(kw):
        return next((j for r in rows[:12] for j, c in enumerate(r)
                     if isinstance(c, str) and _norm(c) == kw), None)
    da_cols = [(cc, j) for kw, cc in _DA_PROJECT_CC for j in [_find_da_col(kw)] if j is not None]

    def vj_exact(lbl_exact, j):
        for r in rows:
            if _lbl(r) == lbl_exact:
                x = r[j] if j < len(r) else None
                return x if isinstance(x, (int, float)) else 0.0
        return 0.0

    def addcc(cc, ten, val):
        if val is not None:
            records.append({"Kỳ (yyyy-mm)": period, "Chỉ tiêu KQKD": ten,
                            "Thực hiện (tỷ)": val, _CC_COL: cc})
    if len(da_cols) >= 6:   # đủ dự án mới tách theo CC; thiếu (đổi layout) -> fallback TỔNG như cũ
        for cc, j in da_cols:
            dt_p = vj_exact("tong doanh thu rong", j)
            giavon_p = vj_exact("gia von", j)
            lngop_p = vj_exact("loi nhuan gop", j)
            cpho_p = vj_exact("chi phi phan bo ho", j)
            laivay_p = vj_exact("chi phi lai vay", j)   # dòng RIÊNG (≠ '...lãi vay HO' đã gộp trong CPHO)
            lntt_p = lngop_p - cpho_p - laivay_p
            tongcp_p = dt_p - lntt_p
            addcc(cc, "Doanh thu thuần", round(dt_p * 1e-9, 9))            # -> HQKD 1000 + DTHU
            addcc(cc, "Tổng chi phí", round(tongcp_p * 1e-9, 9))          # -> HQKD 1047
            addcc(cc, "Giá vốn hàng bán", round(giavon_p * 1e-9, 9))       # -> PNLT
            addcc(cc, "Lợi nhuận gộp", round(lngop_p * 1e-9, 9))          # -> PNLT
            addcc(cc, "Lợi nhuận sau thuế", round(lntt_p * 1e-9, 9))       # -> PNLT (Thuế TNDN=0 trong P&L này -> LNST=LNTT)
        add("Lợi nhuận trước thuế", lntt)              # -> HQKD 1112 (giữ TỔNG, chưa tách CC theo yêu cầu)
    else:
        add("Doanh thu thuần", dt)                       # -> HQKD 1000 + DTHU
        add("Tổng chi phí", round(dt - lntt, 9))         # -> HQKD 1047 (= giá vốn + CP phân bổ HO); 9 số lẻ = giữ đồng (khớp DThu/LNTT), tránh làm tròn nghìn
        add("Lợi nhuận trước thuế", lntt)                # -> HQKD 1112
        add("Giá vốn hàng bán", val_of("gia von"))       # -> PNLT
        add("Lợi nhuận gộp", val_of("loi nhuan gop"))    # -> PNLT
        add("Lợi nhuận sau thuế", val_of("thu nhap rong"))        # -> PNLT (nuôi thẻ LNST)
    add("Chi phí phân bổ HO", val_of("chi phi phan bo ho"))   # -> PNLT (giữ TỔNG, chưa tách CC)
    # Doanh thu HH, DV (#1 bảng 50): tổng các dòng doanh thu HĐ giữa 'Tổng doanh thu ròng' và
    # 'Thu khác' (= S7:S11: DT thực hiện DA + bán dầu + bán vật tư + tiền ăn + thu dương dầu),
    # KHÔNG gồm 'Thu khác'/điều chỉnh. Cột giá trị = val_j ('Tổng dự án'). -> PNLT (metrics #1).
    add("Doanh thu HH, DV", _sum_between("tong doanh thu rong", "thu khac"))
    out = os.path.join(tf.FILLED_DIR, f"KQKD_{period}_{cong_ty or 'NA'}_01_HQKD.xlsx")
    tf.fill("01_HQKD", records, out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    # ---- 02_CHIPHI (cơ cấu chi phí, TẤT ĐỊNH): mục con 'Giá vốn' + mục con 'Chi phí phân bổ HO'
    # (Σ == mã 1047 = Tổng chi phí GỒM giá vốn) -> byNhom KHỚP tile Tổng chi phí & convention
    # byNhom==byKhoi (như XDV). Guide #10 chỉ liệt kê mục con CP phân bổ HO, nhưng byNhom phải
    # reconcile tổng 1047 nên gồm cả mục con giá vốn (khớp checker Đỗ Thu Hằng 87,21). ----
    _cpr = _fill_import_chiphi(_chiphi_recs_duan(rows, val_j, period), period, cong_ty, file_path)
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"),
            "target": "01_HQKD", "value_col_header": "Tổng dự án", "chiphi": _cpr}


def _chiphi_recs_duan(rows, val_j, period):
    """Dòng 02_CHIPHI cho Khối DỰ ÁN (sheet HQKD, cột 'Tổng dự án' = val_j). Cơ cấu chi phí =
    TỪNG MỤC CON của 'Giá vốn' (NVL, nhân công TT, nhiên liệu, khấu hao, bán buôn, thầu phụ, khác
    tại dự án) + TỪNG MỤC CON của 'Chi phí phân bổ HO' (lương+BH, xăng xe HO, khấu hao HO, bảo dưỡng
    HO, NH HO, khác HO, lãi vay HO). Σ == mã 1047 (Giá vốn + CP phân bổ HO). dim1(Nhóm)=dim3(chi
    tiết)=nhãn gốc; dim2 để trống (chuẩn template). Lấy dòng NẰM GIỮA mốc, bỏ dòng 0/None/nhãn rỗng.
    Robust theo file mới: mọi mục con thêm/bớt trong 2 block tự vào cơ cấu (khớp theo TÊN, ko vị trí)."""
    from servers.common import be_bridge as bb
    _norm = lambda v: bb.remove_diacritics("" if v is None else str(v)).strip().lower()  # noqa: E731
    labs = [(_norm(r[1]) if len(r) > 1 and r[1] not in (None, "") else "") for r in rows]

    def _block(start_kw, end_kw, start_exact=False):
        si = next((i for i, k in enumerate(labs)
                   if (k == start_kw if start_exact else start_kw in k)), None)
        if si is None:
            return []
        ei = next((i for i in range(si + 1, len(labs)) if end_kw in labs[i]), len(labs))
        out = []
        for i in range(si + 1, ei):
            r = rows[i]
            ten = str(r[1]).strip() if len(r) > 1 and r[1] not in (None, "") else ""
            v = r[val_j] if val_j < len(r) else None
            if ten and isinstance(v, (int, float)) and v:
                out.append((ten, round(v * 1e-9, 9)))
        return out
    # 'gia von' khớp CHÍNH XÁC (tránh trùng mục con chứa 'giá vốn'); CP phân bổ HO khớp chứa.
    # Block 3 = 'Chi phí lãi vay' RIÊNG giữa 'Thu nhập trước lãi suất & thuế' và 'Thu nhập trước
    # thuế' (một số tháng T02/T04 có; T06 = 0) -> Σ đủ 1047 (= DT − LNTT = giá vốn + CP HO + lãi vay).
    lines = _block("gia von", "loi nhuan gop", start_exact=True) \
        + _block("chi phi phan bo ho", "thu nhap truoc") \
        + _block("thu nhap truoc lai suat", "thu nhap truoc thue")
    return [{"Kỳ (yyyy-mm)": period, "Nhóm CP (chuẩn mực KT)": ten,
             "Khoản mục chi tiết": ten, "Thực hiện (tỷ)": v} for ten, v in lines]


def _fill_import_chiphi(recs, period, cong_ty, file_path):
    """Điền + nạp 02_CHIPHI (report_type CHIPHI). CHIPHI TÁCH BIỆT 01_HQKD nên import cùng
    source_file KHÔNG đè P&L (delete-scope theo report_type). cong_ty resolve từ TÊN FILE GỐC
    ('B.<n>.<MÃ>.') -> đóng dấu ĐÚNG pháp nhân (vd SRVF->TC) khớp CHIPHI đơn vị khác (khác 01_HQKD
    để cong_ty=None vì tên file điền không mang token). Bỏ dòng giá trị 0/None. Trả rows_imported."""
    from servers import template_filler as tf
    from servers.common import contract as C
    recs = [r for r in (recs or [])
            if isinstance(r.get("Thực hiện (tỷ)"), (int, float)) and r.get("Thực hiện (tỷ)")]
    if not recs:
        return {"ok": False, "error": "CHIPHI: không có dòng nhóm chi phí (giá trị != 0)"}
    out = os.path.join(tf.FILLED_DIR, f"CHIPHI_{period}_{cong_ty or 'NA'}_02_CHIPHI.xlsx")
    tf.fill("02_CHIPHI", recs, out)
    cty = C.resolve_company(cong_ty, os.path.basename(file_path)) or cong_ty
    imp = tf.import_filled(out, cong_ty=cty, khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"), "target": "02_CHIPHI"}


def _derive_kqkd_xdv(rows, period, cong_ty, file_path):
    """TẤT ĐỊNH cho P&L Khối XDV (Vinfast Xưởng Dịch Vụ, 'Sheet1' — mã B-series ở cột A: B100/B210
    DThu, B300 giá vốn, B410 LN gộp, B500/B810/B822/B833 chi phí, B840 LNTT, B900 LNST). Cột giá trị
    = 'Kỳ này' (số THÁNG; cột sau là từng chi nhánh XDV). Map -> 01_HQKD: DThu(1000)=B210/B100;
    **Tổng chi phí(1047)=B300+B500+B810+B822+B833** (GỒM giá vốn B300 — CHỐT 2026-07-17 theo audit đối
    soát BCHN TC: chi phí XDV hợp nhất = giá vốn + CP xưởng + cố định + lãi vay + khác = 37,71 khớp
    BCHN, giống mọi khối khác); LNTT(1112)=B840; giá vốn(B300)/LN gộp(B410)/LNST(B900) -> PNLT.
    (Lịch sử: 15/7 gồm giá vốn → 16/7 chốt LOẠI → 17/7 chốt GỒM lại theo BCHN TC hợp nhất.)
    Trả None nếu KHÔNG phải layout XDV (0 ảnh hưởng cty khác)."""
    from servers import template_filler as tf
    from servers.common import be_bridge as bb
    _norm = lambda v: bb.remove_diacritics("" if v is None else str(v)).strip().lower()  # noqa: E731

    def code(r):
        return str(r[0]).strip().upper() if r and r[0] not in (None, "") else ""
    codes = {code(r) for r in rows}
    if "B100" not in codes or ("B840" not in codes and "B900" not in codes):
        return None
    # cột giá trị = header "Kỳ này" (fallback cột C idx 2)
    val_j = next((j for r in rows[:12] for j, c in enumerate(r) if _norm(c) == "ky nay"), None)
    if val_j is None:
        val_j = 2

    def v(*wanted):
        for r in rows:
            if code(r) in wanted:
                x = r[val_j] if val_j < len(r) else None
                return round(x * 1e-9, 9) if isinstance(x, (int, float)) else None
        return None

    def vsum(*wanted):
        s, got = 0.0, False
        for r in rows:
            if code(r) in wanted:
                x = r[val_j] if val_j < len(r) else None
                if isinstance(x, (int, float)):
                    s += x
                    got = True
        return round(s * 1e-9, 9) if got else None
    dt, lntt = (v("B210") or v("B100")), v("B840")
    if dt is None or lntt is None:
        return {"ok": False, "error": "XDV: thiếu B100/B210 hoặc B840"}
    records = []

    def add(ten, val):
        if val is not None:
            records.append({"Kỳ (yyyy-mm)": period, "Chỉ tiêu KQKD": ten, "Thực hiện (tỷ)": val})
    # BREAKDOWN THEO CHI NHÁNH (chốt 2026-07-23, 6 chỉ tiêu): 14 cột chi nhánh (I→V) đứng NGAY SAU
    # cột "Kỳ này" — verify Σ 14 chi nhánh = ĐÚNG cột tổng mọi mã cần dùng (B100/B300/B410/B500/
    # B810/B822/B833/B840/B900), khớp tuyệt đối CẢ 6 THÁNG (khác DUAN — XDV không có lỗi nguồn per-
    # chi-nhánh). Cột ổn định vị trí qua các tháng nhưng vẫn dò theo TÊN cho chắc (nhất quán cách
    # làm DUAN/SRVF). Mã cost center khớp 100% trust_me_bro.xlsx (khối "Khối KD Vinfast - XDV"),
    # tất cả đều pháp nhân TC (không có case chéo pháp nhân như SRVF UB_SR).
    _XDV_BRANCH_CC = [
        ("ocean park", "OCP_XDV"), ("long bien", "LB_XDV"), ("smart city", "SMC_XDV"),
        ("ha long", "HL_XDV"), ("cam pha", "CP_XDV"), ("xuan mai", "XM_XDV"),
        ("uong bi", "UB_XDV"), ("tuyen quang", "TQ_XDV"), ("vinh phuc", "VP_XDV"),
        ("son tay", "ST_XDV"), ("dai tu", "ĐT_XDV"), ("viet tri", "VT_XDV"),
        ("ha khanh", "HK_XDV"), ("ho chi minh", "HCM_XDV"),
    ]

    def _find_xdv_col(kw):
        return next((j for r in rows[:12] for j, c in enumerate(r)
                     if isinstance(c, str) and kw in _norm(c)), None)
    xdv_cols = [(cc, j) for kw, cc in _XDV_BRANCH_CC for j in [_find_xdv_col(kw)] if j is not None]

    def vj(cds, j):
        s, got = 0.0, False
        for r in rows:
            if code(r) in cds:
                x = r[j] if j < len(r) else None
                if isinstance(x, (int, float)):
                    s += x
                    got = True
        return round(s * 1e-9, 9) if got else None

    def addcc(cc, ten, val):
        if val is not None:
            records.append({"Kỳ (yyyy-mm)": period, "Chỉ tiêu KQKD": ten,
                            "Thực hiện (tỷ)": val, _CC_COL: cc})
    if len(xdv_cols) >= 10:   # đủ chi nhánh mới tách theo CC; thiếu (đổi layout) -> fallback TỔNG như cũ
        for cc, j in xdv_cols:
            dt_j = vj(("B210",), j) or vj(("B100",), j)
            addcc(cc, "Doanh thu thuần", dt_j)                                       # -> 1000 + DTHU
            addcc(cc, "Tổng chi phí", vj(("B300", "B500", "B810", "B822", "B833"), j))  # -> 1047 (GỒM giá vốn)
            addcc(cc, "Lợi nhuận trước thuế", vj(("B840",), j))                       # -> 1112
            addcc(cc, "Giá vốn hàng bán", vj(("B300",), j))                          # -> PNLT
            addcc(cc, "Lợi nhuận gộp", vj(("B410",), j))                             # -> PNLT
            addcc(cc, "Lợi nhuận sau thuế", vj(("B900",), j))                        # -> PNLT
    else:
        add("Doanh thu thuần", dt)                              # -> 1000 + DTHU
        # Tổng chi phí(1047) = B300 (giá vốn) + B500 (CP xưởng) + B810 (cố định) + B822 (lãi vay) + B833
        # (khác) — GỒM giá vốn B300 theo CHỐT 2026-07-17 (audit đối soát BCHN TC: 37,71 khớp hợp nhất).
        add("Tổng chi phí", vsum("B300", "B500", "B810", "B822", "B833"))   # -> 1047 (GỒM giá vốn)
        add("Lợi nhuận trước thuế", lntt)                       # -> 1112 (B840)
        add("Giá vốn hàng bán", v("B300"))                      # -> PNLT
        add("Lợi nhuận gộp", v("B410"))                         # -> PNLT
        add("Lợi nhuận sau thuế", v("B900"))                    # -> PNLT (nuôi thẻ LNST)
    add("Doanh thu HH, DV", v("B100"))                     # -> PNLT (#1 bảng 50: B100 'DOANH THU XDV', trước giảm trừ B200; giữ TỔNG, chưa tách CC)
    add("Doanh thu tài chính", v("B821"))                  # -> PNLT (DT tài chính XDV = B821)
    add("Thu nhập khác", vsum("B831", "B832"))             # -> PNLT (thu nhập HĐ khác = DT chiến dịch B831 + thu nhập khác B832; user chốt 2026-07-21)
    out = os.path.join(tf.FILLED_DIR, f"KQKD_{period}_{cong_ty or 'NA'}_01_HQKD.xlsx")
    tf.fill("01_HQKD", records, out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    # ---- 02_CHIPHI: các nhóm CẤU THÀNH 1047 (=B300 + B500 + B810 + B822 + B833, GỒM giá vốn) -> byNhom
    # == byKhoi (1047 gồm giá vốn, chốt 2026-07-17). ----
    name_j = next((j for r in rows[:12] for j, c in enumerate(r) if _norm(c).startswith("chi tieu")), 1)
    _cpr = _fill_import_chiphi(_chiphi_recs_xdv(rows, code, name_j, val_j, period), period, cong_ty, file_path)
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"),
            "target": "01_HQKD", "value_col_header": "Kỳ này", "chiphi": _cpr}


def _chiphi_recs_xdv(rows, code, name_j, val_j, period):
    """Dòng 02_CHIPHI cho XDV: các nhóm cấu thành 1047 (=B300 + B500 + B810 + B822 + B833, GỒM giá vốn —
    khớp định nghĩa 1047 của deriver để byNhom==byKhoi). Tách B500 -> B600 (nhân sự) + B700 (hoạt
    động xưởng) cho chi tiết hơn NẾU B600+B700==B500, không thì để nguyên B500. dim1 (Nhóm) dùng
    NHÃN TỰ NHIÊN (tên XDV generic 'nhân sự/cố định' không hợp keyword _nhom_cp -> sẽ rơi hết 'khác');
    dim3 = tên gốc. Bỏ dòng 0/None. Sum khớp 1047 (đã verify T01-T05)."""
    def val(cd):
        r = next((x for x in rows if code(x) == cd), None)
        v = r[val_j] if r is not None and val_j < len(r) else None
        return v if isinstance(v, (int, float)) else None

    def nm(cd, default):
        r = next((x for x in rows if code(x) == cd), None)
        s = str(r[name_j]).strip() if r is not None and name_j < len(r) and r[name_j] not in (None, "") else ""
        return s or default
    b500, b600, b700 = val("B500"), val("B600"), val("B700")
    groups = [("B300", "Giá vốn hàng bán")]   # GỒM giá vốn (B300) vào tổng chi phí — chốt 2026-07-17 (khớp 1047 mới, BCHN TC)
    if b500 is not None and b600 is not None and b700 is not None and abs((b600 + b700) - b500) < 1000:
        groups += [("B600", "Chi phí nhân sự"), ("B700", "Chi phí hoạt động xưởng")]
    else:
        groups += [("B500", "Chi phí xưởng dịch vụ")]
    groups += [("B810", "Chi phí cố định"), ("B822", "Chi phí tài chính"), ("B833", "Chi phí khác")]
    recs = []
    for cd, nhom in groups:
        x = val(cd)
        if x is None or x == 0:
            continue
        recs.append({"Kỳ (yyyy-mm)": period, "Nhóm CP (chuẩn mực KT)": nhom,
                     "Khoản mục chi tiết": nm(cd, nhom), "Thực hiện (tỷ)": round(x * 1e-9, 9)})
    return recs


def _derive_kqkd_srvf(rows, period, cong_ty, file_path):
    """TẤT ĐỊNH cho P&L QUẢN TRỊ SRVF (Showroom Vinfast, sheet 'KQKD' — mã A-series ở cột 'Mã số'
    idx thường=2, KHÔNG phải TT200). Anchor: A100 TỔNG DOANH THU SHOWROOM, A300 TỔNG CHI PHÍ SHOW
    ROOM, A600 LỢI NHUẬN SHOW ROOM (=U302 LNST); A310 giá vốn, A500 phân bổ HO. Cột giá trị = 'Kỳ
    này' (số THÁNG; cột sau là kỳ trước/lũy kế). Map -> 01_HQKD: DThu(1000)=A100; Tổng chi phí(1047)=
    A300; LNTT(1112)=A600; +giá vốn A310/LNST A600 -> PNLT. Trả None nếu KHÔNG phải layout SRVF
    (0 ảnh hưởng cty khác — A100/A300/A600 là mã riêng showroom). LÝ DO: extractor TT200 không đọc
    được mã A-series -> trước đây chỉ T01 được LLM bóc, T02-T05 rỗng."""
    from servers import template_filler as tf
    from servers.common import be_bridge as bb
    _norm = lambda v: bb.remove_diacritics("" if v is None else str(v)).strip().lower()  # noqa: E731
    ma_j = next((j for r in rows[:8] for j, c in enumerate(r) if _norm(c) == "ma so"), None)
    # CỘT GIÁ TRỊ = THÁNG BÁO CÁO 'T{mm}' (số THÁNG), KHÔNG dùng 'Kỳ này' (= LŨY KẾ quý/năm =
    # tổng các tháng -> lạm phát). Header có 'T01'/'T02'/... theo từng tháng + cột 'Kỳ này' lũy kế.
    # Kỳ đầu năm (T01) lũy kế == tháng nên fallback 'Kỳ này' vẫn đúng cho T01.
    mm = period.split("-")[1] if period and "-" in period else ""
    _cands = (f"t{mm}", f"t{int(mm)}") if mm.isdigit() else ()
    val_j = next((j for r in rows[:8] for j, c in enumerate(r) if _norm(c) in _cands), None)
    if val_j is None:
        val_j = next((j for r in rows[:8] for j, c in enumerate(r) if _norm(c) == "ky nay"), None)
    if ma_j is None or val_j is None:
        return None

    def code(r):
        return str(r[ma_j]).strip().upper() if ma_j < len(r) and r[ma_j] not in (None, "") else ""
    codes = {code(r) for r in rows}
    if not ({"A100", "A300", "A600"} <= codes):
        return None

    def v(*wanted):
        for r in rows:
            if code(r) in wanted:
                x = r[val_j] if val_j < len(r) else None
                return round(x * 1e-9, 9) if isinstance(x, (int, float)) else None
        return None
    dt, cp, lntt = v("A100"), v("A300"), v("A600")
    if dt is None or lntt is None:
        return {"ok": False, "error": "SRVF: thiếu A100/A600 (giá trị)"}
    records = []

    def add(ten, val):
        if val is not None:
            records.append({"Kỳ (yyyy-mm)": period, "Chỉ tiêu KQKD": ten, "Thực hiện (tỷ)": val})
    # BREAKDOWN THEO SHOWROOM (chốt 2026-07-23, 4 chỉ tiêu): sheet có thêm 1 cột/showroom NGAY SAU
    # cột tổng T{mm} (Uông Bí/Vinfast B2B/OceanPark/Long Biên/Smart City/Hạ Long/Cẩm Phả/Vĩnh Phúc/
    # Sơn Tây/Xuân Mai — verify Σ 10 cột = ĐÚNG cột tổng T01, vd A100 490.180.598.030). Cột "CHI NHÁNH
    # VINFAST HÀ NỘI" luôn =0 mọi tháng (nhánh chưa hoạt động) -> bỏ. VỊ TRÍ CỘT DỊCH theo tháng (mỗi
    # tháng sheet cộng thêm 1 cột T0x tổng phía trước) -> PHẢI dò theo TÊN header (không hardcode index),
    # giống cách val_j đã làm. Mã cost center theo trust_me_bro.xlsx (khối Showroom); "Vinfast B2B" là
    # MẢNG kinh doanh (không phải showroom vật lý) -> mã tự đặt B2B_SR (không có trong MD_COSTCENTER,
    # import_filled backfill cong_ty=TC qua fallback — giống pattern HO_XVP). Lưu ý: UB_SR trong master
    # gắn công ty VFQN (khác TC) -> dòng Uông Bí sẽ tự resolve cong_ty=VFQN qua cost_center (ĐÚNG theo
    # SRVF là thư mục ĐA-pháp-nhân, xem comment resolve_company trong template_filler.import_filled).
    _SR_SHOWROOM_CC = [
        ("uong bi", "UB_SR"), ("b2b", "B2B_SR"), ("oceanpark", "OCP_SR"),
        ("long bien", "LB_SR"), ("smart city", "SMC_SR"), ("ha long", "HL_SR"),
        ("cam pha", "CP_SR"), ("vinh phuc", "VP_SR"), ("son tay", "ST_SR"),
        ("xuan mai", "XM_SR"),
    ]

    def _find_sr_col(kw):
        return next((j for r in rows[:8] for j, c in enumerate(r)
                     if isinstance(c, str) and kw in _norm(c)), None)
    sr_cols = [(cc, j) for kw, cc in _SR_SHOWROOM_CC for j in [_find_sr_col(kw)] if j is not None]

    def vj(wanted, j):
        for r in rows:
            if code(r) == wanted:
                x = r[j] if j < len(r) else None
                return round(x * 1e-9, 9) if isinstance(x, (int, float)) else None
        return None

    def addcc(cc, ten, val):
        if val is not None:
            records.append({"Kỳ (yyyy-mm)": period, "Chỉ tiêu KQKD": ten,
                            "Thực hiện (tỷ)": val, _CC_COL: cc})
    if len(sr_cols) >= 8:   # đủ showroom mới tách theo CC; thiếu (đổi layout) -> fallback TỔNG như cũ
        for cc, j in sr_cols:
            addcc(cc, "Doanh thu thuần", vj("A100", j))        # -> 1000 + DTHU
            addcc(cc, "Tổng chi phí", vj("A300", j))           # -> 1047
            addcc(cc, "Lợi nhuận trước thuế", vj("A600", j))   # -> 1112
            addcc(cc, "Giá vốn hàng bán", vj("A310", j))       # -> PNLT
    else:
        add("Doanh thu thuần", dt)                     # -> 1000 + DTHU
        add("Tổng chi phí", cp)                        # -> 1047 (A300 TỔNG CHI PHÍ SHOW ROOM)
        add("Lợi nhuận trước thuế", lntt)              # -> 1112 (A600 = U302 LNST; P&L quản trị ko tách TNDN)
        add("Giá vốn hàng bán", v("A310"))             # -> PNLT
    add("Lợi nhuận sau thuế", v("U302") or lntt)   # -> PNLT (nuôi thẻ LNST; giữ TỔNG, chưa tách CC)
    add("Doanh thu HH, DV", v("A100"))             # -> PNLT (chỉ tiêu #1 bảng 50 = A100 'TỔNG DOANH THU SHOWROOM'
    #   theo Mapping QTTC: T05BC cột M dòng 2. KHÔNG dùng A200 'bán xe XHĐ' (loại DT khác+Claim -> under-count).
    #   Giữ TỔNG (chưa tách CC) — user chốt 2026-07-23 chỉ tách 4 chỉ tiêu trên trước.
    out = os.path.join(tf.FILLED_DIR, f"KQKD_{period}_{cong_ty or 'NA'}_01_HQKD.xlsx")
    tf.fill("01_HQKD", records, out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    _vh = next((str(r[val_j]) for r in rows[:8] if val_j < len(r) and r[val_j] not in (None, "")), f"col{val_j}")
    # ---- 02_CHIPHI (cơ cấu chi phí theo NHÓM, TẤT ĐỊNH) ------------------------------------------
    # Thành phần TRỰC TIẾP của A300 TỔNG CHI PHÍ: lấy mã từ CÔNG THỨC dòng A300 (vd
    # 'A310+A320+A325+A330+A340+A350+A360+A500') -> khớp tổng chính xác (KHÔNG cộng trùng cấp con,
    # KHÔNG sót A325 lương CTV / A500 phân bổ HO như extract_chiphi a_pat=^A3[1-9]0$). Cùng cột giá
    # trị val_j với P&L (tháng, KHÔNG lũy kế). report_type CHIPHI riêng -> KHÔNG đè 01_HQKD.
    _cp = _chiphi_recs_srvf(rows, code, name_j=next(
        (j for r in rows[:8] for j, c in enumerate(r) if _norm(c).startswith("chi tieu")), ma_j + 1),
        val_j=val_j, period=period)
    _cpr = _fill_import_chiphi(_cp, period, cong_ty, file_path)
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"),
            "target": "01_HQKD", "value_col_header": _vh, "chiphi": _cpr}


def _chiphi_recs_srvf(rows, code, name_j, val_j, period):
    """Dòng 02_CHIPHI cho SRVF: các mã thành phần TRỰC TIẾP của A300 (parse công thức A300, fallback
    list). code(r)=mã ở cột 'Mã số'; giá trị lấy tại val_j (cột tháng như P&L). Nhóm CP theo
    _nhom_cp (dùng chung extract_chiphi). Bỏ dòng thiếu/0."""
    import re as _re
    from extract_chiphi import _nhom_cp
    _a300 = next((r for r in rows if code(r) == "A300"), None)
    _f = next((str(c) for c in (_a300 or []) if isinstance(c, str)
               and _re.search(r"A\d{3}\s*\+\s*A\d{3}", str(c))), "")
    comp = _re.findall(r"A\d{3}", _f) or ["A310", "A320", "A325", "A330", "A340", "A350", "A360", "A500"]
    recs = []
    for cd in comp:
        r = next((x for x in rows if code(x) == cd), None)
        if r is None:
            continue
        val = r[val_j] if val_j < len(r) else None
        if not isinstance(val, (int, float)) or val == 0:
            continue
        nm = str(r[name_j]).strip() if name_j < len(r) and r[name_j] not in (None, "") else cd
        recs.append({"Kỳ (yyyy-mm)": period, "Nhóm CP (chuẩn mực KT)": _nhom_cp(nm, cd),
                     "Khoản mục chi tiết": nm, "Thực hiện (tỷ)": round(val * 1e-9, 9)})
    return recs


def _derive_kqkd_ho(rows, period, cong_ty, file_path):
    """HO (Khối hỗ trợ tập đoàn) — sheet 'HO_KQKD' P&L QUẢN TRỊ: cột 1 'Chỉ tiêu' có dòng TỔNG
    ('Tổng Doanh thu'/'Tổng chi phí'/'Tổng lợi nhuận'); cột giá trị theo THÁNG 'T01'/'T02'/… + cột
    'TỔNG 12T' (lũy kế). KHÔNG phải TT200 -> trước rơi LLM, LLM lấy nhầm cột T01 cho MỌI tháng
    (T02-T05 = số T1). Deriver lấy đúng cột 'T{mm}'. Chi phí HO=0 ĐÚNG (phân bổ 100% sang khối vận
    hành). Gate: có dòng 'tong doanh thu'+'tong loi nhuan' (cột 1) -> HO_KQKD-specific, None cty khác."""
    from servers import template_filler as tf
    from servers.common import be_bridge as bb
    _norm = lambda v: bb.remove_diacritics("" if v is None else str(v)).strip().lower()  # noqa: E731
    names = {_norm(r[1]) for r in rows if len(r) > 1 and r[1] not in (None, "")}
    if not ({"tong doanh thu", "tong loi nhuan"} <= names):
        return None
    mm = period.split("-")[1] if period and "-" in period else ""
    cands = (f"t{mm}", f"t{int(mm)}") if mm.isdigit() else ()
    val_j = next((j for r in rows[:8] for j, c in enumerate(r) if _norm(c) in cands), None)
    if val_j is None:
        return {"ok": False, "error": f"HO_KQKD: không thấy cột T{mm}"}

    def rowval(kw):
        for r in rows:
            if len(r) > 1 and _norm(r[1]) == kw:
                x = r[val_j] if val_j < len(r) else None
                return round(x * 1e-9, 9) if isinstance(x, (int, float)) else None
        return None

    def codeval(code_a):
        """Giá trị cột T{mm} của dòng có MÃ (cột A, idx 0) == code_a — HO 'Doanh thu HH, DV' nằm ở
        dòng mã '511_TS', không phải dòng TỔNG (đọc theo tên col1)."""
        for r in rows:
            if r and r[0] not in (None, "") and str(r[0]).strip() == code_a:
                x = r[val_j] if val_j < len(r) else None
                return round(x * 1e-9, 9) if isinstance(x, (int, float)) else None
        return None
    dt, cp, ln = rowval("tong doanh thu"), rowval("tong chi phi"), rowval("tong loi nhuan")
    if dt is None and ln is None:
        return {"ok": False, "error": "HO_KQKD: không đọc được Tổng DT/LN"}
    # Doanh thu thuần = CHỈ mã 511_TS (DT bán hàng/thanh lý), KHÔNG dùng "Tổng doanh thu" (dòng TỔNG
    # gồm cả 515.xx DT tài chính + 7111 thu nhập khác -> thừa, chốt user 2026-07-19 đối chiếu file gốc
    # T4/T6: "Tổng doanh thu" T6 = 0,0659 (gồm 515.01=0,00786 + 7111=0,0581) nhưng 511_TS T6 = None/0).
    # 511_TS None (nhiều tháng không phát sinh) -> coalesce 0.0 (khác "bỏ dòng": giữ #1000 = 0 tường minh
    # thay vì thiếu hẳn dòng, tránh khối biến mất khỏi breakdown-theo-khối các tháng đó).
    dt_thuan = codeval("511_TS") or 0.0
    records = []

    def add(t, v):
        if v is not None:
            records.append({"Kỳ (yyyy-mm)": period, "Chỉ tiêu KQKD": t, "Thực hiện (tỷ)": v})
    add("Doanh thu thuần", dt_thuan)                    # -> 1000 + DTHU
    add("Tổng chi phí", cp if cp is not None else 0.0)  # -> 1047 (HO=0: phân bổ hết sang khối)
    add("Lợi nhuận trước thuế", ln)                     # -> 1112
    add("Lợi nhuận sau thuế", ln)                       # -> PNLT (nuôi thẻ LNST; HO không có thuế -> LNST=LNTT)
    add("Doanh thu HH, DV", dt_thuan)                   # -> PNLT (#1 bảng 50: mã 511_TS 'DT thanh lý bán vật tư, TS')
    add("Lợi nhuận gộp", dt_thuan)                      # -> PNLT (#6: LNG = DT thuần − giá vốn; HO "Không có" giá vốn -> = DT thuần, chốt Mapping 2026-07-18)
    out = os.path.join(tf.FILLED_DIR, f"KQKD_{period}_{cong_ty or 'NA'}_01_HQKD.xlsx")
    tf.fill("01_HQKD", records, out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"),
            "target": "01_HQKD", "value_col_header": f"T{mm}"}


def _bcqt_month_col(rows, period, header_rows=12):
    """Dò cột giá trị THÁNG trong sheet quản trị BCQT/BCQT PT: khớp 'Tháng {mm}' (An KS) hoặc
    'T{mm}/26'/'T{mm}' (An Taxi). Trả index cột hoặc None."""
    from servers.common import be_bridge as bb
    _norm = lambda v: bb.remove_diacritics("" if v is None else str(v)).strip().lower()  # noqa: E731
    mm = period.split("-")[1] if period and "-" in period else ""
    if not mm.isdigit():
        return None
    toks = {f"t{mm}", f"t{int(mm)}", f"thang {mm}", f"thang {int(mm)}"}
    for r in rows[:header_rows]:
        for j, c in enumerate(r):
            n = _norm(c)
            if n in toks or any(n.startswith(t + "/") for t in (f"t{mm}", f"t{int(mm)}")):
                return j
    return None


def _derive_kqkd_antaxi(file_path: str, period: str, cong_ty: str):
    """TẤT ĐỊNH — An Taxi: sheet 'BCQT PT.' (P&L quản trị B02-DN, MÃ SỐ ở cột B: 100 DT bán hàng,
    110 giảm trừ, 120 DT thuần, 130 giá vốn, 140 lãi gộp, 150 CP biến đổi, 170 CP cố định, 182 CP
    tài chính, 192 CP khác, 200 LNTT, 220 LNST). Cột giá trị = tháng 'T{mm}/26'. Map theo SPEC 50 CT:
      01_HQKD: DThu thuần(1000)=Mã120 · Tổng chi phí(1047)=130+150+170+182+192 (GỘP, spec #8, KHÔNG
               cấn trừ thu nhập ngoài) · LNTT(1112)=Mã200.
      PNLT   : 'Doanh thu bán hàng và cung cấp dịch vụ'=Mã100 (GỘP = 'Doanh thu HH,DV' #1) · giảm trừ
               =Mã110 · DT thuần=Mã120 · giá vốn=Mã130 · LN gộp=Mã140 · LNST=Mã220.
      02_CHIPHI: giá vốn(130)+biến đổi(150)+cố định(170)+tài chính(182)+khác(192) — Σ == 1047.
    Trả None nếu KHÔNG phải layout BCQT PT (0 ảnh hưởng file khác)."""
    from openpyxl import load_workbook
    from servers import template_filler as tf
    from servers.common import be_bridge as bb
    _norm = lambda v: bb.remove_diacritics("" if v is None else str(v)).strip().lower()  # noqa: E731
    wb = bb.fast_load_workbook(file_path, read_only=True, data_only=True)
    sh = next((s for s in wb.sheetnames if _norm(s).replace(".", "").replace(" ", "") == "bcqtpt"), None)
    if not sh:
        wb.close()
        return None
    rows = [list(r) for r in wb[sh].iter_rows(values_only=True)]
    wb.close()
    ma_j = next((j for r in rows[:12] for j, c in enumerate(r) if _norm(c) == "ma so"), 1)
    val_j = _bcqt_month_col(rows, period)
    if val_j is None:
        return {"ok": False, "error": f"BCQT PT: không thấy cột tháng {period}"}

    def code(r):
        return str(r[ma_j]).strip() if ma_j < len(r) and r[ma_j] not in (None, "") else ""
    codes = {code(r) for r in rows}
    if not ({"100", "120", "200"} <= codes):
        return None

    def v(c):
        for r in rows:
            if code(r) == c:
                x = r[val_j] if val_j < len(r) else None
                return round(x * 1e-9, 9) if isinstance(x, (int, float)) else None
        return None
    dt_gross, giam_tru, dt_net = v("100"), v("110"), v("120")
    gia_von, ln_gop = v("130"), v("140")
    cp_bd, cp_cd, cp_tc, cp_khac = v("150"), v("170"), v("182"), v("192")
    lntt, lnst = v("200"), v("220")
    if dt_net is None or lntt is None:
        return {"ok": False, "error": "BCQT PT: thiếu Mã 120/200"}
    # #8 Tổng chi phí = TỔNG các dòng chi phí (GỘP, spec — chốt user 2026-07-17), KHÔNG cấn trừ
    # thu nhập tài chính/khác. Khác đẳng thức 1000−1112 đúng phần thu nhập ngoài (chủ đích).
    tong_cp = round(sum(x for x in (gia_von, cp_bd, cp_cd, cp_tc, cp_khac) if x is not None), 9)
    records = []

    def add(ten, val):
        if val is not None:
            records.append({"Kỳ (yyyy-mm)": period, "Chỉ tiêu KQKD": ten, "Thực hiện (tỷ)": val})
    add("Doanh thu thuần", dt_net)                                    # -> 1000 + DTHU (spec #1 net)
    add("Tổng chi phí", tong_cp)                                      # -> 1047 (GỘP, spec #8)
    add("Lợi nhuận trước thuế", lntt)                                 # -> 1112 (Mã200)
    add("Doanh thu bán hàng và cung cấp dịch vụ", dt_gross)           # -> PNLT (Mã100, nuôi Cấu trúc DT)
    add("Doanh thu HH, DV", dt_gross)                                 # -> PNLT (#1 bảng 50 = Mã100 GỘP)
    add("Các khoản giảm trừ doanh thu", giam_tru)                     # -> PNLT (Mã110)
    add("Giá vốn hàng bán", gia_von)                                  # -> PNLT (Mã130)
    add("Lợi nhuận gộp", ln_gop)                                      # -> PNLT (Mã140)
    add("Lợi nhuận sau thuế", lnst)                                   # -> PNLT (Mã220, nuôi thẻ LNST)
    add("Doanh thu tài chính", v("181"))                              # -> PNLT (IX.1 = Mã181)
    add("Thu nhập khác", v("191"))                                    # -> PNLT (X.1 = Mã191)
    out = os.path.join(tf.FILLED_DIR, f"KQKD_{period}_{cong_ty or 'NA'}_01_HQKD.xlsx")
    tf.fill("01_HQKD", records, out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    cp_groups = [("Giá vốn hàng bán", gia_von), ("Chi phí biến đổi", cp_bd),
                 ("Chi phí cố định", cp_cd), ("Chi phí tài chính", cp_tc), ("Chi phí khác", cp_khac)]
    cp_recs = [{"Kỳ (yyyy-mm)": period, "Nhóm CP (chuẩn mực KT)": n, "Khoản mục chi tiết": n,
                "Thực hiện (tỷ)": val} for n, val in cp_groups if val]
    cpr = _fill_import_chiphi(cp_recs, period, cong_ty, file_path) if cp_recs else None
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"),
            "target": "01_HQKD", "value_col_header": f"T{period.split('-')[1]}/26", "chiphi": cpr}


def _derive_kqkd_ankhachsan(file_path: str, period: str, cong_ty: str):
    """TẤT ĐỊNH — An KS: sheet 'BCQT' (P&L quản trị theo MỤC ở cột B: I TỔNG DOANH THU [I.1 DOANH THU
    KHÁCH SẠN], II TỔNG CHI PHÍ [II.1 CHI PHÍ GIÁ VỐN, II.2 CHI PHÍ CHUNG, II.3 CHI PHÍ LƯƠNG + CP
    KHÁC], III LỢI NHUẬN (I−II)). Cột giá trị = 'Tháng {mm}'. BCQT gộp thu nhập khác vào Mục I nên
    DThu=Mục I, Tổng chi phí=Mục II, LNTT=Mục III đã cân đối (An KS không có thuế TNDN -> LNST=LNTT).
      01_HQKD: DThu thuần(1000)=Mục I · Tổng chi phí(1047)=Mục II · LNTT(1112)=Mục III.
      PNLT   : 'Doanh thu bán hàng và cung cấp dịch vụ'=Mục I.1 (Doanh thu HH,DV) · giá vốn=Mục II.1 ·
               LN gộp=Mục I−Mục II.1 · LNST=Mục III.
      02_CHIPHI: II.1 giá vốn + II.2 chi phí chung + II.3 chi phí lương. Trả None nếu KHÔNG phải BCQT."""
    from openpyxl import load_workbook
    from servers import template_filler as tf
    from servers.common import be_bridge as bb
    _norm = lambda v: bb.remove_diacritics("" if v is None else str(v)).strip().lower()  # noqa: E731
    wb = bb.fast_load_workbook(file_path, read_only=True, data_only=True)
    sh = next((s for s in wb.sheetnames if _norm(s).strip() == "bcqt"), None)
    if not sh:
        wb.close()
        return None
    rows = [list(r) for r in wb[sh].iter_rows(values_only=True)]
    wb.close()
    val_j = _bcqt_month_col(rows, period)
    if val_j is None:
        return {"ok": False, "error": f"BCQT: không thấy cột Tháng {period}"}
    # Cột NHÃN = 'Nội dung chi phí'; MỘT SỐ file có cột A trống thừa -> nhãn dời B->C. Dò ĐỘNG.
    name_j = next((j for r in rows[:8] for j, c in enumerate(r) if _norm(c).startswith("noi dung")), None)
    if name_j is None:
        name_j = next((j for r in rows for j, c in enumerate(r) if _norm(c) == "tong doanh thu"), 1)

    def rowval(*starts):
        for r in rows:
            lab = _norm(r[name_j]) if len(r) > name_j else ""
            if lab and any(lab.startswith(s) for s in starts):
                x = r[val_j] if val_j < len(r) else None
                return round(x * 1e-9, 9) if isinstance(x, (int, float)) else None
        return None
    dt = rowval("tong doanh thu")
    dt_ks = rowval("doanh thu khach san")            # I.1 = Doanh thu HH,DV
    cp = rowval("tong chi phi")
    gia_von = rowval("chi phi gia von")              # II.1
    cp_chung = rowval("chi phi chung")               # II.2
    cp_luong = rowval("chi phi luong")               # II.3
    ln = rowval("loi nhuan")                         # III (I−II)
    if ln is None and dt is not None and cp is not None:
        ln = round(dt - cp, 9)
    if dt is None or cp is None:
        return {"ok": False, "error": "BCQT: thiếu Mục I / Mục II"}
    records = []

    def add(ten, val):
        if val is not None:
            records.append({"Kỳ (yyyy-mm)": period, "Chỉ tiêu KQKD": ten, "Thực hiện (tỷ)": val})
    add("Doanh thu thuần", dt)                                        # -> 1000 + DTHU
    add("Tổng chi phí", cp)                                           # -> 1047 (Mục II)
    add("Lợi nhuận trước thuế", ln)                                   # -> 1112 (Mục III)
    add("Doanh thu bán hàng và cung cấp dịch vụ", dt_ks if dt_ks is not None else dt)  # -> PNLT (nuôi Cấu trúc DT)
    add("Doanh thu HH, DV", dt_ks if dt_ks is not None else dt)       # -> PNLT (#1 bảng 50 = Mục I.1 DT Khách sạn)
    add("Giá vốn hàng bán", gia_von)                                  # -> PNLT (II.1)
    add("Lợi nhuận gộp", round(dt - gia_von, 9) if gia_von is not None else None)      # -> PNLT
    add("Lợi nhuận sau thuế", ln)                                     # -> PNLT (An KS ko thuế -> =LNTT)
    out = os.path.join(tf.FILLED_DIR, f"KQKD_{period}_{cong_ty or 'NA'}_01_HQKD.xlsx")
    tf.fill("01_HQKD", records, out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    cp_groups = [("Giá vốn hàng bán", gia_von), ("Chi phí chung", cp_chung),
                 ("Chi phí lương + CP khác cho CNV", cp_luong)]
    cp_recs = [{"Kỳ (yyyy-mm)": period, "Nhóm CP (chuẩn mực KT)": n, "Khoản mục chi tiết": n,
                "Thực hiện (tỷ)": val} for n, val in cp_groups if val]
    cpr = _fill_import_chiphi(cp_recs, period, cong_ty, file_path) if cp_recs else None
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"),
            "target": "01_HQKD", "value_col_header": f"Tháng {period.split('-')[1]}", "chiphi": cpr}


# MÃ SỐ LCTT (BCTC TT200) phân loại Thu/Chi cho HT theo spec "50 chỉ tiêu quản trị" dòng 18-19
# (đã kiểm chứng khớp kế toán tới ĐỒNG trên file B5.HT T01, 2026-07-18). CỐ Ý LOẠI các mã TỔNG
# (20 LC thuần HĐKD / 30 đầu tư / 40 tài chính / 50 LC thuần trong kỳ) và SỐ DƯ (60/61/70) — chỉ
# lấy dòng phát sinh gốc, không đếm đôi. Thu gồm cả các dòng thu HĐ đầu tư/tài chính (22/24/26/27/
# 31/33); Chi gồm chi đầu tư/tài chính (21/23/25/32/34/35) + trả nợ gốc vay (34).
_LCTT_THU_CODES = {1, 5, 6, 22, 24, 26, 27, 31, 33}
_LCTT_CHI_CODES = {2, 3, 4, 7, 21, 23, 25, 32, 34, 35}


def _derive_lctt_ht(file_path: str, sheet: str, period: str, cong_ty: str):
    """HT (Xe tải Hưng Thịnh) — DÒNG TIỀN (03_DONGTIEN) LẤY TỪ sheet 'LCTT' của BCTC theo spec 50
    chỉ tiêu (dòng 18-19), NGUỒN CHUẨN của HT (KHÔNG lấy từ Báo cáo tiền tập đoàn — extract_tien đã
    bỏ qua HT để không nạp đôi + khác nguồn). Đã kiểm chứng khớp kế toán tới đồng (2026-07-18):
      Thu trong kỳ = Σ giá trị cột 'Tháng {mm}' của các MÃ SỐ {1,5,6,22,24,26,27,31,33}
      Chi trong kỳ = Σ |giá trị| các MÃ SỐ {2,3,4,7,21,23,25,32,34,35}
    Cột giá trị = cột 'Tháng {mm}' (mm suy từ period; KHÔNG lấy 'Lũy kế 6 tháng'). Mỗi mã -> 1 dòng
    03_DONGTIEN (Loại Thu/Chi -> dim1 'A.'/'B.' để metrics.build_cashflow gom vào/ra). Chi lưu TRỊ
    TUYỆT ĐỐI (outflow = SUM dim1 'B%' phải dương). cong_ty=HT. Nhờ import_filled delete-scope theo
    (source_file, report_type=THUCHI), lượt này cũng THAY SẠCH mọi dòng lctt lỡ lọt trước đó (T03-05)."""
    import re as _re
    from openpyxl import load_workbook
    from servers import template_filler as tf
    from servers.common import be_bridge as bb
    _norm = lambda v: bb.remove_diacritics("" if v is None else str(v)).strip().lower()  # noqa: E731

    try:
        mm = int(str(period)[-2:])
    except Exception:
        return {"ok": False, "error": "LCTT: period không hợp lệ"}
    wb = bb.fast_load_workbook(file_path, read_only=True, data_only=True)
    try:
        rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()
    # Header = dòng có 'Mã số'. Cột giá trị = 'Tháng {mm}' (khớp CHÍNH XÁC — KHÔNG lấy cột 'Lũy kế').
    hi = next((i for i, r in enumerate(rows[:15]) if any("ma so" in _norm(c) for c in r)), None)
    if hi is None:
        return {"ok": False, "error": "LCTT: không thấy header 'Mã số'"}
    hdr = rows[hi]
    ma_j = next(j for j, c in enumerate(hdr) if "ma so" in _norm(c))
    name_j = next((j for j, c in enumerate(hdr) if _norm(c).startswith("chi tieu")), 0)
    val_j = next((j for j, c in enumerate(hdr) if _norm(c) == f"thang {mm}"), None)
    if val_j is None:
        return {"ok": False, "error": f"LCTT: không thấy cột 'Tháng {mm}'"}
    _COL_KY, _COL_CTY, _COL_LOAI = "Kỳ / Ngày", "Mã Công ty (auto từ CC)", "Loại (Thu/Chi)"
    _COL_KM = "Khoản mục (Thu bán hàng, Thu đầu tư, Chi NCC, Chi tài chính, Chi đầu tư TS…)"
    _COL_TH = "Thực hiện (tỷ)"
    recs = []
    for r in rows[hi + 1:]:
        raw = r[ma_j] if ma_j < len(r) else None
        try:
            code = int(str(raw).strip())
        except (TypeError, ValueError):
            continue
        loai = "Thu" if code in _LCTT_THU_CODES else ("Chi" if code in _LCTT_CHI_CODES else None)
        if loai is None:                       # mã TỔNG (20/30/40/50) / SỐ DƯ (60/61/70) -> bỏ
            continue
        v = r[val_j] if val_j < len(r) else None
        if not isinstance(v, (int, float)) or v == 0:
            continue
        nm = str(r[name_j]).strip() if name_j < len(r) and r[name_j] not in (None, "") else str(code)
        nm = _re.sub(r"^\d+\.?\s*", "", nm)    # bỏ tiền tố '2. ' cho gọn nhãn
        recs.append({_COL_KY: period, _COL_CTY: "HT", _COL_LOAI: loai,
                     _COL_KM: nm, _COL_TH: round(abs(v) * 1e-9, 9)})   # |v|: THU & CHI đều lưu DƯƠNG
    if not recs:
        return {"ok": False, "error": "LCTT: không bóc được dòng thu/chi nào (mã 1-35)"}
    out = os.path.join(tf.FILLED_DIR, f"LCTT_{period}_HT_03_DONGTIEN.xlsx")
    tf.fill("03_DONGTIEN", recs, out)
    imp = tf.import_filled(out, cong_ty=None, source_file=_source_id(file_path))
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"),
            "target": "03_DONGTIEN", "value_col_header": f"Tháng {mm}"}


# Cột "Mã Cost center" của template 01_HQKD — importer resolve cong_ty/khoi/cost_center từ mã CC này.
_CC_COL = "Mã Cost center ◀ NHẬP"


def _xvp_kqkd_totals(file_path: str):
    """Tổng CÔNG TY (tỷ) từ sheet 'KQKD' luật định TT200 của XVP — NGUỒN CHUẨN theo Mapping sheet 9
    (khớp máy tính kế toán + BCHN: T01 35,678 / T02 40,639). HQKD quản trị (tách depot) lệch ~0.2%
    do phân loại chi phí khác (CP HO vs CP bán hàng/QLDN) -> dùng KQKD làm tổng chuẩn, HQKD chỉ để
    chia tỷ lệ depot. Khớp dòng theo NHÃN cột 'Chỉ tiêu', giá trị cột 'Kỳ này'. Trả dict {tên chỉ
    tiêu (khớp add() trong _derive_kqkd_xvp): tổng tỷ}; None nếu không thấy sheet/header."""
    from openpyxl import load_workbook
    from servers.common import be_bridge as bb
    _n = lambda v: bb.remove_diacritics("" if v is None else str(v)).strip().lower()  # noqa: E731
    wb = bb.fast_load_workbook(file_path, read_only=True, data_only=True)
    sh = next((s for s in wb.sheetnames if _n(s) == "kqkd"), None)
    if not sh:
        wb.close()
        return None
    rows = [list(r) for r in wb[sh].iter_rows(values_only=True)]
    wb.close()
    hi = next((i for i, r in enumerate(rows[:12])
               if any("chi tieu" in _n(c) for c in r) and any("ky nay" in _n(c) for c in r)), None)
    if hi is None:
        return None
    hdr = rows[hi]
    kn = next(j for j, c in enumerate(hdr) if "ky nay" in _n(c))
    ct = next(j for j, c in enumerate(hdr) if "chi tieu" in _n(c))

    def find(pred):
        for r in rows[hi + 1:]:
            lab = _n(r[ct]) if ct < len(r) and r[ct] not in (None, "") else ""
            if lab and pred(lab):
                v = r[kn] if kn < len(r) else None
                return round(v * 1e-9, 9) if isinstance(v, (int, float)) else None
        return None
    gv = find(lambda l: "gia von hang ban" in l)
    tc = find(lambda l: "chi phi tai chinh" in l)
    bh = find(lambda l: "chi phi ban hang" in l)
    ql = find(lambda l: "chi phi quan ly" in l)
    ck = find(lambda l: "chi phi khac" in l)               # Mã 32 'Chi phí khác' — Mapping #8 có gồm
    cost = round(sum(x for x in (gv, tc, bh, ql, ck) if x is not None), 9) if any(
        x is not None for x in (gv, tc, bh, ql, ck)) else None
    return {
        "Doanh thu thuần": find(lambda l: "doanh thu thuan" in l),
        "Tổng chi phí": cost,                                             # giá vốn + CP tài chính + CP bán hàng + CP QLDN + CP khác(Mã32)
        "Lợi nhuận trước thuế": find(lambda l: "loi nhuan" in l and "truoc thue" in l),
        "Giá vốn hàng bán": gv,
        "Lợi nhuận gộp": find(lambda l: "loi nhuan gop" in l),
        "Lợi nhuận sau thuế": find(lambda l: "loi nhuan sau thue" in l),
        "Doanh thu HH, DV": find(lambda l: "doanh thu ban hang va cung cap" in l and "thuan" not in l),
        # 4 NHÓM cấu thành Tổng chi phí (Σ == 1047) -> 02_CHIPHI cơ cấu, khớp HQKD 1047. Key có '__'
        # để KHÔNG lọt vào vòng scale theo tên chỉ tiêu ở _derive_kqkd_xvp.
        "__chiphi_nhom__": {"Giá vốn hàng bán": gv, "Chi phí tài chính": tc,
                            "Chi phí bán hàng": bh, "Chi phí QLDN": ql, "Chi phí khác": ck},
    }


def _derive_kqkd_xvp(file_path: str, period: str, cong_ty: str):
    """TẤT ĐỊNH — CÔNG TY XANH VĨNH PHÚC (XVP, pháp nhân 'CP Công nghệ & DV Xanh Vĩnh Phúc', KHÁC
    HTX_XVP): sheet 'HQKD' P&L quản trị TÁCH THEO DEPOT (cost center). Header cột: TỔNG CỘNG(C) ·
    HO Xanh(E) · Depot Phú Thọ(G) · Vĩnh Phúc(I) · Tuyên Quang(K). Mã (cột B): 1 'DT bán hàng'
    (=Doanh thu HH,DV) · 10 DT thuần · 11 giá vốn · 20 LN gộp; LNTT = dòng '…(EBT)' (sau phân bổ HO) ·
    LNST = dòng '…(EAT)'. HQKD chỉ dùng để CHIA TỶ LỆ theo depot; TỔNG mỗi chỉ tiêu được SCALE về số
    sheet 'KQKD' luật định (_xvp_kqkd_totals) — nguồn chuẩn theo Mapping sheet 9, khớp kế toán + BCHN
    (T01 Tổng CP 35,678 / T02 40,639). HQKD quản trị lệch ~0.2% do phân loại chi phí khác nhau.

    Emit P&L THEO 3 DEPOT (cost_center = PT_DP/VP_DP/TQ_DP; importer tự resolve cong_ty XVP + khối 6
    từ MD_COSTCENTER) — KHÔNG emit dòng tổng công ty riêng: tổng công ty = Σ 3 depot (metrics cộng khi
    KHÔNG lọc CC), tránh ĐẾM ĐÔI (flow_sum cộng hết mọi dòng khi không có bộ lọc CC). HO Xanh ≈ 0 nên
    Σ depot = TỔNG CỘNG. Đây cũng THAY chỗ nhánh generic TT200 (sheet 'KQKD') cho XVP -> 1 nguồn P&L.
    Trả None nếu KHÔNG phải layout HQKD-theo-depot (HTX_XVP/HTX_XTQ HQKD 1 cột -> generic mã 1 như cũ)."""
    from openpyxl import load_workbook
    from servers import template_filler as tf
    from servers.common import be_bridge as bb
    _norm = lambda v: bb.remove_diacritics("" if v is None else str(v)).strip().lower()  # noqa: E731
    wb = bb.fast_load_workbook(file_path, read_only=True, data_only=True)
    sh = next((s for s in wb.sheetnames if _norm(s) == "hqkd"), None)
    if not sh:
        wb.close()
        return None
    rows = [list(r) for r in wb[sh].iter_rows(values_only=True)]
    wb.close()
    hdr_i = next((i for i, r in enumerate(rows[:8]) if any(_norm(c) == "tong cong" for c in r)), None)
    if hdr_i is None:
        return None
    hdr = rows[hdr_i]

    def _find_col(*kws):
        return next((j for j, c in enumerate(hdr) if any(k in _norm(c) for k in kws)), None)
    # depot -> cột (khớp theo TÊN header, không hardcode G/I/K). Đủ 3 depot mới nhận (đúng file XVP cty).
    depots = [("PT_DP", _find_col("phu tho")), ("VP_DP", _find_col("vinh phuc")),
              ("TQ_DP", _find_col("tuyen quang"))]
    depots = [(cc, j) for cc, j in depots if j is not None]
    if len(depots) < 3:
        return None
    # HO Xanh (cột E) = chi phí HO công ty CHƯA phân bổ về depot (DT/giá vốn = 0, chỉ có ở LNTT/LNST/
    # Tổng CP). PHẢI emit như 1 cost center ('HO_XVP') để Σ (3 depot + HO Xanh) = TỔNG CỘNG đúng ĐẾN
    # ĐỒNG cho MỌI chỉ tiêu (nếu bỏ, LNTT/LNST công ty = Σ depot bị thiếu phần overhead HO).
    _ho_j = _find_col("ho xanh")
    if _ho_j is not None:
        depots.append(("HO_XVP", _ho_j))
    ma_j = next((j for j, c in enumerate(hdr) if _norm(c) in ("ma so", "ma")), 1)

    def _row_ma(ma):
        return next((r for r in rows[hdr_i + 1:]
                     if ma_j < len(r) and r[ma_j] not in (None, "") and str(r[ma_j]).strip() == ma), None)

    def _row_lbl(*subs):
        return next((r for r in rows[hdr_i + 1:]
                     if r and r[0] not in (None, "") and any(s in _norm(r[0]) for s in subs)), None)
    r_hhdv, r_net, r_gv, r_lng = _row_ma("1"), _row_ma("10"), _row_ma("11"), _row_ma("20")
    r_lntt = _row_lbl("(ebt)")                 # LNTT sau phân bổ HO (EBT), KHÔNG lấy 'trực tiếp'
    r_lnst = _row_lbl("(eat)")
    r_dttc, r_tnk = _row_lbl("doanh thu tai chinh"), _row_lbl("thu nhap khac")  # thu nhập NGOÀI (cộng lại vào CP GỘP)
    if r_net is None or r_lntt is None:
        return {"ok": False, "error": "XVP HQKD: thiếu mã 10 / dòng LNTT (EBT)"}

    def _v(r, j):
        x = r[j] if (r is not None and j < len(r)) else None
        return round(x * 1e-9, 9) if isinstance(x, (int, float)) else None

    def _v0(r, j):                              # như _v nhưng None -> 0.0 (thu nhập ngoài, cộng bù)
        return _v(r, j) or 0.0
    records = []

    def add(cc, ten, val):
        if val is not None:
            records.append({"Kỳ (yyyy-mm)": period, "Chỉ tiêu KQKD": ten,
                            "Thực hiện (tỷ)": val, _CC_COL: cc})
    for cc, j in depots:
        net, lntt = _v(r_net, j), _v(r_lntt, j)
        add(cc, "Doanh thu thuần", net)                              # -> 1000 + DTHU (theo depot)
        if net is not None and lntt is not None:
            # Chi phí THEO DEPOT (proportion) = (DT thuần − LNTT) + DT tài chính + thu nhập khác (GỘP
            # HQKD). CHỈ dùng để CHIA TỶ LỆ giữa depot — TỔNG công ty sẽ được SCALE về số KQKD luật
            # định ở bước dưới (HQKD quản trị lệch ~0.2% do phân loại CP khác nhau).
            add(cc, "Tổng chi phí", round((net - lntt) + _v0(r_dttc, j) + _v0(r_tnk, j), 9))  # -> 1047
        add(cc, "Lợi nhuận trước thuế", lntt)                        # -> 1112 (EBT sau phân bổ HO)
        add(cc, "Giá vốn hàng bán", _v(r_gv, j))                     # -> PNLT
        add(cc, "Lợi nhuận gộp", _v(r_lng, j))                       # -> PNLT
        add(cc, "Lợi nhuận sau thuế", _v(r_lnst, j) if r_lnst is not None else lntt)  # -> PNLT (thẻ LNST)
        add(cc, "Doanh thu HH, DV", _v(r_hhdv, j))                   # -> PNLT (#1 = mã 1 'DT bán hàng' theo depot)
        add(cc, "Doanh thu tài chính", _v(r_dttc, j))                # -> PNLT (DT tài chính theo depot; IX.1)
        add(cc, "Thu nhập khác", _v(r_tnk, j))                       # -> PNLT (thu nhập khác theo depot; X.1)
    if not records:
        return {"ok": False, "error": "XVP HQKD: không bóc được dòng depot nào"}
    # === TỔNG CÔNG TY = HQKD QUẢN TRỊ (Mapping QTTC sheet 9, CẬP NHẬT 2026-07-20) ===
    # LỊCH SỬ: chốt 2026-07-18 từng SCALE tổng về sheet KQKD luật định. Mapping ĐÃ ĐỔI: "KHÔNG chọn
    # (mức công ty) → lấy chỉ tiêu chung Cty Xanh VP (sheet HQKD)" — mọi chỉ tiêu #1/#4/#6/#8/#13/#15
    # đều ghi "sheet HQKD". BCHN TC cũng đối soát theo HQKD (T01 LNST -3,73 / T02 giá vốn 35,41; KQKD
    # ra -3,81 / 35,44 nên lệch ~82tr/27tr). => GIỮ NGUYÊN giá trị depot bóc từ HQKD (Σ depot = TỔNG
    # CỘNG HQKD công ty), KHÔNG scale về KQKD nữa. _kq chỉ còn dùng cho cơ cấu 4 nhóm CP (scale về CP HQKD).
    _kq = _xvp_kqkd_totals(file_path)
    out = os.path.join(tf.FILLED_DIR, f"KQKD_{period}_XVP_01_HQKD.xlsx")
    tf.fill("01_HQKD", records, out)
    # cong_ty='XVP': depot (PT/VP/TQ_DP) tự resolve XVP+khối từ MD_COSTCENTER; riêng 'HO_XVP' KHÔNG có
    # trong MD_COSTCENTER -> import_filled back-fill cong_ty=XVP + khoi cho các dòng còn NULL.
    imp = tf.import_filled(out, cong_ty="XVP", khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    # ---- 02_CHIPHI: cơ cấu 4 nhóm CP (giá vốn/tài chính/bán hàng/QLDN/khác). Lấy TỶ LỆ nhóm từ KQKD
    # (sheet KQKD tách nhóm rõ) nhưng SCALE về Tổng CP HQKD (nay tổng công ty = HQKD) để Σ nhóm == 1047
    # HQKD, khớp màn Chi phí. (Cơ cấu theo nhóm HQKD #10 biến đổi/cố định là follow-up riêng nếu cần.)
    _cp_nhom = dict((_kq or {}).get("__chiphi_nhom__") or {})
    _hqkd_cp = round(sum(r["Thực hiện (tỷ)"] for r in records if r["Chỉ tiêu KQKD"] == "Tổng chi phí"), 9)
    _kq_cp = round(sum(v for v in _cp_nhom.values() if v), 9)
    if _cp_nhom and _kq_cp:
        _f = _hqkd_cp / _kq_cp
        _cp_nhom = {n: (round(v * _f, 9) if v else v) for n, v in _cp_nhom.items()}
    _cp_recs = [{"Kỳ (yyyy-mm)": period, "Nhóm CP (chuẩn mực KT)": n, "Khoản mục chi tiết": n,
                 "Thực hiện (tỷ)": v} for n, v in _cp_nhom.items() if v]
    _cpr = _fill_import_chiphi(_cp_recs, period, "XVP", file_path) if _cp_recs else None
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"),
            "target": "01_HQKD", "value_col_header": "HQKD depot G/I/K",
            "depots": [cc for cc, _ in depots], "chiphi": _cpr}


def _derive_kqkd_htx(file_path: str, period: str, cong_ty: str):
    """HTX Xanh VP/TQ — đọc sheet 'HQKD' QUẢN TRỊ (1 cột 'T{mm}.YYYY'), THAY vì 'BC KQKD' (B02-HTX).
    Lý do: giao dịch bán xe HTKD (giá vốn ≈ doanh thu, pass-through nội bộ) CHỈ có ở BC KQKD; sheet
    HQKD quản trị đã LOẠI phần này -> khớp báo cáo hợp nhất (BCHN). Cùng format HQKD của XVP: '3. Doanh
    thu thuần' / IV giá vốn / V lãi gộp / VI CP biến đổi / VIII CP cố định / IX tài chính / (EBT) LNTT /
    (EAT) LNST. GIỮ NGUYÊN công thức: Tổng CP = DThu thuần − LNTT + DT tài chính + thu khác (=0 cho HTX).
    Gate: THƯ MỤC nguồn 'HTXXANH*' (đúng 2 pháp nhân HTX Taxi Xanh) — KHÔNG gate theo ô A1 vì HTX_XTQ
    dùng nhầm tiêu đề template 'Khối vận tải taxi - XanhVP'. None nếu không phải -> generic chạy tiếp."""
    import re as _re
    from openpyxl import load_workbook
    from servers import template_filler as tf
    from servers.common import be_bridge as bb
    if not _source_id(file_path).split("::", 1)[0].upper().startswith("HTXXANH"):
        return None                               # CHỈ HTX Xanh VP/TQ; cty khác -> generic TT200
    _norm = lambda v: bb.remove_diacritics("" if v is None else str(v)).strip().lower()  # noqa: E731
    wb = bb.fast_load_workbook(file_path, read_only=True, data_only=True)
    sh = next((s for s in wb.sheetnames if _norm(s) == "hqkd"), None)
    if not sh:
        wb.close()
        return None
    rows = [list(r) for r in wb[sh].iter_rows(values_only=True)]
    wb.close()
    hdr_i = next((i for i, r in enumerate(rows[:8]) if any("chi tieu" in _norm(c) for c in r)), None)
    if hdr_i is None:
        return None
    # Cột nhãn = cột 'CHỈ TIÊU' (KHÔNG hardcode 0 — HTX_XTQ T01 lệch +1 cột). Cột giá trị = 'T{mm}.YYYY'.
    lab_j = next(j for j, c in enumerate(rows[hdr_i]) if "chi tieu" in _norm(c))
    mm = period.split("-")[1] if "-" in period else ""
    _mi = str(int(mm)) if mm.isdigit() else mm
    val_j = next((j for j, c in enumerate(rows[hdr_i])
                  if _norm(c).startswith((f"t{_mi}.", f"t{mm}.", f"t0{_mi}."))), None)
    if val_j is None:
        val_j = lab_j + 3                          # sau Chỉ tiêu/Mã số/Thuyết minh

    def _lbl(*subs):
        for r in rows[hdr_i + 1:]:
            if lab_j < len(r) and r[lab_j] not in (None, "") and any(s in _norm(r[lab_j]) for s in subs):
                v = r[val_j] if val_j < len(r) else None
                return round(v * 1e-9, 9) if isinstance(v, (int, float)) else None
        return None
    dt_net = _lbl("3. doanh thu thuan", "doanh thu thuan ve ban hang")
    ebt = _lbl("(ebt)")
    if dt_net is None or ebt is None:
        return {"ok": False, "error": "HTX HQKD: thiếu '3. DThu thuần' hoặc dòng (EBT)"}
    gia_von = _lbl("iv. gia von", "gia von hang ban")
    ln_gop = _lbl("loi nhuan gop ve ban hang", "v. loi nhuan gop")
    eat = _lbl("(eat)")
    hhdv = _lbl("1. doanh thu ban hang va cung cap")
    dttc = _lbl("doanh thu tai chinh") or 0.0
    tnk = _lbl("1. thu nhap khac", "thu nhap khac") or 0.0
    tong_cp = round((dt_net - ebt) + dttc + tnk, 9)      # GIỮ công thức: DThu thuần − LNTT (+ thu ngoài)
    records = []

    def add(ten, val):
        if val is not None:
            records.append({"Kỳ (yyyy-mm)": period, "Chỉ tiêu KQKD": ten, "Thực hiện (tỷ)": val})
    add("Doanh thu thuần", dt_net)                        # -> 1000 + DTHU
    add("Tổng chi phí", tong_cp)                          # -> 1047
    add("Lợi nhuận trước thuế", ebt)                      # -> 1112
    add("Giá vốn hàng bán", gia_von)                      # -> PNLT
    add("Lợi nhuận gộp", ln_gop if ln_gop is not None else                       # #6 = Mã10−Mã11 (mapping)
        (round(dt_net - (gia_von or 0.0), 9) if dt_net is not None else None))
    add("Lợi nhuận sau thuế", eat if eat is not None else ebt)   # -> PNLT (thẻ LNST)
    add("Doanh thu HH, DV", hhdv if hhdv is not None else dt_net)  # -> PNLT (#1)
    if dttc:
        add("Doanh thu tài chính", dttc)                  # -> PNLT (IX.1; chỉ emit khi nguồn có)
    if tnk:
        add("Thu nhập khác", tnk)                         # -> PNLT (X.1)
    out = os.path.join(tf.FILLED_DIR, f"KQKD_{period}_{cong_ty or 'NA'}_01_HQKD.xlsx")
    tf.fill("01_HQKD", records, out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    # 02_CHIPHI (cơ cấu, Mapping #10 HTX = sheet HQKD): IV giá vốn + VI CP biến đổi + VIII CP cố định +
    # CP tài chính (Σ == Tổng CP). Bỏ dòng 0 (giá vốn HTX thường = 0).
    _cp = {"Giá vốn hàng bán": gia_von, "Chi phí biến đổi": _lbl("vi. chi phi bien doi"),
           "Chi phí cố định": _lbl("viii. chi phi co dinh"), "Chi phí tài chính": _lbl("2. chi phi tai chinh")}
    _cp_recs = [{"Kỳ (yyyy-mm)": period, "Nhóm CP (chuẩn mực KT)": n, "Khoản mục chi tiết": n,
                 "Thực hiện (tỷ)": v} for n, v in _cp.items() if v]
    _cpr = _fill_import_chiphi(_cp_recs, period, cong_ty, file_path) if _cp_recs else None
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"),
            "target": "01_HQKD", "value_col_header": f"HQKD T{_mi}", "chiphi": _cpr}


def _derive_kqkd(file_path: str, sheet: str, period: str, cong_ty: str):
    """TẤT ĐỊNH: KQKD chuẩn TT200 -> 01_HQKD -> DTHU/HQKD/PNLT (+ dòng 'Lợi nhuận sau thuế' cho
    thẻ LNST). Đọc theo MÃ SỐ (10/11/20/50/60...), CHỌN cột giá trị nơi Mã 10 (DT thuần) ≠ 0 ->
    tránh đúng lỗi agent chọn nhầm cột rỗng ('Tháng 1' trống, dữ liệu ở 'Kỳ này lũy kế'). Công ty
    dùng mã riêng (XDV B-series / SRVF A-series) -> KHÔNG có Mã 10/50 -> trả not-ok (fallback LLM)."""
    from openpyxl import load_workbook
    from servers import template_filler as tf
    from servers.common import be_bridge as bb
    norm = lambda v: bb.normalize_header(v, True)  # noqa: E731

    wb = bb.fast_load_workbook(file_path, read_only=True, data_only=True)
    try:
        rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()
    # T-SERIES trước (HT: mã T100/T200/T300, cột "Tháng {mm}"). None -> chạy tiếp TT200 bên dưới
    # (cty khác không có mã T-code nên không bị ảnh hưởng).
    _ts = _derive_kqkd_tseries(rows, period, cong_ty, file_path)
    if _ts is not None:
        return _ts
    # DUAN (Khối Dự án): sheet HQKD dòng-có-tên + cột "Tổng dự án". None -> chạy tiếp TT200.
    _du = _derive_kqkd_duan(rows, period, cong_ty, file_path)
    if _du is not None:
        return _du
    # XDV (Vinfast Xưởng Dịch Vụ): 'Sheet1' mã B-series, cột "Kỳ này". None -> chạy tiếp TT200.
    _xd = _derive_kqkd_xdv(rows, period, cong_ty, file_path)
    if _xd is not None:
        return _xd
    # SRVF (Showroom Vinfast): sheet 'KQKD' mã A-series (A100/A300/A600), cột 'Kỳ này'. None -> TT200.
    _sv = _derive_kqkd_srvf(rows, period, cong_ty, file_path)
    if _sv is not None:
        return _sv
    # HO (Khối hỗ trợ tập đoàn): 'HO_KQKD' dòng tổng có tên + cột tháng 'T{mm}'. None -> TT200.
    _ho = _derive_kqkd_ho(rows, period, cong_ty, file_path)
    if _ho is not None:
        return _ho
    # XVP CÔNG TY (Xanh Vĩnh Phúc): sheet 'HQKD' tách 3 depot (cost center). Đọc sheet HQKD riêng
    # (không dùng `rows` của sheet được truyền). None -> file khác (kể cả HTX_XVP HQKD 1 cột) -> TT200.
    _xvp = _derive_kqkd_xvp(file_path, period, cong_ty)
    if _xvp is not None:
        return _xvp
    # HTX Xanh VP/TQ: đọc sheet 'HQKD' quản trị (loại giao dịch nội bộ pass-through) thay 'BC KQKD'.
    # None -> không phải HTX HQKD -> generic TT200 chạy tiếp (0 ảnh hưởng cty khác).
    _htx = _derive_kqkd_htx(file_path, period, cong_ty)
    if _htx is not None:
        return _htx
    hi = next((i for i, r in enumerate(rows[:15])
               if any(norm(c).startswith(("chi tieu", "noi dung")) for c in r)
               and any("ma so" in norm(c) for c in r)), None)
    if hi is None:
        return {"ok": False, "error": "KQKD: không thấy header Chỉ tiêu/Mã số"}
    hdr = rows[hi]
    ct_i = next((j for j, c in enumerate(hdr) if norm(c).startswith(("chi tieu", "noi dung"))), None)
    ma_i = next((j for j, c in enumerate(hdr) if "ma so" in norm(c)), None)
    if ct_i is None or ma_i is None:
        return {"ok": False, "error": "KQKD: thiếu cột Chỉ tiêu/Mã số"}

    def code_of(r):
        return str(r[ma_i]).strip() if ma_i < len(r) and r[ma_i] not in (None, "") else ""
    # CHỌN CỘT giá trị: (1) ƯU TIÊN cột có HEADER kỳ HIỆN TẠI ("kỳ này"/"thực hiện"/"tháng này") —
    # tránh lấy nhầm cột "Kỳ trước"/"cùng kỳ"/"kế hoạch" khi THÁNG NÀY < THÁNG TRƯỚC (heuristic cũ
    # 'cột tổng |giá trị| lớn nhất' chọn SAI: HTX_XVP T03/T05, An KS T03/T05 lấy đúng cột Kỳ trước).
    # (2) KHÔNG có header kỳ-hiện-tại rõ -> fallback cột tổng lớn nhất TRONG các cột KHÔNG phải
    # kỳ-trước/kế-hoạch (giữ xử lý cũ: cột 'Năm nay' của An Taxi/GA, tránh cột lũy kế rỗng / GA lỗ
    # Mã10=0 mà Mã26/50/60≠0). Mã cốt lõi để đo tổng.
    core = [r for r in rows[hi + 1:] if code_of(r) in ("10", "11", "20", "26", "50", "60")]
    if not core:
        return {"ok": False, "error": "KQKD: không thấy mã TT200 cốt lõi — có thể mã riêng, để LLM"}
    _CUR = ("ky nay", "thuc hien", "thang nay", "ky bao cao")
    _PREV = ("ky truoc", "ki truoc", "nam truoc", "cung ky", "dau ky", "dau nam", "ke hoach")

    def _coltot(j):
        return sum(abs(r[j]) for r in core if j < len(r) and isinstance(r[j], (int, float)))
    val_j = next((j for j in range(ma_i + 1, len(hdr))
                  if any(k in norm(hdr[j]) for k in _CUR)
                  and not any(k in norm(hdr[j]) for k in _PREV) and _coltot(j) > 0), None)
    if val_j is None:               # fallback: cột |giá trị| lớn nhất, BỎ cột kỳ-trước/kế-hoạch
        best = 0.0
        for j in range(ma_i + 1, len(hdr)):
            if any(k in norm(hdr[j]) for k in _PREV):
                continue
            tot = _coltot(j)
            if tot > best:
                best, val_j = tot, j
    if val_j is None:
        return {"ok": False, "error": "KQKD: mọi cột giá trị = 0 (nguồn trống)"}

    def num(r):
        v = r[val_j] if val_j < len(r) else None
        return round(v * 1e-9, 9) if isinstance(v, (int, float)) else None

    vals = {}   # base code -> value (tỷ), lấy dòng ĐẦU cho mỗi mã (dòng tổng cấp cao)
    for r in rows[hi + 1:]:
        c = code_of(r)
        if (c in _TT200_CHITIEU or c == "31") and c not in vals:  # mã31 (thu nhập khác) đọc để tính TỔNG DT, không emit riêng
            vals[c] = num(r)
    if "10" not in vals or ("50" not in vals and "60" not in vals):
        return {"ok": False, "error": "KQKD: thiếu Mã 10/50/60 (không phải TT200 chuẩn) -> LLM"}

    # TỔNG DOANH THU = mã10 (DT thuần bán hàng) + mã21 (DT hoạt động tài chính) + mã31 (thu nhập khác)
    # — khớp cách kế toán "Tổng doanh thu = DT chính + DT khác" (vd An KS: DT KS 82.635tr + lãi TG mã21
    # 7.768k). LNTT (mã50) ĐÃ gồm mã21/31 ở vế thu → Tổng chi phí phải = TỔNG DT − LNTT mới cân đối
    # (nếu chỉ lấy mã10 thì chi phí thiếu đúng mã21+mã31). Chốt kế toán 2026-07-16.
    dt_total = round((vals.get("10") or 0.0) + (vals.get("21") or 0.0) + (vals.get("31") or 0.0), 9)
    # Mã 12 (CP quản lý kinh doanh) chỉ có ở mẫu B02-HTX/TT133 (HTX Xanh VP/TQ) — TT200 không dùng.
    # Dùng làm CỜ nhận diện B02-HTX để sửa 'Lợi nhuận gộp' (xem dưới). TT200 -> None -> 0 ảnh hưởng.
    _ma12 = next((num(r) for r in rows[hi + 1:] if code_of(r) == "12"), None)

    records = []
    for c, ten in _TT200_CHITIEU.items():
        if c in vals and vals[c] is not None:
            if c == "10":
                v = dt_total                         # mã10 ("Doanh thu thuần") mang TỔNG DT -> HQKD 1000/DTHU
            elif c == "20" and _ma12 is not None and vals.get("11") is not None:
                # B02-HTX: Mã20 là 'Kết quả HĐSXKD' (=Mã10−Mã11−Mã12, ĐÃ trừ CP QLKD) -> KHÔNG phải
                # lãi gộp. Mapping sheet 13/14 #6: Lợi nhuận gộp = Mã10 − Mã11. TT200 (_ma12=None) giữ Mã20.
                v = round((vals["10"] or 0.0) - (vals["11"] or 0.0), 9)
            else:
                v = vals[c]
            records.append({"Kỳ (yyyy-mm)": period, "Chỉ tiêu KQKD": ten, "Thực hiện (tỷ)": v})
    # 'Tổng chi phí' = TỔNG doanh thu − LNTT (đẳng thức, sinh HQKD 1047). Chỉ khi có cả 10 và 50.
    if vals.get("10") is not None and vals.get("50") is not None:
        records.append({"Kỳ (yyyy-mm)": period, "Chỉ tiêu KQKD": "Tổng chi phí",
                        "Thực hiện (tỷ)": round(dt_total - vals["50"], 9)})  # 9 số lẻ giữ đồng; TỔNG DT gồm mã21/31
    # Doanh thu HH, DV (#1 bảng 50 = doanh thu bán hàng & CCDV GỘP, "Theo BCTC"): mã 01 (TT200:
    # An KS/An Taxi…) HOẶC mã 1 (B02-HTX / mẫu quản trị HQKD của Taxi Xanh XVP·HTX_XVP·HTX_XTQ) —
    # dòng doanh thu GỘP trước giảm trừ, KHÁC mã10 'Doanh thu thuần'. Đọc riêng (không nằm trong
    # _TT200_CHITIEU), cùng cột val_j. -> PNLT 'Doanh thu HH, DV' (metrics_extra đọc #1). None -> bỏ.
    gross_hhdv = next((num(r) for r in rows[hi + 1:] if code_of(r) in ("1", "01")), None)
    if gross_hhdv is not None:
        records.append({"Kỳ (yyyy-mm)": period, "Chỉ tiêu KQKD": "Doanh thu HH, DV",
                        "Thực hiện (tỷ)": gross_hhdv})
    out = os.path.join(tf.FILLED_DIR, f"KQKD_{period}_{cong_ty or 'NA'}_01_HQKD.xlsx")
    tf.fill("01_HQKD", records, out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"),
            "target": "01_HQKD", "value_col_header": str(hdr[val_j])[:30]}


def _derive_thue(file_path: str, sheet: str, period: str, cong_ty: str):
    """TẤT ĐỊNH: từ CĐSPS (cùng cấu trúc 2 tầng Nợ/Có như TK131) lọc TK 133 (GTGT được khấu trừ
    -> PHẢI THU) + TK 333 (thuế phải nộp -> PHẢI NỘP), lấy SỐ DƯ CUỐI KỲ -> 10_THUE -> report_type
    THẬT 'THUE'. Lấy TK cấp cha (133/333) cho TỔNG đúng (không cộng trùng con). value_scale=1e-9."""
    from openpyxl import load_workbook
    from servers import template_filler as tf
    from servers.common import be_bridge as bb
    mp, _meta = _heuristic_tk_mapping(file_path, sheet, "CDPS")   # tái dùng dò 2 tầng Nợ/Có
    if not mp:
        return {"ok": False, "error": "CĐSPS: không dò được header 2 tầng"}
    role_idx = {c["role"]: c["index"] for c in mp["columns"]}
    code_i, cuoi_no, cuoi_co = role_idx.get("entity_code"), role_idx.get("cuoi_ky_no"), role_idx.get("cuoi_ky_co")
    dau_no, dau_co = role_idx.get("dau_ky_no"), role_idx.get("dau_ky_co")
    ps_no, ps_co = role_idx.get("phat_sinh_no"), role_idx.get("phat_sinh_co")
    name_i = role_idx.get("entity_name")
    if code_i is None or cuoi_no is None or cuoi_co is None:
        return {"ok": False, "error": "CĐSPS: thiếu Mã TK / cột cuối kỳ"}
    _partial = bool(_meta.get("partial"))   # đầu kỳ mập mờ -> bỏ đầu (FE lấy cuối kỳ tháng trước)
    wb = bb.fast_load_workbook(file_path, read_only=True, data_only=True)
    try:
        rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()

    def num(r, i):
        return round(r[i] * 1e-9, 9) if (i is not None and i < len(r) and isinstance(r[i], (int, float))) else None

    def _net(r, pos, neg):   # số dư RÒNG (chiều chính − ngược) để bảng cân: cuối = đầu + tăng − giảm
        p, n = num(r, pos), num(r, neg)
        return None if (p is None and n is None) else round((p or 0) - (n or 0), 9)

    records = []
    for ri in range(mp["data_start_row"], len(rows)):
        r = rows[ri]
        code = str(r[code_i]).strip() if code_i < len(r) and r[code_i] not in (None, "") else ""
        if code not in ("133", "333"):     # chỉ TK cha -> tổng đúng, không cộng trùng con
            continue
        ten = bb.parse_text(r[name_i]) if (name_i is not None and name_i < len(r)) else None
        if code == "133":   # GTGT được khấu trừ = dư NỢ (phải thu); tăng = PS Nợ, giảm = PS Có
            pt = "Phải thu"
            cuoi, dau = _net(r, cuoi_no, cuoi_co), _net(r, dau_no, dau_co)
            tang, giam = num(r, ps_no), num(r, ps_co)
        else:               # 333 thuế phải nộp = dư CÓ; tăng = PS Có, giảm = PS Nợ
            pt = "Phải nộp"
            cuoi, dau = _net(r, cuoi_co, cuoi_no), _net(r, dau_co, dau_no)
            tang, giam = num(r, ps_co), num(r, ps_no)
        rec = {"Kỳ": period, "Đơn vị": cong_ty,
               "Loại thuế (GTGT ra/vào, TNCN, TNDN, NK, khác)": ten or code,
               "Phải thu/Phải nộp": pt, "Dư cuối kỳ (tỷ)": cuoi,
               "PS tăng (tỷ)": tang, "PS giảm (tỷ)": giam}
        if not _partial:    # đầu kỳ chỉ điền khi CHẮC (mập mờ -> FE dùng cuối kỳ tháng trước)
            rec["Dư đầu kỳ (tỷ)"] = dau
        records.append(rec)
    if not records:
        return {"ok": False, "error": "CĐSPS: không thấy TK 133/333"}
    out = os.path.join(tf.FILLED_DIR, f"THUE_{period}_{cong_ty or 'NA'}_10_THUE.xlsx")
    tf.fill("10_THUE", records, out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"), "target": "10_THUE"}


def _derive_thue_cdkt(file_path: str, sheet: str, period: str, cong_ty: str):
    """TẤT ĐỊNH (dùng cho GA): thuế lấy từ CĐKT RIÊNG — GA KHÔNG có sheet CĐPS riêng, chỉ có
    'TC_CDPS' (CĐPS HỢP NHẤT Thịnh Cường Group, ĐVT triệu đồng -> sai pháp nhân/đơn vị). Lấy:
      • Mã 152 'Thuế GTGT được khấu trừ'             -> Phải thu (dư Nợ, tài sản)
      • Mã 313 'Thuế và các khoản phải nộp Nhà nước' -> Phải nộp (dư Có, nợ phải trả)
    Số cuối kỳ + Số đầu kỳ (CĐKT không có cột phát sinh -> biến động do metrics suy từ kỳ trước,
    xem _mvrows). report_type THẬT 'THUE'. value_scale=1e-9 (đồng -> tỷ). import_filled đè dữ liệu
    THUE cũ theo source_file (dọn bản LLM cũ dump nhầm sheet hợp nhất, 50 dòng không nhãn)."""
    from openpyxl import load_workbook
    from servers import template_filler as tf
    from servers.common import be_bridge as bb
    import re as _re
    _n = lambda x: bb.remove_diacritics("" if x is None else str(x)).strip().lower()  # noqa: E731
    wb = bb.fast_load_workbook(file_path, read_only=True, data_only=True)
    try:
        rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()
    hdr_i = code_i = name_i = cuoi_i = dau_i = None
    for i, r in enumerate(rows[:40]):     # dò header CĐKT: có 'Mã số' + 'Số cuối kỳ' (GA header ở ~dòng 9)
        cells = [_n(c) for c in r]
        has_cuoi = any(("so cuoi" in c or "cuoi ky" in c or "cuoi ki" in c) for c in cells)
        has_ma = any((c == "ma" or "ma so" in c) for c in cells)
        if has_cuoi and has_ma:
            hdr_i = i
            for j, c in enumerate(cells):
                if name_i is None and c in ("tai san", "noi dung", "chi tieu", "khoan muc"): name_i = j
                if code_i is None and (c == "ma" or "ma so" in c): code_i = j
                if cuoi_i is None and ("so cuoi" in c or "cuoi ky" in c or "cuoi ki" in c): cuoi_i = j
                if dau_i is None and ("so dau" in c or "dau ky" in c or "dau ki" in c): dau_i = j
            break
    if hdr_i is None or code_i is None or cuoi_i is None:
        return {"ok": False, "error": "CĐKT: không dò được header (Mã số / Số cuối kỳ)"}
    if name_i is None:
        name_i = 0

    def _code(v):
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return "" if v in (None, "") else str(v).strip()

    def num(r, idx):
        return round(r[idx] * 1e-9, 9) if (idx is not None and idx < len(r) and isinstance(r[idx], (int, float))) else None

    _MAP = {"152": "Phải thu", "313": "Phải nộp"}   # 152 GTGT được khấu trừ (dư Nợ); 313 thuế phải nộp NN (dư Có)
    records = []
    for r in rows[hdr_i + 1:]:
        code = _code(r[code_i]) if code_i < len(r) else ""
        if code not in _MAP:
            continue
        cuoi = num(r, cuoi_i)
        if cuoi is None:     # dòng thiếu số (vd mã 153/262 rỗng ở GA) -> bỏ
            continue
        ten = bb.parse_text(r[name_i]) if (name_i < len(r) and r[name_i] not in (None, "")) else None
        ten = _re.sub(r"^\s*\d+[.)]\s*", "", ten) if ten else code    # bỏ '2. '/'3. ' đầu dòng
        records.append({"Kỳ": period, "Đơn vị": cong_ty,
                        "Loại thuế (GTGT ra/vào, TNCN, TNDN, NK, khác)": ten,
                        "Phải thu/Phải nộp": _MAP[code],
                        "Dư cuối kỳ (tỷ)": cuoi,
                        "Dư đầu kỳ (tỷ)": num(r, dau_i)})
    if not records:
        return {"ok": False, "error": "CĐKT: không thấy mã 152/313"}
    out = os.path.join(tf.FILLED_DIR, f"THUE_{period}_{cong_ty or 'NA'}_10_THUE.xlsx")
    tf.fill("10_THUE", records, out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"), "target": "10_THUE"}


def _derive_tonkho_cdps(file_path: str, sheet: str, period: str, cong_ty: str):
    """TẤT ĐỊNH: tồn kho từ CĐPS/CĐSPS (bảng cân đối TK) — lọc các TK kho CẤP CHA 151–156 (151 hàng
    đi đường, 152 NVL, 153 CCDC, 154 SPDD, 155 thành phẩm, 156 hàng hóa) lấy số dư Nợ CUỐI KỲ +
    PS Nợ (nhập) / PS Có (xuất) -> 09_TONKHO -> HH. Dùng cho đơn vị KHÔNG có sheet NXT riêng
    (DUAN: NVL 152 công trình; An KS: hàng hóa 156). Lấy TK CHA -> tổng đúng, không cộng con (1561…).
    CHỈ emit dòng có DƯ CUỐI KỲ > 0 (số dư tồn kho THỰC) -> tự loại TK 'chạy-qua' dư cuối=0 (vd An Taxi
    154 'CP dịch vụ dở dang' có phát sinh nhưng cuối kỳ=0 = KHÔNG có tồn kho, spec ghi 'Không có').
    Trả not-ok nếu không TK kho nào còn số dư (đơn vị khác/An Taxi không bị ảnh hưởng)."""
    from openpyxl import load_workbook
    from servers import template_filler as tf
    from servers.common import be_bridge as bb
    mp, _meta = _heuristic_tk_mapping(file_path, sheet, "CDPS")
    if not mp:
        return {"ok": False, "error": "CĐPS: không dò được header 2 tầng"}
    role_idx = {c["role"]: c["index"] for c in mp["columns"]}
    code_i, cuoi_no = role_idx.get("entity_code"), role_idx.get("cuoi_ky_no")
    dau_no = role_idx.get("dau_ky_no")   # Dư ĐẦU KỲ Nợ (TK152 là TS -> dư Nợ) -> tồn đầu kỳ
    ps_no, ps_co, name_i = role_idx.get("phat_sinh_no"), role_idx.get("phat_sinh_co"), role_idx.get("entity_name")
    if code_i is None or cuoi_no is None:
        return {"ok": False, "error": "CĐPS: thiếu Mã TK / cột cuối kỳ"}
    wb = bb.fast_load_workbook(file_path, read_only=True, data_only=True)
    try:
        rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()

    def num(r, i):
        return round(r[i] * 1e-9, 9) if (i is not None and i < len(r) and isinstance(r[i], (int, float))) else None
    _TK_KHO = ("151", "152", "153", "154", "155", "156")
    # Khối DỰ ÁN: hướng dẫn chốt tồn kho = CHỈ TK 152 (Nguyên vật liệu công trình); 151 hàng đi
    # đường / 153 CCDC / 154 CP SXKD dở dang là khoản chạy-qua, KHÔNG tính vào tồn kho dự án (khớp
    # số báo cáo: đầu kỳ/nhập/xuất/cuối chỉ reconcile trên 152). Đơn vị khác giữ nguyên 151–156.
    if "dự án" in (_khoi_of(file_path) or "").lower():
        _TK_KHO = ("152",)
    records = []
    for ri in range(mp["data_start_row"], len(rows)):
        r = rows[ri]
        code = str(r[code_i]).strip() if code_i < len(r) and r[code_i] not in (None, "") else ""
        if code not in _TK_KHO:     # chỉ TK kho cấp CHA 151–156 -> tổng đúng (không cộng con 1521/1561…)
            continue
        cuoi = num(r, cuoi_no)
        _dau = num(r, dau_no)
        # GIỮ TK có tồn ĐẦU hoặc CUỐI kỳ > 0. Bỏ TK chạy-qua (đầu=cuối=0, chỉ có PS — vd An Taxi
        # 154 'CP dịch vụ dở dang'). TRƯỚC chỉ giữ cuối>0 -> MẤT tồn kho TIÊU HẾT trong kỳ (đầu>0,
        # cuối=0, vd HTX_XTQ T02 TK152 'Vật liệu, dụng cụ' đầu 3,5tr / xuất 3,5tr / cuối 0) -> tồn
        # đầu kỳ không lên. Nay đầu>0 vẫn giữ để hiện đầu kỳ (cuối có thể =0).
        if not (abs(cuoi or 0) > 1e-9 or abs(_dau or 0) > 1e-9):
            continue
        ten = bb.parse_text(r[name_i]) if (name_i is not None and name_i < len(r)) else None
        records.append({"Kỳ": period, "Đơn vị": cong_ty, "TK (151-156)": code,
                        "Loại HTK (NVL/Vật tư/Hàng hóa…)": ten or f"Tồn kho (TK{code})",
                        "Dư đầu kỳ (tỷ)": num(r, dau_no),
                        "Dư cuối kỳ (tỷ)": cuoi,
                        "Nhập trong kỳ (tỷ)": num(r, ps_no), "Xuất trong kỳ (tỷ)": num(r, ps_co)})
    if not records:
        return {"ok": False, "error": "CĐPS: không thấy TK kho 151–156 còn số dư"}
    out = os.path.join(tf.FILLED_DIR, f"TONKHO_{period}_{cong_ty or 'NA'}_09_TONKHO.xlsx")
    tf.fill("09_TONKHO", records, out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"), "target": "09_TONKHO"}


def _derive_tscd(file_path: str, sheet: str, period: str, cong_ty: str):
    """TẤT ĐỊNH: bảng KHẤU HAO-PHÂN BỔ -> 08_TSCD -> report_type THẬT 'TS'. Header nhóm: 'Tài sản'
    (nguyên giá) / 'Khấu hao' (hao mòn) / 'Giá trị sổ sách' (GTCL), tầng dưới là các CỘT NGÀY
    (đầu kỳ / + / - / cuối kỳ). Lấy dòng SUBTOTAL theo LOẠI TS (dòng mở đầu 'False <loại>' =
    tổng mỗi loại) -> dim1=Loại TS, neo=GTCL cuối. NG cuối/HM cuối/PS lấy theo vai. Layout khá
    đặc thù -> KHÔNG khớp thì trả not-ok (fallback LLM)."""
    from openpyxl import load_workbook
    from servers import template_filler as tf
    from servers.common import be_bridge as bb
    norm = lambda v: bb.normalize_header(v, True)  # noqa: E731

    wb = bb.fast_load_workbook(file_path, read_only=True, data_only=True)
    try:
        rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()
    # group row: có 'tai san' + 'khau hao'; sub row = ngay dưới
    grp_i = next((i for i, r in enumerate(rows[:8])
                  if any("tai san" in norm(c) for c in r) and any("khau hao" in norm(c) for c in r)), None)
    if grp_i is None:
        return {"ok": False, "error": "TSCĐ: không thấy nhóm Tài sản/Khấu hao"}
    group = _forward_fill(rows[grp_i])
    sub = rows[grp_i + 1]
    def _closing_col(group_kw):
        # cột 'cuối kỳ' của 1 nhóm = cột có group_kw + nhãn sub là NGÀY muộn nhất / chứa '31'
        cands = [j for j in range(len(sub)) if j < len(group) and group_kw in norm(group[j])]
        if not cands:
            return None
        # ưu tiên nhãn sub chứa '31' (cuối tháng); nếu không, cột phải nhất trong nhóm
        end = [j for j in cands if "31" in str(sub[j] or "")]
        return (end or cands)[-1]
    ng_i = _closing_col("tai san")
    hm_i = _closing_col("khau hao")
    gtcl_i = _closing_col("gia tri so sach") or _closing_col("gia tri")
    if ng_i is None and gtcl_i is None:
        return {"ok": False, "error": "TSCĐ: không xác định được cột nguyên giá/GTCL cuối kỳ"}
    data_start = grp_i + 2

    def num(r, i):
        return round(r[i] * 1e-9, 9) if (i is not None and i < len(r) and isinstance(r[i], (int, float))) else None

    records = []
    for ri in range(data_start, len(rows)):
        r = rows[ri]
        c0 = str(r[0]).strip() if r and r[0] not in (None, "") else ""
        if not c0:
            continue
        # dòng SUBTOTAL theo loại: mở đầu 'False <loại>' (đặc thù export này). Bỏ 'Tổng'.
        low = norm(c0)
        if low.startswith("tong"):
            continue
        if not c0.lower().startswith("false"):
            continue     # chỉ lấy dòng loại (subtotal), bỏ dòng tài sản chi tiết -> không đếm trùng
        loai = c0[5:].strip(" :–-") or c0
        ng, gtcl = num(r, ng_i), num(r, gtcl_i)
        hm = num(r, hm_i)
        if ng is None and gtcl is None:
            continue
        rec = {"Kỳ": period, "Đơn vị": cong_ty,
               "Loại TS (Nhà cửa VKT/Máy móc TB/PTVT/Phần mềm-QSDĐ…)": loai,
               "Nguyên giá (tỷ)": ng, "Giá trị còn lại (tỷ)": gtcl}
        if hm is not None:
            rec["Hao mòn lũy kế (tỷ)"] = hm
        records.append(rec)
    if not records:
        return {"ok": False, "error": "TSCĐ: không bóc được dòng loại tài sản (layout lạ -> để LLM)"}
    out = os.path.join(tf.FILLED_DIR, f"TSCD_{period}_{cong_ty or 'NA'}_08_TSCD.xlsx")
    tf.fill("08_TSCD", records, out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"), "target": "08_TSCD"}


def _derive_tscd_cdkt(file_path: str, cdkt_sheet: str, period: str, cong_ty: str):
    """TSCĐ lấy từ CĐKT (SỐ DƯ CUỐI KỲ) — nguồn chuẩn hơn 'Biểu khấu hao' (biểu khấu hao gồm cả
    TS ngoài bảng cân đối → over-state; vd An Taxi biểu 75.45 ≠ CĐKT mã222 64.899). Nguyên giá =
    mã 222(HH)+225(thuê TC)+228(vô hình); hao mòn = 223+226+229 (file lưu ÂM); GTCL = NG − |hao mòn|."""
    from openpyxl import load_workbook
    from servers import template_filler as tf
    from servers.common import be_bridge as bb
    norm = lambda v: bb.normalize_header(v, True)  # noqa: E731
    wb = bb.fast_load_workbook(file_path, read_only=True, data_only=True)
    try:
        rows = [list(r) for r in wb[cdkt_sheet].iter_rows(values_only=True)]
    finally:
        wb.close()
    _NAME_KW = ("noi dung", "tai san", "chi tieu", "khoan muc")

    def _is_name(c):
        n = norm(c)
        return any(n.startswith(k) or k in n for k in _NAME_KW)
    hi = next((i for i, r in enumerate(rows[:15])
               if any(_is_name(c) for c in r) and any("ma so" in norm(c) for c in r)), None)
    if hi is None:
        return {"ok": False, "error": "TSCĐ-CĐKT: không thấy header"}
    hdr = rows[hi]
    ma_i = next((j for j, c in enumerate(hdr) if "ma so" in norm(c)), None)
    cuoi_i = next((j for j, c in enumerate(hdr) if "cuoi" in norm(c)), None)
    if cuoi_i is None:
        cuoi_i = next((j for j, c in enumerate(hdr) if "c. nam" in norm(c) or "ky nay" in norm(c)), None)
    if cuoi_i is None:   # CĐKT layout CỘT-THEO-THÁNG (HT 'BCĐKT hợp nhất': THÁNG{mm}/THANG{mm}, loại
        # cột 'ĐIỀU CHỈNH T{mm}'). Số cuối kỳ = cột tháng của kỳ — cùng logic _derive_cdkt. Chỉ khi
        # match 'cuối kỳ' chuẩn fail -> đơn vị có cột cuối kỳ giữ nguyên (0 ảnh hưởng).
        try:
            _mm = int(str(period)[-2:])
        except Exception:
            _mm = None
        if _mm is not None:
            _cur = (f"thang {_mm}", f"t{_mm:02d}", f"t{_mm}")
            cuoi_i = next((j for j, c in enumerate(hdr)
                           if j > (ma_i or 0) and "dieu chinh" not in norm(c) and norm(c) in _cur), None)
    if ma_i is None or cuoi_i is None:
        return {"ok": False, "error": "TSCĐ-CĐKT: thiếu cột Mã số / cuối kỳ"}
    vals = {}
    for r in rows[hi + 1:]:
        c = str(r[ma_i]).strip() if ma_i < len(r) and r[ma_i] not in (None, "") else ""
        if c in ("222", "225", "228", "223", "226", "229") and c not in vals:
            x = r[cuoi_i] if cuoi_i < len(r) else None
            vals[c] = round(x * 1e-9, 9) if isinstance(x, (int, float)) else None
    ng = round(sum(vals.get(k) or 0.0 for k in ("222", "225", "228")), 9)  # 9 số lẻ = giữ đồng (vals đã 9), tránh làm tròn nghìn
    hm = round(sum(vals.get(k) or 0.0 for k in ("223", "226", "229")), 9)   # ÂM sẵn
    if ng == 0:
        return {"ok": False, "error": "TSCĐ-CĐKT: không thấy mã 222/225/228"}
    gtcl = round(ng + hm, 9)   # hm âm -> cộng = NG − |hao mòn|
    rec = {"Kỳ": period, "Đơn vị": cong_ty,
           "Loại TS (Nhà cửa VKT/Máy móc TB/PTVT/Phần mềm-QSDĐ…)": "TSCĐ (theo CĐKT)",
           "Nguyên giá (tỷ)": ng, "Giá trị còn lại (tỷ)": gtcl, "Hao mòn lũy kế (tỷ)": round(abs(hm), 9)}
    out = os.path.join(tf.FILLED_DIR, f"TSCD_{period}_{cong_ty or 'NA'}_08_TSCD.xlsx")
    tf.fill("08_TSCD", [rec], out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"),
            "target": "08_TSCD", "nguyen_gia": ng, "gtcl": gtcl}


def _derive_tscd_duan(file_path: str, sheet: str, period: str, cong_ty: str):
    """TSCĐ Khối DỰ ÁN lấy từ sheet 'KH TSCĐ' (BẢNG TRÍCH KHẤU HAO) — layout riêng: header có dòng
    MARKER cột '(1)'..'(22)'. Map: (1)=STT · (4)=tên TS · (7)=Nguyên giá · (14)=Khấu hao lũy kế ·
    (21)=GTCL cuối tháng (=(7)-(14)). CỘNG mọi dòng tài sản (STT là số>0) -> 1 record tổng "TSCĐ
    (theo Biểu khấu hao)" -> 08_TSCD -> report_type TS. Chỉ T05/T06 có sheet này (T01-T04 KHÔNG có
    -> deriver không được gọi/không có gì). Trả not-ok nếu không thấy marker/không có dòng TS."""
    from servers import template_filler as tf
    from servers.common import be_bridge as bb
    wb = bb.fast_load_workbook(file_path, read_only=True, data_only=True)
    try:
        rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()

    def _c(x):
        return str(x).strip() if x not in (None, "") else ""
    # Dòng marker: có cả '(7)' (nguyên giá) và '(14)...' (khấu hao lũy kế).
    mrow = next((i for i, r in enumerate(rows[:25])
                 if any(_c(c) == "(7)" for c in r) and any(_c(c).startswith("(14)") for c in r)), None)
    if mrow is None:
        return {"ok": False, "error": "TSCĐ-DUAN: không thấy dòng marker '(7)'/'(14)' trong 'KH TSCĐ'"}
    hdr = rows[mrow]

    def colof(*pats):
        return next((j for j, c in enumerate(hdr) if any(_c(c) == p or _c(c).startswith(p) for p in pats)), None)
    stt_j, ng_j, hm_j, gtcl_j = colof("(1)"), colof("(7)"), colof("(14)"), colof("(21)")
    if None in (stt_j, ng_j, hm_j, gtcl_j):
        return {"ok": False, "error": "TSCĐ-DUAN: thiếu cột marker (1)/(7)/(14)/(21)"}

    def num(r, j):
        v = r[j] if (j is not None and j < len(r)) else None
        return v if isinstance(v, (int, float)) else 0.0
    ng = hm = gtcl = 0.0
    n = 0
    for r in rows[mrow + 1:]:
        s = r[stt_j] if stt_j < len(r) else None
        if not (isinstance(s, (int, float)) and s > 0):   # chỉ dòng tài sản thật (STT số dương)
            continue
        ng += num(r, ng_j)
        hm += num(r, hm_j)
        gtcl += num(r, gtcl_j)
        n += 1
    if n == 0 or ng == 0:
        return {"ok": False, "error": "TSCĐ-DUAN: không có dòng tài sản (STT trống)"}
    rec = {"Kỳ": period, "Đơn vị": cong_ty,
           "Loại TS (Nhà cửa VKT/Máy móc TB/PTVT/Phần mềm-QSDĐ…)": "TSCĐ (theo Biểu khấu hao)",
           "Nguyên giá (tỷ)": round(ng * 1e-9, 9), "Giá trị còn lại (tỷ)": round(gtcl * 1e-9, 9),
           "Hao mòn lũy kế (tỷ)": round(hm * 1e-9, 9)}
    out = os.path.join(tf.FILLED_DIR, f"TSCD_{period}_{cong_ty or 'NA'}_08_TSCD.xlsx")
    tf.fill("08_TSCD", [rec], out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"), "target": "08_TSCD",
            "nguyen_gia": round(ng * 1e-9, 9), "gtcl": round(gtcl * 1e-9, 9), "n_ts": n}


def _derive_kdvh(file_path: str, sheet: str, period: str, cong_ty: str):
    """TẤT ĐỊNH: sheet '5. Bán xe' (báo cáo quản trị bán hàng) -> 12_KDVH -> report_type KDVH.
    Bảng 1 tầng: Bộ phận(showroom)|Loại xe|SL bán thực tế|DT thực tế|Số tiền KH tháng. Neo =
    Doanh thu (DT thực tế). Consolidated đa-công-ty -> KHÔNG ép cong_ty theo dòng (màn vận hành)."""
    from openpyxl import load_workbook
    from servers import template_filler as tf
    from servers.common import be_bridge as bb
    norm = lambda v: bb.normalize_header(v, True)  # noqa: E731

    wb = bb.fast_load_workbook(file_path, read_only=True, data_only=True)
    try:
        rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()
    hi = next((i for i, r in enumerate(rows[:8]) if any("loai xe" in norm(c) for c in r)), None)
    if hi is None:
        return {"ok": False, "error": "Bán xe: không thấy header 'Loại xe'"}
    hdr = rows[hi]
    def col(*kw, exact=None):
        for j, c in enumerate(hdr):
            n = norm(c)
            if exact is not None and n == exact:
                return j
            if exact is None and all(k in n for k in kw):
                return j
        return None
    bp_i = col(exact="bo phan")
    lx_i = col("loai xe")
    sl_i = col("sl", "thuc te") or col("san luong", "thuc te")
    dt_i = col("dt", "thuc te") or col("doanh thu", "thuc te")
    kh_i = col("so tien ke hoach thang") or col("ke hoach", "thang")
    if lx_i is None or dt_i is None:
        return {"ok": False, "error": "Bán xe: thiếu cột Loại xe / DT thực tế"}

    def num(r, i, scale=True):
        if i is None or i >= len(r):
            return None
        v = bb.parse_num(r[i])
        if v is None:
            return None
        return round(v * 1e-9, 9) if scale else v

    records = []
    for r in rows[hi + 1:]:
        lx = bb.parse_text(r[lx_i]) if lx_i < len(r) else None
        if not lx:
            continue
        dt = num(r, dt_i)
        sl = num(r, sl_i, scale=False)
        if not dt and not sl:
            continue     # dòng loại xe không bán -> bỏ
        rec = {"Ngày/Kỳ": period, "Dòng xe (Limo/VF3/VF5…)": lx,
               "Doanh thu (tỷ)": dt, "Sản lượng (xe)": sl}
        if bp_i is not None and bp_i < len(r):
            rec["Mã Cost center/Showroom ◀ NHẬP"] = bb.parse_text(r[bp_i]) or None
        if kh_i is not None:
            rec["Kế hoạch DT (tỷ)"] = num(r, kh_i)
        records.append(rec)
    if not records:
        return {"ok": False, "error": "Bán xe: không bóc được dòng bán xe nào"}
    out = os.path.join(tf.FILLED_DIR, f"KDVH_{period}_{cong_ty or 'NA'}_12_KDVH.xlsx")
    tf.fill("12_KDVH", records, out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"), "target": "12_KDVH"}


def _derive_dautu(file_path: str, sheet: str, period: str, cong_ty: str):
    """TẤT ĐỊNH: sheet '8. Đầu tư' (DANH SÁCH GIAO DỊCH đầu tư) -> 11_DAUTU -> report_type DTU.
    GOM theo dự án: Tăng=Σ Chi (chi ra đầu tư), Giảm=Σ Thu (thu hồi). KHÔNG có số dư đầu/cuối
    trong nguồn (là sổ giao dịch) -> để trống (thẻ Dư cuối trống là ĐÚNG bản chất nguồn)."""
    from openpyxl import load_workbook
    from servers import template_filler as tf
    from servers.common import be_bridge as bb
    norm = lambda v: bb.normalize_header(v, True)  # noqa: E731

    wb = bb.fast_load_workbook(file_path, read_only=True, data_only=True)
    try:
        rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()
    hi = next((i for i, r in enumerate(rows[:8])
               if any("ma dt" in norm(c) for c in r) and any(norm(c) == "thu" or norm(c) == "chi" for c in r)), None)
    if hi is None:
        return {"ok": False, "error": "Đầu tư: không thấy header Mã ĐT/Thu/Chi"}
    hdr = rows[hi]
    da_i = next((j for j, c in enumerate(hdr) if norm(c) == "ten du an"), None) \
        or next((j for j, c in enumerate(hdr) if norm(c) == "du an"), None)
    thu_i = next((j for j, c in enumerate(hdr) if norm(c) == "thu"), None)
    chi_i = next((j for j, c in enumerate(hdr) if norm(c) == "chi"), None)
    if da_i is None or (thu_i is None and chi_i is None):
        return {"ok": False, "error": "Đầu tư: thiếu cột Dự án / Thu / Chi"}

    agg = {}   # du_an -> [tang(Σchi), giam(Σthu)]
    for r in rows[hi + 1:]:
        da = bb.parse_text(r[da_i]) if da_i < len(r) else None
        if not da or norm(da).startswith(("tong", "cong")):
            continue
        chi = bb.parse_num(r[chi_i]) if (chi_i is not None and chi_i < len(r)) else None
        thu = bb.parse_num(r[thu_i]) if (thu_i is not None and thu_i < len(r)) else None
        a = agg.setdefault(da, [0.0, 0.0])
        a[0] += chi or 0.0
        a[1] += thu or 0.0
    records = []
    for da, (tang, giam) in agg.items():
        if not tang and not giam:
            continue
        records.append({"Kỳ": period, "Dự án/Đơn vị nhận đầu tư (Cao Bằng, Núi Pháo, Tân Thịnh…)": da,
                        "Tăng trong kỳ (tỷ)": round(tang * 1e-9, 9) if tang else None,
                        "Giảm trong kỳ (tỷ)": round(giam * 1e-9, 9) if giam else None})
    if not records:
        return {"ok": False, "error": "Đầu tư: không gom được dòng nào"}
    out = os.path.join(tf.FILLED_DIR, f"DAUTU_{period}_{cong_ty or 'NA'}_11_DAUTU.xlsx")
    tf.fill("11_DAUTU", records, out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"), "target": "11_DAUTU"}


def _derive_cdkt(file_path: str, sheet: str, period: str, cong_ty: str):
    """TẤT ĐỊNH: Bảng cân đối kế toán (CDKT) -> điền 07_TAISAN_NV -> report_type THẬT TSNV/BS.
    Cấu trúc 1 tầng: NỘI DUNG | Mã số | Số đầu năm | cuối kỳ. Neo = CUỐI KỲ (thẻ Tổng tài sản
    đọc dòng 'TỔNG CỘNG TÀI SẢN'/mã 270). Đầu = 'số đầu năm' (với BCTC tháng = đầu kỳ; nhãn rõ
    nên lấy chắc). Điền MỌI dòng khoản mục -> _parse_07 tự chọn dòng section/tổng đúng."""
    from openpyxl import load_workbook
    from servers import template_filler as tf
    from servers.common import be_bridge as bb
    norm = lambda v: bb.normalize_header(v, True)  # noqa: E731

    wb = bb.fast_load_workbook(file_path, read_only=True, data_only=True)
    try:
        rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()
    # header: dòng có CỘT TÊN KHOẢN MỤC + 'ma so'. Cột tên khác nhau theo cty: 'NỘI DUNG' (HT),
    # 'TÀI SẢN' (GA), 'CHỈ TIÊU', 'KHOẢN MỤC' — nới nhận diện theo NGHĨA (không cứng 1 chữ).
    _NAME_KW = ("noi dung", "tai san", "chi tieu", "khoan muc", "tai san / nguon von")
    def _is_name(c):
        n = norm(c)
        return any(n.startswith(k) or k in n for k in _NAME_KW)
    hi = next((i for i, r in enumerate(rows[:15])
               if any(_is_name(c) for c in r) and any("ma so" in norm(c) for c in r)), None)
    if hi is None:
        return {"ok": False, "error": "không thấy header (tên khoản mục + Mã số)"}
    hdr = rows[hi]
    nd_i = next((j for j, c in enumerate(hdr) if _is_name(c)), None)
    ma_i = next((j for j, c in enumerate(hdr) if "ma so" in norm(c)), None)
    # 'Số cuối kỳ'/'cuối kỳ tháng' -> cuoi; 'Số đầu kỳ'/'đầu năm' -> dau. FALLBACK dạng viết tắt
    # 'SỐ C. NĂM'/'SỐ Đ NĂM' (XVP T01) CHỈ khi match chính thất bại -> 0 ảnh hưởng cty khác.
    cuoi_i = next((j for j, c in enumerate(hdr) if "cuoi" in norm(c)), None)
    if cuoi_i is None:   # fallback: 'SỐ C. NĂM' (XVP T01) / 'Kỳ này'/'Kì này' (HTX_XTQ 03-06 gõ dấu ì)
        cuoi_i = next((j for j, c in enumerate(hdr) if "c. nam" in norm(c) or "c.nam" in norm(c) or "ky nay" in norm(c) or "ki nay" in norm(c)), None)
    if cuoi_i is None:   # fallback LAYOUT CỘT-THEO-THÁNG (HT 'BCĐKT hợp nhất' sau khi KT dựng lại: KHÔNG
        # có cột 'cuối kỳ', chỉ 'SỐ ĐẦU NĂM' + 'THÁNG 1'..'THÁNG N' + 'ĐIỀU CHỈNH T{n}' + 'THANG {mm}').
        # Số cuối kỳ = CỘT THÁNG của KỲ (khớp 'thang {mm}'/'t{mm}'/'t0{mm}' như deriver KQKD T-series),
        # SAU cột mã (loại nhãn rác), LOẠI cột 'ĐIỀU CHỈNH T{mm}' (số điều chỉnh, KHÔNG phải số dư). Chỉ
        # kích hoạt khi 2 cách trên fail -> cty có cột 'cuối kỳ' chuẩn giữ nguyên 100% (0 ảnh hưởng).
        try:
            _mm = int(str(period)[-2:])
        except Exception:
            _mm = None
        if _mm is not None:
            _cur = (f"thang {_mm}", f"t{_mm:02d}", f"t{_mm}")
            cuoi_i = next((j for j, c in enumerate(hdr)
                           if j > (ma_i or 0) and "dieu chinh" not in norm(c) and norm(c) in _cur), None)
    dau_i = next((j for j, c in enumerate(hdr) if "dau" in norm(c)), None)
    if dau_i is None:    # fallback: 'SỐ Đ NĂM' / 'Kỳ trước'/'Kì trước'
        dau_i = next((j for j, c in enumerate(hdr) if "d nam" in norm(c) or "d. nam" in norm(c) or "ky truoc" in norm(c) or "ki truoc" in norm(c)), None)
    if nd_i is None or cuoi_i is None:
        return {"ok": False, "error": "thiếu cột tên khoản mục / cuối kỳ"}

    def num(r, i):
        return round(r[i] * 1e-9, 9) if (i is not None and i < len(r) and isinstance(r[i], (int, float))) else None

    # Phân NHÓM cơ cấu theo MÃ SỐ chuẩn CĐKT — tag các dòng SECTION CON (1x0/2x0/3x0/400) để dashboard
    # dựng breakdown (Khối 6/7): TS ngắn hạn=110-150, TS dài hạn=210-260, Nợ phải trả=310/330, Vốn chủ=400.
    # KHÔNG tag dòng TỔNG (100/200/270/300/440) & leaf (111/141…) -> tránh đếm đôi (Σ con = tổng nhóm).
    # Mã không chuẩn (GA/SRVF) -> Nhóm trống -> parser fallback name-match (coarse, tương thích ngược).
    _GRP_BY_MA = {"110": "TS ngắn hạn", "120": "TS ngắn hạn", "130": "TS ngắn hạn", "140": "TS ngắn hạn",
                  "150": "TS ngắn hạn", "210": "TS dài hạn", "220": "TS dài hạn", "230": "TS dài hạn",
                  "240": "TS dài hạn", "250": "TS dài hạn", "260": "TS dài hạn",
                  "310": "Nợ phải trả", "330": "Nợ phải trả", "400": "Vốn chủ"}
    # HTX (Thông tư 71/2024, Mẫu B01-HTX): chart PHẲNG — TÀI SẢN không tách ngắn/dài hạn; mã 200='TỔNG
    # CỘNG TÀI SẢN' (TT200 thì 200='TS dài hạn'), 500='TỔNG NGUỒN VỐN', 300/400=Nợ/Vốn (TỔNG). Map RIÊNG:
    # 110-140→TS ngắn hạn, 150-180→TS dài hạn, 300→Nợ phải trả, 400→Vốn chủ; tổng(200/500)+leaf bỏ trống.
    _GRP_BY_MA_HTX = {"110": "TS ngắn hạn", "120": "TS ngắn hạn", "130": "TS ngắn hạn", "140": "TS ngắn hạn",
                      "150": "TS dài hạn", "160": "TS dài hạn", "170": "TS dài hạn", "180": "TS dài hạn",
                      "300": "Nợ phải trả", "400": "Vốn chủ"}
    _NHOM_COL = "Nhóm (TS ngắn hạn/TS dài hạn/Nợ phải trả/Vốn chủ)"
    # CHỈ tag mã khi CĐKT dùng CHART CHUẨN TT200: mã 310 = 'Nợ ngắn hạn' (SECTION). HTX Xanh dùng chart
    # riêng (300=310+…+380, 310='Phải trả người bán' ITEM) -> mã lệch nghĩa -> KHÔNG tag chuẩn.
    def _ma_at(r):
        return str(bb.parse_text(r[ma_i])).strip() if (ma_i is not None and ma_i < len(r)) else ""
    is_std = any(_ma_at(r) == "310" and "no ngan han" in norm(bb.parse_text(r[nd_i]) or "")
                 for r in rows[hi + 1:] if r and nd_i < len(r))
    # HTX: mã 200 = 'TỔNG CỘNG TÀI SẢN' (chart TT71) -> map HTX riêng (chỉ khi KHÔNG phải chuẩn TT200).
    is_htx = (not is_std) and any(_ma_at(r) == "200" and "tong cong tai san" in norm(bb.parse_text(r[nd_i]) or "")
                                  for r in rows[hi + 1:] if r and nd_i < len(r))
    records = []
    for ri in range(hi + 1, len(rows)):
        r = rows[ri]
        nd = bb.parse_text(r[nd_i]) if nd_i < len(r) else None
        if not nd:
            continue
        # bỏ dòng đánh số cột ('1','2','3') ngay dưới header
        if norm(nd) in ("1", "2", "3"):
            continue
        ma = bb.parse_text(r[ma_i]) if (ma_i is not None and ma_i < len(r)) else None
        cuoi = num(r, cuoi_i)
        dau = num(r, dau_i)
        if cuoi is None and dau is None:
            continue
        _ma_key = str(ma).strip().split(".")[0] if ma else ""
        nhom = (_GRP_BY_MA.get(_ma_key) if is_std
                else _GRP_BY_MA_HTX.get(_ma_key) if is_htx else None)
        rec = {"Kỳ": period, "Đơn vị": cong_ty, _NHOM_COL: nhom,
               "Khoản mục (theo CĐKT)": nd, "Cuối kỳ (tỷ)": cuoi}
        # Mã số CĐKT (vd '222'/'225'/'228'/'231' TSCĐ 4 nhóm) -> tái dùng cột "Ghi chú" có sẵn
        # trong template (KHÔNG thêm cột mới vào golden template -> blast radius thấp, chỉ cần
        # importer đọc thêm field này). Dùng cho breakdown "Cơ cấu TSCĐ theo loại" (Khối 12).
        if _ma_key:
            rec["Ghi chú"] = _ma_key
        if dau is not None:
            rec["Đầu kỳ (tỷ)"] = dau
        records.append(rec)
    if not records:
        return {"ok": False, "error": "không bóc được dòng CĐKT nào"}
    _by_ma = {rec.get("Ghi chú"): rec for rec in records if rec.get("Ghi chú") in ("222", "223", "225", "226", "228", "229", "231", "232")}
    _src = _source_id(file_path).split("::", 1)[0].upper()
    _khoi_l = (_khoi_of(file_path) or "").lower()

    # ── Chart 2 (Book1_.xlsx Khối 12 "Biến động tài sản theo loại") — "Tăng NG" = PS Nợ − PS Có
    # trong kỳ của TK 211/212/213/217 (CĐPS), ghi vào cột "PS tăng trong kỳ (tỷ)" có sẵn trong
    # template (importer đã đọc field này từ trước cho mục dở dang -> KHÔNG cần sửa importer).
    # Áp dụng CHUNG cho mọi đơn vị có sheet CĐPS chuẩn (GA: loại sheet 'TC_...' vì đó là CĐPS HỢP
    # NHẤT group, không phải CĐPS riêng GA). SRVF/TRẠM SẠC/XANH VP override riêng NGAY SAU (đè lên).
    def _find_cdps_sheet(wb):
        for s in wb.sheetnames:
            if (s or "").strip().upper().startswith("TC_"):
                continue
            if "cdps" in norm(s).replace(" ", ""):
                return s
        return None

    def _tang_ng_tk_prefix(cdps_rows, hdr_i, tk_i, psno_i, psco_i, prefixes):
        s = 0.0
        for r in cdps_rows[hdr_i + 1:]:
            tk = str(r[tk_i]).strip() if tk_i < len(r) and r[tk_i] not in (None, "") else ""
            if tk[:3] not in prefixes:
                continue
            no = r[psno_i] if psno_i < len(r) and isinstance(r[psno_i], (int, float)) else 0.0
            co = r[psco_i] if psco_i < len(r) and isinstance(r[psco_i], (int, float)) else 0.0
            s += (no - co)
        return round(s * 1e-9, 9)

    def _is_tk_col(c):
        n = norm(c)
        return "tai khoan" in n or n in ("shtk", "ma tk", "matk")

    def _find_ps_no_co(cdps_rows):
        """Trả (hdr_i, tk_i, psno_i, psco_i) — hdr_i = dòng NGAY TRƯỚC dữ liệu. 2 layout CĐPS quan
        sát được: (A) 1 dòng, nhãn gộp SẴN 'Phát sinh nợ'/'Phát sinh có' (SRVF). (B) 2 dòng — dòng
        trên 1 ô nhãn nhóm 'PHÁT SINH'/'SỐ PHÁT SINH'/'Số phát sinh trong kỳ' (span 2 cột), dòng
        NGAY DƯỚI mới có 'Nợ'/'Có' — cột 'Nợ' LUÔN cùng vị trí cột với nhãn nhóm, 'Có' ở cột kế tiếp
        (đã verify khớp HO/DUAN/ANTAXI/ANKHACHSAN/GLOBALAI, layout+tên cột 'mã TK' khác nhau mỗi nơi
        nhưng quy luật vị trí Nợ/Có này ổn định). Cột mã TK dò rộng theo Ý NGHĨA (không cứng 1 tên)."""
        for i, r in enumerate(cdps_rows[:15]):
            if any(_is_tk_col(c) for c in r) and any("phat sinh no" in norm(c) for c in r):
                tk_i = next((j for j, c in enumerate(r) if _is_tk_col(c)), 0)
                psno_i = next((j for j, c in enumerate(r) if "phat sinh no" in norm(c)), None)
                psco_i = next((j for j, c in enumerate(r) if "phat sinh co" in norm(c)), None)
                if psno_i is not None and psco_i is not None:
                    return i, tk_i, psno_i, psco_i
        for i, r in enumerate(cdps_rows[:15]):
            ps_col = next((j for j, c in enumerate(r) if "phat sinh" in norm(c)), None)
            if ps_col is None or i + 1 >= len(cdps_rows):
                continue
            sub = cdps_rows[i + 1]
            if (ps_col + 1 < len(sub) and norm(sub[ps_col]).startswith("no")
                    and norm(sub[ps_col + 1]).startswith("co")):
                tk_i = next((j for j, c in enumerate(r) if _is_tk_col(c)), None)
                if tk_i is not None:
                    return i + 1, tk_i, ps_col, ps_col + 1
        return None, None, None, None

    _wb_self = bb.fast_load_workbook(file_path, read_only=True, data_only=True)
    try:
        _cdps_sheet = _find_cdps_sheet(_wb_self)
        _cdps_rows = [list(r) for r in _wb_self[_cdps_sheet].iter_rows(values_only=True)] if _cdps_sheet else None
    finally:
        _wb_self.close()
    if _cdps_rows:
        _hdr_i, _tk_i, _psno_i, _psco_i = _find_ps_no_co(_cdps_rows)
        if _hdr_i is not None:
            for _ma, _pre in (("222", {"211"}), ("225", {"212"}), ("228", {"213"}), ("231", {"217"})):
                if _ma in _by_ma:
                    _by_ma[_ma]["PS tăng trong kỳ (tỷ)"] = _tang_ng_tk_prefix(
                        _cdps_rows, _hdr_i, _tk_i, _psno_i, _psco_i, _pre)

    # SRVF: mã 222(hữu hình)/228(vô hình)/223(HM hữu hình)/229(HM vô hình) trên CĐKT xếp theo SỐ
    # HIỆU TK chuẩn (211x->hữu hình, 213x->vô hình) nhưng SRVF dùng 213x làm sổ con "SR_..." cho TS
    # hữu hình Showroom (không phải vô hình thật) -> CĐKT tự phân loại SAI Ý NGHĨA (lệch ~4-5 tỷ giữa
    # 2 nhóm, tổng NG gần như không đổi). Book1_.xlsx chốt override bằng CĐPS theo ĐÚNG mã TK con:
    # NG(Dư nợ cuối kỳ)/Tăng NG(PS Nợ-Có): hữu hình=2111+2112+2113+2118+2131+2132+2134,
    # vô hình=2114+2135. HM(Dư có cuối kỳ): hữu hình=21411+21412+21413+21418+21431+21432+21433+21434,
    # vô hình=21414+21435. 225/231 vẫn giữ nguyên theo CĐKT như mọi đơn vị khác.
    if _src == "SRVF" and _cdps_rows:
        _hdr_i2 = next((i for i, r in enumerate(_cdps_rows[:15])
                        if any(norm(c).startswith("tai khoan") for c in r)
                        and any("no cuoi ky" in norm(c) for c in r)), None)
        if _hdr_i2 is not None:
            _hdr3 = _cdps_rows[_hdr_i2]
            _tk_i2 = next((j for j, c in enumerate(_hdr3) if norm(c).startswith("tai khoan")), 0)
            _cuoi_no_i = next((j for j, c in enumerate(_hdr3) if "no cuoi ky" in norm(c)), None)
            _cuoi_co_i = next((j for j, c in enumerate(_hdr3) if "co cuoi ky" in norm(c)), None)
            _psno_i2 = next((j for j, c in enumerate(_hdr3) if "phat sinh no" in norm(c)), None)
            _psco_i2 = next((j for j, c in enumerate(_hdr3) if "phat sinh co" in norm(c)), None)
            _HH_NG = {"2111", "2112", "2113", "2118", "2131", "2132", "2134"}
            _VH_NG = {"2114", "2135"}
            _HH_HM = {"21411", "21412", "21413", "21418", "21431", "21432", "21433", "21434"}
            _VH_HM = {"21414", "21435"}

            def _sum(codes, col_i, sign=1.0):
                s = 0.0
                for cr in _cdps_rows[max(_hdr_i2, _hdr_i2) + 1:]:
                    tk = str(cr[_tk_i2]).strip() if _tk_i2 < len(cr) and cr[_tk_i2] not in (None, "") else ""
                    if tk in codes and col_i is not None and col_i < len(cr) and isinstance(cr[col_i], (int, float)):
                        s += cr[col_i]
                return round(sign * s * 1e-9, 9)
            if _cuoi_no_i is not None:
                if "222" in _by_ma:
                    _by_ma["222"]["Cuối kỳ (tỷ)"] = _sum(_HH_NG, _cuoi_no_i)
                if "228" in _by_ma:
                    _by_ma["228"]["Cuối kỳ (tỷ)"] = _sum(_VH_NG, _cuoi_no_i)
            if _cuoi_co_i is not None:
                if "223" in _by_ma:
                    _by_ma["223"]["Cuối kỳ (tỷ)"] = _sum(_HH_HM, _cuoi_co_i, sign=-1.0)
                if "229" in _by_ma:
                    _by_ma["229"]["Cuối kỳ (tỷ)"] = _sum(_VH_HM, _cuoi_co_i, sign=-1.0)
            if _psno_i2 is not None and _psco_i2 is not None:
                def _tang(codes):
                    s = 0.0
                    for cr in _cdps_rows[_hdr_i2 + 1:]:
                        tk = str(cr[_tk_i2]).strip() if _tk_i2 < len(cr) and cr[_tk_i2] not in (None, "") else ""
                        if tk in codes:
                            no = cr[_psno_i2] if _psno_i2 < len(cr) and isinstance(cr[_psno_i2], (int, float)) else 0.0
                            co = cr[_psco_i2] if _psco_i2 < len(cr) and isinstance(cr[_psco_i2], (int, float)) else 0.0
                            s += (no - co)
                    return round(s * 1e-9, 9)
                if "222" in _by_ma:
                    _by_ma["222"]["PS tăng trong kỳ (tỷ)"] = _tang(_HH_NG)
                if "228" in _by_ma:
                    _by_ma["228"]["PS tăng trong kỳ (tỷ)"] = _tang(_VH_NG)

    # TRẠM SẠC: CĐKT mã222/223 KHÔNG phản ánh đủ (lệch ~5% so với sổ con "Biểu khấu hao" — vài tài
    # sản mới mua (vd 'PC Intel'/'Laptop Dell' T06) đã vào sổ khấu hao nhưng CHƯA lên CĐKT tháng này).
    # Book1_.xlsx chốt TRẠM SẠC lấy TSCĐ hữu hình từ dòng 'Tổng' của sheet 'Biểu khấu hao' (đầy đủ
    # hơn, có PS tăng trực tiếp) thay vì CĐKT. Override CẢ Chart 1 (NG) lẫn Chart 2 (HM/Tăng NG).
    if "tram sac" in _khoi_l or _src == "TRAMSAC":
        _wb3 = bb.fast_load_workbook(file_path, read_only=True, data_only=True)
        try:
            _bkh_sheet = next((s for s in _wb3.sheetnames if "bieu khau hao" in norm(s)), None)
            _bkh_rows = [list(r) for r in _wb3[_bkh_sheet].iter_rows(values_only=True)] if _bkh_sheet else None
        finally:
            _wb3.close()
        if _bkh_rows:
            _tong_i = next((i for i, r in enumerate(_bkh_rows) if r and norm(r[0] or "") == "tong"), None)
            if _tong_i is not None:
                r = _bkh_rows[_tong_i]

                def _n(j):
                    v = r[j] if j < len(r) else None
                    return v if isinstance(v, (int, float)) else 0.0
                _ng_cuoi, _tang_ng, _hm_cuoi = _n(8) * 1e-9, _n(6) * 1e-9, _n(12) * 1e-9
                if "222" in _by_ma:
                    _by_ma["222"]["Cuối kỳ (tỷ)"] = round(_ng_cuoi, 9)
                    _by_ma["222"]["PS tăng trong kỳ (tỷ)"] = round(_tang_ng, 9)
                if "223" in _by_ma:
                    _by_ma["223"]["Cuối kỳ (tỷ)"] = round(-_hm_cuoi, 9)

    # XANH VP: CĐKT header là 'Ngày cuối kỳ'/'Ngày đầu kỳ' — 'đầu kỳ' ở đây là ĐẦU THÁNG (không phải
    # đầu NĂM như đa số đơn vị khác) -> Tăng NG = cuối − đầu tính THẲNG được, không cần CĐPS (XANH VP
    # không có sheet CĐPS riêng). Chỉ áp dụng khi đã có 'Đầu kỳ (tỷ)' (dau_i tìm được ở trên).
    if _src == "XANHVINHPHUC":
        for _ma in ("222", "225", "228", "231"):
            _rec = _by_ma.get(_ma)
            if _rec and "Đầu kỳ (tỷ)" in _rec:
                _rec["PS tăng trong kỳ (tỷ)"] = round(_rec["Cuối kỳ (tỷ)"] - _rec["Đầu kỳ (tỷ)"], 9)

    out = os.path.join(tf.FILLED_DIR, f"CDKT_{period}_{cong_ty or 'NA'}_07_TAISAN_NV.xlsx")
    tf.fill("07_TAISAN_NV", records, out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"), "target": "07_TAISAN_NV"}


def _derive_tonkho(file_path: str, sheet: str, period: str, cong_ty: str):
    """TẤT ĐỊNH: sổ Nhập-Xuất-Tồn (TONKHO) -> điền 09_TONKHO -> report_type THẬT 'HH'.
    Header 2 tầng: nhóm (Tồn đầu/Nhập/Xuất/Tồn cuối) x tầng dưới (Số lượng/THÀNH TIỀN). Lấy
    cột 'THÀNH TIỀN' theo VAI; neo = TỒN CUỐI (số thẻ Tồn kho đọc). Tồn-đầu mập mờ (vd 'TỒN ĐẦU
    KỲ t3') -> để agent. value_scale=1e-9. Số dư cuối = Σ 'thành tiền' tồn cuối (khớp Tổng cộng)."""
    import re as _re
    from servers import template_filler as tf
    from servers.common import be_bridge as bb
    norm = lambda v: bb.normalize_header(v, True)  # noqa: E731

    wb = bb.fast_load_workbook(file_path, read_only=True, data_only=True)
    try:
        rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()
    # sub_idx: dòng phụ đề có >=2 ô 'thành tiền'
    sub_idx = next((i for i, r in enumerate(rows[:20])
                    if sum(1 for c in r if "thanh tien" in norm(c)) >= 2), None)
    if sub_idx is None or sub_idx == 0:
        return {"ok": False, "error": "không thấy header 2 tầng Số lượng/Thành tiền"}
    group = _forward_fill(rows[sub_idx - 1])
    sub = rows[sub_idx]
    hdr = rows[sub_idx - 1]
    code_i = next((j for j, c in enumerate(hdr) if norm(c).startswith("ma")), None)
    name_i = next((j for j, c in enumerate(hdr) if norm(c).startswith("ten")), None)
    if name_i is None:
        return {"ok": False, "error": "không thấy cột Tên vật tư"}
    fp_month = _guess_month_token(file_path)
    roles = {}   # role -> col index (chỉ cột THÀNH TIỀN)
    dau_cands = []
    for j in range(len(sub)):
        if "thanh tien" not in norm(sub[j]):
            continue
        gt = norm(group[j]) if j < len(group) else ""
        if "cuoi" in gt:
            roles["cuoi"] = j
        elif "nhap" in gt:
            roles["nhap"] = j
        elif "xuat" in gt:
            roles["xuat"] = j
        elif "dau" in gt:
            dau_cands.append((j, gt))
    if "cuoi" not in roles:
        return {"ok": False, "error": "không thấy cột Tồn cuối (Thành tiền)"}
    # đầu kỳ: chắc khi 1 cột, hoặc khớp tháng file; nhiều/mập mờ -> bỏ (để agent)
    dau_partial = False
    if len(dau_cands) == 1:
        roles["dau"] = dau_cands[0][0]
    elif dau_cands:
        # Cột 'đầu kỳ' phải CÓ SỐ THẬT. HT 'nxt tk 156' kèm cột MẪU 'TỒN ĐẦU KỲ t3' (RỖNG mọi kỳ)
        # song song 'TỒN ĐẦU KỲ' (số dư chạy thật). Heuristic khớp-tháng bên dưới chọn nhầm 't3'
        # ĐÚNG kỳ T3 -> đầu kỳ=0 (bug CHỈ tháng 3; T1/T2/T4+ 't3' không khớp nên rơi xuống nhánh
        # 'trơn' đúng). -> ƯU TIÊN cột có dữ liệu; hoà mới xét nhãn tháng/trơn.
        def _colsum(j):
            return sum(abs(r[j]) for r in rows[sub_idx + 1:]
                       if j < len(r) and isinstance(r[j], (int, float)))
        with_data = [j for j, _ in dau_cands if _colsum(j) > 1e3]   # >1000đ = cột có số thật
        if len(with_data) == 1:
            roles["dau"] = with_data[0]
        else:
            cands = [(j, gt) for j, gt in dau_cands if j in with_data] if with_data else dau_cands
            hit = [j for j, gt in cands if fp_month and _label_has_month(gt, fp_month)]
            if len(hit) == 1:
                roles["dau"] = hit[0]
            else:
                # Nhiều cột 'đầu' + không khớp tháng file: loại cột gắn nhãn tháng KHÁC (vd 'TỒN ĐẦU
                # KỲ t3' khi kỳ là T6 = số dư kỳ trước sót lại, thường rỗng). Còn đúng 1 nhãn 'đầu kỳ'
                # TRƠN (không token tháng nào ≠ tháng kỳ) -> đó là đầu kỳ NÀY. Vẫn nhiều -> bỏ (để agent).
                plain = [j for j, gt in cands
                         if not any(_label_has_month(gt, m) for m in range(1, 13) if m != (fp_month or 0))]
                if len(plain) == 1:
                    roles["dau"] = plain[0]
                else:
                    dau_partial = True

    data_start = sub_idx + 1
    while data_start < len(rows):
        nm = norm(rows[data_start][name_i]) if name_i < len(rows[data_start]) else ""
        if nm.startswith("tong cong") or nm == "cong" or nm.startswith("cong "):
            data_start += 1
        else:
            break

    # Mã TK cho CẢ SHEET: nếu tên sheet TỰ KHAI TK (vd "nxt tk 156" -> mọi dòng trong đó VỐN thuộc
    # TK156 theo kế toán, sheet riêng cho từng TK) thì lấy TK đó cho MỌI record — đáng tin hơn cột
    # "Mã" (thường là SỐ KHUNG/SKU vật tư, KHÔNG PHẢI mã TK, vd HT "K503-20373"/"LZZ5EXSA4GN..." —
    # trước đây bị nhét nhầm vào field TK -> hỏng phân loại "Tồn kho theo loại"). Không có TK trong
    # tên sheet (đơn vị khác, sheet đặt tên khác) -> fallback cột "Mã" NHƯNG chỉ nhận nếu giá trị
    # THỰC SỰ giống mã TK (đúng 3 chữ số 151-156), tránh nuốt garbage vào field TK.
    _tk_sheet = next(iter(_re.findall(r"\b(15[1-6])\b", sheet)), None)

    def num(r, i):
        return round(r[i] * 1e-9, 9) if (i is not None and i < len(r) and isinstance(r[i], (int, float))) else None

    records = []
    for ri in range(data_start, len(rows)):
        r = rows[ri]
        name = bb.parse_text(r[name_i]) if name_i < len(r) else None
        if not name or norm(name).startswith(("tong cong", "cong")):
            continue
        rec = {"Kỳ": period, "Đơn vị": cong_ty,
               "Loại HTK (NVL/Vật tư/Hàng hóa…)": name, "Dư cuối kỳ (tỷ)": num(r, roles["cuoi"])}
        if _tk_sheet:
            rec["TK (151-156)"] = _tk_sheet
        elif code_i is not None and code_i < len(r):
            _code_raw = bb.parse_text(r[code_i])
            rec["TK (151-156)"] = _code_raw if (_code_raw and _re.fullmatch(r"15[1-6]", _code_raw.strip())) else None
        if "nhap" in roles:
            rec["Nhập trong kỳ (tỷ)"] = num(r, roles["nhap"])
        if "xuat" in roles:
            rec["Xuất trong kỳ (tỷ)"] = num(r, roles["xuat"])
        if not dau_partial and "dau" in roles:
            rec["Dư đầu kỳ (tỷ)"] = num(r, roles["dau"])
        if rec["Dư cuối kỳ (tỷ)"] not in (None, 0) or rec.get("Nhập trong kỳ (tỷ)") or rec.get("Xuất trong kỳ (tỷ)"):
            records.append(rec)
    if not records:
        return {"ok": False, "error": "không bóc được dòng tồn kho nào"}
    out = os.path.join(tf.FILLED_DIR, f"TONKHO_{period}_{cong_ty or 'NA'}_09_TONKHO.xlsx")
    tf.fill("09_TONKHO", records, out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"),
            "target": "09_TONKHO", "partial_dau_ky": dau_partial}


def _derive_tonkho_cdkt(file_path: str, cdkt_sheet: str, period: str, cong_ty: str):
    """Tồn kho lấy từ CĐKT mã 140 'Hàng tồn kho' (SỐ DƯ) — cho đơn vị KHÔNG có sheet NXT/156/CĐSPS
    riêng (vd Trạm sạc: tồn kho chỉ nằm ở BCĐKT). -> 09_TONKHO -> report_type HH (màn Tồn kho đọc HH).
    Header/cột dùng CHUNG logic với _derive_tscd_cdkt (đã kiểm hoạt động cho các CĐKT này). Trả not-ok
    nếu mã140 = 0/thiếu (vd GA phần mềm) -> KHÔNG tạo dòng tồn kho rỗng. Chạy SAU _derive_tonkho: chỉ
    gọi khi đơn vị không có sheet kho (gate ở cmd_autofill)."""
    from openpyxl import load_workbook
    from servers import template_filler as tf
    from servers.common import be_bridge as bb
    norm = lambda v: bb.normalize_header(v, True)  # noqa: E731
    wb = bb.fast_load_workbook(file_path, read_only=True, data_only=True)
    try:
        rows = [list(r) for r in wb[cdkt_sheet].iter_rows(values_only=True)]
    finally:
        wb.close()
    _NAME_KW = ("noi dung", "tai san", "chi tieu", "khoan muc")

    def _is_name(c):
        n = norm(c)
        return any(n.startswith(k) or k in n for k in _NAME_KW)
    hi = next((i for i, r in enumerate(rows[:15])
               if any(_is_name(c) for c in r) and any("ma so" in norm(c) for c in r)), None)
    if hi is None:
        return {"ok": False, "error": "TONKHO-CĐKT: không thấy header"}
    hdr = rows[hi]
    ma_i = next((j for j, c in enumerate(hdr) if "ma so" in norm(c)), None)
    cuoi_i = next((j for j, c in enumerate(hdr) if "cuoi" in norm(c)), None)
    if cuoi_i is None:
        cuoi_i = next((j for j, c in enumerate(hdr) if "c. nam" in norm(c) or "ky nay" in norm(c)), None)
    dau_i = next((j for j, c in enumerate(hdr) if "dau" in norm(c)), None)
    if dau_i is None:
        dau_i = next((j for j, c in enumerate(hdr) if "d nam" in norm(c) or "d. nam" in norm(c) or "ky truoc" in norm(c)), None)
    if ma_i is None or cuoi_i is None:
        return {"ok": False, "error": "TONKHO-CĐKT: thiếu cột Mã số / cuối kỳ"}

    def num(r, i):
        return round(r[i] * 1e-9, 9) if (i is not None and i < len(r) and isinstance(r[i], (int, float))) else None
    # mã 140 = 'Hàng tồn kho' (tổng nhóm); fallback 141 (leaf) nếu 140 trống.
    row140 = next((r for r in rows[hi + 1:] if ma_i < len(r) and str(r[ma_i]).strip() in ("140", "141")), None)
    cuoi = num(row140, cuoi_i) if row140 else None
    if not cuoi:   # 0/None -> đơn vị KHÔNG có tồn kho (vd GA phần mềm) -> KHÔNG tạo dòng rỗng
        return {"ok": False, "error": "TONKHO-CĐKT: mã140 = 0/thiếu (đơn vị không có tồn kho)"}
    # KHÔNG gán TK cụ thể: mã140 là TỔNG GỘP mọi TK kho con (151-154 theo TT200), không tách được
    # theo loại -> để trống "TK (151-156)" cho payload.tk=None, donut "Tồn kho theo loại" tự gộp
    # đúng vào "Hàng hóa khác" (generic) thay vì gán nhầm cứng vào 1 loại cụ thể như "156" trước đây.
    rec = {"Kỳ": period, "Đơn vị": cong_ty, "Loại HTK (NVL/Vật tư/Hàng hóa…)": "Hàng tồn kho (theo CĐKT)",
           "Dư cuối kỳ (tỷ)": cuoi}
    dau = num(row140, dau_i)
    if dau is not None:
        rec["Dư đầu kỳ (tỷ)"] = dau
    out = os.path.join(tf.FILLED_DIR, f"TONKHO_{period}_{cong_ty or 'NA'}_09_TONKHO.xlsx")
    tf.fill("09_TONKHO", [rec], out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"),
            "target": "09_TONKHO", "cuoi": cuoi}


def _derive_congno(file_path: str, sheet: str, canonical_kind: str, period: str, cong_ty: str):
    """TẤT ĐỊNH (không LLM): sổ tổng hợp công nợ TK131/TK331 -> điền template 05/06 -> report_type
    THẬT (PTHU/PTRA) mà thẻ dashboard đọc sẵn. Dùng _heuristic_tk_mapping để CHỌN cột theo VAI
    (đầu kỳ/phát sinh/cuối kỳ x Nợ/Có), rồi áp đẳng thức kế toán theo loại TK. Số dư CUỐI KỲ luôn
    lấy chắc (thẻ chính); cột đầu kỳ MẬP MỜ (meta.partial) thì BỎ TRỐNG để agent bổ sung sau —
    KHÔNG đoán. value_scale=1e-9 (nguồn VND). Trả {ok, rows, target, partial}."""
    from openpyxl import load_workbook
    from servers import template_filler as tf
    from servers.common import be_bridge as bb

    spec = _CONGNO_IDENTITY.get(canonical_kind)
    if not spec:
        return {"ok": False, "error": f"chưa hỗ trợ identity cho {canonical_kind}"}
    mp, meta = _heuristic_tk_mapping(file_path, sheet, canonical_kind)
    if not mp:
        return {"ok": False, "error": "detector không dựng được mapping (cấu trúc lạ)"}
    role_idx = {c["role"]: c["index"] for c in mp["columns"]}
    code_i, name_i = role_idx.get("entity_code"), role_idx.get("entity_name")
    bal_i = role_idx.get(spec["bal"])
    if bal_i is None or name_i is None:
        return {"ok": False, "error": "thiếu cột cuối kỳ / tên đối tượng"}
    open_i, inc_i, dec_i = role_idx.get(spec["open"]), role_idx.get(spec["inc"]), role_idx.get(spec["dec"])

    wb = bb.fast_load_workbook(file_path, read_only=True, data_only=True)
    try:
        rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()
    norm = lambda v: bb.normalize_header(v, True)  # noqa: E731

    # SỔ TỔNG HỢP GỘP NHIỀU TK (vd HTX 'THCN PHẢI THU' = TK 131+138; 'THCN PHẢI TRẢ' = 331+338): có
    # CỘT 'TÀI KHOẢN' đánh dấu TK từng dòng. Phải LỌC đúng TK gốc (131 phải thu / 331 phải trả), loại
    # 138/338/133/336… (phải thu/phải trả KHÁC — báo cáo tách riêng). Dò cột theo GIÁ TRỊ (dòng dữ liệu
    # là mã TK 3 số cùng LỚP: '13x' phải thu / '33x' phải trả) — KHÔNG theo header vì có tháng gắn nhãn
    # SAI ('Tên ĐT' thay vì 'TÀI KHOẢN', vd HTX_XVP T04). Cột nào nhiều mã TK nhất (>2) = cột tài khoản.
    # Sheet 1-TK (HT: col đầu là mã NCC, không phải mã TK) -> không cột nào đạt -> tk_col=None -> lấy hết.
    _cls = {"TK131": "13", "TK331": "33"}.get(canonical_kind)
    _tk_target = {"TK131": "131", "TK331": "331"}.get(canonical_kind)
    tk_col = None
    if _cls:
        _ds = mp["data_start_row"]
        _sample = rows[_ds:_ds + 150]
        _best = 2
        for _j in range(min(4, max((len(r) for r in _sample), default=0))):
            _cnt = sum(1 for r in _sample if _j < len(r) and r[_j] is not None
                       and (_s := str(r[_j]).strip())[:2] == _cls and _s[:3].isdigit() and len(_s) <= 6)
            if _cnt > _best:
                _best, tk_col = _cnt, _j
        # Header lỗi (HTX_XVP T04: col 'Tên ĐT' TRÙNG cột TK) -> heuristic map tên NHẦM vào cột TK.
        # Nếu name_i CHÍNH LÀ tk_col, chọn lại cột tên THẬT (nhiều chuỗi dài, ≠ mã TK, ≠ tk_col/code_i).
        if tk_col is not None and name_i == tk_col:
            _cand, _bl = None, 0
            for _j in range(min(6, max((len(r) for r in _sample), default=0))):
                if _j in (tk_col, code_i):
                    continue
                _txt = sum(1 for r in _sample if _j < len(r) and isinstance(r[_j], str)
                           and len(r[_j].strip()) > 3 and not r[_j].strip()[:3].isdigit())
                if _txt > _bl:
                    _bl, _cand = _txt, _j
            if _cand is not None:
                name_i = _cand

    def num(r, i):
        if i is None or i >= len(r) or not isinstance(r[i], (int, float)):
            return None
        return round(r[i] * 1e-9, 9)   # VND -> tỷ

    def net(r, pos_role, neg_role):
        """Số dư RÒNG = chiều chính − chiều ngược (guide #30/#36). None nếu THIẾU cả 2 cột (giữ
        tương thích sheet 1 chiều); có 1 cột -> cột kia coi như 0."""
        p, n = num(r, role_idx.get(pos_role)), num(r, role_idx.get(neg_role))
        if p is None and n is None:
            return None
        return round((p or 0) - (n or 0), 9)

    records = []
    for ri in range(mp["data_start_row"], len(rows)):
        r = rows[ri]
        if tk_col is not None and _tk_target:   # lọc đúng TK gốc, loại TK khác (138/338/…)
            _tkv = str(r[tk_col]).strip() if tk_col < len(r) and r[tk_col] is not None else ""
            if _tkv and not _tkv.startswith(_tk_target):
                continue
        name = bb.parse_text(r[name_i]) if name_i < len(r) else None
        if not name:
            continue
        _rawnm = str(name).strip().lower()   # PHÂN BIỆT DẤU: 'cộng'(tổng) ≠ 'công'(công ty)
        # Dòng TỔNG ('Cộng'/'Tổng cộng') KHÔNG có Mã đối tượng. NCC/KH thật 'Cộng tác viên …' CÓ mã
        # (vd NCC111908) -> KHÔNG được bỏ. Trước đây startswith('cộng ') bắt nhầm -> mất 0,675 tỷ
        # (HT T05 'Cộng tác viên cty long sơn'). Chỉ nới nhánh 'cộng <x>' khi thiếu mã; 'cộng'/'tổng
        # cộng' vẫn luôn bỏ (không bao giờ là đối tượng thật) -> giữ nguyên fix đếm đôi 131 cũ.
        _code_val = bb.parse_text(r[code_i]) if (code_i is not None and code_i < len(r)) else None
        # Dòng TỔNG có thể nằm ở CỘT MÃ (vd HTX_XVP T04 header lỗi -> tên map nhầm cột TK, mã='Tổng':
        # tổng-con mỗi TK bị đếm ĐÔI với chi tiết). Bắt 'cộng'/'tổng'/'tổng cộng' ở CẢ tên lẫn mã.
        _code_low = str(_code_val).strip().lower() if _code_val else ""
        if (_rawnm in ("cộng", "tổng") or _rawnm.startswith("tổng cộng")
                or _code_low in ("cộng", "tổng") or _code_low.startswith("tổng cộng")
                or (_rawnm.startswith("cộng ") and not _code_val)):
            continue   # bỏ dòng tổng (an toàn kép, ngoài data_start đã bỏ)
        rec = {"Kỳ": period, "Đơn vị": cong_ty, spec["ten_col"]: name,
               "Dư cuối kỳ (tỷ)": net(r, spec["bal"], spec["bal_opp"])}   # RÒNG (Có−Nợ / Nợ−Có)
        if code_i is not None and code_i < len(r):
            rec["Mã đối tượng"] = bb.parse_text(r[code_i]) or None
        if not meta.get("partial") and open_i is not None:   # đầu kỳ chỉ điền khi CHẮC
            rec["Dư đầu kỳ (tỷ)"] = net(r, spec["open"], spec["open_opp"])   # RÒNG
        if inc_i is not None:
            rec[spec["inc_col"]] = num(r, inc_i)
        if dec_i is not None:
            rec[spec["dec_col"]] = num(r, dec_i)
        if any(v not in (None, "") for k, v in rec.items() if k not in ("Kỳ", "Đơn vị")):
            records.append(rec)
    if not records:
        return {"ok": False, "error": "không bóc được dòng công nợ nào"}

    out = os.path.join(tf.FILLED_DIR, f"{canonical_kind}_{period}_{cong_ty or 'NA'}_{spec['target']}.xlsx")
    tf.fill(spec["target"], records, out)
    imp = tf.import_filled(out, cong_ty=cong_ty, khoi=_khoi_of(file_path), source_file=_source_id(file_path))
    return {"ok": bool(imp.get("rows_imported")), "rows": imp.get("rows_imported"),
            "target": spec["target"], "report_type": spec["target"],
            "partial_dau_ky": bool(meta.get("partial")), "out": out}


def _analyst_propose_prompt(file_path: str, sheet: str) -> str:
    return f"""Bạn là subagent analyst (SKILL: Dashboard_Agent/agents/analyst/SKILL.md).

Phân tích file Excel và TRẢ VỀ DUY NHẤT 1 khối ```json SheetMapping cho sheet '{sheet}'.

File: {file_path}
Sheet: {sheet}

Quy trình BẮT BUỘC (dùng MCP tools, đọc guideline qua glossary — KHÔNG hard-code tên cột):
1. sheet_profile(file_path='{file_path}', sheet='{sheet}') — đọc row_sample/col_sample, phát hiện header 1-2 dòng, orientation row_major/column_major.
2. report_spec_search(sheet='{sheet}') và report_spec_search(canonical_kind=...) nếu sheet_profile/canonical đoán được loại (TK131, TK331, KQKD, CDKT, TSCD...).
3. glossary_lookup(term=<tên sheet hoặc canonical_kind>) — khớp kpi_glossary.json (nguồn từ guideline.xlsx): xem nguon_du_lieu, chi_tieu, công thức để biết sheet phục vụ chỉ số nào.
4. Đối chiếu display_contract.json — field/màn hình FE nào cần dữ liệu từ sheet này.
5. Dựng SheetMapping: orientation, data_start_row (0-based), columns/entities+row_roles, target_report_type (tiền tố GEN_), canonical_kind, matched_kpi_ids.

Chỉ số 0-based. role: entity_code|entity_name|label|skip|<do_luong>.
KHÔNG giải thích ngoài khối JSON."""


# 1 SESSION CO DINH cho analyst ingest (khop _INGEST_SID ben DashBoard_AI/source_bridge.py). Truyen
# --session-id TUONG MINH: bo trong thi Gateway de 'gateway-fallback-<uuid>' MOI luot (khong gom
# duoc). RESET = xoa file '<id>.*' trong thu muc sessions cua agent analyst.
_INGEST_SID = "dashboard-ingest"


def _reset_ingest_session():
    """Xoa session ingest co dinh (reset moi file — yeu cau user). Best-effort, loi thi bo qua."""
    try:
        subprocess.run(["docker", "exec", "openclaw", "sh", "-c",
                        f'rm -f "$HOME"/.openclaw/agents/analyst/sessions/{_INGEST_SID}.*'],
                       capture_output=True, text=True, timeout=20)
    except Exception:  # noqa: BLE001
        pass


def _propose_mapping(file_path: str, sheet: str, timeout: int = 240):
    """Goi analyst agent (LLM) qua docker exec -> tra (mapping|None, reply_text).

    DUNG CHUNG 1 session co dinh (--session-id _INGEST_SID) — yeu cau 2026-07-13 "1 session thoi"
    khi chay OpenClaw + 9router. Cac luot propose giờ TUAN TU (xem cmd_autobatch) nen khong ghi
    dong thoi cung 1 session. Kien thuc hoc duoc (mapping) van duoc memory.report_spec_save() ghi
    ra dia ngay trong tool, doc lap voi lich su session."""
    prompt = _analyst_propose_prompt(file_path, sheet)
    cmd = ["docker", "exec", "openclaw", "openclaw", "agent", "--agent", "analyst",
           "--session-id", _INGEST_SID, "--json", "--timeout", str(timeout), "-m", prompt]
    try:
        p = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout + 30)
    except subprocess.TimeoutExpired:
        return None, "analyst agent timeout"
    reply = _extract_agent_text(p.stdout)
    mapping = _extract_json_block(reply) if reply else None
    return mapping, reply


def cmd_propose(args):
    from servers.common import canonical

    ck = canonical.guess_canonical_kind(args.sheet)
    scope_guess = canonical.guess_scope(os.path.basename(args.file))
    if ck in _HEURISTIC_CANONICAL_KINDS:
        mapping, meta = _heuristic_tk_mapping(args.file, args.sheet, ck)
        if mapping:
            _out({"ok": True, "proposed_mapping": mapping, "agent_text": None, "source": "heuristic",
                  "scope_guess": scope_guess, "basis_guess": meta.get("basis_guess")})
            return
        # heuristic không khớp cấu trúc mong đợi -> fallback LLM như bình thường bên dưới.
    mapping, reply = _propose_mapping(args.file, args.sheet, args.timeout)
    _out({"ok": bool(mapping), "proposed_mapping": mapping, "agent_text": reply, "source": "llm",
          "scope_guess": scope_guess, "basis_guess": None})


def cmd_autobatch(args):
    """Nap TAT CA sheet wired trong 1 lan: mapping da hoc dung luon; chua hoc thi propose (neu
    --propose) qua analyst LLM. Moi sheet: generic_import_execute. Tra ket qua tung sheet.
    LUU Y: sheet chua hoc phu thuoc chat luong propose (model yeu co the fail -> status=need_manual)."""
    from servers import ingest_server as ing
    from servers.common import memory

    prof = ing.sheet_profile(args.file)
    all_specs = memory.report_spec_search()
    gl_cache: dict = {}

    # PHASE 1 (PHAN TICH) - xac dinh mapping cho tung sheet wired. Sheet da hoc/heuristic co
    # mapping ngay; sheet chua hoc -> can goi analyst LLM (_propose_mapping). Tu 2026-07-13 cac
    # lan propose DUNG CHUNG session main (het doc lap) -> mac dinh TUAN TU (ANALYST_MAX_PARALLEL
    # = 1; chi tang lai khi chac chan OpenClaw khoa session an toan). Buoc GHI (PHASE 2) van tuan tu.
    plan = []  # thu tu theo sheet, moi entry: {name, ck, mapping, source} hoac needs_propose=True
    for s in prof.get("sheets", []):
        name = s.get("sheet")
        ck = s.get("canonical_kind_guess")
        learned = _learned_sheet_mappings(name, ck, all_specs)
        kpi_hints = _kpi_hints_from_guideline(name, ck, gl_cache)
        if _dashboard_relevance(learned, kpi_hints, ck) != "wired":
            continue
        mapping = learned[0] if learned else None
        source = "learned" if mapping else None
        if mapping is None and ck in _HEURISTIC_CANONICAL_KINDS:
            mapping, _heur_meta = _heuristic_tk_mapping(args.file, name, ck)
            source = "heuristic" if mapping else None
        needs_propose = mapping is None and args.propose
        plan.append({"name": name, "ck": ck, "mapping": mapping,
                     "source": source, "needs_propose": needs_propose})

    # PROPOSE TUAN TU (2026-07-13): _propose_mapping DUNG CHUNG 1 session co dinh (_INGEST_SID).
    # Chay SONG SONG se ban NHIEU turn openclaw DONG THOI vao CUNG 1 file session (.jsonl + .lock)
    # -> tuan tu, chap nhan lau hon (yeu cau user "1 session thoi"). RESET session TRUOC file nay.
    _to_propose = [x for x in plan if x["needs_propose"]]
    if _to_propose:
        _reset_ingest_session()
    for e in _to_propose:
        try:
            mapping, reply = _propose_mapping(args.file, e["name"], args.timeout)
        except Exception as exc:  # noqa: BLE001 - 1 propose hong khong lam sap ca batch
            mapping, reply = None, f"{type(exc).__name__}: {exc}"
        e["mapping"] = mapping
        e["source"] = "proposed" if mapping else None
        if mapping is None:
            e["fail_reason"] = (reply or "")[:300]

    # PHASE 2 (GHI) - TUAN TU tung sheet mot, tranh dung do dataset chung ky giua cac cong ty.
    results = []
    for e in plan:
        name, ck, mapping, source = e["name"], e["ck"], e["mapping"], e["source"]
        if mapping is None:
            results.append({"sheet": name, "canonical_kind": ck, "status": "need_manual",
                            "reason": e.get("fail_reason") or "chua co mapping (bat --propose hoac lam thu cong)"})
            continue
        try:
            r = ing.generic_import_execute(args.file, sheet=name, mapping=mapping,
                                           period=args.period, cong_ty=args.cong_ty, dry_run=args.dry_run)
            results.append({"sheet": name, "canonical_kind": ck,
                            "status": "preview" if args.dry_run else ("duplicate" if r.get("skipped_duplicate") else "imported"),
                            "mapping_source": source, "row_count": r.get("row_count"),
                            "target_report_type": r.get("target_report_type"), "dataset_id": r.get("dataset_id")})
        except Exception as e:
            results.append({"sheet": name, "canonical_kind": ck, "status": "error", "error": f"{type(e).__name__}: {e}"})
    imported = sum(1 for r in results if r["status"] in ("imported", "preview"))
    _out({"ok": True, "file_name": prof.get("file_name"), "wired_total": len(results),
          "imported": imported, "results": results})


def cmd_execute(args):
    from servers import ingest_server as ing

    with open(args.mapping_file, encoding="utf-8") as fh:
        mapping = json.load(fh)
    r = ing.generic_import_execute(
        args.file, sheet=args.sheet, mapping=mapping, period=args.period,
        cong_ty=args.cong_ty, dry_run=args.dry_run,
    )
    r["ok"] = True
    _out(r)


def cmd_confirm(args):
    """Lưu 1 quyết định phân loại (scope/basis/màn hình) do NGƯỜI xác nhận cho 1 canonical_kind.
    Dùng lại cho MỌI file/công ty cùng loại báo cáo sau này (agent_bridge /import/agent/confirm).
    KHÔNG tự ghi số vào KPI nào — chỉ là metadata phân loại, xem servers/common/memory.py."""
    from servers.common import memory

    rec = memory.binding_save(args.canonical_kind, scope=args.scope, basis=args.basis,
                              target_screen=args.target_screen, chi_tieu=args.chi_tieu,
                              kpi_id=args.kpi_id)
    _out({"ok": True, "binding": rec})


def _extract_agent_text(stdout: str):
    try:
        env = json.loads(stdout)
    except json.JSONDecodeError:
        return None
    # tim finalAssistantVisibleText o bat ky do sau nao
    stack = [env]
    while stack:
        node = stack.pop()
        if isinstance(node, dict):
            for k in ("finalAssistantVisibleText", "finalAssistantRawText"):
                if isinstance(node.get(k), str) and node[k].strip():
                    return node[k]
            stack.extend(node.values())
        elif isinstance(node, list):
            stack.extend(node)
    return None


def _extract_json_block(text: str):
    """Trích object JSON từ reply của analyst. Thử theo thứ tự ưu tiên:
    1. TỪNG code fence ```json (KHÔNG chỉ fence đầu — model hay in 2 khối: giải thích + đáp án),
       parse NGUYÊN VĂN nội dung fence (không regex lazy `\\{.*?\\}` như trước — lazy backtrack
       có thể bắt OBJECT CON của JSON lồng rồi trả mapping SAI âm thầm).
    2. Khối { ... } lớn nhất ngoài fence.
    Trả None nếu không khối nào parse được (caller giữ reply để chẩn đoán)."""
    import re
    candidates = []
    for m in re.finditer(r"```(?:json)?\s*\n?(.*?)```", text, re.DOTALL):
        body = m.group(1).strip()
        if body.startswith("{"):
            candidates.append(body)
    i, j = text.find("{"), text.rfind("}")
    if 0 <= i < j:
        candidates.append(text[i:j + 1])
    parsed = []
    for raw in candidates:
        try:
            obj = json.loads(raw)
        except json.JSONDecodeError:
            continue
        if isinstance(obj, dict):
            parsed.append(obj)
    if not parsed:
        return None
    # Nhiều khối parse được -> lấy khối NHIỀU KEY NHẤT (mapping đầy đủ thắng ví dụ/mảnh nhỏ).
    return max(parsed, key=lambda o: len(o))


def cmd_autofill(args):
    """LOOP TẤT ĐỊNH điền template vàng — thay 1-prompt-nhồi-tất-cả. Mỗi sheet rơi vào ĐÚNG 1
    'bucket' (đảm bảo phủ, không sót): filled_learned (đã học mapping -> điền ngay, KHÔNG LLM) |
    need_llm (đã route ra sheet đích nhưng chưa học -> analyst đề xuất) | need_human (chưa nhận
    diện được loại) | skip (metadata) | empty. dry_run=true chỉ preview, không ghi DB."""
    from servers import ingest_server as ing
    from servers import template_filler as tf
    from servers.common import memory

    routes = ing.sheet_routes(args.file).get("routes", [])
    with open(args.file, "rb") as fh:
        data = fh.read()
    fname = os.path.basename(args.file)
    period = args.period or tf.guess_period(fname)
    headers = tf._all_sheet_headers(data)   # mở workbook 1 LẦN (file nặng load chậm)

    # File THU CHI (có sheet SD TIỀN): dòng tiền 03_DONGTIEN/03B/04_VAY được các EXTRACTOR
    # TẤT ĐỊNH (extract_thuchi/sodu_tien/vay ở mục 'DẪN XUẤT' bên dưới) nạp trọn 1 lần, KHÔNG
    # cần LLM. Nếu vẫn để các sheet 'BC THU CHI_T*_*' vào need_llm thì LLM xử lý LẶP từng công
    # ty (lặp context 5 lần) VÀ ghi đè/đếm đôi lên dữ liệu extractor -> loại chúng ra khỏi need_llm.
    _has_sd = any("sd tien" in tf._norm(s) or "so du tien" in tf._norm(s) for s in headers)
    _derived_targets = {"03_DONGTIEN", "03B_SODU_TIEN", "04_VAY"} if _has_sd else set()

    # LUẬT PER-CÔNG-TY TỪ GUIDE (knowledge/<MA>.yaml -> don_vi) — ADDITIVE: guide không khai báo
    # thì cả 2 biến dưới giữ hành vi cũ 100%.
    #  - sheets_theo_y_nghia.p_and_l = WHITELIST sheet P&L. Vd HT: file T01-04 có sheet 'kqkd'
    #    (TT200, số lũy kế ĐÓNG BĂNG — HT.yaml CẤM đọc) đứng TRƯỚC 'kqkdQT' (T-series, ĐÚNG) ->
    #    extractor TT200 parse được sheet cấm và claim nhầm 01_HQKD, sheet đúng bị skip_dup.
    #    So khớp: bỏ dấu + lower + bỏ MỌI khoảng trắng rồi so BẰNG (không substring — 'kqkd'
    #    không được ăn theo 'kqkdqt').
    #  - che_do_phan_tich: thuan_llm = file của công ty này THUẦN AI (mọi sheet routed -> LLM,
    #    tắt extractor tất định) — trừ 3 target tiền (_derived_targets, file thu chi) vẫn tất định.
    import re as _re
    # Guide dùng để GATE (whitelist p_and_l / thuan_llm) phải resolve công ty theo TÊN FILE ưu tiên
    # (prefer_file_name=True): raw --cong-ty có thể là mã HỢP LỆ NHƯNG SAI do receiver copy sidecar
    # (tiền lệ 2026-07-09). Nếu để raw thắng, 1 file non-HT bị gắn nhầm company='HT' sẽ bị áp chế độ
    # HT (skip 'kqkd' + thuần LLM) -> mất số câm; ngược lại file HT bị gắn nhầm 'GA' sẽ mất chế độ HT.
    # Tên file 'B<khối>.<mã cty>.' do nghiệp vụ đặt -> đáng tin hơn để quyết chế độ pipeline.
    _kqkd_ok, _cdkt_ok, _guide_thuan_llm = None, None, False
    try:
        from servers.common import contract as _contract
        from servers.common.extraction import load_guide as _load_guide
        _guide_co = _contract.resolve_company(args.cong_ty, fname, prefer_file_name=True)
        _dv = (((_load_guide(_guide_co, fname) or {}).get("content") or {}).get("don_vi") or {})
        _styn = _dv.get("sheets_theo_y_nghia") or {}
        _pl = _styn.get("p_and_l")
        if isinstance(_pl, list) and _pl:
            _kqkd_ok = {_re.sub(r"\s+", "", tf._norm(x)) for x in _pl}
        #  - sheets_theo_y_nghia.balance_sheet = WHITELIST sheet Bảng cân đối (CĐKT). Vd HO: file có
        #    'CĐKT' (BCTC RIÊNG hội sở, ĐVT VND — ĐÚNG) VÀ 'TC_CĐKT' (hợp nhất "Thịnh Cường Group",
        #    ĐVT triệu đồng, kỳ khác) — CẢ HAI route canonical_kind=CDKT. Không whitelist thì
        #    _derive_cdkt nạp cả hai vào 07_TAISAN_NV; import_filled delete-scope (source_file+cong_ty)
        #    khiến sheet chạy SAU (TC_CĐKT) GHI ĐÈ sheet đúng -> BS phình 49->298 dòng, sai đơn vị/kỳ,
        #    mất cân đối 270=300+400. Cùng cơ chế whitelist p_and_l (khớp bỏ dấu + bỏ khoảng trắng).
        _bs = _styn.get("balance_sheet")
        if isinstance(_bs, list) and _bs:
            _cdkt_ok = {_re.sub(r"\s+", "", tf._norm(x)) for x in _bs}
        _guide_thuan_llm = str(_dv.get("che_do_phan_tich") or "").strip().lower() == "thuan_llm"
    except Exception:
        _kqkd_ok, _cdkt_ok, _guide_thuan_llm = None, None, False   # guide lỗi/thiếu -> KHÔNG đổi hành vi cũ

    def _env_on(name):
        # Cờ env kiểu ON/OFF: '0'/'false'/'off'/'no'/'' -> TẮT (bool(str) coi '0' là True -> sai).
        return str(os.environ.get(name, "")).strip().lower() in ("1", "true", "yes", "on")

    _env_force = _env_on("ANALYST_FORCE_LLM")
    _force_llm = _env_force or _guide_thuan_llm
    _force_reason = ("FORCE_LLM: ép gửi LLM" if _env_force
                     else "guide che_do_phan_tich=thuan_llm: thuần AI phân tích")
    # WHITELIST P&L có mặt trong FILE? (có sheet KQKD khớp p_and_l). Dùng để làm whitelist THỰC SỰ
    # ƯU TIÊN khi file có NHIỀU sheet KQKD cùng có T-code (HT: 'kqkd tổng hợp nhất' HỢP NHẤT +
    # 'XETAITC_KQKD'/'XETAIHT_KQKD' PHÁP NHÂN CON). Trước đây lưới an toàn T-code giữ CẢ 3 -> 01_HQKD
    # phụ thuộc THỨ TỰ sheet (kqkd_done lấy sheet đầu); đảo thứ tự sẽ lấy nhầm 1 pháp nhân con (13.98
    # thay hợp nhất 88.74 tỷ). Khi CÓ sheet khớp whitelist -> sheet KQKD KHÔNG khớp bị bỏ (kể cả có
    # T-code). Lưới T-code CHỈ còn tác dụng khi KHÔNG sheet nào khớp whitelist (phòng đổi tên sheet).
    _kqkd_wl_present = bool(_kqkd_ok) and any(
        r.get("canonical_kind") == "KQKD" and r.get("status") == "routed"
        and _re.sub(r"\s+", "", tf._norm(r.get("sheet"))) in _kqkd_ok for r in routes)

    ledger = []
    congno_todo = []   # sheet TK131/TK331 -> _derive_congno (tất định, chạy ở mục DẪN XUẤT)
    tonkho_todo = []   # sheet TONKHO (nhập-xuất-tồn) -> _derive_tonkho
    cdkt_todo = []     # sheet CDKT (bảng cân đối) -> _derive_cdkt
    thue_todo = []     # sheet CDPS (cân đối số PS) -> _derive_thue (lọc TK 133/333)
    tscd_todo = []     # sheet TSCD (khấu hao) -> _derive_tscd
    kdvh_todo = []     # sheet '5. Bán xe' (KDVH) -> _derive_kdvh
    dautu_todo = []    # sheet '8. Đầu tư' (DAUTU) -> _derive_dautu
    derived = []       # kết quả extractor tất định (KQKD chạy inline; còn lại ở mục DẪN XUẤT dưới)
    kqkd_done = False   # 01_HQKD đã có nguồn -> sheet KQKD thứ 2 (kqkdQT/quý) BỎ, tránh đè/đôi
    # sổ chi tiết giao dịch bỏ qua theo NGUỒN (vd SRVF) — xem _SKIP_DETAIL_SHEETS_BY_SOURCE
    _detail_skip_src = _SKIP_DETAIL_SHEETS_BY_SOURCE.get(_source_id(args.file).split("::", 1)[0]) or set()
    for r in routes:
        sheet, status, target = r["sheet"], r["status"], r.get("target_sheet")
        ck = r.get("canonical_kind")
        if status in ("skip_metadata", "empty"):
            ledger.append({"sheet": sheet, "bucket": status.replace("skip_metadata", "skip"),
                           "target_sheet": None})
            continue
        # GUARD SHEET SIÊU RỘNG: bỏ qua sheet nhiều cột bệnh lý (vd CĐPS của HO = 16350 cột) TRƯỚC
        # mọi nhánh route -> không deriver nào materialize nổi ~5.6M ô (thue + tonkho_cdps = 2 lượt
        # ×25s). Không mất dữ liệu thật: các sheet này là phantom-column, extractor luôn fail ("không
        # dò được header"). Rộng chính đáng KHÔNG bao giờ >1000 cột. Log 'skip_wide' để không nuốt thầm.
        _ncols = max((len(_hr) for _hr in (headers.get(sheet) or [])), default=0)
        if _ncols > _MAX_SHEET_COLS:
            ledger.append({"sheet": sheet, "bucket": "skip_wide", "target_sheet": None,
                           "canonical_kind": ck, "cols": _ncols,
                           "reason": f"sheet {_ncols} cột (>{_MAX_SHEET_COLS}) — phantom/rác, bỏ qua tránh nghẽn autofill"})
            continue
        # BỎ QUA SỔ NHẬT KÝ CHUNG (NKC) theo TÊN — sổ cái thô, không phải nguồn chỉ tiêu; đụng vào chỉ
        # tốn thời gian (vd Trạm sạc). Chuẩn hoá bỏ khoảng trắng để khớp 'Sổ nhật ký chung'/'NKC'/'SoNKC'.
        if _re.sub(r"\s+", "", tf._norm(sheet)) in {_re.sub(r"\s+", "", s) for s in _SKIP_SHEET_NAMES}:
            ledger.append({"sheet": sheet, "bucket": "skip_name", "target_sheet": None,
                           "canonical_kind": ck, "reason": "Sổ nhật ký chung (NKC) — sổ cái thô, bỏ qua tăng tốc"})
            continue
        # GA: BỎ QUA sheet 'Data' (dữ liệu thô backing công thức, KHÔNG phải nguồn chỉ tiêu) -> tăng tốc.
        if tf._norm(sheet) == "data" and "công nghệ" in (_khoi_of(args.file) or "").lower():
            ledger.append({"sheet": sheet, "bucket": "skip_name", "target_sheet": None,
                           "canonical_kind": ck, "reason": "GA: sheet 'Data' thô, bỏ qua tăng tốc"})
            continue
        if _detail_skip_src and tf._norm(sheet).replace(" ", "") in _detail_skip_src:
            ledger.append({"sheet": sheet, "bucket": "skip_name", "target_sheet": None,
                           "canonical_kind": ck,
                           "reason": "sổ chi tiết giao dịch (không phải nguồn chỉ tiêu) — bỏ qua tăng tốc"})
            continue
        if status == "routed" and ck == "KQKD" and _kqkd_ok is not None \
                and _re.sub(r"\s+", "", tf._norm(sheet)) not in _kqkd_ok \
                and (_kqkd_wl_present or not _sheet_has_tcodes(args.file, sheet)):
            # Guide công ty CẤM đọc sheet KQKD này (không nằm trong whitelist p_and_l) -> bỏ hẳn:
            # không extractor, không LLM (đặt TRƯỚC nhánh FORCE_LLM vì là luật nghiệp vụ, áp mọi
            # chế độ). target=None để mục DẪN XUẤT 02_CHIPHI không chọn nhầm sheet này làm nguồn.
            # LƯỚI AN TOÀN: sheet P&L T-series (có mã T100/T200/T300) LUÔN được nhận dù tên không nằm
            # trong whitelist — phòng HT đổi tên sheet tháng mới (đã từng: kqkdQT->kqkd hưng thịnh).
            # Sheet 'kqkd' TT200 đóng băng KHÔNG có T-code nên vẫn bị chặn.
            ledger.append({"sheet": sheet, "bucket": "skip_guide", "target_sheet": None,
                           "canonical_kind": ck,
                           "reason": "guide công ty loại sheet KQKD này (sheets_theo_y_nghia.p_and_l)"})
            continue
        if status == "routed" and target and _force_llm and target not in _derived_targets:
            # CHẾ ĐỘ THUẦN LLM (env ANALYST_FORCE_LLM toàn cục, hoặc guide công ty khai
            # don_vi.che_do_phan_tich: thuan_llm — vd HT: BCTC riêng thuần AI phân tích).
            # Bỏ extractor tất định, đẩy sheet cho LLM (BE xử 1-sheet-1-lượt). GIỮ các override
            # target của nhánh tất định (CDPS->10_THUE, KDVH->12_KDVH, DAUTU->11_DAUTU) — nếu để
            # target gốc của route, cđps và cđkt cùng trỏ 07_TAISAN_NV: sheet tới sau bị bỏ
            # ("target đã có") còn 10_THUE không ai điền. 3 target TIỀN (_derived_targets — file
            # thu chi) KHÔNG bao giờ đưa LLM: giữ extractor tất định, 2 nguồn không đè nhau.
            ledger.append({"sheet": sheet, "bucket": "need_llm",
                           "target_sheet": {"CDPS": "10_THUE", "KDVH": "12_KDVH",
                                            "DAUTU": "11_DAUTU"}.get(ck, target),
                           "canonical_kind": ck, "reason": _force_reason})
            continue
        if status == "routed" and ck in _CONGNO_IDENTITY:
            # SRVF: công nợ số dư LẤY TỔNG HỢP TỪ CĐKT (mã 131 phải thu / 311 phải trả) qua
            # derive_srvf_cdps — KHÔNG nạp sheet '131'/'331' per-NCC (chứa dòng 'Tổng cộng' tự
            # tham chiếu -> cộng đôi: 342,58×2=685,15 thay vì 370,22). Chốt user + guide C Điệp
            # 2026-07-19 (#30/#36 = CĐKT). Bỏ qua đây để nhánh SRVF (derive_srvf_cdps) là nguồn duy nhất.
            if _source_id(args.file).split("::", 1)[0].upper() == "SRVF":
                ledger.append({"sheet": sheet, "bucket": "skip_dup", "target_sheet": target,
                               "canonical_kind": ck,
                               "reason": "SRVF: công nợ số dư từ CĐKT (derive_srvf_cdps), bỏ sheet per-NCC tránh đếm đôi"})
                continue
            # Sổ công nợ TK131/331: detector tất định + identity kế toán -> report_type THẬT
            # (PTHU/PTRA). KHÔNG cho LLM đụng (model hay dựng sai mapping / quên value_scale).
            congno_todo.append((sheet, ck))
            ledger.append({"sheet": sheet, "bucket": "derived", "target_sheet": target,
                           "canonical_kind": ck, "reason": "công nợ nạp bằng detector tất định"})
            continue
        if status == "routed" and ck == "TONKHO":
            tonkho_todo.append(sheet)
            ledger.append({"sheet": sheet, "bucket": "derived", "target_sheet": target,
                           "canonical_kind": ck, "reason": "tồn kho nạp bằng detector tất định"})
            continue
        if status == "routed" and ck == "CDKT":
            if _cdkt_ok is not None and _re.sub(r"\s+", "", tf._norm(sheet)) not in _cdkt_ok:
                # Guide loại sheet CĐKT này khỏi BS (không nằm trong sheets_theo_y_nghia.balance_sheet)
                # -> KHÔNG nạp 07_TAISAN_NV. Vd HO: chặn 'TC_CĐKT' (hợp nhất group) đè lên 'CĐKT' riêng.
                # target=None để không nhánh dẫn xuất nào chọn nhầm sheet này làm nguồn BS.
                ledger.append({"sheet": sheet, "bucket": "skip_guide", "target_sheet": None,
                               "canonical_kind": ck,
                               "reason": "guide loại sheet CĐKT này (sheets_theo_y_nghia.balance_sheet)"})
                continue
            cdkt_todo.append(sheet)
            ledger.append({"sheet": sheet, "bucket": "derived", "target_sheet": target,
                           "canonical_kind": ck, "reason": "CĐKT nạp bằng detector tất định"})
            continue
        if status == "routed" and ck == "CDPS":
            # GA: 'TC_CDPS' là CĐPS HỢP NHẤT Thịnh Cường Group (ĐVT triệu đồng) — KHÔNG phải CĐPS riêng
            # GA. Thuế GA lấy từ CĐKT riêng (mã 152/313) ở block GLOBALAI (_derive_thue_cdkt). Bỏ để không
            # route nhầm sang thuế/tồn kho (deriver fail layout -> trước flip need_llm -> analyst đã chết).
            if _guide_co == "GA" and (sheet or "").strip().upper().startswith("TC_"):
                ledger.append({"sheet": sheet, "bucket": "skip_guide", "target_sheet": None,
                               "canonical_kind": ck,
                               "reason": "GA: TC_CDPS là CĐPS hợp nhất Thịnh Cường Group -> thuế lấy từ CĐKT riêng"})
                continue
            thue_todo.append(sheet)   # CĐSPS -> thuế (TK 133/333) tất định
            ledger.append({"sheet": sheet, "bucket": "derived", "target_sheet": "10_THUE",
                           "canonical_kind": ck, "reason": "thuế (TK133/333) từ CĐSPS tất định"})
            continue
        if status == "routed" and ck == "TSCD":
            tscd_todo.append(sheet)
            ledger.append({"sheet": sheet, "bucket": "derived", "target_sheet": target,
                           "canonical_kind": ck, "reason": "TSCĐ (khấu hao) tất định"})
            continue
        if status == "routed" and ck == "KDVH":
            kdvh_todo.append(sheet)
            ledger.append({"sheet": sheet, "bucket": "derived", "target_sheet": "12_KDVH",
                           "canonical_kind": ck, "reason": "bán xe (KDVH) tất định"})
            continue
        if status == "routed" and ck == "DAUTU":
            dautu_todo.append(sheet)
            ledger.append({"sheet": sheet, "bucket": "derived", "target_sheet": "11_DAUTU",
                           "canonical_kind": ck, "reason": "đầu tư (gom giao dịch) tất định"})
            continue
        if status == "routed" and ck == "KQKD":
            if kqkd_done:   # 01_HQKD đã có nguồn tất định từ sheet KQKD trước -> BỎ (tránh đè/đôi)
                ledger.append({"sheet": sheet, "bucket": "skip_dup", "target_sheet": "01_HQKD",
                               "canonical_kind": ck, "reason": "01_HQKD đã có nguồn KQKD tất định khác"})
                continue
            if not args.dry_run:
                # THỬ extractor TT200; CHỈ claim khi THÀNH CÔNG (không claim trước -> sheet KQKD tốt
                # không bị sheet đầu 'HQKD HỢP NHẤT'/mã-riêng chặn). Fail -> fall through need_llm
                # (đánh dấu ck=KQKD; hậu-loop reconcile để không nhân đôi 01_HQKD).
                try:
                    _rk = _derive_kqkd(args.file, sheet, period, args.cong_ty)
                except Exception as ex:  # noqa: BLE001
                    _rk = {"ok": False, "error": str(ex)[:150]}
                if _rk.get("ok"):
                    kqkd_done = True
                    derived.append({"kind": "01_HQKD", "sheet": sheet, "ok": True,
                                    "rows": _rk.get("rows"), "value_col": _rk.get("value_col_header"),
                                    "chiphi": _rk.get("chiphi")})
                    ledger.append({"sheet": sheet, "bucket": "derived", "target_sheet": "01_HQKD",
                                   "canonical_kind": ck, "reason": "KQKD TT200 tất định"})
                    continue
            # dry_run HOẶC KQKD không phải TT200 -> need_llm (LLM đọc extraction_guide). KHÔNG rơi
            # xuống fill_spec: 1 sheet KQKD (vd 'HQKD HỢP NHẤT') có fill_spec CŨ (từ lần LLM sai) sẽ
            # replay cột rỗng -> đè lên bản tất định. Hậu-loop reconcile theo kqkd_done để chỉ 1 nguồn.
            ledger.append({"sheet": sheet, "bucket": "need_llm", "target_sheet": "01_HQKD",
                           "canonical_kind": ck, "reason": "KQKD layout lạ -> LLM theo extraction_guide"})
            continue
        if status == "routed" and ck == "LCTT":
            # DÒNG TIỀN của MỌI pháp nhân (GỒM HT) lấy từ Báo cáo tiền tập đoàn — sheet 'BC THU CHI_T*_
            # <CTY>' qua extract_thuchi, KHÔNG từ lctt. CẬP NHẬT 2026-07-21 (guide TC-SRVF dòng tiền):
            # HT "Tiền vào/ra trong kỳ" = Mã I/II sheet 'BC THU CHI_T*_HUNGTHINH' (gồm 'Giảm gốc vay')
            # → T02 thu 5,85 / chi 8,51 khớp báo cáo; trước lấy lctt (_derive_lctt_ht) ra 5,83/6,93
            # (thiếu gốc vay + lệch nền 20tr). GIỮ hàm _derive_lctt_ht (không xoá) nhưng KHÔNG gọi nữa
            # để HT lấy dòng tiền CÙNG NGUỒN với các cty khác, tránh nạp ĐÔI 03_DONGTIEN khác nguồn.
            ledger.append({"sheet": sheet, "bucket": "skip_role", "target_sheet": None, "canonical_kind": ck,
                           "reason": "LCTT: dòng tiền lấy từ file thu chi (BC THU CHI) cho MỌI pháp nhân"})
            continue
        if status == "routed" and target in _derived_targets:
            # Đã có extractor tất định phủ target này (chạy ở mục DẪN XUẤT) -> KHÔNG cho LLM đụng.
            ledger.append({"sheet": sheet, "bucket": "derived", "target_sheet": target,
                           "reason": "dòng tiền nạp bằng extractor tất định (không cần LLM)"})
            continue
        if status == "skip_role":
            # Sheet thuộc vai file KHÁC (ma trận Book1) — KHÔNG đưa vào need_llm (target=None,
            # gọi LLM sẽ vô nghĩa/hỏng). Vd lctt trong file BCTC: dòng tiền chỉ lấy từ file thu chi.
            ledger.append({"sheet": sheet, "bucket": "skip_role", "target_sheet": None,
                           "reason": "sheet thuộc loại file khác (xem ma trận nguồn dữ liệu)"})
            continue
        if status == "unknown":
            ledger.append({"sheet": sheet, "bucket": "need_human", "target_sheet": None,
                           "reason": "chưa nhận diện được loại báo cáo"})
            continue
        head = headers.get(sheet) or []
        hr = tf._header_row_of(head)
        cols = head[hr] if hr < len(head) else []
        spec = memory.fill_spec_find(memory.source_fingerprint(sheet, cols))
        if not spec:
            ledger.append({"sheet": sheet, "bucket": "need_llm", "target_sheet": target,
                           "canonical_kind": r.get("canonical_kind"), "route_via": r.get("route_via"),
                           "reason": "chưa học mapping — cần analyst đề xuất"})
            continue
        res = tf.fill_from_source(data, sheet, spec["target_sheet"], spec["mapping"],
                                  period=period, cong_ty=args.cong_ty, file_name=fname,
                                  dry_run=args.dry_run, auto_import=not args.dry_run, learn=False,
                                  value_scale=spec.get("value_scale", 1.0), constants=spec.get("constants"),
                                  rename_rows=spec.get("rename_rows"), source_path=args.file)
        ledger.append({"sheet": sheet, "bucket": "filled_learned", "target_sheet": spec["target_sheet"],
                       "dry_run": args.dry_run, "row_count": res.get("row_count"),
                       "rows_written": res.get("rows_written"),
                       "imported": (res.get("import") or {}).get("ok")})
    # DẪN XUẤT TẤT ĐỊNH (chỉ tiêu computed KHÔNG có sheet nguồn riêng — bóc từ chính file này):
    #  - 02_CHIPHI (cơ cấu chi phí): từ dòng A3xx của sheet KQKD (nếu file có KQKD -> 01_HQKD).
    #  - 03B_SODU_TIEN (số dư tiền): từ sheet 'SD TIỀN' (nếu file là Báo cáo Thu Chi).
    # Chạy SAU khi nạp xong (report_type CHIPHI/SDT riêng biệt -> KHÔNG đè dữ liệu autofill).
    # (KQKD đã chạy inline trong loop; 'derived' khởi tạo trước loop.)
    if not args.dry_run:
        for sheet, ck in congno_todo:   # TK131->PTHU, TK331->PTRA (tất định, report_type THẬT)
            try:
                rc = _derive_congno(args.file, sheet, ck, period, args.cong_ty)
                derived.append({"kind": _CONGNO_IDENTITY[ck]["target"], "sheet": sheet,
                                "ok": rc.get("ok"), "rows": rc.get("rows"),
                                "partial_dau_ky": rc.get("partial_dau_ky"), "error": rc.get("error")})
            except Exception as ex:  # noqa: BLE001
                derived.append({"kind": ck, "sheet": sheet, "ok": False, "error": str(ex)[:150]})
        tonkho_done = False             # 09_TONKHO đã có nguồn tất định -> sheet tồn kho khác BỎ
        for sheet in tonkho_todo:       # TONKHO -> 09_TONKHO -> HH (tất định)
            # File có thể có 2 sheet cùng TONKHO (HT: 'nxt tk 156' NGUỒN CHÍNH + 'bảng tính tồn kho
            # hợp nhất' phụ). source_file theo FILE (không theo sheet) + import delete-scope -> sheet
            # sau ĐÈ sheet trước; sheet phụ fail còn bị flip need_llm -> LLM nạp lại 09_TONKHO (đôi).
            # Lấy sheet ĐẦU derive THÀNH CÔNG (thứ tự workbook: 'nxt tk 156' đúng spec đứng trước).
            if tonkho_done:
                derived.append({"kind": "09_TONKHO", "sheet": sheet, "ok": True, "skipped_dup": True,
                                "reason": "09_TONKHO đã có nguồn tồn kho tất định khác"})
                continue
            try:
                rc = _derive_tonkho(args.file, sheet, period, args.cong_ty)
                if rc.get("ok"):
                    tonkho_done = True
                derived.append({"kind": "09_TONKHO", "sheet": sheet, "ok": rc.get("ok"),
                                "rows": rc.get("rows"), "partial_dau_ky": rc.get("partial_dau_ky"),
                                "error": rc.get("error")})
            except Exception as ex:  # noqa: BLE001
                derived.append({"kind": "09_TONKHO", "sheet": sheet, "ok": False, "error": str(ex)[:150]})
        for sheet in cdkt_todo:         # CDKT -> 07_TAISAN_NV -> TSNV/BS (tất định)
            try:
                rc = _derive_cdkt(args.file, sheet, period, args.cong_ty)
                derived.append({"kind": "07_TAISAN_NV", "sheet": sheet, "ok": rc.get("ok"),
                                "rows": rc.get("rows"), "error": rc.get("error")})
            except Exception as ex:  # noqa: BLE001
                derived.append({"kind": "07_TAISAN_NV", "sheet": sheet, "ok": False, "error": str(ex)[:150]})
        for sheet in thue_todo:         # CDPS -> 10_THUE -> THUE (TK 133/333)
            try:
                rc = _derive_thue(args.file, sheet, period, args.cong_ty)
                derived.append({"kind": "10_THUE", "sheet": sheet, "ok": rc.get("ok"),
                                "rows": rc.get("rows"), "error": rc.get("error")})
            except Exception as ex:  # noqa: BLE001
                derived.append({"kind": "10_THUE", "sheet": sheet, "ok": False, "error": str(ex)[:150]})
        # KHÔNG có sheet NXT riêng -> tồn kho TK 151–156 từ CĐPS (DUAN NVL 152, An KS hàng hóa 156).
        # Loại Trạm sạc/HO: 2 khối này lấy tồn kho từ CĐKT mã140 (nhánh dưới) -> tránh trùng nguồn.
        if not tonkho_todo and not any(k in (_khoi_of(args.file) or "").lower() for k in ("trạm sạc", "hỗ trợ")):
            for sheet in thue_todo:
                try:
                    rc = _derive_tonkho_cdps(args.file, sheet, period, args.cong_ty)
                    if rc.get("ok"):
                        derived.append({"kind": "09_TONKHO", "sheet": sheet, "ok": True,
                                        "rows": rc.get("rows"), "error": None, "via": "CĐPS TK152"})
                except Exception:  # noqa: BLE001
                    pass
        for sheet in tscd_todo:         # TSCD (khấu hao) -> 08_TSCD -> TS
            try:
                # DUAN: 'KH TSCĐ' layout marker '(n)' riêng -> deriver DUAN; đơn vị khác -> generic.
                if "dự án" in (_khoi_of(args.file) or "").lower():
                    rc = _derive_tscd_duan(args.file, sheet, period, args.cong_ty)
                else:
                    rc = _derive_tscd(args.file, sheet, period, args.cong_ty)
                derived.append({"kind": "08_TSCD", "sheet": sheet, "ok": rc.get("ok"),
                                "rows": rc.get("rows"), "via": rc.get("n_ts") and "KH TSCĐ", "error": rc.get("error")})
            except Exception as ex:  # noqa: BLE001
                derived.append({"kind": "08_TSCD", "sheet": sheet, "ok": False, "error": str(ex)[:150]})
        # SRVF (Chi nhánh VinFast Showroom) — FORMAT RIÊNG, deriver chuyên biệt (cong_ty ÉP 'TC',
        # alias folder SRVF->TC). CĐKT chuẩn TT200 -> _derive_cdkt; P&L ở sheet 'T{mm}BC' (KHÔNG phải
        # 'KQKD') -> _derive_kqkd_srvf; công nợ/thuế/tồn kho (TK152/153/154/156) từ CĐPS 1-tầng -> derive_srvf_cdps.
        if _source_id(args.file).split("::", 1)[0].upper() == "SRVF":
            from servers.common import be_bridge as bb
            _mm = period.split("-")[-1]
            try:
                _wb = bb.fast_load_workbook(args.file, read_only=True, data_only=True)
                _shs = _wb.sheetnames
                _ck = next((s for s in _shs if "CĐKT" in s or s.upper().replace(" ", "") == "CDKT"), None)
                _bc = next((s for s in _shs if s.upper().replace(" ", "") == f"T{_mm}BC"), None)
                _bcrows = [list(r) for r in _wb[_bc].iter_rows(values_only=True)] if _bc else None
                _wb.close()
                if _ck:                                    # CĐKT -> TSNV (số dư, công nợ, tồn kho, tiền)
                    rc = _derive_cdkt(args.file, _ck, period, "TC")
                    derived.append({"kind": "07_TAISAN_NV", "sheet": _ck, "ok": rc.get("ok"),
                                    "rows": rc.get("rows"), "via": "SRVF"})
                    rc2 = _derive_tscd_cdkt(args.file, _ck, period, "TC")   # TSCĐ (mã 222/223) -> TS
                    derived.append({"kind": "08_TSCD", "sheet": _ck, "ok": rc2.get("ok"),
                                    "rows": rc2.get("rows"), "via": "SRVF CĐKT mã220"})
                if _bcrows:                                # T{mm}BC -> 01_HQKD + 02_CHIPHI + PNLT (A-series)
                    rk = _derive_kqkd_srvf(_bcrows, period, "TC", args.file)
                    derived.append({"kind": "01_HQKD", "sheet": _bc, "ok": bool(rk and rk.get("ok")),
                                    "rows": (rk or {}).get("rows"), "via": "SRVF T{mm}BC"})
                from derive_srvf_cdps import extract as _srvf_cdps   # công nợ + thuế + tồn kho xe
                rd = _srvf_cdps(args.file, period, "TC")
                derived.append({"kind": "SRVF CĐPS (PTHU/PTRA/THUE/HH)", "ok": rd.get("ok"),
                                "rows": {k: rd[k] for k in ("pthu", "ptra", "thue", "tonkho") if k in rd},
                                "error": rd.get("error")})
            except Exception as ex:  # noqa: BLE001
                derived.append({"kind": "SRVF", "ok": False, "error": str(ex)[:150]})
        # GA (Global AI, Khối Công nghệ) — công nợ sheet 'PThu'/'PTra' layout "BÁO CÁO CÔNG NỢ" 2-tầng
        # (deriver chung TK131/331 KHÔNG nhận -> trước để LLM, chỉ nạp T01/T02). Deriver chuyên biệt
        # derive_ga_congno: DƯ CUỐI Nợ->PTHU, Có->PTRA, per-KH. Đánh dấu 2 sheet 'derived' để LLM
        # không nạp đè (idempotent theo source_file). KQKD/CĐKT GA theo deriver chuẩn ở trên; THUẾ GA
        # lấy từ CĐKT riêng (mã 152/313) — xem _derive_thue_cdkt ngay dưới (KHÔNG dùng TC_CDPS hợp nhất).
        if _source_id(args.file).split("::", 1)[0].upper() == "GLOBALAI":
            try:
                from derive_ga_congno import extract as _ga_congno
                rg = _ga_congno(args.file, period, args.cong_ty or "GA")
                derived.append({"kind": "GA công nợ (PTHU/PTRA)", "ok": rg.get("ok"),
                                "rows": {k: rg[k] for k in ("pthu", "ptra", "pthu_tong", "ptra_tong") if k in rg}})
                if rg.get("ok"):
                    for e in ledger:
                        if (e.get("sheet") or "").strip().lower() in ("pthu", "ptra"):
                            e["bucket"] = "derived"
                            e["reason"] = "công nợ GA nạp bằng derive_ga_congno (tất định)"
            except Exception as ex:  # noqa: BLE001
                derived.append({"kind": "GA công nợ", "ok": False, "error": str(ex)[:150]})
            # THUẾ GA: TK133/333 lấy từ CĐKT RIÊNG (mã 152 'Thuế GTGT được khấu trừ' -> Phải thu; mã 313
            # 'Thuế và các khoản phải nộp' -> Phải nộp). KHÔNG dùng 'TC_CDPS' (CĐPS hợp nhất Thịnh Cường
            # Group, triệu đồng): _derive_thue fail 'header 2 tầng' -> trước flip need_llm -> analyst, mà
            # analyst container KHÔNG mount received_reports nên luôn rỗng. Chạy sau _derive_cdkt; đè bản
            # THUE cũ (LLM dump nhầm 50 dòng không nhãn) theo source_file.
            _cdkt_sheet = next((e["sheet"] for e in ledger
                                if e.get("target_sheet") == "07_TAISAN_NV" and (e.get("sheet") or "")), None)
            if _cdkt_sheet:
                try:
                    rtx = _derive_thue_cdkt(args.file, _cdkt_sheet, period, args.cong_ty or "GA")
                    derived.append({"kind": "10_THUE", "sheet": _cdkt_sheet, "ok": rtx.get("ok"),
                                    "rows": rtx.get("rows"), "error": rtx.get("error"), "via": "CĐKT mã152/313"})
                except Exception as ex:  # noqa: BLE001
                    derived.append({"kind": "10_THUE", "ok": False, "error": str(ex)[:150]})
        # An Taxi / An KS (nhóm AAG) — P&L QUẢN TRỊ theo spec 50 chỉ tiêu: An Taxi đọc sheet 'BCQT PT.'
        # (mã 100 DT bán hàng='Doanh thu HH,DV', 120 DT thuần, 130 giá vốn, 150/170/182/192 chi phí,
        # 200 LNTT, 220 LNST); An KS đọc sheet 'BCQT' (Mục I/II/III). Deriver riêng -> 01_HQKD (DThu
        # thuần / Tổng chi phí GỘP #8 / LNTT) + PNLT (DT bán hàng gross, giá vốn, LN gộp, LNST) +
        # 02_CHIPHI. Idempotent theo source_file -> ĐÈ bản KQKD generic (tránh dt_total gộp thu nhập
        # ngoài vào DThu, và lấy đúng #8 = tổng dòng chi phí GỘP). Chạy SAU vòng _derive_kqkd.
        _an_folder = _source_id(args.file).split("::", 1)[0].upper()
        if _an_folder == "ANTAXI":
            try:
                rk = _derive_kqkd_antaxi(args.file, period, args.cong_ty or "AAG")
                derived.append({"kind": "01_HQKD", "ok": bool(rk and rk.get("ok")), "via": "An Taxi BCQT PT",
                                "rows": (rk or {}).get("rows"), "chiphi": (rk or {}).get("chiphi"),
                                "error": (rk or {}).get("error")})
            except Exception as ex:  # noqa: BLE001
                derived.append({"kind": "An Taxi BCQT PT", "ok": False, "error": str(ex)[:150]})
        if _an_folder == "ANKHACHSAN":
            try:
                rk = _derive_kqkd_ankhachsan(args.file, period, args.cong_ty or "AAG")
                derived.append({"kind": "01_HQKD", "ok": bool(rk and rk.get("ok")), "via": "An KS BCQT",
                                "rows": (rk or {}).get("rows"), "chiphi": (rk or {}).get("chiphi"),
                                "error": (rk or {}).get("error")})
            except Exception as ex:  # noqa: BLE001
                derived.append({"kind": "An KS BCQT", "ok": False, "error": str(ex)[:150]})
        # TSCĐ TỪ CĐKT SỐ CUỐI KỲ (mã 222+225+228 NG, 223+226+229 hao mòn = mã 220, TIE Tổng TS
        # mã 270). Dùng cho đơn vị KHÔNG có 'Biểu khấu hao' đọc được -> CĐKT là nguồn TSCĐ:
        #   • An Taxi: 'Biểu khấu hao' over-state (75.45≠64.9) -> ghi ĐÈ (chạy SAU _derive_tscd).
        #   • HO: sheet KHTS layout lạ (_derive_tscd fail); GA: không có sheet khấu hao -> THÊM mới.
        # import_filled delete-scope (report_type=TS,source_file,cong_ty) ghi 1 record 'TSCĐ (theo
        # CĐKT)'. Đơn vị KHÁC giữ breakdown theo loại TS -> KHÔNG đụng.
        _kn = (_khoi_of(args.file) or "").lower()
        if any(k in _kn for k in ("an taxi", "an ks", "hỗ trợ", "công nghệ", "trạm sạc", "xe tải",
                                  "dự án", "taxi xanh")):   # taxi xanh (XVP): Biểu KH 972 over-state 2× CĐKT 475
            # Xe tải (HT): 'Biểu khấu hao' (khấu hao-phân bổ) UNDER-state hao mòn (T01: 41tr thay vì
            # CĐKT mã223 1.077 tỷ) -> lấy TSCĐ từ CĐKT (nguyên giá mã222, hao mòn mã223) khớp BC.
            for sheet in cdkt_todo:
                try:
                    rc = _derive_tscd_cdkt(args.file, sheet, period, args.cong_ty)
                    derived.append({"kind": "08_TSCD", "sheet": sheet, "ok": rc.get("ok"),
                                    "rows": rc.get("rows"), "via": "CĐKT mã220 (ghi đè khấu hao)",
                                    "error": rc.get("error")})
                except Exception as ex:  # noqa: BLE001
                    derived.append({"kind": "08_TSCD", "sheet": sheet, "ok": False,
                                    "via": "CĐKT mã220", "error": str(ex)[:150]})
        # TỒN KHO TỪ CĐKT mã 140 -> 09_TONKHO -> HH. CHỈ khi đơn vị KHÔNG có sheet kho riêng
        # (tonkho_todo rỗng) -> tránh đè _derive_tonkho. Trạm sạc (mã140 ~3.7 tỷ) + HO (mã140 ~11.9 tỷ
        # tĩnh) chỉ có tồn kho ở CĐKT nên màn Tồn kho trước đây trống. Deriver tự bỏ nếu mã140=0
        # (GA phần mềm -> không tạo dòng). Mở gate cho các khối này (self-exclude khi mã140=0).
        if not tonkho_todo and any(k in _kn for k in ("trạm sạc", "hỗ trợ")):
            for sheet in cdkt_todo:
                try:
                    rc = _derive_tonkho_cdkt(args.file, sheet, period, args.cong_ty)
                    derived.append({"kind": "09_TONKHO", "sheet": sheet, "ok": rc.get("ok"),
                                    "rows": rc.get("rows"), "via": "CĐKT mã140", "error": rc.get("error")})
                except Exception as ex:  # noqa: BLE001
                    derived.append({"kind": "09_TONKHO", "sheet": sheet, "ok": False,
                                    "via": "CĐKT mã140", "error": str(ex)[:150]})
        for sheet in kdvh_todo:         # '5. Bán xe' -> 12_KDVH -> KDVH
            try:
                rc = _derive_kdvh(args.file, sheet, period, args.cong_ty)
                derived.append({"kind": "12_KDVH", "sheet": sheet, "ok": rc.get("ok"),
                                "rows": rc.get("rows"), "error": rc.get("error")})
            except Exception as ex:  # noqa: BLE001
                derived.append({"kind": "12_KDVH", "sheet": sheet, "ok": False, "error": str(ex)[:150]})
        for sheet in dautu_todo:        # '8. Đầu tư' -> 11_DAUTU -> DTU
            try:
                rc = _derive_dautu(args.file, sheet, period, args.cong_ty)
                derived.append({"kind": "11_DAUTU", "sheet": sheet, "ok": rc.get("ok"),
                                "rows": rc.get("rows"), "error": rc.get("error")})
            except Exception as ex:  # noqa: BLE001
                derived.append({"kind": "11_DAUTU", "sheet": sheet, "ok": False, "error": str(ex)[:150]})
        kqkd_entry = next((e for e in ledger if e.get("target_sheet") == "01_HQKD"), None)
        kqkd = kqkd_entry["sheet"] if kqkd_entry else None
        # Deriver P&L layout-riêng (SRVF/XDV/HO) ĐÃ emit 02_CHIPHI cùng lúc (đúng cột tháng, đủ nhóm
        # khớp tổng) -> KHÔNG chạy extract_chiphi post-loop, tránh ĐÈ bản đầy đủ bằng bản a_pat TT200
        # thiếu nhóm (A325/A500) hoặc sai layout (B-series/dòng-tên). Chỉ áp khi deriver báo ok.
        _dcp = next((d.get("chiphi") for d in derived
                     if d.get("kind") == "01_HQKD" and (d.get("chiphi") or {}).get("ok")), None)
        if _dcp:
            derived.append({"kind": "02_CHIPHI", "ok": True, "rows": _dcp.get("rows"),
                            "via": "deriver P&L (emit cùng 01_HQKD, khớp tổng)"})
        elif kqkd:
            chiphi_ok = False
            try:
                from extract_chiphi import extract as _derive_chiphi
                r = _derive_chiphi(args.file, kqkd, period, args.cong_ty)
                chiphi_ok = bool(r.get("ok"))
                derived.append({"kind": "02_CHIPHI", "ok": r.get("ok"), "rows": r.get("rows"),
                                "error": r.get("error")})
            except Exception as ex:  # noqa: BLE001
                derived.append({"kind": "02_CHIPHI", "ok": False, "error": str(ex)[:120]})
            # Extractor tất định (dòng A3xx chuẩn TT200) KHÔNG khớp VÀ P&L đã đi LLM (thuan_llm/
            # force_llm/layout lạ) -> ĐỪNG để 02_CHIPHI mất câm (không sheet nguồn riêng nên không
            # tự lọt vào need_llm qua sheet_routes): thêm 1 ENTRY LLM lấy CHÍNH sheet P&L đó — cùng
            # session ingest đã mở cho file này (BE lặp thêm 1 lượt, KHÔNG mở session mới). Không
            # gắn canonical_kind="KQKD" (tránh bị vòng reconcile skip_dup KQKD ở dưới bắt nhầm).
            if not chiphi_ok and kqkd_entry.get("bucket") == "need_llm":
                ledger.append({"sheet": kqkd, "bucket": "need_llm", "target_sheet": "02_CHIPHI",
                              "reason": "02_CHIPHI: extractor TT200 không khớp layout, P&L đã đi "
                                        "LLM -> LLM điền luôn cơ cấu chi phí từ cùng sheet"})
        try:  # KHỐI TIỀN GỘP 1 MỐI (SDT + VAY + THUCHI) từ Báo cáo tiền tập đoàn -> extract_tien.
            from servers.common import be_bridge as bb
            _wb = bb.fast_load_workbook(args.file, read_only=True, data_only=True)
            has_sd = any("SD TI" in s.upper() or "SỐ DƯ TI" in s.upper() for s in _wb.sheetnames)
            _wb.close()
            if has_sd:
                from extract_tien import extract as _derive_tien
                rt = _derive_tien(args.file, period, args.cong_ty)
                for _kind, _key in (("03B_SODU_TIEN", "sdt"), ("04_VAY", "vay"), ("03_DONGTIEN", "thuchi")):
                    derived.append({"kind": _kind, "ok": rt.get(_key) is not None,
                                    "rows": rt.get(_key), "error": rt.get("sd_error")})
                # #3 (2026-07-23) WIRE PIPELINE VAY: sau extract_tien (base VAY kỳ này), PATCH NGAY từ
                # báo cáo ngân hàng (THUCHI/baocaonganhang) — nợ đến hạn + lãi vay + tách kỳ hạn — SCOPE
                # theo period (only_period) để KHÔNG tách lại kỳ khác đã split (mangle). Hết bẫy mất-patch
                # khi re-ingest báo cáo tiền tập đoàn. Chỉ khi có VAY + KHÔNG dry-run. Thứ tự bắt buộc:
                # den_han -> lãi -> tách kỳ hạn (kyhan đọc dòng đã có den_han/lãi rồi redistribute).
                if rt.get("vay") and not args.dry_run:
                    for _mod, _lbl in (("extract_no_den_han", "nợ đến hạn"),
                                       ("extract_lai_vay", "lãi vay"),
                                       ("extract_vay_kyhan", "tách kỳ hạn")):
                        try:
                            _rv = __import__(_mod).apply(only_period=period, commit=True)
                            derived.append({"kind": f"VAY+ {_lbl}", "ok": True, "via": _mod, "matched": _rv})
                        except Exception as _ex:  # noqa: BLE001
                            derived.append({"kind": f"VAY+ {_lbl}", "ok": False, "error": str(_ex)[:150]})
        except Exception as ex:  # noqa: BLE001
            derived.append({"kind": "khối tiền (extract_tien)", "ok": False, "error": str(ex)[:120]})

        # HẬU-XỬ LÝ: detector tất định FAIL cho sheet nào (layout riêng như GA PThu/CĐKT) ->
        # ĐỪNG để mất âm thầm: chuyển bucket 'derived' của sheet đó sang 'need_llm' -> BE gọi LLM
        # (đọc extraction_guide/GA.yaml mà điền theo hướng dẫn công ty). Map target -> tên sheet.
        # target đã có ÍT NHẤT 1 sheet nạp THÀNH CÔNG -> KHÔNG flip sheet fail cùng target sang LLM
        # (tránh nạp đôi: vd HT 2 sheet TONKHO — 'nxt tk 156' ok, 'bảng tính tồn kho' fail; nếu flip
        # sheet fail thì LLM nạp lại 09_TONKHO chồng lên bản tất định).
        _ok_targets = {d.get("kind") for d in derived if d.get("ok")}
        _fail_targets = {}   # target_sheet -> sheet nguồn (từ derived fail)
        for d in derived:
            if d.get("ok") is False and d.get("sheet") and d.get("kind") not in _ok_targets:
                _fail_targets[d.get("kind")] = d["sheet"]
        for e in ledger:
            if e.get("bucket") == "derived" and e.get("target_sheet") in _fail_targets \
                    and e.get("sheet") == _fail_targets[e["target_sheet"]]:
                e["bucket"] = "need_llm"
                e["reason"] = "detector tất định không đọc được layout -> LLM theo extraction_guide"
        # KQKD reconcile: nếu 1 sheet KQKD đã tất định (kqkd_done) mà sheet KQKD khác lỡ vào
        # need_llm (fail) -> đổi sang skip_dup (tránh LLM nạp đôi 01_HQKD). Nếu CHƯA có bản tất
        # định -> giữ ĐÚNG 1 sheet KQKD ở need_llm (cái đầu), còn lại skip_dup.
        kq_llm = [e for e in ledger if e.get("canonical_kind") == "KQKD" and e.get("bucket") == "need_llm"]
        if kqkd_done:
            for e in kq_llm:
                e["bucket"], e["reason"] = "skip_dup", "01_HQKD đã có nguồn KQKD tất định"
        elif len(kq_llm) > 1:
            for e in kq_llm[1:]:
                e["bucket"], e["reason"] = "skip_dup", "01_HQKD chỉ cần 1 sheet KQKD"

    summary = {}
    for e in ledger:
        summary[e["bucket"]] = summary.get(e["bucket"], 0) + 1
    _out({"ok": True, "file": fname, "period": period, "summary": summary, "ledger": ledger,
          "derived": derived})


def cmd_reset_learning(args):
    """XÓA phần PHÂN TÍCH ĐÃ HỌC của agent (reset sạch): fill_specs (mapping tự điền), discoveries
    (layout đã khám phá), report_specs, bindings. GIỮ inventory file (source_catalog/imports_ledger)
    để file vẫn liệt kê ở tab Nguồn (trạng thái sẽ về 'chờ phân tích'). Không đụng DB dashboard."""
    import glob
    import os as _os
    root = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "memory")
    removed = {}
    for sub_dir in ("fill_specs", "discoveries", "report_specs", "bindings"):
        d = os.path.join(root, sub_dir)
        files = glob.glob(os.path.join(d, "*.json"))
        for f in files:
            try:
                _os.remove(f)
            except OSError:
                pass
        removed[sub_dir] = len(files)
    _out({"ok": True, "removed": removed})


def cmd_forget_file(args):
    """QUÊN dấu vết import của các file (theo content_hash = sha1 bytes file) khỏi imports_ledger,
    để cho phép PHÂN TÍCH LẠI đường generic/GEN_* (dedup chặn theo content_hash). Dùng khi XOÁ HẲN
    1 file. KHÔNG đụng fill_specs/report_specs (kiến thức layout dùng chung) — chỉ mở khoá ledger."""
    import hashlib
    from servers import ingest_server as ing
    hashes, done = set(), []
    for p in (args.path or []):
        try:
            with open(p, "rb") as fh:
                h = hashlib.sha1(fh.read()).hexdigest()
            hashes.add(h)
            done.append({"path": p, "content_hash": h})
        except OSError as e:
            done.append({"path": p, "error": str(e)})
    removed = ing.ledger_remove_by_content_hash(hashes)
    _out({"ok": True, "files": done, "ledger_removed": removed})


def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)
    p = sub.add_parser("reset-learning"); p.set_defaults(fn=cmd_reset_learning)
    p = sub.add_parser("forget-file"); p.add_argument("--path", nargs="+", required=True); p.set_defaults(fn=cmd_forget_file)

    p = sub.add_parser("profile"); p.add_argument("file"); p.set_defaults(fn=cmd_profile)
    p = sub.add_parser("plan"); p.add_argument("file"); p.set_defaults(fn=cmd_plan)
    p = sub.add_parser("propose"); p.add_argument("file"); p.add_argument("--sheet", required=True)
    p.add_argument("--timeout", type=int, default=240); p.set_defaults(fn=cmd_propose)
    p = sub.add_parser("execute"); p.add_argument("file"); p.add_argument("--sheet", required=True)
    p.add_argument("--mapping-file", required=True); p.add_argument("--period"); p.add_argument("--cong-ty", dest="cong_ty")
    p.add_argument("--dry-run", action="store_true"); p.set_defaults(fn=cmd_execute)
    p = sub.add_parser("autobatch"); p.add_argument("file"); p.add_argument("--period"); p.add_argument("--cong-ty", dest="cong_ty")
    p.add_argument("--propose", action="store_true"); p.add_argument("--dry-run", action="store_true")
    p.add_argument("--timeout", type=int, default=240); p.set_defaults(fn=cmd_autobatch)
    p = sub.add_parser("autofill"); p.add_argument("file"); p.add_argument("--period")
    p.add_argument("--cong-ty", dest="cong_ty"); p.add_argument("--dry-run", action="store_true")
    p.set_defaults(fn=cmd_autofill)
    p = sub.add_parser("confirm"); p.add_argument("canonical_kind")
    p.add_argument("--scope"); p.add_argument("--basis"); p.add_argument("--target-screen", dest="target_screen")
    p.add_argument("--chi-tieu", dest="chi_tieu"); p.add_argument("--kpi-id", dest="kpi_id")
    p.set_defaults(fn=cmd_confirm)

    args = ap.parse_args()
    try:
        args.fn(args)
    except Exception as e:
        import traceback
        _out({"ok": False, "error": f"{type(e).__name__}: {e}", "trace": traceback.format_exc()[-800:]})
        sys.exit(1)


if __name__ == "__main__":
    main()
