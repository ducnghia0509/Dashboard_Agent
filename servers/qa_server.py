# -*- coding: utf-8 -*-
"""MCP server luong 2 (qa): sql_query, glossary_lookup, discovery_search, source_inspect.
Chay: `python -m servers.qa_server` (stdio MCP server).

KHONG dung RAG/vector - moi tra cuu la deterministic (SQL that, JSON tra cuu chinh xac)."""
import json
import os

from mcp.server.fastmcp import FastMCP

from .common import be_bridge as bb
from .common import guardrails, introspect, memory
from .common.db_ro import get_ro_db

mcp = FastMCP("dashboard_qa")

_HERE = os.path.dirname(os.path.abspath(__file__))
_AGENT_ROOT = os.path.normpath(os.path.join(_HERE, ".."))
_KPI_GLOSSARY_PATH = os.path.join(_AGENT_ROOT, "kpi_glossary.json")

_kpi_cache = None


def _kpi_glossary() -> list:
    global _kpi_cache
    if _kpi_cache is None:
        if os.path.exists(_KPI_GLOSSARY_PATH):
            with open(_KPI_GLOSSARY_PATH, encoding="utf-8") as fh:
                _kpi_cache = json.load(fh)
        else:
            _kpi_cache = []
    return _kpi_cache


def _norm(s) -> str:
    return bb.normalize_header(s or "", True)


@mcp.tool()
def sql_query(sql: str, params: list = None) -> dict:
    """Chạy 1 câu SELECT/WITH read-only trên raw_rows (Postgres, role read-only).
    Guardrails tự chặn DML/DDL/multi-statement và ép LIMIT. Trả rows + sql_executed."""
    conn = get_ro_db()
    rows, safe_sql = guardrails.run_readonly(conn, sql, params)
    return {"rows": rows, "sql_executed": safe_sql, "row_count": len(rows)}


@mcp.tool()
def schema_describe() -> str:
    """Mô tả cấu trúc bảng raw_rows + report_type hợp lệ (dùng để tự sinh SQL)."""
    return introspect.schema_describe()


@mcp.tool()
def glossary_lookup(term: str) -> dict:
    """Tra cứu định nghĩa/công thức/nguồn dữ liệu theo từ khoá (không dấu, không phân
    biệt hoa/thường) trong: master_data (công ty/khối/cost center/chiều phân tích),
    FIELD_DEFS/FIELD_LABELS/REPORT_LABELS (schema 9 báo cáo), và kpi_glossary.json
    (50 chỉ số quản trị từ guideline.xlsx, kèm công thức + cảnh báo đỏ + nguồn)."""
    nt = _norm(term)
    out = {"kpi_glossary": [], "field_defs": [], "master_data": [], "report_types": []}

    for rec in _kpi_glossary():
        hay = _norm(" ".join([
            rec.get("chi_tieu", ""), rec.get("nhom_bao_cao", ""), rec.get("nhom_con", ""),
            rec.get("chieu_phan_tich", ""), rec.get("canh_bao_do", ""),
            rec.get("nguon_du_lieu", ""), rec.get("cong_thuc", ""),
        ]))
        if nt in hay:
            out["kpi_glossary"].append(rec)

    for rt, fields in bb.FIELD_DEFS.items():
        for key, typ, required, _idx, _extra in fields:
            label = bb.FIELD_LABELS.get(key, key)
            if nt in _norm(key) or nt in _norm(label):
                out["field_defs"].append({
                    "report_type": rt, "report_label": bb.REPORT_LABELS.get(rt, rt),
                    "key": key, "label": label, "type": typ, "required": required,
                })

    for code, label in bb.REPORT_LABELS.items():
        if nt in _norm(code) or nt in _norm(label):
            out["report_types"].append({"code": code, "label": label})

    md = bb.master_data()
    for section in ("companies", "khoi", "costCenters", "chieuPhanTich"):
        for item in md.get(section, []):
            text = json.dumps(item, ensure_ascii=False)
            if nt in _norm(text):
                out["master_data"].append({"section": section, "item": item})

    out["total_matches"] = sum(len(v) for v in out.values() if isinstance(v, list))
    return out


@mcp.tool()
def discovery_search(query: str = None, report_type: str = None) -> list:
    """Tìm trong discovery memory: số này/file này từng được phân tích chưa, đến từ
    sheet/cột nào, report_type gì, mapping ra sao."""
    return memory.discovery_search(query=query, report_type=report_type)


@mcp.tool()
def report_spec_search(query: str = None, sheet: str = None, target_report_type: str = None,
                        canonical_kind: str = None) -> list:
    """Tìm trong catalog SheetMapping đã học (Extension 2 - sheet lạ không khớp 9 report_type
    cố định, vd '131'/'331'/'Biểu khấu hao'): sheet này đã có cách lấy dữ liệu (mapping) chưa,
    report_type GEN_* này lấy từ sheet/cột nào. canonical_kind (vd 'TK131') tìm được mapping
    đã học ở CÔNG TY KHÁC dù tên sheet/file khác nhau, miễn cùng loại báo cáo. Dùng trước khi
    phân tích lại từ đầu bằng sheet_profile, và dùng làm ngữ cảnh khi qa cần giải thích 1 số
    liệu GEN_*."""
    return memory.report_spec_search(query=query, sheet=sheet, target_report_type=target_report_type,
                                      canonical_kind=canonical_kind)


def _input_dir() -> str:
    input_dir = os.environ.get("INPUT_DIR") or "../Data_test_dashboard"
    return os.path.normpath(os.path.join(_AGENT_ROOT, input_dir))


@mcp.tool()
def source_inspect(file_name: str, sheet: str = None, max_rows: int = 200) -> dict:
    """Mở file gốc (chỉ đọc) trong INPUT_DIR để đào sâu số CHƯA hiển thị trên dashboard.
    Chặn path traversal - file phải nằm trong INPUT_DIR. Giới hạn max_rows dòng trả về."""
    from openpyxl import load_workbook

    base = _input_dir()
    target = os.path.normpath(os.path.join(base, file_name))
    if os.path.commonpath([base, target]) != base:
        raise ValueError(f"'{file_name}' nằm ngoài INPUT_DIR - không được phép đọc.")
    if not os.path.exists(target):
        raise FileNotFoundError(f"Không tìm thấy '{file_name}' trong {base}")

    wb = load_workbook(target, data_only=True, read_only=True)
    try:
        sheet_names = wb.sheetnames
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.worksheets[0]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append([("" if c is None else c) for c in row])
        return {
            "file_name": file_name, "sheet_used": ws.title, "all_sheets": sheet_names,
            "row_count_returned": len(rows), "truncated": len(rows) >= max_rows,
            "rows": rows,
        }
    finally:
        wb.close()


@mcp.tool()
def unmapped_cc_list(include_resolved: bool = False) -> list:
    """Cost center chưa khớp MD_COSTCENTER khi điền template (admin cần bổ sung danh mục
    để lần sau tự roll-up đúng khối). Mỗi mục: raw, cong_ty, sheets, count, first/last_seen."""
    return memory.unmapped_cc_list(include_resolved=include_resolved)


@mcp.tool()
def reconcile_status(dataset_id: str = None) -> dict:
    """Soi lỗ hổng pipeline: file đã kéo về chưa import (uningested_files), KPI/màn FE thiếu nguồn
    (missing_report_types/dark_kpis), và tóm tắt pipeline (collected/received/ingested). Không ghi."""
    from .common import reconcile
    return reconcile.status(dataset_id)


@mcp.tool()
def pipeline_state() -> dict:
    """View hợp nhất từng file: collected (available_metadata) -> received (indexed) -> ingested."""
    from .common import reconcile
    return reconcile.pipeline_state()


@mcp.tool()
def catalog_search(query: str = None, company: str = None, canonical_kind: str = None,
                   sheet: str = None, only_uningested: bool = False) -> list:
    """Tra CATALOG toàn bộ file đã kéo về (Connect_VPS/received_reports) — con trỏ lossless,
    trả lời 'có file/sheet/cột nào' tức thì (kể cả file CHƯA import). Không mở file.

    Mỗi mục: file, path, company, report_type, month, ingested, sheets:[{name,columns,nrows,
    canonical_kind}]. Định vị được file rồi dùng source_inspect đọc chi tiết ô gốc.
    """
    from .common import source_catalog
    return source_catalog.search(query=query, company=company, canonical_kind=canonical_kind,
                                 sheet=sheet, only_uningested=only_uningested)


if __name__ == "__main__":
    mcp.run()
