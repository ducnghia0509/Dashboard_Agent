# -*- coding: utf-8 -*-
"""ENGINE ĐIỀN TEMPLATE VÀNG.

Nhận số liệu nguồn + mapping (cột nguồn -> cột nhập liệu của template) -> ghi vào 1
BẢN SAO Template_chuan.xlsx (giữ nguyên công thức per-row: C/D auto cty/khoi từ cost
center, các cột % tự tính). File điền xong đi qua importer_template.py sẵn có -> raw_rows.

Nguyên tắc:
- CHỈ ghi cột nhập liệu (theo tên header); KHÔNG đụng cột công thức.
- Mỗi lần điền bắt đầu từ template gốc SẠCH (rows dữ liệu trống) -> không lẫn dữ liệu cũ.
- Cost center: resolve về mã chuẩn (contract.resolve_costcenter). Không khớp -> ghi
  unmapped_cc + để trống CC -> backend suy khối = NULL -> FE gom "(Chưa phân bổ)".
"""
import io
import os

from openpyxl import load_workbook

from .common import be_bridge as bb
from .common import contract as C
from .common import memory

_HERE = os.path.dirname(os.path.abspath(__file__))
_AGENT_ROOT = os.path.normpath(os.path.join(_HERE, ".."))
FILLED_DIR = os.path.normpath(os.path.join(_AGENT_ROOT, "..", "template_trust", "filled"))


def _norm(s) -> str:
    return bb.remove_diacritics("" if s is None else str(s)).strip().lower()


def _col_index(sheet_spec) -> dict:
    """{norm(header) -> 1-based column} cho sheet đích."""
    return {_norm(h): j + 1 for j, h in enumerate(sheet_spec["columns"]) if h}


def _is_costcenter_col(header: str) -> bool:
    n = _norm(header)
    return "cost center" in n or "costcenter" in n


def fill(target_sheet: str, records: list, out_path: str, source_template: str = None) -> dict:
    """Ghi list record (dict {template_header: value}) vào bản sao template.

    records: mỗi phần tử là 1 dòng nhập liệu; key = TÊN CỘT template (input), value = số/chuỗi.
    CHỈ ghi cột có trong records (bỏ qua cột công thức). Trả {rows_written, out_path}.
    """
    src = source_template or C.GOLDEN_TEMPLATE
    spec = C.data_sheets().get(target_sheet)
    if spec is None:
        raise ValueError(f"Sheet đích không hợp lệ: {target_sheet}. Hợp lệ: {list(C.data_sheets())}")
    cidx = _col_index(spec)
    start_row = spec["header_row"] + 2  # 1-based, ngay sau dòng header

    wb = load_workbook(src)  # data_only=False -> giữ công thức
    try:
        ws = wb[target_sheet]
        written = 0
        for i, rec in enumerate(records):
            r = start_row + i
            for key, val in rec.items():
                if val is None or val == "":
                    continue
                ci = cidx.get(_norm(key))
                if ci:  # chỉ ghi cột nhập liệu khớp header; cột công thức bị bỏ qua
                    ws.cell(row=r, column=ci, value=val)
            written += 1
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        wb.save(out_path)
        return {"rows_written": written, "out_path": out_path, "target_sheet": target_sheet}
    finally:
        wb.close()


def _read_source(data: bytes, sheet: str):
    """Đọc sheet nguồn -> (header:list[str], rows:list[list]) (data_only)."""
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    try:
        ws = wb[sheet] if sheet in wb.sheetnames else wb.worksheets[0]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    return rows


def _all_sheet_headers(data: bytes, n: int = 20) -> dict:
    """Mở workbook 1 LẦN, trả {sheet: [n dòng đầu]} — đủ để dò header/fingerprint.
    QUAN TRỌNG: file BCTC nặng (vd HO 8MB, sheet CĐPS 16350 cột) tốn ~18s/lần load;
    mở lại theo TỪNG sheet sẽ treo. Load 1 lần + đọc header mọi sheet chỉ ~0.03s."""
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    try:
        out = {}
        for ws in wb.worksheets:
            rows = []
            for i, r in enumerate(ws.iter_rows(values_only=True)):
                if i >= n:
                    break
                rows.append(list(r))
            out[ws.title] = rows
        return out
    finally:
        wb.close()


def _is_money_col(header: str) -> bool:
    """Cột tiền của template (đơn vị tỷ) — header chứa '(tỷ)'."""
    return "(ty)" in _norm(header)


def build_records(source_rows: list, header_row: int, mapping: dict,
                   target_sheet: str, period: str = None, file_name: str = None,
                   value_scale: float = 1.0, constants: dict = None) -> dict:
    """Từ dữ liệu nguồn + mapping -> list record cho template + resolve cost center.

    mapping: {template_header: source_header}  (cột nguồn nào điền vào cột template nào).
    constants: {template_header: giá trị HẰNG} — gán cứng cho mọi dòng (vd cost center/công ty khi
      sheet nguồn cấp công ty, không có CC theo dòng). Áp SAU mapping, TRƯỚC resolve CC.
    period: điền cột 'Kỳ' nếu template có và mapping không có.
    value_scale: nhân giá trị SỐ ở cột tiền '(tỷ)' (vd nguồn VND -> tỷ dùng 1e-9). Cột text không đổi.
    Trả {records, unresolved:[...], resolved_cc:int}.
    """
    constants = constants or {}
    hdr = [ _norm(h) for h in source_rows[header_row] ]
    src_idx = {h: j for j, h in enumerate(hdr) if h}
    # cột template nào là cost center (để resolve)
    spec = C.data_sheets()[target_sheet]
    cc_headers = [c for c in spec["columns"] if _is_costcenter_col(c)]
    period_headers = [c for c in spec["columns"] if _norm(c).startswith("ky") or _norm(c).startswith("ngay")]

    records, unresolved = [], []
    resolved_cc = 0
    for r in source_rows[header_row + 1:]:
        if not r or all(c in (None, "") for c in r):
            continue
        rec = {}
        for tmpl_h, src_h in mapping.items():
            j = src_idx.get(_norm(src_h))
            if j is not None and j < len(r):
                val = r[j]
                if value_scale != 1.0 and _is_money_col(tmpl_h) and isinstance(val, (int, float)):
                    val = val * value_scale   # VND -> tỷ (giữ cột text nguyên vẹn)
                rec[tmpl_h] = val
        for tmpl_h, cval in constants.items():   # hằng số (CC/công ty) gán cứng mọi dòng
            rec[tmpl_h] = cval
        if period and period_headers:
            rec.setdefault(period_headers[0], period)
        # resolve cost center trong record (nếu có cột CC được map)
        for cc_h in cc_headers:
            raw_cc = rec.get(cc_h)
            if raw_cc in (None, ""):
                continue
            res = C.resolve_costcenter(raw_cc)
            if res:
                rec[cc_h] = res["cc"]        # chuẩn hoá về mã CC chuẩn
                resolved_cc += 1
            else:
                memory.unmapped_cc_record(str(raw_cc), sheet=target_sheet, file_name=file_name)
                unresolved.append(str(raw_cc))
        if any(v not in (None, "") for v in rec.values()):
            records.append(rec)
    return {"records": records, "unresolved": sorted(set(unresolved)), "resolved_cc": resolved_cc}


def import_filled(path: str, cong_ty: str = None, khoi: str = None) -> dict:
    """Nạp 1 file ĐÃ ĐIỀN (template chuẩn) vào raw_rows.

    GỘP ĐA-CÔNG-TY: mỗi KỲ (tháng) = 1 dataset dùng CHUNG cho mọi công ty; import chỉ thay
    dữ liệu của (kỳ, cong_ty) NÀY (không xoá công ty khác) rồi ĐÓNG DẤU cong_ty cho các dòng
    vừa nạp -> dashboard tập đoàn thấy đủ công ty, filter theo công ty hoạt động.
    (Không truyền cong_ty -> quay lại hành vi cũ: thay cả kỳ.)"""
    with open(path, "rb") as fh:
        data = fh.read()
    _wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    try:
        is_tmpl = bb.detect_template(_wb)
    finally:
        _wb.close()
    if not is_tmpl:
        return {"ok": False, "error": "File điền không nhận diện là template chuẩn (thiếu sheet 01..12?)."}
    parsed = bb.template_parse(data)
    if not parsed["tuples"]:
        return {"ok": False, "error": "Template chưa có dòng dữ liệu nào (kiểm tra cột Kỳ)."}
    grain = parsed["grain"]
    period = parsed.get("period")
    db = bb.get_db()

    # CÔNG TY chỉ nhận MÃ HỢP LỆ trong MD_CONGTY (suy từ tên file B.<n>.<MÃ>. nếu cong_ty là
    # folder-token như 'ANTAXI'/'DUAN'); None nếu không hợp lệ -> KHÔNG đóng dấu bậy vào chiều công ty.
    cong_ty = C.resolve_company(cong_ty, os.path.basename(path))

    if grain == "month" and period:
        file_ccs = sorted({t[5] for t in parsed["tuples"] if t[5]})
        has_null_cc = any(not t[5] for t in parsed["tuples"])
        if file_ccs or cong_ty:
            # 1 dataset/kỳ dùng chung: tìm dataset kỳ này, chưa có thì tạo (KHÔNG xoá cả kỳ).
            existing = [d for d in bb.repo.list_datasets("month") if d.get("period") == period]
            target = existing[0]["id"] if existing else bb.repo.create_dataset(period, kind="month", period=period)["id"]
            bb.repo.set_active(target)
            # idempotent theo (kỳ, COST CENTER của file): chỉ thay đúng các CC file mang tới
            # (nhiều CC chung 1 công ty như HO_TC & VG_TS đều TC -> KHÔNG đè nhau).
            if file_ccs:
                ph = ",".join(["?"] * len(file_ccs))
                db.execute(f"DELETE FROM raw_rows WHERE dataset_id=? AND cost_center IN ({ph})",
                           [target, *file_ccs])
            # dòng CC trống của CÔNG TY HỢP LỆ này: thay theo (kỳ, công ty, LOẠI BÁO CÁO trong file)
            # -> KHÔNG đè loại khác. (Trước đây xoá theo mỗi cong_ty -> nạp tuần tự từng sheet của cùng
            # 1 công ty null-CC làm sheet sau xoá sạch loại của sheet trước = mất KPI.)
            if has_null_cc and cong_ty:
                file_types = sorted({t[0] for t in parsed["tuples"]})
                if file_types:
                    tph = ",".join(["?"] * len(file_types))
                    db.execute(f"DELETE FROM raw_rows WHERE dataset_id=? AND cong_ty=? "
                               f"AND (cost_center IS NULL OR cost_center='') AND report_type IN ({tph})",
                               [target, cong_ty, *file_types])
            db.commit()
        else:
            # File KHÔNG có cost center lẫn công ty PARAM nhưng cty nằm THEO DÒNG (báo cáo đa-công-ty
            # như SD TIỀN/THU CHI tập đoàn). TÁI DÙNG dataset kỳ (KHÔNG wipe cả kỳ -> tránh xoá dữ liệu
            # công ty/loại khác) + xoá theo LOẠI BÁO CÁO của file rồi chèn lại.
            existing = [d for d in bb.repo.list_datasets("month") if d.get("period") == period]
            target = existing[0]["id"] if existing else bb.repo.create_dataset(period, kind="month", period=period)["id"]
            bb.repo.set_active(target)
            file_types = sorted({t[0] for t in parsed["tuples"]})
            if file_types:
                tph = ",".join(["?"] * len(file_types))
                db.execute(f"DELETE FROM raw_rows WHERE dataset_id=? AND report_type IN ({tph})",
                           [target, *file_types])
                db.commit()
    elif grain == "day" and parsed["ngay"]:
        name = f"Ngày {parsed['ngay']}"
        bb.delete_by_key("day", name=name)
        target = bb.new_dataset(kind="day", period=period, name=name)["id"]
    else:
        target = bb.new_dataset(kind=grain, period=period, name=None)["id"]

    before_id = db.execute("SELECT COALESCE(MAX(id),0) m FROM raw_rows").fetchone()["m"]
    result = bb.template_import_parsed(target, parsed)
    # ĐÓNG DẤU công ty CHỈ cho dòng CHƯA suy được công ty từ CC (fallback) -> KHÔNG ghi đè
    # công ty đã resolve đúng từ cost center (tránh folder-token 'ANTAXI' đè công ty thật 'AAG').
    if cong_ty:
        db.execute("UPDATE raw_rows SET cong_ty=? WHERE dataset_id=? AND id>? "
                   "AND (cong_ty IS NULL OR cong_ty='')", [cong_ty, target, before_id])
        db.commit()
    # ĐÓNG DẤU KHỐI (HIỆU QUẢ KD gán theo Khối suy từ đường dẫn nguồn) CHỈ cho dòng CHƯA có khối
    # (khối suy từ cost center vẫn giữ nguyên). KHÔNG stamp cho dòng tiền THUCHI (theo pháp nhân).
    if khoi:
        # canonical hoá về ĐÚNG tên khối metric dùng (master.khoi_names) — khớp hoa/thường (Template
        # 'Xe điện' vs master 'xe điện') để không tạo card khối "lạ" tách khỏi danh mục.
        canon = {n.strip().upper(): n for n in bb.khoi_names()}
        khoi = canon.get(str(khoi).strip().upper(), khoi)
        db.execute("UPDATE raw_rows SET khoi=? WHERE dataset_id=? AND id>? "
                   "AND (khoi IS NULL OR khoi='') AND report_type<>'THUCHI'", [khoi, target, before_id])
        db.commit()
    if grain == "month" and result.get("period"):
        bb.set_period(target, result["period"])
    return {"ok": True, "dataset_id": target, "grain": grain, "cong_ty": cong_ty,
            "rows_imported": result["rows"], "by_type": result["by_type"],
            "period": result.get("period"), "ngay": result.get("ngay")}


def _derive_kqkd_summary(records: list, columns: list) -> list:
    """Sinh 2 dòng chuẩn cho 01_HQKD để KPI Chi phí & LNTT sáng (nguồn KQKD liệt kê chi phí rời,
    không có dòng 'Tổng chi phí'; dòng LNTT tên 'Tổng lợi nhuận... trước thuế' không khớp code).

    - Tổng chi phí = Σ Thực hiện các dòng chi phí (giá vốn/CPTC/CPBH/CPQLDN/CP khác).
    - Lợi nhuận trước thuế = Thực hiện dòng chứa 'loi nhuan'+'truoc thue'.
    Nhóm theo (Kỳ, Cost center). KHÔNG thêm nếu đã có sẵn dòng chuẩn (tránh trùng)."""
    chi = next((h for h in columns if "chi tieu" in _norm(h)), None)
    th = next((h for h in columns if "thuc hien" in _norm(h)), None)
    if not chi or not th:
        return records
    ky = next((h for h in columns if _norm(h).startswith(("ky", "ngay"))), None)
    cc = next((h for h in columns if _is_costcenter_col(h)), None)
    from collections import defaultdict
    G = defaultdict(lambda: {"cost": 0.0, "lntt": None, "has_cp": False, "has_lntt": False, "proto": None})
    for r in records:
        ct = _norm(r.get(chi))
        key = (r.get(ky) if ky else None, r.get(cc) if cc else None)
        g = G[key]
        if g["proto"] is None:
            g["proto"] = r
        if ct == "tong chi phi":
            g["has_cp"] = True
        if ct == "loi nhuan truoc thue":
            g["has_lntt"] = True
        v = r.get(th)
        if isinstance(v, (int, float)):
            if any(p in ct for p in C.KQKD_COST_PATTERNS):
                g["cost"] += v
            if all(p in ct for p in C.KQKD_LNTT_REQUIRE):
                g["lntt"] = v
    derived = []
    for (ky_v, cc_v), g in G.items():
        def _mk(name, val):
            rec = {chi: name, th: val}
            if ky:
                rec[ky] = ky_v
            if cc and cc_v is not None:
                rec[cc] = cc_v
            return rec
        if g["cost"] and not g["has_cp"]:
            derived.append(_mk("Tổng chi phí", round(g["cost"], 6)))
        if g["lntt"] is not None and not g["has_lntt"]:
            derived.append(_mk("Lợi nhuận trước thuế", round(g["lntt"], 6)))
    return records + derived


def _header_row_of(rows) -> int:
    """Dòng header = dòng CÓ NHIỀU Ô NHẤT trong 20 dòng đầu (dòng đầu tiên đạt max).
    Bền với báo cáo tài chính thật có nhiều dòng tiêu đề/metadata (Đơn vị/Mẫu số...) ở trên."""
    best_i, best_n = 0, -1
    for i, r in enumerate(rows[:20]):
        n = sum(1 for c in r if c not in (None, ""))
        if n > best_n:
            best_n, best_i = n, i
    return best_i


def _apply_rename_rows(records: list, target_sheet: str, rename_rows: dict) -> int:
    """ÁNH XẠ TÊN CHỈ TIÊU về CHUẨN cho ĐÚNG dòng analyst chỉ định (row-level, khác mapping cột).
    rename_rows = {tên chỉ tiêu NGUỒN (khớp không dấu, chứa-trong): tên chuẩn}. Dùng cho 01_HQKD:
    analyst chọn ĐÚNG 1 dòng tổng doanh thu -> 'Doanh thu thuần', 1 dòng tổng chi phí -> 'Tổng chi phí',
    1 dòng LNTT -> 'Lợi nhuận trước thuế' (chọn 1 dòng/chỉ tiêu để KHỎI cộng trùng). Trả số dòng đã đổi."""
    if not rename_rows:
        return 0
    chi = next((h for h in (records[0].keys() if records else []) if "chi tieu" in _norm(h)), None)
    if not chi:
        return 0
    n = 0
    for src, canonical in rename_rows.items():
        if not src or not canonical:
            continue
        sn = _norm(src)
        if not sn:
            continue
        # Mỗi key đổi ĐÚNG 1 dòng: ưu tiên khớp CHÍNH XÁC; nếu không có, lấy dòng chứa-trong NGẮN
        # NHẤT (tổng cấp cao nhất, tránh khớp nhầm dòng con dài hơn) -> KHÔNG cộng trùng.
        exact = [r for r in records if _norm(r.get(chi)) == sn]
        cands = exact or [r for r in records if sn in _norm(r.get(chi))]
        if not cands:
            continue
        best = min(cands, key=lambda r: len(_norm(r.get(chi))))
        if best.get(chi) != canonical:
            best[chi] = canonical
            n += 1
    return n


def fill_from_source(data: bytes, source_sheet: str, target_sheet: str, mapping: dict,
                     period: str = None, cong_ty: str = None, file_name: str = None,
                     dry_run: bool = True, auto_import: bool = False, learn: bool = True,
                     value_scale: float = 1.0, constants: dict = None,
                     normalize_kqkd: bool = True, rename_rows: dict = None,
                     khoi: str = None, source_path: str = None) -> dict:
    """Pipeline: đọc nguồn -> build records (+resolve CC) -> (dry_run: preview | ghi file điền
    [+ auto_import: nạp raw_rows]). value_scale nhân cột tiền (VND->tỷ dùng 1e-9). constants gán cứng
    (vd cost center). rename_rows: đổi tên CHỈ TIÊU của đúng dòng nguồn về tên chuẩn (xem _apply_rename_rows
    — cần cho 01_HQKD để KPI Doanh thu/LNTT sáng). normalize_kqkd: với 01_HQKD tự sinh dòng 'Tổng chi phí'
    & 'Lợi nhuận trước thuế' NẾU nguồn chưa có. learn=True: khi ghi thật, LƯU mapping+scale+constants."""
    rows = _read_source(data, source_sheet)
    if not rows:
        return {"ok": False, "error": f"Sheet nguồn rỗng: {source_sheet}"}
    header_row = _header_row_of(rows)
    src_cols = [c for c in rows[header_row]] if header_row < len(rows) else []
    fp = memory.source_fingerprint(source_sheet, src_cols)
    built = build_records(rows, header_row, mapping, target_sheet, period=period,
                          file_name=file_name, value_scale=value_scale, constants=constants)
    renamed_n = _apply_rename_rows(built["records"], target_sheet, rename_rows)  # TRƯỚC derive
    # Sheet đích grain THÁNG: ÉP cột Kỳ = kỳ tháng (đè nếu mapping lỡ ghi NGÀY vào Kỳ) -> tránh
    # grain bị lật thành 'day' rồi import vào dataset day (bị delete_by_key day đè, mất dữ liệu).
    if period and C.grain_for(target_sheet) == "month":
        chi_ky = next((h for h in (built["records"][0].keys() if built["records"] else [])
                       if _norm(h).startswith("ky")), None)
        if chi_ky:
            for r in built["records"]:
                r[chi_ky] = period
    derived_n = 0
    if normalize_kqkd and target_sheet == "01_HQKD":
        cols = C.data_sheets().get("01_HQKD", {}).get("columns", [])
        before = len(built["records"])
        built["records"] = _derive_kqkd_summary(built["records"], cols)
        derived_n = len(built["records"]) - before
    result = {
        "ok": True, "target_sheet": target_sheet, "grain": C.grain_for(target_sheet),
        "source_fingerprint": fp,
        "row_count": len(built["records"]),
        "renamed_rows": renamed_n,  # số dòng đã đổi tên về chuẩn (rename_rows)
        "derived_kqkd": derived_n,  # số dòng chuẩn tự sinh (Tổng chi phí / LNTT)
        "resolved_cc": built["resolved_cc"], "unresolved_cc": built["unresolved"],
        # sample: kèm các dòng suy diễn ở cuối để kiểm
        "sample": (built["records"][:4] + built["records"][-derived_n:]) if derived_n else built["records"][:5],
        "dry_run": dry_run,
    }
    if dry_run:
        return result
    tag = "_".join(x for x in [cong_ty, period, target_sheet] if x) or target_sheet
    out_path = os.path.join(FILLED_DIR, f"{tag}.xlsx".replace("/", "-"))
    w = fill(target_sheet, built["records"], out_path)
    result["out_path"] = w["out_path"]
    result["rows_written"] = w["rows_written"]
    if learn and built["records"]:
        memory.fill_spec_save(fp, target_sheet, mapping, source_sheet=source_sheet,
                              value_scale=value_scale, constants=constants, rename_rows=rename_rows)
        result["learned"] = True
    if auto_import:
        # resolve công ty HỢP LỆ từ TÊN FILE NGUỒN (import_filled chỉ thấy tên file đã điền,
        # không có mã 'B.7.AAG' để suy) -> ANTAXI/DUAN gắn đúng AAG/TC thay vì folder-token.
        # khoi: suy từ ĐƯỜNG DẪN nguồn (file_name có thể là path) -> stamp Khối cho dòng khoi NULL.
        result["import"] = import_filled(out_path, cong_ty=C.resolve_company(cong_ty, file_name),
                                         khoi=khoi or C.khoi_from_path(source_path or file_name))
    return result


def _source_sheets(data: bytes) -> list:
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def guess_period(file_name: str):
    """Suy kỳ 'YYYY-MM' từ tên file: 'YYYYMM' (202601) hoặc 'MM.YYYY' (05.2026)."""
    import re
    if not file_name:
        return None
    m = re.search(r"(20\d{2})[.\-_]?(0[1-9]|1[0-2])", file_name)   # 2026-01 / 202601
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    m = re.search(r"(0[1-9]|1[0-2])[.\-_](20\d{2})", file_name)     # 05.2026
    if m:
        return f"{m.group(2)}-{m.group(1)}"
    return None


def autofill_file(path: str, period: str = None, cong_ty: str = None) -> dict:
    """TỰ ĐỘNG (không LLM): với mỗi sheet nguồn có fill_spec ĐÃ HỌC (theo fingerprint) ->
    điền template + import raw_rows. Sheet chưa học -> bỏ qua (cần analyst). Dùng trong orchestrator."""
    with open(path, "rb") as fh:
        data = fh.read()
    fname = os.path.basename(path)
    period = period or guess_period(fname)
    processed, skipped = [], []
    headers = _all_sheet_headers(data)  # mở workbook 1 LẦN (file nặng: load rất chậm)
    for sheet, head in headers.items():
        if not head:
            continue
        cols = head[_header_row_of(head)]
        spec = memory.fill_spec_find(memory.source_fingerprint(sheet, cols))
        if not spec:
            skipped.append(sheet)
            continue
        r = fill_from_source(data, sheet, spec["target_sheet"], spec["mapping"],
                             period=period, cong_ty=cong_ty, file_name=fname,
                             dry_run=False, auto_import=True, learn=False,
                             value_scale=spec.get("value_scale", 1.0),
                             constants=spec.get("constants"))
        processed.append({"sheet": sheet, "target": spec["target_sheet"],
                          "rows": r.get("rows_written"), "import": r.get("import", {}).get("ok")})
    return {"ok": True, "file": fname, "period": period,
            "processed": processed, "skipped_sheets": skipped,
            "any_processed": bool(processed)}
