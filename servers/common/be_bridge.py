# -*- coding: utf-8 -*-
"""Cau noi sang backend/app cua DashBoard_AI: sys.path.insert roi import lai nguyen
logic detect/auto_map/prepare/commit/import_workbook/master_data.

KHONG viet lai logic nghiep vu - moi thu goi thang vao module app.* cua DashBoard_AI.
"""
import os
import sys

from dotenv import load_dotenv

_HERE = os.path.dirname(os.path.abspath(__file__))
_AGENT_ROOT = os.path.normpath(os.path.join(_HERE, "..", ".."))
load_dotenv(os.path.join(_AGENT_ROOT, ".env"))

BACKEND_PATH = os.environ.get("BACKEND_PATH") or os.path.normpath(
    os.path.join(_AGENT_ROOT, "..", "DashBoard_AI", "backend")
)
BACKEND_PATH = os.path.normpath(os.path.join(_AGENT_ROOT, BACKEND_PATH)) \
    if not os.path.isabs(BACKEND_PATH) else BACKEND_PATH

if BACKEND_PATH not in sys.path:
    sys.path.insert(0, BACKEND_PATH)

try:
    from app import (schemas, importer, importer_month, importer_ledger, importer_template,
                     db, repo, master, metrics, text_util)
except ModuleNotFoundError as exc:  # pragma: no cover - lỗi cấu hình, không phải logic
    raise RuntimeError(
        f"Không import được app.* từ BACKEND_PATH={BACKEND_PATH!r}. "
        f"Kiểm tra .env (BACKEND_PATH) hoặc chạy `pip install -r {BACKEND_PATH}/../requirements.txt`."
    ) from exc

# Re-export ngắn gọn cho 2 MCP server dùng.
detect = schemas.detect
auto_map = schemas.auto_map
build_schemas = schemas.build_schemas
FIELD_DEFS = schemas.FIELD_DEFS
FIELD_LABELS = schemas.FIELD_LABELS
REPORT_CODES = schemas.REPORT_CODES
REPORT_LABELS = schemas.REPORT_LABELS
SNAPSHOT = schemas.SNAPSHOT

prepare = importer.prepare
commit = importer.commit
import_workbook = importer_month.import_workbook
detect_ledger = importer_ledger.detect_ledger

# Template chuẩn v3 (13 sheet) — để import file điền từ template vàng.
detect_template = importer_template.detect_template
template_parse = importer_template.parse_workbook
template_import_parsed = importer_template.import_parsed
delete_by_key = repo.delete_by_key
set_period = repo.set_period

master_data = master.master_data
report_templates = master.report_templates
khoi_names = master.khoi_names
cc_to_khoi = master.cc_to_khoi

get_db = db.get_db
DB_SCHEMA = db.SCHEMA

normalize_header = text_util.normalize_header
remove_diacritics = text_util.remove_diacritics
parse_num = text_util.parse_num
parse_date = text_util.parse_date
parse_text = text_util.parse_text

fingerprint = repo.fingerprint
find_profile = repo.find_profile
save_profile = repo.save_profile
get_dataset = repo.get_dataset
list_datasets = repo.list_datasets
new_dataset = repo.new_dataset


# Ngưỡng styles.xml "phình": file sạch bình thường <300KB; GA/HO/XVP quan sát thấy 7-12MB.
_STYLES_BLOAT_THRESHOLD = 1_000_000

# Cache CẤP TIẾN TRÌNH Workbook ĐÃ MỞ theo (path, mtime, size, read_only, data_only) — 1 file
# nguồn thường bị MỞ LẶP LẠI nhiều lần/lượt autofill (mỗi deriver _derive_* tự mở riêng bằng
# đường dẫn, không truyền tay wb cho nhau: đo thực tế GA 31 lần/file). Cache BYTES (đã cắt style)
# chỉ né được phần rebuild zip; phần CÒN LẠI — openpyxl tự parse sharedStrings/sheet index — vẫn
# mất ~1-2s MỖI LẦN mở dù bytes đã rẻ (test: cache bytes không giảm 31 lượt kế tiếp). Cache thẳng
# OBJECT đã mở né toàn bộ chi phí đó cho lượt 2+. Nhiều "deriver" đọc CÙNG wb (khác sheet hoặc
# cùng sheet) đã verify AN TOÀN (kết quả giống hệt mở riêng); rủi ro DUY NHẤT là code cũ luôn tự
# `wb.close()` trong finally — patch `.close` thành no-op TRÊN INSTANCE cache (không đụng class)
# để lượt sau vẫn dùng được. Áp dụng cho MỌI file (không riêng GA/HO/XVP phình) vì việc mở lặp là
# kiến trúc chung của cmd_autofill. Cache theo (mtime,size) tự invalidate khi file đổi (re-pull).
_workbook_cache = {}


def fast_load_workbook(source, **kwargs):
    """Wrapper quanh openpyxl.load_workbook, 2 lớp tăng tốc CỘNG DỒN:
    (1) Cache Workbook ĐÃ MỞ theo định danh nguồn (path+mtime+size, HOẶC md5 nội dung nếu nguồn
        là bytes/BytesIO — vd ingest_server đọc lại file từ đĩa mỗi lượt profile/route) — né việc
        1 file bị mở lại hàng chục lần/lượt autofill (mỗi deriver tự mở độc lập, không truyền tay
        wb cho nhau). `.close()` bị vô hiệu hoá TRÊN INSTANCE cache (đã verify nhiều lượt đọc/
        nhiều sheet trên cùng object cho kết quả giống hệt mở riêng — code gọi vẫn `finally:
        wb.close()` an toàn, không đụng class/instance khác).
    (2) Với cache-miss: một số nguồn (GA/HO/XVP quan sát thấy) có xl/styles.xml PHÌNH TO bất
        thường (7-12MB so với <300KB bình thường) do Excel tích luỹ <cellStyleXfs>/<cellStyles>
        (catalog "Cell Styles" cho UI Excel, KHÔNG ảnh hưởng giá trị/kiểu ô — <numFmts>/<cellXfs>
        giữ NGUYÊN). Đo thực tế: GA T06 load_workbook 13.16s -> 1.00s (giảm ~92%) sau khi cắt 2
        block này. File không phình -> mở thẳng, không tốn overhead.

    source: đường dẫn (str) hoặc bytes/BytesIO (như openpyxl.load_workbook nhận). kwargs truyền
    thẳng cho load_workbook (read_only=, data_only=, ...)."""
    import hashlib
    import io
    import os
    import re
    import zipfile

    import openpyxl

    # CHỈ cache khi read_only=True. Cache lưu Y NGUYÊN 1 Workbook OBJECT dùng chung nhiều lượt gọi;
    # caller mở KHÔNG read_only (vd tf.fill() nạp golden template rồi GHI đè ô + wb.save(out_path)
    # khác đường dẫn) coi object là "bản nháp riêng, dùng 1 lần" — nếu cache/tái dùng, lượt gọi SAU
    # (khác record, có thể THIẾU field mà lượt TRƯỚC đã ghi) sẽ GIỮ LẠI giá trị Ô CŨ của lượt trước
    # tại cùng vị trí dòng/cột -> rò rỉ dữ liệu chéo giữa các đơn vị/lượt fill khác nhau (đã bắt được
    # thực tế: TRẠM SẠC ghi "PS tăng" xong, ANTAXI fill sau đó — record ANTAXI không có key này ở
    # đúng dòng đó -> ANTAXI bị đè nhầm giá trị "PS tăng" của TRẠM SẠC). read_only=True vẫn cache
    # bình thường (nhiều reader cùng đọc 1 object bất biến là AN TOÀN, đây là ca tăng tốc chính).
    is_path = isinstance(source, str)
    if is_path:
        stream = source   # path str: zipfile/openpyxl tự mở lại từ đĩa, không cần seek
        if kwargs.get("read_only") is not True:
            cache_key = None
        else:
            try:
                st = os.stat(source)
                cache_key = (source, st.st_mtime_ns, st.st_size,
                             kwargs.get("read_only"), kwargs.get("data_only"))
            except OSError:
                cache_key = None
    else:
        # bytes/BytesIO: đọc RA bytes THẬT (1 lần) để (a) băm làm cache key, (b) tự chủ 1 BytesIO
        # RIÊNG cho mọi thao tác sau — không phụ thuộc vị trí đọc/seek của object gốc caller đưa.
        raw = bytes(source) if isinstance(source, (bytes, bytearray)) else source.read()
        stream = io.BytesIO(raw)
        cache_key = (("#content", hashlib.md5(raw).hexdigest(), len(raw),
                     kwargs.get("read_only"), kwargs.get("data_only"))
                     if kwargs.get("read_only") is True else None)

    if cache_key is not None and cache_key in _workbook_cache:
        return _workbook_cache[cache_key]

    def _rewound():
        if not is_path:
            stream.seek(0)
        return stream

    def _cache_and_return(wb):
        if cache_key is not None:
            wb.close = lambda: None   # no-op TRÊN INSTANCE — code gọi vẫn `finally: wb.close()` an toàn
            _workbook_cache[cache_key] = wb
        return wb

    try:
        zin = zipfile.ZipFile(stream)
        info = zin.getinfo("xl/styles.xml")
    except (KeyError, zipfile.BadZipFile, OSError):
        return _cache_and_return(openpyxl.load_workbook(_rewound(), **kwargs))

    if info.file_size <= _STYLES_BLOAT_THRESHOLD:
        return _cache_and_return(openpyxl.load_workbook(_rewound(), **kwargs))

    styles = zin.read("xl/styles.xml").decode("utf-8", errors="ignore")
    stub_by_tag = {
        "cellStyleXfs": '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>',
        "cellStyles": '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>',
    }
    stripped, inserts = styles, []
    for tag, stub in stub_by_tag.items():
        stripped, n = re.subn(rf"<{tag}\b[^>]*/>|<{tag}\b[^>]*>.*?</{tag}>", "", stripped,
                              count=1, flags=re.S)
        if n:   # chỉ chèn lại stub nếu THỰC SỰ vừa cắt (không đoán bừa cấu trúc file lạ)
            inserts.append(stub)
    if not inserts:   # phình do chỗ khác (không phải 2 block đã biết) -> AN TOÀN, mở nguyên bản
        return _cache_and_return(openpyxl.load_workbook(_rewound(), **kwargs))
    stripped = stripped.replace("</styleSheet>", "".join(inserts) + "</styleSheet>")

    buf = io.BytesIO()
    # ZIP_STORED (không nén): buffer chỉ tồn tại trong RAM, dùng 1 lần rồi bỏ -> nén DEFLATE tốn
    # CPU vô ích cho sheet to (vd TC_CDPS 12MB) copy-qua không đổi nội dung. STORED vừa ghi vừa
    # đọc lại (openpyxl) đều nhanh hơn, đổi lấy buffer lớn hơn — chấp nhận được vì không ghi đĩa.
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_STORED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/styles.xml":
                data = stripped.encode("utf-8")
            zout.writestr(item, data)
    buf.seek(0)
    return _cache_and_return(openpyxl.load_workbook(buf, **kwargs))
