# -*- coding: utf-8 -*-
"""CONTRACT của template vàng (Template_chuan.xlsx).

Nạp & cache: các sheet nhập liệu (01_HQKD..12_KDVH) + cột đích + quy tắc kỳ
(ngày/tháng), danh mục chuẩn MD_CONGTY/MD_KHOIKD/MD_COSTCENTER + DM_DANHMUC.
Cung cấp resolve_costcenter() để suy Mã Công ty + Mã Khối từ 1 chuỗi nguồn
(mã CC / tên CC / tên bộ phận) — đúng mô hình "auto từ CC" của template.

KHÔNG ghi gì; chỉ đọc. template_filler.py dùng để điền số vào bản sao template.
"""
import difflib
import os
import re

from openpyxl import load_workbook

from . import be_bridge as bb
from . import canonical

_HERE = os.path.dirname(os.path.abspath(__file__))
_AGENT_ROOT = os.path.normpath(os.path.join(_HERE, "..", ".."))

# Vị trí template vàng: ưu tiên env GOLDEN_TEMPLATE, mặc định ~/template_trust/Template_chuan.xlsx
GOLDEN_TEMPLATE = os.environ.get("GOLDEN_TEMPLATE") or os.path.normpath(
    os.path.join(_AGENT_ROOT, "..", "template_trust", "Template_chuan.xlsx")
)

# Các sheet nhập liệu (bỏ README/00_50CHITIEU/MAP_FIELD/DM_/MD_).
_DATA_SHEET_RE = re.compile(r"^\d{2}[A-Z]?_")

_cache = None
_cache_mtime = None


def _norm(s) -> str:
    if s is None:
        return ""
    return bb.remove_diacritics(str(s)).strip().lower()


def _header_row(rows) -> int:
    """Dòng tiêu đề cột của sheet nhập liệu: dòng đầu (trong 6 dòng đầu) có >=4 ô
    và ô đầu bắt đầu bằng 'ky'/'ngay' (Kỳ / Ngày / Kỳ / Ngày...)."""
    for i, r in enumerate(rows[:6]):
        cells = [c for c in r if c not in (None, "")]
        if len(cells) >= 4:
            head = _norm(r[0])
            if head.startswith("ky") or head.startswith("ngay") or "ky" in head.split()[:1]:
                return i
    return 1  # fallback: dòng 2


def _grain(title: str) -> str:
    """Quy tắc kỳ đọc từ tiêu đề sheet: 'day' | 'month' | 'both'."""
    n = _norm(title)
    seg = n.split("ky nhap", 1)[-1] if "ky nhap" in n else n
    has_month = "thang" in seg or "yyyy-mm" in seg
    has_day = "ngay" in seg or "yyyy-mm-dd" in seg
    if has_month and has_day:
        return "both"
    if has_day:
        return "day"
    return "month"  # đa số sheet snapshot


def _sheet_values(ws):
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _mark_input_vs_formula(data_sheets: dict):
    """Đọc lại template với data_only=False: cột có công thức ở dòng dữ liệu đầu = cột AUTO
    (khỏi map vào); còn lại = cột NHẬP LIỆU. Giúp analyst chỉ map vào đúng cột nhập."""
    wb = load_workbook(GOLDEN_TEMPLATE, read_only=True, data_only=False)
    try:
        for name, spec in data_sheets.items():
            d0 = spec["header_row"] + 2  # 1-based dòng dữ liệu đầu
            n = len(spec["columns"])
            is_formula = [False] * n
            is_vlookup = [False] * n   # công thức tra cứu DIMENSION (cty/khối auto) -> KHÔNG map
            for row in wb[name].iter_rows(min_row=d0, max_row=d0 + 5, values_only=True):
                for i in range(n):
                    v = row[i] if i < len(row) else None
                    if isinstance(v, str) and v.startswith("="):
                        is_formula[i] = True
                        if "VLOOKUP" in v.upper():
                            is_vlookup[i] = True
            cols = spec["columns"]
            spec["input_columns"] = [h for i, h in enumerate(cols) if h and not is_formula[i]]
            # DIMENSION auto (VLOOKUP cty/khối): TUYỆT ĐỐI không map.
            spec["auto_dim_columns"] = [h for i, h in enumerate(cols) if h and is_vlookup[i]]
            # Công thức GIÁ TRỊ (Dư cuối=ĐK+Tăng−Giảm, %...): openpyxl điền KHÔNG có giá trị cache ->
            # importer đọc =0. Nếu NGUỒN có sẵn giá trị này (vd Cuối kỳ) thì NÊN map đè để nhất quán.
            spec["calc_columns"] = [h for i, h in enumerate(cols) if h and is_formula[i] and not is_vlookup[i]]
            spec["formula_columns"] = spec["auto_dim_columns"] + spec["calc_columns"]  # tương thích cũ
    finally:
        wb.close()


def load_contract(force: bool = False) -> dict:
    """Nạp (cache theo mtime) toàn bộ contract từ template vàng."""
    global _cache, _cache_mtime
    if not os.path.exists(GOLDEN_TEMPLATE):
        raise FileNotFoundError(f"Không tìm thấy template vàng: {GOLDEN_TEMPLATE}")
    mtime = os.path.getmtime(GOLDEN_TEMPLATE)
    if _cache is not None and not force and _cache_mtime == mtime:
        return _cache

    wb = load_workbook(GOLDEN_TEMPLATE, read_only=True, data_only=True)
    try:
        names = wb.sheetnames
        data_sheets = {}
        for name in names:
            # 00_50CHITIEU là sheet tham chiếu KPI, không phải sheet nhập liệu -> bỏ.
            if not _DATA_SHEET_RE.match(name) or name.startswith("00_"):
                continue
            rows = _sheet_values(wb[name])
            if not rows:
                continue
            hr = _header_row(rows)
            title = " ".join(str(c) for c in (rows[0] or []) if c)
            cols = [("" if c is None else str(c).strip()) for c in rows[hr]]
            # bỏ cột rỗng ở đuôi
            while cols and cols[-1] == "":
                cols.pop()
            data_sheets[name] = {
                "title": title,
                "grain": _grain(title),
                "header_row": hr,
                "columns": cols,
            }

        md_congty = _kv(wb, "MD_CONGTY")          # {ma: ten}
        md_khoikd = _kv(wb, "MD_KHOIKD")          # {ma_khoi: ten_khoi}
        cc = _cost_centers(wb, md_khoikd)         # {norm(code|name): entry}
        aliases = _dm_cc_aliases(wb)              # bổ sung alias CC từ DM_DANHMUC

        _mark_input_vs_formula(data_sheets)       # cột nào NHẬP, cột nào công thức (auto)

        _cache = {
            "path": GOLDEN_TEMPLATE,
            "sheets": names,
            "data_sheets": data_sheets,
            "md_congty": md_congty,
            "md_khoikd": md_khoikd,
            "costcenter_index": cc,
            "cc_aliases": aliases,
        }
        _cache_mtime = mtime
        return _cache
    finally:
        wb.close()


def _kv(wb, sheet) -> dict:
    """Sheet 2 cột (mã | tên), dòng 0 là header."""
    if sheet not in wb.sheetnames:
        return {}
    out = {}
    for r in list(wb[sheet].iter_rows(values_only=True))[1:]:
        if r and r[0] not in (None, "") and len(r) > 1 and r[1]:
            out[str(r[0]).strip()] = str(r[1]).strip()
    return out


def _cost_centers(wb, md_khoikd) -> dict:
    """MD_COSTCENTER: Mã CC | Tên CC | Mã Công ty | Mã Khối KD.
    Trả index tra cứu theo norm(mã) và norm(tên) -> entry {cc, cong_ty, khoi_ma, khoi_ten, cc_name}."""
    idx = {}
    if "MD_COSTCENTER" not in wb.sheetnames:
        return idx
    for r in list(wb["MD_COSTCENTER"].iter_rows(values_only=True))[1:]:
        if not r or r[0] in (None, ""):
            continue
        code = str(r[0]).strip()
        name = str(r[1]).strip() if len(r) > 1 and r[1] else ""
        cty = str(r[2]).strip() if len(r) > 2 and r[2] else None
        khoi_ma = str(r[3]).strip() if len(r) > 3 and r[3] is not None else None
        entry = {
            "cc": code, "cc_name": name, "cong_ty": cty,
            "khoi_ma": khoi_ma, "khoi_ten": md_khoikd.get(khoi_ma) if khoi_ma else None,
        }
        idx[_norm(code)] = entry
        if name:
            idx.setdefault(_norm(name), entry)
    return idx


def _dm_cc_aliases(wb) -> dict:
    """DM_DANHMUC có cột 'Mã Costcenter' + 'Cost center' -> alias tên->mã (bổ sung)."""
    out = {}
    if "DM_DANHMUC" not in wb.sheetnames:
        return out
    rows = _sheet_values(wb["DM_DANHMUC"])
    # tìm dòng header chứa 'costcenter'
    hr = None
    for i, r in enumerate(rows[:5]):
        joined = _norm(" ".join(str(c) for c in r if c))
        if "costcenter" in joined or "cost center" in joined:
            hr = i
            break
    if hr is None:
        return out
    header = [_norm(c) for c in rows[hr]]
    ci_code = next((j for j, h in enumerate(header) if "ma costcenter" in h or "ma cost center" in h), None)
    ci_name = next((j for j, h in enumerate(header) if h == "cost center"), None)
    for r in rows[hr + 1:]:
        code = str(r[ci_code]).strip() if ci_code is not None and ci_code < len(r) and r[ci_code] else None
        name = str(r[ci_name]).strip() if ci_name is not None and ci_name < len(r) and r[ci_name] else None
        if name and code:
            out[_norm(name)] = code
    return out


def data_sheets() -> dict:
    return load_contract()["data_sheets"]


def grain_for(sheet: str) -> str:
    return load_contract()["data_sheets"].get(sheet, {}).get("grain", "month")


# Chọn sheet đích theo loại báo cáo nguồn (canonical_kind -> sheet template).
# NGUỒN SỰ THẬT DUY NHẤT cho auto-route: 1 workbook nguồn nhiều sheet -> mỗi sheet vào đúng
# sheet nhập liệu (thay vì chỉ nạp mỗi sheet KQKD như luồng cũ). Mở rộng: chỉ thêm dòng ở đây
# + alias tương ứng trong canonical.CANONICAL_KINDS.
SHEET_FOR_KIND = {
    "KQKD": "01_HQKD", "CDKT": "07_TAISAN_NV", "CDPS": "07_TAISAN_NV",
    "TK131": "05_PHAITHU", "TK331": "06_PHAITRA", "TSCD": "08_TSCD",
    "LCTT": "03_DONGTIEN", "THUCHI": "03_DONGTIEN", "SODU_TIEN": "03B_SODU_TIEN",
    "VAY": "04_VAY", "TONKHO": "09_TONKHO", "THUE": "10_THUE",
    "DAUTU": "11_DAUTU", "KDVH": "12_KDVH",
}


def route_sheet(sheet_name: str, title_text: str = ""):
    """Try-route (tất định, KHÔNG LLM): -> (target_sheet, canonical_kind, via).
    Ưu tiên khớp theo TÊN sheet; nếu tên không nhận diện được thì thử theo NỘI DUNG
    (tiêu đề trong sheet — bắt các tên khó như '156'='BÁO CÁO NHẬP XUẤT TỒN',
    'KHTS'='BẢNG TÍNH KHẤU HAO'). via = 'name'|'content'|None. Trả target=None nếu là sheet
    metadata (skip) hoặc không nhận diện được loại (unknown — KHÔNG drop im lặng, để caller xử)."""
    if canonical.is_skip_sheet(sheet_name):
        return None, None, "skip"
    ck = canonical.guess_canonical_kind(sheet_name)
    via = "name"
    if not ck and title_text:
        ck = canonical.guess_kind_from_content(title_text)   # strict: chỉ cụm tiêu đề đặc trưng
        via = "content"
    if not ck:
        return None, None, None
    return SHEET_FOR_KIND.get(ck), ck, via

# NHẤT QUÁN FE↔BE (đã audit importer_template + metrics*.py + main.py _SCREENS 2026-07-04):
# mỗi sheet template -> report_type importer sinh -> màn FE đọc. Analyst CHỈ nên nhắm sheet
# có trong bảng này (đều WIRED end-to-end), tránh đẻ dữ liệu không màn nào đọc.
SHEET_WIRING = {
    "01_HQKD":       {"report_type": ["DTHU", "HQKD", "PNLT", "TREND"], "man_hinh": "Tổng quan · Doanh thu · Chi phí · Lợi nhuận"},
    "02_CHIPHI":     {"report_type": ["CHIPHI"], "man_hinh": "Chi phí (cơ cấu chi tiết)"},
    "03_DONGTIEN":   {"report_type": ["THUCHI"], "man_hinh": "Dòng tiền · Tổng quan"},
    "03B_SODU_TIEN": {"report_type": ["SDT"], "man_hinh": "Dòng tiền (số dư tiền) · Tổng quan"},
    "04_VAY":        {"report_type": ["VAY"], "man_hinh": "Dòng tiền – Vay"},
    "05_PHAITHU":    {"report_type": ["PTHU"], "man_hinh": "Công nợ · Tổng quan"},
    "06_PHAITRA":    {"report_type": ["PTRA"], "man_hinh": "Công nợ"},
    "07_TAISAN_NV":  {"report_type": ["TSNV", "BS"], "man_hinh": "Tài sản – Nguồn vốn (Finance) · Tổng quan"},
    "08_TSCD":       {"report_type": ["TS"], "man_hinh": "Tài sản · Tổng quan"},
    "09_TONKHO":     {"report_type": ["HH"], "man_hinh": "Tồn kho · Tổng quan"},
    "10_THUE":       {"report_type": ["THUE"], "man_hinh": "Thuế"},
    "11_DAUTU":      {"report_type": ["DTU"], "man_hinh": "Đầu tư"},
    "12_KDVH":       {"report_type": ["KDVH"], "man_hinh": "Vận hành KDVH"},
}
# Tên chỉ tiêu 01_HQKD phải đặt ĐÚNG để backend nhận diện -> KPI sáng.
KQKD_CANONICAL = {
    "Doanh thu thuần": "-> KPI Doanh thu (report_type DTHU/HQKD 1000)",
    "Tổng chi phí": "-> KPI Chi phí (HQKD 1047)",
    "Lợi nhuận trước thuế": "-> KPI LNTT (HQKD 1112)",
}

# Chuẩn hoá chỉ tiêu KQKD (công thức từ 'Nguồn dữ liệu lên Dashboard.xlsx'):
#   Tổng chi phí = Giá vốn + CP tài chính + CP bán hàng + CP QLDN + CP khác
#   LN trước thuế = dòng "Tổng lợi nhuận kế toán trước thuế" của báo cáo.
# -> template_filler tự SINH 2 dòng chuẩn "Tổng chi phí" & "Lợi nhuận trước thuế" khi điền 01_HQKD.
KQKD_COST_PATTERNS = ["gia von", "chi phi tai chinh", "chi phi lai vay",
                      "chi phi ban hang", "chi phi quan ly", "chi phi khac"]
# LNTT: chỉ tiêu chứa CẢ "loi nhuan" và "truoc thue" (vd "tong loi nhuan ke toan truoc thue").
KQKD_LNTT_REQUIRE = ("loi nhuan", "truoc thue")


def guide() -> dict:
    """Bản MÔ TẢ TEMPLATE VÀNG đầy đủ cho analyst: đơn vị, cột nhập vs auto, cost center,
    chỉ tiêu chuẩn, chọn sheet. Analyst đọc cái này TRƯỚC khi dựng mapping."""
    c = load_contract()
    sheets = {n: {"muc_dich": (s["title"] or "")[:130], "grain": s["grain"],
                  "report_type": SHEET_WIRING.get(n, {}).get("report_type", []),
                  "man_hinh_FE": SHEET_WIRING.get(n, {}).get("man_hinh", "(chưa nối FE)"),
                  "cot_nhap_lieu": s.get("input_columns", s["columns"]),
                  "cot_KHONG_map": s.get("auto_dim_columns", []),   # VLOOKUP cty/khối auto
                  "cot_tinh_toan_dien_neu_nguon_co": s.get("calc_columns", [])}  # Dư cuối/%...: map đè nếu nguồn có
              for n, s in c["data_sheets"].items()}
    ccs = sorted({e["cc"] for e in c["costcenter_index"].values() if e.get("cc")})
    return {
        "don_vi": "TỶ ĐỒNG. Nguồn tính bằng VND -> BẮT BUỘC truyền value_scale=1e-9 khi gọi template_fill.",
        "quy_tac": [
            "Map vào 'cot_nhap_lieu'. TUYỆT ĐỐI KHÔNG map 'cot_KHONG_map' (VLOOKUP cty/khối auto từ CC).",
            "'cot_tinh_toan_dien_neu_nguon_co' (Dư cuối kỳ, %...): nếu NGUỒN có sẵn giá trị đó (vd cột 'Cuối kỳ') thì PHẢI map đè vào (file điền không tự tính công thức -> để trống importer đọc = 0).",
            "Cost center: sheet nguồn cấp CÔNG TY (không có CC theo dòng) -> dùng constants={'Mã Cost center ◀ NHẬP':'<mã CC hợp lệ>'}; cty/khối tự suy từ CC.",
            "Sheet 01_HQKD: map cột 'Chỉ tiêu KQKD' từ cột chỉ tiêu nguồn. ĐỂ KPI Doanh thu/Chi phí/LNTT SÁNG, phải dùng rename_rows đổi ĐÚNG tên dòng nguồn về CHUẨN: dòng tổng doanh thu (cấp cao nhất) -> 'Doanh thu thuần'; dòng tổng chi phí -> 'Tổng chi phí'; dòng lợi nhuận trước thuế -> 'Lợi nhuận trước thuế'. Chọn 1 dòng TỔNG/chỉ tiêu (KHÔNG chọn dòng con) để khỏi cộng trùng. Nếu nguồn KHÔNG có dòng tổng chi phí/LNTT sẵn, hệ thống tự sinh từ các dòng chi phí rời.",
            "Chọn sheet đích theo loại báo cáo nguồn (xem sheet_theo_loai / canonical_kind).",
            "Kỳ: tháng 'YYYY-MM', ngày 'YYYY-MM-DD'. Grain mỗi sheet xem 'grain'.",
        ],
        "sheet_theo_loai": SHEET_FOR_KIND,
        "chi_tieu_KQKD_chuan": KQKD_CANONICAL,
        "cong_ty": c["md_congty"],
        "khoi": c["md_khoikd"],
        "cost_center_ma_hop_le": ccs,
        "sheets": sheets,
    }


def resolve_company(raw=None, file_name: str = None) -> str | None:
    """Suy MÃ CÔNG TY HỢP LỆ (CHỈ trong MD_CONGTY của Cty_Khoi_Costcenter/template).
    Ưu tiên: (1) raw đã là mã công ty hợp lệ (vd 'AAG','TC'); (2) mã trong tên file
    'B.<n>.<MÃ>.' nếu hợp lệ (vd 'B.7.AAG.'->AAG, 'B.4.TC.'->TC). Trả None nếu KHÔNG suy
    được mã hợp lệ — KHÔNG bịa folder-token ('ANTAXI'/'DUAN') thành công ty."""
    import re
    md = load_contract()["md_congty"]
    valid = {str(k).strip().upper(): str(k).strip() for k in md}
    if raw:
        r = str(raw).strip().upper()
        if r in valid:
            return valid[r]
    if file_name:
        m = re.search(r"\bB\.\d+\.([A-Za-z_]+)\.", str(file_name))
        if m and m.group(1).strip().upper() in valid:
            return valid[m.group(1).strip().upper()]
    return None


# Thư mục nguồn (Khối KD) -> Mã Khối (MD_KHOIKD). HIỆU QUẢ KD gán theo Khối suy từ ĐƯỜNG DẪN nguồn
# (khớp ngữ cảnh bo_sung_nguoi_dung.txt). Dùng khi nguồn không có cost center theo dòng (vd KQKD
# showroom cấp công ty) -> stamp khoi cho dòng khoi NULL.
_PATH_KHOI = {
    "SRVF": "8", "XDV": "5", "TRAMSAC": "4", "TRAM SAC": "4", "DUAN": "3", "DU AN": "3",
    "HO": "9", "XANHVINHPHUC": "6", "XANH VINH PHUC": "6", "ANTAXI": "6", "AN TAXI": "6",
    "ANKHACHSAN": "2", "AN KHACH SAN": "2", "HUNGTHINH": "7", "HUNG THINH": "7",
}


def khoi_from_path(path: str) -> str | None:
    """Suy TÊN KHỐI (MD_KHOIKD) từ đường dẫn nguồn bằng cách khớp CHÍNH XÁC từng SEGMENT thư mục
    (vd .../THINHCUONG/BAOCAOTAICHINH/SRVF/... -> 'Khối KD Xe điện Vinfast - SR'). None nếu không khớp."""
    if not path:
        return None
    import re
    segs = [s.strip().upper().replace(" ", "").replace("_", "") for s in re.split(r"[\\/]+", str(path)) if s.strip()]
    norm_map = {k.replace(" ", "").replace("_", ""): v for k, v in _PATH_KHOI.items()}
    md = load_contract()["md_khoikd"]   # {ma_khoi: ten_khoi}
    for s in segs:
        if s in norm_map:
            return md.get(norm_map[s], norm_map[s])   # trả TÊN khối (không phải mã)
    return None


def resolve_costcenter(raw) -> dict | None:
    """Suy (cc, cong_ty, khoi) từ 1 chuỗi nguồn: mã CC, tên CC, hoặc tên bộ phận.

    exact(mã) -> exact(tên) -> alias DM -> fuzzy(cutoff .86). None nếu không khớp.
    """
    if raw in (None, ""):
        return None
    c = load_contract()
    idx = c["costcenter_index"]
    n = _norm(raw)
    if not n:
        return None
    if n in idx:
        return {**idx[n], "matched_by": "exact"}
    alias = c["cc_aliases"].get(n)
    if alias and _norm(alias) in idx:
        return {**idx[_norm(alias)], "matched_by": "alias"}
    keys = list(idx.keys())
    hit = difflib.get_close_matches(n, keys, n=1, cutoff=0.86)
    if hit:
        return {**idx[hit[0]], "matched_by": f"fuzzy:{hit[0]}"}
    return None
