# -*- coding: utf-8 -*-
"""SOURCE CATALOG — index (lossless pointer) mọi file xlsx đã kéo về (Connect_VPS/received_reports).

Bronze layer = FILE TRÊN ĐĨA (không nhân bản vào DB). Catalog chỉ lưu con trỏ + cấu trúc
(sheet/cột/số dòng/canonical_kind) để QA/analyst biết "có file/sheet/cột nào" tức thì, kể cả
file CHƯA import vào raw_rows. Chi tiết ô đọc on-demand bằng source_inspect.

Index dựng lúc file land (P3). Truy vấn qua catalog_search (qa_server).
"""
import glob
import hashlib
import json
import os
from datetime import datetime, timezone

from openpyxl import load_workbook

from . import be_bridge as bb
from . import canonical

_HERE = os.path.dirname(os.path.abspath(__file__))
_AGENT_ROOT = os.path.normpath(os.path.join(_HERE, "..", ".."))

# Thư mục file kéo về (đổi tên Connect_VPN -> Connect_VPS). Override qua env RECEIVED_DIR.
RECEIVED_DIR = os.environ.get("RECEIVED_DIR") or os.path.normpath(
    os.path.join(_AGENT_ROOT, "..", "Connect_VPS", "received_reports"))
CATALOG = os.path.join(_AGENT_ROOT, "memory", "source_catalog.json")


def _norm(s) -> str:
    return bb.remove_diacritics("" if s is None else str(s)).strip().lower()


def _load() -> dict:
    if not os.path.exists(CATALOG):
        return {}
    try:
        with open(CATALOG, encoding="utf-8") as fh:
            return json.load(fh)
    except (OSError, json.JSONDecodeError):
        return {}


def _save(cat: dict):
    os.makedirs(os.path.dirname(CATALOG), exist_ok=True)
    with open(CATALOG, "w", encoding="utf-8") as fh:
        json.dump(cat, fh, ensure_ascii=False, indent=2)


def _file_key(path: str) -> str:
    return hashlib.sha1(os.path.abspath(path).encode("utf-8")).hexdigest()[:16]


def _sidecar(path: str) -> dict:
    """File .json cùng tên (do receiver ghi) chứa company/month/report_type."""
    j = os.path.splitext(path)[0] + ".json"
    if os.path.exists(j):
        try:
            with open(j, encoding="utf-8") as fh:
                return json.load(fh)
        except (OSError, json.JSONDecodeError):
            pass
    return {}


def _from_path(path: str) -> dict:
    """Suy company/report_type từ cấu trúc received_reports/{company}/{report_type}/file."""
    rel = os.path.relpath(path, RECEIVED_DIR)
    parts = rel.split(os.sep)
    return {"company": parts[0] if len(parts) >= 3 else None,
            "report_type": parts[1] if len(parts) >= 3 else None}


def index_file(path: str) -> dict:
    """Index 1 file xlsx -> entry {file, path, company, report_type, month, sheets:[...], ...}."""
    side = _sidecar(path)
    meta = _from_path(path)
    entry = {
        "file": os.path.basename(path),
        "path": os.path.abspath(path),
        "company": side.get("company") or meta.get("company"),
        "report_type": side.get("report_type") or meta.get("report_type"),
        "month": side.get("month"),
        "period_type": side.get("period_type"),
        "sheets": [],
        "mtime": os.path.getmtime(path),
        "indexed_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "ingested": False,
    }
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            # CHỈ đọc ~30 dòng đầu để lấy header (KHÔNG duyệt hết — sheet BCTC có thể tới
            # cả triệu dòng phantom -> duyệt hết là treo). Số dòng lấy từ ws.max_row.
            header = []
            for r in ws.iter_rows(min_row=1, max_row=30, values_only=True):
                if sum(1 for c in r if c not in (None, "")) >= 2:
                    header = [("" if c is None else str(c).strip()) for c in r]
                    break
            entry["sheets"].append({
                "name": ws.title,
                "columns": [h for h in header if h][:40],
                "nrows": ws.max_row,          # xấp xỉ (dimension Excel), tránh duyệt toàn bộ
                "canonical_kind": canonical.guess_canonical_kind(ws.title),
            })
    finally:
        wb.close()
    cat = _load()
    key = _file_key(path)
    if key in cat:                      # giữ cờ ingested nếu đã có
        entry["ingested"] = cat[key].get("ingested", False)
    cat[key] = entry
    _save(cat)
    return entry


def index_dir(root: str = None) -> dict:
    """Quét thư mục received_reports, index mọi .xlsx (bỏ file tạm ~$)."""
    root = root or RECEIVED_DIR
    if not os.path.isdir(root):
        return {"ok": False, "error": f"Không thấy thư mục: {root}", "indexed": 0}
    files = [f for f in glob.glob(os.path.join(root, "**", "*.xlsx"), recursive=True)
             if not os.path.basename(f).startswith("~$")]
    cat = _load()
    done = 0
    for f in files:
        try:
            key = _file_key(f)
            prev = cat.get(key)
            # BỎ QUA file đã index & chưa đổi (mtime khớp) -> tránh mở lại file lớn (17MB) mỗi lần.
            if prev and prev.get("mtime") == os.path.getmtime(f):
                continue
            index_file(f)
            done += 1
        except Exception:  # 1 file hỏng không chặn cả mẻ
            pass
    return {"ok": True, "scanned": len(files), "indexed_new": done, "total_in_catalog": len(_load())}


def mark_ingested(path: str, ingested: bool = True):
    cat = _load()
    key = _file_key(path)
    if key in cat:
        cat[key]["ingested"] = ingested
        _save(cat)


def search(query: str = None, company: str = None, canonical_kind: str = None,
           sheet: str = None, only_uningested: bool = False) -> list:
    """Tìm trong catalog (không mở file). Lọc theo tên/công ty/canonical_kind/sheet."""
    q, cmp_ = _norm(query), _norm(company)
    ck, sh = _norm(canonical_kind), _norm(sheet)
    out = []
    for e in _load().values():
        if only_uningested and e.get("ingested"):
            continue
        if cmp_ and cmp_ not in _norm(e.get("company")):
            continue
        hay = _norm(e.get("file")) + " " + _norm(e.get("company")) + " " + _norm(e.get("report_type"))
        sheets_norm = " ".join(_norm(s["name"]) for s in e.get("sheets", []))
        cks = " ".join(_norm(s.get("canonical_kind")) for s in e.get("sheets", []) if s.get("canonical_kind"))
        if q and q not in hay and q not in sheets_norm:
            continue
        if ck and ck not in cks:
            continue
        if sh and sh not in sheets_norm:
            continue
        out.append(e)
    return out
