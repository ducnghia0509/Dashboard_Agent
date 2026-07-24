# -*- coding: utf-8 -*-
"""MCP server luong 1 (ingest): template_analyze, discovery_record, import_plan_validate,
import_execute. Chay: `python -m servers.ingest_server` (stdio MCP server).

Tai su dung nguyen logic nghiep vu cua DashBoard_AI qua be_bridge - khong viet lai
detect/auto_map/commit/import_workbook."""
import hashlib
import io
import json
import os
import re

from mcp.server.fastmcp import FastMCP

from .common import be_bridge as bb
from .common import canonical
from .common import source_catalog as canonical_SC
from .common import contract
from .common import extraction
from .common import memory
from . import template_filler

mcp = FastMCP("dashboard_ingest")

_HERE = os.path.dirname(os.path.abspath(__file__))
_AGENT_ROOT = os.path.normpath(os.path.join(_HERE, ".."))
_WORKSPACE_ROOT = os.path.dirname(_AGENT_ROOT)  # thư mục cha (vd d:\Company\ThinhCuong) - chứa
                                                  # DashBoard_Agent/DashBoard_AI/Data_test_dashboard
_IMPORTS_LEDGER = os.path.join(_AGENT_ROOT, "memory", "imports_ledger.json")
_GEN_MONEY_SANITY_MAX_TY = 1_000_000.0  # khớp template_filler._MONEY_SANITY_MAX_TY


def _input_dir() -> str:
    input_dir = os.environ.get("INPUT_DIR") or "."
    return os.path.normpath(os.path.join(_AGENT_ROOT, input_dir))


def _resolve_path(file_path: str) -> str:
    if not os.path.isabs(file_path):
        file_path = os.path.join(_input_dir(), file_path)
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Không tìm thấy file: {file_path}")
    return file_path


def _read_file(file_path: str) -> bytes:
    with open(_resolve_path(file_path), "rb") as fh:
        return fh.read()


def _jsonable(v):
    """Excel cell value -> kiểu JSON-serializable an toàn (datetime/Decimal -> str)."""
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    return str(v)


def _all_sheets_columns(data: bytes) -> dict:
    wb = bb.fast_load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    try:
        out = {}
        for ws in wb.worksheets:
            first_row = next(ws.iter_rows(max_row=1, values_only=True), [])
            out[ws.title] = [("" if c is None else str(c)) for c in first_row]
        return out
    finally:
        wb.close()


@mcp.tool()
def discover_files(root_dir: str = None, pattern: str = r".*\.xlsx$") -> dict:
    """Quét thư mục tìm workbook Excel — dùng khi chưa biết cần đọc file nào (nhiều
    công ty/nhiều file). RẺ: chỉ liệt kê tên file + đoán company/canonical_kind từ
    tên thư mục/tên file, KHÔNG mở nội dung workbook (mở file 17MB tốn 10-15s/lần,
    xem sheet_profile/generic_import_execute cho bước đọc nội dung).

    root_dir=None: mặc định quét INPUT_DIR. Có thể trỏ nơi khác nhưng PHẢI nằm trong
    workspace (cùng thư mục cha với DashBoard_Agent) — chặn quét ra ngoài phạm vi dự án.

    company_guess: nếu file nằm trong thư mục con của root_dir, dùng TÊN thư mục con đó;
    nếu không, thử regex trên tên file dạng 'X.N.MÃCTY.rest...' (vd 'B.7.AAG.TCKT...' -> 'AAG').
    canonical_kind_guess: đoán loại báo cáo từ TÊN FILE (xem servers/common/canonical.py) —
    file lớn/nhiều sheet (vd báo cáo tài chính riêng tổng hợp) thường không đoán được từ tên
    file mà phải gọi sheet_profile(sheet=None) để xem danh sách sheet bên trong."""
    base = os.path.normpath(os.path.join(_AGENT_ROOT, root_dir)) if root_dir else _input_dir()
    if os.path.commonpath([_WORKSPACE_ROOT, base]) != _WORKSPACE_ROOT:
        raise ValueError(f"'{root_dir}' nằm ngoài workspace — không được phép quét.")
    if not os.path.isdir(base):
        raise FileNotFoundError(f"Thư mục không tồn tại: {base}")

    rx = re.compile(pattern, re.IGNORECASE)
    code_rx = re.compile(r"^[A-Za-z]\.\d+\.([A-Za-z0-9]{2,8})\.")
    files = []
    for dirpath, _dirnames, filenames in os.walk(base):
        rel_dir = os.path.relpath(dirpath, base)
        for fn in filenames:
            if not rx.match(fn):
                continue
            if rel_dir == ".":
                company_guess = None
                m = code_rx.match(fn)
                if m:
                    company_guess = m.group(1)
            else:
                # B21/B30: KHÔNG dùng thẳng segment path làm mã công ty (segment là tên folder
                # phân loại báo cáo do sender đặt — vd 'THUCHI'/'HO' — không phải pháp nhân, có
                # lúc còn lẫn nhiều công ty trong 1 folder). Validate qua resolve_company(), ƯU
                # TIÊN tên FILE (quy ước 'B.<khối>.<mã cty>.', đáng tin hơn folder) qua
                # prefer_file_name=True; không khớp gì cả -> None, để caller/analyst tự xác định.
                company_guess = contract.resolve_company(raw=rel_dir.split(os.sep)[0], file_name=fn,
                                                           prefer_file_name=True)
            rel_path = os.path.normpath(os.path.join(rel_dir, fn)) if rel_dir != "." else fn
            full_path = os.path.join(dirpath, fn)
            files.append({
                "path": rel_path.replace("\\", "/"),
                "file_name": fn,
                "company_guess": company_guess,
                "canonical_kind_guess": canonical.guess_canonical_kind(fn),
                "size_bytes": os.path.getsize(full_path),
            })
    companies = sorted({f["company_guess"] for f in files if f["company_guess"]})
    return {"root_dir": base, "companies": companies, "files": files}


@mcp.tool()
def sheet_profile(file_path: str, sheet: str = None, max_rows: int = 10,
                   max_cols: int = 8, col_depth: int = 30) -> dict:
    """Khám phá cấu trúc THÔ của 1 workbook/sheet khi template_analyze không nhận diện
    được report_type cố định nào (report_type=None). Dùng khi cần tự suy luận layout
    của 1 sheet lạ (khác 9 báo cáo THUCHI/SDT/HQKD/... đã biết).

    sheet=None: chỉ liệt kê tên các sheet trong workbook (rẻ, không đọc nội dung), kèm
    canonical_kind_guess đoán từ TÊN sheet (xem servers/common/canonical.py) để analyst
    biết ngay sheet nào đáng đào sâu trước khi đọc nội dung.
    sheet=<tên>: đọc 1 lượt sheet đó, trả 2 góc nhìn để tự nhận biết layout:
      - row_sample: max_rows dòng đầu, MỌI cột - phát hiện header nằm ngang/nhiều dòng
        (vd 2 dòng header như sheet '131'/'331': 'Đầu kỳ|Phát sinh|Cuối kỳ' rồi 'Nợ|Có').
      - col_sample: max_cols cột đầu, mỗi cột lấy col_depth dòng - phát hiện layout
        "thực thể nằm theo cột" (vd nhiều công ty làm cột như CĐKT hợp nhất).
      - merged_ranges: vùng ô merge (gợi ý ranh giới header nhiều dòng/nhiều cột).
    KHÔNG ghi gì, không cần DB - chỉ đọc file trong INPUT_DIR (qua _read_file)."""
    data = _read_file(file_path)
    wb = bb.fast_load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    try:
        if sheet is None:
            return {
                "file_name": os.path.basename(file_path),
                "sheets": [{"sheet": s, "canonical_kind_guess": canonical.guess_canonical_kind(s)}
                           for s in wb.sheetnames],
            }

        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' không tồn tại. Các sheet có: {wb.sheetnames}")
        ws = wb[sheet]

        try:
            dimensions = ws.dimensions
        except Exception:
            dimensions = None

        try:
            merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        except Exception:
            merged_ranges = []  # read_only worksheet có thể không có merged_cells

        depth = max(max_rows, col_depth)
        rows = []
        for row in ws.iter_rows(max_row=depth, values_only=True):
            rows.append([_jsonable(c) for c in row])

        row_sample = rows[:max_rows]
        col_sample = {}
        for ci in range(min(max_cols, max((len(r) for r in rows), default=0))):
            from openpyxl.utils import get_column_letter
            col_sample[get_column_letter(ci + 1)] = [
                (r[ci] if ci < len(r) else None) for r in rows[:col_depth]
            ]

        return {
            "file_name": os.path.basename(file_path), "sheet": sheet, "dimensions": dimensions,
            "merged_ranges": merged_ranges[:50], "row_sample": row_sample, "col_sample": col_sample,
            "n_rows_scanned": len(rows),
        }
    finally:
        wb.close()


@mcp.tool()
def sheet_routes(file_path: str) -> dict:
    """Try-route TẤT ĐỊNH cả workbook (mở 1 LẦN): phân loại MỖI sheet -> đích + trạng thái.
    Khác sheet_profile(sheet=None): CÓ đọc tiêu đề trong sheet để route theo NỘI DUNG (bắt tên
    sheet khó như '156'='BÁO CÁO NHẬP XUẤT TỒN', 'KHTS'='BẢNG TÍNH KHẤU HAO') và KHÔNG bỏ sót
    im lặng — mỗi sheet rơi vào đúng 1 nhóm:
      - routed        : target_sheet + canonical_kind + route_via ('name'|'content').
      - unknown       : có dữ liệu nhưng chưa nhận diện loại -> caller đẩy analyst/người.
      - skip_metadata : Master Data / User / DS báo cáo / BC THU CHI (TỔNG)...
      - empty         : không đủ dòng dữ liệu."""
    data = _read_file(file_path)
    wb = bb.fast_load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    try:
        out = []
        for name in wb.sheetnames:
            rows = [row for row in wb[name].iter_rows(max_row=25, values_only=True)]
            non_empty = sum(1 for r in rows if any(c not in (None, "") for c in r))
            # Tiêu đề = ô ĐẦU TIÊN không rỗng của mỗi dòng trong ~8 dòng đầu (cột tiêu đề/merge),
            # KHÔNG lấy cả lưới -> tránh keyword vãng lai trong header/ô số khiến route nhầm.
            title_cells = []
            for r in rows[:8]:
                cell = next((c for c in r if c not in (None, "")), None)
                if cell is not None:
                    title_cells.append(str(cell))
            target, ck, via = contract.route_sheet(name, " ".join(title_cells))
            if via == "skip":
                status = "skip_metadata"
            elif target and not contract.role_allows(file_path, target):
                # Sheet thuộc vai file KHÁC (ma trận Book1) — không đưa vào need_llm/autofill
                # để khỏi tốn LLM vào sheet mà fill_from_source sẽ chặn.
                status = "skip_role"
                target = None
            elif target:
                status = "routed"
            elif non_empty >= 4:
                status = "unknown"
            else:
                status = "empty"
            out.append({"sheet": name, "canonical_kind": ck, "target_sheet": target,
                        "route_via": (via if via in ("name", "content") else None),
                        "status": status})
        return {"file_name": os.path.basename(file_path), "routes": out}
    finally:
        wb.close()


@mcp.tool()
def template_analyze(file_path: str) -> dict:
    """Phân tích 1 file xlsx chưa biết template: dò report_type/header/mapping,
    liệt kê toàn bộ sheet+cột, ghi vào discovery memory. Trả về TemplateSpec (dict)."""
    data = _read_file(file_path)
    file_name = os.path.basename(file_path)
    sheets_columns = _all_sheets_columns(data)

    try:
        prep = bb.prepare(data, file_name)
    except ValueError as e:
        # Không nhận diện được report_type - vẫn ghi discovery để qa biết đã từng thấy file này.
        record = memory.discovery_record(
            file_name=file_name, fingerprint="", sheets=list(sheets_columns.keys()),
            columns_per_sheet=sheets_columns, detected_report_type=None,
            anomalies=[str(e)],
        )
        return {
            "file_name": file_name, "report_type": None, "header_row": 0, "headers": [],
            "mapping": {}, "confidence": 0.0, "missing_required": [], "low_confidence_fields": [],
            "anomalies": [str(e)], "sample_rows": [], "all_sheets": list(sheets_columns.keys()),
        }

    confidence = 1.0 if prep["from_profile"] else (0.0 if prep["missing_required"] else 0.7)
    if prep["low_confidence"]:
        confidence = min(confidence, 0.5)
    fp = prep["fingerprint"]

    memory.discovery_record(
        file_name=file_name, fingerprint=fp, sheets=list(sheets_columns.keys()),
        columns_per_sheet=sheets_columns, detected_report_type=prep["report_type"],
        header_row=prep["header_row"], mapping=prep["mapping"], confidence=confidence,
        anomalies=([f"Thiếu field bắt buộc: {prep['missing_required']}"] if prep["missing_required"] else []),
    )

    return {
        "file_name": file_name,
        "report_type": prep["report_type"],
        "header_row": prep["header_row"],
        "headers": prep["headers"],
        "mapping": prep["mapping"],
        "confidence": confidence,
        "missing_required": prep["missing_required"],
        "low_confidence_fields": prep["low_confidence"],
        "anomalies": [],
        "sample_rows": prep["sample_rows"],
        "all_sheets": list(sheets_columns.keys()),
        "fingerprint": fp,
    }


@mcp.tool()
def discovery_record_tool(
    file_name: str, fingerprint: str, sheets: list, columns_per_sheet: dict,
    detected_report_type: str = None, header_row: int = None, mapping: dict = None,
    period: str = None, confidence: float = 0.0, anomalies: list = None,
) -> dict:
    """Ghi thủ công 1 bản ghi vào discovery memory (dùng khi agent muốn bổ sung
    ghi chú/period sau khi đã template_analyze)."""
    return memory.discovery_record(
        file_name, fingerprint, sheets, columns_per_sheet, detected_report_type,
        header_row, mapping, period, confidence, anomalies,
    )


@mcp.tool()
def import_plan_validate(
    file_name: str, report_type: str, mapping: dict, period: str = None,
    dataset_kind: str = "day",
) -> dict:
    """Kiểm ImportPlan trước khi ghi: field bắt buộc theo FIELD_DEFS, report_type hợp lệ,
    period hợp lệ (YYYY-MM) nếu kind=month. KHÔNG ghi DB."""
    errors, warnings = [], []
    if report_type not in bb.FIELD_DEFS:
        errors.append(f"report_type '{report_type}' không hợp lệ (phải là 1 trong {bb.REPORT_CODES}).")
        return {"ok": False, "errors": errors, "warnings": warnings}

    fields = bb.FIELD_DEFS[report_type]
    missing = [f[0] for f in fields if f[2] and f[0] not in mapping]
    if missing:
        errors.append(f"Thiếu mapping cho field bắt buộc: {missing}")

    if dataset_kind == "month":
        import re
        if not period or not re.match(r"^\d{4}-\d{2}$", period):
            warnings.append("period không đúng định dạng YYYY-MM hoặc bị thiếu (sẽ tự suy từ tên file/dữ liệu).")
    elif dataset_kind not in ("day", "month"):
        errors.append("dataset_kind phải là 'day' hoặc 'month'.")

    return {"ok": not errors, "errors": errors, "warnings": warnings}


def _ledger_load() -> list:
    if not os.path.exists(_IMPORTS_LEDGER):
        return []
    try:
        with open(_IMPORTS_LEDGER, encoding="utf-8") as fh:
            return json.load(fh)
    except (OSError, json.JSONDecodeError):
        return []


def _ledger_save(entries: list):
    memory.atomic_dump_json(entries, _IMPORTS_LEDGER)


def _ledger_append(entry: dict):
    """Thêm 1 bản ghi vào sổ import: lock cả chu trình load-append-save (chống lost-update
    khi nhiều import chạy song song) + ghi atomic (chống rách file làm mất cả sổ dedup)."""
    with memory.locked_json(_IMPORTS_LEDGER):
        entries = _ledger_load()
        entries.append(entry)
        _ledger_save(entries)


def ledger_remove_by_content_hash(content_hashes) -> int:
    """Xoá MỌI bản ghi imports_ledger có content_hash thuộc tập cho trước — dùng khi XOÁ HẲN 1 file
    để MỞ KHOÁ dedup (generic_import_execute/import_execute chặn nạp lại theo content_hash), cho
    phép phân tích lại. content_hash = sha1(bytes file) phân biệt ĐÚNG từng file (kể cả trùng
    basename như 3 nguồn XVP). KHÔNG đụng fill_specs/report_specs (kiến thức layout dùng chung).
    Lock cả chu trình + ghi atomic. Trả số bản ghi đã xoá."""
    hs = {h for h in (content_hashes or []) if h}
    if not hs:
        return 0
    with memory.locked_json(_IMPORTS_LEDGER):
        entries = _ledger_load()
        keep = [e for e in entries if e.get("content_hash") not in hs]
        removed = len(entries) - len(keep)
        if removed:
            _ledger_save(keep)
    return removed


_ENTITY_CODE_HINTS = ("ma_doi_tuong", "ma_dt", "ma_kh", "ma_ncc", "ma_khach", "code", "ma")
_ENTITY_NAME_HINTS = ("ten_doi_tuong", "ten_dt", "ten_kh", "ten_ncc", "ten_khach", "name", "ten", "dien_giai")


def _normalize_columns(columns: list) -> list:
    """Chuẩn hoá 'columns' về schema ColumnRole {index, role, label} — khoan dung với biến thể
    mà LLM hay sinh: 'col_letter' (A/B/..) -> index; 'field_name' -> role nếu thiếu 'role'.
    Nếu role không phải từ khoá đặc biệt nhưng field_name gợi ý mã/tên đối tượng thì map về
    entity_code/entity_name (nếu chưa có cột nào giữ vai trò đó) để melt gắn đúng thực thể."""
    from openpyxl.utils import column_index_from_string

    norm = []
    have_code = have_name = False
    for c in columns:
        idx = c.get("index")
        if idx is None and c.get("col_letter"):
            idx = column_index_from_string(str(c["col_letter"]).strip().upper()) - 1
        if idx is None:
            continue
        role = c.get("role") or c.get("field_name") or ""
        label = c.get("label", "")
        norm.append({"index": int(idx), "role": role, "label": label})

    reserved = {"entity_code", "entity_name", "label", "skip"}
    have_code = any(c["role"] == "entity_code" for c in norm)
    have_name = any(c["role"] == "entity_name" for c in norm)
    for c in norm:
        if c["role"] in reserved:
            continue
        low = c["role"].lower()
        if not have_code and any(low == h or low.startswith(h) for h in _ENTITY_CODE_HINTS):
            c["role"] = "entity_code"; have_code = True
        elif not have_name and any(low == h or low.startswith(h) for h in _ENTITY_NAME_HINTS):
            c["role"] = "entity_name"; have_name = True
    return norm


def _prior_dataset_ids(dataset_kind: str, file_name: str, exclude_id=None) -> list:
    """Các dataset_id đã import trước đó của CÙNG file (theo ledger) — phạm vi dọn B26.
    KHÔNG dọn theo (kind, period) toàn kỳ như trước: nhánh month toàn-kỳ sẽ xoá nhầm cả
    dataset template dùng chung nhiều công ty của import_filled (cùng kind='month', cùng
    period); nhánh day toàn-kỳ xoá lây báo cáo ngày khác cùng tháng."""
    return [e["dataset_id"] for e in _ledger_load()
            if e.get("dataset_kind") == dataset_kind and e.get("file_name") == file_name
            and e.get("dataset_id") and e.get("dataset_id") != exclude_id]


def _ledger_find(dataset_kind, report_type, period, fingerprint, content_hash=None,
                 match_fingerprint=False):
    """Tìm bản ghi import trùng. Khoá dedup = (dataset_kind, content_hash) — content_hash
    (sha1 nội dung file) định danh file duy nhất: CÙNG nội dung file -> coi là trùng, bỏ qua.

    match_fingerprint=True (dùng cho GENERIC nạp TỪNG SHEET): phải trùng CẢ content_hash LẪN
    fingerprint (fingerprint gồm sheet+layout) -> sheet thứ 2/3 của CÙNG file KHÔNG bị coi trùng
    với sheet đầu (fix B6). Với day/month (cả file) giữ nguyên chỉ theo content_hash."""
    if not content_hash:
        return None
    for e in _ledger_load():
        if e.get("dataset_kind") != dataset_kind or e.get("content_hash") != content_hash:
            continue
        if match_fingerprint and e.get("fingerprint") != fingerprint:
            continue
        return e
    return None


@mcp.tool()
def import_execute(
    file_path: str, dataset_kind: str = "day", report_type: str = None,
    mapping: dict = None, period: str = None, dry_run: bool = True,
) -> dict:
    """Ghi dữ liệu tinh vào raw_rows (idempotent - kiểm memory/imports_ledger.json
    theo khoá (dataset_kind, report_type, period, fingerprint) trước khi ghi thật).

    dry_run=True (mặc định): chỉ preview số dòng dự kiến, KHÔNG ghi DB.
    dataset_kind='day' -> tái dùng importer.prepare/commit (9 báo cáo, cần report_type+mapping).
    dataset_kind='month' -> tái dùng importer_month.import_workbook (nhiều sheet, tự suy period).
    """
    data = _read_file(file_path)
    file_name = os.path.basename(file_path)

    if dataset_kind == "month":
        wb = bb.fast_load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        try:
            is_month = bb.detect_ledger(wb)
        finally:
            wb.close()
        if not is_month:
            return {"dry_run": dry_run, "by_type": {}, "skipped_duplicate": False,
                    "message": f"'{file_name}' không phải template Tháng (thiếu sheet DATA dạng sổ cái)."}
        fp = bb.fingerprint([file_name])  # fingerprint theo tên file cho kind=month (không có header cột chung)
        content_hash = hashlib.sha1(data).hexdigest()
        dup = _ledger_find("month", "MONTH", period, fp, content_hash)
        if dup and not dry_run:
            return {"dry_run": False, "dataset_id": dup["dataset_id"], "by_type": {},
                    "skipped_duplicate": True,
                    "message": f"Đã import trước đó (dataset_id={dup['dataset_id']}, period={dup.get('period')}). Bỏ qua để tránh trùng."}
        if dry_run:
            return {"dry_run": True, "by_type": {}, "skipped_duplicate": bool(dup),
                    "message": "Dry-run: sẽ gọi importer_month.import_workbook (số dòng thực tế biết sau khi ghi)."}
        ds = bb.new_dataset(kind="month", period=period, name=file_name)
        result = bb.import_workbook(ds["id"], data, file_name)
        pr = result.get("period") or period
        if result.get("period"):
            bb.repo.set_period(ds["id"], result["period"])
        # B26: dọn dataset mồ côi của CHÍNH file này (re-export khác bytes) -> không phình DB.
        # Scope theo file (tra ledger), KHÔNG delete_by_key toàn kỳ — tránh xoá chéo dataset
        # template đa-công-ty của import_filled cùng kind='month' cùng kỳ.
        for old_id in _prior_dataset_ids("month", file_name, exclude_id=ds["id"]):
            bb.repo.delete_dataset(old_id)
        _ledger_append({"dataset_kind": "month", "report_type": "MONTH", "period": result.get("period") or period,
                        "fingerprint": fp, "content_hash": content_hash, "dataset_id": ds["id"], "file_name": file_name})
        return {"dry_run": False, "dataset_id": ds["id"], "by_type": result["by_type"],
                "skipped_duplicate": False, "message": "Đã ghi raw_rows.", "period": result.get("period")}

    # dataset_kind == "day"
    try:
        prep = bb.prepare(data, file_name)
    except ValueError as e:
        return {"dry_run": dry_run, "by_type": {}, "skipped_duplicate": False, "message": str(e)}

    final_map = dict(prep["mapping"])
    if mapping:
        final_map.update({k: int(v) for k, v in mapping.items() if v is not None})
    rt = report_type or prep["report_type"]
    fp = prep["fingerprint"]
    content_hash = hashlib.sha1(data).hexdigest()

    dup = _ledger_find("day", rt, period, fp, content_hash)
    if dup and not dry_run:
        return {"dry_run": False, "dataset_id": dup["dataset_id"], "by_type": {},
                "skipped_duplicate": True,
                "message": f"File cùng report_type+fingerprint đã import trước đó (dataset_id={dup['dataset_id']}). Bỏ qua để tránh trùng."}

    if dry_run:
        n_rows = sum(1 for r in prep["rows"][prep["header_row"] + 1:] if r and any(c is not None for c in r))
        return {"dry_run": True, "by_type": {rt: n_rows}, "skipped_duplicate": bool(dup),
                "message": f"Dry-run: dự kiến ghi ~{n_rows} dòng report_type={rt}."}

    ds = bb.new_dataset(kind="day", period=period, name=file_name)
    result = bb.commit(ds["id"], rt, prep["header_row"], prep["headers"], final_map, prep["rows"],
                       source_file=file_name)  # B10
    # B26: dọn dataset day mồ côi của CHÍNH file này (scope theo file qua ledger — trước đây
    # delete_by_key("day", period=THÁNG) xoá lây mọi báo cáo ngày khác trong cùng tháng).
    for old_id in _prior_dataset_ids("day", file_name, exclude_id=ds["id"]):
        bb.repo.delete_dataset(old_id)
    _ledger_append({"dataset_kind": "day", "report_type": rt, "period": period,
                    "fingerprint": fp, "content_hash": content_hash, "dataset_id": ds["id"], "file_name": file_name})
    return {"dry_run": False, "dataset_id": ds["id"], "by_type": {rt: result["rows"]},
            "skipped_duplicate": False, "message": "Đã ghi raw_rows.", "issues": result["issues"]}


@mcp.tool()
def generic_import_execute(
    file_path: str, sheet: str, mapping: dict, period: str = None, ngay: str = None,
    cong_ty: str = None, dry_run: bool = True, value_scale: float = 1.0,
) -> dict:
    """Ghi dữ liệu theo 1 SheetMapping khai báo (xem servers/common/models.py SheetMapping) —
    dùng cho sheet KHÔNG khớp 9 report_type cố định (vd '131'/'331'/'Biểu khấu hao' trong file
    báo cáo tài chính riêng). KHÔNG sinh/chạy code tuỳ ý — chỉ diễn giải 'columns' (row_major)
    hoặc 'entities'+'row_roles' (column_major) theo đúng 1 thuật toán cố định: "melt" bảng rộng
    thành raw_rows dạng dài (tidy), tổng quát hoá pattern importer_month._parse_cdkt.

    mapping (dict) tối thiểu cần: orientation ('row_major'|'column_major'),
    target_report_type (PHẢI có tiền tố 'GEN_'), data_start_row, và columns[] hoặc
    entities[]+row_roles[] tương ứng orientation (xem models.SheetMapping). Có thể kèm
    canonical_kind (vd 'TK131') để report_spec_search tìm lại được ở công ty khác.

    cong_ty: tên công ty/đơn vị sở hữu file này (từ discover_files hoặc SheetMapping.company)
    — CHỈ dùng cho orientation='row_major' (mỗi file thường thuộc 1 công ty); với
    'column_major' công ty đã lấy từ tên thực thể (entities[].entity_name) nên bỏ qua tham số này.

    LUÔN trả sample_mapped_rows (5 dòng đầu đã map) bất kể dry_run=True/False, để người dùng
    soi trước khi orchestrator cho approve ghi thật (suy luận LLM, rủi ro map sai cột cao hơn
    auto_map cũ).

    value_scale: hệ số quy đổi ĐƠN VỊ VỀ TỶ ĐỒNG (khớp đơn vị hiển thị FE) — nguồn VND thì
    truyền value_scale=1e-9 (tool KHÔNG tự đổi). Có GUARD biên độ (scale_warning): nếu giá trị
    lớn nhất sau scale vẫn > 1 triệu tỷ, tool trả error và KHÔNG ghi thật (nghi quên value_scale).

    dry_run=True KHÔNG ghi DB, nhưng CÓ THỂ ghi 1 file spec học sớm (verified=False) vào
    memory/report_specs/ khi melt thành công không cảnh báo — cố ý (khép vòng học sớm,
    autobatch không phải gọi lại LLM cho layout đã dry-run đúng)."""
    target_rt = mapping.get("target_report_type") or ""
    if not target_rt.startswith("GEN_"):
        raise ValueError("target_report_type phải có tiền tố 'GEN_' để tách biệt report_type cố định.")
    # CHUẨN HOÁ mã GEN_ về dạng ổn định (single source: bảng alias canonical) — cùng khái niệm
    # LUÔN ra 1 mã dù LLM/heuristic đặt tên khác nhau ('GEN_10THUE'/'GEN_10_THUE'->'GEN_THUE',
    # 'GEN_PTRA'->'GEN_TK331'). Nhờ vậy tập chiều ổn định giữa các tháng và idempotent theo
    # (report_type, source_file) hoạt động đúng (2 lượt nạp cùng sheet -> xoá đè, không nạp đúp).
    target_rt = canonical.canonical_gen_code(target_rt)
    mapping = {**mapping, "target_report_type": target_rt}
    orientation = mapping.get("orientation")
    if orientation not in ("row_major", "column_major"):
        raise ValueError("orientation phải là 'row_major' hoặc 'column_major'.")

    # B22: VALIDATE cong_ty TRƯỚC KHI GHI DB — đây là đường ghi raw_rows KHÔNG qua
    # template_filler (không có resolve_company như import_filled), từng để lọt cong_ty='THUCHI'
    # (row_major, company_guess từ path sai) hoặc entity_name tự bịa (column_major) vào dữ liệu
    # thật (2026-07-09). KHÔNG âm thầm ghi mã không hợp lệ — chặn cứng, bắt caller sửa mapping.
    if orientation == "row_major" and cong_ty:
        resolved = contract.resolve_company(raw=cong_ty)
        if not resolved:
            raise ValueError(
                f"cong_ty='{cong_ty}' không khớp mã pháp nhân hợp lệ nào (MD_CONGTY). "
                f"KHÔNG được tự suy đoán/dùng token thư mục làm cong_ty — xác định lại pháp nhân "
                f"thật (đối chiếu companies.yaml / cost_center_ma_hop_le) trước khi ghi.")
        cong_ty = resolved
    elif orientation == "column_major":
        bad_entities = []
        for ent in mapping.get("entities", []):
            name = ent.get("entity_name")
            if name and not contract.resolve_company(raw=name):
                bad_entities.append(name)
        if bad_entities:
            raise ValueError(
                f"entities[].entity_name có giá trị không khớp mã pháp nhân hợp lệ nào (MD_CONGTY): "
                f"{sorted(set(bad_entities))}. Với column_major, entity_name CHÍNH LÀ cong_ty ghi vào "
                f"raw_rows — KHÔNG được tự bịa/suy đoán, phải là 1 trong companies.yaml.")

    data = _read_file(file_path)
    wb = bb.fast_load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' không tồn tại. Các sheet có: {wb.sheetnames}")
        rows = [[_jsonable(c) for c in r] for r in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()

    # ĐỊNH DANH NGUỒN '<công_ty_thư_mục>::<tên_file>' (khớp source_key UI) để idempotent-delete +
    # trạng thái tách ĐÚNG từng nguồn; file ngoài received_reports -> basename như cũ.
    file_name = canonical_SC.source_id_from_path(file_path)
    if "data_start_row" not in mapping:
        print(f"[generic_import_execute] WARN: mapping thiếu data_start_row cho sheet '{sheet}' "
              f"-> mặc định 0, có thể melt nhầm dòng header thành dữ liệu (B25).", flush=True)
    data_start = mapping.get("data_start_row", 0)
    ngay_val = ngay or (f"{period}-01" if period and len(period) == 7 else None)
    period_val = period or (ngay_val[:7] if ngay_val else None)

    def _blank_cell(v) -> bool:
        """Ô 'không có số liệu' (None/rỗng/'-' kế toán) — KHÁC với số 0 tường minh.
        Trước đây lọc bằng `if not val` sau parse_num nên số dư/phát sinh = 0 THẬT cũng bị
        bỏ khỏi raw_rows -> FE không phân biệt được '0' với 'chưa có dữ liệu'."""
        return v is None or (isinstance(v, str) and v.strip() in ("", "-", "–"))

    # (row_index_nguồn, entity_code, entity_name, role, khoi, cong_ty, amount, payload)
    melted = []
    if orientation == "row_major":
        cols = {c["index"]: c for c in _normalize_columns(mapping.get("columns", []))}
        code_idx = next((i for i, c in cols.items() if c["role"] == "entity_code"), None)
        name_idx = next((i for i, c in cols.items() if c["role"] == "entity_name"), None)
        for ri in range(data_start, len(rows)):
            row = rows[ri]
            if not row or all(v is None for v in row):
                continue
            entity_code = bb.parse_text(row[code_idx]) if code_idx is not None and code_idx < len(row) else None
            entity_name = bb.parse_text(row[name_idx]) if name_idx is not None and name_idx < len(row) else None
            row_payload = {c["role"]: bb.parse_text(row[idx]) for idx, c in cols.items()
                           if idx < len(row) and row[idx] is not None}
            for idx, c in cols.items():
                role = c["role"]
                if role in ("entity_code", "entity_name", "label", "skip") or idx >= len(row):
                    continue
                if _blank_cell(row[idx]):
                    continue
                val = bb.parse_num(row[idx])
                if value_scale != 1.0:
                    val = val * value_scale   # B20: VND -> tỷ khi value_scale=1e-9
                melted.append((ri, entity_code, entity_name, role, None, cong_ty, val, row_payload))
    else:  # column_major
        for rr in mapping.get("row_roles", []):
            ri = rr["row_index"]
            if ri >= len(rows):
                continue
            row = rows[ri]
            for ent in mapping.get("entities", []):
                ci = ent["col_index"]
                if ci >= len(row) or _blank_cell(row[ci]):
                    continue
                val = bb.parse_num(row[ci])
                if value_scale != 1.0:
                    val = val * value_scale   # B20
                payload = {"row_label": rr.get("label", ""), "entity": ent["entity_name"]}
                # B19: vị trí thứ 5 = khoi phải để None (trước đây bị nhét TÊN công ty vào khoi ->
                # nhiễu GROUP BY khoi). Chỉ vị trí 6 = cong_ty mang tên thực thể.
                melted.append((ri, rr["role"], rr.get("label", ""), None,
                               None, ent["entity_name"], val, payload))

    sample_mapped_rows = [
        {"source_row": m[0], "dim1": m[1], "dim2": m[2], "dim3": m[3],
         "khoi": m[4], "cong_ty": m[5], "amount": m[6]}
        for m in melted[:5]
    ]

    # GUARD BIÊN ĐỘ (đồng bộ với template_filler._MONEY_SANITY_MAX_TY): amount đã scale mà vẫn
    # vượt xa quy mô "tỷ đồng" hợp lý -> gần như chắc quên value_scale=1e-9 (nguồn là VND).
    max_money = max((abs(m[6]) for m in melted), default=0.0)
    scale_warn = max_money > _GEN_MONEY_SANITY_MAX_TY

    # B24: KHÔNG đưa file_name vào fingerprint -> mapping học được TÁI DÙNG cho file khác cùng layout
    # (sheet+orientation+columns). Định danh file duy nhất vẫn do content_hash lo ở dedup.
    fp = bb.fingerprint([sheet, orientation,
                         json.dumps(mapping.get("columns") or mapping.get("row_roles") or [],
                                    ensure_ascii=False)])

    if dry_run:
        result = {"dry_run": True, "target_report_type": target_rt, "row_count": len(melted),
                  "sample_mapped_rows": sample_mapped_rows, "skipped_duplicate": False,
                  "max_money_ty": round(max_money, 4), "scale_warning": scale_warn,
                  "message": f"Dry-run: dự kiến ghi {len(melted)} dòng report_type={target_rt}."}
        if scale_warn:
            result["error"] = (
                f"BIÊN ĐỘ BẤT THƯỜNG: giá trị lớn nhất sau scale = {max_money:,.0f} "
                f"(> {_GEN_MONEY_SANITY_MAX_TY:,.0f}). Nhiều khả năng nguồn là VND và QUÊN "
                f"value_scale=1e-9. Kiểm tra lại trước khi ghi thật.")
        # Khép vòng học sớm: dry-run thành công (có dòng melt được, KHÔNG cảnh báo biên độ) -> lưu
        # mapping ngay (verified=False), không cần chờ ai bấm "ghi thật". Nhờ vậy autobatch/propose
        # không phải chạy lại LLM cho cùng 1 sheet đã verify đúng qua dry-run trước đó.
        if melted and not scale_warn:
            memory.report_spec_save(fp, mapping, verified=False)
        return result

    if scale_warn:
        # KHÔNG ghi raw_rows / KHÔNG học spec khi biên độ bất thường (khớp guard của template_fill).
        return {"dry_run": False, "row_count": 0, "sample_mapped_rows": sample_mapped_rows,
                "skipped_duplicate": False, "ok": False,
                "max_money_ty": round(max_money, 4), "scale_warning": True,
                "error": (
                    f"BIÊN ĐỘ BẤT THƯỜNG: giá trị lớn nhất sau scale = {max_money:,.0f} "
                    f"(> {_GEN_MONEY_SANITY_MAX_TY:,.0f}). Nhiều khả năng nguồn là VND và QUÊN "
                    f"value_scale=1e-9. Không ghi dữ liệu."),
                "message": "Bị chặn ghi do biên độ giá trị bất thường (xem error)."}

    content_hash = hashlib.sha1(data).hexdigest()
    dup = _ledger_find("generic", target_rt, period_val, fp, content_hash, match_fingerprint=True)
    if dup:
        return {"dry_run": False, "dataset_id": dup["dataset_id"], "row_count": 0,
                "sample_mapped_rows": sample_mapped_rows, "skipped_duplicate": True,
                "message": f"Sheet này (cùng file+layout) đã nạp trước đó, dataset_id={dup['dataset_id']}."}

    # B7: GỘP generic theo KỲ vào 1 dataset dùng chung (KHÔNG new_dataset mỗi sheet -> set_active
    # đè khiến chỉ sheet cuối hiện). Idempotent theo (report_type, source_file) -> nạp lại thay đúng chỗ.
    existing = [d for d in bb.repo.list_datasets("generic") if d.get("period") == period_val]
    ds_id = existing[0]["id"] if existing else bb.repo.create_dataset(
        f"GEN {period_val}" if period_val else f"GEN {file_name}", kind="generic", period=period_val)["id"]
    bb.repo.set_active(ds_id)
    db = bb.get_db()
    # Idempotent theo (report_type, source_file) — source_file là BASENAME nên 2 file trùng
    # tên ở 2 thư mục/công ty khác nhau sẽ xoá nhầm dòng của nhau; với row_major đã biết
    # cong_ty -> scope thêm theo công ty (khớp cách import_filled scope source_file+cong_ty).
    _del_conds = ["dataset_id=?", "report_type=?", "source_file=?"]
    _del_params = [ds_id, target_rt, file_name]
    if orientation == "row_major" and cong_ty:
        _del_conds.append("cong_ty=?")
        _del_params.append(cong_ty)
    db.execute("DELETE FROM raw_rows WHERE " + " AND ".join(_del_conds), _del_params)
    to_insert = [
        (ds_id, target_rt, ri, ngay_val, cong_ty_v, khoi_v, None, period_val, val, None,
         entity_code, entity_name, role, json.dumps(payload, ensure_ascii=False), file_name)  # B10
        for ri, entity_code, entity_name, role, khoi_v, cong_ty_v, val, payload in melted
    ]
    db.executemany(
        "INSERT INTO raw_rows(dataset_id,report_type,row_index,ngay,cong_ty,khoi,"
        "cost_center,period_month,amount,amount2,dim1,dim2,dim3,payload,source_file) "
        "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        to_insert,
    )
    db.commit()

    _ledger_append({"dataset_kind": "generic", "report_type": target_rt, "period": period_val,
                    "fingerprint": fp, "content_hash": content_hash, "dataset_id": ds_id,
                    "file_name": file_name, "sheet": sheet})
    memory.report_spec_save(fp, mapping, verified=True)

    return {"dry_run": False, "dataset_id": ds_id, "row_count": len(to_insert),
            "sample_mapped_rows": sample_mapped_rows, "skipped_duplicate": False,
            "message": f"Đã ghi {len(to_insert)} dòng report_type={target_rt}."}


@mcp.tool()
def template_contract_info(target_sheet: str = None) -> dict:
    """MÔ TẢ TEMPLATE VÀNG cho analyst — ĐỌC TRƯỚC KHI DỰNG MAPPING.

    Trả guide: đơn vị (tỷ đồng -> value_scale khi nguồn VND), quy tắc map, mỗi sheet
    {mục đích, grain, cot_nhap_lieu (CHỈ map vào đây), cot_KHONG_map (công thức, bỏ)},
    danh sách công ty/khối/mã cost center hợp lệ, chỉ tiêu KQKD chuẩn, sheet đích theo loại.

    target_sheet: TRUYỀN MÃ SHEET TEMPLATE ĐÍCH bạn đang điền (vd '05_PHAITHU', '01_HQKD' —
    KHÔNG phải tên sheet nguồn như '131') để nhận guide GỌN cho riêng sheet đó (nhanh hơn, ít
    nhiễu). Chưa biết đích -> gọi không tham số để xem toàn bộ 13 sheet + chọn."""
    return contract.guide(target_sheet)


@mcp.tool()
def extraction_guide(ma_congty: str = None, file_name: str = None) -> dict:
    """HƯỚNG DẪN LẤY DỮ LIỆU per-đơn vị (Lớp 2 — knowledge/{ma_congty}.yaml hoặc _chung.yaml).

    GỌI SAU khi xác định ma_congty (từ tên file/thư mục, đối chiếu companies.yaml) và TRƯỚC KHI
    dựng mapping cho template_fill — để biết ĐÚNG sheet nguồn, mã số/cột, quy đổi đơn vị cho công
    ty này. Nếu công ty có layout đặc thù (vd AAG: An KS + An Taxi khác chuẩn TT200), trả về guide
    RIÊNG của công ty đó (THAY THẾ hoàn toàn, KHÔNG merge với hướng dẫn chung) — nếu không có guide
    riêng, trả hướng dẫn chung (_chung.yaml).

    Không truyền ma_congty (chưa xác định) nhưng có file_name -> tự thử suy mã công ty từ tên file.
    Trả {"source": tên file guide đã dùng, "ma_congty", "is_specific": bool, "content": {...}}.
    """
    return extraction.load_guide(ma_congty, file_name)


@mcp.tool()
def template_fill(file_path: str, source_sheet: str, target_sheet: str, mapping: dict,
                  period: str = None, cong_ty: str = None, dry_run: bool = True,
                  auto_import: bool = False, value_scale: float = 1.0,
                  constants: dict = None, normalize_kqkd: bool = True,
                  rename_rows: dict = None) -> dict:
    """Điền số liệu từ file nguồn vào BẢN SAO template vàng, theo mapping cột.

    - mapping: {tên cột TEMPLATE (input): tên cột NGUỒN}. Xem template_contract_info() để biết cột.
      Tên cột TEMPLATE phải khớp ĐÚNG header (sai -> tool trả ok:false kèm danh sách cột hợp lệ).
      PIVOT bảng dọc: giá trị mapping có thể là dict {'src': 'CỘT GIÁ TRỊ', 'khi': {CỘT ĐK: đk}}
      — đk: chuỗi (khớp chứa-trong không dấu, trên giá trị forward-fill nên chịu được cột section
      chỉ ghi 1 lần), '' (ô phải RỖNG), '*' (ô phải KHÁC rỗng). Vd TC01_SD TIỀN (dòng = LOẠI
      TIỀN×công ty×ngân hàng) -> 03B_SODU_TIEN:
      {'Đơn vị': 'CÔNG TY',
       'Tiền mặt (tỷ)':   {'src': 'ĐẾN NGÀY HIỆN TẠI', 'khi': {'LOẠI TIỀN': 'TIỀN MẶT', 'CÔNG TY': '*', 'NGÂN HÀNG': ''}},
       'Tiền gửi NH (tỷ)': {'src': 'ĐẾN NGÀY HIỆN TẠI', 'khi': {'LOẠI TIỀN': 'TIỀN GỬI', 'CÔNG TY': '*', 'NGÂN HÀNG': ''}}}
    - constants: {tên cột TEMPLATE: giá trị HẰNG} — gán cứng mọi dòng (vd cost center khi sheet nguồn
      cấp CÔNG TY không có CC theo dòng: constants={'Mã Cost center ◀ NHẬP':'ST_GD'}).
    - rename_rows: {tên chỉ tiêu NGUỒN: tên chỉ tiêu CHUẨN} — ĐỔI TÊN đúng vài dòng về tên chuẩn để KPI
      sáng (KHÁC mapping: mapping đổi CỘT; rename_rows đổi GIÁ TRỊ tên của DÒNG). BẮT BUỘC cho 01_HQKD khi
      nguồn đặt tên riêng: chọn ĐÚNG 1 dòng tổng doanh thu -> 'Doanh thu thuần', 1 dòng tổng chi phí ->
      'Tổng chi phí', 1 dòng lợi nhuận trước thuế -> 'Lợi nhuận trước thuế'. Khớp KHÔNG DẤU + chứa-trong;
      CHỌN 1 DÒNG TỔNG CẤP CAO NHẤT / chỉ tiêu để KHÔNG cộng trùng với dòng con.
    - value_scale: nhân cột tiền '(tỷ)'. NGUỒN VND -> template TỶ ĐỒNG phải dùng value_scale=1e-9.
      (tool KHÔNG tự đổi đơn vị — phải truyền value_scale đúng, nếu không số sai 1 tỷ lần.)
    - Cost center resolve về mã chuẩn; không khớp -> unmapped_cc + để nguyên (khối=NULL -> "(Chưa phân bổ)").
    - dry_run=True: preview. dry_run=False: ghi file điền -> out_path; auto_import=True: nạp raw_rows.
    """
    data = _read_file(file_path)
    return template_filler.fill_from_source(
        data, source_sheet, target_sheet, mapping, period=period, cong_ty=cong_ty,
        file_name=os.path.basename(file_path), dry_run=dry_run, auto_import=auto_import,
        value_scale=value_scale, constants=constants, normalize_kqkd=normalize_kqkd,
        rename_rows=rename_rows, source_path=file_path)


@mcp.tool()
def autofill_run(file_path: str, cong_ty: str = None, period: str = None, dry_run: bool = True) -> dict:
    """RẺ, TẤT ĐỊNH (không LLM): với MỖI sheet nguồn trong file_path, tra fill_spec ĐÃ HỌC theo
    fingerprint (từ lần `template_fill`/`autofill_run` trước đó học được) — nếu khớp, điền template
    vàng bằng CHÍNH mapping đã học. dry_run=True (mặc định): CHỈ xem trước — trả mỗi sheet khớp kèm
    `row_count`/`sample`/`unresolved_cc`/`scale_warning`, KHÔNG ghi gì. dry_run=False: ghi file điền
    + nạp raw_rows luôn cho các sheet khớp (dùng SAU khi người dùng đã xác nhận qua bước dry_run).

    Dùng làm BƯỚC 0 của `analyst` cho MỌI file trước khi phân tích: nếu `any_processed=true` và
    `skipped_sheets` rỗng thì layout này ĐÃ TỪNG học — báo ImportPlan thẳng từ kết quả này, khỏi
    phải lặp lại `template_contract_info`/`extraction_guide`/`sheet_profile`/dựng mapping từ đầu.
    `skipped_sheets` khác rỗng = sheet đó chưa có spec, vẫn cần đi tiếp luồng phân tích cho RIÊNG
    sheet đó (các sheet đã khớp trong `processed` thì bỏ qua, không phân tích lại).
    """
    return template_filler.autofill_file(_resolve_path(file_path), period=period, cong_ty=cong_ty,
                                         dry_run=dry_run)


@mcp.tool()
def catalog_reindex(root_dir: str = None) -> dict:
    """Quét lại thư mục file đã kéo về (Connect_VPS/received_reports mặc định) -> cập nhật
    source_catalog. Trả {indexed, total_in_catalog}. QA tra qua catalog_search."""
    from .common import source_catalog
    return source_catalog.index_dir(root_dir)


@mcp.tool()
def template_import(filled_path: str, cong_ty: str = None) -> dict:
    """Nạp 1 file ĐÃ ĐIỀN template chuẩn (do template_fill tạo trong template_trust/filled/) vào
    raw_rows. cong_ty: GỘP đa-công-ty theo kỳ (1 dataset/kỳ, idempotent theo (kỳ,cong_ty), đóng dấu
    cong_ty). Trả {dataset_id, grain, by_type, rows_imported, period}."""
    if not os.path.isabs(filled_path):
        filled_path = os.path.join(template_filler.FILLED_DIR, filled_path)
    # Truyền source_file (tên file điền) để import_filled xoá-thay theo PHẠM VI file, KHÔNG wipe
    # toàn kỳ (fix B3). Kèm cong_ty -> scope thêm theo công ty.
    return template_filler.import_filled(filled_path, cong_ty=cong_ty,
                                         source_file=os.path.basename(filled_path))


if __name__ == "__main__":
    mcp.run()
